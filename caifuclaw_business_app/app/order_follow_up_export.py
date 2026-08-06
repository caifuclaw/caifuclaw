# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

from __future__ import annotations

import asyncio
import copy
import hashlib
import logging
import os
import shutil
import socket
import subprocess
import sys
import tempfile
import threading
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import asc, func, or_, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .database import SessionLocal, engine
from .models import (
    LocalUser,
    Order,
    OrderFollowUpExportArtifact,
    OrderFollowUpExportItem,
    OrderFollowUpExportJob,
    OrderItem,
    PlatformAccount,
    ScheduledTaskRun,
    ScheduledTaskRunOrder,
)
from .product_models import Product, ProductInventory, ProductShopMapping
from .order_operation_logs import ORDER_LOG_SYSTEM_SOURCE, SYSTEM_OPERATOR, add_order_operation_logs
from .order_types import (
    order_has_bsi_draft,
    order_is_joom_fbj_warehouse,
)
from .settings import Settings, get_settings


logger = logging.getLogger(__name__)

LOCAL_TIMEZONE = ZoneInfo("Asia/Shanghai")
WORKBOOK_LOCK_KEY = 2026071901
WORKER_NAME = f"order-follow-up-export:{socket.gethostname()}:{os.getpid()}"

JOB_PENDING = "pending"
JOB_RUNNING = "running"
JOB_RETRY_WAIT = "retry_wait"
JOB_SUCCESS = "success"
JOB_SKIPPED = "skipped"
JOB_FAILED = "failed"
TERMINAL_JOB_STATUSES = {JOB_SUCCESS, JOB_SKIPPED, JOB_FAILED}

ITEM_PENDING = "pending"
ITEM_SUCCESS = "success"
ITEM_SKIPPED = "skipped"

ARTIFACT_WORKBOOK = "workbook"
ARTIFACT_PURCHASE_PLAN = "purchase_plan"

ORDER_SHEET_NAME = "订单总表"
DIRECT_PLAN_SHEET_NAME = "直发货采购计划"
EXPORT_LOG_SHEET_NAME = "_caifuclaw_export_log"
DIRECT_EXPORT_LOG_SHEET_NAME = "_caifuclaw_direct_log"
PURCHASE_META_SHEET_NAME = "_caifuclaw_export_meta"
REQUIRED_SHEETS = {ORDER_SHEET_NAME, DIRECT_PLAN_SHEET_NAME, "产品目录"}

ORDER_HEADERS = [
    "配货日",
    "平台",
    "店铺名",
    "创建时间",
    "订单编号",
    "国家二字码",
    "客户姓名",
    "SKU",
    "商品数量",
    "商品销售单价",
    "币种",
    "自选物流",
    "物流渠道",
    "最后发货期限",
    "间隔",
    "货运单号",
    "发出截止时间",
    "产品中文名称",
]

_worker_task: asyncio.Task | None = None
_process_lock = threading.Lock()


@dataclass(frozen=True)
class ExportPaths:
    root: Path
    sync_dir: Path
    backup_dir: Path
    purchase_plan_dir: Path
    workbook: Path
    template: Path


@dataclass(frozen=True)
class WorkbookWriteResult:
    file_path: Path
    item_rows: dict[int, int]
    changed_item_count: int
    direct_row_count: int


@dataclass(frozen=True)
class PurchasePlanWriteResult:
    file_path: Path | None
    row_count: int


@dataclass(frozen=True)
class ExistingWorkbookOrder:
    row_number: int
    mapping_missing: bool
    managed_by_caifuclaw: bool


class ExportLockBusy(RuntimeError):
    pass


def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _local_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    aware = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    return aware.astimezone(LOCAL_TIMEZONE).replace(tzinfo=None)


def _local_date(value: datetime | None, fallback: datetime) -> date:
    return (_local_datetime(value) or _local_datetime(fallback) or fallback).date()


def _normalized_sku(value: str | None) -> str:
    return (value or "").strip().lower()


def _normalized_platform(value: str | None) -> str:
    platform = (value or "").strip().lower()
    return "joom" if platform == "joom_logistics" else platform


def _normalized_order_number(platform: str | None, value: str | None) -> str:
    number = (value or "").strip().lower()
    return number.replace("-", "") if _normalized_platform(platform) == "allegro" else number


def _unique_texts(*values) -> list[str]:
    result: list[str] = []
    for value in values:
        text_value = str(value or "").strip()
        if text_value and text_value not in result:
            result.append(text_value)
    return result


def _nested_value(payload: dict, *path: str):
    current = payload
    for key in path:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _order_item_sku_aliases(item: OrderItem, platform: str | None) -> list[str]:
    raw_payload = item.raw_payload if isinstance(item.raw_payload, dict) else {}
    source_payload = raw_payload.get("raw_payload")
    source_payload = source_payload if isinstance(source_payload, dict) else {}
    external_ids: list[object] = []
    if _normalized_platform(platform) == "allegro":
        external_ids.extend(
            [
                _nested_value(raw_payload, "offer", "external", "id"),
                _nested_value(source_payload, "offer", "external", "id"),
                raw_payload.get("external_id"),
                source_payload.get("external_id"),
                _nested_value(raw_payload, "allegro_product_info", "external_id"),
            ]
        )
    return _unique_texts(*external_ids, item.sku)


def _export_sku(item: OrderItem, platform: str | None) -> str:
    aliases = _order_item_sku_aliases(item, platform)
    return aliases[0] if aliases else (item.sku or "").strip()


def _export_order_number(order: Order) -> str:
    platform = _normalized_platform(order.platform)
    if platform == "ozon":
        return order.posting_number or order.platform_order_no or order.platform_order_id or ""
    number = order.platform_order_no or order.posting_number or order.platform_order_id or ""
    return number.replace("-", "") if platform == "allegro" else number


def _order_number_aliases(order: Order) -> list[str]:
    return _unique_texts(
        _export_order_number(order),
        order.platform_order_no,
        order.posting_number,
        order.platform_order_id,
        order.shipment_tracking_number,
    )


def _export_buyer_name(order: Order) -> str:
    payload = order.raw_payload if isinstance(order.raw_payload, dict) else {}
    platform = _normalized_platform(order.platform)
    if platform == "allegro":
        first_name = _nested_value(payload, "delivery", "address", "firstName")
        last_name = _nested_value(payload, "delivery", "address", "lastName")
        full_name = " ".join(str(part).strip() for part in (first_name, last_name) if str(part or "").strip())
        candidates = [
            full_name,
            _nested_value(payload, "delivery", "address", "name"),
            _nested_value(payload, "shipping", "receiver_address", "name"),
        ]
    elif platform == "mercadolibre":
        candidates = [
            _nested_value(payload, "shipment", "destination", "receiver_name"),
            _nested_value(payload, "shipping", "destination", "receiver_name"),
            _nested_value(payload, "shipping", "receiver_address", "receiver_name"),
            _nested_value(payload, "shipping", "receiver_address", "name"),
            _nested_value(payload, "shipping", "receiver_name"),
        ]
    else:
        candidates = []
    return next((str(value).strip() for value in candidates if str(value or "").strip()), order.buyer_name or "")


def _export_logistics_channel(order: Order, settings: Settings) -> str:
    current = (order.logistics_channel or "").strip()
    if current:
        return current
    fallbacks = settings.order_follow_up_export_logistics_channel_fallbacks
    platform = (order.platform or "").strip().lower()
    return str(fallbacks.get(platform) or fallbacks.get(_normalized_platform(platform)) or "").strip()


def _safe_int(value, default: int = 0) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return default


def _safe_float(value):
    if value in {None, ""}:
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, TypeError, ValueError):
        return str(value)


def _snapshot_datetime(value: datetime | None) -> str:
    local = _local_datetime(value)
    return local.isoformat(timespec="seconds") if local else ""


def _parse_snapshot_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _export_paths(settings: Settings | None = None) -> ExportPaths:
    current = settings or get_settings()
    root = current.order_follow_up_export_data_path
    sync_dir = root / current.order_follow_up_export_sync_dir
    backup_dir = root / current.order_follow_up_export_backup_dir
    purchase_plan_dir = root / current.order_follow_up_export_purchase_plan_dir
    return ExportPaths(
        root=root,
        sync_dir=sync_dir,
        backup_dir=backup_dir,
        purchase_plan_dir=purchase_plan_dir,
        workbook=sync_dir / current.order_follow_up_export_workbook_name,
        template=sync_dir / current.order_follow_up_export_template_workbook_name,
    )


def _ensure_export_directories(paths: ExportPaths) -> None:
    paths.sync_dir.mkdir(parents=True, exist_ok=True)
    paths.backup_dir.mkdir(parents=True, exist_ok=True)
    paths.purchase_plan_dir.mkdir(parents=True, exist_ok=True)


def enqueue_order_follow_up_export(scheduled_task_run_id: int) -> bool:
    settings = get_settings()
    if not settings.order_follow_up_export_enabled:
        return False
    if scheduled_task_run_id <= settings.order_follow_up_export_cutover_run_id:
        return False

    db = SessionLocal()
    try:
        run = db.get(ScheduledTaskRun, scheduled_task_run_id)
        if not run or run.status != "success" or run.task_type != "auto_order_pipeline":
            return False
        existing = db.scalar(
            select(OrderFollowUpExportJob.id).where(
                OrderFollowUpExportJob.scheduled_task_run_id == scheduled_task_run_id
            )
        )
        if existing:
            return False
        db.add(
            OrderFollowUpExportJob(
                scheduled_task_run_id=scheduled_task_run_id,
                workbook_key=settings.order_follow_up_export_workbook_name,
                status=JOB_PENDING,
                max_attempts=max(1, settings.order_follow_up_export_max_attempts),
                created_at=_utc_now(),
                updated_at=_utc_now(),
            )
        )
        db.commit()
        return True
    except IntegrityError:
        db.rollback()
        return False
    finally:
        db.close()


def reconcile_order_follow_up_export_jobs(limit: int = 100) -> int:
    settings = get_settings()
    if not settings.order_follow_up_export_enabled:
        return 0
    db = SessionLocal()
    try:
        run_ids = list(
            db.scalars(
                select(ScheduledTaskRun.id)
                .outerjoin(
                    OrderFollowUpExportJob,
                    OrderFollowUpExportJob.scheduled_task_run_id == ScheduledTaskRun.id,
                )
                .where(
                    ScheduledTaskRun.id > settings.order_follow_up_export_cutover_run_id,
                    ScheduledTaskRun.task_type == "auto_order_pipeline",
                    ScheduledTaskRun.status == "success",
                    OrderFollowUpExportJob.id.is_(None),
                )
                .order_by(asc(ScheduledTaskRun.id))
                .limit(max(1, limit))
            ).all()
        )
    finally:
        db.close()

    created = 0
    for run_id in run_ids:
        try:
            created += int(enqueue_order_follow_up_export(int(run_id)))
        except Exception:
            logger.exception("Failed to reconcile order follow up export run_id=%s", run_id)
    return created


def _recover_expired_jobs(db: Session, now: datetime) -> int:
    rows = db.scalars(
        select(OrderFollowUpExportJob).where(
            OrderFollowUpExportJob.status == JOB_RUNNING,
            OrderFollowUpExportJob.lease_until.is_not(None),
            OrderFollowUpExportJob.lease_until < now,
        )
    ).all()
    for row in rows:
        row.claimed_by = ""
        row.claimed_at = None
        row.lease_until = None
        row.heartbeat_at = None
        if row.attempt_count >= row.max_attempts:
            row.status = JOB_FAILED
            row.finished_at = now
            row.error_message = row.error_message or "导出执行器租约超时且已达到最大重试次数"
        else:
            row.status = JOB_RETRY_WAIT
            row.next_retry_at = now
            row.error_message = row.error_message or "导出执行器租约超时，任务已重新排队"
        row.updated_at = now
    if rows:
        db.commit()
    return len(rows)


def _claim_next_job(db: Session) -> OrderFollowUpExportJob | None:
    settings = get_settings()
    now = _utc_now()
    _recover_expired_jobs(db, now)
    oldest_id = db.scalar(
        select(func.min(OrderFollowUpExportJob.id)).where(
            OrderFollowUpExportJob.status.notin_(TERMINAL_JOB_STATUSES)
        )
    )
    if not oldest_id:
        return None
    row = db.scalar(
        select(OrderFollowUpExportJob)
        .where(
            OrderFollowUpExportJob.id == oldest_id,
            or_(
                OrderFollowUpExportJob.status == JOB_PENDING,
                (
                    (OrderFollowUpExportJob.status == JOB_RETRY_WAIT)
                    & (
                        OrderFollowUpExportJob.next_retry_at.is_(None)
                        | (OrderFollowUpExportJob.next_retry_at <= now)
                    )
                ),
            ),
        )
        .with_for_update(skip_locked=True)
    )
    if not row:
        return None
    row.status = JOB_RUNNING
    row.attempt_count = _safe_int(row.attempt_count) + 1
    row.claimed_by = WORKER_NAME
    row.claimed_at = now
    row.heartbeat_at = now
    row.lease_until = now + timedelta(seconds=max(60, settings.order_follow_up_export_lease_seconds))
    row.next_retry_at = None
    row.error_message = ""
    row.started_at = row.started_at or now
    row.updated_at = now
    db.commit()
    db.refresh(row)
    return row


def _mapping_rank(updated_at: datetime | None, created_at: datetime | None, mapping_id: int) -> tuple[datetime, int]:
    return updated_at or created_at or datetime.min, mapping_id


def _direct_plan_eligible(order: Order) -> bool:
    if bool(order.is_overseas_warehouse):
        return False
    fulfillment = (order.fulfillment_type or "").strip().upper()
    if fulfillment in {"FBO", "FBP", "FBJ", "OVERSEAS", "OVERSEAS_WAREHOUSE"}:
        return False
    return (order.biz_status or "") != "已作废"


def _processed_run_order_ids(db: Session, run_id: int) -> list[int]:
    rows = db.scalars(
        select(ScheduledTaskRunOrder)
        .where(ScheduledTaskRunOrder.run_id == run_id, ScheduledTaskRunOrder.order_id > 0)
        .order_by(asc(ScheduledTaskRunOrder.id))
    ).all()
    return list(
        dict.fromkeys(
            row.order_id
            for row in rows
            if (row.status_after or row.purchase_order_id or row.error_message)
        )
    )


def _collect_snapshots(
    db: Session,
    job: OrderFollowUpExportJob,
    settings: Settings | None = None,
) -> list[dict]:
    current_settings = settings or get_settings()
    if not job.scheduled_task_run_id:
        return []
    run = db.get(ScheduledTaskRun, job.scheduled_task_run_id)
    if not run:
        return []
    order_ids = _processed_run_order_ids(db, run.id)
    if not order_ids:
        return []

    order_item_rows = db.execute(
        select(Order, OrderItem)
        .join(OrderItem, OrderItem.order_id == Order.id)
        .where(Order.id.in_(order_ids))
        .order_by(asc(Order.id), asc(OrderItem.id))
    ).all()
    if not order_item_rows:
        return []

    account_rows = db.execute(
        select(PlatformAccount.id, PlatformAccount.platform, PlatformAccount.account_id)
    ).all()
    account_ids = {
        ((platform or "").strip().lower(), (account_id or "").strip()): int(shop_db_id)
        for shop_db_id, platform, account_id in account_rows
    }
    shop_db_ids = sorted(
        {
            account_ids.get(((order.platform or "").strip().lower(), (order.shop_id or "").strip()))
            for order, _item in order_item_rows
        }
        - {None}
    )
    item_sku_aliases = {
        int(item.id): _order_item_sku_aliases(item, order.platform)
        for order, item in order_item_rows
    }
    normalized_aliases = sorted(
        {
            _normalized_sku(alias)
            for aliases in item_sku_aliases.values()
            for alias in aliases
            if _normalized_sku(alias)
        }
    )

    exact_mappings: dict[tuple[int, str], tuple[tuple[datetime, int], dict]] = {}
    normalized_mappings: dict[tuple[int, str], tuple[tuple[datetime, int], dict]] = {}
    global_mappings: dict[str, tuple[set[int], tuple[tuple[datetime, int], dict]]] = {}
    if shop_db_ids or normalized_aliases:
        mapping_filters = []
        if shop_db_ids:
            mapping_filters.append(ProductShopMapping.shop_id.in_(shop_db_ids))
        if normalized_aliases:
            mapping_filters.append(func.lower(func.trim(ProductShopMapping.shop_sku)).in_(normalized_aliases))
        mapping_rows = db.execute(
            select(
                ProductShopMapping.id,
                ProductShopMapping.shop_id,
                ProductShopMapping.shop_sku,
                ProductShopMapping.created_at,
                ProductShopMapping.updated_at,
                Product.id.label("product_id"),
                Product.internal_name,
                Product.cost,
                ProductInventory.stock_qty,
                LocalUser.display_name,
                LocalUser.username,
            )
            .join(Product, Product.id == ProductShopMapping.product_id)
            .outerjoin(ProductInventory, ProductInventory.product_id == Product.id)
            .outerjoin(LocalUser, LocalUser.id == Product.buyer_user_id)
            .where(or_(*mapping_filters))
        ).all()
        for row in mapping_rows:
            mapping_id = int(row.id)
            payload = {
                "product_id": int(row.product_id),
                "product_name": row.internal_name or "",
                "cost": _safe_float(row.cost),
                "stock_qty": _safe_int(row.stock_qty),
                "buyer": row.display_name or row.username or "",
            }
            rank = _mapping_rank(row.updated_at, row.created_at, mapping_id)
            exact_key = (int(row.shop_id), row.shop_sku or "")
            normalized_key = (int(row.shop_id), _normalized_sku(row.shop_sku))
            if exact_key not in exact_mappings or rank > exact_mappings[exact_key][0]:
                exact_mappings[exact_key] = (rank, payload)
            if normalized_key not in normalized_mappings or rank > normalized_mappings[normalized_key][0]:
                normalized_mappings[normalized_key] = (rank, payload)
            normalized_sku = normalized_key[1]
            if normalized_sku in normalized_aliases:
                if normalized_sku not in global_mappings:
                    global_mappings[normalized_sku] = ({int(row.product_id)}, (rank, payload))
                else:
                    product_ids, best = global_mappings[normalized_sku]
                    product_ids.add(int(row.product_id))
                    if rank > best[0]:
                        global_mappings[normalized_sku] = (product_ids, (rank, payload))

    fallback_time = run.ended_at or job.created_at or _utc_now()
    snapshots: list[dict] = []
    for order, item in order_item_rows:
        aliases = item_sku_aliases.get(int(item.id), [])
        sku = _export_sku(item, order.platform)
        shop_db_id = account_ids.get(
            ((order.platform or "").strip().lower(), (order.shop_id or "").strip())
        )
        mapped = None
        if shop_db_id:
            for alias in aliases:
                candidate = exact_mappings.get((shop_db_id, alias)) or normalized_mappings.get(
                    (shop_db_id, _normalized_sku(alias))
                )
                if candidate:
                    mapped = candidate[1]
                    break
        if mapped is None:
            global_candidates = [
                global_mappings[_normalized_sku(alias)][1]
                for alias in aliases
                if _normalized_sku(alias) in global_mappings
                and len(global_mappings[_normalized_sku(alias)][0]) == 1
            ]
            product_ids = {int(candidate[1]["product_id"]) for candidate in global_candidates}
            if len(product_ids) == 1:
                mapped = max(global_candidates, key=lambda candidate: candidate[0])[1]
        mapping_status = "mapped" if mapped and mapped.get("product_name") else "missing"
        product_name = mapped.get("product_name") if mapped else ""
        display_name = product_name or sku
        allocation_date = _local_date(order.picking_at, fallback_time)
        order_type = (
            f"{(order.platform or '').upper()}海外仓"
            if order.is_overseas_warehouse
            else 0
        )
        snapshots.append(
            {
                "order_id": int(order.id),
                "order_item_id": int(item.id),
                "allocation_date": allocation_date.isoformat(),
                "platform": order.platform or "",
                "shop_name": order.shop_name or "",
                "created_time": _snapshot_datetime(order.payment_at or order.platform_created_at or order.created_at),
                "order_number": _export_order_number(order),
                "order_number_aliases": _order_number_aliases(order),
                "country_code": order.country_code or "",
                "buyer_name": _export_buyer_name(order),
                "sku": sku,
                "source_sku": (item.sku or "").strip(),
                "sku_aliases": aliases,
                "quantity": max(0, _safe_int(item.quantity)),
                "unit_price": _safe_float(item.unit_price),
                "currency": item.currency or order.currency or "",
                "buyer_selected_logistics": order.buyer_selected_logistics or "",
                "logistics_channel": _export_logistics_channel(order, current_settings),
                "shipping_deadline": _snapshot_datetime(order.shipping_deadline_at),
                "tracking_number": order.shipment_tracking_number or order.posting_number or "",
                "dispatch_deadline": _snapshot_datetime(order.dispatch_deadline_at),
                "product_id": mapped.get("product_id") if mapped else None,
                "product_name": product_name,
                "display_product_name": display_name,
                "mapping_status": mapping_status,
                "stock_qty": mapped.get("stock_qty", 0) if mapped else 0,
                "cost": mapped.get("cost") if mapped else None,
                "buyer": mapped.get("buyer", "") if mapped else "",
                "order_type": order_type,
                "is_overseas_warehouse": bool(order.is_overseas_warehouse),
                "direct_plan_eligible": _direct_plan_eligible(order),
                "source_status": order.biz_status or "",
                "source_run_id": int(run.id),
            }
        )
    return snapshots


def _workbook_identity_keys(platform, order_number, sku) -> set[tuple[str, str, str]]:
    normalized_order = _normalized_order_number(platform, str(order_number or ""))
    normalized_sku = _normalized_sku(str(sku or ""))
    if not normalized_order or not normalized_sku:
        return set()
    return {(_normalized_platform(str(platform or "")), normalized_order, normalized_sku)}


def _existing_workbook_orders(path: Path) -> dict[tuple[str, str, str], ExistingWorkbookOrder]:
    if not path.is_file():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    try:
        if ORDER_SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError(f"导出工作簿缺少工作表：{ORDER_SHEET_NAME}")
        managed_rows: set[int] = set()
        if EXPORT_LOG_SHEET_NAME in workbook.sheetnames:
            log_sheet = workbook[EXPORT_LOG_SHEET_NAME]
            managed_rows = {
                _safe_int(row[4] if len(row) > 4 else None)
                for row in log_sheet.iter_rows(min_row=2, max_col=5, values_only=True)
                if _safe_int(row[4] if len(row) > 4 else None) > 0
            }
        sheet = workbook[ORDER_SHEET_NAME]
        result: dict[tuple[str, str, str], ExistingWorkbookOrder] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, max_col=18, values_only=True),
            start=2,
        ):
            platform = row[1] if len(row) > 1 else None
            order_number = row[4] if len(row) > 4 else None
            sku = row[7] if len(row) > 7 else None
            product_name = row[17] if len(row) > 17 else None
            existing = ExistingWorkbookOrder(
                row_number=row_number,
                mapping_missing=str(product_name or "").strip() in {"", "未登记目录"},
                managed_by_caifuclaw=row_number in managed_rows,
            )
            for key in _workbook_identity_keys(platform, order_number, sku):
                previous = result.get(key)
                if previous is None or (existing.managed_by_caifuclaw and not previous.managed_by_caifuclaw):
                    result[key] = existing
        return result
    finally:
        workbook.close()


def _matching_workbook_order(
    snapshot: dict,
    existing_orders: dict[tuple[str, str, str], ExistingWorkbookOrder],
) -> ExistingWorkbookOrder | None:
    order_numbers = _unique_texts(
        snapshot.get("order_number"),
        *(snapshot.get("order_number_aliases") or []),
        snapshot.get("tracking_number"),
    )
    skus = _unique_texts(
        snapshot.get("sku"),
        snapshot.get("source_sku"),
        *(snapshot.get("sku_aliases") or []),
    )
    for order_number in order_numbers:
        for sku in skus:
            for key in _workbook_identity_keys(snapshot.get("platform"), order_number, sku):
                if key in existing_orders:
                    return existing_orders[key]
    return None


def _workbook_match_applies(snapshot: dict, existing: ExistingWorkbookOrder | None) -> bool:
    if existing is None:
        return False
    return not snapshot.get("is_overseas_warehouse") or existing.managed_by_caifuclaw


def _prepare_job_items(
    db: Session,
    job: OrderFollowUpExportJob,
    settings: Settings | None = None,
) -> list[OrderFollowUpExportItem]:
    existing = db.scalars(
        select(OrderFollowUpExportItem)
        .where(OrderFollowUpExportItem.job_id == job.id)
        .order_by(asc(OrderFollowUpExportItem.id))
    ).all()
    if existing:
        return list(existing)

    current_settings = settings or get_settings()
    snapshots = _collect_snapshots(db, job, current_settings)
    existing_orders: dict[tuple[str, str, str], ExistingWorkbookOrder] = {}
    workbook_path = _export_paths(current_settings).workbook
    if current_settings.order_follow_up_export_dedupe_existing_workbook:
        existing_orders = _existing_workbook_orders(workbook_path)
    item_ids = [int(snapshot["order_item_id"]) for snapshot in snapshots]
    prior_rows: dict[int, OrderFollowUpExportItem] = {}
    if item_ids:
        previous = db.scalars(
            select(OrderFollowUpExportItem)
            .where(
                OrderFollowUpExportItem.order_item_id.in_(item_ids),
                OrderFollowUpExportItem.job_id < job.id,
                OrderFollowUpExportItem.status == ITEM_SUCCESS,
                OrderFollowUpExportItem.action.in_(["append", "update"]),
            )
            .order_by(asc(OrderFollowUpExportItem.id))
        ).all()
        for row in previous:
            prior_rows[row.order_item_id] = row

    created: list[OrderFollowUpExportItem] = []
    for snapshot in snapshots:
        prior = prior_rows.get(int(snapshot["order_item_id"]))
        action = "append"
        worksheet_row = None
        if prior:
            worksheet_row = prior.worksheet_row
            if prior.mapping_status == "missing" and snapshot["mapping_status"] == "mapped" and worksheet_row:
                action = "update"
            else:
                action = "skip"
                snapshot["skip_reason"] = "previous_export_item"
        else:
            workbook_order = _matching_workbook_order(snapshot, existing_orders)
            if _workbook_match_applies(snapshot, workbook_order):
                worksheet_row = workbook_order.row_number
                if workbook_order.mapping_missing and snapshot["mapping_status"] == "mapped":
                    action = "update"
                else:
                    action = "skip"
                    snapshot["skip_reason"] = "existing_workbook_order"
        row = OrderFollowUpExportItem(
            job_id=job.id,
            order_id=int(snapshot["order_id"]),
            order_item_id=int(snapshot["order_item_id"]),
            action=action,
            status=ITEM_SKIPPED if action == "skip" else ITEM_PENDING,
            mapping_status=str(snapshot["mapping_status"]),
            worksheet_row=worksheet_row,
            snapshot_json=snapshot,
            created_at=_utc_now(),
            updated_at=_utc_now(),
        )
        db.add(row)
        created.append(row)
    db.commit()
    for row in created:
        db.refresh(row)
    return created


def _copy_row_template(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        if isinstance(source.value, str) and source.value.startswith("="):
            try:
                target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
            except Exception:
                target.value = source.value
        else:
            target.value = None


def _last_data_row(ws, columns: tuple[int, ...], minimum_row: int) -> int:
    for row_number in range(ws.max_row, minimum_row - 1, -1):
        if any(ws.cell(row_number, column).value not in {None, ""} for column in columns):
            return row_number
    return minimum_row


def _order_sheet_values(snapshot: dict) -> dict[int, object]:
    mapping_missing = snapshot.get("mapping_status") != "mapped"
    return {
        1: date.fromisoformat(snapshot["allocation_date"]),
        2: snapshot.get("platform", ""),
        3: snapshot.get("shop_name", ""),
        4: _parse_snapshot_datetime(snapshot.get("created_time")),
        5: snapshot.get("order_number", ""),
        6: snapshot.get("country_code", ""),
        7: snapshot.get("buyer_name", ""),
        8: snapshot.get("sku", ""),
        9: _safe_int(snapshot.get("quantity")),
        10: snapshot.get("unit_price"),
        11: snapshot.get("currency", ""),
        12: snapshot.get("buyer_selected_logistics", ""),
        13: snapshot.get("logistics_channel", ""),
        14: _parse_snapshot_datetime(snapshot.get("shipping_deadline")),
        15: None,
        16: snapshot.get("tracking_number", ""),
        17: _parse_snapshot_datetime(snapshot.get("dispatch_deadline")),
        18: "未登记目录" if mapping_missing else snapshot.get("product_name", ""),
        22: snapshot.get("order_type", 0),
    }


def _write_order_row(ws, row_number: int, snapshot: dict, *, append: bool, template_row: int) -> None:
    if append:
        _copy_row_template(ws, template_row, row_number)
    for column, value in _order_sheet_values(snapshot).items():
        ws.cell(row_number, column).value = value


def _direct_group_key(snapshot: dict) -> str:
    allocation_date = str(snapshot.get("allocation_date") or "")
    product_id = _safe_int(snapshot.get("product_id"))
    if product_id:
        identity = f"product:{product_id}"
    else:
        platform = _normalized_platform(str(snapshot.get("platform") or ""))
        sku = _normalized_sku(str(snapshot.get("source_sku") or snapshot.get("sku") or ""))
        identity = f"sku:{platform}:{sku}"
    return f"{allocation_date}|{identity}"


def _direct_groups(items: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}
    for item in items:
        snapshot = item["snapshot"]
        if item["action"] != "append" or not snapshot.get("direct_plan_eligible"):
            continue
        product_name = snapshot.get("display_product_name") or snapshot.get("sku") or ""
        key = (snapshot.get("allocation_date") or "", product_name)
        if key not in grouped:
            grouped[key] = {
                "allocation_date": key[0],
                "product_name": key[1],
                "quantity": 0,
                "stock_qty": _safe_int(snapshot.get("stock_qty")),
                "cost": snapshot.get("cost"),
                "buyer": snapshot.get("buyer", ""),
                "mapping_status": snapshot.get("mapping_status", "missing"),
                "item_ids": [],
                "group_keys": [],
            }
        group = grouped[key]
        group["quantity"] += _safe_int(snapshot.get("quantity"))
        group["item_ids"].append(int(snapshot["order_item_id"]))
        group["group_keys"].append(_direct_group_key(snapshot))
        if snapshot.get("mapping_status") != "mapped":
            group["mapping_status"] = "missing"
    return sorted(grouped.values(), key=lambda row: (row["allocation_date"], row["product_name"]))


def _append_direct_rows(ws, groups: list[dict], exported_at: datetime) -> list[int]:
    if not groups:
        return []
    template_row = _last_data_row(ws, (5, 6), 4)
    next_row = template_row + 1
    for offset, group in enumerate(groups):
        row_number = next_row + offset
        _copy_row_template(ws, template_row, row_number)
        quantity = _safe_int(group["quantity"])
        stock_qty = _safe_int(group["stock_qty"])
        ws.cell(row_number, 5).value = group["allocation_date"]
        ws.cell(row_number, 6).value = group["product_name"]
        ws.cell(row_number, 7).value = quantity
        ws.cell(row_number, 8).value = stock_qty
        ws.cell(row_number, 9).value = exported_at.strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row_number, 10).value = quantity - stock_qty
        ws.cell(row_number, 11).value = group.get("cost")
        ws.cell(row_number, 12).value = None
        ws.cell(row_number, 13).value = group.get("buyer", "")
        for column in range(14, 18):
            ws.cell(row_number, column).value = None
        ws.cell(row_number, 18).value = (
            "未登记目录，仅提示，不可自动采购" if group.get("mapping_status") != "mapped" else ""
        )
    return [next_row + offset for offset in range(len(groups))]


def _update_direct_row(ws, row_number: int, snapshot: dict, exported_at: datetime) -> bool:
    if row_number < 4 or row_number > ws.max_row:
        return False
    quantity = _safe_int(ws.cell(row_number, 7).value)
    stock_qty = _safe_int(snapshot.get("stock_qty"))
    ws.cell(row_number, 5).value = snapshot.get("allocation_date") or ""
    ws.cell(row_number, 6).value = snapshot.get("display_product_name") or snapshot.get("sku") or ""
    ws.cell(row_number, 8).value = stock_qty
    ws.cell(row_number, 9).value = exported_at.strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row_number, 10).value = quantity - stock_qty
    ws.cell(row_number, 11).value = snapshot.get("cost")
    ws.cell(row_number, 13).value = snapshot.get("buyer", "")
    ws.cell(row_number, 18).value = (
        "未登记目录，仅提示，不可自动采购" if snapshot.get("mapping_status") != "mapped" else ""
    )
    return True


def _ensure_export_log_sheet(workbook) -> object:
    if EXPORT_LOG_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[EXPORT_LOG_SHEET_NAME]
    else:
        sheet = workbook.create_sheet(EXPORT_LOG_SHEET_NAME)
        sheet.append(["token", "job_id", "order_item_id", "action", "order_sheet_row", "exported_at"])
    sheet.sheet_state = "veryHidden"
    return sheet


def _ensure_direct_export_log_sheet(workbook) -> object:
    if DIRECT_EXPORT_LOG_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[DIRECT_EXPORT_LOG_SHEET_NAME]
    else:
        sheet = workbook.create_sheet(DIRECT_EXPORT_LOG_SHEET_NAME)
        sheet.append(
            [
                "token",
                "job_id",
                "order_item_id",
                "action",
                "direct_plan_row",
                "group_key",
                "quantity_delta",
                "exported_at",
            ]
        )
    sheet.sheet_state = "veryHidden"
    return sheet


def _existing_direct_export_log(workbook) -> dict[int, int]:
    if DIRECT_EXPORT_LOG_SHEET_NAME not in workbook.sheetnames:
        return {}
    result: dict[int, int] = {}
    sheet = workbook[DIRECT_EXPORT_LOG_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_id = _safe_int(row[2] if len(row) > 2 else None)
        direct_plan_row = _safe_int(row[4] if len(row) > 4 else None)
        if item_id and direct_plan_row:
            result[item_id] = direct_plan_row
    return result


def _existing_export_log(workbook, job_id: int) -> dict[int, int]:
    if EXPORT_LOG_SHEET_NAME not in workbook.sheetnames:
        return {}
    result: dict[int, int] = {}
    sheet = workbook[EXPORT_LOG_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _safe_int(row[1] if len(row) > 1 else None) != job_id:
            continue
        item_id = _safe_int(row[2] if len(row) > 2 else None)
        sheet_row = _safe_int(row[4] if len(row) > 4 else None)
        if item_id:
            result[item_id] = sheet_row
    return result


def _mark_full_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
    for sheet in workbook.worksheets:
        for pivot in getattr(sheet, "_pivots", []):
            cache = getattr(pivot, "cache", None)
            if cache is not None:
                cache.enableRefresh = True
                cache.refreshOnLoad = True
                cache.saveData = True


def _pivot_part_count(path: Path) -> int:
    with ZipFile(path) as archive:
        return sum(1 for name in archive.namelist() if name.startswith("xl/pivotTables/") and name.endswith(".xml"))


def _validate_main_workbook(path: Path, expected_job_id: int, expected_item_ids: set[int], minimum_pivots: int) -> dict[int, int]:
    if not path.is_file() or path.stat().st_size <= 0:
        raise RuntimeError(f"导出工作簿不存在或为空：{path}")
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    try:
        missing_sheets = REQUIRED_SHEETS - set(workbook.sheetnames)
        if missing_sheets:
            raise RuntimeError(f"导出工作簿缺少工作表：{sorted(missing_sheets)}")
        order_sheet = workbook[ORDER_SHEET_NAME]
        headers = [order_sheet.cell(1, column).value for column in range(1, len(ORDER_HEADERS) + 1)]
        if headers != ORDER_HEADERS:
            raise RuntimeError("订单总表表头不符合预期")
        logged = _existing_export_log(workbook, expected_job_id)
        missing_items = expected_item_ids - set(logged)
        if missing_items:
            raise RuntimeError(f"导出日志缺少订单商品：{sorted(missing_items)[:20]}")
    finally:
        workbook.close()
    cached_workbook = load_workbook(path, read_only=True, data_only=True, keep_links=True)
    try:
        order_sheet = cached_workbook[ORDER_SHEET_NAME]
        row_numbers = {row_number for row_number in logged.values() if row_number > 0}
        if row_numbers:
            for row_number, row in enumerate(
                order_sheet.iter_rows(
                    min_row=min(row_numbers),
                    max_row=max(row_numbers),
                    min_col=17,
                    max_col=24,
                    values_only=True,
                ),
                start=min(row_numbers),
            ):
                if row_number not in row_numbers:
                    continue
                for value in (row[0], row[1], row[2], row[6], row[7]):
                    if isinstance(value, str) and value.startswith("#"):
                        raise RuntimeError(f"订单总表第 {row_number} 行存在公式错误：{value}")
    finally:
        cached_workbook.close()
    pivot_count = _pivot_part_count(path)
    if pivot_count < minimum_pivots:
        raise RuntimeError(f"数据透视表结构丢失：预期至少 {minimum_pivots}，实际 {pivot_count}")
    return logged


def _recalculate_workbook(path: Path, settings: Settings) -> None:
    command = settings.order_follow_up_export_recalculate_command.strip()
    timeout = max(30, settings.order_follow_up_export_recalculate_timeout_seconds)
    if command:
        subprocess.run(command.format(file=str(path)), shell=True, check=True, timeout=timeout)
        return
    engine_name = settings.order_follow_up_export_recalculate_engine.strip().lower()
    if engine_name in {"", "none", "off", "disabled", "false"}:
        return
    if sys.platform.startswith("win"):
        _recalculate_with_wps(path, timeout)
        return
    if engine_name not in {"auto", "system", "default", "libreoffice", "soffice", "openoffice"}:
        raise RuntimeError(f"不支持的 order_follow_up_export.recalculate_engine：{engine_name}")
    _recalculate_with_libreoffice(path, timeout)


def _recalculate_with_libreoffice(path: Path, timeout: int) -> None:
    command = shutil.which("soffice") or shutil.which("libreoffice")
    if not command:
        macos_command = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
        command = str(macos_command) if macos_command.is_file() else ""
    if not command:
        raise RuntimeError("未找到 LibreOffice/soffice，无法重算 Order follow up 工作簿")
    with tempfile.TemporaryDirectory(prefix="caifuclaw_order_follow_up_recalc_") as temp_dir:
        output_dir = Path(temp_dir)
        result = subprocess.run(
            [
                command,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nofirststartwizard",
                "--nolockcheck",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(path),
            ],
            check=True,
            timeout=timeout,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        converted = output_dir / f"{path.stem}.xlsx"
        if not converted.is_file():
            candidates = list(output_dir.glob("*.xlsx"))
            if len(candidates) == 1:
                converted = candidates[0]
        if not converted.is_file():
            raise RuntimeError(f"LibreOffice 重算未生成文件：{result.stdout} {result.stderr}")
        shutil.copy2(converted, path)


def _recalculate_with_wps(path: Path, timeout: int) -> None:  # pragma: no cover - Windows COM only
    del timeout
    try:
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError("WPS/Excel COM 重算需要安装 pywin32") from exc
    app = None
    workbook = None
    last_error = None
    for prog_id in ("Ket.Application", "ET.Application", "WPS.Application", "Excel.Application"):
        try:
            app = win32com.client.DispatchEx(prog_id)
            break
        except Exception as exc:
            last_error = exc
    if app is None:
        raise RuntimeError(f"无法启动 WPS/Excel COM：{last_error}")
    try:
        app.Visible = False
        app.DisplayAlerts = False
        workbook = app.Workbooks.Open(str(path))
        try:
            app.CalculateFull()
        except Exception:
            app.Calculate()
        workbook.RefreshAll()
        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=True)
        app.Quit()


def _backup_workbook(paths: ExportPaths, job_id: int, now: datetime) -> Path:
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    backup = paths.backup_dir / f"{paths.workbook.stem}_{timestamp}_job{job_id}{paths.workbook.suffix}"
    shutil.copy2(paths.workbook, backup)
    return backup


def _cleanup_old_backups(paths: ExportPaths, settings: Settings, now: datetime) -> None:
    retention_days = settings.order_follow_up_export_backup_retention_days
    if retention_days <= 0:
        return
    threshold = now.timestamp() - retention_days * 86400
    pattern = f"{paths.workbook.stem}_*_job*{paths.workbook.suffix}"
    for path in paths.backup_dir.glob(pattern):
        try:
            if path.stat().st_mtime < threshold:
                path.unlink()
        except OSError:
            logger.warning("Failed to remove old order follow up backup: %s", path, exc_info=True)


def _ensure_canonical_workbook(paths: ExportPaths) -> None:
    _ensure_export_directories(paths)
    if paths.workbook.is_file():
        return
    if not paths.template.is_file():
        raise FileNotFoundError(f"Order follow up 模板不存在：{paths.template}")
    shutil.copy2(paths.template, paths.workbook)


def _write_main_workbook(job_id: int, items: list[dict], settings: Settings) -> WorkbookWriteResult:
    paths = _export_paths(settings)
    _ensure_canonical_workbook(paths)
    minimum_pivots = _pivot_part_count(paths.workbook)
    expected_ids = {int(item["order_item_id"]) for item in items if item["action"] != "skip"}

    current = load_workbook(paths.workbook, read_only=True, data_only=False, keep_links=True)
    try:
        already_logged = _existing_export_log(current, job_id)
    finally:
        current.close()
    pending = [item for item in items if item["action"] != "skip" and item["order_item_id"] not in already_logged]
    if not pending:
        return WorkbookWriteResult(paths.workbook, already_logged, 0, 0)

    now = _local_datetime(_utc_now()) or _utc_now()
    _backup_workbook(paths, job_id, now)
    temp_path = paths.sync_dir / f".{paths.workbook.stem}.job{job_id}.{os.getpid()}.tmp.xlsx"
    shutil.copy2(paths.workbook, temp_path)
    try:
        workbook = load_workbook(temp_path, data_only=False, keep_links=True)
        try:
            missing_sheets = REQUIRED_SHEETS - set(workbook.sheetnames)
            if missing_sheets:
                raise RuntimeError(f"模板缺少工作表：{sorted(missing_sheets)}")
            order_sheet = workbook[ORDER_SHEET_NAME]
            direct_sheet = workbook[DIRECT_PLAN_SHEET_NAME]
            direct_item_rows = _existing_direct_export_log(workbook)
            template_row = _last_data_row(order_sheet, (1, 5, 8), 2)
            next_order_row = template_row + 1
            item_rows = dict(already_logged)
            for item in pending:
                snapshot = item["snapshot"]
                if item["action"] == "update" and item.get("worksheet_row"):
                    row_number = int(item["worksheet_row"])
                    _write_order_row(order_sheet, row_number, snapshot, append=False, template_row=template_row)
                else:
                    row_number = next_order_row
                    next_order_row += 1
                    _write_order_row(order_sheet, row_number, snapshot, append=True, template_row=template_row)
                item_rows[int(item["order_item_id"])] = row_number

            direct_log_rows: list[list[object]] = []
            for item in pending:
                snapshot = item["snapshot"]
                if item["action"] != "update" or not snapshot.get("direct_plan_eligible"):
                    continue
                item_id = int(item["order_item_id"])
                direct_plan_row = direct_item_rows.get(item_id, 0)
                if direct_plan_row and _update_direct_row(direct_sheet, direct_plan_row, snapshot, now):
                    direct_log_rows.append(
                        [
                            f"{job_id}:{item_id}:update",
                            job_id,
                            item_id,
                            "update",
                            direct_plan_row,
                            _direct_group_key(snapshot),
                            0,
                            now,
                        ]
                    )

            groups = _direct_groups(pending)
            appended_direct_rows = _append_direct_rows(direct_sheet, groups, now)
            item_quantities = {
                int(item["order_item_id"]): _safe_int(item["snapshot"].get("quantity")) for item in pending
            }
            for group, direct_plan_row in zip(groups, appended_direct_rows, strict=True):
                for item_id, group_key in zip(group["item_ids"], group["group_keys"], strict=True):
                    direct_log_rows.append(
                        [
                            f"{job_id}:{item_id}:append",
                            job_id,
                            item_id,
                            "append",
                            direct_plan_row,
                            group_key,
                            item_quantities[item_id],
                            now,
                        ]
                    )
            direct_row_count = len(appended_direct_rows)
            log_sheet = _ensure_export_log_sheet(workbook)
            for item in pending:
                item_id = int(item["order_item_id"])
                token = f"{job_id}:{item_id}:{item['action']}"
                log_sheet.append([token, job_id, item_id, item["action"], item_rows[item_id], now])
            direct_log_sheet = _ensure_direct_export_log_sheet(workbook)
            for row in direct_log_rows:
                direct_log_sheet.append(row)
            _mark_full_recalculation(workbook)
            workbook.save(temp_path)
        finally:
            workbook.close()

        _validate_main_workbook(temp_path, job_id, expected_ids, minimum_pivots)
        _recalculate_workbook(temp_path, settings)
        logged = _validate_main_workbook(temp_path, job_id, expected_ids, minimum_pivots)
        os.replace(temp_path, paths.workbook)
        _cleanup_old_backups(paths, settings, now)
        return WorkbookWriteResult(paths.workbook, logged, len(pending), direct_row_count)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _purchase_plan_filename(job: OrderFollowUpExportJob, directory: Path) -> Path:
    created = _local_datetime(job.created_at) or datetime.now()
    base = directory / f"caifuclaw_pur_plan_{created.strftime('%Y%m%d_%H%M')}.xlsx"
    if not base.exists() or _purchase_plan_matches_job(base, job.id):
        return base
    return directory / f"caifuclaw_pur_plan_{created.strftime('%Y%m%d_%H%M')}_job{job.id}.xlsx"


def _purchase_plan_matches_job(path: Path, job_id: int) -> bool:
    if not path.is_file():
        return False
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if PURCHASE_META_SHEET_NAME not in workbook.sheetnames:
                return False
            return _safe_int(workbook[PURCHASE_META_SHEET_NAME]["B1"].value) == job_id
        finally:
            workbook.close()
    except Exception:
        return False


def _write_purchase_plan(job: OrderFollowUpExportJob, items: list[dict], settings: Settings) -> PurchasePlanWriteResult:
    paths = _export_paths(settings)
    _ensure_export_directories(paths)
    groups = _direct_groups(items)
    if not groups:
        return PurchasePlanWriteResult(None, 0)
    output_path = _purchase_plan_filename(job, paths.purchase_plan_dir)
    if _purchase_plan_matches_job(output_path, job.id):
        return PurchasePlanWriteResult(output_path, len(groups))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "直发货采购计划"
    headers = ["配货日", "产品名称", "采购数量（当日来单）", "库存数", "待采购数量", "导出时间", "映射状态"]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    exported_at = _local_datetime(_utc_now()) or _utc_now()
    for group in groups:
        quantity = _safe_int(group["quantity"])
        stock_qty = _safe_int(group["stock_qty"])
        sheet.append(
            [
                date.fromisoformat(group["allocation_date"]),
                group["product_name"],
                quantity,
                stock_qty,
                quantity - stock_qty,
                exported_at,
                "已映射" if group.get("mapping_status") == "mapped" else "未登记目录",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    widths = {"A": 13, "B": 42, "C": 22, "D": 12, "E": 15, "F": 20, "G": 15}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row_number, 6).number_format = "yyyy-mm-dd hh:mm:ss"

    meta = workbook.create_sheet(PURCHASE_META_SHEET_NAME)
    meta["A1"] = "job_id"
    meta["B1"] = job.id
    meta["A2"] = "scheduled_task_run_id"
    meta["B2"] = job.scheduled_task_run_id
    meta["A3"] = "order_item_ids"
    meta["B3"] = ",".join(str(item_id) for group in groups for item_id in group["item_ids"])
    meta.sheet_state = "veryHidden"
    _mark_full_recalculation(workbook)

    temp_path = paths.purchase_plan_dir / f".{output_path.stem}.{os.getpid()}.tmp.xlsx"
    try:
        workbook.save(temp_path)
        workbook.close()
        if not _purchase_plan_matches_job(temp_path, job.id):
            raise RuntimeError("采购计划文件元数据校验失败")
        os.replace(temp_path, output_path)
    finally:
        workbook.close()
        if temp_path.exists():
            temp_path.unlink()
    return PurchasePlanWriteResult(output_path, len(groups))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _artifact_values(path: Path | None, row_count: int) -> dict:
    if path is None:
        return {
            "status": JOB_SKIPPED,
            "file_path": "",
            "filename": "",
            "sha256": "",
            "size_bytes": 0,
            "row_count": row_count,
        }
    return {
        "status": JOB_SUCCESS,
        "file_path": str(path),
        "filename": path.name,
        "sha256": _sha256(path),
        "size_bytes": path.stat().st_size,
        "row_count": row_count,
    }


def _upsert_artifact(db: Session, job_id: int, artifact_type: str, values: dict, now: datetime) -> None:
    row = db.scalar(
        select(OrderFollowUpExportArtifact).where(
            OrderFollowUpExportArtifact.job_id == job_id,
            OrderFollowUpExportArtifact.artifact_type == artifact_type,
        )
    )
    if row is None:
        row = OrderFollowUpExportArtifact(job_id=job_id, artifact_type=artifact_type, created_at=now)
        db.add(row)
    row.status = values["status"]
    row.file_path = values["file_path"]
    row.filename = values["filename"]
    row.sha256 = values["sha256"]
    row.size_bytes = values["size_bytes"]
    row.row_count = values["row_count"]
    row.error_message = ""
    row.finished_at = now


def _job_item_payloads(db: Session, job_id: int) -> list[dict]:
    rows = db.scalars(
        select(OrderFollowUpExportItem)
        .where(OrderFollowUpExportItem.job_id == job_id)
        .order_by(asc(OrderFollowUpExportItem.id))
    ).all()
    return [
        {
            "id": row.id,
            "order_item_id": row.order_item_id,
            "action": row.action,
            "worksheet_row": row.worksheet_row,
            "snapshot": dict(row.snapshot_json or {}),
        }
        for row in rows
    ]


def confirmed_order_ids_for_follow_up_export(
    db: Session,
    order_ids: list[int],
    *,
    export_job_id: int | None = None,
) -> set[int]:
    """Return orders whose export items have canonical workbook rows."""
    unique_order_ids = sorted({int(order_id) for order_id in order_ids if int(order_id) > 0})
    if not unique_order_ids:
        return set()

    if export_job_id:
        item_states_by_order: dict[int, list[tuple[str, int | None]]] = defaultdict(list)
        for order_id, status, worksheet_row in db.execute(
            select(
                OrderFollowUpExportItem.order_id,
                OrderFollowUpExportItem.status,
                OrderFollowUpExportItem.worksheet_row,
            ).where(
                OrderFollowUpExportItem.job_id == export_job_id,
                OrderFollowUpExportItem.order_id.in_(unique_order_ids),
            )
        ).all():
            item_states_by_order[int(order_id)].append((str(status or ""), worksheet_row))
        return {
            order_id
            for order_id, item_states in item_states_by_order.items()
            if item_states
            and all(status in {ITEM_SUCCESS, ITEM_SKIPPED} and worksheet_row is not None for status, worksheet_row in item_states)
        }

    item_ids_by_order: dict[int, set[int]] = defaultdict(set)
    for order_id, item_id in db.execute(
        select(OrderItem.order_id, OrderItem.id).where(OrderItem.order_id.in_(unique_order_ids))
    ).all():
        item_ids_by_order[int(order_id)].add(int(item_id))

    confirmed_item_ids_by_order: dict[int, set[int]] = defaultdict(set)
    for order_id, item_id in db.execute(
        select(OrderFollowUpExportItem.order_id, OrderFollowUpExportItem.order_item_id).where(
            OrderFollowUpExportItem.order_id.in_(unique_order_ids),
            OrderFollowUpExportItem.status.in_((ITEM_SUCCESS, ITEM_SKIPPED)),
            OrderFollowUpExportItem.worksheet_row.is_not(None),
        )
    ).all():
        confirmed_item_ids_by_order[int(order_id)].add(int(item_id))

    return {
        order_id
        for order_id, item_ids in item_ids_by_order.items()
        if item_ids and item_ids.issubset(confirmed_item_ids_by_order.get(order_id, set()))
    }


def confirmed_joom_fbj_order_ids_for_follow_up_export(
    db: Session,
    order_ids: list[int],
    *,
    export_job_id: int | None = None,
) -> set[int]:
    """Backward-compatible FBJ alias for the canonical export check."""
    return confirmed_order_ids_for_follow_up_export(db, order_ids, export_job_id=export_job_id)


def mark_joom_fbj_orders_shipped_after_follow_up_export(
    db: Session,
    order_ids: list[int],
    *,
    export_job_id: int | None = None,
    operated_at: datetime | None = None,
) -> list[Order]:
    """Mark only fully registered FBJ orders shipped after the workbook succeeds."""
    confirmed_ids = confirmed_order_ids_for_follow_up_export(db, order_ids, export_job_id=export_job_id)
    if not confirmed_ids:
        return []

    rows = [
        row
        for row in db.scalars(select(Order).where(Order.id.in_(confirmed_ids))).all()
        if order_is_joom_fbj_warehouse(row)
        and row.biz_status == "待处理"
    ]
    if not rows:
        return []

    now = operated_at or _utc_now()
    status_before = {row.id: row.biz_status or "" for row in rows}
    for row in rows:
        row.biz_status = "已发货"
        row.local_status = "shipped"
        row.shipped_at = row.shipped_at or now
        row.marked_shipped_at = row.marked_shipped_at or now
        row.error_message = ""
        row.updated_at = now

    add_order_operation_logs(
        db,
        rows,
        operation_type="fbj_follow_up_export_shipped",
        operation_attribute="FBJ登记跟进表后发货",
        description=lambda row: (
            f"Joom FBJ订单 {_export_order_number(row)} 已登记至Order follow up，"
            f"状态：{status_before[row.id]} -> 已发货；不获取平台面单、不打印、不生成采购单"
        ),
        operator=SYSTEM_OPERATOR,
        source=ORDER_LOG_SYSTEM_SOURCE,
        operated_at=now,
        event_key=lambda row: f"fbj_follow_up_export_shipped:{export_job_id or 0}:{row.id}",
        extra=lambda row: {
            "export_job_id": export_job_id,
            "status_before": status_before[row.id],
            "status_after": "已发货",
        },
    )
    return rows


def mark_bsi_orders_shipped_after_follow_up_export(
    db: Session,
    order_ids: list[int],
    *,
    export_job_id: int | None = None,
    operated_at: datetime | None = None,
) -> list[Order]:
    """Close BSI draft orders only after their follow-up rows are registered."""
    confirmed_ids = confirmed_order_ids_for_follow_up_export(db, order_ids, export_job_id=export_job_id)
    if not confirmed_ids:
        return []

    rows = [
        row
        for row in db.scalars(select(Order).where(Order.id.in_(confirmed_ids))).all()
        if order_has_bsi_draft(row)
        and row.biz_status == "待处理"
    ]
    if not rows:
        return []

    now = operated_at or _utc_now()
    status_before = {row.id: row.biz_status or "" for row in rows}
    for row in rows:
        row.biz_status = "已发货"
        row.local_status = "shipped"
        row.shipped_at = row.shipped_at or now
        row.marked_shipped_at = row.marked_shipped_at or now
        row.error_message = ""
        row.updated_at = now

    add_order_operation_logs(
        db,
        rows,
        operation_type="bsi_follow_up_export_shipped",
        operation_attribute="BSI登记跟进表后发货",
        description=lambda row: (
            f"BSI订单 {_export_order_number(row)} 已登记至Order follow up，"
            f"状态：{status_before[row.id]} -> 已发货；不获取平台面单、不打印、不生成采购单"
        ),
        operator=SYSTEM_OPERATOR,
        source=ORDER_LOG_SYSTEM_SOURCE,
        operated_at=now,
        event_key=lambda row: f"bsi_follow_up_export_shipped:{export_job_id or 0}:{row.id}",
        extra=lambda row: {
            "export_job_id": export_job_id,
            "bsi_order_no": row.bsi_order_no,
            "status_before": status_before[row.id],
            "status_after": "已发货",
        },
    )
    return rows


def _finalize_success(
    job_id: int,
    workbook_result: WorkbookWriteResult,
    purchase_result: PurchasePlanWriteResult,
) -> None:
    db = SessionLocal()
    try:
        now = _utc_now()
        job = db.get(OrderFollowUpExportJob, job_id)
        if not job:
            return
        items = db.scalars(
            select(OrderFollowUpExportItem).where(OrderFollowUpExportItem.job_id == job_id)
        ).all()
        for item in items:
            if item.action == "skip":
                item.status = ITEM_SKIPPED
            else:
                item.status = ITEM_SUCCESS
                item.worksheet_row = workbook_result.item_rows.get(item.order_item_id, item.worksheet_row)
            item.error_message = ""
            item.exported_at = now
            item.updated_at = now
        db.flush()
        fbj_shipped_rows = mark_joom_fbj_orders_shipped_after_follow_up_export(
            db,
            [int(item.order_id) for item in items],
            export_job_id=job_id,
            operated_at=now,
        )
        bsi_shipped_rows = mark_bsi_orders_shipped_after_follow_up_export(
            db,
            [int(item.order_id) for item in items],
            export_job_id=job_id,
            operated_at=now,
        )
        workbook_values = _artifact_values(workbook_result.file_path, workbook_result.changed_item_count)
        purchase_values = _artifact_values(purchase_result.file_path, purchase_result.row_count)
        _upsert_artifact(db, job_id, ARTIFACT_WORKBOOK, workbook_values, now)
        _upsert_artifact(db, job_id, ARTIFACT_PURCHASE_PLAN, purchase_values, now)
        job.status = JOB_SUCCESS
        job.error_message = ""
        job.claimed_by = ""
        job.claimed_at = None
        job.lease_until = None
        job.heartbeat_at = None
        job.next_retry_at = None
        job.finished_at = now
        job.updated_at = now
        job.stats_json = {
            "item_count": len(items),
            "changed_item_count": workbook_result.changed_item_count,
            "skipped_item_count": sum(1 for item in items if item.action == "skip"),
            "direct_row_count": workbook_result.direct_row_count,
            "purchase_plan_row_count": purchase_result.row_count,
            "missing_mapping_count": sum(1 for item in items if item.mapping_status == "missing"),
            "fbj_shipped_count": len(fbj_shipped_rows),
            "bsi_shipped_count": len(bsi_shipped_rows),
        }
        db.commit()
    finally:
        db.close()


def _finalize_empty(job_id: int) -> None:
    db = SessionLocal()
    try:
        now = _utc_now()
        job = db.get(OrderFollowUpExportJob, job_id)
        if not job:
            return
        job.status = JOB_SKIPPED
        job.error_message = ""
        job.claimed_by = ""
        job.claimed_at = None
        job.lease_until = None
        job.heartbeat_at = None
        job.next_retry_at = None
        job.finished_at = now
        job.updated_at = now
        job.stats_json = {"item_count": 0, "reason": "no_processed_order_items"}
        db.commit()
    finally:
        db.close()


def _finalize_failure(job_id: int, exc: Exception) -> None:
    db = SessionLocal()
    try:
        settings = get_settings()
        now = _utc_now()
        job = db.get(OrderFollowUpExportJob, job_id)
        if not job:
            return
        message = str(exc)[:10000]
        job.error_message = message
        job.claimed_by = ""
        job.claimed_at = None
        job.lease_until = None
        job.heartbeat_at = None
        job.updated_at = now
        if job.attempt_count >= job.max_attempts:
            job.status = JOB_FAILED
            job.next_retry_at = None
            job.finished_at = now
        else:
            job.status = JOB_RETRY_WAIT
            job.next_retry_at = now + timedelta(
                seconds=max(10, settings.order_follow_up_export_retry_delay_seconds)
            )
        items = db.scalars(
            select(OrderFollowUpExportItem).where(
                OrderFollowUpExportItem.job_id == job_id,
                OrderFollowUpExportItem.status == ITEM_PENDING,
            )
        ).all()
        for item in items:
            item.error_message = message
            item.updated_at = now
        db.commit()
    finally:
        db.close()


def _try_workbook_advisory_lock(connection) -> bool:
    if connection.dialect.name != "postgresql":
        return True
    return bool(
        connection.execute(
            text("SELECT pg_try_advisory_lock(:key)"),
            {"key": WORKBOOK_LOCK_KEY},
        ).scalar()
    )


def _release_workbook_advisory_lock(connection) -> None:
    if connection.dialect.name == "postgresql":
        connection.execute(
            text("SELECT pg_advisory_unlock(:key)"),
            {"key": WORKBOOK_LOCK_KEY},
        )


def _process_claimed_job(job_id: int) -> None:
    settings = get_settings()
    db = SessionLocal()
    try:
        job = db.get(OrderFollowUpExportJob, job_id)
        if not job or job.status != JOB_RUNNING:
            return
        prepared = _prepare_job_items(db, job, settings)
        if not prepared:
            _finalize_empty(job_id)
            return
        payloads = _job_item_payloads(db, job_id)
        job_snapshot = OrderFollowUpExportJob(
            id=job.id,
            scheduled_task_run_id=job.scheduled_task_run_id,
            created_at=job.created_at,
        )
    finally:
        db.close()

    with _process_lock:
        with engine.connect() as connection:
            if not _try_workbook_advisory_lock(connection):
                raise ExportLockBusy("Order follow up 工作簿正在由其他执行器处理")
            try:
                workbook_result = _write_main_workbook(job_id, payloads, settings)
                purchase_result = _write_purchase_plan(job_snapshot, payloads, settings)
            finally:
                _release_workbook_advisory_lock(connection)
    _finalize_success(job_id, workbook_result, purchase_result)


def process_next_order_follow_up_export_job() -> bool:
    settings = get_settings()
    if not settings.order_follow_up_export_enabled:
        return False
    reconcile_order_follow_up_export_jobs()
    db = SessionLocal()
    try:
        job = _claim_next_job(db)
        job_id = job.id if job else None
    finally:
        db.close()
    if not job_id:
        return False
    try:
        _process_claimed_job(job_id)
    except Exception as exc:
        logger.exception("Order follow up export failed job_id=%s", job_id)
        _finalize_failure(job_id, exc)
    return True


async def _worker_loop() -> None:
    logger.info("Order follow up export worker started")
    try:
        while True:
            try:
                processed = await asyncio.to_thread(process_next_order_follow_up_export_job)
                settings = get_settings()
                delay = (
                    settings.order_follow_up_export_worker_poll_seconds
                    if processed
                    else settings.order_follow_up_export_worker_idle_seconds
                )
            except asyncio.CancelledError:
                raise
            except Exception:
                logger.exception("Order follow up export worker iteration failed")
                delay = 30
            await asyncio.sleep(max(1, delay))
    except asyncio.CancelledError:
        logger.info("Order follow up export worker stopped")
        raise


def start_order_follow_up_export_worker() -> None:
    global _worker_task
    if not get_settings().order_follow_up_export_enabled:
        return
    if _worker_task and not _worker_task.done():
        return
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        logger.debug("Order follow up export worker start skipped: no running event loop")
        return
    _worker_task = loop.create_task(_worker_loop(), name="order-follow-up-export-worker")


async def stop_order_follow_up_export_worker() -> None:
    global _worker_task
    if not _worker_task:
        return
    task = _worker_task
    _worker_task = None
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass
