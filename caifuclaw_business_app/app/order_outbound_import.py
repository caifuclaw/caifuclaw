from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .database import SessionLocal
from .models import Order, Shipment
from .order_operation_logs import ORDER_LOG_SYSTEM_SOURCE, SYSTEM_OPERATOR, add_order_operation_log


SHEET_NAME = "订单出库"
TRACKING_HEADER = "物流单号"
SHIPPED_DATE_HEADER = "发货日期"
ORDER_STATUS_PICKING = "配货中"
ORDER_STATUS_SHIPPED = "已发货"
LOCAL_STATUS_SHIPPED = "shipped"
LOCAL_TIME_OFFSET = timedelta(hours=8)
OPERATION_TYPE = "order_follow_up_outbound_import"
OPERATION_ATTRIBUTE = "订单出库导入"


@dataclass(frozen=True)
class OutboundEntry:
    row_number: int
    tracking_number: str
    shipped_at: datetime


@dataclass(frozen=True)
class OutboundReadIssue:
    row_number: int
    tracking_number: str
    reason: str
    detail: str = ""


@dataclass
class OutboundImportStats:
    mode: str
    rows_seen: int = 0
    unique_tracking_numbers: int = 0
    updated_orders: int = 0
    would_update_orders: int = 0
    skipped_not_found: int = 0
    skipped_not_picking: int = 0
    skipped_invalid_rows: int = 0
    skipped_date_conflicts: int = 0
    log: str = ""


def clean_tracking_number(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return "".join(str(value).strip().split())


def tracking_lookup_key(value: object) -> str:
    return clean_tracking_number(value).lower()


def parse_excel_shipped_at(value: object, *, epoch: datetime) -> datetime | None:
    parsed: datetime | None = None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch)
        except (TypeError, ValueError, OverflowError):
            return None
        if isinstance(converted, datetime):
            parsed = converted
        elif isinstance(converted, date):
            parsed = datetime.combine(converted, time.min)
    elif value is not None:
        text_value = str(value).strip()
        if not text_value:
            return None
        normalized = text_value.replace("/", "-").replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            for date_format in ("%Y-%m-%d", "%Y%m%d"):
                try:
                    parsed = datetime.strptime(normalized, date_format)
                    break
                except ValueError:
                    continue
    if parsed is None or parsed.year < 2000:
        return None
    if parsed.tzinfo is not None:
        return parsed.astimezone(timezone.utc).replace(tzinfo=None, microsecond=0)
    return (parsed - LOCAL_TIME_OFFSET).replace(microsecond=0)


def read_outbound_entries(path: Path) -> tuple[list[OutboundEntry], list[OutboundReadIssue], int]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Excel 缺少页签: {SHEET_NAME}")
        worksheet = workbook[SHEET_NAME]
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1)), start=1)
            if cell.value is not None
        }
        missing_headers = [header for header in (TRACKING_HEADER, SHIPPED_DATE_HEADER) if header not in headers]
        if missing_headers:
            raise ValueError(f"{SHEET_NAME} 缺少字段: {', '.join(missing_headers)}")

        tracking_column = headers[TRACKING_HEADER]
        date_column = headers[SHIPPED_DATE_HEADER]
        rows_seen = 0
        values: dict[str, list[OutboundEntry]] = defaultdict(list)
        issues: list[OutboundReadIssue] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=max(tracking_column, date_column), values_only=True),
            start=2,
        ):
            tracking_number = clean_tracking_number(row[tracking_column - 1])
            shipped_value = row[date_column - 1]
            if not tracking_number and shipped_value in (None, ""):
                continue
            rows_seen += 1
            if not tracking_number:
                issues.append(OutboundReadIssue(row_number, "", "missing_tracking_number"))
                continue
            shipped_at = parse_excel_shipped_at(shipped_value, epoch=workbook.epoch)
            if shipped_at is None:
                issues.append(
                    OutboundReadIssue(row_number, tracking_number, "invalid_shipped_date", str(shipped_value or ""))
                )
                continue
            values[tracking_lookup_key(tracking_number)].append(
                OutboundEntry(row_number, tracking_number, shipped_at)
            )

        entries: list[OutboundEntry] = []
        for grouped_entries in values.values():
            shipped_dates = {entry.shipped_at for entry in grouped_entries}
            if len(shipped_dates) > 1:
                detail = " | ".join(sorted(value.isoformat() for value in shipped_dates))
                issues.append(
                    OutboundReadIssue(
                        grouped_entries[0].row_number,
                        grouped_entries[0].tracking_number,
                        "shipped_date_conflict",
                        detail,
                    )
                )
                continue
            entries.append(grouped_entries[-1])
        return entries, issues, rows_seen
    finally:
        workbook.close()


def _raw_tracking_values(order: Order) -> list[str]:
    payload = order.raw_payload if isinstance(order.raw_payload, dict) else {}
    values = [
        payload.get("tracking_number"),
        payload.get("trackingNumber"),
        payload.get("shipment_tracking_number"),
    ]
    for container_name in ("shipment", "shipping", "logistics"):
        container = payload.get(container_name)
        if isinstance(container, dict):
            values.extend((container.get("tracking_number"), container.get("trackingNumber")))
    tracking = payload.get("tracking")
    if isinstance(tracking, dict):
        values.append(tracking.get("number"))
    return [tracking_lookup_key(value) for value in values if tracking_lookup_key(value)]


def _matching_orders(db: Session, tracking_keys: set[str]) -> dict[str, list[Order]]:
    result: dict[str, dict[int, Order]] = defaultdict(dict)
    if not tracking_keys:
        return {}

    direct_rows = db.scalars(
        select(Order).where(func.lower(func.trim(Order.shipment_tracking_number)).in_(tracking_keys))
    ).all()
    for order in direct_rows:
        result[tracking_lookup_key(order.shipment_tracking_number)][order.id] = order

    shipment_rows = db.execute(
        select(Shipment.tracking_number, Order)
        .join(Order, Order.id == Shipment.order_id)
        .where(func.lower(func.trim(Shipment.tracking_number)).in_(tracking_keys))
    ).all()
    for tracking_number, order in shipment_rows:
        result[tracking_lookup_key(tracking_number)][order.id] = order

    picking_rows = db.scalars(select(Order).where(Order.biz_status == ORDER_STATUS_PICKING)).all()
    for order in picking_rows:
        for key in _raw_tracking_values(order):
            if key in tracking_keys:
                result[key][order.id] = order
    return {key: list(rows.values()) for key, rows in result.items()}


def _write_log(
    path: Path,
    result_rows: list[dict],
    issues: list[OutboundReadIssue],
    stats: OutboundImportStats,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "type",
                "row",
                "tracking_number",
                "shipped_at",
                "order_id",
                "order_number",
                "status_before",
                "reason",
                "detail",
            ]
        )
        for row in result_rows:
            writer.writerow(
                [
                    row["type"],
                    row["row"],
                    row["tracking_number"],
                    row["shipped_at"],
                    row.get("order_id", ""),
                    row.get("order_number", ""),
                    row.get("status_before", ""),
                    row.get("reason", ""),
                    row.get("detail", ""),
                ]
            )
        for issue in issues:
            writer.writerow(
                ["skipped", issue.row_number, issue.tracking_number, "", "", "", "", issue.reason, issue.detail]
            )
        writer.writerow([])
        for key, value in asdict(stats).items():
            writer.writerow(["summary", key, value])


def _order_number(order: Order) -> str:
    return order.platform_order_no or order.posting_number or order.platform_order_id or str(order.id)


def run_order_outbound_import(
    path: Path,
    *,
    apply: bool,
    log_path: Path,
    session_factory: Callable = SessionLocal,
) -> OutboundImportStats:
    entries, issues, rows_seen = read_outbound_entries(path)
    stats = OutboundImportStats(
        mode="apply" if apply else "dry-run",
        rows_seen=rows_seen,
        unique_tracking_numbers=len(entries),
        skipped_invalid_rows=sum(issue.reason != "shipped_date_conflict" for issue in issues),
        skipped_date_conflicts=sum(issue.reason == "shipped_date_conflict" for issue in issues),
        log=str(log_path),
    )
    result_rows: list[dict] = []
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    with session_factory() as db:
        orders_by_tracking = _matching_orders(db, {tracking_lookup_key(entry.tracking_number) for entry in entries})
        for entry in entries:
            matched_orders = orders_by_tracking.get(tracking_lookup_key(entry.tracking_number), [])
            if not matched_orders:
                stats.skipped_not_found += 1
                result_rows.append(
                    {
                        "type": "skipped",
                        "row": entry.row_number,
                        "tracking_number": entry.tracking_number,
                        "shipped_at": entry.shipped_at.isoformat(),
                        "reason": "order_not_found",
                    }
                )
                continue
            picking_orders = [order for order in matched_orders if order.biz_status == ORDER_STATUS_PICKING]
            if not picking_orders:
                stats.skipped_not_picking += 1
                statuses = "; ".join(sorted({order.biz_status or "-" for order in matched_orders}))
                result_rows.append(
                    {
                        "type": "skipped",
                        "row": entry.row_number,
                        "tracking_number": entry.tracking_number,
                        "shipped_at": entry.shipped_at.isoformat(),
                        "status_before": statuses,
                        "reason": "order_not_picking",
                    }
                )
                continue

            for order in picking_orders:
                stats.would_update_orders += 1
                result_rows.append(
                    {
                        "type": "updated" if apply else "would_update",
                        "row": entry.row_number,
                        "tracking_number": entry.tracking_number,
                        "shipped_at": entry.shipped_at.isoformat(),
                        "order_id": order.id,
                        "order_number": _order_number(order),
                        "status_before": order.biz_status,
                    }
                )
                if not apply:
                    continue
                order.biz_status = ORDER_STATUS_SHIPPED
                order.local_status = LOCAL_STATUS_SHIPPED
                order.shipped_at = entry.shipped_at
                order.marked_shipped_at = entry.shipped_at
                order.updated_at = now
                add_order_operation_log(
                    db,
                    order_id=order.id,
                    operation_type=OPERATION_TYPE,
                    operation_attribute=OPERATION_ATTRIBUTE,
                    description=(
                        f"Order follow up 订单出库导入：物流单号 {entry.tracking_number}，"
                        f"状态：{ORDER_STATUS_PICKING} -> {ORDER_STATUS_SHIPPED}"
                    ),
                    operator=SYSTEM_OPERATOR,
                    source=ORDER_LOG_SYSTEM_SOURCE,
                    operated_at=now,
                    event_key=f"{OPERATION_TYPE}:{order.id}:{entry.shipped_at.isoformat()}",
                    extra={
                        "tracking_number": entry.tracking_number,
                        "excel_row": entry.row_number,
                        "shipped_at": entry.shipped_at.isoformat(),
                        "status_before": ORDER_STATUS_PICKING,
                        "status_after": ORDER_STATUS_SHIPPED,
                    },
                )
                stats.updated_orders += 1
        if apply:
            db.commit()
        else:
            db.rollback()

    _write_log(log_path, result_rows, issues, stats)
    return stats
