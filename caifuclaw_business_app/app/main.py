# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

import asyncio
import base64
import hashlib
import hmac
import io
import json
import logging
import platform
import re
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import uuid
import zipfile
import websockets
from copy import deepcopy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date, datetime, time, timedelta, timezone
from functools import lru_cache
from pathlib import Path
from time import perf_counter, sleep
from typing import Any
from urllib.parse import quote, urljoin, urlparse

import httpx
from apscheduler.triggers.cron import CronTrigger
from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from jwt.exceptions import PyJWTError as JWTError
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.middleware.gzip import GZipMiddleware
from websockets.exceptions import ConnectionClosed
from sqlalchemy import Integer, and_, asc, bindparam, case, desc, exists, func, or_, select, text, union_all, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload, defer, aliased, selectinload

from .config_loader import optional as config_optional
from .api.dependencies import create_access_dependencies, create_internal_service_dependency
from .api.router import create_api_router
from .api.routes.dashboard import DashboardRouteServices
from .api.routes.sync_settings import SyncSettingsRouteServices
from .connector_client import ConnectorRuntimeClient
from .credential_manager import get_credential_manager
from .bsi_sdms import BSI_CARRIER_CODE, verify_bsi_authorization
from .country_mapping import country_name_cn, country_name_to_code
from .database import Base, SessionLocal, engine, get_db
from .api_logger import log_api_call
from .image_storage import ImageStorageError, load_oss_config, upload_file_to_oss
from .ai_image_service import (
    AiImageError,
    AiImageUpstreamError,
    ImageApiConfig,
    build_ai_split_messages,
    call_image_api,
    merge_images,
    output_suffix,
    parse_ai_split_regions,
    refine_ai_split_regions,
    split_image,
    validate_image_file,
    write_api_images,
)
from .models import (
    ApiRequestLog,
    DashboardPlatformSetting,
    EmailSmtpSetting,
    ExchangeRate,
    ExchangeRateCurrencySetting,
    LabelFile,
    LocalUser,
    LogisticsAuthorization,
    LogisticsMatchRule,
    ModelEndpoint,
    ModelSetting,
    Order,
    OrderItem,
    OrderOperationLog,
    OrderRiskHandling,
    OutboundScanRecord,
    PlatformPrintSetting,
    PlatformSetting,
    ShippingDeadlineSetting,
    PlatformAccount,
    Role,
    RoleMenuPermission,
    ScheduledTask,
    SchedulerHeartbeat,
    ScheduledTaskRun,
    ScheduledTaskRunOrder,
    ScheduledTaskRunStep,
    Shipment,
    SyncAccountState,
    SyncAuditLog,
    SyncSetting,
    TranslationProviderSetting,
    UserRole,
    UserTablePreference,
    WeComRobotSetting,
    generate_internal_order_no,
)
from .openclaw_browser_relay import (
    DEFAULT_OPENCLAW_CONFIG_PATH,
    DEFAULT_OPENCLAW_RELAY_STATUS_URL,
    OPENCLAW_RELAY_AUTH_HEADER,
    OPENCLAW_RELAY_HOST,
    OPENCLAW_RELAY_PORT,
    OPENCLAW_RELAY_STATUS_TIMEOUT_SECONDS,
    derive_openclaw_relay_token as _derive_openclaw_relay_token,
    read_openclaw_gateway_token as _read_openclaw_gateway_token,
)
from .product_matching import mapping_choice_for_order_item
from .deadline_settings import (
    BASE_DATE_PAYMENT_AT,
    BASE_DATE_PLATFORM_CREATED,
    BASE_DATE_SHIPPING_DEADLINE,
    OTHER_PLATFORM,
    backfill_order_dispatch_deadlines,
    calculate_dispatch_deadline,
    canonical_deadline_platform,
    load_shipping_deadline_settings,
    normalize_base_date_field,
    seed_default_shipping_deadline_settings,
)
from .dashboard_settings import (
    OTHER_DASHBOARD_PLATFORM,
    canonical_dashboard_platform,
    dashboard_receipt_rate_for,
    load_dashboard_receipt_rates,
    seed_default_dashboard_platform_settings,
)
from .exchange_rates import sync_exchange_rates_from_provider
from .oauth_authorization import (
    build_authorize_url,
    callback_html,
    create_authorization_session,
    exchange_authorization_code,
    get_authorization_session,
    mark_authorization_failed,
    mark_authorization_success,
)
from .platform_product_catalog import (
    CATALOG_MAIN_IMAGE_CONTENT_TYPES,
    CATALOG_SUPPORTED_PLATFORMS,
    PlatformProductCatalogItem,
    PlatformProductPricingRule,
    _normalize_platform as _normalize_catalog_platform,
    catalog_main_image_display_url,
    catalog_main_image_file_path,
    map_catalog_item,
    recalculate_catalog,
    synchronize_platform_catalog,
)
from .product_models import (
    Product,
    ProductInventory,
    ProductShopMapping,
    PurchaseOrder,
    PurchaseOrderEditLock,
    PurchaseOrderItem,
    PurchaseOrderLog,
    PurchaseOrderSource,
)
from .email_service import (
    apply_provider_preset,
    encrypt_auth_code,
    get_email_setting,
    list_email_provider_presets,
    notification_recipient_values,
    parse_recipients,
    send_email,
)
from .scheduler import reload_jobs, run_scheduled_task_now, start_scheduler, stop_scheduler
from .sync_runtime import audit_sync_event, sync_health_snapshot
from .wecom_service import (
    DEFAULT_WECOM_PROMPT,
    decrypt_wecom_webhook_url,
    dumps_int_list,
    dumps_string_list,
    encrypt_wecom_webhook_url,
    get_wecom_robot_setting,
    loads_int_list,
    loads_string_list,
    mask_wecom_webhook_url,
    mentioned_mobile_list_from_user_ids,
    normalize_int_list,
    normalize_string_list,
    send_wecom_robot_test_message,
    validate_wecom_webhook_url,
)
from .translation_settings import (
    DEFAULT_TRANSLATION_PROVIDER,
    build_translation_client_from_setting,
    decrypt_translation_secret_key,
    dumps_translation_provider_options,
    encrypt_translation_secret_key,
    get_translation_provider_setting,
    list_translation_language_presets,
    list_translation_provider_presets,
    mask_translation_secret_key,
    normalize_translation_provider,
    translation_provider_endpoint,
    translation_provider_name,
    translation_provider_options_dict,
)
from .purchase_order_notification import enqueue_purchase_order_wecom_notification
from .connectors.base import NormalizedOrder
from .connectors.base import ShipmentResult
from .label_platforms import label_shipment_id_for_order
from .label_storage import is_real_label_pdf, save_label_pdf
from .label_tracking import apply_label_result_tracking, clean_tracking_number
from .wanbang import (
    WANBANG_CARRIER_NAME,
    fetch_wanbang_label_for_order,
    order_uses_wanbang,
    run_wanbang_test_flow_for_order,
)
from .logistics_rules import (
    LOGISTICS_MATCH_STATUS_MANUAL,
    apply_logistics_match_result,
    apply_logistics_rules,
    apply_manual_logistics_channel,
    load_enabled_logistics_rules,
    match_logistics_rule,
    normalize_country_codes,
    normalize_platform_code,
    normalize_shop_names,
    split_logistics_rule_eligible_orders,
)
from .order_types import (
    infer_fulfillment_type,
    infer_is_overseas_warehouse,
    order_is_joom_fbj_warehouse,
    order_is_joom_offline_shipping,
    order_is_joom_overseas_warehouse,
    order_is_logistics_label_exempt,
    order_is_overseas_warehouse,
    wildberries_payload_country_code,
)
from .pdf_tools import merge_pdf_parts, orient_pdf_bytes
from .chinese_label_pdf import (
    ChineseLabelRow,
    generate_chinese_label_pdf,
    resolve_chinese_label_deadline,
)
from .sync_engine import _connector_for_account, _platform_endpoint, backfill_wanbang_tracking_to_platform
from .oauth_tokens import ensure_access_token
from .order_operation_logs import (
    ORDER_LOG_HISTORY_SOURCE,
    ORDER_LOG_MANUAL_SOURCE,
    ORDER_LOG_SYSTEM_SOURCE,
    SYSTEM_OPERATOR,
    add_order_operation_log,
    add_order_operation_logs,
    operator_name,
    safe_exception_message,
)
from .order_follow_up_export import (
    start_order_follow_up_export_worker,
    stop_order_follow_up_export_worker,
)
from .print_options import (
    PRINT_ORIENTATION_AUTO,
    PRINT_ORIENTATION_LABELS,
    PRINT_PLATFORM_CHINESE_LABEL,
    PRINT_PLATFORM_CHINESE_LABEL_NAME,
    label_orientation_for_platform,
    label_size_mm_for_platform,
    is_valid_print_orientation,
    normalize_print_orientation,
)
from .printer_identity import PrinterIdentity, printer_fingerprint
from .schemas import (
    ApiRequestLogDto,
    ApiRequestLogListResponse,
    ApiRequestLogSummaryDto,
    ApiRequestLogSummaryListResponse,
    CredentialsRequest,
    DashboardAnalyticsResponse,
    DashboardDailySalesDto,
    DashboardHotSkuDto,
    DashboardMonthlySalesDto,
    DashboardMtdComparisonDto,
    DashboardOverviewResponse,
    DashboardPlatformSettingDto,
    DashboardRiskBucketDto,
    DashboardRiskResponse,
    DashboardRiskShopDto,
    DashboardRiskSkuDto,
    DashboardSalesResponse,
    DashboardSettingsResponse,
    DashboardSettingsUpdateRequest,
    DashboardSettingsUpdateResponse,
    DashboardShopSalesDto,
    DashboardSkuResponse,
    OperationsCustomerComplaintDto,
    OperationsDailyOrderPointDto,
    OperationsDailyReportResponse,
    OperationsDailyShopDto,
    OperationsFulfillmentRiskDto,
    ManualSyncRequest,
    OrderBatchRequest,
    OrderBatchResponse,
    OrderDetailDto,
    OrderDetailItemDto,
    OrderExportRequest,
    OrderLogisticsChannelBatchRequest,
    OrderDto,
    OrderListResponse,
    OrderOperationLogChangeDto,
    OrderOperationLogDto,
    OrderOperationLogListResponse,
    OrderRiskHandlingRequest,
    OrderSearchRequest,
    OrderSearchSummary,
    OrderWanbangTestItemDto,
    OrderWanbangTestResponse,
    OrderSummaryDto,
    OrderSummaryResponse,
    OutboundScanListResponse,
    OutboundScanRecordDto,
    OutboundScanRequest,
    OutboundScanResponse,
    OutboundScanStatsResponse,
    InventoryDto,
    InventoryListResponse,
    InventoryUpsertRequest,
    ProductBatchRequest,
    ProductDto,
    ProductListResponse,
    ProductShopDto,
    ProductUpsertRequest,
    PurchaseOrderDetailDto,
    PurchaseOrderEditLockDto,
    PurchaseOrderEditLockRequest,
    PurchaseDetailDto,
    PurchaseDetailListResponse,
    PlatformPrintSettingDto,
    PlatformPrintSettingUpsertRequest,
    PlatformSettingDto,
    PlatformSettingToggleRequest,
    PrinterMonitorRequest,
    PrinterMonitorResultDto,
    PrinterDto,
    ShippingDeadlineSettingDto,
    ShippingDeadlineSettingsUpdateRequest,
    ShippingDeadlineSettingsUpdateResponse,
    EmailProviderDto,
    EmailSmtpSettingDto,
    EmailSmtpSettingUpdateRequest,
    EmailTestRequest,
    AiImageAssetDto,
    AiImageBatchDownloadRequest,
    AiImageProcessResponse,
    ModelConnectionTestResponse,
    ModelEndpointDto,
    ModelEndpointUpsertRequest,
    ModelSettingDto,
    ModelSettingUpsertRequest,
    WeComRobotSettingDto,
    WeComRobotTestRequest,
    WeComRobotTestResponse,
    WeComRobotSettingUpdateRequest,
    TranslationProviderOptionDto,
    TranslationLanguageOptionDto,
    TranslationProviderSettingDto,
    TranslationProviderSettingUpdateRequest,
    TranslationProviderTestRequest,
    TranslationProviderTestResponse,
    TextTranslationRequest,
    TextTranslationResponse,
    ExchangeRateDto,
    ExchangeRateListResponse,
    ExchangeRateCurrencySettingDto,
    ExchangeRateCurrencySettingUpdateRequest,
    ExchangeRateSyncResult,
    PlatformProductCatalogItemDto,
    PlatformProductCatalogListResponse,
    PlatformProductCatalogMappingRequest,
    PlatformProductCatalogOptionsResponse,
    PlatformProductCatalogRecalculateRequest,
    PlatformProductCatalogRecalculateResult,
    PlatformProductCatalogSyncRequest,
    PlatformProductCatalogSyncResult,
    PlatformProductPricingRuleDto,
    PlatformProductPricingRuleInput,
    PurchaseOrderDto,
    PurchaseOrderGenerateRequest,
    PurchaseOrderItemDto,
    PurchaseOrderItemUpdateRequest,
    PurchaseOrderListResponse,
    PurchaseOrderSourceDto,
    PurchaseOrderUpdateRequest,
    LogisticsAuthorizationDto,
    LogisticsAuthorizationUpdateRequest,
    LogisticsAuthorizationVerifyResponse,
    LogisticsChannelOptionDto,
    LogisticsMatchRuleDto,
    LogisticsMatchRuleListResponse,
    LogisticsMatchRulePayload,
    LogisticsRematchRequest,
    LogisticsRematchResponse,
    MenuDto,
    UserCreateRequest,
    UserDto,
    UserOptionDto,
    WeComMentionUserOptionDto,
    RoleDto,
    RoleCreateRequest,
    RoleUpdateRequest,
    UserResetPasswordRequest,
    UserUpdateRequest,
    ShopAuthorizationRequest,
    ShopCreateRequest,
    ShopDto,
    ShopOAuthCompleteRequest,
    ShopOAuthStartRequest,
    ShopUpdateRequest,
    ScheduledTaskDto,
    ScheduledTaskRunPdfDownloadLinkDto,
    ScheduledTaskRunDto,
    ScheduledTaskRunListResponse,
    ScheduledTaskRunOrderDto,
    ScheduledTaskRunPlatformDto,
    ScheduledTaskRunStepDto,
    ScheduledTaskUpsertRequest,
    SyncSettingDto,
)
from .security import (
    create_filebrowser_session_token,
    create_scheduled_task_run_pdf_download_token,
    create_user_token,
    decode_filebrowser_session_token,
    hash_password,
    verify_password,
)
from .settings import DEFAULT_ALLOWED_ORIGINS, get_settings, validate_security_settings
from .sync_engine import (
    VOIDED_PLATFORM_STATUSES,
    refresh_configs,
    sync_enabled_accounts,
)
from .task_runner import list_task_run_orders, list_task_run_steps, list_task_runs, monitor_printer_status, process_due_task_retries, refresh_reprint_candidates, retry_run_order_print, retry_run_platform_print
from .traffic_analytics import (
    TrafficSyncRequest,
    create_traffic_sync_runs,
    list_traffic_accounts,
    mark_interrupted_traffic_runs,
    query_categories as query_traffic_categories,
    query_category_sku_comparison as query_traffic_category_sku_comparison,
    query_category_sku_focus_analysis as query_traffic_category_sku_focus_analysis,
    query_comparison as query_traffic_comparison,
    query_negative_reviews_daily as query_traffic_negative_reviews_daily,
    query_rankings as query_traffic_rankings,
    query_summary as query_traffic_summary,
    run_traffic_sync_runs,
    validate_period as validate_traffic_period,
)
def _configured_cors_origins() -> list[str]:
    try:
        origins = get_settings().allowed_origins
    except Exception as exc:
        logger = logging.getLogger(__name__)
        logger.warning("Falling back to local CORS origins because settings are unavailable: %s", exc)
        origins = DEFAULT_ALLOWED_ORIGINS
    return [str(origin).strip() for origin in origins if str(origin).strip()]


app = FastAPI(title="CaifuClaw AI", version="0.3.0")
logger = logging.getLogger(__name__)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_configured_cors_origins(),
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-CaifuClaw-Retry-Attempt"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000, compresslevel=5)

ORDER_STATUS_PENDING = "待处理"
ORDER_STATUS_WAITING_PRINT = "待打印"
ORDER_STATUS_WAITING_PURCHASE = "待采购"
ORDER_STATUS_PICKING = "配货中"
ORDER_STATUS_SHIPPED = "已发货"
ORDER_STATUS_AWAITING_PICKUP = "待揽收"
ORDER_STATUS_DELIVERED = "已妥投"
ORDER_STATUS_VOIDED = "已作废"
ORDER_STATUS_COMPLETED = "已完成"
ORDER_STATUS_ALL = "all"
LEGACY_FBJ_EXPORT_STATUSES = ("FBJ待导出", "FBJ已导出")
PURCHASE_DELETE_ROLLBACK_PAYMENT_START_DATE = date(2026, 6, 1)
ORDER_STATUS_PENDING_KEY = "pending"
ORDER_STATUS_WAITING_PRINT_KEY = "waiting_print"
ORDER_STATUS_WAITING_PURCHASE_KEY = "waiting_purchase"
ORDER_STATUS_PICKING_KEY = "picking"
ORDER_STATUS_SHIPPED_KEY = "shipped"
ORDER_STATUS_AWAITING_PICKUP_KEY = "awaiting_pickup"
ORDER_STATUS_DELIVERED_KEY = "delivered"
ORDER_STATUS_VOIDED_KEY = "voided"
ORDER_STATUS_COMPLETED_KEY = "completed"
ORDER_RISK_ALL = "all"
ORDER_RISK_UNHANDLED = "unhandled"
ORDER_RISK_HANDLED = "handled"
ORDER_RISK_OVERDUE = "overdue"
ORDER_RISK_DUE_24 = "due_24"
ORDER_RISK_FILTERS = {
    ORDER_RISK_ALL,
    ORDER_RISK_UNHANDLED,
    ORDER_RISK_HANDLED,
    ORDER_RISK_OVERDUE,
    ORDER_RISK_DUE_24,
}
ORDER_RISK_BIZ_STATUSES = (
    ORDER_STATUS_PENDING,
    ORDER_STATUS_WAITING_PRINT,
    ORDER_STATUS_WAITING_PURCHASE,
    ORDER_STATUS_PICKING,
)

ORDER_STATUS_KEY_TO_LABEL = {
    ORDER_STATUS_PENDING_KEY: ORDER_STATUS_PENDING,
    ORDER_STATUS_WAITING_PRINT_KEY: ORDER_STATUS_WAITING_PRINT,
    ORDER_STATUS_WAITING_PURCHASE_KEY: ORDER_STATUS_WAITING_PURCHASE,
    ORDER_STATUS_PICKING_KEY: ORDER_STATUS_PICKING,
    ORDER_STATUS_SHIPPED_KEY: ORDER_STATUS_SHIPPED,
    ORDER_STATUS_AWAITING_PICKUP_KEY: ORDER_STATUS_AWAITING_PICKUP,
    ORDER_STATUS_DELIVERED_KEY: ORDER_STATUS_DELIVERED,
    ORDER_STATUS_VOIDED_KEY: ORDER_STATUS_VOIDED,
    ORDER_STATUS_COMPLETED_KEY: ORDER_STATUS_COMPLETED,
}
ORDER_STATUS_LABEL_TO_KEY = {label: key for key, label in ORDER_STATUS_KEY_TO_LABEL.items()}
ORDER_STATUS_LABEL_TO_KEY.update(
    {
        "待配送": ORDER_STATUS_AWAITING_PICKUP_KEY,
    }
)
ORDER_STATUS_AWAITING_PICKUP_LABELS = {ORDER_STATUS_AWAITING_PICKUP, "待配送"}
ORDER_VOIDED_PLATFORM_STATUS_VALUES = tuple(sorted(VOIDED_PLATFORM_STATUSES))
SHOP_AUTH_SUCCESS = "success"
SHOP_AUTH_FAILED = "failed"
SHOP_AUTH_UNAUTHORIZED = "unauthorized"
LOGISTICS_AUTH_SUCCESS = "success"
LOGISTICS_AUTH_FAILED = "failed"
LOGISTICS_AUTH_UNAUTHORIZED = "unauthorized"
LOGISTICS_AUTH_FIELD_SCHEMAS: dict[str, list[str]] = {
    "qianhai_weishi": ["token", "account"],
    "wanbang_suda_new": ["customer_code", "token"],
    BSI_CARRIER_CODE: ["app_id", "customer_code", "customer_secret"],
}
LOGISTICS_AUTH_SEED_DATA = [
    {
        "carrier_code": "qianhai_weishi",
        "carrier_name": "深圳前海纬狮物流网络科技有限公司",
        "account_name": "DEMO-CARRIER-1",
        "credential_type": "api_key",
        "credentials": {
            "token": "",
            "account": "DEMO-CARRIER-1",
        },
        "config_json": {
            "company_name_en": "",
            "same_address_doorplate": False,
            "auto_recipient_phone": "",
            "auto_recipient_email": "",
        },
    },
    {
        "carrier_code": "wanbang_suda_new",
        "carrier_name": "万邦速达(新)",
        "account_name": "DEMO-CARRIER",
        "credential_type": "api_key",
        "credentials": {
            "customer_code": "DEMO-CARRIER-2",
            "token": "",
        },
        "config_json": {
            "company_name_en": "Allegro",
            "label_fields": [],
            "same_address_doorplate": False,
            "auto_recipient_phone": "",
            "auto_recipient_email": "",
            "base_url": "https://api.wanbexpress.com",
            "warehouse_code": "",
            "shipping_method": "",
            "item_type": "SPX",
            "with_battery_type": "NOBattery",
            "default_weight_kg": "0.2",
            "length_cm": "1",
            "width_cm": "1",
            "height_cm": "1",
            "default_declared_name_en": "goods",
            "default_declared_name_cn": "goods",
            "default_declared_value": "1",
            "default_declared_currency": "USD",
            "production_company_name": "",
            "production_company_uscc": "",
            "auto_confirm": True,
        },
    },
    {
        "carrier_code": BSI_CARRIER_CODE,
        "carrier_name": "BSI海外仓",
        "account_name": "DEMO-CARRIER-3",
        "credential_type": "api_sign",
        "credentials": {
            "app_id": "",
            "customer_code": "DEMO-CARRIER-3",
            "customer_secret": "",
        },
        "config_json": {
            "auto_create_drafts": False,
            "base_url": "https://gateway.gotofreight.com/sdmspanel",
            "warehouse_code": "DEMO-WAREHOUSE",
            "callback_url": "",
            "poland_channel_id": 1061,
            "poland_channel_name": "",
            "pan_eu_channel_id": 3102,
            "pan_eu_channel_name": "",
        },
    },
]
PLATFORM_CATALOG = [
    {"platform": "ozon", "display_name": "Ozon", "auth_type": "api_key", "base_url": "https://api-seller.ozon.ru"},
    {"platform": "wildberries", "display_name": "Wildberries", "auth_type": "api_key", "base_url": "https://marketplace-api.wildberries.ru"},
    {"platform": "joom_logistics", "display_name": "Joom", "auth_type": "oauth2", "base_url": "https://api-merchant.joom.com/api/v3"},
    {"platform": "allegro", "display_name": "Allegro", "auth_type": "oauth2", "base_url": "https://api.allegro.pl"},
    {"platform": "mercadolibre", "display_name": "MercadoLibre", "auth_type": "oauth2", "base_url": "https://api.mercadolibre.com"},
    {"platform": "amazon", "display_name": "Amazon", "auth_type": "oauth2_sigv4", "base_url": "https://sellingpartnerapi-na.amazon.com"},
    {"platform": "shopee", "display_name": "Shopee", "auth_type": "oauth2_hmac", "base_url": "https://partner.shopeemobile.com"},
    {"platform": "tiktok_shop", "display_name": "TikTok Shop", "auth_type": "oauth2_hmac", "base_url": "https://open-api.tiktokglobalshop.com"},
    {"platform": "aliexpress", "display_name": "AliExpress", "auth_type": "oauth2_top", "base_url": "https://api-sg.aliexpress.com/sync"},
    {"platform": "lazada", "display_name": "Lazada", "auth_type": "oauth2_top", "base_url": "https://api.lazada.com/rest"},
    {"platform": "shopify", "display_name": "Shopify", "auth_type": "oauth2_admin_api", "base_url": "https://{shop_domain}/admin/api/2026-04"},
    {"platform": "ebay", "display_name": "eBay", "auth_type": "oauth2", "base_url": "https://api.ebay.com"},
    {"platform": "walmart", "display_name": "Walmart", "auth_type": "oauth2_client_credentials", "base_url": "https://marketplace.walmartapis.com"},
    {"platform": "temu", "display_name": "Temu", "auth_type": "oauth2_hmac", "base_url": "https://openapi-b-us.temu.com"},
    {"platform": "shein", "display_name": "SHEIN", "auth_type": "hmac_openapi", "base_url": "https://openapi.sheincorp.com"},
    {"platform": "coupang", "display_name": "Coupang", "auth_type": "hmac_openapi", "base_url": "https://api-gateway.coupang.com"},
    {"platform": "wayfair", "display_name": "Wayfair", "auth_type": "oauth2_client_credentials_graphql", "base_url": "https://api.wayfair.com/v1/graphql"},
    {"platform": "dmsmatrix", "display_name": "DMSMatrix", "auth_type": "api_key", "base_url": "https://api.dmsmatrix.net/apis"},
]
PLATFORM_ALIASES = {
    "joom": "joom_logistics",
    "joomlogistics": "joom_logistics",
    "mercado": "mercadolibre",
    "mercado_global": "mercadolibre",
    "mercadoglobal": "mercadolibre",
    "mercado_libre": "mercadolibre",
    "tiktok": "tiktok_shop",
    "tiktokshop": "tiktok_shop",
    "ali_express": "aliexpress",
    "shopify_admin": "shopify",
    "ebay_sell": "ebay",
    "walmart_marketplace": "walmart",
    "shein_open": "shein",
    "coupang_openapi": "coupang",
    "wayfair_partner": "wayfair",
    "dms_matrix": "dmsmatrix",
    "dms-matrix": "dmsmatrix",
    "dms_matrix_erp": "dmsmatrix",
    "dmsmatrix_erp": "dmsmatrix",
}
PLATFORM_DISPLAY_NAMES = {
    "ozon": "Ozon",
    "joom_logistics": "Joom",
    "allegro": "Allegro",
    "wildberries": "Wildberries",
    "mercadolibre": "MercadoLibre",
    "amazon": "Amazon",
    "shopee": "Shopee",
    "tiktok_shop": "TikTok Shop",
    "aliexpress": "AliExpress",
    "lazada": "Lazada",
    "shopify": "Shopify",
    "ebay": "eBay",
    "walmart": "Walmart",
    "temu": "Temu",
    "shein": "SHEIN",
    "coupang": "Coupang",
    "wayfair": "Wayfair",
    "dmsmatrix": "DMSMatrix",
    PRINT_PLATFORM_CHINESE_LABEL: PRINT_PLATFORM_CHINESE_LABEL_NAME,
}
SHOP_CODE_PREFIXES = {
    "joom_logistics": "joom",
    "tiktok_shop": "tiktok",
    "aliexpress": "ali",
    "shopify": "shopify",
    "walmart": "walmart",
    "coupang": "coupang",
    "wayfair": "wayfair",
    "dmsmatrix": "dms",
}
PRINT_DOCUMENT_TYPE_LABEL = "label"
PRINT_DOCUMENT_TYPES = {
    PRINT_DOCUMENT_TYPE_LABEL: "面单打印",
}
PRINT_ORIENTATION_TYPES = PRINT_ORIENTATION_LABELS
ORDER_LIST_TABLE_KEY = "orders.list"
ORDER_LIST_PRIMARY_COLUMN_KEY = "platform_order_no"
ORDER_LIST_EXPORT_COLUMNS = [
    {"key": "platform_order_no", "title": "订单编号", "required": True, "fixed": "left"},
    {"key": "platform", "title": "平台"},
    {"key": "shop_name", "title": "店铺"},
    {"key": "transaction_id", "title": "交易号"},
    {"key": "posting_number", "title": "交运单号"},
    {"key": "tracking_number", "title": "货运单号"},
    {"key": "status", "title": "状态"},
    {"key": "platform_status", "title": "平台状态"},
    {"key": "fulfillment_type", "title": "履约类型"},
    {"key": "country_name_cn", "title": "国家"},
    {"key": "logistics_channel", "title": "物流渠道"},
    {"key": "logistics_match_rule_name", "title": "匹配规则"},
    {"key": "order_amount", "title": "订单金额"},
    {"key": "payment_at", "title": "付款日期"},
    {"key": "handover_at", "title": "交运时间"},
    {"key": "remaining_shipping_time", "title": "剩余发货时间"},
    {"key": "created_at", "title": "订单导入时间"},
    {"key": "actions", "title": "操作", "settingsHidden": True, "fixed": "right"},
]
ORDER_SUMMARY_TABLE_KEY = "order-summary.list"
ORDER_SUMMARY_PRIMARY_COLUMN_KEY = "order_no"
ORDER_SUMMARY_EXPORT_COLUMNS = [
    {"key": "picking_at", "title": "配货日"},
    {"key": "platform", "title": "平台"},
    {"key": "shop_name", "title": "店铺名"},
    {"key": "platform_created_at", "title": "创建时间"},
    {"key": "order_no", "title": "订单编号", "required": True, "fixed": False},
    {"key": "status", "title": "状态"},
    {"key": "platform_status", "title": "平台状态"},
    {"key": "country_name_cn", "title": "国家"},
    {"key": "customer_name", "title": "客户姓名"},
    {"key": "sku", "title": "SKU"},
    {"key": "platform_product_name", "title": "产品名称"},
    {"key": "quantity", "title": "数量"},
    {"key": "unit_price", "title": "单价"},
    {"key": "currency", "title": "币种"},
    {"key": "buyer_selected_logistics", "title": "自选物流"},
    {"key": "shipping_deadline_at", "title": "最后发货期限"},
    {"key": "shipment_tracking_number", "title": "货运单号"},
    {"key": "dispatch_deadline_at", "title": "发出截止时间"},
    {"key": "product_name", "title": "产品中文名称"},
    {"key": "customer_confirm", "title": "客户确认"},
    {"key": "warning", "title": "预警"},
    {"key": "purchase_no", "title": "采购单号"},
    {"key": "shipping_time", "title": "Shipping time"},
]
INVENTORY_TABLE_KEY = "inventory.list"
INVENTORY_PRIMARY_COLUMN_KEY = "product_code"
INVENTORY_EXPORT_COLUMNS = [
    {"key": "product_code", "title": "产品编号", "required": True, "fixed": "left"},
    {"key": "product_name", "title": "产品名称"},
    {"key": "stock_qty", "title": "库存数量"},
    {"key": "last_count_qty", "title": "上次盘点"},
    {"key": "safety_stock", "title": "安全库存"},
    {"key": "stock_status", "title": "库存状态"},
    {"key": "remark", "title": "备注"},
    {"key": "updated_at", "title": "更新时间"},
    {"key": "actions", "title": "操作", "settingsHidden": True, "fixed": "right"},
]
OUTBOUND_SCANS_TABLE_KEY = "outbound-scans.list"
OUTBOUND_SCANS_PRIMARY_COLUMN_KEY = "tracking_number"
OUTBOUND_SCANS_EXPORT_COLUMNS = [
    {"key": "scanned_at", "title": "扫描时间"},
    {"key": "result", "title": "结果"},
    {"key": "tracking_number", "title": "货运单号", "required": True, "fixed": False},
    {"key": "platform", "title": "平台"},
    {"key": "shop_name", "title": "店铺"},
    {"key": "platform_order_no", "title": "订单编号"},
    {"key": "order_status", "title": "订单状态"},
    {"key": "platform_status", "title": "平台状态"},
    {"key": "message", "title": "提示"},
    {"key": "scanned_by", "title": "扫描人"},
]
SHIPPING_DEADLINE_BASE_DATE_LABELS = {
    BASE_DATE_PAYMENT_AT: "付款时间",
    BASE_DATE_PLATFORM_CREATED: "创建时间",
    BASE_DATE_SHIPPING_DEADLINE: "最后发货期限",
}
ROLE_ADMIN = "admin"
ROLE_USER = "user"
ROLE_PURCHASE = "purchase"
ROLE_SALES = "sales"
ROLE_CUSTOMER_SERVICE = "customer_service"

ROLE_DEFINITIONS = [
    {"code": ROLE_ADMIN, "name": "超级管理员"},
    {"code": ROLE_PURCHASE, "name": "采购"},
    {"code": ROLE_SALES, "name": "销售"},
    {"code": ROLE_CUSTOMER_SERVICE, "name": "客服"},
]
ROLE_CODES = {item["code"] for item in ROLE_DEFINITIONS}

MENU_DEFINITIONS = [
    {"code": "dashboard", "label": "工作台", "path": "/dashboard"},
    {"code": "operations-daily-report", "label": "运营日报表", "path": "/operations-daily-report"},
    {"code": "traffic-analytics", "label": "流量分析", "path": "/traffic-analytics"},
    {"code": "platform-product-catalog", "label": "平台产品目录", "path": "/platform-product-catalog"},
    {"code": "ai-image-processing", "label": "图片处理", "path": "/ai-image-processing"},
    {"code": "text-translation", "label": "文字翻译", "path": "/text-translation"},
    {"code": "traffic-sync-status", "label": "流量同步状态", "path": "/traffic-sync-status"},
    {"code": "orders", "label": "订单列表", "path": "/orders"},
    {"code": "order-summary", "label": "订单明细表", "path": "/order-summary"},
    {"code": "purchase-orders", "label": "采购单管理", "path": "/purchase-orders"},
    {"code": "purchase-details", "label": "采购明细表", "path": "/purchase-details"},
    {"code": "scan-outbound", "label": "扫码出库", "path": "/scan-outbound"},
    {"code": "outbound-scans", "label": "扫码记录", "path": "/outbound-scans"},
    {"code": "inventory", "label": "产品库存", "path": "/inventory"},
    {"code": "products", "label": "产品管理", "path": "/products"},
    {"code": "logistics-authorizations", "label": "物流授权", "path": "/logistics-authorizations"},
    {"code": "logistics-rules", "label": "物流规则", "path": "/logistics-rules"},
    {"code": "shops", "label": "店铺管理", "path": "/shops"},
    {"code": "users", "label": "用户管理", "path": "/users"},
    {"code": "permissions", "label": "权限管理", "path": "/permissions"},
    {"code": "system-settings", "label": "系统配置", "path": "/system-settings"},
    {"code": "exchange-rates", "label": "汇率管理", "path": "/exchange-rates"},
    {"code": "scheduled-task-logs", "label": "定时任务日志", "path": "/scheduled-task-logs"},
    {"code": "sync-api-logs", "label": "同步接口日志", "path": "/sync-api-logs"},
]
MENU_CODES = {item["code"] for item in MENU_DEFINITIONS}
MENU_CODE_ALIASES = {"order-outbound": "outbound-scans"}
ADMIN_MENU_CODES = [item["code"] for item in MENU_DEFINITIONS]
DEFAULT_ROLE_MENUS = {
    ROLE_PURCHASE: ["order-summary", "products", "inventory", "purchase-orders", "purchase-details"],
    ROLE_SALES: [
        "operations-daily-report",
        "traffic-analytics",
        "platform-product-catalog",
        "traffic-sync-status",
        "orders",
        "order-summary",
        "outbound-scans",
        "logistics-rules",
        "products",
    ],
    ROLE_CUSTOMER_SERVICE: ["orders", "order-summary", "scan-outbound", "outbound-scans", "logistics-rules"],
}

API_MENU_RULES: list[tuple[str, set[str]]] = [
    ("/api/v1/users", {"users"}),
    ("/api/v1/menus", {"permissions"}),
    ("/api/v1/roles", {"permissions"}),
    ("/api/v1/dashboard/operations", {"operations-daily-report"}),
    ("/api/v1/dashboard", {"dashboard"}),
    ("/api/v1/traffic-analytics", {"traffic-analytics"}),
    ("/api/v1/platform-product-catalog", {"platform-product-catalog"}),
    ("/api/v1/ai-image", {"ai-image-processing"}),
    ("/api/v1/ai-translation", {"text-translation"}),
    ("/api/v1/purchase-details", {"purchase-details"}),
    ("/api/v1/purchase-orders", {"purchase-orders"}),
    ("/api/v1/inventory", {"inventory"}),
    ("/api/v1/products", {"products"}),
    ("/api/v1/product-shops", {"products"}),
    ("/api/v1/order-summary", {"order-summary"}),
    ("/api/v1/orders/batch", {"orders"}),
    ("/api/v1/orders", {"orders"}),
    ("/api/orders", {"orders"}),
    ("/api/v1/outbound-scan", {"scan-outbound"}),
    ("/api/v1/logistics-rules", {"logistics-rules"}),
    ("/api/v1/logistics-authorizations", {"logistics-authorizations"}),
    ("/api/v1/shops", {"shops"}),
    ("/api/shops", {"shops"}),
    ("/api/v1/platforms", {"shops", "system-settings"}),
    ("/api/v1/sync-settings", {"shops"}),
    ("/api/sync-settings", {"shops"}),
    ("/api/v1/sync-api-logs", {"sync-api-logs"}),
    ("/api/v1/system-settings/scheduled-task-runs", {"system-settings", "scheduled-task-logs"}),
    ("/api/v1/system-settings/scheduled-task-run-orders", {"system-settings", "scheduled-task-logs"}),
    ("/api/v1/system-settings/exchange-rates", {"system-settings", "exchange-rates"}),
    ("/api/v1/system-settings", {"system-settings"}),
]


def _ensure_legacy_columns() -> None:
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE IF EXISTS local_users ADD COLUMN IF NOT EXISTS display_name VARCHAR(120) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS local_users ADD COLUMN IF NOT EXISTS wecom_mobile VARCHAR(20) DEFAULT ''"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS roles (
                    id SERIAL PRIMARY KEY,
                    code VARCHAR(80) NOT NULL UNIQUE,
                    name VARCHAR(120) NOT NULL,
                    description TEXT DEFAULT '',
                    is_system BOOLEAN DEFAULT FALSE,
                    enabled BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS local_users ADD COLUMN IF NOT EXISTS role_id INTEGER REFERENCES roles(id) ON DELETE SET NULL"))
        conn.execute(text("ALTER TABLE IF EXISTS local_users ADD COLUMN IF NOT EXISTS enabled BOOLEAN DEFAULT TRUE"))
        conn.execute(text("ALTER TABLE IF EXISTS local_users ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP"))
        conn.execute(text("UPDATE local_users SET display_name = COALESCE(display_name, username, '')"))
        conn.execute(text("UPDATE local_users SET wecom_mobile = COALESCE(wecom_mobile, '')"))
        conn.execute(text("UPDATE local_users SET enabled = COALESCE(enabled, TRUE)"))
        conn.execute(text("UPDATE local_users SET updated_at = COALESCE(updated_at, created_at, timezone('UTC', NOW()))"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS roles (
                    id SERIAL PRIMARY KEY,
                    code VARCHAR(80) NOT NULL UNIQUE,
                    name VARCHAR(120) NOT NULL,
                    description TEXT DEFAULT '',
                    is_system BOOLEAN DEFAULT FALSE,
                    enabled BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_roles_code ON roles(code)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS role_menu_permissions (
                    id SERIAL PRIMARY KEY,
                    role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                    menu_code VARCHAR(80) NOT NULL,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_role_menu_permission UNIQUE(role_id, menu_code)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_role_menu_permissions_role_id ON role_menu_permissions(role_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_role_menu_permissions_menu_code ON role_menu_permissions(menu_code)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS user_roles (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES local_users(id) ON DELETE CASCADE,
                    role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_user_roles_user_role UNIQUE(user_id, role_id)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_user_roles_user_id ON user_roles(user_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_user_roles_role_id ON user_roles(role_id)"))
        conn.execute(text("ALTER TABLE IF EXISTS user_roles ALTER COLUMN created_at SET DEFAULT timezone('UTC', NOW())"))
        conn.execute(
            text(
                """
                INSERT INTO user_roles (user_id, role_id, created_at)
                SELECT id, role_id, COALESCE(updated_at, created_at, timezone('UTC', NOW()))
                FROM local_users
                WHERE role_id IS NOT NULL
                ON CONFLICT (user_id, role_id) DO NOTHING
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS credential_type VARCHAR(40) DEFAULT 'api_key'"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS encrypted_credentials BYTEA"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS status VARCHAR(40) DEFAULT 'active'"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS authorization_status VARCHAR(40) DEFAULT 'unauthorized'"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS token_valid BOOLEAN"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS token_message TEXT"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS last_authorized_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS authorization_expires_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS session_expires_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS last_sync_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS last_sync_status VARCHAR(255)"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS credentials_version VARCHAR(80) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts DROP COLUMN IF EXISTS backup_status"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts DROP COLUMN IF EXISTS backup_error"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts DROP COLUMN IF EXISTS backed_up_at"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS created_by VARCHAR(80)"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_accounts ADD COLUMN IF NOT EXISTS created_at TIMESTAMP"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS logistics_authorizations (
                    id SERIAL PRIMARY KEY,
                    carrier_code VARCHAR(80) NOT NULL,
                    carrier_name VARCHAR(160) DEFAULT '',
                    account_name VARCHAR(160) DEFAULT '',
                    enabled BOOLEAN DEFAULT TRUE,
                    authorization_status VARCHAR(40) DEFAULT 'unauthorized',
                    token_valid BOOLEAN,
                    token_message TEXT,
                    credential_type VARCHAR(40) DEFAULT 'api_key',
                    encrypted_credentials BYTEA,
                    config_json JSONB DEFAULT '{}'::jsonb,
                    settings_json JSONB DEFAULT '{}'::jsonb,
                    last_authorized_at TIMESTAMP,
                    authorization_expires_at TIMESTAMP,
                    credentials_version VARCHAR(80) DEFAULT '',
                    created_by VARCHAR(80),
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_logistics_authorizations_carrier_account UNIQUE(carrier_code, account_name)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_logistics_authorizations_carrier_code ON logistics_authorizations(carrier_code)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_logistics_authorizations_enabled ON logistics_authorizations(enabled)"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS carrier_code VARCHAR(80)"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS carrier_name VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS account_name VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS enabled BOOLEAN DEFAULT TRUE"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS authorization_status VARCHAR(40) DEFAULT 'unauthorized'"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS token_valid BOOLEAN"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS token_message TEXT"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS credential_type VARCHAR(40) DEFAULT 'api_key'"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS encrypted_credentials BYTEA"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS config_json JSONB DEFAULT '{}'::jsonb"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS settings_json JSONB DEFAULT '{}'::jsonb"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS last_authorized_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS authorization_expires_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS credentials_version VARCHAR(80) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS created_by VARCHAR(80)"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS created_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_authorizations ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS logistics_match_rules (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(160) NOT NULL,
                    platform VARCHAR(40) DEFAULT '',
                    priority INTEGER DEFAULT 10,
                    enabled BOOLEAN DEFAULT TRUE,
                    shop_names JSONB DEFAULT '[]'::jsonb,
                    is_overseas_warehouse BOOLEAN,
                    country_codes JSONB DEFAULT '[]'::jsonb,
                    logistics_channel VARCHAR(160) DEFAULT '',
                    carrier_code VARCHAR(80) DEFAULT '',
                    remark TEXT DEFAULT '',
                    created_by VARCHAR(80),
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS name VARCHAR(160)"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS platform VARCHAR(40) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS priority INTEGER DEFAULT 10"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS enabled BOOLEAN DEFAULT TRUE"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS shop_names JSONB DEFAULT '[]'::jsonb"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS is_overseas_warehouse BOOLEAN"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS country_codes JSONB DEFAULT '[]'::jsonb"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS logistics_channel VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS carrier_code VARCHAR(80) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS remark TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS created_by VARCHAR(80)"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS created_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS logistics_match_rules ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_logistics_match_rules_enabled_priority ON logistics_match_rules(enabled, priority)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_logistics_match_rules_platform_priority ON logistics_match_rules(platform, enabled, priority)"))
        conn.execute(text("ALTER TABLE IF EXISTS api_request_logs ADD COLUMN IF NOT EXISTS operation VARCHAR(80) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS api_request_logs ADD COLUMN IF NOT EXISTS status VARCHAR(40) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS api_request_logs ADD COLUMN IF NOT EXISTS request_id VARCHAR(120) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS api_request_logs ADD COLUMN IF NOT EXISTS extra JSONB DEFAULT '{}'::jsonb"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_api_request_logs_operation ON api_request_logs(operation)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_api_request_logs_status ON api_request_logs(status)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_api_request_logs_request_id ON api_request_logs(request_id)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS scheduler_heartbeats (
                    id SERIAL PRIMARY KEY,
                    owner_id VARCHAR(160) NOT NULL,
                    host VARCHAR(160) DEFAULT '',
                    pid INTEGER,
                    is_leader BOOLEAN DEFAULT FALSE,
                    started_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    last_seen_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    message TEXT DEFAULT '',
                    CONSTRAINT uq_scheduler_heartbeats_owner UNIQUE(owner_id)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduler_heartbeats_last_seen ON scheduler_heartbeats(last_seen_at)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS sync_account_states (
                    id SERIAL PRIMARY KEY,
                    platform VARCHAR(40) NOT NULL,
                    account_id VARCHAR(120) NOT NULL,
                    job_type VARCHAR(80) DEFAULT 'sync_orders',
                    last_started_at TIMESTAMP,
                    last_finished_at TIMESTAMP,
                    last_success_at TIMESTAMP,
                    last_failed_at TIMESTAMP,
                    next_due_at TIMESTAMP,
                    last_status VARCHAR(40) DEFAULT '',
                    consecutive_failures INTEGER DEFAULT 0,
                    overdue_since TIMESTAMP,
                    catchup_required BOOLEAN DEFAULT FALSE,
                    catchup_from TIMESTAMP,
                    catchup_to TIMESTAMP,
                    last_message TEXT DEFAULT '',
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    CONSTRAINT uq_sync_account_state UNIQUE(platform, account_id, job_type)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_account_states_platform_account ON sync_account_states(platform, account_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_account_states_next_due ON sync_account_states(next_due_at)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_account_states_last_success ON sync_account_states(last_success_at)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_account_states_status ON sync_account_states(last_status)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS sync_audit_logs (
                    id SERIAL PRIMARY KEY,
                    event_type VARCHAR(80) NOT NULL,
                    platform VARCHAR(40) DEFAULT '',
                    account_id VARCHAR(120) DEFAULT '',
                    job_type VARCHAR(80) DEFAULT '',
                    status VARCHAR(40) DEFAULT '',
                    message TEXT DEFAULT '',
                    owner_id VARCHAR(160) DEFAULT '',
                    extra JSONB DEFAULT '{}'::jsonb,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_audit_logs_created ON sync_audit_logs(created_at)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_audit_logs_event_type ON sync_audit_logs(event_type)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_sync_audit_logs_account ON sync_audit_logs(platform, account_id)"))
        conn.execute(
            text(
                """
                UPDATE api_request_logs
                SET status = CASE
                    WHEN error_message IS NOT NULL AND error_message <> '' THEN 'failed'
                    WHEN response_status IS NOT NULL AND response_status >= 400 THEN 'failed'
                    ELSE 'success'
                END
                WHERE status IS NULL OR status = ''
                """
            )
        )
        conn.execute(
            text(
                """
                UPDATE api_request_logs
                SET operation = CASE
                    WHEN lower(url) LIKE '%package-label%' OR lower(url) LIKE '%label%' OR lower(url) LIKE '%sticker%' THEN 'fetch_label'
                    WHEN lower(url) LIKE '%shipment%' OR lower(url) LIKE '%delivery%' THEN 'shipment'
                    WHEN lower(url) LIKE '%order%' OR lower(url) LIKE '%posting%' THEN 'order_api'
                    ELSE 'platform_api'
                END
                WHERE operation IS NULL OR operation = ''
                """
            )
        )

        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS shop_id VARCHAR(120)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS internal_order_no VARCHAR(32)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ALTER COLUMN internal_order_no TYPE VARCHAR(32)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS shop_name VARCHAR(160)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS site VARCHAR(80)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS buyer_id VARCHAR(120)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS buyer_name VARCHAR(160)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS biz_status VARCHAR(40)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS fulfillment_type VARCHAR(40) DEFAULT 'FBS'"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS is_overseas_warehouse BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS bsi_order_no VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS bsi_submitted_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS platform_handover_deadline TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS country_code VARCHAR(8)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS country_name_cn VARCHAR(80)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS buyer_selected_logistics VARCHAR(160)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS order_amount VARCHAR(40)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS currency VARCHAR(16)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS platform_created_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS payment_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS shipping_deadline_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS dispatch_deadline_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS shipment_tracking_number VARCHAR(160)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_channel VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_carrier_code VARCHAR(80) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_match_rule_id INTEGER"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_match_rule_name VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_match_status VARCHAR(40) DEFAULT 'unmatched'"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_match_reason TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_matched_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS picking_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS marked_shipped_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS label_printed_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS handover_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS shipped_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS logistics_last_synced_at TIMESTAMP"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_logistics_last_synced_at ON orders(logistics_last_synced_at)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_logistics_match_rule_id ON orders(logistics_match_rule_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_logistics_match_status ON orders(logistics_match_status)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_fulfillment_type ON orders(fulfillment_type)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_is_overseas_warehouse ON orders(is_overseas_warehouse)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_bsi_order_no ON orders(bsi_order_no)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS order_operation_logs (
                    id SERIAL PRIMARY KEY,
                    order_id INTEGER NOT NULL REFERENCES orders(id) ON DELETE CASCADE,
                    operation_type VARCHAR(80) DEFAULT '',
                    operation_attribute VARCHAR(120) DEFAULT '',
                    description TEXT DEFAULT '',
                    operator VARCHAR(80) DEFAULT '',
                    source VARCHAR(40) DEFAULT 'manual',
                    event_key VARCHAR(180) DEFAULT '',
                    extra JSONB DEFAULT '{}'::jsonb,
                    operated_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_order_operation_logs_order_id ON order_operation_logs(order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_order_operation_logs_operated_at ON order_operation_logs(operated_at)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_order_operation_logs_event_key ON order_operation_logs(event_key)"))
        conn.execute(
            text(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS uq_order_operation_logs_event_key
                ON order_operation_logs(event_key)
                WHERE event_key IS NOT NULL AND event_key <> ''
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS exchange_rates (
                    id SERIAL PRIMARY KEY,
                    rate_date DATE NOT NULL,
                    currency_code VARCHAR(12) NOT NULL,
                    currency_name VARCHAR(80) DEFAULT '',
                    rate NUMERIC(20, 8) NOT NULL,
                    source_updated_at TIMESTAMP,
                    synced_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    CONSTRAINT uq_exchange_rates_date_currency UNIQUE(rate_date, currency_code)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_exchange_rates_rate_date ON exchange_rates(rate_date)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_exchange_rates_currency_code ON exchange_rates(currency_code)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_exchange_rates_updated_at ON exchange_rates(updated_at)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS shipping_deadline_settings (
                    id SERIAL PRIMARY KEY,
                    platform VARCHAR(40) NOT NULL,
                    base_date_field VARCHAR(40) NOT NULL DEFAULT 'platform_created_at',
                    offset_days INTEGER NOT NULL DEFAULT 0,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    enabled BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    CONSTRAINT uq_shipping_deadline_settings_platform UNIQUE(platform)
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS shipping_deadline_settings ADD COLUMN IF NOT EXISTS sort_order INTEGER DEFAULT 0"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_shipping_deadline_settings_platform ON shipping_deadline_settings(platform)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_shipping_deadline_settings_sort_order ON shipping_deadline_settings(sort_order)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_orders_dispatch_deadline_at ON orders(dispatch_deadline_at)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS exchange_rate_currency_settings (
                    id SERIAL PRIMARY KEY,
                    currency_code VARCHAR(12) NOT NULL,
                    currency_name VARCHAR(80) DEFAULT '',
                    enabled BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()) NOT NULL,
                    CONSTRAINT uq_exchange_rate_currency_settings_code UNIQUE(currency_code)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_exchange_rate_currency_settings_enabled ON exchange_rate_currency_settings(enabled)"))

        conn.execute(
            text(
                """
                UPDATE platform_accounts
                SET authorization_status = CASE
                    WHEN authorization_status IS NOT NULL THEN authorization_status
                    WHEN status = 'active' THEN 'success'
                    WHEN status IN ('auth_error', 'expired') THEN 'failed'
                    ELSE 'unauthorized'
                END
                WHERE authorization_status IS NULL
                """
            )
        )
        conn.execute(text("UPDATE platform_accounts SET token_message = COALESCE(token_message, '未验证')"))
        conn.execute(text("ALTER TABLE IF EXISTS product_shop_mappings DROP CONSTRAINT IF EXISTS uq_product_shop_mapping"))
        if conn.dialect.name == "postgresql":
            conn.execute(text("ALTER TABLE IF EXISTS products ALTER COLUMN product_code TYPE VARCHAR(40)"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS buyer_user_id INTEGER REFERENCES local_users(id) ON DELETE SET NULL"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS english_name VARCHAR(255) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS gross_weight NUMERIC(12,3)"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS package_length NUMERIC(12,2)"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS package_width NUMERIC(12,2)"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS package_height NUMERIC(12,2)"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS ean VARCHAR(64) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS description TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS main_image_url TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS products ADD COLUMN IF NOT EXISTS is_slow_moving_material BOOLEAN DEFAULT FALSE"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_products_buyer_user_id ON products(buyer_user_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_products_is_slow_moving_material ON products(is_slow_moving_material)"))
        conn.execute(text("UPDATE platform_accounts SET created_at = COALESCE(created_at, updated_at, timezone('UTC', NOW()))"))
        conn.execute(text("UPDATE orders SET shop_id = COALESCE(shop_id, account_id) WHERE shop_id IS NULL"))
        conn.execute(text("UPDATE orders SET fulfillment_type = COALESCE(NULLIF(fulfillment_type, ''), 'FBS')"))
        conn.execute(text("UPDATE orders SET is_overseas_warehouse = COALESCE(is_overseas_warehouse, FALSE)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS order_items (
                    id SERIAL PRIMARY KEY,
                    order_id INTEGER NOT NULL REFERENCES orders(id) ON DELETE CASCADE,
                    sku VARCHAR(255) DEFAULT '',
                    platform_product_name VARCHAR(500) DEFAULT '',
                    quantity INTEGER DEFAULT 1,
                    unit_price VARCHAR(40),
                    currency VARCHAR(16) DEFAULT '',
                    raw_payload JSONB DEFAULT '{}'::jsonb,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_order_items_order_id ON order_items(order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_order_items_sku ON order_items(sku)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_order_items_order_sku ON order_items(order_id, sku)"))
        conn.execute(text("ALTER TABLE IF EXISTS order_items ADD COLUMN IF NOT EXISTS platform_product_name VARCHAR(500) DEFAULT ''"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS product_inventory (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
                    product_name VARCHAR(255) NOT NULL,
                    stock_qty INTEGER DEFAULT 0,
                    last_count_qty INTEGER DEFAULT 0,
                    remark TEXT,
                    updated_by VARCHAR(80),
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_product_inventory_product_id UNIQUE(product_id)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_product_inventory_product_id ON product_inventory(product_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_product_inventory_product_name ON product_inventory(product_name)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS platform_print_settings (
                    id SERIAL PRIMARY KEY,
                    platform VARCHAR(40) NOT NULL,
                    document_type VARCHAR(40) DEFAULT 'label',
                    printer_name VARCHAR(255) DEFAULT '',
                    printer_system VARCHAR(40) DEFAULT '',
                    printer_device_uri VARCHAR(500) DEFAULT '',
                    printer_driver_name VARCHAR(255) DEFAULT '',
                    printer_port_name VARCHAR(255) DEFAULT '',
                    printer_fingerprint VARCHAR(80) DEFAULT '',
                    page_orientation VARCHAR(20) DEFAULT 'auto',
                    enabled BOOLEAN DEFAULT TRUE,
                    remark TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS document_type VARCHAR(40) DEFAULT 'label'"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS printer_system VARCHAR(40) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS printer_device_uri VARCHAR(500) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS printer_driver_name VARCHAR(255) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS printer_port_name VARCHAR(255) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS printer_fingerprint VARCHAR(80) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings ADD COLUMN IF NOT EXISTS page_orientation VARCHAR(20) DEFAULT 'auto'"))
        conn.execute(text("UPDATE platform_print_settings SET document_type = 'label' WHERE document_type IS NULL OR document_type = ''"))
        conn.execute(text("UPDATE platform_print_settings SET page_orientation = 'auto' WHERE page_orientation IS NULL OR page_orientation = ''"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings DROP CONSTRAINT IF EXISTS uq_platform_print_settings_platform"))
        conn.execute(text("ALTER TABLE IF EXISTS platform_print_settings DROP CONSTRAINT IF EXISTS platform_print_settings_platform_key"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_platform_print_settings_platform ON platform_print_settings(platform)"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_platform_print_settings_platform_document ON platform_print_settings(platform, document_type)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS platform_settings (
                    id SERIAL PRIMARY KEY,
                    platform VARCHAR(40) NOT NULL,
                    platform_name VARCHAR(160) DEFAULT '',
                    enabled BOOLEAN DEFAULT TRUE,
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_platform_settings_platform UNIQUE(platform)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_platform_settings_platform ON platform_settings(platform)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_platform_settings_enabled ON platform_settings(enabled)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS email_smtp_settings (
                    id INTEGER PRIMARY KEY,
                    provider VARCHAR(40) DEFAULT 'qq',
                    enabled BOOLEAN DEFAULT FALSE,
                    smtp_host VARCHAR(255) DEFAULT 'smtp.qq.com',
                    smtp_port INTEGER DEFAULT 465,
                    use_ssl BOOLEAN DEFAULT TRUE,
                    sender_email VARCHAR(255) DEFAULT '',
                    sender_name VARCHAR(120) DEFAULT '',
                    encrypted_auth_code BYTEA,
                    notification_recipients JSONB DEFAULT '{}'::jsonb,
                    last_test_at TIMESTAMP,
                    last_test_status VARCHAR(40) DEFAULT '',
                    last_test_message TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS email_smtp_settings ADD COLUMN IF NOT EXISTS provider VARCHAR(40) DEFAULT 'qq'"))
        conn.execute(text("ALTER TABLE IF EXISTS email_smtp_settings ADD COLUMN IF NOT EXISTS notification_recipients JSONB DEFAULT '{}'::jsonb"))
        conn.execute(
            text(
                """
                UPDATE email_smtp_settings
                SET notification_recipients = jsonb_build_object(
                    'wanbang_tracking_failure', 'demo@example.invalid',
                    'bsi_address_anomaly', 'demo@example.invalid'
                )
                WHERE id = 1
                  AND (notification_recipients IS NULL OR notification_recipients = '{}'::jsonb)
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO email_smtp_settings (
                    id,
                    provider,
                    enabled,
                    smtp_host,
                    smtp_port,
                    use_ssl,
                    sender_email,
                    sender_name,
                    last_test_status,
                    last_test_message,
                    created_at,
                    updated_at
                )
                VALUES (
                    1,
                    'qq',
                    FALSE,
                    'smtp.qq.com',
                    465,
                    TRUE,
                    '',
                    '',
                    '',
                    '',
                    timezone('UTC', NOW()),
                    timezone('UTC', NOW())
                )
                ON CONFLICT (id) DO NOTHING
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS wecom_robot_settings (
                    id INTEGER PRIMARY KEY,
                    encrypted_webhook_url BYTEA,
                    timeout_seconds INTEGER DEFAULT 30,
                    max_retries INTEGER DEFAULT 2,
                    rate_limit_per_minute INTEGER DEFAULT 20,
                    default_mentioned_user_ids TEXT DEFAULT '[]',
                    default_mentioned_list TEXT DEFAULT '[]',
                    default_mentioned_mobile_list TEXT DEFAULT '[]',
                    default_prompt TEXT DEFAULT '你有新的任务，请处理',
                    purchase_order_notify_enabled BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS encrypted_webhook_url BYTEA"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS timeout_seconds INTEGER DEFAULT 30"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS max_retries INTEGER DEFAULT 2"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS rate_limit_per_minute INTEGER DEFAULT 20"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS default_mentioned_user_ids TEXT DEFAULT '[]'"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS default_mentioned_list TEXT DEFAULT '[]'"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS default_mentioned_mobile_list TEXT DEFAULT '[]'"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS default_prompt TEXT DEFAULT '你有新的任务，请处理'"))
        conn.execute(text("ALTER TABLE IF EXISTS wecom_robot_settings ADD COLUMN IF NOT EXISTS purchase_order_notify_enabled BOOLEAN DEFAULT FALSE"))
        conn.execute(
            text(
                """
                INSERT INTO wecom_robot_settings (
                    id,
                    timeout_seconds,
                    max_retries,
                    rate_limit_per_minute,
                    default_mentioned_user_ids,
                    default_mentioned_list,
                    default_mentioned_mobile_list,
                    default_prompt,
                    purchase_order_notify_enabled,
                    created_at,
                    updated_at
                )
                VALUES (
                    1,
                    30,
                    2,
                    20,
                    '[]',
                    '[]',
                    '[]',
                    '你有新的任务，请处理',
                    FALSE,
                    timezone('UTC', NOW()),
                    timezone('UTC', NOW())
                )
                ON CONFLICT (id) DO NOTHING
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS translation_provider_settings (
                    id SERIAL PRIMARY KEY,
                    provider VARCHAR(40) NOT NULL DEFAULT 'baidu',
                    provider_name VARCHAR(80) DEFAULT '百度翻译',
                    enabled BOOLEAN DEFAULT FALSE,
                    app_id VARCHAR(160) DEFAULT '',
                    encrypted_secret_key BYTEA,
                    endpoint TEXT DEFAULT '',
                    source_language VARCHAR(20) DEFAULT 'auto',
                    timeout_seconds INTEGER DEFAULT 30,
                    max_retries INTEGER DEFAULT 2,
                    batch_size INTEGER DEFAULT 80,
                    batch_chars INTEGER DEFAULT 5000,
                    provider_options_json TEXT DEFAULT '{}',
                    last_test_at TIMESTAMP,
                    last_test_status VARCHAR(40) DEFAULT '',
                    last_test_message TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_translation_provider_settings_provider UNIQUE(provider)
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS provider VARCHAR(40) DEFAULT 'baidu'"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS provider_name VARCHAR(80) DEFAULT '百度翻译'"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS enabled BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS app_id VARCHAR(160) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS encrypted_secret_key BYTEA"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS endpoint TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS source_language VARCHAR(20) DEFAULT 'auto'"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS timeout_seconds INTEGER DEFAULT 30"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS max_retries INTEGER DEFAULT 2"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS batch_size INTEGER DEFAULT 80"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS batch_chars INTEGER DEFAULT 5000"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS provider_options_json TEXT DEFAULT '{}'"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS last_test_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS last_test_status VARCHAR(40) DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS translation_provider_settings ADD COLUMN IF NOT EXISTS last_test_message TEXT DEFAULT ''"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_translation_provider_settings_provider ON translation_provider_settings(provider)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_translation_provider_settings_enabled ON translation_provider_settings(enabled)"))
        conn.execute(
            text(
                """
                INSERT INTO translation_provider_settings (
                    provider,
                    provider_name,
                    enabled,
                    app_id,
                    endpoint,
                    source_language,
                    timeout_seconds,
                    max_retries,
                    batch_size,
                    batch_chars,
                    provider_options_json,
                    last_test_status,
                    last_test_message,
                    created_at,
                    updated_at
                )
                VALUES (
                    'baidu',
                    '百度翻译',
                    FALSE,
                    '',
                    'https://fanyi-api.baidu.com/api/trans/vip/translate',
                    'auto',
                    30,
                    2,
                    80,
                    5000,
                    '{}',
                    '',
                    '',
                    timezone('UTC', NOW()),
                    timezone('UTC', NOW())
                )
                ON CONFLICT (provider) DO NOTHING
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS model_endpoints (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(160) NOT NULL,
                    base_url TEXT DEFAULT '',
                    encrypted_api_key BYTEA,
                    enabled BOOLEAN DEFAULT TRUE,
                    remark TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_model_endpoints_name UNIQUE(name)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_model_endpoints_enabled ON model_endpoints(enabled)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS model_settings (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(160) NOT NULL,
                    model VARCHAR(160) NOT NULL,
                    endpoint_id INTEGER NOT NULL REFERENCES model_endpoints(id),
                    is_default BOOLEAN DEFAULT FALSE,
                    supports_vision BOOLEAN DEFAULT FALSE,
                    enabled BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_model_settings_name UNIQUE(name)
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS model_settings ADD COLUMN IF NOT EXISTS supports_vision BOOLEAN DEFAULT FALSE"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_model_settings_enabled ON model_settings(enabled)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_model_settings_is_default ON model_settings(is_default)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_model_settings_endpoint_id ON model_settings(endpoint_id)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS scheduled_tasks (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(120) NOT NULL,
                    task_type VARCHAR(80) DEFAULT 'auto_order_pipeline',
                    cron_expr VARCHAR(120) NOT NULL,
                    enabled BOOLEAN DEFAULT TRUE,
                    settings JSONB DEFAULT '{}'::jsonb,
                    remark TEXT DEFAULT '',
                    last_run_at TIMESTAMP,
                    last_status VARCHAR(40) DEFAULT '',
                    last_message TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_tasks_enabled ON scheduled_tasks(enabled)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_tasks_task_type ON scheduled_tasks(task_type)"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_tasks ALTER COLUMN task_type SET DEFAULT 'auto_order_pipeline'"))
        conn.execute(text("UPDATE scheduled_tasks SET task_type = 'auto_order_pipeline' WHERE task_type IS NULL OR task_type = '' OR task_type = 'placeholder'"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS scheduled_task_runs (
                    id SERIAL PRIMARY KEY,
                    scheduled_task_id INTEGER,
                    task_type VARCHAR(80) NOT NULL,
                    trigger_mode VARCHAR(40) DEFAULT 'scheduler',
                    status VARCHAR(40) DEFAULT 'running',
                    summary TEXT DEFAULT '',
                    stats_json JSONB DEFAULT '{}'::jsonb,
                    started_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    ended_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_runs_task_id ON scheduled_task_runs(scheduled_task_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_runs_status ON scheduled_task_runs(status)"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS attempt_no INTEGER DEFAULT 0"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS max_retry_count INTEGER DEFAULT 0"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS parent_run_id INTEGER"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS original_run_id INTEGER"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS next_retry_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS retry_reason TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS email_sent BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_runs ADD COLUMN IF NOT EXISTS email_error TEXT DEFAULT ''"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_runs_parent_run_id ON scheduled_task_runs(parent_run_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_runs_original_run_id ON scheduled_task_runs(original_run_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_runs_next_retry_at ON scheduled_task_runs(next_retry_at)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS scheduled_task_run_steps (
                    id SERIAL PRIMARY KEY,
                    run_id INTEGER NOT NULL REFERENCES scheduled_task_runs(id) ON DELETE CASCADE,
                    step_code VARCHAR(80) NOT NULL,
                    step_name VARCHAR(120) NOT NULL,
                    status VARCHAR(40) DEFAULT 'running',
                    message TEXT DEFAULT '',
                    stats_json JSONB DEFAULT '{}'::jsonb,
                    payload_json JSONB DEFAULT '{}'::jsonb,
                    started_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    ended_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_steps_run_id ON scheduled_task_run_steps(run_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_steps_step_code ON scheduled_task_run_steps(step_code)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS scheduled_task_run_orders (
                    id SERIAL PRIMARY KEY,
                    run_id INTEGER NOT NULL REFERENCES scheduled_task_runs(id) ON DELETE CASCADE,
                    order_id INTEGER NOT NULL,
                    platform VARCHAR(40) DEFAULT '',
                    purchase_order_id INTEGER,
                    pdf_generated BOOLEAN DEFAULT FALSE,
                    pdf_file_path TEXT DEFAULT '',
                    printer_name VARCHAR(255) DEFAULT '',
                    print_job_name VARCHAR(255) DEFAULT '',
                    print_submitted BOOLEAN DEFAULT FALSE,
                    print_message TEXT DEFAULT '',
                    status_before VARCHAR(40) DEFAULT '',
                    status_after VARCHAR(40) DEFAULT '',
                    needs_reprint BOOLEAN DEFAULT FALSE,
                    error_message TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_scheduled_task_run_orders_run_order_platform UNIQUE(run_id, order_id, platform)
                )
                """
            )
        )
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_run_orders DROP CONSTRAINT IF EXISTS uq_scheduled_task_run_orders_run_order"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_run_orders DROP CONSTRAINT IF EXISTS uq_scheduled_task_run_orders_run_order_platform"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_orders_run_id ON scheduled_task_run_orders(run_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_orders_order_id ON scheduled_task_run_orders(order_id)"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_scheduled_task_run_orders_run_order_platform ON scheduled_task_run_orders(run_id, order_id, platform)"))
        conn.execute(text("ALTER TABLE IF EXISTS scheduled_task_run_orders ADD COLUMN IF NOT EXISTS print_job_name VARCHAR(255) DEFAULT ''"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS purchase_orders (
                    id SERIAL PRIMARY KEY,
                    purchase_no VARCHAR(40) NOT NULL UNIQUE,
                    status VARCHAR(40) DEFAULT '草稿',
                    purchase_date DATE DEFAULT CURRENT_DATE,
                    source_count INTEGER DEFAULT 0,
                    item_count INTEGER DEFAULT 0,
                    total_required_qty INTEGER DEFAULT 0,
                    created_by VARCHAR(80),
                    remark TEXT,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_orders_purchase_no ON purchase_orders(purchase_no)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_orders_status ON purchase_orders(status)"))
        conn.execute(text("ALTER TABLE IF EXISTS purchase_orders ADD COLUMN IF NOT EXISTS purchase_date DATE"))
        conn.execute(text("UPDATE purchase_orders SET purchase_date = COALESCE(purchase_date, DATE(created_at + INTERVAL '8 hours'))"))
        conn.execute(text("UPDATE purchase_orders SET status = '完成' WHERE status = '已完成'"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_orders_purchase_date ON purchase_orders(purchase_date)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS purchase_order_edit_locks (
                    id SERIAL PRIMARY KEY,
                    purchase_order_id INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
                    locked_by VARCHAR(80) NOT NULL,
                    locked_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    expires_at TIMESTAMP NOT NULL,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    CONSTRAINT uq_purchase_order_edit_locks_order UNIQUE(purchase_order_id)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_edit_locks_order ON purchase_order_edit_locks(purchase_order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_edit_locks_expires_at ON purchase_order_edit_locks(expires_at)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS purchase_order_items (
                    id SERIAL PRIMARY KEY,
                    purchase_order_id INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
                    product_id INTEGER REFERENCES products(id) ON DELETE SET NULL,
                    product_name VARCHAR(255) NOT NULL,
                    required_qty INTEGER DEFAULT 0,
                    buyer_user_id INTEGER REFERENCES local_users(id) ON DELETE SET NULL,
                    buyer VARCHAR(120),
                    total_cost_record NUMERIC(12, 2),
                    purchase_cost NUMERIC(12, 2),
                    purchase_channel VARCHAR(160),
                    purchase_qty INTEGER DEFAULT 0,
                    remark TEXT,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    updated_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_items_purchase_order_id ON purchase_order_items(purchase_order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_items_product_name ON purchase_order_items(product_name)"))
        conn.execute(text("ALTER TABLE IF EXISTS purchase_order_items ADD COLUMN IF NOT EXISTS buyer_user_id INTEGER REFERENCES local_users(id) ON DELETE SET NULL"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_items_buyer_user_id ON purchase_order_items(buyer_user_id)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS purchase_order_sources (
                    id SERIAL PRIMARY KEY,
                    purchase_order_id INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
                    purchase_order_item_id INTEGER NOT NULL REFERENCES purchase_order_items(id) ON DELETE CASCADE,
                    order_id INTEGER NOT NULL,
                    order_item_id INTEGER NOT NULL UNIQUE,
                    product_id INTEGER,
                    product_name VARCHAR(255) NOT NULL,
                    quantity INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_sources_purchase_order_id ON purchase_order_sources(purchase_order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_sources_order_item_id ON purchase_order_sources(order_item_id)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS purchase_order_logs (
                    id SERIAL PRIMARY KEY,
                    purchase_order_id INTEGER,
                    purchase_no VARCHAR(40) NOT NULL,
                    action VARCHAR(40) NOT NULL,
                    operator VARCHAR(80),
                    snapshot JSONB DEFAULT '{}'::jsonb,
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_logs_purchase_order_id ON purchase_order_logs(purchase_order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_purchase_order_logs_purchase_no ON purchase_order_logs(purchase_no)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS outbound_scan_records (
                    id SERIAL PRIMARY KEY,
                    tracking_number VARCHAR(160) NOT NULL,
                    raw_input VARCHAR(255) DEFAULT '',
                    order_id INTEGER REFERENCES orders(id),
                    platform VARCHAR(40) DEFAULT '',
                    shop_name VARCHAR(160) DEFAULT '',
                    platform_order_no VARCHAR(160) DEFAULT '',
                    posting_number VARCHAR(160) DEFAULT '',
                    order_status VARCHAR(40) DEFAULT '',
                    platform_status VARCHAR(80) DEFAULT '',
                    result VARCHAR(40) DEFAULT '',
                    message TEXT DEFAULT '',
                    scanned_by VARCHAR(80) DEFAULT '',
                    scanned_at TIMESTAMP DEFAULT timezone('UTC', NOW()),
                    created_at TIMESTAMP DEFAULT timezone('UTC', NOW())
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_outbound_scan_records_tracking_number ON outbound_scan_records(tracking_number)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_outbound_scan_records_order_id ON outbound_scan_records(order_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_outbound_scan_records_scanned_at ON outbound_scan_records(scanned_at)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_outbound_scan_records_result ON outbound_scan_records(result)"))
        conn.execute(
            text(
                """
                UPDATE orders
                SET biz_status = CASE
                    WHEN biz_status IS NOT NULL THEN biz_status
                    WHEN local_status IN ('shipped') THEN '已发货'
                    WHEN local_status IN ('awaiting_pickup') THEN '待揽收'
                    WHEN local_status IN ('delivered') THEN '已妥投'
                    WHEN local_status IN ('voided', 'cancelled') THEN '已作废'
                    WHEN local_status IN ('shipment_creating', 'label_downloading', 'label_saved', 'shipment_created', 'picking') THEN '配货中'
                    ELSE '待处理'
                END
                WHERE biz_status IS NULL
                """
            )
        )
        conn.execute(
            text(
                """
                UPDATE orders
                SET biz_status = '已作废',
                    updated_at = COALESCE(updated_at, created_at, timezone('UTC', NOW()))
                WHERE LOWER(COALESCE(platform_status, '')) IN :voided_platform_statuses
                  AND COALESCE(biz_status, '') <> '已作废'
                """
            ).bindparams(bindparam("voided_platform_statuses", expanding=True)),
            {"voided_platform_statuses": ORDER_VOIDED_PLATFORM_STATUS_VALUES},
        )
        conn.execute(
            text(
                """
                UPDATE orders
                SET picking_at = COALESCE(updated_at, created_at, timezone('UTC', NOW()))
                WHERE picking_at IS NULL
                  AND (
                    biz_status = '配货中'
                    OR local_status IN ('shipment_creating', 'label_downloading', 'label_saved', 'shipment_created', 'picking')
                  )
                """
            )
        )


def _backfill_order_internal_order_no() -> None:
    with SessionLocal() as db:
        rows = db.execute(select(Order.id, Order.internal_order_no).order_by(Order.id)).all()
        seen_values: set[str] = set()
        changed = 0
        for order_id, current_value in rows:
            current = str(current_value or "").strip()
            if current and current not in seen_values:
                seen_values.add(current)
                continue
            internal_order_no = generate_internal_order_no()
            while internal_order_no in seen_values:
                internal_order_no = generate_internal_order_no()
            seen_values.add(internal_order_no)
            db.execute(
                text("UPDATE orders SET internal_order_no = :internal_order_no WHERE id = :order_id"),
                {"internal_order_no": internal_order_no, "order_id": order_id},
            )
            changed += 1
        if changed:
            db.commit()
            logger.info("Backfilled internal order numbers for %s orders", changed)


def _normalize_legacy_fbj_export_statuses() -> None:
    """Return old FBJ export statuses to the standard pending workflow."""
    with SessionLocal() as db:
        result = db.execute(
            update(Order)
            .where(Order.biz_status.in_(LEGACY_FBJ_EXPORT_STATUSES))
            .values(
                biz_status=ORDER_STATUS_PENDING,
                local_status="new",
                error_message="",
                updated_at=datetime.utcnow(),
            )
        )
        changed = int(result.rowcount or 0)
        if changed:
            db.commit()
            logger.info("Normalized %s legacy FBJ export orders to pending", changed)


def _ensure_order_number_indexes() -> None:
    statements = [
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_orders_internal_order_no ON orders(internal_order_no)",
        "CREATE EXTENSION IF NOT EXISTS pg_trgm",
        "CREATE INDEX IF NOT EXISTS ix_orders_platform_order_no_trgm ON orders USING gin (platform_order_no gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_orders_platform_order_id_trgm ON orders USING gin (platform_order_id gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_orders_posting_number_trgm ON orders USING gin (posting_number gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_orders_shipment_tracking_number_trgm ON orders USING gin (shipment_tracking_number gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_shipments_tracking_number_trgm ON shipments USING gin (tracking_number gin_trgm_ops)",
    ]
    for statement in statements:
        try:
            with engine.begin() as conn:
                conn.execute(text(statement))
        except Exception as exc:  # pragma: no cover - depends on database permissions/extensions
            logger.warning("Skipping order number fuzzy index setup: %s", exc)
            return


def _ensure_dashboard_indexes() -> None:
    statements = [
        "CREATE INDEX IF NOT EXISTS ix_orders_payment_at_id ON orders(payment_at, id)",
        "CREATE INDEX IF NOT EXISTS ix_orders_payment_month_shop ON orders(payment_at, platform, shop_id, shop_name)",
        "CREATE INDEX IF NOT EXISTS ix_orders_biz_deadline ON orders(biz_status, dispatch_deadline_at, shipping_deadline_at)",
        "CREATE INDEX IF NOT EXISTS ix_orders_pending_payment ON orders(biz_status, payment_at) WHERE biz_status IN ('待处理', '配货中')",
        "CREATE INDEX IF NOT EXISTS ix_order_items_order_quantity ON order_items(order_id, quantity)",
        "CREATE INDEX IF NOT EXISTS ix_order_items_order_currency_id ON order_items(order_id, id, currency)",
        "CREATE INDEX IF NOT EXISTS ix_exchange_rates_currency_date_updated ON exchange_rates(currency_code, rate_date DESC, updated_at DESC)",
    ]
    for statement in statements:
        try:
            with engine.begin() as conn:
                conn.execute(text(statement))
        except Exception as exc:  # pragma: no cover - depends on database permissions
            logger.warning("Skipping dashboard index setup for %s: %s", statement, exc)


def _ensure_traffic_analytics_indexes() -> None:
    statements = [
        "ALTER TABLE IF EXISTS traffic_metrics ADD COLUMN IF NOT EXISTS buyers BIGINT",
        "ALTER TABLE IF EXISTS traffic_metrics ADD COLUMN IF NOT EXISTS units_sold BIGINT",
        "ALTER TABLE IF EXISTS traffic_metrics ADD COLUMN IF NOT EXISTS negative_reviews BIGINT",
        "CREATE INDEX IF NOT EXISTS ix_traffic_metrics_grain_stat_date ON traffic_metrics(grain, stat_date)",
        "CREATE INDEX IF NOT EXISTS ix_traffic_metrics_account_grain_stat_date ON traffic_metrics(platform_account_id, grain, stat_date)",
        "CREATE INDEX IF NOT EXISTS ix_traffic_metrics_grain_period_account ON traffic_metrics(grain, period_start, period_end, platform_account_id)",
        "CREATE INDEX IF NOT EXISTS ix_traffic_sync_runs_account_latest ON traffic_sync_runs(platform_account_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_orders_traffic_account_payment ON orders(platform, account_id, payment_at) WHERE payment_at IS NOT NULL",
        "CREATE INDEX IF NOT EXISTS ix_orders_traffic_account_created ON orders(platform, account_id, platform_created_at) WHERE payment_at IS NULL",
    ]
    for statement in statements:
        try:
            with engine.begin() as conn:
                conn.execute(text(statement))
        except Exception as exc:  # pragma: no cover - depends on database permissions
            logger.warning("Skipping traffic analytics index setup for %s: %s", statement, exc)


def _ensure_order_workflow_indexes() -> None:
    statements = [
        "CREATE INDEX IF NOT EXISTS ix_orders_status_payment_page ON orders(biz_status, payment_at DESC, created_at DESC, updated_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_orders_platform_payment_page ON orders(platform, payment_at DESC, created_at DESC, updated_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_orders_payment_page ON orders(payment_at DESC, created_at DESC, updated_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_orders_picking_payment_page ON orders(picking_at, payment_at DESC, created_at DESC, updated_at DESC, id DESC) WHERE picking_at IS NOT NULL",
        "CREATE INDEX IF NOT EXISTS ix_orders_customer_history ON orders(shop_id, buyer_id, platform_created_at, id) WHERE buyer_id IS NOT NULL AND buyer_id <> ''",
        "CREATE INDEX IF NOT EXISTS ix_order_items_order_id_id ON order_items(order_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_shipments_order_id_id ON shipments(order_id, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_label_files_shipment_id_id ON label_files(shipment_id, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_success_order_scanned ON outbound_scan_records(order_id, scanned_at) WHERE result = 'success'",
        "CREATE INDEX IF NOT EXISTS ix_platform_accounts_platform_account ON platform_accounts(platform, account_id)",
        "CREATE INDEX IF NOT EXISTS ix_product_shop_mappings_shop_sku_product ON product_shop_mappings(shop_id, shop_sku, product_id)",
        "CREATE INDEX IF NOT EXISTS ix_product_shop_mappings_shop_sku_lower_product ON product_shop_mappings(shop_id, lower(trim(shop_sku)), product_id)",
        "CREATE INDEX IF NOT EXISTS ix_product_shop_mappings_product_shop ON product_shop_mappings(product_id, shop_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_products_enabled_updated_id ON products(enabled, updated_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_products_updated_id ON products(updated_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_order_sources_order_item_purchase ON purchase_order_sources(order_item_id, purchase_order_id)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_created_id ON api_request_logs(created_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_filters_created ON api_request_logs(platform, operation, status, created_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_log_date_group ON api_request_logs(log_date, platform, account_id, operation)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_account_created ON api_request_logs(account_id, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_url_trgm ON api_request_logs USING gin (url gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_account_trgm ON api_request_logs USING gin (account_id gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_operation_trgm ON api_request_logs USING gin (operation gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_api_request_logs_error_trgm ON api_request_logs USING gin (error_message gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_product_inventory_product_id_id ON product_inventory(product_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_product_inventory_stock_product ON product_inventory(stock_qty, product_id)",
        "CREATE INDEX IF NOT EXISTS ix_product_inventory_product_name_trgm ON product_inventory USING gin (product_name gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_products_product_code_trgm ON products USING gin (product_code gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_products_internal_name_trgm ON products USING gin (internal_name gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_order_items_sku_trgm ON order_items USING gin (sku gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_order_items_platform_product_name_trgm ON order_items USING gin (platform_product_name gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_orders_created_id ON purchase_orders(created_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_orders_purchase_date_created ON purchase_orders(purchase_date, created_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_orders_purchase_no_trgm ON purchase_orders USING gin (purchase_no gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_order_items_purchase_order_id_id ON purchase_order_items(purchase_order_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_order_items_product_name_trgm ON purchase_order_items USING gin (product_name gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_order_items_buyer_trgm ON purchase_order_items USING gin (buyer gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_purchase_order_sources_purchase_item_order ON purchase_order_sources(purchase_order_id, purchase_order_item_id, order_id)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_scanned_page ON outbound_scan_records(scanned_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_result_scanned ON outbound_scan_records(result, scanned_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_platform_scanned ON outbound_scan_records(platform, scanned_at DESC, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_tracking_trgm ON outbound_scan_records USING gin (tracking_number gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_shop_name_trgm ON outbound_scan_records USING gin (shop_name gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_platform_order_no_trgm ON outbound_scan_records USING gin (platform_order_no gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_posting_number_trgm ON outbound_scan_records USING gin (posting_number gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_outbound_scan_scanned_by_trgm ON outbound_scan_records USING gin (scanned_by gin_trgm_ops)",
        "CREATE INDEX IF NOT EXISTS ix_scheduled_task_runs_task_id_id ON scheduled_task_runs(scheduled_task_id, id DESC)",
        "CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_steps_run_id_id ON scheduled_task_run_steps(run_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_orders_run_id_id ON scheduled_task_run_orders(run_id, id)",
        "CREATE INDEX IF NOT EXISTS ix_scheduled_task_run_orders_run_reprint_id ON scheduled_task_run_orders(run_id, needs_reprint, id)",
    ]
    for statement in statements:
        try:
            with engine.begin() as conn:
                conn.execute(text(statement))
        except Exception as exc:  # pragma: no cover - depends on database permissions
            logger.warning("Skipping order workflow index setup for %s: %s", statement, exc)


def _backfill_order_operation_logs() -> None:
    with SessionLocal() as db:
        existing_history = db.scalar(
            select(OrderOperationLog.id)
            .where(OrderOperationLog.source == ORDER_LOG_HISTORY_SOURCE)
            .limit(1)
        )
        if existing_history:
            return
        rows = db.scalars(select(Order)).all()
        for row in rows:
            order_label = row.platform_order_no or row.posting_number or row.platform_order_id or str(row.id)
            if row.created_at:
                add_order_operation_log(
                    db,
                    order_id=row.id,
                    operation_type="order_imported",
                    operation_attribute="订单同步",
                    description=f"历史补充：订单 {order_label} 已同步到系统",
                    operator=SYSTEM_OPERATOR,
                    source=ORDER_LOG_HISTORY_SOURCE,
                    operated_at=row.created_at,
                    event_key=f"history:{row.id}:order_imported",
                )
            if row.picking_at:
                add_order_operation_log(
                    db,
                    order_id=row.id,
                    operation_type="to_picking",
                    operation_attribute="修改订单基础信息",
                    description="历史补充：订单转入配货中",
                    operator=SYSTEM_OPERATOR,
                    source=ORDER_LOG_HISTORY_SOURCE,
                    operated_at=row.picking_at,
                    event_key=f"history:{row.id}:to_picking",
                )
            if row.handover_at or row.shipment_tracking_number:
                add_order_operation_log(
                    db,
                    order_id=row.id,
                    operation_type="sync_logistics",
                    operation_attribute="同步物流信息",
                    description="历史补充：订单物流信息已同步",
                    operator=SYSTEM_OPERATOR,
                    source=ORDER_LOG_HISTORY_SOURCE,
                    operated_at=row.handover_at or row.updated_at or row.created_at,
                    event_key=f"history:{row.id}:sync_logistics",
                )
            if row.label_printed_at:
                add_order_operation_log(
                    db,
                    order_id=row.id,
                    operation_type="print_label",
                    operation_attribute="打印面单",
                    description="历史补充：订单面单已打印",
                    operator=SYSTEM_OPERATOR,
                    source=ORDER_LOG_HISTORY_SOURCE,
                    operated_at=row.label_printed_at,
                    event_key=f"history:{row.id}:print_label",
                )
            if row.marked_shipped_at or row.shipped_at:
                add_order_operation_log(
                    db,
                    order_id=row.id,
                    operation_type="mark_shipped",
                    operation_attribute="修改订单基础信息",
                    description="历史补充：订单已标记发货",
                    operator=SYSTEM_OPERATOR,
                    source=ORDER_LOG_HISTORY_SOURCE,
                    operated_at=row.marked_shipped_at or row.shipped_at,
                    event_key=f"history:{row.id}:mark_shipped",
                )

        outbound_rows = db.scalars(
            select(OutboundScanRecord).where(
                OutboundScanRecord.order_id.is_not(None),
                OutboundScanRecord.result == "success",
            )
        ).all()
        for scan in outbound_rows:
            add_order_operation_log(
                db,
                order_id=int(scan.order_id),
                operation_type="outbound_scan",
                operation_attribute="扫码出库",
                description=f"历史补充：货运单号 {scan.tracking_number or '-'} 扫码出库成功",
                operator=scan.scanned_by or SYSTEM_OPERATOR,
                source=ORDER_LOG_HISTORY_SOURCE,
                operated_at=scan.scanned_at or scan.created_at,
                event_key=f"history:outbound_scan:{scan.id}",
                extra={"tracking_number": scan.tracking_number or "", "scan_record_id": scan.id},
            )
        db.commit()


def _repair_joom_orders() -> None:
    db = next(get_db())
    try:
        rows = db.scalars(
            select(Order).where(Order.platform == "joom_logistics", Order.raw_payload != {})
        ).all()
        changed = 0
        for row in rows:
            extracted = _extract_order_fields(row.raw_payload or {})
            field_changed = False
            for attr in ("platform_created_at", "payment_at", "shipping_deadline_at", "platform_handover_deadline"):
                value = extracted.get(attr)
                if value and getattr(row, attr) != value:
                    setattr(row, attr, value)
                    field_changed = True
            tracking_number = extracted.get("shipment_tracking_number")
            if tracking_number and row.shipment_tracking_number != tracking_number:
                row.shipment_tracking_number = tracking_number
                field_changed = True
            if row.platform_order_id and row.platform_order_no != row.platform_order_id:
                row.platform_order_no = row.platform_order_id
                field_changed = True
            if field_changed:
                row.updated_at = datetime.utcnow()
                changed += 1
        if changed:
            db.commit()
            logger.info("Repaired Joom orders: %s", changed)
    finally:
        db.close()


def _repair_wildberries_order_numbers() -> None:
    db = next(get_db())
    try:
        rows = db.scalars(
            select(Order).where(Order.platform == "wildberries", Order.raw_payload != {})
        ).all()
        changed = 0
        for row in rows:
            raw_payload = dict(row.raw_payload) if isinstance(row.raw_payload, dict) else {}
            order_id = _to_str(raw_payload.get("id") or row.platform_order_id).strip()
            if not order_id:
                continue
            legacy_ids = {
                _to_str(raw_payload.get("orderUid")).strip(),
                _to_str(raw_payload.get("rid")).strip(),
            }
            legacy_ids.discard("")
            field_changed = False
            if row.platform_order_no in legacy_ids or not row.platform_order_no:
                row.platform_order_no = order_id
                field_changed = True
            if row.posting_number in ("", None) or row.posting_number in legacy_ids:
                row.posting_number = order_id
                field_changed = True
            if row.buyer_id in legacy_ids:
                row.buyer_id = None
                field_changed = True
            if row.buyer_name in legacy_ids:
                row.buyer_name = None
                field_changed = True
            for key in ("buyer_id", "customer_id"):
                if raw_payload.get(key) in legacy_ids:
                    raw_payload.pop(key, None)
                    field_changed = True
            if field_changed:
                row.raw_payload = raw_payload
                row.updated_at = datetime.utcnow()
                changed += 1
        if changed:
            db.commit()
            logger.info("Repaired Wildberries order numbers: %s", changed)
    finally:
        db.close()


def _iso(value: datetime | None) -> str | None:
    if not value:
        return None
    if value.tzinfo is None:
        return f"{value.isoformat()}Z"
    return value.isoformat()


def _platform_time_iso(value: datetime | None) -> str | None:
    if not value:
        return None
    if value.tzinfo is not None:
        value = value.astimezone(timezone.utc).replace(tzinfo=None)
    return f"{value.replace(microsecond=0).isoformat()}Z"


def _platform_time_text(value) -> str | None:
    if value in (None, ""):
        return None
    parsed = _parse_platform_datetime(value)
    if parsed:
        return _platform_time_iso(parsed)
    text_value = str(value).strip()
    return text_value or None


def _parse_platform_datetime(value) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text_value = str(value).strip()
        if not text_value:
            return None
        text_value = text_value.replace(" ", "T").replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text_value)
        except ValueError:
            return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed.replace(microsecond=0)


def _is_mercado_payload(raw_payload: dict) -> bool:
    return bool(
        str(raw_payload.get("marketplace") or "").lower() == "mercadolibre"
        or raw_payload.get("mercado_api_mode")
        or raw_payload.get("mercado_store_type")
    )


def _shipping_deadline_value(raw_payload: dict, order: dict, shipping: dict, shipment: dict):
    mercado_expiration_date = _first_value(raw_payload.get("expiration_date"), order.get("expiration_date")) if _is_mercado_payload(raw_payload) else None
    return _first_value(
        raw_payload.get("last_ship_date"),
        raw_payload.get("ship_by_date"),
        raw_payload.get("delivery_date_end"),
        raw_payload.get("shipping_deadline_at"),
        mercado_expiration_date,
        ((shipping.get("lead_time") or {}).get("estimated_delivery_time") or {}).get("pay_before") if isinstance(shipping.get("lead_time"), dict) else None,
        ((shipment.get("lead_time") or {}).get("estimated_delivery_time") or {}).get("pay_before") if isinstance(shipment.get("lead_time"), dict) else None,
    )


def _platform_import_time(raw_payload: dict, field: str, fallback: datetime | None = None) -> str | None:
    raw_payload = raw_payload or {}
    order = _first_order(raw_payload)
    payment = _first_payment(raw_payload)
    shipping = raw_payload.get("shipping") if isinstance(raw_payload.get("shipping"), dict) else {}
    shipment = raw_payload.get("shipment") if isinstance(raw_payload.get("shipment"), dict) else {}
    if field == "payment_at":
        value = _first_value(
            payment.get("date_approved"),
            payment.get("date_created"),
            order.get("date_closed"),
            raw_payload.get("in_process_at"),
            raw_payload.get("payment_at"),
            raw_payload.get("created_at"),
            raw_payload.get("order_date"),
            raw_payload.get("date_created"),
            order.get("date_created"),
        )
    elif field == "platform_created_at":
        value = _first_value(
            raw_payload.get("created_at"),
            raw_payload.get("order_date"),
            raw_payload.get("date_created"),
            order.get("date_created"),
            raw_payload.get("in_process_at"),
        )
    elif field == "platform_handover_deadline":
        value = _first_value(
            raw_payload.get("shipment_date"),
            raw_payload.get("platform_handover_deadline"),
            raw_payload.get("ship_by_date"),
            raw_payload.get("delivery_date_begin"),
            shipping.get("date_first_printed"),
        )
    elif field == "shipping_deadline_at":
        value = _shipping_deadline_value(raw_payload, order, shipping, shipment)
    else:
        value = None
    return _platform_time_text(value) or _platform_time_iso(fallback)


def _effective_shipping_deadline(row: Order, extracted: dict) -> datetime | None:
    raw_payload = row.raw_payload or {}
    if _is_mercado_payload(raw_payload) and extracted.get("shipping_deadline_at"):
        return extracted["shipping_deadline_at"]
    return row.shipping_deadline_at or extracted.get("shipping_deadline_at")


def _order_extra_time(db: Session | None, order_id: int, column: str) -> datetime | None:
    if db is None:
        return None
    if column not in {"picking_at", "marked_shipped_at", "label_printed_at"}:
        return None
    return db.execute(text(f"SELECT {column} FROM orders WHERE id = :order_id"), {"order_id": order_id}).scalar_one_or_none()


def _coalesce_model_time(row: Order, db: Session | None, column: str) -> datetime | None:
    return getattr(row, column, None) or _order_extra_time(db, row.id, column)


def _shop_id(payload_or_row) -> str:
    return getattr(payload_or_row, "shop_id", None) or getattr(payload_or_row, "account_id", None) or ""


def _platform_display_name(platform: str, fallback: str = "") -> str:
    return PLATFORM_DISPLAY_NAMES.get(platform, fallback or platform)


def _platform_catalog_entries() -> list[tuple[str, str, int]]:
    seen: set[str] = set()
    entries: list[tuple[str, str, int]] = []
    for index, item in enumerate(PLATFORM_CATALOG):
        platform = _canonical_platform(str(item.get("platform") or ""))
        if not platform or platform in seen:
            continue
        seen.add(platform)
        entries.append((platform, _platform_display_name(platform, str(item.get("display_name") or platform)), index))
    return entries


def seed_default_platform_settings(db: Session) -> int:
    existing = {
        row.platform: row
        for row in db.scalars(select(PlatformSetting)).all()
    }
    now = datetime.utcnow()
    changed = 0
    for platform, platform_name, sort_order in _platform_catalog_entries():
        row = existing.get(platform)
        if row:
            if (row.platform_name or "") != platform_name or int(row.sort_order or 0) != sort_order:
                row.platform_name = platform_name
                row.sort_order = sort_order
                row.updated_at = now
                changed += 1
            continue
        db.add(
            PlatformSetting(
                platform=platform,
                platform_name=platform_name,
                enabled=True,
                sort_order=sort_order,
                created_at=now,
                updated_at=now,
            )
        )
        changed += 1
    return changed


def _shop_code_prefix(platform: str) -> str:
    return SHOP_CODE_PREFIXES.get(platform, platform)


def _generate_shop_account_id(db: Session, platform: str) -> str:
    prefix = _shop_code_prefix(platform)
    pattern = re.compile(rf"^{re.escape(prefix)}(\d{{4}})$")
    account_ids = db.scalars(select(PlatformAccount.account_id).where(PlatformAccount.platform == platform)).all()
    existing = set(account_ids)
    max_serial = 0
    for account_id in account_ids:
        match = pattern.match(account_id or "")
        if match:
            max_serial = max(max_serial, int(match.group(1)))
    serial = max_serial + 1
    while serial <= 9999:
        candidate = f"{prefix}{serial:04d}"
        if candidate not in existing:
            return candidate
        serial += 1
    raise HTTPException(status_code=400, detail="店铺编码流水号已用完")


def _parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    normalized = normalized.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed
    return parsed.astimezone(timezone.utc).replace(tzinfo=None)


def _first_value(*values):
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _as_list(value) -> list:
    return value if isinstance(value, list) else []


def _first_dict(*values) -> dict:
    for value in values:
        if isinstance(value, dict):
            return value
    return {}


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def _tracking_number_from_payload(raw_payload: dict) -> str:
    if not isinstance(raw_payload, dict):
        return ""
    shipping = raw_payload.get("shipping") if isinstance(raw_payload.get("shipping"), dict) else {}
    shipment = raw_payload.get("shipment") if isinstance(raw_payload.get("shipment"), dict) else {}
    logistics = raw_payload.get("logistics") if isinstance(raw_payload.get("logistics"), dict) else {}
    tracking = raw_payload.get("tracking") if isinstance(raw_payload.get("tracking"), dict) else {}
    delivery = raw_payload.get("delivery") if isinstance(raw_payload.get("delivery"), dict) else {}
    shipments_payload = raw_payload.get("shipments_payload") if isinstance(raw_payload.get("shipments_payload"), dict) else {}

    def first_waybill(values) -> str:
        if not isinstance(values, list):
            return ""
        for item in values:
            if not isinstance(item, dict):
                continue
            value = _first_value(
                item.get("waybill"),
                item.get("waybillNumber"),
                item.get("waybill_number"),
                item.get("trackingNumber"),
                item.get("tracking_number"),
                item.get("trackingNo"),
                item.get("tracking_no"),
            )
            text = _to_str(value).strip()
            if text:
                return text
        return ""
    tracking_number = _to_str(
        _first_value(
            raw_payload.get("shipment_tracking_number"),
            raw_payload.get("tracking_number"),
            raw_payload.get("track_number"),
            raw_payload.get("trackingNumber"),
            raw_payload.get("trackNumber"),
            raw_payload.get("trackingNo"),
            raw_payload.get("tracking_no"),
            raw_payload.get("waybillNumber"),
            raw_payload.get("waybill_number"),
            shipment.get("shipment_tracking_number"),
            shipment.get("tracking_number"),
            shipment.get("track_number"),
            shipment.get("trackingNumber"),
            shipment.get("trackNumber"),
            shipment.get("trackingNo"),
            shipment.get("tracking_no"),
            shipment.get("waybillNumber"),
            shipment.get("waybill_number"),
            shipping.get("shipment_tracking_number"),
            shipping.get("tracking_number"),
            shipping.get("track_number"),
            shipping.get("trackingNumber"),
            shipping.get("trackNumber"),
            shipping.get("trackingNo"),
            shipping.get("tracking_no"),
            shipping.get("waybillNumber"),
            shipping.get("waybill_number"),
            logistics.get("shipment_tracking_number"),
            logistics.get("tracking_number"),
            logistics.get("trackingNumber"),
            logistics.get("number"),
            tracking.get("shipment_tracking_number"),
            tracking.get("tracking_number"),
            tracking.get("trackingNumber"),
            tracking.get("number"),
            delivery.get("trackingNumber"),
            delivery.get("tracking_number"),
            first_waybill(raw_payload.get("shipments")),
            first_waybill(shipping.get("shipments")),
            first_waybill(delivery.get("shipments")),
            first_waybill(shipments_payload.get("shipments")),
        )
    ).strip()
    return clean_tracking_number(tracking_number, raw_payload)


def _to_int_list(value: str | None) -> list[int]:
    if not value:
        return []
    result: list[int] = []
    for item in value.split(","):
        item = item.strip()
        if not item:
            continue
        try:
            result.append(int(item))
        except ValueError:
            continue
    return result


def _shop_authorization_message(valid: bool) -> str:
    return "Token 授权成功" if valid else "令牌无效，请检查"


def _oauth_platform(platform: str) -> str:
    normalized = (platform or "").lower()
    if normalized in {"joom", "joomlogistics", "joom_logistics"}:
        return "joom_logistics"
    if normalized == "mercadolibre":
        return "mercadolibre"
    if normalized == "allegro":
        return "allegro"
    return normalized


def _canonical_platform(platform: str) -> str:
    normalized = (platform or "").strip().lower()
    return PLATFORM_ALIASES.get(normalized, normalized)


def _platform_lookup_codes(platform: str) -> set[str]:
    canonical = _canonical_platform(platform)
    return {canonical} if canonical else set()


def _find_shop(db: Session, platform: str, account_id: str) -> PlatformAccount | None:
    rows = db.scalars(
        select(PlatformAccount).where(
            PlatformAccount.platform.in_(_platform_lookup_codes(platform)),
            PlatformAccount.account_id == account_id,
        )
    ).all()
    if not rows:
        return None
    requested = (platform or "").strip().lower()
    canonical = _canonical_platform(platform)
    return (
        next((row for row in rows if row.platform == requested), None)
        or next((row for row in rows if row.platform == canonical), None)
        or rows[0]
    )


def _oauth_client_secret(credentials: dict) -> str:
    return str(credentials.get("client_secret") or credentials.get("api_key") or "").strip()


def _normalize_shop_credentials(platform: str, credentials: dict | None) -> dict:
    data = dict(credentials or {})
    if _oauth_platform(platform) in {"joom_logistics", "mercadolibre", "allegro"} and not data.get("client_secret") and data.get("api_key"):
        data["client_secret"] = data.get("api_key")
    return data


def _merge_non_empty_credentials(base: dict | None, updates: dict | None) -> dict:
    result = dict(base or {})
    for key, value in (updates or {}).items():
        if value not in (None, ""):
            result[key] = value
    return result


def _is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _scheduled_task_pdf_allowed_roots() -> list[Path]:
    return [get_settings().label_storage_path.resolve()]


def _safe_scheduled_task_pdf_path(raw_path: str, allowed_roots: list[Path]) -> Path | None:
    raw_path = (raw_path or "").strip()
    if not raw_path:
        return None
    path = Path(raw_path)
    if not path.is_absolute():
        path = allowed_roots[0] / path
    try:
        resolved = path.resolve()
    except (OSError, ValueError):
        return None
    if not any(_is_relative_to(resolved, root) for root in allowed_roots):
        return None
    if resolved.suffix.lower() != ".pdf" or not resolved.is_file():
        return None
    return resolved


def _scheduled_task_run_pdf_entries(db: Session, run_id: int) -> list[tuple[str, Path]]:
    rows = db.scalars(
        select(ScheduledTaskRunOrder)
        .where(
            ScheduledTaskRunOrder.run_id == run_id,
            ScheduledTaskRunOrder.pdf_generated == True,
            ScheduledTaskRunOrder.pdf_file_path != "",
        )
        .order_by(asc(ScheduledTaskRunOrder.id))
    ).all()

    allowed_roots = _scheduled_task_pdf_allowed_roots()
    pdf_entries: list[tuple[str, Path]] = []
    seen_paths: set[Path] = set()
    for row in rows:
        resolved = _safe_scheduled_task_pdf_path(row.pdf_file_path or "", allowed_roots)
        if not resolved or resolved in seen_paths:
            continue
        seen_paths.add(resolved)
        platform = (row.platform or "unknown").strip() or "unknown"
        pdf_entries.append((platform, resolved))

    backup_name = f"run-{run_id}.pdf"
    for root in allowed_roots:
        backup_root = root / "system" / "scheduled-task"
        if not backup_root.is_dir():
            continue
        for path in sorted(backup_root.glob(f"*/**/{backup_name}")):
            try:
                resolved = path.resolve()
            except (OSError, ValueError):
                continue
            if resolved in seen_paths or not _is_relative_to(resolved, root) or not resolved.is_file():
                continue
            parts = resolved.relative_to(backup_root).parts
            platform = parts[0] if parts else "unknown"
            seen_paths.add(resolved)
            pdf_entries.append((platform, resolved))
    return pdf_entries


def _label_print_filename_stem(timestamp: str) -> str:
    return f"label_print_{timestamp}"


def _scheduled_task_run_pdf_filename(timestamp: str) -> str:
    return f"{_label_print_filename_stem(timestamp)}.zip"


def _safe_archive_filename_stem(value: str, fallback: str = "unknown") -> str:
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", (value or "").strip())
    stem = re.sub(r"\s+", " ", stem).strip(" ._")
    return stem or fallback


def _content_disposition_attachment(filename: str) -> str:
    ascii_filename = "".join(ch if ord(ch) < 128 and ch not in {'"', "\\"} else "_" for ch in filename) or "download"
    return f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{quote(filename)}"


def _scheduled_task_run_pdf_entries_or_raise(run_id: int, db: Session) -> list[tuple[str, Path]]:
    run = db.get(ScheduledTaskRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="任务运行记录不存在")

    pdf_entries = _scheduled_task_run_pdf_entries(db, run_id)
    if not pdf_entries:
        raise HTTPException(status_code=404, detail=f"任务运行 {run_id} 没有可导出的PDF")
    return pdf_entries


def _build_scheduled_task_run_pdf_archive(run_id: int, db: Session) -> tuple[io.BytesIO, str]:
    pdf_entries = _scheduled_task_run_pdf_entries_or_raise(run_id, db)
    output = io.BytesIO()
    used_names: set[str] = set()
    platform_name_counts: dict[str, int] = {}
    timestamp = _local_now().strftime("%Y%m%d_%H%M%S")
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for platform, path in pdf_entries:
            platform_name = _safe_archive_filename_stem(_platform_display_name(platform, platform))
            platform_name_counts[platform_name] = platform_name_counts.get(platform_name, 0) + 1
            duplicate_index = platform_name_counts[platform_name]
            arcname = f"{platform_name}_{timestamp}{f'_{duplicate_index}' if duplicate_index > 1 else ''}.pdf"
            while arcname.lower() in used_names:
                duplicate_index += 1
                arcname = f"{platform_name}_{timestamp}_{duplicate_index}.pdf"
            used_names.add(arcname.lower())
            archive.write(path, arcname)
    output.seek(0)
    return output, _scheduled_task_run_pdf_filename(timestamp)


def _scheduled_task_run_pdf_response(run_id: int, db: Session) -> Response:
    output, filename = _build_scheduled_task_run_pdf_archive(run_id, db)
    content = output.getvalue()
    headers = {
        "Content-Disposition": _content_disposition_attachment(filename),
        "Content-Length": str(len(content)),
        "Cache-Control": "no-store",
    }
    return Response(content=content, media_type="application/octet-stream", headers=headers)


def _verify_shop_credentials(platform: str, credentials: dict) -> tuple[bool, list[str], str]:
    oauth_platform = _oauth_platform(platform)
    if oauth_platform == "mercadolibre":
        missing = []
        if not credentials.get("access_token"):
            missing.append("access_token")
        if not credentials.get("seller_id"):
            missing.append("seller_id")
        valid = not missing
        if valid:
            return True, [], "Token 授权成功"
        return False, missing, "请发起 MercadoLibre 在线授权"
    if oauth_platform in {"joom_logistics", "mercadolibre", "allegro"}:
        missing = []
        if not credentials.get("access_token"):
            missing.append("access_token")
        if oauth_platform == "mercadolibre" and not credentials.get("seller_id"):
            missing.append("seller_id")
        valid = not missing
        if valid:
            return True, [], "Token 授权成功"
        if "access_token" in missing:
            return False, missing, f"请发起 {oauth_platform} 在线授权"
        return False, missing, "令牌无效，请检查"
    required = {
        "ozon": ["client_id", "api_key"],
        "wildberries": ["api_key"],
        "allegro": ["access_token"],
        "mercadolibre": ["access_token", "seller_id"],
        "amazon": ["lwa_client_id", "lwa_client_secret", "refresh_token", "aws_access_key_id", "aws_secret_access_key"],
        "shopee": ["partner_id", "partner_key", "shop_id", "access_token"],
        "tiktok_shop": ["app_key", "app_secret", "shop_cipher", "access_token"],
        "aliexpress": ["app_key", "app_secret", "access_token"],
        "lazada": ["app_key", "app_secret", "access_token"],
        "shopify": ["shop_domain", "access_token"],
        "ebay": ["access_token"],
        "walmart": ["client_id", "client_secret"],
        "temu": ["app_key", "app_secret", "access_token"],
        "shein": ["open_key_id", "secret_key"],
        "coupang": ["access_key", "secret_key", "vendor_id"],
        "wayfair": ["client_id", "client_secret"],
        "dmsmatrix": ["client_name", "client_id", "client_secret", "channel_code"],
    }.get(platform, [])
    missing = [key for key in required if not credentials.get(key)]
    valid = not missing
    return valid, missing, _shop_authorization_message(valid)


def _apply_shop_authorization_result(
    row: PlatformAccount,
    credentials: dict,
    authorization_expires_at: str | None = None,
) -> tuple[bool, list[str]]:
    valid, missing, message = _verify_shop_credentials(row.platform, credentials)
    oauth_pending = _oauth_platform(row.platform) in {"joom_logistics", "mercadolibre", "allegro"} and "access_token" in missing
    row.status = "active" if valid or oauth_pending else "auth_error"
    row.authorization_status = SHOP_AUTH_SUCCESS if valid else (SHOP_AUTH_UNAUTHORIZED if oauth_pending else SHOP_AUTH_FAILED)
    row.token_valid = valid if not oauth_pending else None
    row.token_message = message
    if valid:
        row.last_authorized_at = datetime.utcnow()
    expires_at = _parse_datetime(authorization_expires_at)
    if expires_at:
        row.authorization_expires_at = expires_at
        row.session_expires_at = expires_at
    return valid, missing


def _shop_dto(row: PlatformAccount) -> ShopDto:
    return ShopDto(
        id=row.id,
        platform=_canonical_platform(row.platform),
        account_id=row.account_id,
        shop_id=row.account_id,
        display_name=row.display_name,
        enabled=row.enabled,
        credential_type=row.credential_type,
        status=row.status,
        authorization_status=row.authorization_status or SHOP_AUTH_UNAUTHORIZED,
        token_valid=row.token_valid,
        token_message=row.token_message or "未验证",
        last_authorized_at=_iso(row.last_authorized_at),
        authorization_expires_at=_iso(row.authorization_expires_at),
        session_expires_at=_iso(row.session_expires_at),
        last_sync_at=_iso(row.last_sync_at),
        last_sync_status=row.last_sync_status,
        credentials_version=row.credentials_version or "",
        created_by=row.created_by,
        created_at=_iso(row.created_at),
        settings=row.settings or {},
        updated_at=_iso(row.updated_at) or "",
    )


def _normalize_logistics_carrier_code(value: str) -> str:
    normalized = (value or "").strip().lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return normalized


def _mask_secret_value(value: object) -> str:
    text_value = "" if value is None else str(value)
    if not text_value:
        return ""
    if len(text_value) <= 8:
        return "*" * len(text_value)
    return f"{text_value[:3]}{'*' * max(4, len(text_value) - 7)}{text_value[-4:]}"


def _mask_credentials(credentials: dict) -> dict:
    masked: dict = {}
    for key, value in (credentials or {}).items():
        lowered = str(key).lower()
        if any(token in lowered for token in ("token", "secret", "key", "password", "令牌", "密钥")):
            masked[key] = _mask_secret_value(value)
        else:
            masked[key] = value
    return masked


def _logistics_credentials(row: LogisticsAuthorization) -> dict:
    return get_credential_manager().decrypt_credentials(row.encrypted_credentials) if row.encrypted_credentials else {}


def _verify_logistics_credentials(carrier_code: str, credentials: dict) -> tuple[bool, list[str], str]:
    required = LOGISTICS_AUTH_FIELD_SCHEMAS.get(carrier_code, [])
    missing = [field for field in required if not str((credentials or {}).get(field) or "").strip()]
    if missing:
        return False, missing, f"缺少授权字段：{', '.join(missing)}"
    if not credentials:
        return False, [], "缺少授权信息"
    return True, [], "物流授权信息完整"


def _apply_logistics_authorization_result(
    row: LogisticsAuthorization,
    credentials: dict,
    authorization_expires_at: str | None = None,
) -> tuple[bool, list[str], str]:
    valid, missing, message = _verify_logistics_credentials(row.carrier_code, credentials)
    row.authorization_status = LOGISTICS_AUTH_SUCCESS if valid else LOGISTICS_AUTH_FAILED
    row.token_valid = valid
    row.token_message = message
    if valid:
        row.last_authorized_at = datetime.utcnow()
    expires_at = _parse_datetime(authorization_expires_at)
    if expires_at:
        row.authorization_expires_at = expires_at
    return valid, missing, message


def _logistics_authorization_dto(row: LogisticsAuthorization) -> LogisticsAuthorizationDto:
    credentials = _logistics_credentials(row)
    return LogisticsAuthorizationDto(
        id=row.id,
        carrier_code=row.carrier_code,
        carrier_name=row.carrier_name or "",
        account_name=row.account_name or "",
        enabled=row.enabled,
        authorization_status=row.authorization_status or LOGISTICS_AUTH_UNAUTHORIZED,
        token_valid=row.token_valid,
        token_message=row.token_message,
        credential_type=row.credential_type or "api_key",
        credentials_masked=_mask_credentials(credentials),
        config_json=row.config_json or {},
        settings_json=row.settings_json or {},
        last_authorized_at=_iso(row.last_authorized_at),
        authorization_expires_at=_iso(row.authorization_expires_at),
        credentials_version=row.credentials_version or "",
        created_by=row.created_by,
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _logistics_channel_option_dto(row: LogisticsAuthorization) -> LogisticsChannelOptionDto:
    carrier_name = (row.carrier_name or "").strip()
    carrier_code = (row.carrier_code or "").strip()
    account_name = (row.account_name or "").strip()
    value = " / ".join(item for item in [carrier_name or carrier_code, account_name] if item)
    details = [carrier_code] if carrier_code and carrier_code != carrier_name else []
    return LogisticsChannelOptionDto(
        value=value,
        label=f"{value} / {' / '.join(details)}" if details else value,
        carrier_code=carrier_code,
        carrier_name=carrier_name,
        account_name=account_name,
    )


def _enabled_logistics_channel_options(rows: list[LogisticsAuthorization]) -> list[LogisticsChannelOptionDto]:
    return [
        _logistics_channel_option_dto(row)
        for row in rows
        if row.enabled and (row.carrier_name or row.carrier_code or "").strip()
    ]


def _add_logistics_shop_option(options: dict[str, dict[str, str]], value: object, label: object = "") -> None:
    option_value = str(value or "").strip()
    if not option_value:
        return
    option_label = str(label or "").strip() or option_value
    key = option_value.casefold()
    if key not in options:
        options[key] = {"value": option_value, "label": option_label}


def _logistics_shop_options_for_platform(
    db: Session,
    platform: str,
) -> list[dict[str, str]]:
    platform_code = normalize_platform_code(platform)
    if not platform_code:
        raise HTTPException(status_code=400, detail="请选择平台")

    options: dict[str, dict[str, str]] = {}
    platform_expr = func.lower(func.trim(PlatformAccount.platform))
    accounts = db.scalars(
        select(PlatformAccount)
        .where(platform_expr == platform_code)
        .order_by(asc(PlatformAccount.display_name), asc(PlatformAccount.account_id), asc(PlatformAccount.id))
    ).all()
    for account in accounts:
        display_name = (account.display_name or "").strip()
        account_id = (account.account_id or "").strip()
        value = display_name or account_id
        label = f"{display_name} / {account_id}" if display_name and account_id and display_name != account_id else value
        _add_logistics_shop_option(options, value, label)

    order_platform_expr = func.lower(func.trim(Order.platform))
    order_rows = db.execute(
        select(Order.shop_name, Order.shop_id, Order.account_id)
        .where(order_platform_expr == platform_code)
        .distinct()
        .order_by(asc(Order.shop_name), asc(Order.shop_id), asc(Order.account_id))
        .limit(1000)
    ).all()
    for shop_name, shop_id, account_id in order_rows:
        display_name = str(shop_name or "").strip()
        shop_code = str(shop_id or account_id or "").strip()
        value = display_name or shop_code
        label = f"{display_name} / {shop_code}" if display_name and shop_code and display_name != shop_code else value
        _add_logistics_shop_option(options, value, label)

    return sorted(options.values(), key=lambda item: item["label"].casefold())


def _logistics_match_rule_dto(row: LogisticsMatchRule) -> LogisticsMatchRuleDto:
    return LogisticsMatchRuleDto(
        id=row.id,
        name=row.name or "",
        platform=normalize_platform_code(row.platform or ""),
        priority=int(row.priority or 10),
        enabled=bool(row.enabled),
        shop_names=normalize_shop_names(row.shop_names or []),
        is_overseas_warehouse=row.is_overseas_warehouse,
        country_codes=normalize_country_codes(row.country_codes or []),
        logistics_channel=row.logistics_channel or "",
        carrier_code=_normalize_logistics_carrier_code(row.carrier_code or ""),
        remark=row.remark or "",
        created_by=row.created_by,
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _logistics_authorization_for_rule_channel(db: Session, logistics_channel: str) -> LogisticsAuthorization | None:
    rows = db.scalars(
        select(LogisticsAuthorization)
        .where(LogisticsAuthorization.enabled == True)
        .order_by(asc(LogisticsAuthorization.id))
    ).all()
    for authorization in rows:
        if _logistics_channel_option_dto(authorization).value == logistics_channel:
            return authorization
    return None


def _apply_logistics_match_rule_payload(
    db: Session,
    row: LogisticsMatchRule,
    payload: LogisticsMatchRulePayload,
) -> None:
    name = (payload.name or "").strip()
    platform = normalize_platform_code(payload.platform)
    logistics_channel = (payload.logistics_channel or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="规则名称不能为空")
    if not platform:
        raise HTTPException(status_code=400, detail="请选择平台")
    if not logistics_channel:
        raise HTTPException(status_code=400, detail="物流渠道不能为空")
    authorization = _logistics_authorization_for_rule_channel(db, logistics_channel)
    if not authorization:
        raise HTTPException(status_code=400, detail="请选择已启用的物流授权")
    row.name = name
    row.platform = platform
    row.priority = int(payload.priority or 10)
    row.enabled = bool(payload.enabled)
    row.shop_names = normalize_shop_names(payload.shop_names or [])
    row.is_overseas_warehouse = payload.is_overseas_warehouse
    row.country_codes = normalize_country_codes(payload.country_codes or [])
    row.logistics_channel = logistics_channel
    row.carrier_code = _normalize_logistics_carrier_code(authorization.carrier_code)
    row.remark = (payload.remark or "").strip()
    row.updated_at = datetime.utcnow()


def _rematch_orders(
    db: Session,
    rows: list[Order],
    *,
    include_manual: bool = False,
) -> LogisticsRematchResponse:
    rules = load_enabled_logistics_rules(db)
    matched = 0
    unmatched = 0
    skipped = 0
    now = datetime.utcnow()
    for row in rows:
        if not include_manual and row.logistics_match_status == LOGISTICS_MATCH_STATUS_MANUAL:
            skipped += 1
            continue
        result = match_logistics_rule(row, rules)
        apply_logistics_match_result(row, result, matched_at=now)
        if result.status == "matched":
            matched += 1
        else:
            unmatched += 1
    return LogisticsRematchResponse(
        matched=matched,
        unmatched=unmatched,
        skipped=skipped,
        total=len(rows),
        message=f"已重新匹配 {len(rows)} 个订单，命中 {matched} 个，未匹配 {unmatched} 个，跳过 {skipped} 个",
    )


def _seed_default_logistics_authorizations(db: Session) -> None:
    for item in LOGISTICS_AUTH_SEED_DATA:
        carrier_code = item["carrier_code"]
        account_name = item["account_name"]
        config_json = {}
        existing = db.scalar(
            select(LogisticsAuthorization).where(
                LogisticsAuthorization.carrier_code == carrier_code,
                LogisticsAuthorization.account_name == account_name,
            )
        )
        if existing:
            existing.enabled = False
            existing.encrypted_credentials = None
            existing.config_json = {}
            existing.settings_json = {}
            existing.authorization_status = "unauthorized"
            existing.token_valid = None
            existing.token_message = ""
            existing.updated_at = datetime.utcnow()
            continue
        credentials: dict[str, str] = {}
        row = LogisticsAuthorization(
            carrier_code=carrier_code,
            carrier_name=item["carrier_name"],
            account_name=account_name,
            enabled=False,
            credential_type=item.get("credential_type") or "api_key",
            encrypted_credentials=None,
            config_json=config_json,
            settings_json=dict(item.get("settings_json") or {}),
            credentials_version=datetime.utcnow().isoformat(),
            created_by="system",
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        _apply_logistics_authorization_result(row, credentials)
        db.add(row)


def _backfill_logistics_rule_carrier_codes(db: Session) -> int:
    """Populate durable carrier codes for rules created before carrier codes were stored."""
    authorizations = db.scalars(select(LogisticsAuthorization).order_by(asc(LogisticsAuthorization.id))).all()
    channel_codes = {
        _logistics_channel_option_dto(row).value: _normalize_logistics_carrier_code(row.carrier_code)
        for row in authorizations
        if _logistics_channel_option_dto(row).value and _normalize_logistics_carrier_code(row.carrier_code)
    }
    changed = 0
    for rule in db.scalars(select(LogisticsMatchRule)).all():
        carrier_code = channel_codes.get((rule.logistics_channel or "").strip(), "")
        if carrier_code and _normalize_logistics_carrier_code(rule.carrier_code) != carrier_code:
            rule.carrier_code = carrier_code
            rule.updated_at = datetime.utcnow()
            changed += 1
    return changed


def _backfill_order_logistics_carrier_codes(db: Session) -> int:
    """Preserve the carrier selected by legacy matched and manually routed orders."""
    authorizations = db.scalars(select(LogisticsAuthorization).order_by(asc(LogisticsAuthorization.id))).all()
    channel_codes = {
        _logistics_channel_option_dto(row).value: _normalize_logistics_carrier_code(row.carrier_code)
        for row in authorizations
        if _logistics_channel_option_dto(row).value and _normalize_logistics_carrier_code(row.carrier_code)
    }
    rule_codes = {
        row.id: _normalize_logistics_carrier_code(row.carrier_code)
        for row in db.scalars(select(LogisticsMatchRule)).all()
        if row.id and _normalize_logistics_carrier_code(row.carrier_code)
    }
    changed = 0
    rows = db.scalars(
        select(Order).where(or_(Order.logistics_carrier_code.is_(None), Order.logistics_carrier_code == ""))
    ).all()
    for row in rows:
        carrier_code = rule_codes.get(row.logistics_match_rule_id)
        if not carrier_code:
            carrier_code = channel_codes.get((row.logistics_channel or "").strip(), "")
        if not carrier_code:
            continue
        row.logistics_carrier_code = carrier_code
        row.updated_at = datetime.utcnow()
        changed += 1
    return changed


def _derive_order_status(row: Order) -> str:
    if _is_voided_platform_status(getattr(row, "platform_status", None)):
        return ORDER_STATUS_VOIDED
    if row.biz_status:
        return row.biz_status
    mapping = {
        "shipped": ORDER_STATUS_SHIPPED,
        "awaiting_pickup": ORDER_STATUS_AWAITING_PICKUP,
        "delivered": ORDER_STATUS_DELIVERED,
        "voided": ORDER_STATUS_VOIDED,
        "cancelled": ORDER_STATUS_VOIDED,
        "shipment_creating": ORDER_STATUS_PICKING,
        "label_downloading": ORDER_STATUS_PICKING,
        "label_saved": ORDER_STATUS_PICKING,
        "shipment_created": ORDER_STATUS_PICKING,
        "picking": ORDER_STATUS_PICKING,
    }
    return mapping.get(row.local_status, ORDER_STATUS_PENDING)


def _is_voided_platform_status(platform_status: str | None) -> bool:
    return (platform_status or "").strip().lower() in VOIDED_PLATFORM_STATUSES


def _voided_platform_status_condition():
    return func.lower(func.coalesce(Order.platform_status, "")).in_(ORDER_VOIDED_PLATFORM_STATUS_VALUES)


def _has_purchase_order_for_order(order_id: int):
    return exists().where(PurchaseOrderSource.order_id == order_id)


def _waiting_purchase_condition():
    return and_(
        Order.biz_status == ORDER_STATUS_WAITING_PURCHASE,
        or_(Order.label_printed_at.is_not(None), Order.is_overseas_warehouse == True),
        ~_has_purchase_order_for_order(Order.id),
    )


def _normalize_order_status_filter(status_filter: str | None) -> str | None:
    status_filter = (status_filter or "").strip()
    if not status_filter or status_filter in {ORDER_STATUS_ALL, "全部订单"}:
        return None
    if status_filter in {"awaiting_delivery", "待配送"}:
        return ORDER_STATUS_AWAITING_PICKUP
    return ORDER_STATUS_KEY_TO_LABEL.get(status_filter, status_filter)


def _status_key_for_label(label: str | None) -> str:
    if not label:
        return ORDER_STATUS_PENDING_KEY
    return ORDER_STATUS_LABEL_TO_KEY.get(label, ORDER_STATUS_PENDING_KEY)


def _should_show_remaining_shipping(status: str | None) -> bool:
    return _status_key_for_label(status) in {
        ORDER_STATUS_PENDING_KEY,
        ORDER_STATUS_WAITING_PRINT_KEY,
        ORDER_STATUS_WAITING_PURCHASE_KEY,
        ORDER_STATUS_PICKING_KEY,
    }


def _normalize_order_risk_filter(value: str | None) -> str | None:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return None
    if normalized not in ORDER_RISK_FILTERS:
        raise HTTPException(status_code=400, detail="风险筛选条件无效")
    return normalized


def _risk_deadline_expression():
    return func.coalesce(Order.dispatch_deadline_at, Order.shipping_deadline_at)


def _risk_filter_conditions(risk_filter: str | None, shop: str | None = None) -> list:
    normalized = _normalize_order_risk_filter(risk_filter)
    if not normalized:
        return []

    now = datetime.utcnow()
    deadline = _risk_deadline_expression()
    deadline_24h = now + timedelta(hours=24)
    handling_exists = exists().where(OrderRiskHandling.order_id == Order.id)
    conditions = [
        Order.biz_status.in_(ORDER_RISK_BIZ_STATUSES),
        deadline.is_not(None),
        deadline < deadline_24h,
    ]
    if normalized == ORDER_RISK_UNHANDLED:
        conditions.append(~handling_exists)
    elif normalized == ORDER_RISK_HANDLED:
        conditions.append(handling_exists)
    elif normalized == ORDER_RISK_OVERDUE:
        conditions.append(deadline < now)
    elif normalized == ORDER_RISK_DUE_24:
        conditions.extend([deadline >= now, deadline < deadline_24h])
    if shop and shop.strip():
        shop_value = shop.strip()
        conditions.append(or_(Order.shop_id == shop_value, Order.shop_name == shop_value))
    return conditions


def _risk_shop_scope_conditions(shop_keys: list[tuple[str, str]] | None) -> list:
    conditions = []
    for platform, account_id in shop_keys or []:
        platform_values = {
            platform.lower(),
            *(alias.lower() for alias, canonical in PLATFORM_ALIASES.items() if canonical == platform),
        }
        conditions.append(
            and_(
                func.lower(func.trim(func.coalesce(Order.platform, ""))).in_(platform_values),
                or_(Order.account_id == account_id, Order.shop_id == account_id),
            )
        )
    return [or_(*conditions)] if conditions else []


def _risk_shop_keys_from_ids(db: Session, risk_filter: str | None, shop_ids: str | None) -> list[tuple[str, str]]:
    if not _normalize_order_risk_filter(risk_filter):
        return []
    return _shop_keys_from_ids(db, shop_ids)


def _shop_keys_from_ids(db: Session, shop_ids: str | None) -> list[tuple[str, str]]:
    selected_ids = _to_int_list(shop_ids)
    if not selected_ids:
        return []
    return list(_dashboard_shop_scope(db, selected_ids).keys)


def _risk_bucket_from_seconds(seconds: int | None) -> str:
    if seconds is None:
        return "no_deadline"
    hours = seconds / 3600
    if hours < -48:
        return "overdue_48"
    if hours < -24:
        return "overdue_24_48"
    if hours < 0:
        return "overdue_0_24"
    if hours < 24:
        return "due_24"
    if hours < 48:
        return "due_48"
    return "due_later"


def _format_remaining_seconds(seconds: int | None) -> str:
    if seconds is None:
        return ""
    if seconds <= 0:
        return "0小时"
    days, remain = divmod(seconds, 86400)
    hours, remain = divmod(remain, 3600)
    minutes, _ = divmod(remain, 60)
    if days > 0:
        return f"{days}天{hours}小时{minutes}分"
    if hours > 0:
        return f"{hours}小时{minutes}分"
    return f"{max(1, minutes)}分"


def _format_remaining_delta(deadline: datetime | None) -> str:
    if not deadline:
        return ""
    try:
        if deadline.tzinfo is not None:
            deadline = deadline.astimezone(timezone.utc).replace(tzinfo=None)
        deadline_local = deadline.replace(tzinfo=None) + LOCAL_TIME_OFFSET
        seconds = int((deadline_local - _local_now()).total_seconds())
    except TypeError:
        return ""
    return _format_remaining_seconds(seconds)


def _country_name_cn(code: str | None) -> str:
    return country_name_cn(code)


def _country_name_to_code(name: str | None) -> str:
    return country_name_to_code(name)


def _extract_order_items(raw_payload: dict) -> list[dict]:
    for value in (raw_payload.get("products"), raw_payload.get("items"), raw_payload.get("order_items")):
        items = [item for item in _as_list(value) if isinstance(item, dict)]
        if items:
            return items
    nested_items: list[dict] = []
    for order in _as_list(raw_payload.get("orders")):
        if not isinstance(order, dict):
            continue
        for value in (order.get("products"), order.get("items"), order.get("order_items")):
            nested_items.extend(item for item in _as_list(value) if isinstance(item, dict))
    if nested_items:
        return nested_items
    config = raw_payload.get("config") if isinstance(raw_payload.get("config"), dict) else {}
    return [item for item in _as_list(config.get("items")) if isinstance(item, dict)]


def _item_sku(item: dict, platform: str = "") -> str:
    source_item = _first_dict(item.get("item"), item.get("product"), item.get("offer"), item)
    variation = _first_dict(item.get("variation"), source_item.get("variation"))
    raw_item = item.get("raw_payload") if isinstance(item.get("raw_payload"), dict) else {}
    raw_offer = raw_item.get("offer") if isinstance(raw_item.get("offer"), dict) else {}
    offer = _first_dict(item.get("offer"), raw_offer, source_item.get("offer"), source_item)
    external = offer.get("external") if isinstance(offer.get("external"), dict) else {}
    if _to_str(platform).strip().lower() == "allegro":
        seller_sku = _to_str(
            _first_value(
                external.get("id"),
                item.get("seller_sku"),
                item.get("sellerSku"),
                item.get("seller_custom_field"),
                raw_item.get("seller_sku"),
                raw_item.get("sellerSku"),
                raw_item.get("sku"),
                offer.get("seller_sku"),
                offer.get("sellerSku"),
                offer.get("sku"),
            )
        ).strip()
        if seller_sku:
            return seller_sku
    return _to_str(
        _first_value(
            item.get("offer_id"),
            item.get("offerId"),
            item.get("seller_sku"),
            item.get("seller_custom_field"),
            item.get("sku"),
            source_item.get("seller_sku"),
            source_item.get("seller_custom_field"),
            source_item.get("sku"),
            variation.get("seller_sku"),
            variation.get("seller_custom_field"),
            source_item.get("id"),
            item.get("item_id"),
            item.get("id"),
        )
    ).strip()


def _item_platform_product_name(item: dict | None) -> str:
    item = item or {}
    source_item = _first_dict(item.get("item"), item.get("product"), item.get("offer"), item)
    variant = _first_dict(item.get("variant"), item.get("variation"), source_item.get("variant"), source_item.get("variation"))
    return _to_str(
        _first_value(
            item.get("platform_product_name"),
            item.get("product_name"),
            item.get("name"),
            item.get("title"),
            item.get("item_title"),
            item.get("productName"),
            item.get("product_name"),
            item.get("offer_name"),
            item.get("goods_name"),
            item.get("goodsName"),
            item.get("subject"),
            item.get("description"),
            source_item.get("name"),
            source_item.get("title"),
            source_item.get("item_title"),
            source_item.get("productName"),
            source_item.get("product_name"),
            source_item.get("offer_name"),
            source_item.get("subject"),
            variant.get("name"),
            variant.get("title"),
        )
    ).strip()


def _item_quantity(item: dict) -> int:
    value = _first_value(item.get("quantity"), item.get("count"), 1)
    try:
        return int(value or 1)
    except (TypeError, ValueError):
        return 1


def _item_unit_price(item: dict) -> str | None:
    source_item = _first_dict(item.get("item"), item.get("product"), item.get("offer"), item)
    value = _first_value(item.get("price"), item.get("unit_price"), item.get("full_unit_price"), item.get("price_unit"), item.get("priceWithoutCommission"), source_item.get("price"))
    if isinstance(value, dict):
        value = _first_value(value.get("amount"), value.get("value"), value.get("price"))
    return None if value in (None, "") else _to_str(value)


def _item_currency(item: dict, fallback: str = "") -> str:
    source_item = _first_dict(item.get("item"), item.get("product"), item.get("offer"), item)
    price = item.get("price")
    price_currency = price.get("currency") if isinstance(price, dict) else None
    return _to_str(_first_value(item.get("currency_code"), item.get("currency_id"), item.get("currency"), price_currency, source_item.get("currency_code"), source_item.get("currency_id"), fallback))


def _item_display_name(raw_payload: dict | None) -> str:
    return _item_platform_product_name(raw_payload)


def _first_order(raw_payload: dict) -> dict:
    return _first_dict(*_as_list(raw_payload.get("orders")))


def _first_payment(raw_payload: dict) -> dict:
    order = _first_order(raw_payload)
    return _first_dict(*_as_list(order.get("payments")), *_as_list(raw_payload.get("payments")))


def _shipping_destination_address(shipping: dict) -> dict:
    destination = shipping.get("destination") if isinstance(shipping.get("destination"), dict) else {}
    return _first_dict(
        destination.get("shipping_address"),
        destination,
        shipping.get("receiver_address"),
        shipping.get("shipping_address"),
    )


def _normalized_order_item_payloads(raw_payload: dict, fallback_currency: str = "", platform: str = "") -> list[dict]:
    items = _extract_order_items(raw_payload or {})
    if not items:
        return [{"sku": "", "platform_product_name": "", "quantity": 1, "unit_price": None, "currency": fallback_currency or "", "raw_payload": {}}]
    return [
        {
            "sku": _item_sku(item, platform),
            "platform_product_name": _item_platform_product_name(item),
            "quantity": _item_quantity(item),
            "unit_price": _item_unit_price(item),
            "currency": _item_currency(item, fallback_currency),
            "raw_payload": item,
        }
        for item in items
    ]


def _replace_order_items(db: Session, order: Order) -> None:
    db.query(OrderItem).filter(OrderItem.order_id == order.id).delete(synchronize_session=False)
    for item in _normalized_order_item_payloads(order.raw_payload or {}, order.currency or "", order.platform):
        db.add(OrderItem(order_id=order.id, **item))


def _backfill_order_items(db: Session) -> None:
    time_changed = False
    time_rows = db.scalars(select(Order).where(Order.raw_payload != {}).limit(10000)).all()
    for row in time_rows:
        extracted = _extract_order_fields(row.raw_payload or {})
        for attr in ("platform_created_at", "platform_handover_deadline", "payment_at", "shipping_deadline_at"):
            value = extracted.get(attr)
            if value and getattr(row, attr) != value:
                setattr(row, attr, value)
                time_changed = True
        fulfillment_type = infer_fulfillment_type(row.platform, row.raw_payload or {}, row.fulfillment_type)
        is_overseas_warehouse = infer_is_overseas_warehouse(row.platform, fulfillment_type, row.raw_payload or {})
        if row.fulfillment_type != fulfillment_type:
            row.fulfillment_type = fulfillment_type
            time_changed = True
        if bool(row.is_overseas_warehouse) != is_overseas_warehouse:
            row.is_overseas_warehouse = is_overseas_warehouse
            time_changed = True

    missing_created_rows = db.scalars(
        select(Order).where(Order.platform_created_at.is_(None)).limit(5000)
    ).all()
    for row in missing_created_rows:
        extracted = _extract_order_fields(row.raw_payload or {})
        row.platform_created_at = extracted["platform_created_at"] or row.platform_created_at

    missing_order_ids = db.scalars(
        select(Order.id).where(~exists().where(OrderItem.order_id == Order.id)).limit(5000)
    ).all()
    malformed_order_ids = db.scalars(
        select(OrderItem.order_id).where(OrderItem.unit_price.like("{%")).distinct().limit(5000)
    ).all()
    stale_sku_order_ids = []
    item_rows = db.execute(
        select(OrderItem.order_id, OrderItem.sku, OrderItem.raw_payload, Order.platform)
        .join(Order, Order.id == OrderItem.order_id)
        .limit(10000)
    ).all()
    for order_id, sku, raw_payload, platform in item_rows:
        normalized_sku = _item_sku(raw_payload or {}, platform)
        if normalized_sku and normalized_sku != (sku or ""):
            stale_sku_order_ids.append(order_id)
    stale_name_order_ids = []
    name_rows = db.execute(select(OrderItem.order_id, OrderItem.platform_product_name, OrderItem.raw_payload).limit(10000)).all()
    for order_id, platform_product_name, raw_payload in name_rows:
        normalized_name = _item_platform_product_name(raw_payload or {})
        if normalized_name and normalized_name != (platform_product_name or ""):
            stale_name_order_ids.append(order_id)
    empty_placeholder_order_ids = db.scalars(
        select(Order.id)
        .join(OrderItem, OrderItem.order_id == Order.id)
        .where(OrderItem.sku == "", OrderItem.raw_payload == {})
        .limit(5000)
    ).all()
    target_order_ids = list(dict.fromkeys([*missing_order_ids, *malformed_order_ids, *stale_sku_order_ids, *stale_name_order_ids]))
    for order_id in empty_placeholder_order_ids:
        if order_id not in target_order_ids:
            target_order_ids.append(order_id)
    if not target_order_ids:
        if missing_created_rows or time_changed:
            db.commit()
        return
    rows = db.scalars(select(Order).where(Order.id.in_(target_order_ids))).all()
    for row in rows:
        _replace_order_items(db, row)
    db.commit()


def _extract_order_fields(raw_payload: dict) -> dict:
    customer = raw_payload.get("customer") or raw_payload.get("buyer") or raw_payload.get("user") or {}
    delivery = raw_payload.get("delivery_method") or raw_payload.get("delivery") or {}
    analytics = raw_payload.get("analytics_data") or {}
    shipping = raw_payload.get("shipping") or {}
    shipment = raw_payload.get("shipment") if isinstance(raw_payload.get("shipment"), dict) else {}
    order = _first_order(raw_payload)
    payment = _first_payment(raw_payload)
    destination_address = _shipping_destination_address(shipping) or _shipping_destination_address(shipment)
    receiver = shipping.get("receiver_address") or destination_address or raw_payload.get("address") or {}
    customer_address = customer.get("address") or {}
    products = raw_payload.get("products") or []
    financial = raw_payload.get("financial_data") or {}
    financial_products = financial.get("products") or []
    order_items = _extract_order_items(raw_payload)

    buyer_name = _first_value(
        customer.get("name"),
        customer.get("full_name"),
        customer.get("nickname"),
        (order.get("buyer") or {}).get("nickname") if isinstance(order.get("buyer"), dict) else None,
        " ".join(
            part.strip()
            for part in [
                _to_str((order.get("buyer") or {}).get("first_name") if isinstance(order.get("buyer"), dict) else ""),
                _to_str((order.get("buyer") or {}).get("last_name") if isinstance(order.get("buyer"), dict) else ""),
            ]
            if part and part.strip()
        ),
        shipping.get("receiver_name"),
        receiver.get("name"),
        receiver.get("receiver_name"),
        (shipping.get("destination") or {}).get("receiver_name") if isinstance(shipping.get("destination"), dict) else None,
    )

    # --- country ---
    country_code = _first_value(
        receiver.get("country_code"),
        (receiver.get("country") or {}).get("id") if isinstance(receiver.get("country"), dict) else None,
        (destination_address.get("country") or {}).get("id") if isinstance(destination_address.get("country"), dict) else None,
        shipping.get("country_id"),
        shipment.get("country_id"),
        delivery.get("country_code"),
        raw_payload.get("country_code"),
    )
    country_code = _to_str(country_code).upper()
    if not country_code:
        country_name_raw = _first_value(
            customer_address.get("country"),
            receiver.get("country"),
            (receiver.get("country") or {}).get("name") if isinstance(receiver.get("country"), dict) else None,
            (destination_address.get("country") or {}).get("name") if isinstance(destination_address.get("country"), dict) else None,
            raw_payload.get("country"),
        )
        country_code = _country_name_to_code(country_name_raw)
    # analytics_data.region may be a region name, only use if 2-letter code
    if not country_code:
        region = _to_str(analytics.get("region")).upper()
        if len(region) == 2 and region.isalpha():
            country_code = region
    wb_country_code = wildberries_payload_country_code(raw_payload)
    if wb_country_code:
        country_code = wb_country_code

    # --- amount & currency ---
    amount = _first_value(
        raw_payload.get("order_amount"),
        raw_payload.get("sum"),
        raw_payload.get("total_amount"),
        raw_payload.get("amount"),
        order.get("paid_amount"),
        payment.get("total_paid_amount"),
        payment.get("transaction_amount"),
    )
    currency = _first_value(
        raw_payload.get("currency_code"),
        raw_payload.get("currency"),
        raw_payload.get("money_currency"),
        order.get("currency_id"),
        payment.get("currency_id"),
        shipping.get("currency_id"),
        shipment.get("currency_id"),
    )
    if not amount and products:
        try:
            total = sum(float(p.get("price", 0)) * int(p.get("quantity", 1)) for p in products)
            amount = f"{total:.2f}"
        except (ValueError, TypeError):
            pass
    if not currency and products:
        currency = _first_value(*(_item_currency(p) for p in products if isinstance(p, dict)))
    if not amount and financial_products:
        try:
            total = sum(float(fp.get("price", 0)) * int(fp.get("quantity", 1)) for fp in financial_products)
            amount = f"{total:.2f}"
        except (ValueError, TypeError):
            pass
    if not currency and financial_products:
        currency = _first_value(*(_item_currency(fp) for fp in financial_products if isinstance(fp, dict)))

    platform_handover_deadline = _parse_platform_datetime(
        _first_value(
            raw_payload.get("shipment_date"),
            raw_payload.get("platform_handover_deadline"),
            raw_payload.get("ship_by_date"),
            raw_payload.get("delivery_date_begin"),
            shipping.get("date_first_printed"),
        )
    )
    platform_created_at = _parse_platform_datetime(
        _first_value(
            raw_payload.get("created_at"),
            raw_payload.get("order_date"),
            raw_payload.get("date_created"),
            order.get("date_created"),
            raw_payload.get("in_process_at"),
        )
    )
    payment_at = _parse_platform_datetime(
        _first_value(
            payment.get("date_approved"),
            payment.get("date_created"),
            order.get("date_closed"),
            raw_payload.get("in_process_at"),
            raw_payload.get("payment_at"),
            raw_payload.get("created_at"),
            raw_payload.get("order_date"),
            raw_payload.get("date_created"),
            order.get("date_created"),
        )
    )
    shipping_deadline_at = _parse_platform_datetime(_shipping_deadline_value(raw_payload, order, shipping, shipment))
    if not shipping_deadline_at and payment_at:
        shipping_deadline_at = payment_at + timedelta(days=5)

    handover_at = _parse_datetime(
        _first_value(
            raw_payload.get("shipped_at"),
            raw_payload.get("handover_at"),
            raw_payload.get("delivering_date"),
            raw_payload.get("last_ship_date"),
        )
    )

    return {
        "site": _to_str(_first_value(raw_payload.get("site"), raw_payload.get("marketplace"), raw_payload.get("domain"))),
        "buyer_id": _to_str(_first_value(customer.get("id"), customer.get("customer_id"), raw_payload.get("customer_id"), raw_payload.get("buyer_id"))),
        "buyer_name": _to_str(buyer_name),
        "platform_handover_deadline": platform_handover_deadline,
        "platform_created_at": platform_created_at,
        "handover_at": handover_at,
        "country_code": country_code,
        "country_name_cn": _country_name_cn(country_code),
        "buyer_selected_logistics": _to_str(_first_value(
            delivery.get("name"),
            shipping.get("tracking_method"),
            shipment.get("tracking_method"),
            ((shipping.get("lead_time") or {}).get("shipping_method") or {}).get("name") if isinstance(shipping.get("lead_time"), dict) else None,
            shipping.get("shipping_mode"),
            shipping.get("logistic_type"),
            (shipping.get("logistic") or {}).get("type") if isinstance(shipping.get("logistic"), dict) else None,
            raw_payload.get("buyer_selected_logistics"),
        )),
        "order_amount": _to_str(amount),
        "currency": _to_str(currency),
        "shipment_tracking_number": _tracking_number_from_payload(raw_payload),
        "payment_at": payment_at,
        "shipping_deadline_at": shipping_deadline_at,
        "items": order_items,
    }


def _latest_shipment(db: Session, order_id: int) -> Shipment | None:
    return db.scalar(select(Shipment).where(Shipment.order_id == order_id).order_by(Shipment.id.desc()))


def _latest_label(db: Session, shipment_id: int | None) -> LabelFile | None:
    if not shipment_id:
        return None
    return db.scalar(select(LabelFile).where(LabelFile.shipment_id == shipment_id).order_by(LabelFile.id.desc()))


def _is_real_label_pdf(data: bytes) -> bool:
    return is_real_label_pdf(data)


def _label_file_is_real(label: LabelFile | None) -> bool:
    if not label or not label.file_path:
        return False
    p = Path(label.file_path)
    if not p.exists() or p.stat().st_size <= 0:
        return False
    try:
        return _is_real_label_pdf(p.read_bytes())
    except Exception:
        return False


def _pdf_first_page_size_mm_from_bytes(data: bytes) -> tuple[float, float] | None:
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data))
        page = reader.pages[0]
        width = float(page.mediabox.width) * 25.4 / 72.0
        height = float(page.mediabox.height) * 25.4 / 72.0
        rotation = int(page.get("/Rotate", 0) or 0) % 180
        if rotation == 90:
            return height, width
        return width, height
    except Exception:
        return None


def _is_wildberries_cross_border_order(row: Order) -> bool:
    if str(row.platform or "").strip().lower() != "wildberries":
        return False
    payload = row.raw_payload if isinstance(row.raw_payload, dict) else {}
    country_code = str(payload.get("country_code") or payload.get("countryCode") or row.country_code or "").strip().upper()
    if country_code and country_code != "RU":
        return True
    supply = payload.get("supply") if isinstance(payload.get("supply"), dict) else {}
    cross_border_type = str(_first_non_empty(payload.get("crossBorderType"), supply.get("crossBorderType")) or "").strip()
    return cross_border_type == "1"


def _is_legacy_wildberries_small_sticker(row: Order, data: bytes) -> bool:
    if not _is_wildberries_cross_border_order(row):
        return False
    page_size = _pdf_first_page_size_mm_from_bytes(data)
    if not page_size:
        return False
    width, height = page_size
    return abs(width - 58.0) <= 1.0 and abs(height - 40.0) <= 1.0


def _candidate_label_paths(row: Order, filename_stem: str) -> list[Path]:
    if not filename_stem:
        return []
    account_id = row.account_id or row.shop_id or ""
    roots = [get_settings().label_storage_path.resolve()]

    month_values: list[str] = []
    for value in (row.created_at, row.platform_created_at, row.payment_at, row.updated_at):
        if value:
            month = value.strftime("%Y%m")
            if month not in month_values:
                month_values.append(month)
    current_month = datetime.utcnow().strftime("%Y%m")
    if current_month not in month_values:
        month_values.append(current_month)

    tenant_values = [row.tenant_id or ""]
    if "tenant-demo" not in tenant_values:
        tenant_values.append("tenant-demo")

    candidates: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        for tenant in tenant_values:
            if not tenant:
                continue
            for month in month_values:
                path = root / tenant / row.platform / account_id / month / f"{filename_stem}.pdf"
                resolved = path.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    candidates.append(path)
    return candidates


def _find_existing_label_pdf(row: Order) -> tuple[Path, bytes] | None:
    stems = [
        row.posting_number or "",
        row.shipment_tracking_number or "",
        row.platform_order_no or "",
        row.platform_order_id or "",
    ]
    seen_stems: set[str] = set()
    for stem in stems:
        stem = str(stem or "").strip()
        if not stem or stem in seen_stems:
            continue
        seen_stems.add(stem)
        for path in _candidate_label_paths(row, stem):
            try:
                if path.exists() and path.stat().st_size > 0:
                    data = path.read_bytes()
                    if _is_real_label_pdf(data) and not _is_legacy_wildberries_small_sticker(row, data):
                        return path, data
            except Exception:
                continue
    return None


def _ozon_label_error_message(row: Order, exc: Exception) -> str:
    message = str(exc)
    if row.platform != "ozon":
        return message
    lowered = message.lower()
    if "invalid_argument" in lowered and "package-label" in lowered:
        return (
            f"Ozon 返回 INVALID_ARGUMENT，posting_number={row.posting_number or row.platform_order_id}。"
            "该订单当前可能不支持重新下载面单；已发货订单通常需要使用发货前缓存的本地 PDF。"
        )
    return message


def _ozon_label_not_ready(row: Order) -> bool:
    if row.platform != "ozon":
        return False
    tracking_number = clean_tracking_number(row.shipment_tracking_number, row.raw_payload or {}, row.platform)
    if tracking_number:
        return False
    payload = row.raw_payload if isinstance(row.raw_payload, dict) else {}
    status = str(row.platform_status or payload.get("status") or "").strip().lower()
    substatus = str(payload.get("substatus") or "").strip().lower()
    return status in {"awaiting_packaging", "awaiting_registration"} or substatus in {"posting_created", "posting_awaiting_registration"}


def _allegro_label_fetch_unavailable_message(value) -> bool:
    text = str(value or "")
    markers = (
        "Allegro 订单 shipment 面单接口不可用",
        "Feature unavailable",
        "只有运单号时无法从 Allegro 下载面单 PDF",
        "没有可用于下载面单的 shipmentId",
        "没有可用于下载面单的 Allegro shipmentId",
        "只有通过 Ship with Allegro/WZA 创建的 shipment 才能下载面单",
        "Allegro WZA 面单接口返回 406 Not Acceptable",
        "当前 shipment 没有可下载平台面单",
    )
    return any(marker in text for marker in markers)


async def _ensure_labels_cached(
    db: Session,
    rows: list[Order],
    load_bytes: bool = True,
) -> tuple[dict[int, bytes], int, int, int]:
    """确保给定订单在本地有有效面单 PDF。

    本地已有面单时直接命中；缺失时按店铺分组从平台拉取并落盘。
    返回: (order_id -> pdf_bytes 或 b"", cached 数, fetched 数, failed 数)
    """
    result: dict[int, bytes] = {}
    cached = 0
    fetched = 0
    failed = 0
    missing_rows: list[Order] = []

    for row in rows:
        if order_is_overseas_warehouse(row) or order_is_logistics_label_exempt(row):
            result[row.id] = b""
            continue
        if order_is_joom_offline_shipping(row):
            result[row.id] = b""
            continue
        shipment = _latest_shipment(db, row.id)
        label = _latest_label(db, shipment.id if shipment else None)
        local_ok = False
        if label and label.file_path:
            p = Path(label.file_path)
            if p.exists() and p.stat().st_size > 0:
                try:
                    data = p.read_bytes()
                    if _is_real_label_pdf(data) and not _is_legacy_wildberries_small_sticker(row, data):
                        local_ok = True
                        cached += 1
                        row.error_message = ""
                        if shipment and not shipment.tracking_number and row.shipment_tracking_number:
                            shipment.tracking_number = row.shipment_tracking_number
                        if load_bytes:
                            result[row.id] = data
                    else:
                        local_ok = False
                except Exception:
                    local_ok = False
                if local_ok and not load_bytes:
                    result[row.id] = b""
        if not local_ok:
            existing_pdf = _find_existing_label_pdf(row)
            if existing_pdf:
                path, data = existing_pdf
                shipment = _latest_shipment(db, row.id)
                posting_number = row.posting_number or row.platform_order_id
                if shipment is None:
                    shipment = Shipment(
                        order_id=row.id,
                        platform_shipment_id=posting_number,
                        tracking_number=row.shipment_tracking_number or "",
                        carrier="Ozon" if row.platform == "ozon" else row.platform,
                        status="label_ready",
                    )
                    db.add(shipment)
                    db.flush()
                digest = __import__("hashlib").sha256(data).hexdigest()
                db.add(
                    LabelFile(
                        shipment_id=shipment.id,
                        file_path=str(path),
                        content_type="application/pdf",
                        sha256=digest,
                    )
                )
                if row.local_status in {"new", "failed_retryable", ""}:
                    row.local_status = "label_saved"
                row.error_message = ""
                if shipment and not shipment.tracking_number and row.shipment_tracking_number:
                    shipment.tracking_number = row.shipment_tracking_number
                cached += 1
                result[row.id] = data if load_bytes else b""
            else:
                missing_rows.append(row)

    if not missing_rows:
        return result, cached, fetched, failed

    # 按 (platform, account_id) 分组，为缺失面单的订单分别创建连接器。
    groups: dict[tuple[str, str], list[Order]] = {}
    for row in missing_rows:
        groups.setdefault((row.platform, row.account_id), []).append(row)

    for (platform, account_id), group_rows in groups.items():
        wanbang_rows = [row for row in group_rows if order_uses_wanbang(row)]
        for r in wanbang_rows:
            try:
                label_result, shipment_result = await fetch_wanbang_label_for_order(db, r)
                content = label_result.content
                if not _is_real_label_pdf(content):
                    raise RuntimeError("万邦返回非有效真实面单 PDF")

                posting_number = shipment_result.platform_shipment_id or r.posting_number or r.platform_order_id
                file_path, sha256 = save_label_pdf(
                    r.tenant_id, platform, account_id, posting_number, content
                )

                shipment = _latest_shipment(db, r.id)
                if shipment is None:
                    shipment = Shipment(
                        order_id=r.id,
                        platform_shipment_id=shipment_result.platform_shipment_id or posting_number,
                        tracking_number=shipment_result.tracking_number or r.shipment_tracking_number or "",
                        carrier=shipment_result.carrier or WANBANG_CARRIER_NAME,
                        status="label_ready",
                    )
                    db.add(shipment)
                    db.flush()
                else:
                    if shipment_result.platform_shipment_id:
                        shipment.platform_shipment_id = shipment_result.platform_shipment_id
                    if shipment_result.tracking_number:
                        shipment.tracking_number = shipment_result.tracking_number
                    if shipment_result.carrier:
                        shipment.carrier = shipment_result.carrier
                    shipment.status = "label_ready"
                apply_label_result_tracking(r, shipment, label_result)
                if shipment_result.tracking_number:
                    r.shipment_tracking_number = shipment_result.tracking_number
                    shipment.tracking_number = shipment_result.tracking_number
                    await backfill_wanbang_tracking_to_platform(
                        db,
                        r,
                        tracking_number=shipment_result.tracking_number,
                        source="label_fetch",
                    )
                db.add(
                    LabelFile(
                        shipment_id=shipment.id,
                        file_path=file_path,
                        content_type=label_result.content_type or "application/pdf",
                        sha256=sha256,
                    )
                )
                if r.local_status in {"new", "failed_retryable", ""}:
                    r.local_status = "label_saved"
                r.error_message = ""
                fetched += 1
                result[r.id] = content if load_bytes else b""
            except Exception as exc:
                failed += 1
                r.error_message = f"万邦面单同步失败：{str(exc)[:500]}"
                result[r.id] = b""
                continue

        group_rows = [row for row in group_rows if not order_uses_wanbang(row)]
        if not group_rows:
            continue

        try:
            local_setting = db.scalar(
                select(SyncSetting).where(
                    SyncSetting.platform == platform,
                    SyncSetting.account_id == account_id,
                )
            )
            connector = _connector_for_account(db, platform, account_id, local_setting)
            # 强制调用平台真实 API，不走 dry-run。
            if hasattr(connector, "settings") and isinstance(connector.settings, dict):
                connector.settings["dry_run_fulfillment"] = False
        except Exception as exc:
            for r in group_rows:
                failed += 1
                r.error_message = f"面单同步失败：连接器初始化失败：{str(exc)[:500]}"
                result[r.id] = b""
            continue

        for r in group_rows:
            try:
                if _ozon_label_not_ready(r):
                    r.error_message = "Ozon 面单尚未生成，等待平台状态进入可打印阶段"
                    result[r.id] = b""
                    continue
                posting_number = r.posting_number or r.platform_order_id
                shipment = _latest_shipment(db, r.id)
                platform_shipment_id, unavailable_reason = label_shipment_id_for_order(r, shipment)
                if unavailable_reason:
                    r.error_message = unavailable_reason
                    result[r.id] = b""
                    continue
                shipment_result = ShipmentResult(
                    platform_shipment_id=platform_shipment_id,
                    tracking_number=r.shipment_tracking_number or posting_number,
                    carrier="Ozon" if platform == "ozon" else platform,
                    status="label_ready",
                    raw_payload=r.raw_payload or {},
                )
                normalized = NormalizedOrder(
                    platform_order_id=r.platform_order_id,
                    platform_order_no=r.platform_order_no or "",
                    posting_number=r.posting_number or "",
                    platform_status=r.platform_status or "",
                    raw_payload=r.raw_payload or {},
                    fulfillment_type=r.fulfillment_type or infer_fulfillment_type(platform, r.raw_payload or {}),
                    is_overseas_warehouse=order_is_overseas_warehouse(r),
                )
                try:
                    label_result = await connector.fetch_label(shipment_result, normalized)
                except Exception as exc:
                    if "401" not in str(exc) or platform not in {"mercadolibre", "joom_logistics", "allegro"}:
                        raise
                    account = db.scalar(
                        select(PlatformAccount).where(
                            PlatformAccount.platform == platform,
                            PlatformAccount.account_id == account_id,
                        )
                    )
                    if not account or not account.encrypted_credentials:
                        raise
                    credentials = get_credential_manager().decrypt_credentials(account.encrypted_credentials)
                    settings = dict(account.settings or {})
                    settings["account_id"] = account_id
                    refreshed = ensure_access_token(db, account, credentials, settings, force=True)
                    if not refreshed.get("access_token"):
                        raise
                    connector = _connector_for_account(db, platform, account_id, local_setting)
                    if hasattr(connector, "settings") and isinstance(connector.settings, dict):
                        connector.settings["dry_run_fulfillment"] = False
                    label_result = await connector.fetch_label(shipment_result, normalized)
                content = label_result.content
                if not _is_real_label_pdf(content):
                    raise RuntimeError("平台返回非有效真实面单 PDF")

                file_path, sha256 = save_label_pdf(
                    r.tenant_id, platform, account_id, posting_number, content
                )

                if shipment is None:
                    shipment = Shipment(
                        order_id=r.id,
                        platform_shipment_id=platform_shipment_id,
                        tracking_number=r.shipment_tracking_number or "",
                        carrier="Ozon" if platform == "ozon" else platform,
                        status="label_ready",
                    )
                    db.add(shipment)
                    db.flush()
                elif platform_shipment_id and getattr(shipment, "platform_shipment_id", "") != platform_shipment_id:
                    shipment.platform_shipment_id = platform_shipment_id
                apply_label_result_tracking(r, shipment, label_result)
                if shipment and not shipment.tracking_number and r.shipment_tracking_number:
                    shipment.tracking_number = r.shipment_tracking_number
                db.add(
                    LabelFile(
                        shipment_id=shipment.id,
                        file_path=file_path,
                        content_type=label_result.content_type or "application/pdf",
                        sha256=sha256,
                    )
                )
                if r.local_status in {"new", "failed_retryable", ""}:
                    r.local_status = "label_saved"
                r.error_message = ""
                fetched += 1
                result[r.id] = content if load_bytes else b""
            except Exception as exc:
                if platform == "allegro" and _allegro_label_fetch_unavailable_message(exc):
                    r.error_message = str(exc)[:500]
                    result[r.id] = b""
                    continue
                failed += 1
                r.error_message = f"面单同步失败：{_ozon_label_error_message(r, exc)[:500]}"
                result[r.id] = b""
                continue

    db.commit()
    return result, cached, fetched, failed


def _display_order_amount(row: Order, extracted: dict) -> str:
    amount = row.order_amount or extracted["order_amount"]
    if row.platform != "wildberries" or not amount:
        return amount
    try:
        numeric = Decimal(str(amount))
    except (InvalidOperation, ValueError):
        return amount
    if numeric == numeric.to_integral_value() and abs(numeric) >= Decimal("1000"):
        return format((numeric / Decimal("100")).quantize(Decimal("0.01")).normalize(), "f")
    return amount


def _display_country_code(row: Order, extracted: dict) -> str:
    country_code = _to_str(extracted["country_code"] or row.country_code).strip().upper()
    if country_code:
        return country_code
    country_name = _to_str(row.country_name_cn).strip()
    return _country_name_to_code(country_name)


def _display_country_name_cn(row: Order, extracted: dict) -> str:
    country_code = _display_country_code(row, extracted)
    if country_code:
        return _country_name_cn(country_code)
    country_name = _to_str(row.country_name_cn).strip()
    code_from_name = _country_name_to_code(country_name)
    if code_from_name:
        return _country_name_cn(code_from_name)
    if country_name:
        return country_name
    return extracted["country_name_cn"] or _country_name_cn(country_code)


def _clean_joom_label_customer_name(value: str | None) -> str:
    text = re.sub(r"\s+", " ", _to_str(value)).strip(" :")
    if not text:
        return ""
    lowered = text.lower()
    if lowered.startswith(("tel", "phone")):
        return ""
    if text.upper() in {"TO", "SHIP TO"}:
        return ""
    if re.fullmatch(r"[A-Z]{2}", text):
        return ""
    if re.fullmatch(r"\d[\d\s/.-]*", text):
        return ""
    return text


def _joom_label_customer_name_from_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    for index, line in enumerate(lines):
        direct = re.match(r"^(?:TO|SHIP TO)\s*:\s*(.+)$", line, flags=re.IGNORECASE)
        if direct:
            name = re.split(r"\s+Tel\s*:", direct.group(1), maxsplit=1, flags=re.IGNORECASE)[0]
            cleaned = _clean_joom_label_customer_name(name)
            if cleaned:
                return cleaned
        if re.match(r"^(?:SHIP TO|TO)\s*:\s*$", line, flags=re.IGNORECASE):
            for candidate in lines[index + 1:index + 8]:
                cleaned = _clean_joom_label_customer_name(candidate)
                if cleaned:
                    return cleaned
    return ""


@lru_cache(maxsize=2048)
def _joom_label_customer_name_from_pdf(path_text: str) -> str:
    path = Path(path_text)
    if not path.exists():
        return ""
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        logger.debug("Failed to extract Joom label customer name from %s", path, exc_info=True)
        return ""
    return _joom_label_customer_name_from_text(text)


def _label_filename_stem(file_path: str | None) -> str:
    if not file_path:
        return ""
    filename = str(file_path).replace("\\", "/").rsplit("/", 1)[-1]
    return Path(filename).stem


def _joom_label_customer_name(row: Order, label_path: str | None = None) -> str:
    if row.platform != "joom_logistics":
        return ""
    candidates: list[Path] = []
    if label_path:
        path = Path(label_path)
        candidates.append(path)
        stem = _label_filename_stem(label_path)
        if stem:
            candidates.extend(_candidate_label_paths(row, stem))
    else:
        stem = row.posting_number or row.platform_order_no or row.platform_order_id
        if stem:
            candidates.extend(_candidate_label_paths(row, stem))
    for candidate in candidates:
        if candidate.exists():
            name = _joom_label_customer_name_from_pdf(str(candidate.resolve()))
            if name:
                return name
    return ""


def _display_customer_name(row: Order, extracted: dict, label_path: str | None = None) -> str:
    buyer_ids = {
        _to_str(row.buyer_id).strip(),
        _to_str(extracted.get("buyer_id")).strip(),
    }
    for value in (row.buyer_name, extracted.get("buyer_name")):
        text = _to_str(value).strip()
        if text and text not in buyer_ids:
            return text
    label_name = _joom_label_customer_name(row, label_path)
    if label_name and label_name not in buyer_ids:
        return label_name
    return ""


def _first_successful_outbound_scan_at(db: Session | None, order_id: int | None) -> datetime | None:
    if not db or not order_id:
        return None
    return db.scalar(
        select(func.min(OutboundScanRecord.scanned_at)).where(
            OutboundScanRecord.order_id == order_id,
            OutboundScanRecord.result == "success",
        )
    )


def _order_dto(
    row: Order,
    shipment: Shipment | None = None,
    label: LabelFile | None = None,
    db: Session = None,
    outbound_scanned_at: datetime | None = None,
    risk_view: bool = False,
    risk_handling: OrderRiskHandling | None = None,
) -> OrderDto:
    # 如果没有预加载，则手动查询，兼容旧调用。
    if shipment is None and db:
        shipment = _latest_shipment(db, row.id)
    if label is None and shipment and db:
        label = _latest_label(db, shipment.id)
    if outbound_scanned_at is None and db:
        outbound_scanned_at = _first_successful_outbound_scan_at(db, row.id)
    extracted = _extract_order_fields(row.raw_payload or {})
    status = _derive_order_status(row)
    shipping_deadline_at = _effective_shipping_deadline(row, extracted)
    stored_dispatch_deadline_at = row.dispatch_deadline_at
    dispatch_deadline_at = stored_dispatch_deadline_at
    if dispatch_deadline_at is None and db:
        dispatch_deadline_at = calculate_dispatch_deadline(row, load_shipping_deadline_settings(db))
    # 风险队列与仪表盘都以数据库已落库的发出截止时间优先，避免列表临时计算导致口径漂移。
    risk_deadline_at = (stored_dispatch_deadline_at or shipping_deadline_at) if risk_view else None
    remaining_deadline_at = risk_deadline_at if risk_view else shipping_deadline_at
    remaining_seconds = None
    if (_should_show_remaining_shipping(status) or risk_view) and remaining_deadline_at:
        try:
            deadline_utc = (
                remaining_deadline_at.astimezone(timezone.utc).replace(tzinfo=None)
                if remaining_deadline_at.tzinfo
                else remaining_deadline_at
            )
            remaining_seconds = int((deadline_utc - datetime.utcnow()).total_seconds())
        except TypeError:
            remaining_seconds = None
    tracking_number = (
        clean_tracking_number(row.shipment_tracking_number, row.raw_payload or {}, row.platform)
        or extracted["shipment_tracking_number"]
        or (clean_tracking_number(shipment.tracking_number, row.raw_payload or {}, row.platform) if shipment else "")
        or ""
    )
    has_label = _label_file_is_real(label)
    fulfillment_type = row.fulfillment_type or infer_fulfillment_type(row.platform, row.raw_payload or {})
    is_overseas_warehouse = order_is_overseas_warehouse(row)
    is_joom_offline_shipping = order_is_joom_offline_shipping(row)
    is_logistics_label_exempt = order_is_logistics_label_exempt(row)
    if is_logistics_label_exempt:
        tracking_number = ""
    return OrderDto(
        id=row.id,
        platform=row.platform,
        account_id=row.account_id,
        shop_id=row.shop_id or row.account_id,
        shop_name=row.shop_name or row.account_id,
        site=row.site or extracted["site"],
        platform_order_id=row.platform_order_id,
        platform_order_no=row.platform_order_no or "",
        posting_number=row.posting_number or "",
        transaction_id=row.platform_order_no or row.platform_order_id,
        customer_id=row.buyer_id or extracted["buyer_id"],
        customer_name=_display_customer_name(row, extracted, label.file_path if label else None),
        status=status,
        local_status=row.local_status,
        platform_status="" if row.platform_status in (None, "None") else row.platform_status,
        fulfillment_type=fulfillment_type,
        is_overseas_warehouse=is_overseas_warehouse,
        is_joom_offline_shipping=is_joom_offline_shipping,
        logistics_label_exempt=is_logistics_label_exempt,
        bsi_order_no=row.bsi_order_no or "",
        bsi_submitted_at=_iso(row.bsi_submitted_at),
        platform_handover_deadline=_platform_import_time(
            row.raw_payload or {},
            "platform_handover_deadline",
            row.platform_handover_deadline or extracted["platform_handover_deadline"],
        ),
        country_name_cn=_display_country_name_cn(row, extracted),
        country_code=_display_country_code(row, extracted),
        buyer_selected_logistics=row.buyer_selected_logistics or extracted["buyer_selected_logistics"],
        order_amount=_display_order_amount(row, extracted),
        currency=row.currency or extracted["currency"],
        payment_at=_platform_import_time(row.raw_payload or {}, "payment_at", row.payment_at or extracted["payment_at"]),
        shipment_tracking_number=tracking_number,
        tracking_number=tracking_number,
        logistics_channel=row.logistics_channel or "",
        logistics_match_rule_id=row.logistics_match_rule_id,
        logistics_match_rule_name=row.logistics_match_rule_name or "",
        logistics_match_status=row.logistics_match_status or "unmatched",
        logistics_match_reason=row.logistics_match_reason or "",
        logistics_matched_at=_iso(row.logistics_matched_at),
        picking_at=_iso(_coalesce_model_time(row, db, "picking_at")),
        marked_shipped_at=_iso(_coalesce_model_time(row, db, "marked_shipped_at")),
        label_printed_at=_iso(_coalesce_model_time(row, db, "label_printed_at")),
        handover_at=_iso(
            row.handover_at
            or (shipment.created_at if shipment else None)
            or extracted["handover_at"]
            or outbound_scanned_at
        ),
        shipped_at=_iso(row.shipped_at),
        shipping_deadline_at=_platform_import_time(row.raw_payload or {}, "shipping_deadline_at", shipping_deadline_at),
        dispatch_deadline_at=_platform_time_iso(dispatch_deadline_at),
        remaining_shipping_seconds=remaining_seconds,
        remaining_shipping_time=_format_remaining_seconds(remaining_seconds),
        risk_deadline_at=_platform_time_iso(risk_deadline_at),
        risk_bucket=_risk_bucket_from_seconds(remaining_seconds) if risk_view else "",
        risk_handled=bool(risk_handling),
        risk_handled_at=_iso(risk_handling.handled_at) if risk_handling else None,
        risk_handled_by=risk_handling.handled_by if risk_handling else "",
        risk_handling_note=risk_handling.note if risk_handling else "",
        has_label=has_label,
        label_path=label.file_path if has_label and label else "",
        created_at=_iso(row.created_at) or "",
        updated_at=_iso(row.updated_at) or "",
    )


def _sku_lookup_key(value: str | None) -> str:
    return (value or "").strip().lower()


def _mapping_rank(updated_at: datetime | None, created_at: datetime | None, mapping_id: int | None) -> tuple[datetime, int]:
    return (updated_at or created_at or datetime.min, int(mapping_id or 0))


def _order_item_product_matches(
    db: Session,
    shop_db_id: int | None,
    items: list[OrderItem],
) -> tuple[dict[str, dict], dict[str, dict]]:
    if not shop_db_id:
        return {}, {}

    exact_skus = {item.sku for item in items if item.sku}
    normalized_skus = {_sku_lookup_key(item.sku) for item in items if _sku_lookup_key(item.sku)}
    if not exact_skus and not normalized_skus:
        return {}, {}

    normalized_mapping_sku = func.lower(func.trim(func.coalesce(ProductShopMapping.shop_sku, "")))
    conditions = []
    if exact_skus:
        conditions.append(ProductShopMapping.shop_sku.in_(exact_skus))
    if normalized_skus:
        conditions.append(normalized_mapping_sku.in_(normalized_skus))

    rows = db.execute(
        select(
            ProductShopMapping.id,
            ProductShopMapping.shop_sku,
            ProductShopMapping.created_at,
            ProductShopMapping.updated_at,
            Product.product_code,
            Product.internal_name,
            Product.cost,
            Product.weight,
        )
        .join(Product, Product.id == ProductShopMapping.product_id)
        .where(
            ProductShopMapping.shop_id == shop_db_id,
            or_(*conditions),
        )
    ).all()

    exact_matches: dict[str, tuple[tuple[datetime, int], dict]] = {}
    normalized_matches: dict[str, tuple[tuple[datetime, int], dict]] = {}
    for mapping_id, shop_sku, created_at, updated_at, product_code, product_name, product_cost, product_weight in rows:
        match = {
            "product_code": product_code or "",
            "product_name": product_name or "",
            "product_cost": product_cost,
            "product_weight": product_weight,
        }
        rank = _mapping_rank(updated_at, created_at, mapping_id)
        exact_key = shop_sku or ""
        if exact_key in exact_skus and (exact_key not in exact_matches or rank > exact_matches[exact_key][0]):
            exact_matches[exact_key] = (rank, match)
        normalized_key = _sku_lookup_key(shop_sku)
        if normalized_key in normalized_skus and (
            normalized_key not in normalized_matches or rank > normalized_matches[normalized_key][0]
        ):
            normalized_matches[normalized_key] = (rank, match)

    return (
        {sku: match for sku, (_rank, match) in exact_matches.items()},
        {sku: match for sku, (_rank, match) in normalized_matches.items()},
    )


def _order_item_detail_rows(db: Session, row: Order) -> list[OrderDetailItemDto]:
    item_rows = db.scalars(
        select(OrderItem)
        .where(OrderItem.order_id == row.id)
        .order_by(OrderItem.id)
    ).all()
    if not item_rows:
        return []

    platform_account = db.scalar(
        select(PlatformAccount).where(
            PlatformAccount.platform == row.platform,
            PlatformAccount.account_id == (row.shop_id or row.account_id),
        )
    )
    shop_db_id = platform_account.id if platform_account else None
    exact_matches, normalized_matches = _order_item_product_matches(db, shop_db_id, item_rows)

    items: list[OrderDetailItemDto] = []
    for item in item_rows:
        match = exact_matches.get(item.sku or "") or normalized_matches.get(_sku_lookup_key(item.sku)) or {}
        product_cost = match.get("product_cost")
        product_weight = match.get("product_weight")
        items.append(
            OrderDetailItemDto(
                id=item.id,
                sku=item.sku or "",
                platform_product_name=item.platform_product_name or _item_platform_product_name(item.raw_payload),
                quantity=item.quantity or 1,
                unit_price=item.unit_price or "",
                currency=item.currency or row.currency or "",
                product_code=match.get("product_code") or "",
                product_name=match.get("product_name") or "",
                product_cost=float(product_cost) if product_cost is not None else None,
                product_weight=float(product_weight) if product_weight is not None else None,
            )
        )
    return items


def _order_chinese_product_name_map(db: Session, rows: list[Order]) -> dict[int, str]:
    if not rows:
        return {}

    order_ids = [row.id for row in rows]
    account_pairs = {(row.platform, row.shop_id or row.account_id) for row in rows}
    account_id_map: dict[tuple[str, str], int] = {}
    if account_pairs:
        platform_values = {platform for platform, _account_id in account_pairs}
        account_values = {account_id for _platform, account_id in account_pairs}
        accounts = db.scalars(
            select(PlatformAccount).where(
                PlatformAccount.platform.in_(platform_values),
                PlatformAccount.account_id.in_(account_values),
            )
        ).all()
        account_id_map = {(account.platform, account.account_id): account.id for account in accounts}

    shop_db_whens = [
        (OrderItem.order_id == row.id, account_id_map.get((row.platform, row.shop_id or row.account_id)))
        for row in rows
        if account_id_map.get((row.platform, row.shop_id or row.account_id)) is not None
    ]
    if not shop_db_whens:
        return {}

    shop_db_case = case(*shop_db_whens, else_=None)
    mapping_choice = mapping_choice_for_order_item(shop_db_case)
    product_name = func.coalesce(mapping_choice["exact_product"].internal_name, mapping_choice["insensitive_product"].internal_name)
    stmt = (
        select(
            OrderItem.order_id,
            product_name.label("product_name"),
        )
        .select_from(OrderItem)
        .outerjoin(mapping_choice["exact_mapping"], mapping_choice["exact_condition"])
        .outerjoin(mapping_choice["exact_product"], mapping_choice["exact_product"].id == mapping_choice["exact_mapping"].product_id)
        .outerjoin(
            mapping_choice["insensitive_mapping"],
            mapping_choice["insensitive_condition"],
        )
        .outerjoin(
            mapping_choice["insensitive_product"],
            mapping_choice["insensitive_product"].id == mapping_choice["insensitive_mapping"].product_id,
        )
        .where(OrderItem.order_id.in_(order_ids))
        .order_by(OrderItem.order_id, OrderItem.id)
    )
    names: dict[int, list[str]] = {order_id: [] for order_id in order_ids}
    for order_id, product_name in db.execute(stmt).all():
        name = (product_name or "").strip()
        if name:
            names.setdefault(order_id, []).append(name)
    return {
        order_id: "；".join(dict.fromkeys(row_names))
        for order_id, row_names in names.items()
        if row_names
    }


def _order_operation_log_dto(row: OrderOperationLog) -> OrderOperationLogDto:
    extra = row.extra if isinstance(row.extra, dict) else {}
    raw_changes = extra.get("changes") if isinstance(extra.get("changes"), list) else []
    changes = []
    for change in raw_changes[:20]:
        if not isinstance(change, dict):
            continue
        changes.append(
            OrderOperationLogChangeDto(
                field=str(change.get("field") or ""),
                label=str(change.get("label") or change.get("field") or ""),
                before=str(change.get("before") or "-"),
                after=str(change.get("after") or "-"),
            )
        )
    description = row.description or ""
    raw_result = str(extra.get("result") or "").strip().lower()
    if raw_result in {"failed", "failure", "error"} or "失败" in description:
        result = "failed"
    elif raw_result in {"warning", "skipped"} or any(keyword in description for keyword in ("超时", "跳过", "未匹配")):
        result = "warning"
    elif raw_result in {"unchanged", "no_change"} or "无变化" in description or "更新 0 条" in description:
        result = "unchanged"
    else:
        result = "success"

    operator = (row.operator or "").strip()
    if row.source == ORDER_LOG_SYSTEM_SOURCE and operator in {"", "超级管理员"}:
        operator = SYSTEM_OPERATOR
    elif row.source == ORDER_LOG_HISTORY_SOURCE and operator in {"", "超级管理员", SYSTEM_OPERATOR}:
        operator = "历史补充"

    def optional_int(*keys: str) -> int | None:
        for key in keys:
            value = extra.get(key)
            if isinstance(value, int):
                return value
            if isinstance(value, str) and value.isdigit():
                return int(value)
        return None

    return OrderOperationLogDto(
        id=row.id,
        operation_type=row.operation_type or "",
        operation_attribute=row.operation_attribute or "",
        description=description,
        operator=operator,
        source=row.source or "",
        result=result,
        changes=changes,
        task_run_id=optional_int("task_run_id", "run_id"),
        sync_job_log_id=optional_int("sync_job_log_id", "job_log_id"),
        operated_at=_iso(row.operated_at) or "",
        created_at=_iso(row.created_at) or "",
    )


def _display_log_value(value) -> str:
    text_value = str(value or "").strip()
    return text_value or "-"


def _legacy_order_log_description(db: Session, row: OrderOperationLog, order: Order) -> str:
    description = (row.description or "").strip()
    if row.operation_type == "order_sync" and description in {"订单同步新增", "订单同步更新"}:
        action = "首次同步" if description == "订单同步新增" else "同步更新"
        return f"历史记录：订单 {_order_display_number(order)} {action}，当时状态和变更内容未保存"
    return description


def _order_operation_log_rows(db: Session, order: Order) -> list[OrderOperationLogDto]:
    rows = db.scalars(
        select(OrderOperationLog)
        .where(OrderOperationLog.order_id == order.id)
        .order_by(desc(OrderOperationLog.operated_at), desc(OrderOperationLog.id))
    ).all()
    dtos = []
    for row in rows:
        dto = _order_operation_log_dto(row)
        dto.description = _legacy_order_log_description(db, row, order)
        dtos.append(dto)
    return dtos


def _order_detail_dto(row: Order, db: Session) -> OrderDetailDto:
    base = _order_dto(row, db=db).model_dump()
    return OrderDetailDto(
        **base,
        internal_order_no=row.internal_order_no or "",
        items=_order_item_detail_rows(db, row),
        operation_logs=[],
    )


def _clean_tracking_number(value: str | None) -> str:
    if not value:
        return ""
    return "".join(str(value).strip().split())


def _tracking_number_lookup_key(value: str | None) -> str:
    return _clean_tracking_number(value).lower()


def _tracking_number_matches(column, lookup_key: str):
    return func.lower(func.btrim(column)) == lookup_key


LOCAL_TIME_OFFSET = timedelta(hours=8)


def _local_now() -> datetime:
    return datetime.utcnow() + LOCAL_TIME_OFFSET


def _local_today() -> date:
    return _local_now().date()


def _local_date_start_utc(local_day: date) -> datetime:
    return datetime.combine(local_day, time.min) - LOCAL_TIME_OFFSET


def _local_dates_utc_bounds(start_day: date, end_exclusive_day: date) -> tuple[datetime, datetime]:
    return _local_date_start_utc(start_day), _local_date_start_utc(end_exclusive_day)


def _platform_date_bounds(start_day: date, end_exclusive_day: date) -> tuple[datetime, datetime]:
    return _local_dates_utc_bounds(start_day, end_exclusive_day)


def _local_day_utc_bounds(value: str | date | None = None) -> tuple[datetime, datetime]:
    if isinstance(value, str) and value:
        local_day = datetime.fromisoformat(value).date()
    elif isinstance(value, date):
        local_day = value
    else:
        local_day = _local_today()
    start_utc = _local_date_start_utc(local_day)
    return start_utc, start_utc + timedelta(days=1)


def _dashboard_number(value, digits: int = 2) -> float:
    if value is None:
        return 0
    return round(float(value), digits)


def _dashboard_int(value) -> int:
    return int(value or 0)


@dataclass(frozen=True)
class DashboardShopScope:
    """Resolved dashboard shop scope shared by all dashboard aggregations."""

    shop_ids: tuple[int, ...] = ()
    keys: tuple[tuple[str, str], ...] = ()

    @property
    def is_filtered(self) -> bool:
        return bool(self.keys)

    def params(self) -> dict[str, str | None]:
        if not self.keys:
            return {"dashboard_shop_scope": None}
        return {
            "dashboard_shop_scope": json.dumps(
                [{"platform": platform, "account_id": account_id} for platform, account_id in self.keys],
                ensure_ascii=False,
            )
        }


def _dashboard_shop_scope(db: Session, shop_ids: list[int] | None) -> DashboardShopScope:
    raw_shop_ids = shop_ids if isinstance(shop_ids, (list, tuple, set)) else []
    selected_ids = tuple(sorted({int(shop_id) for shop_id in raw_shop_ids}))
    if not selected_ids:
        return DashboardShopScope()
    if len(selected_ids) > 100:
        raise HTTPException(status_code=422, detail="最多只能选择 100 家店铺")

    rows = db.scalars(select(PlatformAccount).where(PlatformAccount.id.in_(selected_ids))).all()
    rows_by_id = {row.id: row for row in rows}
    missing_ids = [shop_id for shop_id in selected_ids if shop_id not in rows_by_id]
    if missing_ids:
        raise HTTPException(status_code=422, detail="所选店铺不存在或已不可用")

    keys = tuple(
        sorted(
            {
                (_canonical_platform(row.platform), str(row.account_id).strip())
                for row in rows
                if _canonical_platform(row.platform) and str(row.account_id or "").strip()
            }
        )
    )
    if not keys:
        raise HTTPException(status_code=422, detail="所选店铺缺少有效的平台账号标识")
    return DashboardShopScope(shop_ids=selected_ids, keys=keys)


def _dashboard_order_platform_sql(order_alias: str = "") -> str:
    prefix = f"{order_alias}." if order_alias else "orders."
    raw_platform = f"LOWER(BTRIM(COALESCE({prefix}platform, '')))"
    aliases = sorted(PLATFORM_ALIASES.items())
    when_clauses = " ".join(
        f"WHEN {raw_platform} = '{alias}' THEN '{canonical}'" for alias, canonical in aliases
    )
    return f"CASE {when_clauses} ELSE {raw_platform} END"


def _dashboard_shop_scope_sql(order_alias: str = "") -> str:
    prefix = f"{order_alias}." if order_alias else "orders."
    return f"""
                      AND (
                          CAST(:dashboard_shop_scope AS jsonb) IS NULL
                          OR EXISTS (
                              SELECT 1
                              FROM jsonb_to_recordset(CAST(:dashboard_shop_scope AS jsonb))
                                   AS selected_shop(platform text, account_id text)
                              WHERE selected_shop.platform = {_dashboard_order_platform_sql(order_alias)}
                                AND selected_shop.account_id IN (
                                    NULLIF(BTRIM({prefix}account_id), ''),
                                    NULLIF(BTRIM({prefix}shop_id), '')
                                )
                          )
                      )
    """


def _dashboard_currency_code_sql(order_alias: str = "") -> str:
    prefix = f"{order_alias}." if order_alias else "orders."
    currency_column = f"{prefix}currency"
    payload_column = f"{prefix}raw_payload"
    id_column = f"{prefix}id"
    return (
        f"UPPER(NULLIF(BTRIM(COALESCE(\n"
        f"                               NULLIF(BTRIM({currency_column}), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} ->> 'currency_code'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} ->> 'currency'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} ->> 'money_currency'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{products,0,currency_code}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{products,0,currency_id}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{products,0,currency}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{products,0,price,currency}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{financial_data,products,0,currency_code}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{financial_data,products,0,currency_id}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{financial_data,products,0,currency}}'), ''),\n"
        f"                               NULLIF(BTRIM({payload_column} #>> '{{financial_data,products,0,price,currency}}'), ''),\n"
        f"                               (\n"
        f"                                   SELECT NULLIF(BTRIM(oi.currency), '')\n"
        f"                                   FROM order_items oi\n"
        f"                                   WHERE oi.order_id = {id_column}\n"
        f"                                     AND NULLIF(BTRIM(oi.currency), '') IS NOT NULL\n"
        f"                                   ORDER BY oi.id\n"
        f"                                   LIMIT 1\n"
        f"                               )\n"
        f"                           )), ''))"
    )


def _dashboard_amount_fields_sql(order_alias: str = "") -> str:
    prefix = f"{order_alias}." if order_alias else "orders."
    amount_column = f"{prefix}order_amount"
    return (
        f"CASE WHEN {amount_column} ~ '^[+-]?[0-9]+(\\.[0-9]+)?$' "
        f"THEN {amount_column}::numeric ELSE NULL END AS amount_numeric,\n"
        f"                           {_dashboard_currency_code_sql(order_alias)} AS currency_code"
    )


def _dashboard_order_count_key_sql(order_alias: str = "orders") -> str:
    prefix = f"{order_alias}." if order_alias else ""
    return (
        "CASE WHEN LOWER(BTRIM(COALESCE("
        f"{prefix}platform, ''))) IN ('joom', 'joomlogistics', 'joom_logistics') "
        "THEN CONCAT('joom:', COALESCE(NULLIF(BTRIM("
        f"{prefix}account_id), ''), NULLIF(BTRIM({prefix}shop_id), ''), ''), ':', COALESCE("
        f"NULLIF(BTRIM({prefix}raw_payload ->> 'transactionId'), ''), "
        f"NULLIF(BTRIM({prefix}raw_payload ->> 'transaction_id'), ''), "
        f"NULLIF(BTRIM({prefix}platform_order_id), ''), {prefix}id::text)) "
        f"ELSE CONCAT('row:', {prefix}id::text) END"
    )


_DASHBOARD_CNY_AMOUNT_CTE_SQL = """
converted AS (
    SELECT base.*,
           CASE
               WHEN base.amount_numeric IS NULL THEN NULL
               WHEN base.currency_code IN ('CNY', 'RMB') THEN base.amount_numeric
               WHEN fx.rate IS NOT NULL THEN base.amount_numeric * fx.rate
               ELSE NULL
           END AS cny_amount
    FROM base
    LEFT JOIN LATERAL (
        SELECT er.rate
        FROM exchange_rates er
        WHERE er.currency_code = base.currency_code
        ORDER BY CASE
                     WHEN base.order_date IS NULL THEN 0
                     ELSE ABS(er.rate_date - base.order_date)
                 END ASC,
                 er.rate_date DESC,
                 er.updated_at DESC
        LIMIT 1
    ) fx ON base.currency_code IS NOT NULL AND base.currency_code NOT IN ('CNY', 'RMB')
)
"""

_DASHBOARD_RECEIPT_AMOUNT_CTE_SQL = """
receipted AS (
    SELECT converted.*,
           COALESCE(exact_rate.receipt_rate, default_rate.receipt_rate, 1) AS receipt_rate,
           cny_amount * COALESCE(exact_rate.receipt_rate, default_rate.receipt_rate, 1) AS expected_receipt
    FROM converted
    LEFT JOIN dashboard_platform_settings exact_rate
      ON exact_rate.platform = CASE LOWER(BTRIM(COALESCE(converted.platform, '')))
          WHEN 'joom' THEN 'joom_logistics'
          WHEN 'joomlogistics' THEN 'joom_logistics'
          WHEN 'mercado' THEN 'mercadolibre'
          WHEN 'mercado_libre' THEN 'mercadolibre'
          WHEN 'dms_matrix' THEN 'dmsmatrix'
          WHEN 'dms-matrix' THEN 'dmsmatrix'
          ELSE LOWER(BTRIM(COALESCE(converted.platform, '')))
      END
    LEFT JOIN dashboard_platform_settings default_rate ON default_rate.platform = 'other'
)
"""


def _dashboard_text_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _dashboard_text_datetime(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=" ")
    return str(value)


def _dashboard_growth_pct(current: float, previous: float) -> float:
    if not previous:
        return 0
    return round((current - previous) * 100 / previous, 2)


def _dashboard_period(fallback_end: date, date_from: date | None, date_to: date | None) -> tuple[date, date, date, date]:
    date_from = date_from if isinstance(date_from, date) else None
    date_to = date_to if isinstance(date_to, date) else None
    end_day = date_to or fallback_end
    start_day = date_from or end_day.replace(day=1)
    if start_day > end_day:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    if (end_day - start_day).days > 366:
        raise HTTPException(status_code=400, detail="统计日期范围不能超过367天")
    previous_end = start_day - timedelta(days=1)
    previous_start = previous_end - (end_day - start_day)
    return start_day, end_day, previous_start, previous_end


def _dashboard_comparison_period(
    start_day: date,
    end_day: date,
    default_start: date,
    default_end: date,
    compare_from: date | None,
    compare_to: date | None,
) -> tuple[date, date]:
    compare_from = compare_from if isinstance(compare_from, date) else None
    compare_to = compare_to if isinstance(compare_to, date) else None
    if (compare_from is None) != (compare_to is None):
        raise HTTPException(status_code=400, detail="对比周期需要同时提供开始和结束日期")
    if compare_from is None or compare_to is None:
        return default_start, default_end
    if compare_from > compare_to:
        raise HTTPException(status_code=400, detail="对比周期开始日期不能晚于结束日期")
    if compare_to - compare_from != end_day - start_day:
        raise HTTPException(status_code=400, detail="对比周期天数必须与统计周期一致")
    return compare_from, compare_to


def _dashboard_base_params() -> dict[str, str]:
    return {
        "pending": ORDER_STATUS_PENDING,
        "waiting_print": ORDER_STATUS_WAITING_PRINT,
        "waiting_purchase": ORDER_STATUS_WAITING_PURCHASE,
        "picking": ORDER_STATUS_PICKING,
        "shipped": ORDER_STATUS_SHIPPED,
        "delivered": ORDER_STATUS_DELIVERED,
        "voided": ORDER_STATUS_VOIDED,
    }


def _dashboard_summary(db: Session, scope: DashboardShopScope | None = None):
    scope = scope or DashboardShopScope()
    return db.execute(
        text(
            f"""
            SELECT COUNT(*) AS total_orders,
                   MIN(payment_at::date) AS first_order_date,
                   MAX(payment_at::date) AS last_order_date,
                   COUNT(*) FILTER (WHERE currency_code IS NULL) AS blank_currency_orders
            FROM (
                SELECT payment_at,
                       {_dashboard_currency_code_sql()} AS currency_code
                FROM orders
                WHERE 1 = 1
                {_dashboard_shop_scope_sql()}
            ) base
            """
        ),
        scope.params(),
    ).one()


def _dashboard_last_order_date(db: Session, scope: DashboardShopScope | None = None) -> date:
    scope = scope or DashboardShopScope()
    return db.execute(
        text(
            f"""
            SELECT MAX(payment_at::date) AS last_order_date
            FROM orders
            WHERE 1 = 1
            {_dashboard_shop_scope_sql()}
            """
        ),
        scope.params(),
    ).scalar_one() or _local_today()


def _dashboard_date_start(value: date) -> datetime:
    return datetime.combine(value, time.min)


def _dashboard_context(
    db: Session,
    date_from: date | None = None,
    date_to: date | None = None,
    compare_from: date | None = None,
    compare_to: date | None = None,
    scope: DashboardShopScope | None = None,
):
    summary = _dashboard_summary(db, scope)
    fallback_end = summary.last_order_date or _local_today()
    start_day, end_day, previous_start, previous_end = _dashboard_period(fallback_end, date_from, date_to)
    previous_start, previous_end = _dashboard_comparison_period(
        start_day,
        end_day,
        previous_start,
        previous_end,
        compare_from,
        compare_to,
    )
    return summary, start_day, end_day, previous_start, previous_end


def _dashboard_mtd_comparison(
    db: Session,
    start_day: date,
    end_day: date,
    previous_start: date,
    previous_end: date,
    scope: DashboardShopScope | None = None,
) -> DashboardMtdComparisonDto:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}

    def period_stats(start_day: date, end_day: date):
        return db.execute(
            text(
                f"""
                WITH base AS (
                    SELECT orders.id,
                           orders.platform,
                           {_dashboard_order_count_key_sql()} AS order_count_key,
                           payment_at::date AS order_date,
                           biz_status,
                           {_dashboard_amount_fields_sql()}
                    FROM orders
                    WHERE payment_at >= :start_at
                      AND payment_at < :end_at
                      {_dashboard_shop_scope_sql()}
                ),
                {_DASHBOARD_CNY_AMOUNT_CTE_SQL},
                {_DASHBOARD_RECEIPT_AMOUNT_CTE_SQL}
                SELECT COUNT(DISTINCT order_count_key) AS orders,
                       COALESCE(SUM(cny_amount) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS raw_amount,
                       COALESCE(SUM(expected_receipt) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS expected_receipt,
                       COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :pending) AS pending,
                       COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :voided) AS voided
                FROM receipted
                """
            ),
            {
                **base_params,
                "start_at": _dashboard_date_start(start_day),
                "end_at": _dashboard_date_start(end_day + timedelta(days=1)),
            },
        ).one()

    current_period = period_stats(start_day, end_day)
    previous_period = period_stats(previous_start, previous_end)
    current_orders = _dashboard_int(current_period.orders)
    previous_orders = _dashboard_int(previous_period.orders)
    current_amount = _dashboard_number(current_period.raw_amount)
    previous_amount = _dashboard_number(previous_period.raw_amount)
    current_receipt = _dashboard_number(current_period.expected_receipt)
    previous_receipt = _dashboard_number(previous_period.expected_receipt)
    return DashboardMtdComparisonDto(
        current_label=f"{start_day.isoformat()}~{end_day.isoformat()}",
        previous_label=f"{previous_start.isoformat()}~{previous_end.isoformat()}",
        current_orders=current_orders,
        previous_orders=previous_orders,
        order_growth_pct=_dashboard_growth_pct(current_orders, previous_orders),
        current_amount=current_amount,
        previous_amount=previous_amount,
        amount_growth_pct=_dashboard_growth_pct(current_amount, previous_amount),
        current_receipt=current_receipt,
        previous_receipt=previous_receipt,
        receipt_growth_pct=_dashboard_growth_pct(current_receipt, previous_receipt),
        current_pending=_dashboard_int(current_period.pending),
        current_voided=_dashboard_int(current_period.voided),
        previous_pending=_dashboard_int(previous_period.pending),
        previous_voided=_dashboard_int(previous_period.voided),
    )


def _dashboard_monthly_sales(
    db: Session,
    start_day: date,
    end_day: date,
    scope: DashboardShopScope | None = None,
) -> list[DashboardMonthlySalesDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    monthly_rows = db.execute(
        text(
            f"""
                WITH base AS (
                    SELECT orders.id,
                           orders.platform,
                           {_dashboard_order_count_key_sql()} AS order_count_key,
                       payment_at::date AS order_date,
                       biz_status,
                       {_dashboard_amount_fields_sql()}
                FROM orders
                    WHERE payment_at >= :start_at
                      AND payment_at < :end_at
                      {_dashboard_shop_scope_sql()}
                ),
                {_DASHBOARD_CNY_AMOUNT_CTE_SQL},
                {_DASHBOARD_RECEIPT_AMOUNT_CTE_SQL}
            SELECT TO_CHAR(DATE_TRUNC('month', order_date), 'YYYY-MM') AS month,
                   COUNT(DISTINCT order_count_key) AS orders,
                   ROUND(COUNT(DISTINCT order_count_key)::numeric / NULLIF(MAX(order_date) - MIN(order_date) + 1, 0), 2) AS avg_daily_orders,
                   COALESCE(SUM(cny_amount) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS raw_amount,
                   COALESCE(SUM(cny_amount) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0)
                     / NULLIF(COUNT(DISTINCT order_count_key) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS raw_aov,
                   COALESCE(SUM(expected_receipt) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS expected_receipt,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :pending) AS pending,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :picking) AS picking,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :shipped) AS shipped,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :delivered) AS delivered,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :voided) AS voided,
                   ROUND(100.0 * COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :voided)
                     / NULLIF(COUNT(DISTINCT order_count_key), 0), 2) AS voided_rate,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE currency_code IS NULL) AS blank_currency_orders
            FROM receipted
            GROUP BY DATE_TRUNC('month', order_date)
            ORDER BY DATE_TRUNC('month', order_date)
            """
        ),
        {
            **base_params,
            "start_at": _dashboard_date_start(start_day),
            "end_at": _dashboard_date_start(end_day + timedelta(days=1)),
        },
    ).all()
    return [
        DashboardMonthlySalesDto(
            month=row.month,
            orders=_dashboard_int(row.orders),
            avg_daily_orders=_dashboard_number(row.avg_daily_orders),
            raw_amount=_dashboard_number(row.raw_amount),
            raw_aov=_dashboard_number(row.raw_aov),
            expected_receipt=_dashboard_number(row.expected_receipt),
            pending=_dashboard_int(row.pending),
            picking=_dashboard_int(row.picking),
            shipped=_dashboard_int(row.shipped),
            delivered=_dashboard_int(row.delivered),
            voided=_dashboard_int(row.voided),
            voided_rate=_dashboard_number(row.voided_rate),
            blank_currency_orders=_dashboard_int(row.blank_currency_orders),
        )
        for row in monthly_rows
    ]


def _dashboard_shop_sales(
    db: Session,
    month_start: date,
    last_order_date: date,
    scope: DashboardShopScope | None = None,
) -> list[DashboardShopSalesDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    shop_sales_rows = db.execute(
        text(
            f"""
            WITH base AS (
                SELECT orders.id,
                       COALESCE(NULLIF(BTRIM(orders.platform), ''), '未记录平台') AS platform,
                       COALESCE(NULLIF(BTRIM(orders.shop_name), ''), NULLIF(BTRIM(orders.shop_id), ''), '未命名店铺') AS shop,
                       {_dashboard_order_count_key_sql()} AS order_count_key,
                       payment_at::date AS order_date,
                       biz_status,
                       {_dashboard_amount_fields_sql()}
                FROM orders
                WHERE payment_at >= :start_at
                  AND payment_at < :end_at
                  {_dashboard_shop_scope_sql()}
                ),
            {_DASHBOARD_CNY_AMOUNT_CTE_SQL},
            {_DASHBOARD_RECEIPT_AMOUNT_CTE_SQL}
            SELECT platform,
                   shop,
                   COUNT(DISTINCT order_count_key) AS orders,
                   COALESCE(SUM(cny_amount) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS raw_amount,
                   COALESCE(SUM(cny_amount) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0)
                     / NULLIF(COUNT(DISTINCT order_count_key) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS raw_aov,
                   COALESCE(SUM(expected_receipt) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS expected_receipt,
                   MAX(receipt_rate) * 100 AS receipt_rate_pct,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :voided) AS voided,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE currency_code IS NULL) AS blank_currency_orders
            FROM receipted
            GROUP BY platform, shop
            ORDER BY raw_amount DESC, orders DESC
            LIMIT 12
            """
        ),
        {
            **base_params,
            "start_at": _dashboard_date_start(month_start),
            "end_at": _dashboard_date_start(last_order_date + timedelta(days=1)),
        },
    ).all()
    return [
        DashboardShopSalesDto(
            platform=row.platform or "",
            shop=row.shop or "未命名店铺",
            orders=_dashboard_int(row.orders),
            raw_amount=_dashboard_number(row.raw_amount),
            raw_aov=_dashboard_number(row.raw_aov),
            expected_receipt=_dashboard_number(row.expected_receipt),
            receipt_rate_pct=_dashboard_number(row.receipt_rate_pct, 2),
            voided=_dashboard_int(row.voided),
            blank_currency_orders=_dashboard_int(row.blank_currency_orders),
        )
        for row in shop_sales_rows
    ]


def _dashboard_daily_sales(
    db: Session,
    start_day: date,
    end_day: date,
    scope: DashboardShopScope | None = None,
) -> list[DashboardDailySalesDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    daily_rows = db.execute(
        text(
            f"""
                WITH base AS (
                    SELECT orders.id,
                           orders.platform,
                           {_dashboard_order_count_key_sql()} AS order_count_key,
                       payment_at::date AS order_date,
                       biz_status,
                       {_dashboard_amount_fields_sql()}
                FROM orders
                    WHERE payment_at >= :start_at
                      AND payment_at < :end_at
                      {_dashboard_shop_scope_sql()}
                ),
            {_DASHBOARD_CNY_AMOUNT_CTE_SQL},
            {_DASHBOARD_RECEIPT_AMOUNT_CTE_SQL}
            SELECT order_date,
                   COUNT(DISTINCT order_count_key) AS orders,
                   COALESCE(SUM(cny_amount) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS raw_amount,
                   COALESCE(SUM(expected_receipt) FILTER (WHERE COALESCE(biz_status, '') <> :voided), 0) AS expected_receipt,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :pending) AS pending,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :voided) AS voided
            FROM receipted
            GROUP BY order_date
            ORDER BY order_date
            """
        ),
        {
            **base_params,
            "start_at": _dashboard_date_start(start_day),
            "end_at": _dashboard_date_start(end_day + timedelta(days=1)),
        },
    ).all()
    return [
        DashboardDailySalesDto(
            date=_dashboard_text_date(row.order_date) or "",
            orders=_dashboard_int(row.orders),
            raw_amount=_dashboard_number(row.raw_amount),
            expected_receipt=_dashboard_number(row.expected_receipt),
            pending=_dashboard_int(row.pending),
            voided=_dashboard_int(row.voided),
        )
        for row in daily_rows
    ]


def _dashboard_risk_buckets(
    db: Session,
    scope: DashboardShopScope | None = None,
) -> list[DashboardRiskBucketDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    risk_bucket_labels = {
        "overdue_48": "已超48h+",
        "overdue_24_48": "已超24-48h",
        "overdue_0_24": "已超0-24h",
        "due_24": "24h内到期",
        "due_48": "24-48h到期",
        "due_later": "48h后到期",
        "no_deadline": "无截止时间",
    }
    risk_rows = db.execute(
        text(
            f"""
            WITH base AS (
                SELECT orders.id,
                       {_dashboard_order_count_key_sql()} AS order_count_key,
                       payment_at::date AS order_date,
                       {_dashboard_amount_fields_sql()},
                       COALESCE(dispatch_deadline_at, shipping_deadline_at) AS deadline_at,
                       CASE WHEN COALESCE(dispatch_deadline_at, shipping_deadline_at) IS NULL THEN NULL
                            ELSE EXTRACT(EPOCH FROM (COALESCE(dispatch_deadline_at, shipping_deadline_at) - timezone('UTC', NOW()))) / 3600
                       END AS hours_to_deadline
                FROM orders
                WHERE biz_status IN (:pending, :waiting_print, :waiting_purchase, :picking)
                  AND COALESCE(dispatch_deadline_at, shipping_deadline_at) IS NOT NULL
                  AND COALESCE(dispatch_deadline_at, shipping_deadline_at) < timezone('UTC', NOW()) + INTERVAL '24 hours'
                  AND NOT EXISTS (
                      SELECT 1 FROM order_risk_handlings risk_handling
                      WHERE risk_handling.order_id = orders.id
                  )
                  {_dashboard_shop_scope_sql()}
            ),
            {_DASHBOARD_CNY_AMOUNT_CTE_SQL},
            bucketed AS (
                SELECT *,
                       CASE
                           WHEN deadline_at IS NULL THEN 'no_deadline'
                           WHEN hours_to_deadline < -48 THEN 'overdue_48'
                           WHEN hours_to_deadline < -24 THEN 'overdue_24_48'
                           WHEN hours_to_deadline < 0 THEN 'overdue_0_24'
                           WHEN hours_to_deadline < 24 THEN 'due_24'
                           WHEN hours_to_deadline < 48 THEN 'due_48'
                           ELSE 'due_later'
                       END AS risk_key
                FROM converted
            )
            SELECT risk_key,
                   COUNT(DISTINCT order_count_key) AS orders,
                   COALESCE(SUM(cny_amount), 0) AS raw_amount,
                   MIN(deadline_at) AS earliest_deadline,
                   MAX(deadline_at) AS latest_deadline
            FROM bucketed
            GROUP BY risk_key
            ORDER BY CASE risk_key
                WHEN 'overdue_48' THEN 1
                WHEN 'overdue_24_48' THEN 2
                WHEN 'overdue_0_24' THEN 3
                WHEN 'due_24' THEN 4
                WHEN 'due_48' THEN 5
                WHEN 'due_later' THEN 6
                ELSE 7
            END
            """
        ),
        base_params,
    ).all()
    return [
        DashboardRiskBucketDto(
            key=row.risk_key,
            label=risk_bucket_labels.get(row.risk_key, row.risk_key),
            orders=_dashboard_int(row.orders),
            raw_amount=_dashboard_number(row.raw_amount),
            earliest_deadline=_dashboard_text_datetime(row.earliest_deadline),
            latest_deadline=_dashboard_text_datetime(row.latest_deadline),
        )
        for row in risk_rows
    ]


def _dashboard_risk_shops(
    db: Session,
    scope: DashboardShopScope | None = None,
) -> list[DashboardRiskShopDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    risk_shop_rows = db.execute(
        text(
            f"""
            WITH pending_orders AS (
                SELECT o.id,
                       o.platform,
                       {_dashboard_order_count_key_sql("o")} AS order_count_key,
                       COALESCE(NULLIF(o.shop_name, ''), o.shop_id) AS shop,
                       o.payment_at::date AS order_date,
                       {_dashboard_amount_fields_sql("o")},
                       COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) AS deadline_at,
                       EXTRACT(EPOCH FROM (COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) - timezone('UTC', NOW()))) / 3600 AS hours_to_deadline
                FROM orders o
                WHERE o.biz_status IN (:pending, :waiting_print, :waiting_purchase, :picking)
                  AND COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) IS NOT NULL
                  AND COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) < timezone('UTC', NOW()) + INTERVAL '24 hours'
                  AND NOT EXISTS (
                      SELECT 1 FROM order_risk_handlings risk_handling
                      WHERE risk_handling.order_id = o.id
                  )
                  {_dashboard_shop_scope_sql("o")}
            ),
            item_agg AS (
                SELECT oi.order_id, SUM(oi.quantity) AS units
                FROM order_items oi
                JOIN pending_orders po ON po.id = oi.order_id
                GROUP BY oi.order_id
            ),
            base AS (
                SELECT po.*,
                       COALESCE(ia.units, 0) AS units
                FROM pending_orders po
                LEFT JOIN item_agg ia ON ia.order_id = po.id
            ),
            {_DASHBOARD_CNY_AMOUNT_CTE_SQL}
            SELECT platform,
                   shop,
                   COUNT(DISTINCT order_count_key) AS pending_orders,
                   COALESCE(SUM(units), 0) AS pending_units,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE hours_to_deadline < 0) AS overdue_orders,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE hours_to_deadline >= 0 AND hours_to_deadline < 24) AS due_24h,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE hours_to_deadline >= 24 AND hours_to_deadline < 48) AS due_48h,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE hours_to_deadline >= 48) AS due_later,
                   COALESCE(SUM(cny_amount), 0) AS raw_amount,
                   MIN(hours_to_deadline) AS min_hours_to_deadline,
                   MIN(deadline_at) AS earliest_deadline
            FROM converted
            GROUP BY platform, shop
            ORDER BY overdue_orders DESC, pending_orders DESC
            LIMIT 8
            """
        ),
        base_params,
    ).all()
    return [
        DashboardRiskShopDto(
            platform=row.platform,
            shop=row.shop,
            pending_orders=_dashboard_int(row.pending_orders),
            pending_units=_dashboard_int(row.pending_units),
            overdue_orders=_dashboard_int(row.overdue_orders),
            due_24h=_dashboard_int(row.due_24h),
            due_48h=_dashboard_int(row.due_48h),
            due_later=_dashboard_int(row.due_later),
            raw_amount=_dashboard_number(row.raw_amount),
            min_hours_to_deadline=_dashboard_number(row.min_hours_to_deadline, 1) if row.min_hours_to_deadline is not None else None,
            earliest_deadline=_dashboard_text_datetime(row.earliest_deadline),
        )
        for row in risk_shop_rows
    ]


def _dashboard_risk_skus(
    db: Session,
    scope: DashboardShopScope | None = None,
) -> list[DashboardRiskSkuDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    risk_sku_rows = db.execute(
        text(
            f"""
            WITH normalized_product_mappings AS MATERIALIZED (
                SELECT DISTINCT ON (
                           psm.shop_id,
                           LOWER(BTRIM(COALESCE(psm.shop_sku, '')))
                       )
                       psm.shop_id,
                       LOWER(BTRIM(COALESCE(psm.shop_sku, ''))) AS normalized_sku,
                       p.internal_name AS product_name
                FROM product_shop_mappings psm
                JOIN products p ON p.id = psm.product_id
                WHERE NULLIF(BTRIM(COALESCE(psm.shop_sku, '')), '') IS NOT NULL
                ORDER BY psm.shop_id,
                         LOWER(BTRIM(COALESCE(psm.shop_sku, ''))),
                         COALESCE(psm.updated_at, psm.created_at) DESC,
                         psm.id DESC
            ),
            pending_orders AS (
                SELECT id,
                       platform,
                       {_dashboard_order_count_key_sql()} AS order_count_key,
                       COALESCE(NULLIF(shop_id, ''), account_id) AS account_id,
                       COALESCE(NULLIF(shop_name, ''), shop_id) AS shop,
                       COALESCE(dispatch_deadline_at, shipping_deadline_at) AS deadline_at
                FROM orders
                WHERE biz_status IN (:pending, :waiting_print, :waiting_purchase, :picking)
                  AND COALESCE(dispatch_deadline_at, shipping_deadline_at) IS NOT NULL
                  AND COALESCE(dispatch_deadline_at, shipping_deadline_at) < timezone('UTC', NOW()) + INTERVAL '24 hours'
                  AND NOT EXISTS (
                      SELECT 1 FROM order_risk_handlings risk_handling
                      WHERE risk_handling.order_id = orders.id
                  )
                  {_dashboard_shop_scope_sql()}
            )
            SELECT oi.sku,
                   COALESCE(
                       NULLIF(MAX(exact_product.internal_name), ''),
                       NULLIF(MAX(normalized_mapping.product_name), ''),
                       ''
                   ) AS product_name,
                   COUNT(DISTINCT po.order_count_key) AS pending_orders,
                   COALESCE(SUM(oi.quantity), 0) AS pending_units,
                   COUNT(DISTINCT po.shop) AS shops,
                   COUNT(DISTINCT po.order_count_key) FILTER (WHERE po.deadline_at < timezone('UTC', NOW())) AS overdue_orders,
                   MIN(po.deadline_at) AS earliest_deadline
            FROM pending_orders po
            JOIN order_items oi ON oi.order_id = po.id
            LEFT JOIN platform_accounts pa
              ON pa.platform = po.platform
             AND pa.account_id = po.account_id
            LEFT JOIN product_shop_mappings exact_mapping
              ON exact_mapping.shop_id = pa.id
             AND exact_mapping.shop_sku = oi.sku
            LEFT JOIN products exact_product ON exact_product.id = exact_mapping.product_id
            LEFT JOIN normalized_product_mappings normalized_mapping
              ON normalized_mapping.shop_id = pa.id
             AND normalized_mapping.normalized_sku = LOWER(BTRIM(COALESCE(oi.sku, '')))
            GROUP BY oi.sku
            ORDER BY overdue_orders DESC, pending_units DESC, pending_orders DESC
            LIMIT 10
            """
        ),
        base_params,
    ).all()
    return [
        DashboardRiskSkuDto(
            sku=row.sku,
            product_name=row.product_name or "",
            pending_orders=_dashboard_int(row.pending_orders),
            pending_units=_dashboard_int(row.pending_units),
            shops=_dashboard_int(row.shops),
            overdue_orders=_dashboard_int(row.overdue_orders),
            earliest_deadline=_dashboard_text_datetime(row.earliest_deadline),
        )
        for row in risk_sku_rows
    ]


def _operations_daily_report_active_shops(db: Session) -> list[tuple[str, str, str]]:
    rows = db.scalars(
        select(PlatformAccount)
        .where(PlatformAccount.enabled.is_(True))
        .order_by(PlatformAccount.platform, PlatformAccount.display_name, PlatformAccount.account_id)
    ).all()
    shops: list[tuple[str, str, str]] = []
    for row in rows:
        platform = _canonical_platform(row.platform)
        account_id = str(row.account_id or "").strip()
        if not platform or not account_id:
            continue
        shop = str(row.display_name or "").strip() or account_id
        shops.append((platform, account_id, shop))
    return shops


def _operations_daily_report_customer_complaints(
    db: Session,
    issue_day: date,
) -> list[OperationsCustomerComplaintDto]:
    complaint_rows = db.execute(
        text(
            """
            WITH active_accounts AS (
                SELECT pa.id AS platform_account_id,
                       pa.platform,
                       pa.account_id,
                       COALESCE(
                           NULLIF(BTRIM(pa.display_name), ''),
                           NULLIF(BTRIM(pa.account_id), ''),
                           ''
                       ) AS shop
                FROM platform_accounts pa
                WHERE pa.enabled IS TRUE
            )
            SELECT aa.platform,
                   aa.platform_account_id,
                   aa.account_id,
                   aa.shop,
                   COALESCE(SUM(tm.negative_reviews), 0) AS negative_review_count,
                   MAX(tm.stat_date) AS latest_issue_at
            FROM traffic_metrics tm
            JOIN active_accounts aa ON aa.platform_account_id = tm.platform_account_id
            WHERE tm.grain = 'daily'
              AND tm.stat_date = :issue_day
              AND COALESCE(tm.negative_reviews, 0) > 0
              AND aa.shop <> ''
            GROUP BY aa.platform, aa.platform_account_id, aa.account_id, aa.shop
            HAVING COALESCE(SUM(tm.negative_reviews), 0) > 0
            ORDER BY negative_review_count DESC, latest_issue_at DESC, aa.platform, aa.shop
            """
        ),
        {"issue_day": issue_day},
    ).all()
    return [
        OperationsCustomerComplaintDto(
            platform=_canonical_platform(row.platform),
            shop=str(row.shop or "").strip(),
            count=_dashboard_int(row.negative_review_count),
            latest_issue_at=_dashboard_text_date(row.latest_issue_at),
        )
        for row in complaint_rows
    ]


def _operations_daily_report(
    db: Session,
    days: int = 7,
    report_date: date | None = None,
) -> OperationsDailyReportResponse:
    today = _local_today()
    latest_complete_day = today - timedelta(days=1)
    end_day = report_date or latest_complete_day
    if end_day > latest_complete_day:
        raise HTTPException(status_code=422, detail="统计日期不能晚于昨天")
    start_day = end_day - timedelta(days=days - 1)
    day_labels = [(start_day + timedelta(days=index)).isoformat() for index in range(days)]
    shops = _operations_daily_report_active_shops(db)
    shop_keys = {(platform, account_id) for platform, account_id, _ in shops}

    daily_rows = db.execute(
        text(
            f"""
            WITH base AS (
                SELECT {_dashboard_order_platform_sql('o')} AS platform,
                       COALESCE(NULLIF(BTRIM(o.account_id), ''), NULLIF(BTRIM(o.shop_id), '')) AS account_id,
                       (o.payment_at + INTERVAL '8 hours')::date AS order_date,
                       {_dashboard_order_count_key_sql('o')} AS order_count_key,
                       {_dashboard_amount_fields_sql('o')}
                FROM orders o
                WHERE o.payment_at >= :start_at
                  AND o.payment_at < :end_at
                  AND COALESCE(o.biz_status, '') <> :voided
            ),
            {_DASHBOARD_CNY_AMOUNT_CTE_SQL}
            SELECT platform,
                   account_id,
                   order_date,
                   COUNT(DISTINCT order_count_key) AS orders,
                   COALESCE(SUM(cny_amount), 0) AS revenue_cny
            FROM converted
            WHERE account_id IS NOT NULL
            GROUP BY platform, account_id, order_date
            ORDER BY platform, account_id, order_date
            """
        ),
        {
            "start_at": _local_date_start_utc(start_day),
            "end_at": _local_date_start_utc(end_day + timedelta(days=1)),
            "voided": ORDER_STATUS_VOIDED,
        },
    ).all()
    daily_metrics_by_shop: dict[tuple[str, str], dict[str, tuple[int, float]]] = {}
    for row in daily_rows:
        key = (str(row.platform or ""), str(row.account_id or ""))
        if key not in shop_keys:
            continue
        daily_metrics_by_shop.setdefault(key, {})[_dashboard_text_date(row.order_date) or ""] = (
            _dashboard_int(row.orders),
            _dashboard_number(row.revenue_cny),
        )

    risk_rows = db.execute(
        text(
            f"""
            WITH risk_orders AS (
                SELECT {_dashboard_order_platform_sql('o')} AS platform,
                       {_dashboard_order_count_key_sql('o')} AS order_count_key,
                       COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) AS deadline_at
                FROM orders o
                WHERE o.biz_status IN (:pending, :picking)
                  AND COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) IS NOT NULL
                  AND COALESCE(o.dispatch_deadline_at, o.shipping_deadline_at) < timezone('UTC', NOW()) + INTERVAL '24 hours'
                  AND NOT EXISTS (
                      SELECT 1
                      FROM order_risk_handlings risk_handling
                      WHERE risk_handling.order_id = o.id
                  )
            )
            SELECT platform,
                   COUNT(DISTINCT order_count_key) FILTER (WHERE deadline_at < timezone('UTC', NOW())) AS overdue_orders,
                   COUNT(DISTINCT order_count_key) FILTER (
                       WHERE deadline_at >= timezone('UTC', NOW())
                         AND deadline_at < timezone('UTC', NOW()) + INTERVAL '24 hours'
                   ) AS due_soon_orders
            FROM risk_orders
            GROUP BY platform
            """
        ),
        {"pending": ORDER_STATUS_PENDING, "picking": ORDER_STATUS_PICKING},
    ).all()
    risk_by_platform = {
        str(row.platform or ""): (_dashboard_int(row.overdue_orders), _dashboard_int(row.due_soon_orders))
        for row in risk_rows
    }
    active_platforms = sorted({platform for platform, _, _ in shops})
    customer_complaints = _operations_daily_report_customer_complaints(db, end_day)

    return OperationsDailyReportResponse(
        generated_at=f"{_local_now().replace(microsecond=0).isoformat()}+08:00",
        date_from=start_day.isoformat(),
        date_to=end_day.isoformat(),
        shop_daily_orders=[
            OperationsDailyShopDto(
                platform=platform,
                account_id=account_id,
                shop=shop,
                days=[
                    OperationsDailyOrderPointDto(
                        date=day_label,
                        orders=daily_metrics_by_shop.get((platform, account_id), {}).get(day_label, (0, 0))[0],
                        revenue_cny=daily_metrics_by_shop.get((platform, account_id), {}).get(day_label, (0, 0))[1],
                    )
                    for day_label in day_labels
                ],
                total_orders=sum(metric[0] for metric in daily_metrics_by_shop.get((platform, account_id), {}).values()),
                total_revenue_cny=sum(metric[1] for metric in daily_metrics_by_shop.get((platform, account_id), {}).values()),
            )
            for platform, account_id, shop in shops
        ],
        fulfillment_risk=[
            OperationsFulfillmentRiskDto(
                platform=platform,
                overdue_orders=risk_by_platform.get(platform, (0, 0))[0],
                due_soon_orders=risk_by_platform.get(platform, (0, 0))[1],
            )
            for platform in active_platforms
        ],
        customer_complaints=customer_complaints,
        customer_complaints_data_status="negative_reviews",
    )


def _dashboard_hot_skus(
    db: Session,
    start_day: date,
    end_day: date,
    previous_start: date,
    previous_end: date,
    scope: DashboardShopScope | None = None,
) -> list[DashboardHotSkuDto]:
    scope = scope or DashboardShopScope()
    base_params = {**_dashboard_base_params(), **scope.params()}
    hot_sku_rows = db.execute(
        text(
            f"""
            WITH normalized_product_mappings AS MATERIALIZED (
                SELECT DISTINCT ON (
                           psm.shop_id,
                           LOWER(BTRIM(COALESCE(psm.shop_sku, '')))
                       )
                       psm.shop_id,
                       LOWER(BTRIM(COALESCE(psm.shop_sku, ''))) AS normalized_sku,
                       p.internal_name AS product_name
                FROM product_shop_mappings psm
                JOIN products p ON p.id = psm.product_id
                WHERE NULLIF(BTRIM(COALESCE(psm.shop_sku, '')), '') IS NOT NULL
                ORDER BY psm.shop_id,
                         LOWER(BTRIM(COALESCE(psm.shop_sku, ''))),
                         COALESCE(psm.updated_at, psm.created_at) DESC,
                         psm.id DESC
            ),
            recent_base AS (
                SELECT oi.sku,
                       COALESCE(
                           NULLIF(exact_product.internal_name, ''),
                           NULLIF(normalized_mapping.product_name, ''),
                           ''
                       ) AS product_name,
                       oi.quantity,
                       o.id AS order_id,
                       {_dashboard_order_count_key_sql("o")} AS order_count_key,
                       o.payment_at::date AS order_date,
                       COALESCE(NULLIF(o.shop_name, ''), o.shop_id) AS shop,
                       o.platform,
                       o.biz_status
                FROM order_items oi
                JOIN orders o ON o.id = oi.order_id
                LEFT JOIN platform_accounts pa
                  ON pa.platform = o.platform
                 AND pa.account_id = COALESCE(NULLIF(o.shop_id, ''), o.account_id)
                LEFT JOIN product_shop_mappings exact_mapping
                  ON exact_mapping.shop_id = pa.id
                 AND exact_mapping.shop_sku = oi.sku
                LEFT JOIN products exact_product ON exact_product.id = exact_mapping.product_id
                LEFT JOIN normalized_product_mappings normalized_mapping
                  ON normalized_mapping.shop_id = pa.id
                 AND normalized_mapping.normalized_sku = LOWER(BTRIM(COALESCE(oi.sku, '')))
                WHERE COALESCE(o.biz_status, '') <> :voided
                  AND o.payment_at >= :start_at
                  AND o.payment_at < :end_at
                  {_dashboard_shop_scope_sql("o")}
            ),
            recent_agg AS (
                SELECT sku,
                       MAX(product_name) AS product_name,
                       SUM(quantity) FILTER (
                           WHERE order_date >= :current_start_day
                             AND order_date <= :current_end_day
                       ) AS units_7d,
                       SUM(quantity) FILTER (
                           WHERE order_date >= :previous_start_day
                             AND order_date <= :previous_end_day
                       ) AS units_prev_7d,
                       COUNT(DISTINCT order_count_key) FILTER (WHERE biz_status = :pending) AS pending_orders
                FROM recent_base
                GROUP BY sku
            ),
            top_skus AS (
                SELECT *
                FROM recent_agg
                ORDER BY COALESCE(units_7d, 0) DESC, COALESCE(units_7d, 0) + COALESCE(units_prev_7d, 0) DESC
                LIMIT 12
            ),
            all_base AS (
                SELECT oi.sku,
                       COALESCE(
                           NULLIF(exact_product.internal_name, ''),
                           NULLIF(normalized_mapping.product_name, ''),
                           ''
                       ) AS product_name,
                       oi.quantity,
                       o.id AS order_id,
                       {_dashboard_order_count_key_sql("o")} AS order_count_key,
                       COALESCE(NULLIF(o.shop_name, ''), o.shop_id) AS shop,
                       o.platform
                FROM order_items oi
                JOIN top_skus ts ON ts.sku = oi.sku
                JOIN orders o ON o.id = oi.order_id
                LEFT JOIN platform_accounts pa
                  ON pa.platform = o.platform
                 AND pa.account_id = COALESCE(NULLIF(o.shop_id, ''), o.account_id)
                LEFT JOIN product_shop_mappings exact_mapping
                  ON exact_mapping.shop_id = pa.id
                 AND exact_mapping.shop_sku = oi.sku
                LEFT JOIN products exact_product ON exact_product.id = exact_mapping.product_id
                LEFT JOIN normalized_product_mappings normalized_mapping
                  ON normalized_mapping.shop_id = pa.id
                 AND normalized_mapping.normalized_sku = LOWER(BTRIM(COALESCE(oi.sku, '')))
                WHERE COALESCE(o.biz_status, '') <> :voided
                  {_dashboard_shop_scope_sql("o")}
            ),
            all_agg AS (
                SELECT sku,
                       MAX(product_name) AS product_name,
                       SUM(quantity) AS units_all,
                       COUNT(DISTINCT order_count_key) AS orders_all,
                       COUNT(DISTINCT shop) AS shops,
                       STRING_AGG(DISTINCT platform, ', ' ORDER BY platform) AS platforms
                FROM all_base
                GROUP BY sku
            )
            SELECT ts.sku,
                   COALESCE(NULLIF(aa.product_name, ''), NULLIF(ts.product_name, ''), '') AS product_name,
                   COALESCE(aa.units_all, 0) AS units_all,
                   COALESCE(aa.orders_all, 0) AS orders_all,
                   COALESCE(ts.units_7d, 0) AS units_7d,
                   COALESCE(ts.units_prev_7d, 0) AS units_prev_7d,
                   COALESCE(ts.units_7d, 0) - COALESCE(ts.units_prev_7d, 0) AS units_7d_delta,
                   COALESCE(aa.shops, 0) AS shops,
                   aa.platforms,
                   COALESCE(ts.pending_orders, 0) AS pending_orders
            FROM top_skus ts
            LEFT JOIN all_agg aa ON aa.sku = ts.sku
            ORDER BY ts.units_7d DESC, aa.units_all DESC
            """
        ),
        {
            **base_params,
            "previous_start_day": previous_start,
            "previous_end_day": previous_end,
            "current_start_day": start_day,
            "current_end_day": end_day,
            "start_at": _dashboard_date_start(previous_start),
            "end_at": _dashboard_date_start(end_day + timedelta(days=1)),
        },
    ).all()
    return [
        DashboardHotSkuDto(
            sku=row.sku,
            product_name=row.product_name or "",
            units_all=_dashboard_int(row.units_all),
            orders_all=_dashboard_int(row.orders_all),
            units_7d=_dashboard_int(row.units_7d),
            units_prev_7d=_dashboard_int(row.units_prev_7d),
            units_7d_delta=_dashboard_int(row.units_7d_delta),
            shops=_dashboard_int(row.shops),
            platforms=row.platforms or "",
            pending_orders=_dashboard_int(row.pending_orders),
        )
        for row in hot_sku_rows
    ]


def _outbound_scan_dto(row: OutboundScanRecord) -> OutboundScanRecordDto:
    return OutboundScanRecordDto(
        id=row.id,
        tracking_number=row.tracking_number or "",
        raw_input=row.raw_input or "",
        order_id=row.order_id,
        platform=row.platform or "",
        shop_name=row.shop_name or "",
        platform_order_no=row.platform_order_no or "",
        posting_number=row.posting_number or "",
        order_status=row.order_status or "",
        platform_status=row.platform_status or "",
        result=row.result or "",
        message=row.message or "",
        scanned_by=row.scanned_by or "",
        scanned_at=_iso(row.scanned_at) or "",
        created_at=_iso(row.created_at) or "",
    )


def _find_order_by_tracking_number(db: Session, tracking_number: str) -> Order | None:
    lookup_key = _tracking_number_lookup_key(tracking_number)
    if not lookup_key:
        return None

    row = db.scalar(
        select(Order)
        .where(_tracking_number_matches(Order.shipment_tracking_number, lookup_key))
        .order_by(desc(Order.updated_at), desc(Order.id))
    )
    if row:
        return row
    shipment = db.scalar(
        select(Shipment)
        .where(_tracking_number_matches(Shipment.tracking_number, lookup_key))
        .order_by(desc(Shipment.id))
    )
    if shipment:
        return db.scalar(select(Order).where(Order.id == shipment.order_id))
    raw_order_id = db.execute(
        text(
            """
            SELECT id
            FROM orders
            WHERE LOWER(BTRIM(raw_payload->>'tracking_number')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->>'trackingNumber')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->>'shipment_tracking_number')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'shipment'->>'tracking_number')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'shipment'->>'trackingNumber')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'shipping'->>'tracking_number')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'shipping'->>'trackingNumber')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'logistics'->>'tracking_number')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'logistics'->>'trackingNumber')) = :tracking_lookup_key
               OR LOWER(BTRIM(raw_payload->'tracking'->>'number')) = :tracking_lookup_key
            ORDER BY updated_at DESC, id DESC
            LIMIT 1
            """
        ),
        {"tracking_lookup_key": lookup_key},
    ).scalar_one_or_none()
    if raw_order_id:
        return db.get(Order, raw_order_id)
    return None


def _has_successful_outbound_scan(db: Session, tracking_number: str) -> bool:
    lookup_key = _tracking_number_lookup_key(tracking_number)
    if not lookup_key:
        return False
    return bool(
        db.scalar(
            select(OutboundScanRecord.id)
            .where(
                _tracking_number_matches(OutboundScanRecord.tracking_number, lookup_key),
                OutboundScanRecord.result == "success",
            )
            .limit(1)
        )
    )


def _create_outbound_scan_record(
    db: Session,
    *,
    tracking_number: str,
    raw_input: str,
    result: str,
    message: str,
    scanned_by: str,
    order: Order | None = None,
) -> OutboundScanRecord:
    record = OutboundScanRecord(
        tracking_number=tracking_number,
        raw_input=raw_input,
        order_id=order.id if order else None,
        platform=order.platform if order else "",
        shop_name=(order.shop_name or order.account_id) if order else "",
        platform_order_no=(order.platform_order_no or order.platform_order_id) if order else "",
        posting_number=order.posting_number if order else "",
        order_status=_derive_order_status(order) if order else "",
        platform_status=order.platform_status if order else "",
        result=result,
        message=message,
        scanned_by=scanned_by,
        scanned_at=datetime.utcnow(),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def _outbound_scan_number_search_condition(term: str):
    pattern = f"%{term.strip()}%"
    return or_(
        OutboundScanRecord.tracking_number.ilike(pattern),
        OutboundScanRecord.platform_order_no.ilike(pattern),
        exists().where(
            Order.id == OutboundScanRecord.order_id,
            or_(
                Order.platform_order_no.ilike(pattern),
                Order.platform_order_id.ilike(pattern),
                Order.shipment_tracking_number.ilike(pattern),
                Order.raw_payload["shipment_tracking_number"].astext.ilike(pattern),
                Order.raw_payload["tracking_number"].astext.ilike(pattern),
                Order.raw_payload["trackingNumber"].astext.ilike(pattern),
                Order.raw_payload["shipment"]["tracking_number"].astext.ilike(pattern),
                Order.raw_payload["shipment"]["trackingNumber"].astext.ilike(pattern),
                Order.raw_payload["shipping"]["tracking_number"].astext.ilike(pattern),
                Order.raw_payload["shipping"]["trackingNumber"].astext.ilike(pattern),
                Order.raw_payload["logistics"]["tracking_number"].astext.ilike(pattern),
                Order.raw_payload["logistics"]["trackingNumber"].astext.ilike(pattern),
                Order.raw_payload["tracking"]["number"].astext.ilike(pattern),
                exists().where(Shipment.order_id == Order.id, Shipment.tracking_number.ilike(pattern)),
            ),
        ),
    )


def _build_outbound_scans_query(
    number: str | None,
    platform: str | None,
    shop_name: str | None,
    result: str | None,
    scanned_by: str | None,
    scanned_start: str | None,
    scanned_end: str | None,
):
    stmt = select(OutboundScanRecord)
    if number and number.strip():
        stmt = stmt.where(_outbound_scan_number_search_condition(number.strip()))
    if platform:
        stmt = stmt.where(OutboundScanRecord.platform.ilike(f"%{platform.strip()}%"))
    if shop_name:
        stmt = stmt.where(OutboundScanRecord.shop_name.ilike(f"%{shop_name.strip()}%"))
    if result:
        stmt = stmt.where(OutboundScanRecord.result == result)
    if scanned_by:
        stmt = stmt.where(OutboundScanRecord.scanned_by.ilike(f"%{scanned_by.strip()}%"))
    try:
        if scanned_start:
            start, _ = _local_day_utc_bounds(scanned_start)
            stmt = stmt.where(OutboundScanRecord.scanned_at >= start)
        if scanned_end:
            _, end = _local_day_utc_bounds(scanned_end)
            stmt = stmt.where(OutboundScanRecord.scanned_at < end)
    except (TypeError, ValueError):
        pass
    return stmt.order_by(desc(OutboundScanRecord.scanned_at), desc(OutboundScanRecord.id))


def _build_outbound_scans_count_query(
    number: str | None,
    platform: str | None,
    shop_name: str | None,
    result: str | None,
    scanned_by: str | None,
    scanned_start: str | None,
    scanned_end: str | None,
):
    return select(func.count()).select_from(
        _build_outbound_scans_query(
            number,
            platform,
            shop_name,
            result,
            scanned_by,
            scanned_start,
            scanned_end,
        )
        .order_by(None)
        .subquery()
    )


def _build_shop_query(
    display_name: str | None,
    platform: str | None,
    enabled: bool | None,
    sort_by: str | None,
    sort_order: str,
):
    stmt = select(PlatformAccount)
    if display_name:
        stmt = stmt.where(PlatformAccount.display_name.ilike(f"%{display_name.strip()}%"))
    if platform:
        stmt = stmt.where(PlatformAccount.platform.in_(_platform_lookup_codes(platform)))
    if enabled is not None:
        stmt = stmt.where(PlatformAccount.enabled == enabled)
    sort_map = {
        "display_name": PlatformAccount.display_name,
        "platform": PlatformAccount.platform,
        "enabled": PlatformAccount.enabled,
    }
    sort_column = sort_map.get(sort_by or "", PlatformAccount.created_at)
    order_clause = desc(sort_column) if (sort_order or "desc").lower() == "desc" else asc(sort_column)
    if sort_column is not PlatformAccount.created_at:
        stmt = stmt.order_by(order_clause, desc(PlatformAccount.created_at))
    else:
        stmt = stmt.order_by(order_clause)
    return stmt


def _enabled_product_shops(db: Session) -> list[PlatformAccount]:
    return db.scalars(
        select(PlatformAccount)
        .where(PlatformAccount.enabled == True)
        .order_by(asc(PlatformAccount.display_name), asc(PlatformAccount.id))
    ).all()


def _product_shop_dto(row: PlatformAccount) -> ProductShopDto:
    return ProductShopDto(id=row.id, display_name=row.display_name or row.account_id, platform=row.platform)


def _enabled_buyer_users(db: Session) -> list[LocalUser]:
    rows = db.scalars(
        select(LocalUser)
        .where(LocalUser.enabled == True)
        .order_by(asc(LocalUser.display_name), asc(LocalUser.username))
    ).all()
    return [user for user in rows if not _is_admin_user(user, db)]


def _buyer_user_by_payload(db: Session, user_id: int | None) -> LocalUser | None:
    if not user_id:
        return None
    user = db.get(LocalUser, user_id)
    if not user or not user.enabled:
        raise HTTPException(status_code=400, detail="采购人不存在或已停用")
    if _is_admin_user(user, db):
        raise HTTPException(status_code=400, detail="采购人不能是管理员角色")
    return user


def _user_lookup_for_import(db: Session) -> dict[str, LocalUser]:
    lookup: dict[str, LocalUser] = {}
    for user in _enabled_buyer_users(db):
        for value in {user.username, user.display_name or ""}:
            key = (value or "").strip()
            if key:
                lookup[key] = user
    return lookup


def _decimal_or_none(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise HTTPException(status_code=400, detail=f"数字格式错误: {value}")


def _int_or_none(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        raise HTTPException(status_code=400, detail=f"整数格式错误: {value}")


def _normalize_optional_text(value: str | None, max_length: int | None = None) -> str:
    text_value = (value or "").strip()
    if max_length is not None and len(text_value) > max_length:
        raise HTTPException(status_code=400, detail=f"文本长度不能超过 {max_length} 个字符")
    return text_value


def _product_dto(row: Product, shops: list[PlatformAccount] | None = None) -> ProductDto:
    grouped: dict[str, list[str]] = {}
    for mapping in sorted(row.mappings, key=lambda item: (item.shop_id, item.id)):
        sku = (mapping.shop_sku or "").strip()
        if sku:
            grouped.setdefault(str(mapping.shop_id), []).append(sku)
    mappings = grouped
    if shops:
        mappings = {str(shop.id): mappings.get(str(shop.id), []) for shop in shops}
    return ProductDto(
        id=row.id,
        product_code=row.product_code,
        internal_name=row.internal_name,
        english_name=row.english_name or "",
        cost=float(row.cost) if row.cost is not None else None,
        weight=float(row.weight) if row.weight is not None else None,
        gross_weight=float(row.gross_weight) if row.gross_weight is not None else None,
        package_length=float(row.package_length) if row.package_length is not None else None,
        package_width=float(row.package_width) if row.package_width is not None else None,
        package_height=float(row.package_height) if row.package_height is not None else None,
        ean=row.ean or "",
        description=row.description or "",
        main_image_url=row.main_image_url or "",
        is_slow_moving_material=bool(row.is_slow_moving_material),
        safety_stock=row.safety_stock,
        buyer_user_id=row.buyer_user_id,
        buyer_name=_user_display_name(row.buyer_user),
        enabled=row.enabled,
        mappings=mappings,
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _product_filter_conditions(
    product_code: str | None = None,
    internal_name: str | None = None,
    english_name: str | None = None,
    ean: str | None = None,
    shop_sku: str | None = None,
    enabled: bool | None = None,
    is_slow_moving_material: bool | None = None,
    keyword: str | None = None,
) -> list:
    conditions = []
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        matched_mapping_ids = select(ProductShopMapping.product_id).where(ProductShopMapping.shop_sku.ilike(pattern))
        conditions.append(
            or_(
                Product.product_code.ilike(pattern),
                Product.internal_name.ilike(pattern),
                Product.english_name.ilike(pattern),
                Product.ean.ilike(pattern),
                Product.id.in_(matched_mapping_ids),
            )
        )
    if product_code and product_code.strip():
        conditions.append(Product.product_code.ilike(f"%{product_code.strip()}%"))
    if internal_name and internal_name.strip():
        conditions.append(Product.internal_name.ilike(f"%{internal_name.strip()}%"))
    if english_name and english_name.strip():
        conditions.append(Product.english_name.ilike(f"%{english_name.strip()}%"))
    if ean and ean.strip():
        conditions.append(Product.ean.ilike(f"%{ean.strip()}%"))
    if enabled is not None:
        conditions.append(Product.enabled == enabled)
    if is_slow_moving_material is not None:
        conditions.append(Product.is_slow_moving_material == is_slow_moving_material)
    if shop_sku and shop_sku.strip():
        matched_ids = select(ProductShopMapping.product_id).where(ProductShopMapping.shop_sku.ilike(f"%{shop_sku.strip()}%"))
        conditions.append(Product.id.in_(matched_ids))
    return conditions


def _product_list_base_query(
    product_code: str | None = None,
    internal_name: str | None = None,
    english_name: str | None = None,
    ean: str | None = None,
    shop_sku: str | None = None,
    enabled: bool | None = None,
    is_slow_moving_material: bool | None = None,
    keyword: str | None = None,
):
    stmt = select(Product)
    conditions = _product_filter_conditions(
        product_code,
        internal_name,
        english_name,
        ean,
        shop_sku,
        enabled,
        is_slow_moving_material,
        keyword,
    )
    if conditions:
        stmt = stmt.where(*conditions)
    return stmt


def _generate_product_code(db: Session) -> str:
    max_code = db.scalar(select(func.max(Product.product_code)).where(Product.product_code.like("P________")))
    next_number = 1
    if max_code and len(max_code) == 9 and max_code[1:].isdigit():
        next_number = int(max_code[1:]) + 1
    return f"P{next_number:08d}"


def _normalize_product_name(value: str | None) -> str:
    name = (value or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="中文名称不能为空")
    return name


def _product_name_match_key(value: str | None) -> str:
    return _normalize_product_name(value).casefold()


def _normalize_mapping_values(values: list[str] | None) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        sku = str(value or "").strip()
        if not sku or sku.upper() == "#N/A" or sku in seen:
            continue
        normalized.append(sku)
        seen.add(sku)
    return normalized


def _upsert_product_mappings(
    db: Session,
    product: Product,
    mappings: dict[str, list[str]],
    replace_shop_ids: set[int] | None = None,
) -> None:
    active_shops = {shop.id for shop in _enabled_product_shops(db)}
    replace_shops = active_shops if replace_shop_ids is None else active_shops & replace_shop_ids
    target: dict[int, list[str]] = {}
    for shop_id_text, values in (mappings or {}).items():
        try:
            shop_id = int(shop_id_text)
        except (TypeError, ValueError):
            continue
        if shop_id not in replace_shops:
            continue
        target[shop_id] = _normalize_mapping_values(values)

    for shop_id, sku_list in target.items():
        for sku in sku_list:
            conflict = db.scalar(
                select(ProductShopMapping).where(
                    ProductShopMapping.shop_id == shop_id,
                    ProductShopMapping.shop_sku == sku,
                    ProductShopMapping.product_id != product.id,
                )
            )
            if conflict:
                shop = db.get(PlatformAccount, shop_id)
                shop_name = shop.display_name if shop else str(shop_id)
                raise HTTPException(status_code=400, detail=f"{shop_name} 的 SKU 已绑定其他产品: {sku}")

    current: dict[int, dict[str, ProductShopMapping]] = {}
    for mapping in product.mappings:
        current.setdefault(mapping.shop_id, {})[mapping.shop_sku] = mapping

    for shop_id, rows in current.items():
        if shop_id not in replace_shops:
            continue
        desired = set(target.get(shop_id, []))
        for sku, mapping in list(rows.items()):
            if shop_id not in target or sku not in desired:
                db.delete(mapping)

    for shop_id, sku_list in target.items():
        existing_by_sku = current.get(shop_id, {})
        for sku in sku_list:
            existing = existing_by_sku.get(sku)
            if existing:
                existing.updated_at = datetime.utcnow()
            else:
                db.add(ProductShopMapping(product_id=product.id, shop_id=shop_id, shop_sku=sku))


def _product_mapping_groups(product: Product) -> dict[int, list[str]]:
    grouped: dict[int, list[str]] = {}
    for mapping in sorted(product.mappings, key=lambda item: (item.shop_id, item.id)):
        sku = (mapping.shop_sku or "").strip()
        if sku:
            grouped.setdefault(mapping.shop_id, []).append(sku)
    return grouped


def _product_export_rows(product: Product, shops: list[PlatformAccount]) -> list[list[object]]:
    mappings = _product_mapping_groups(product)
    max_rows = max((len(mappings.get(shop.id, [])) for shop in shops), default=0) or 1
    rows: list[list[object]] = []
    for index in range(max_rows):
        rows.append(
            [
                product.product_code,
                product.internal_name,
                product.english_name or "",
                *[(mappings.get(shop.id, [])[index] if index < len(mappings.get(shop.id, [])) else "") for shop in shops],
                product.cost if product.cost is not None else "",
                product.weight if product.weight is not None else "",
                product.gross_weight if product.gross_weight is not None else "",
                product.package_length if product.package_length is not None else "",
                product.package_width if product.package_width is not None else "",
                product.package_height if product.package_height is not None else "",
                product.ean or "",
                product.description or "",
                product.main_image_url or "",
                "是" if product.is_slow_moving_material else "否",
                _user_display_name(product.buyer_user),
                product.safety_stock if product.safety_stock is not None else "",
                "启用" if product.enabled else "禁用",
            ]
        )
    return rows


def _product_export_column_count(shops: list[PlatformAccount], *, include_code: bool = True) -> int:
    base_headers_count = 3 if include_code else 2
    standard_headers_count = 7
    tail_headers_count = 6
    return base_headers_count + len(shops) + standard_headers_count + tail_headers_count


def _style_product_workbook(worksheet, shops: list[PlatformAccount], *, header_rows: int = 2, include_code: bool = True) -> None:
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return

    total_columns = _product_export_column_count(shops, include_code=include_code)
    group_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="EAF4FF")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_side = Side(style="thin", color="D9E2EC")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    required_headers = {"产品中文名"}
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=total_columns):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.border = border
            if cell.row <= header_rows:
                cell.alignment = center_alignment
                cell.font = Font(bold=True, color="1F2937")
                cell.fill = group_fill if cell.row == 1 else header_fill
            else:
                cell.alignment = body_alignment

    for column_index in range(1, total_columns + 1):
        header_cell = worksheet.cell(row=2, column=column_index)
        merged_header_cell = worksheet.cell(row=1, column=column_index)
        if header_cell.value in required_headers or merged_header_cell.value in required_headers:
            worksheet.cell(row=1, column=column_index).fill = required_fill
            if not isinstance(header_cell, MergedCell):
                header_cell.fill = required_fill

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:{get_column_letter(total_columns)}{max(worksheet.max_row, 2)}"
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[2].height = 30
    for row_index in range(3, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 24

    for column_index, column_cells in enumerate(worksheet.iter_cols(min_col=1, max_col=total_columns), start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 3, 12), 45)


def _append_product_export_header(worksheet, shops: list[PlatformAccount], *, include_code: bool = True) -> None:
    shop_headers = [shop.display_name or shop.account_id for shop in shops]
    base_headers = ['产品中文名', '产品英文名']
    if include_code:
        base_headers = ["编码", *base_headers]
    standard_headers = ["成本", "净重", "毛重", "包装长", "包装宽", "包装高", "EAN"]
    tail_headers = ['描述', '图片链接', '是否呆滞料', '采购人', '安全库存数', '状态']
    first_row = [
        *base_headers,
        *(["SKU CODE", *["" for _ in shop_headers[1:]]] if shop_headers else []),
        "标准参数",
        *["" for _ in standard_headers[1:]],
        *tail_headers,
    ]
    second_row = [
        *["" for _ in base_headers],
        *shop_headers,
        *standard_headers,
        *["" for _ in tail_headers],
    ]
    worksheet.append(first_row)
    worksheet.append(second_row)
    try:
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return
    total_columns = len(first_row)
    for row in worksheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=total_columns):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
    for column_index in range(1, len(base_headers) + 1):
        worksheet.merge_cells(start_row=1, start_column=column_index, end_row=2, end_column=column_index)
    cursor = len(base_headers) + 1
    if shop_headers:
        shop_end = cursor + len(shop_headers) - 1
        if shop_end > cursor:
            worksheet.merge_cells(start_row=1, start_column=cursor, end_row=1, end_column=shop_end)
        cursor = shop_end + 1
    standard_end = cursor + len(standard_headers) - 1
    worksheet.merge_cells(start_row=1, start_column=cursor, end_row=1, end_column=standard_end)
    cursor = standard_end + 1
    for column_index in range(cursor, total_columns + 1):
        worksheet.merge_cells(start_row=1, start_column=column_index, end_row=2, end_column=column_index)


def _product_workbook_response(workbook, filename: str) -> StreamingResponse:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_product_import_template_workbook(shops: list[PlatformAccount], db: Session):
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "产品目录"
    _append_product_export_header(worksheet, shops, include_code=False)
    buyers = _enabled_buyer_users(db)
    buyer_index = 2 + len(shops) + 7 + 3
    buyer_column = buyer_index + 1
    worksheet.cell(row=1, column=1).comment = Comment("必填。导入时按产品中文名忽略大小写匹配，存在则更新，不存在则新增。", "CaifuClaw AI Demo")
    worksheet.cell(row=1, column=buyer_column).comment = Comment("可填写采购人的用户名或显示名。", "CaifuClaw AI Demo")

    guide_sheet = workbook.create_sheet("填写说明")
    guide_sheet.append(["说明项", "填写说明"])
    guide_sheet.append(["产品中文名", "必填。导入时按产品中文名忽略大小写匹配，存在则更新，不存在则新增。"])
    guide_sheet.append(["店铺 SKU", "店铺列来自当前启用店铺；同一产品同一店铺可在多行填写多个 SKU。"])
    guide_sheet.append(["是否呆滞料", "可填写：是/否、true/false、yes/no、1/0。"])
    guide_sheet.append(["采购人", "可填写已启用采购人的用户名或显示名。"])
    _style_lookup_sheet(guide_sheet)


    if buyers:
        buyer_sheet = workbook.create_sheet("可选采购人")
        buyer_sheet.append(["采购人", "用户名"])
        for buyer in buyers:
            buyer_sheet.append([_user_display_name(buyer), buyer.username])
        _style_lookup_sheet(buyer_sheet)
        buyer_validation = DataValidation(
            type="list",
            formula1="'可选采购人'!$A$2:$A$1048576",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="采购人无效",
            error="请选择可选采购人中的名称，或填写已启用采购人的用户名。",
        )
        worksheet.add_data_validation(buyer_validation)
        buyer_validation.add(f"{worksheet.cell(row=1, column=buyer_column).column_letter}3:{worksheet.cell(row=1, column=buyer_column).column_letter}1048576")

    _style_product_workbook(worksheet, shops, include_code=False)
    return workbook


def _style_lookup_sheet(worksheet) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return
    header_fill = PatternFill("solid", fgColor="EAF4FF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)


def _build_inventory_import_template_workbook():
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "库存"
    worksheet.append(["产品名称", "库存数量", "上次盘点", "安全库存", "备注"])

    required_fill = PatternFill("solid", fgColor="FFF2CC")
    optional_fill = PatternFill("solid", fgColor="EAF4FF")
    for column_index, cell in enumerate(worksheet[1], start=1):
        cell.fill = required_fill if column_index in {1, 4} else optional_fill
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].comment = Comment("必填。请填写产品管理中已有的产品名称，名称需完全一致。", "CaifuClaw AI Demo")
    worksheet["B1"].comment = Comment("选填。未填写时新记录按 0 导入，已有库存保留原值。", "CaifuClaw AI Demo")
    worksheet["C1"].comment = Comment("选填。未填写时按 0 导入。", "CaifuClaw AI Demo")
    worksheet["D1"].comment = Comment("必填。请填写大于或等于 0 的整数。", "CaifuClaw AI Demo")
    worksheet["E1"].comment = Comment("选填，最多 500 个字符。", "CaifuClaw AI Demo")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:E1"
    worksheet.row_dimensions[1].height = 30
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 42

    optional_quantity_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="库存数量无效",
        error="请填写大于或等于 0 的整数。",
    )
    worksheet.add_data_validation(optional_quantity_validation)
    optional_quantity_validation.add("B2:C1048576")
    safety_stock_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="安全库存无效",
        error="安全库存为必填项，请填写大于或等于 0 的整数。",
    )
    worksheet.add_data_validation(safety_stock_validation)
    safety_stock_validation.add("D2:D1048576")

    guide_sheet = workbook.create_sheet("填写说明")
    guide_sheet.append(["字段", "是否必填", "填写说明"])
    guide_sheet.append(["产品名称", "是", "填写产品管理中已有的产品名称，名称需完全一致。"])
    guide_sheet.append(["库存数量", "否", "填写大于或等于 0 的整数；未填写时新记录按 0 导入，已有库存保留原值。"])
    guide_sheet.append(["上次盘点", "否", "填写大于或等于 0 的整数；未填写时按 0 导入。"])
    guide_sheet.append(["安全库存", "是", "填写大于或等于 0 的整数。"])
    guide_sheet.append(["备注", "否", "最多 500 个字符。"])
    _style_lookup_sheet(guide_sheet)
    return workbook


def _create_product_from_payload(db: Session, payload: ProductUpsertRequest) -> Product:
    buyer_user = _buyer_user_by_payload(db, payload.buyer_user_id)
    product = Product(
        product_code=_generate_product_code(db),
        internal_name=_normalize_product_name(payload.internal_name),
        english_name=_normalize_optional_text(payload.english_name, 255),
        cost=_decimal_or_none(payload.cost),
        weight=_decimal_or_none(payload.weight),
        gross_weight=_decimal_or_none(payload.gross_weight),
        package_length=_decimal_or_none(payload.package_length),
        package_width=_decimal_or_none(payload.package_width),
        package_height=_decimal_or_none(payload.package_height),
        ean=_normalize_optional_text(payload.ean, 64),
        description=_normalize_optional_text(payload.description),
        main_image_url=_normalize_optional_text(payload.main_image_url),
        is_slow_moving_material=bool(payload.is_slow_moving_material),
        safety_stock=_int_or_none(payload.safety_stock),
        buyer_user_id=buyer_user.id if buyer_user else None,
        enabled=payload.enabled,
    )
    db.add(product)
    db.flush()
    _upsert_product_mappings(db, product, payload.mappings)
    return product


def _payment_range_bounds(
    payment_time_range: str | None = None,
    payment_start: str | None = None,
    payment_end: str | None = None,
) -> tuple[datetime, datetime] | None:
    # Custom date range takes priority over preset
    if payment_start and payment_end:
        try:
            start_day = datetime.fromisoformat(payment_start).date()
            end_day = datetime.fromisoformat(payment_end).date() + timedelta(days=1)
            return _platform_date_bounds(start_day, end_day)
        except (ValueError, TypeError):
            return None
    if not payment_time_range:
        return None
    today = _local_today()
    if payment_time_range == "today":
        start = today
        end = today + timedelta(days=1)
    elif payment_time_range == "yesterday":
        start = today - timedelta(days=1)
        end = today
    elif payment_time_range == "before_yesterday":
        start = today - timedelta(days=2)
        end = today - timedelta(days=1)
    elif payment_time_range == "last_3_days_exclude_today":
        start = today - timedelta(days=3)
        end = today
    elif payment_time_range == "last_7_days_exclude_today":
        start = today - timedelta(days=7)
        end = today
    elif payment_time_range == "last_30_days_exclude_today":
        start = today - timedelta(days=30)
        end = today
    else:
        return None
    return _platform_date_bounds(start, end)


MAX_BATCH_ORDER_NUMBERS = 100
MAX_BATCH_ORDER_NUMBER_LENGTH = 160
MAX_BATCH_ORDER_NUMBER_INPUTS = 500


def _normalize_batch_order_numbers(numbers: list[str]) -> tuple[int, list[str]]:
    if len(numbers) > MAX_BATCH_ORDER_NUMBER_INPUTS:
        raise HTTPException(status_code=400, detail=f"单次提交的单号不能超过 {MAX_BATCH_ORDER_NUMBER_INPUTS} 个")

    submitted_numbers = [str(value).strip() for value in numbers if str(value).strip()]
    unique_numbers = list(dict.fromkeys(submitted_numbers))
    if not unique_numbers:
        raise HTTPException(status_code=400, detail="请至少输入一个单号")
    if len(unique_numbers) > MAX_BATCH_ORDER_NUMBERS:
        raise HTTPException(status_code=400, detail=f"批量查询单号不能超过 {MAX_BATCH_ORDER_NUMBERS} 个")

    too_long = [value for value in unique_numbers if len(value) > MAX_BATCH_ORDER_NUMBER_LENGTH]
    if too_long:
        raise HTTPException(
            status_code=400,
            detail=f"单号不能超过 {MAX_BATCH_ORDER_NUMBER_LENGTH} 个字符：{too_long[0][:40]}",
        )
    return len(submitted_numbers), unique_numbers


def _order_number_search_condition(term: str):
    pattern = f"%{term.strip()}%"
    return or_(
        Order.platform_order_no.ilike(pattern),
        Order.platform_order_id.ilike(pattern),
        Order.posting_number.ilike(pattern),
        Order.shipment_tracking_number.ilike(pattern),
        Order.raw_payload["shipment_tracking_number"].astext.ilike(pattern),
        Order.raw_payload["tracking_number"].astext.ilike(pattern),
        Order.raw_payload["trackingNumber"].astext.ilike(pattern),
        Order.raw_payload["shipment"]["tracking_number"].astext.ilike(pattern),
        Order.raw_payload["shipment"]["trackingNumber"].astext.ilike(pattern),
        Order.raw_payload["shipping"]["tracking_number"].astext.ilike(pattern),
        Order.raw_payload["shipping"]["trackingNumber"].astext.ilike(pattern),
        Order.raw_payload["logistics"]["tracking_number"].astext.ilike(pattern),
        Order.raw_payload["logistics"]["trackingNumber"].astext.ilike(pattern),
        Order.raw_payload["tracking"]["number"].astext.ilike(pattern),
        exists().where(Shipment.order_id == Order.id, Shipment.tracking_number.ilike(pattern)),
    )


def _order_number_exact_condition(numbers: list[str]):
    values = tuple(numbers)
    return or_(
        Order.platform_order_no.in_(values),
        Order.platform_order_id.in_(values),
        Order.posting_number.in_(values),
        Order.shipment_tracking_number.in_(values),
        Order.raw_payload["shipment_tracking_number"].astext.in_(values),
        Order.raw_payload["tracking_number"].astext.in_(values),
        Order.raw_payload["trackingNumber"].astext.in_(values),
        Order.raw_payload["shipment"]["tracking_number"].astext.in_(values),
        Order.raw_payload["shipment"]["trackingNumber"].astext.in_(values),
        Order.raw_payload["shipping"]["tracking_number"].astext.in_(values),
        Order.raw_payload["shipping"]["trackingNumber"].astext.in_(values),
        Order.raw_payload["logistics"]["tracking_number"].astext.in_(values),
        Order.raw_payload["logistics"]["trackingNumber"].astext.in_(values),
        Order.raw_payload["tracking"]["number"].astext.in_(values),
        exists().where(Shipment.order_id == Order.id, Shipment.tracking_number.in_(values)),
    )


def _raw_payload_order_number_values(raw_payload: dict | None) -> list[str]:
    payload = raw_payload if isinstance(raw_payload, dict) else {}
    shipment = payload.get("shipment") if isinstance(payload.get("shipment"), dict) else {}
    shipping = payload.get("shipping") if isinstance(payload.get("shipping"), dict) else {}
    logistics = payload.get("logistics") if isinstance(payload.get("logistics"), dict) else {}
    tracking = payload.get("tracking") if isinstance(payload.get("tracking"), dict) else {}
    return [
        str(value)
        for value in (
            payload.get("shipment_tracking_number"),
            payload.get("tracking_number"),
            payload.get("trackingNumber"),
            shipment.get("tracking_number"),
            shipment.get("trackingNumber"),
            shipping.get("tracking_number"),
            shipping.get("trackingNumber"),
            logistics.get("tracking_number"),
            logistics.get("trackingNumber"),
            tracking.get("number"),
        )
        if value is not None
    ]


def _batch_order_search_summary(
    db: Session,
    submitted: int,
    numbers: list[str],
    conditions: list,
) -> OrderSearchSummary:
    rows = db.execute(
        select(
            Order.id,
            Order.platform_order_no,
            Order.platform_order_id,
            Order.posting_number,
            Order.shipment_tracking_number,
            Order.raw_payload,
        ).where(*conditions)
    ).all()
    matched_values: set[str] = set()
    order_ids: list[int] = []
    for row in rows:
        order_ids.append(row.id)
        matched_values.update(
            str(value)
            for value in (
                row.platform_order_no,
                row.platform_order_id,
                row.posting_number,
                row.shipment_tracking_number,
            )
            if value is not None
        )
        matched_values.update(_raw_payload_order_number_values(row.raw_payload))

    if order_ids:
        matched_values.update(
            str(value)
            for value in db.scalars(
                select(Shipment.tracking_number).where(
                    Shipment.order_id.in_(order_ids),
                    Shipment.tracking_number.in_(numbers),
                )
            ).all()
            if value is not None
        )

    unmatched_numbers = [value for value in numbers if value not in matched_values]
    return OrderSearchSummary(
        submitted=submitted,
        unique=len(numbers),
        matched=len(numbers) - len(unmatched_numbers),
        unmatched_numbers=unmatched_numbers,
    )


def _order_product_keyword_match_item_ids(term: str):
    pattern = f"%{term.strip()}%"
    mapping_choice = mapping_choice_for_order_item()
    product_name = func.coalesce(mapping_choice["exact_product"].internal_name, mapping_choice["insensitive_product"].internal_name)
    mapped_item_ids = (
        select(OrderItem.id)
        .select_from(OrderItem)
        .join(Order, Order.id == OrderItem.order_id)
        .join(
            PlatformAccount,
            (PlatformAccount.platform == Order.platform) & (PlatformAccount.account_id == Order.shop_id),
        )
        .outerjoin(mapping_choice["exact_mapping"], mapping_choice["exact_condition"])
        .outerjoin(mapping_choice["exact_product"], mapping_choice["exact_product"].id == mapping_choice["exact_mapping"].product_id)
        .outerjoin(mapping_choice["insensitive_mapping"], mapping_choice["insensitive_condition"])
        .outerjoin(
            mapping_choice["insensitive_product"],
            mapping_choice["insensitive_product"].id == mapping_choice["insensitive_mapping"].product_id,
        )
        .where(product_name.ilike(pattern))
    )
    return union_all(
        select(OrderItem.id).where(OrderItem.sku.ilike(pattern)),
        select(OrderItem.id).where(OrderItem.platform_product_name.ilike(pattern)),
        mapped_item_ids,
    ).subquery()


def _order_product_keyword_match_order_ids(term: str):
    pattern = f"%{term.strip()}%"
    mapping_choice = mapping_choice_for_order_item()
    product_name = func.coalesce(mapping_choice["exact_product"].internal_name, mapping_choice["insensitive_product"].internal_name)
    mapped_order_ids = (
        select(OrderItem.order_id)
        .select_from(OrderItem)
        .join(Order, Order.id == OrderItem.order_id)
        .join(
            PlatformAccount,
            (PlatformAccount.platform == Order.platform) & (PlatformAccount.account_id == Order.shop_id),
        )
        .outerjoin(mapping_choice["exact_mapping"], mapping_choice["exact_condition"])
        .outerjoin(mapping_choice["exact_product"], mapping_choice["exact_product"].id == mapping_choice["exact_mapping"].product_id)
        .outerjoin(mapping_choice["insensitive_mapping"], mapping_choice["insensitive_condition"])
        .outerjoin(
            mapping_choice["insensitive_product"],
            mapping_choice["insensitive_product"].id == mapping_choice["insensitive_mapping"].product_id,
        )
        .where(product_name.ilike(pattern))
    )
    return union_all(
        select(OrderItem.order_id).where(OrderItem.sku.ilike(pattern)),
        select(OrderItem.order_id).where(OrderItem.platform_product_name.ilike(pattern)),
        mapped_order_ids,
    ).subquery()


def _order_product_keyword_condition(term: str):
    order_ids = _order_product_keyword_match_order_ids(term)
    return Order.id.in_(select(order_ids.c.order_id))


def _order_filter_conditions(
    status_filter: str | None,
    platform: str | None,
    transaction_id: str | None,
    order_no: str | None = None,
    payment_time_range: str | None = None,
    payment_start: str | None = None,
    payment_end: str | None = None,
    number: str | None = None,
    product_keyword: str | None = None,
    numbers: list[str] | None = None,
    risk_filter: str | None = None,
    shop: str | None = None,
    risk_shop_keys: list[tuple[str, str]] | None = None,
    shop_keys: list[tuple[str, str]] | None = None,
) -> list:
    conditions = []
    normalized_status_filter = _normalize_order_status_filter(status_filter)
    if normalized_status_filter:
        voided_platform_status = _voided_platform_status_condition()
        if normalized_status_filter == ORDER_STATUS_WAITING_PURCHASE:
            conditions.append(and_(_waiting_purchase_condition(), ~voided_platform_status))
        elif normalized_status_filter == ORDER_STATUS_AWAITING_PICKUP:
            conditions.append(and_(Order.biz_status.in_(ORDER_STATUS_AWAITING_PICKUP_LABELS), ~voided_platform_status))
        elif normalized_status_filter == ORDER_STATUS_VOIDED:
            conditions.append(or_(Order.biz_status == ORDER_STATUS_VOIDED, voided_platform_status))
        else:
            conditions.append(and_(Order.biz_status == normalized_status_filter, ~voided_platform_status))
    if platform:
        conditions.append(Order.platform == platform)
    number_terms = [value.strip() for value in (number, transaction_id, order_no) if value and value.strip()]
    for term in dict.fromkeys(number_terms):
        conditions.append(_order_number_search_condition(term))
    if numbers:
        conditions.append(_order_number_exact_condition(numbers))
    if product_keyword and product_keyword.strip():
        conditions.append(_order_product_keyword_condition(product_keyword.strip()))
    bounds = _payment_range_bounds(payment_time_range, payment_start, payment_end)
    if bounds:
        start, end = bounds
        conditions.extend([Order.payment_at >= start, Order.payment_at < end])
    conditions.extend(_risk_filter_conditions(risk_filter, shop))
    if risk_filter:
        conditions.extend(_risk_shop_scope_conditions(risk_shop_keys))
    conditions.extend(_risk_shop_scope_conditions(shop_keys))
    return conditions


def _build_orders_query(
    status_filter: str | None,
    platform: str | None,
    transaction_id: str | None,
    order_no: str | None = None,
    payment_time_range: str | None = None,
    payment_start: str | None = None,
    payment_end: str | None = None,
    number: str | None = None,
    product_keyword: str | None = None,
    numbers: list[str] | None = None,
    risk_filter: str | None = None,
    shop: str | None = None,
    risk_shop_keys: list[tuple[str, str]] | None = None,
    shop_keys: list[tuple[str, str]] | None = None,
):
    conditions = _order_filter_conditions(
        status_filter,
        platform,
        transaction_id,
        order_no,
        payment_time_range,
        payment_start,
        payment_end,
        number,
        product_keyword,
        numbers,
        risk_filter,
        shop,
        risk_shop_keys,
        shop_keys,
    )
    stmt = select(Order)
    if conditions:
        stmt = stmt.where(*conditions)
    normalized_risk_filter = _normalize_order_risk_filter(risk_filter)
    if normalized_risk_filter:
        deadline = _risk_deadline_expression()
        now = datetime.utcnow()
        stmt = stmt.order_by(
            case((deadline < now, 0), else_=1),
            deadline.asc().nulls_last(),
            desc(Order.payment_at).nulls_last(),
            desc(Order.created_at),
            desc(Order.updated_at),
        )
    else:
        stmt = stmt.order_by(desc(Order.payment_at).nulls_last(), desc(Order.created_at), desc(Order.updated_at))
    # 懒加载大 JSONB 列，避免列表查询传输 last_api_payload。
    stmt = stmt.options(defer(Order.last_api_payload))
    return stmt


def _query_orders(
    db: Session,
    status_filter: str | None,
    platform: str | None,
    transaction_id: str | None,
    order_no: str | None,
    number: str | None,
    payment_time_range: str | None,
    payment_start: str | None = None,
    payment_end: str | None = None,
    page: int = 1,
    page_size: int = 50,
    product_keyword: str | None = None,
    numbers: list[str] | None = None,
    submitted_number_count: int | None = None,
    risk_filter: str | None = None,
    shop: str | None = None,
    risk_shop_keys: list[tuple[str, str]] | None = None,
    shop_keys: list[tuple[str, str]] | None = None,
) -> OrderListResponse:
    normalized_risk_filter = _normalize_order_risk_filter(risk_filter)
    stmt = _build_orders_query(
        status_filter,
        platform,
        transaction_id,
        order_no,
        payment_time_range,
        payment_start,
        payment_end,
        number=number,
        product_keyword=product_keyword,
        numbers=numbers,
        risk_filter=normalized_risk_filter,
        shop=shop,
        risk_shop_keys=risk_shop_keys,
        shop_keys=shop_keys,
    )
    conditions = _order_filter_conditions(
        status_filter,
        platform,
        transaction_id,
        order_no,
        payment_time_range,
        payment_start,
        payment_end,
        number,
        product_keyword,
        numbers,
        normalized_risk_filter,
        shop,
        risk_shop_keys,
        shop_keys,
    )
    count_stmt = select(func.count(Order.id))
    if conditions:
        count_stmt = count_stmt.where(*conditions)
    total = db.scalar(count_stmt) or 0

    search_summary = None
    if numbers:
        search_summary = _batch_order_search_summary(
            db,
            submitted=submitted_number_count if submitted_number_count is not None else len(numbers),
            numbers=numbers,
            conditions=conditions,
        )

    rows = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()

    # 批量预加载 shipment 和 label 数据，避免 N+1 查询。
    order_ids = [row.id for row in rows]
    shipments_map = {}
    labels_map = {}
    outbound_scan_map = {}
    risk_handling_map = {}

    if order_ids:
        # 批量查询所有 shipment。
        shipments = db.scalars(
            select(Shipment).where(Shipment.order_id.in_(order_ids)).order_by(Shipment.id.desc())
        ).all()

        # 每个订单只取最新的 shipment。
        for shipment in shipments:
            if shipment.order_id not in shipments_map:
                shipments_map[shipment.order_id] = shipment

        # 批量查询所有 label。
        shipment_ids = [s.id for s in shipments_map.values()]
        if shipment_ids:
            labels = db.scalars(
                select(LabelFile).where(LabelFile.shipment_id.in_(shipment_ids)).order_by(LabelFile.id.desc())
            ).all()

            # 每个 shipment 只取最新的 label。
            for label in labels:
                if label.shipment_id not in labels_map:
                    labels_map[label.shipment_id] = label

        outbound_scans = db.execute(
            select(
                OutboundScanRecord.order_id,
                func.min(OutboundScanRecord.scanned_at),
            )
            .where(
                OutboundScanRecord.order_id.in_(order_ids),
                OutboundScanRecord.result == "success",
            )
            .group_by(OutboundScanRecord.order_id)
        ).all()
        outbound_scan_map = {order_id: scanned_at for order_id, scanned_at in outbound_scans}

        if normalized_risk_filter:
            risk_handling_rows = db.scalars(
                select(OrderRiskHandling).where(OrderRiskHandling.order_id.in_(order_ids))
            ).all()
            risk_handling_map = {row.order_id: row for row in risk_handling_rows}

    # 构建 DTO 列表。
    items = []
    for row in rows:
        shipment = shipments_map.get(row.id)
        label = labels_map.get(shipment.id) if shipment else None
        items.append(
            _order_dto(
                row,
                shipment,
                label,
                outbound_scanned_at=outbound_scan_map.get(row.id),
                risk_view=bool(normalized_risk_filter),
                risk_handling=risk_handling_map.get(row.id),
            )
        )

    return OrderListResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        search_summary=search_summary,
    )


def _dispatch_deadline(
    row: Order,
    item: OrderItem,
    shipping_deadline_at: datetime | None = None,
    settings: dict[str, ShippingDeadlineSetting] | None = None,
) -> datetime | None:
    if not item.sku:
        return None
    if row.dispatch_deadline_at:
        return row.dispatch_deadline_at
    if shipping_deadline_at is not None:
        row.shipping_deadline_at = shipping_deadline_at
    return calculate_dispatch_deadline(row, settings or {})


def _summary_warning(
    purchase_generated: bool,
    product_name: str,
    shipping_time: datetime | None,
    dispatch_deadline: datetime | None,
) -> str:
    if purchase_generated or not product_name or not dispatch_deadline:
        return ""
    if shipping_time:
        return "Delivered"
    today = datetime.combine(_local_today(), time.min)
    if dispatch_deadline.tzinfo is not None:
        dispatch_deadline = dispatch_deadline.astimezone(timezone.utc).replace(tzinfo=None)
    dispatch_deadline_local = dispatch_deadline.replace(tzinfo=None) + LOCAL_TIME_OFFSET
    deadline_day = datetime.combine(dispatch_deadline_local.date(), time.min)
    if today > deadline_day:
        return "Delayed"
    if today > deadline_day - timedelta(days=2):
        return "Urgent"
    return _format_remaining_delta(dispatch_deadline)


def _normalize_summary_warning_filter(value: str | None) -> str:
    normalized = (value or "").strip()
    if not normalized or normalized.lower() == "all":
        return ""
    aliases = {
        "urgent": "Urgent",
        "delayed": "Delayed",
        "delivered": "Delivered",
    }
    return aliases.get(normalized.lower(), normalized)


def _summary_mapping_timestamp(updated_at: datetime | None, created_at: datetime | None) -> datetime:
    return updated_at or created_at or datetime.min


def _normalize_summary_sku(value: str | None) -> str:
    return (value or "").strip().lower()


def _order_summary_product_lookup(
    db: Session,
    order_item_rows: list[tuple[Order, OrderItem]],
) -> dict[int, str]:
    shop_keys = {
        (row.platform, row.shop_id)
        for row, item in order_item_rows
        if row.platform and row.shop_id and (item.sku or "").strip()
    }
    if not shop_keys:
        return {}

    account_filter = or_(
        *[
            and_(PlatformAccount.platform == platform, PlatformAccount.account_id == shop_id)
            for platform, shop_id in sorted(shop_keys)
        ]
    )
    account_rows = db.execute(
        select(PlatformAccount.platform, PlatformAccount.account_id, PlatformAccount.id).where(account_filter)
    ).all()
    account_ids = {(platform, account_id): shop_db_id for platform, account_id, shop_db_id in account_rows}

    item_keys: list[tuple[int, int, str, str]] = []
    for row, item in order_item_rows:
        sku = (item.sku or "").strip()
        shop_db_id = account_ids.get((row.platform, row.shop_id))
        if not sku or not shop_db_id:
            continue
        item_keys.append((item.id, shop_db_id, sku, _normalize_summary_sku(sku)))
    if not item_keys:
        return {}

    shop_db_ids = sorted({shop_db_id for _item_id, shop_db_id, _sku, _normalized_sku in item_keys})
    exact_skus = sorted({sku for _item_id, _shop_db_id, sku, _normalized_sku in item_keys})
    normalized_skus = sorted({normalized_sku for _item_id, _shop_db_id, _sku, normalized_sku in item_keys})
    normalized_mapping_sku = func.lower(func.trim(func.coalesce(ProductShopMapping.shop_sku, "")))
    mapping_rows = db.execute(
        select(
            ProductShopMapping.id,
            ProductShopMapping.shop_id,
            ProductShopMapping.shop_sku,
            ProductShopMapping.created_at,
            ProductShopMapping.updated_at,
            Product.id.label("product_id"),
            Product.internal_name,
        )
        .join(Product, Product.id == ProductShopMapping.product_id)
        .where(
            ProductShopMapping.shop_id.in_(shop_db_ids),
            or_(ProductShopMapping.shop_sku.in_(exact_skus), normalized_mapping_sku.in_(normalized_skus)),
        )
    ).all()

    exact_candidates: dict[tuple[int, str], tuple[datetime, int, int | None, str]] = {}
    insensitive_candidates: dict[tuple[int, str], tuple[datetime, int, int | None, str]] = {}
    for mapping_id, shop_db_id, shop_sku, created_at, updated_at, product_id, product_name in mapping_rows:
        rank = (_summary_mapping_timestamp(updated_at, created_at), int(mapping_id or 0), product_id, product_name or "")
        exact_key = (int(shop_db_id), shop_sku or "")
        if exact_key not in exact_candidates or rank[:2] > exact_candidates[exact_key][:2]:
            exact_candidates[exact_key] = rank
        insensitive_key = (int(shop_db_id), _normalize_summary_sku(shop_sku))
        if insensitive_key not in insensitive_candidates or rank[:2] > insensitive_candidates[insensitive_key][:2]:
            insensitive_candidates[insensitive_key] = rank

    result: dict[int, str] = {}
    for item_id, shop_db_id, sku, normalized_sku in item_keys:
        candidate = exact_candidates.get((shop_db_id, sku)) or insensitive_candidates.get((shop_db_id, normalized_sku))
        if candidate:
            result[item_id] = candidate[3]
    return result


ORDER_SUMMARY_INCLUDED_STATUSES = {
    ORDER_STATUS_WAITING_PURCHASE,
    ORDER_STATUS_PICKING,
    ORDER_STATUS_SHIPPED,
    ORDER_STATUS_DELIVERED,
    *ORDER_STATUS_AWAITING_PICKUP_LABELS,
}
ORDER_SUMMARY_EXCLUDED_STATUSES = {
    ORDER_STATUS_PENDING,
}


def _query_order_summary(
    db: Session,
    status_filter: str | None,
    platform: str | None,
    transaction_id: str | None,
    tracking_number: str | None,
    number: str | None,
    payment_time_range: str | None,
    payment_start: str | None = None,
    payment_end: str | None = None,
    picking_start: str | None = None,
    picking_end: str | None = None,
    old_customer_only: bool = False,
    page: int = 1,
    page_size: int = 50,
    warning_filter: str | None = None,
    product_keyword: str | None = None,
    item_ids: list[int] | None = None,
    paginate: bool = True,
    lazy: bool = False,
    shop_keys: list[tuple[str, str]] | None = None,
) -> OrderSummaryResponse:
    normalized_warning_filter = _normalize_summary_warning_filter(warning_filter)
    normalized_product_keyword = (product_keyword or "").strip()
    order_conditions = _order_filter_conditions(
        status_filter,
        platform,
        transaction_id,
        None,
        payment_time_range,
        payment_start,
        payment_end,
        number,
        shop_keys=shop_keys,
    )
    summary_order_by = (
        desc(Order.payment_at).nulls_last(),
        desc(Order.created_at),
        desc(Order.updated_at),
        Order.id,
        OrderItem.id,
    )

    item_scope_stmt = select(OrderItem.id.label("item_id")).join(Order, Order.id == OrderItem.order_id)
    if order_conditions:
        item_scope_stmt = item_scope_stmt.where(*order_conditions)
    item_scope_stmt = item_scope_stmt.where(
        or_(
            Order.picking_at.is_not(None),
            Order.biz_status.in_(ORDER_SUMMARY_INCLUDED_STATUSES),
        ),
        Order.biz_status.notin_(ORDER_SUMMARY_EXCLUDED_STATUSES),
    )
    if normalized_product_keyword:
        matched_item_ids = _order_product_keyword_match_item_ids(normalized_product_keyword)
        item_scope_stmt = item_scope_stmt.where(OrderItem.id.in_(select(matched_item_ids.c.id)))

    picking_start_dt = None
    picking_end_dt = None
    if picking_start:
        picking_start_dt, _ = _local_day_utc_bounds(picking_start)
    if picking_end:
        _, picking_end_dt = _local_day_utc_bounds(picking_end)
    if picking_start_dt:
        item_scope_stmt = item_scope_stmt.where(Order.picking_at >= picking_start_dt)
    if picking_end_dt:
        item_scope_stmt = item_scope_stmt.where(Order.picking_at < picking_end_dt)
    if tracking_number and tracking_number.strip():
        item_scope_stmt = item_scope_stmt.where(_order_number_search_condition(tracking_number.strip()))
    if item_ids:
        item_scope_stmt = item_scope_stmt.where(OrderItem.id.in_(item_ids))
    if old_customer_only:
        prior_order_filter = aliased(Order)
        item_scope_stmt = item_scope_stmt.where(
            Order.buyer_id.is_not(None),
            Order.buyer_id != "",
            Order.platform_created_at.is_not(None),
            exists().where(
                prior_order_filter.shop_id == Order.shop_id,
                prior_order_filter.buyer_id == Order.buyer_id,
                prior_order_filter.id != Order.id,
                prior_order_filter.platform_created_at < Order.platform_created_at,
            ),
        )
    item_scope_stmt = item_scope_stmt.order_by(*summary_order_by)

    page_item_ids: list[int] | None = None
    lazy_has_more = False
    if normalized_warning_filter or not paginate:
        total = 0
    elif lazy:
        page_item_ids = list(db.scalars(item_scope_stmt.offset((page - 1) * page_size).limit(page_size + 1)).all())
        lazy_has_more = len(page_item_ids) > page_size
        if lazy_has_more:
            page_item_ids = page_item_ids[:page_size]
        total = ((page - 1) * page_size) + len(page_item_ids) + (1 if lazy_has_more else 0)
        if not page_item_ids:
            return OrderSummaryResponse(items=[], total=(page - 1) * page_size, page=page, page_size=page_size, has_more=False)
    else:
        count_scope = item_scope_stmt.order_by(None).subquery()
        total = db.scalar(select(func.count()).select_from(count_scope)) or 0
        page_item_ids = list(db.scalars(item_scope_stmt.offset((page - 1) * page_size).limit(page_size)).all())
        if not page_item_ids:
            return OrderSummaryResponse(items=[], total=total, page=page, page_size=page_size, has_more=False)

    shipping_time = (
        select(func.min(OutboundScanRecord.scanned_at))
        .where(
            OutboundScanRecord.order_id == Order.id,
            OutboundScanRecord.result == "success",
        )
        .scalar_subquery()
    )
    latest_tracking_number = (
        select(Shipment.tracking_number)
        .where(Shipment.order_id == Order.id)
        .order_by(Shipment.id.desc())
        .limit(1)
        .scalar_subquery()
    )
    latest_label_path = (
        select(LabelFile.file_path)
        .join(Shipment, Shipment.id == LabelFile.shipment_id)
        .where(Shipment.order_id == Order.id)
        .order_by(LabelFile.id.desc())
        .limit(1)
        .scalar_subquery()
    )
    prior_order = aliased(Order)
    has_prior_order = exists().where(
        prior_order.shop_id == Order.shop_id,
        prior_order.buyer_id == Order.buyer_id,
        prior_order.id != Order.id,
        prior_order.platform_created_at < Order.platform_created_at,
    )
    stmt = (
        select(
            Order,
            OrderItem,
            PurchaseOrder.purchase_no,
            shipping_time.label("shipping_time"),
            latest_tracking_number.label("latest_tracking_number"),
            latest_label_path.label("latest_label_path"),
            has_prior_order.label("has_prior_order"),
        )
        .select_from(OrderItem)
        .join(Order, Order.id == OrderItem.order_id)
        .outerjoin(PurchaseOrderSource, PurchaseOrderSource.order_item_id == OrderItem.id)
        .outerjoin(PurchaseOrder, PurchaseOrder.id == PurchaseOrderSource.purchase_order_id)
        .order_by(*summary_order_by)
        .options(defer(Order.last_api_payload))
    )
    if normalized_warning_filter or not paginate:
        item_scope = item_scope_stmt.order_by(None).subquery()
        stmt = stmt.join(item_scope, item_scope.c.item_id == OrderItem.id)
    else:
        stmt = stmt.where(OrderItem.id.in_(page_item_ids))
    rows = db.execute(stmt).all()
    product_lookup = _order_summary_product_lookup(db, [(row, item) for row, item, *_rest in rows])
    deadline_settings = load_shipping_deadline_settings(db)

    items: list[OrderSummaryDto] = []
    for row, item, purchase_no, shipping_time, latest_tracking, latest_label, has_prior_order in rows:
        product_name = product_lookup.get(item.id, "")
        product_text = product_name or ""
        platform_product_name = item.platform_product_name or _item_platform_product_name(item.raw_payload)
        extracted = _extract_order_fields(row.raw_payload or {})
        tracking_number = (
            clean_tracking_number(row.shipment_tracking_number, row.raw_payload or {}, row.platform)
            or extracted["shipment_tracking_number"]
            or clean_tracking_number(latest_tracking, row.raw_payload or {}, row.platform)
            or ""
        )
        if order_is_logistics_label_exempt(row):
            tracking_number = ""
        shipping_deadline_at = _effective_shipping_deadline(row, extracted)
        dispatch_deadline = _dispatch_deadline(row, item, shipping_deadline_at, deadline_settings)
        warning = _summary_warning(bool(purchase_no), product_text, shipping_time, dispatch_deadline)
        if normalized_warning_filter and warning != normalized_warning_filter:
            continue
        items.append(
            OrderSummaryDto(
                order_id=row.id,
                item_id=item.id,
                picking_at=_iso(row.picking_at),
                platform=row.platform,
                shop_name=row.shop_name or "",
                platform_created_at=_platform_import_time(row.raw_payload or {}, "platform_created_at", row.platform_created_at),
                order_no=row.posting_number or row.platform_order_no or "",
                status=_derive_order_status(row),
                platform_status="" if row.platform_status in (None, "None") else row.platform_status,
                platform_order_no=row.platform_order_no or "",
                platform_order_id=row.platform_order_id or "",
                posting_number=row.posting_number or "",
                country_code=_display_country_code(row, extracted),
                country_name_cn=_display_country_name_cn(row, extracted),
                customer_name=_display_customer_name(row, extracted, latest_label),
                sku=item.sku or "",
                platform_product_name=platform_product_name,
                quantity=item.quantity or 1,
                unit_price=item.unit_price or "",
                currency=item.currency or row.currency or "",
                buyer_selected_logistics=row.buyer_selected_logistics or "",
                shipping_deadline_at=_platform_import_time(
                    row.raw_payload or {},
                    "shipping_deadline_at",
                    shipping_deadline_at,
                ),
                shipment_tracking_number=tracking_number,
                tracking_number=tracking_number,
                dispatch_deadline_at=_platform_time_iso(dispatch_deadline),
                product_name=product_text,
                customer_confirm="老客户" if row.buyer_id and has_prior_order else ("新客户" if row.buyer_id else ""),
                warning=warning,
                shipping_time=_iso(shipping_time),
                purchase_generated=bool(purchase_no),
                purchase_no=purchase_no or "",
            )
        )

    if normalized_warning_filter:
        total = len(items)
        if paginate:
            start = (page - 1) * page_size
            items = items[start:start + page_size]
    elif not paginate:
        total = len(items)

    return OrderSummaryResponse(items=items, total=total, page=page, page_size=page_size, has_more=lazy_has_more)


def _load_orders_by_ids(db: Session, order_ids: list[int]) -> list[Order]:
    if not order_ids:
        raise HTTPException(status_code=400, detail="请先选择订单")
    rows = db.scalars(select(Order).where(Order.id.in_(order_ids))).all()
    row_map = {row.id: row for row in rows}
    if len(row_map) != len(set(order_ids)):
        raise HTTPException(status_code=404, detail="部分订单不存在")
    return [row_map[order_id] for order_id in order_ids]


def _platform_display_name(value: str | None, fallback: str = "") -> str:
    platform = _to_str(value).strip()
    if not platform:
        return fallback
    normalized = PLATFORM_ALIASES.get(platform.lower(), platform.lower())
    catalog_names = {item["platform"]: item["display_name"] for item in PLATFORM_CATALOG}
    return (
        PLATFORM_DISPLAY_NAMES.get(platform)
        or PLATFORM_DISPLAY_NAMES.get(platform.lower())
        or PLATFORM_DISPLAY_NAMES.get(normalized)
        or catalog_names.get(platform)
        or catalog_names.get(platform.lower())
        or catalog_names.get(normalized)
        or fallback
        or platform
    )


def _format_order_list_money(amount: str | int | float | Decimal | None, currency: str | None) -> str:
    if amount is None or amount == "":
        return "-"
    try:
        numeric = Decimal(str(amount).strip())
    except (InvalidOperation, ValueError):
        return "-"
    currency_code = _to_str(currency or "CNY").strip().upper() or "CNY"
    formatted = f"{numeric:,.2f}"
    if currency_code == "CNY":
        symbol = "¥"
    elif currency_code == "RMB":
        symbol = f"{currency_code}\u00a0"
    elif currency_code == "USD":
        symbol = "US$"
    elif currency_code == "EUR":
        symbol = "€"
    elif currency_code == "GBP":
        symbol = "£"
    elif currency_code == "MXN":
        symbol = "MX$"
    else:
        symbol = f"{currency_code}\u00a0"
    if numeric < 0 and formatted.startswith("-"):
        return f"-{symbol}{formatted[1:]}"
    return f"{symbol}{formatted}"


def _outbound_scan_result_label(value: str | None) -> str:
    return {
        "success": "成功",
        "duplicate": "重复",
        "not_found": "未找到",
        "invalid": "无效",
        "error": "异常",
    }.get(value or "", value or "-")


def _table_export_columns(
    table_key: str,
    primary_column_key: str,
    export_columns: list[dict],
    user: LocalUser,
    db: Session,
    column_keys: str | None = None,
) -> list[dict]:
    system_columns = []
    for index, column in enumerate(export_columns):
        next_column = dict(column)
        next_column["order"] = column.get("order", index + 1)
        next_column["visible"] = column.get("visible", True)
        if next_column["key"] == primary_column_key:
            next_column["required"] = True
            if next_column.get("fixed") is None:
                next_column["fixed"] = "left"
        if next_column["key"] == "actions":
            next_column["settingsHidden"] = True
            if next_column.get("fixed") is None:
                next_column["fixed"] = "right"
        system_columns.append(next_column)

    column_map = {column["key"]: column for column in system_columns}
    requested_keys = [key.strip() for key in (column_keys or "").split(",") if key.strip()]
    requested_columns = [
        column_map[key]
        for key in requested_keys
        if key in column_map and not column_map[key].get("settingsHidden")
    ]
    if requested_columns:
        return requested_columns

    preference = db.scalar(
        select(UserTablePreference).where(
            UserTablePreference.user_id == user.id,
            UserTablePreference.table_key == table_key,
        )
    )
    user_columns = {
        column.get("key"): column
        for column in ((preference.config_json or {}).get("columns") if preference and preference.config_json else []) or []
        if column.get("key")
    }
    has_user_columns = bool(user_columns)
    max_user_order = max(
        [int(column.get("order") or 0) for column in user_columns.values() if isinstance(column.get("order"), (int, float))] or [0]
    )
    appended_order = max_user_order
    merged_columns = []
    for column in system_columns:
        user_column = user_columns.get(column["key"])
        required = bool(column.get("required")) or column["key"] == primary_column_key
        settings_hidden = bool(column.get("settingsHidden"))
        if settings_hidden:
            order = column["order"]
        elif user_column:
            order = user_column.get("order") if isinstance(user_column.get("order"), (int, float)) else column["order"]
        elif has_user_columns:
            appended_order += 1
            order = appended_order
        else:
            order = column["order"]
        visible = True if required or settings_hidden else (user_column.get("visible") if user_column else column.get("visible", True))
        merged_columns.append({**column, "visible": bool(visible), "order": int(order or column["order"])})

    return [
        column
        for column in sorted(merged_columns, key=lambda item: item["order"])
        if column.get("visible") and not column.get("settingsHidden")
    ]


def _order_list_export_columns(user: LocalUser, db: Session, column_keys: str | None = None) -> list[dict]:
    return _table_export_columns(ORDER_LIST_TABLE_KEY, ORDER_LIST_PRIMARY_COLUMN_KEY, ORDER_LIST_EXPORT_COLUMNS, user, db, column_keys)


def _order_list_export_value(column_key: str, dto: OrderDto) -> str:
    value_getters = {
        "platform_order_no": lambda: dto.platform_order_no or "-",
        "platform": lambda: _platform_display_name(dto.platform) or "-",
        "shop_name": lambda: dto.shop_name,
        "transaction_id": lambda: dto.transaction_id,
        "posting_number": lambda: dto.posting_number or "-",
        "tracking_number": lambda: dto.tracking_number or "-",
        "status": lambda: dto.status or "-",
        "platform_status": lambda: dto.platform_status,
        "fulfillment_type": lambda: "海外仓" if dto.is_overseas_warehouse else ("无需平台面单" if dto.logistics_label_exempt else (dto.fulfillment_type or "-")),
        "country_name_cn": lambda: dto.country_name_cn or dto.country_code or "-",
        "logistics_channel": lambda: dto.logistics_channel or ("未匹配" if dto.logistics_match_status == "unmatched" else "-"),
        "logistics_match_rule_name": lambda: dto.logistics_match_rule_name or "-",
        "order_amount": lambda: _format_order_list_money(dto.order_amount, dto.currency),
        "payment_at": lambda: _excel_datetime(dto.payment_at) or "-",
        "handover_at": lambda: _excel_datetime(dto.handover_at) or "-",
        "remaining_shipping_time": lambda: dto.remaining_shipping_time or ("-" if _should_show_remaining_shipping(dto.status) else ""),
        "created_at": lambda: _excel_datetime(dto.created_at) or "-",
    }
    getter = value_getters.get(column_key)
    if not getter:
        return ""
    value = getter()
    return "" if value is None else str(value)


def _order_summary_export_value(column_key: str, dto: OrderSummaryDto) -> str:
    value_getters = {
        "picking_at": lambda: _excel_datetime(dto.picking_at) or "-",
        "platform": lambda: _platform_display_name(dto.platform, dto.platform) or "-",
        "shop_name": lambda: dto.shop_name or "-",
        "platform_created_at": lambda: _excel_datetime(dto.platform_created_at) or "-",
        "order_no": lambda: dto.order_no or "-",
        "status": lambda: dto.status or "-",
        "platform_status": lambda: dto.platform_status or "-",
        "country_name_cn": lambda: dto.country_name_cn or dto.country_code or "-",
        "customer_name": lambda: dto.customer_name or "-",
        "sku": lambda: dto.sku or "-",
        "platform_product_name": lambda: dto.platform_product_name or "-",
        "quantity": lambda: dto.quantity,
        "unit_price": lambda: dto.unit_price or "-",
        "currency": lambda: dto.currency or "-",
        "buyer_selected_logistics": lambda: dto.buyer_selected_logistics or "-",
        "shipping_deadline_at": lambda: _excel_datetime(dto.shipping_deadline_at) or "-",
        "shipment_tracking_number": lambda: dto.shipment_tracking_number or "-",
        "dispatch_deadline_at": lambda: _excel_datetime(dto.dispatch_deadline_at) or "-",
        "product_name": lambda: dto.product_name or "-",
        "customer_confirm": lambda: dto.customer_confirm or "-",
        "warning": lambda: dto.warning or "-",
        "purchase_no": lambda: dto.purchase_no or "-",
        "shipping_time": lambda: _excel_datetime(dto.shipping_time) or "-",
    }
    getter = value_getters.get(column_key)
    if not getter:
        return ""
    value = getter()
    return "" if value is None else str(value)


def _inventory_export_value(column_key: str, dto: InventoryDto) -> str:
    value_getters = {
        "product_code": lambda: dto.product_code or "-",
        "product_name": lambda: dto.product_name or "-",
        "stock_qty": lambda: dto.stock_qty,
        "last_count_qty": lambda: dto.last_count_qty,
        "safety_stock": lambda: dto.safety_stock if dto.safety_stock is not None else "",
        "stock_status": lambda: dto.stock_status or "-",
        "remark": lambda: dto.remark or "",
        "updated_at": lambda: _excel_datetime(dto.updated_at) or "-",
    }
    getter = value_getters.get(column_key)
    if not getter:
        return ""
    value = getter()
    return "" if value is None else str(value)


def _outbound_scan_export_value(column_key: str, dto: OutboundScanRecordDto) -> str:
    value_getters = {
        "scanned_at": lambda: _excel_datetime(dto.scanned_at) or "-",
        "result": lambda: _outbound_scan_result_label(dto.result),
        "tracking_number": lambda: dto.tracking_number or "-",
        "platform": lambda: _platform_display_name(dto.platform, dto.platform) or "-",
        "shop_name": lambda: dto.shop_name or "-",
        "platform_order_no": lambda: dto.platform_order_no or "-",
        "posting_number": lambda: dto.posting_number or "-",
        "order_status": lambda: dto.order_status or "-",
        "platform_status": lambda: dto.platform_status or "-",
        "message": lambda: dto.message or "-",
        "scanned_by": lambda: dto.scanned_by or "-",
    }
    getter = value_getters.get(column_key)
    if not getter:
        return ""
    value = getter()
    return "" if value is None else str(value)


def _ensure_order_statuses(rows: list[Order], allowed_statuses: set[str]) -> None:
    invalid = [row.platform_order_id for row in rows if _derive_order_status(row) not in allowed_statuses]
    if invalid:
        raise HTTPException(status_code=400, detail=f"以下订单状态不允许当前操作: {', '.join(invalid[:10])}")


def _platform_tracking_number_from_posting(row: Order) -> str:
    if str(getattr(row, "platform", "") or "").lower() == "ozon":
        payload = getattr(row, "raw_payload", None) if isinstance(getattr(row, "raw_payload", None), dict) else {}
        status = str(getattr(row, "platform_status", "") or payload.get("status") or "").strip().lower()
        substatus = str(payload.get("substatus") or "").strip().lower()
        if status in {"awaiting_packaging", "awaiting_registration"} or substatus == "posting_created":
            return ""
        if not status:
            return ""
        return str(getattr(row, "posting_number", "") or "").strip()
    return ""


def _order_tracking_number_value(db: Session, row: Order) -> str:
    row_payload = getattr(row, "raw_payload", None) or {}
    row_platform = getattr(row, "platform", None)
    tracking_number = clean_tracking_number(getattr(row, "shipment_tracking_number", None), row_payload, row_platform)
    if tracking_number:
        return tracking_number
    payload_tracking = _tracking_number_from_payload(row_payload)
    if payload_tracking:
        return payload_tracking
    shipment = _latest_shipment(db, row.id)
    if shipment:
        shipment_tracking = clean_tracking_number(shipment.tracking_number, row_payload, row_platform)
        if shipment_tracking:
            return shipment_tracking
    platform_tracking = _platform_tracking_number_from_posting(row)
    if platform_tracking:
        return platform_tracking
    return ""


def _order_tracking_log_suffix(db: Session, row: Order) -> str:
    tracking_number = _order_tracking_number_value(db, row)
    return f"，货运单号：{tracking_number}" if tracking_number else ""


def _order_display_number(row: Order) -> str:
    return row.platform_order_no or row.posting_number or row.platform_order_id or str(row.id)


def _order_skips_platform_label(row: Order) -> bool:
    return order_is_overseas_warehouse(row) or order_is_logistics_label_exempt(row)


def _split_platform_label_exempt_rows(rows: list[Order]) -> tuple[list[Order], list[Order]]:
    exempt_rows = [row for row in rows if _order_skips_platform_label(row)]
    exempt_ids = {row.id for row in exempt_rows}
    return [row for row in rows if row.id not in exempt_ids], exempt_rows


def _mark_platform_label_exempt_rows_as_shipped(
    db: Session,
    rows: list[Order],
    operator: str,
    *,
    source: str,
    operated_at: datetime | None = None,
    extra: dict | None = None,
) -> None:
    if not rows:
        return
    now = operated_at or datetime.utcnow()
    status_before = {row.id: _derive_order_status(row) for row in rows}
    for row in rows:
        row.biz_status = ORDER_STATUS_SHIPPED
        row.local_status = "shipped"
        row.label_printed_at = row.label_printed_at or now
        row.shipped_at = getattr(row, "shipped_at", None) or now
        row.marked_shipped_at = getattr(row, "marked_shipped_at", None) or now
        row.updated_at = now
    add_order_operation_logs(
        db,
        rows,
        operation_type="label_exempt_shipped",
        operation_attribute="无需平台面单",
        description=lambda order: (
            f"订单 {_order_display_number(order)} 无需平台物流/面单和采购，"
            f"状态：{status_before.get(order.id) or '-'} -> {ORDER_STATUS_SHIPPED}"
        ),
        operator=operator,
        source=source,
        operated_at=now,
        extra={
            "skipped_reason": "label_exempt",
            "order_ids": [row.id for row in rows],
            **(extra or {}),
        },
    )


def _mark_logistics_rule_unmatched_as_shipped(
    db: Session,
    rows: list[Order],
    operator: str,
    *,
    source: str,
    operated_at: datetime | None = None,
    extra: dict | None = None,
) -> None:
    if not rows:
        return
    now = operated_at or datetime.utcnow()
    status_before = {row.id: _derive_order_status(row) for row in rows}
    for row in rows:
        row.biz_status = ORDER_STATUS_SHIPPED
        row.local_status = "shipped"
        row.shipped_at = row.shipped_at or now
        row.marked_shipped_at = row.marked_shipped_at or now
        row.updated_at = now
    add_order_operation_logs(
        db,
        rows,
        operation_type="logistics_rule_unmatched_shipped",
        operation_attribute="物流规则未匹配",
        description=lambda order: (
            f"订单 {_order_display_number(order)} 所属平台已启用物流规则但未匹配，"
            f"跳过同步物流、打印面单和生成采购单，状态：{status_before.get(order.id) or '-'} -> {ORDER_STATUS_SHIPPED}"
        ),
        operator=operator,
        source=source,
        operated_at=now,
        extra={
            "skipped_reason": "logistics_rule_unmatched",
            "order_ids": [row.id for row in rows],
            **(extra or {}),
        },
    )


def _order_item_ids_for_orders(db: Session, rows: list[Order]) -> list[int]:
    order_ids = [row.id for row in rows]
    if not order_ids:
        return []
    return db.scalars(
        select(OrderItem.id)
        .where(OrderItem.order_id.in_(order_ids))
        .order_by(asc(OrderItem.order_id), asc(OrderItem.id))
    ).all()


def _purchase_order_from_existing_sources(
    db: Session,
    order_item_ids: list[int],
) -> tuple[PurchaseOrder | None, bool]:
    expected_item_ids = {int(item_id) for item_id in order_item_ids}
    if not expected_item_ids:
        return None, False
    source_rows = db.execute(
        select(PurchaseOrderSource.order_item_id, PurchaseOrderSource.purchase_order_id)
        .where(PurchaseOrderSource.order_item_id.in_(expected_item_ids))
    ).all()
    source_item_ids = {int(order_item_id) for order_item_id, _purchase_order_id in source_rows}
    purchase_order_ids = {int(purchase_order_id) for _order_item_id, purchase_order_id in source_rows if purchase_order_id}
    if source_item_ids == expected_item_ids and len(purchase_order_ids) == 1:
        purchase = db.get(PurchaseOrder, next(iter(purchase_order_ids)))
        if purchase:
            return purchase, True
    return None, bool(source_item_ids)


def _generate_purchase_order_for_orders(
    db: Session,
    rows: list[Order],
    operator: str | None,
    remark: str = "",
    *,
    allow_existing: bool = False,
) -> PurchaseOrder:
    order_item_ids = _order_item_ids_for_orders(db, rows)
    if not order_item_ids:
        raise HTTPException(status_code=400, detail="所选订单没有商品明细，无法生成采购单")
    if allow_existing:
        purchase, has_existing_sources = _purchase_order_from_existing_sources(db, order_item_ids)
        if purchase:
            return purchase
        if has_existing_sources:
            raise HTTPException(status_code=400, detail="所选订单部分或多张采购单已存在，请刷新后重试")
    try:
        purchase, _ = _generate_or_append_purchase_order_for_item_ids(db, order_item_ids, operator, remark)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return purchase


def _move_orders_to_picking_after_purchase(
    db: Session,
    rows: list[Order],
    purchase: PurchaseOrder,
    operator: str,
    *,
    source: str,
    description: str,
) -> None:
    now = datetime.utcnow()
    for row in rows:
        row.biz_status = ORDER_STATUS_PICKING
        row.local_status = "picking"
        row.picking_at = row.picking_at or now
        row.updated_at = now
    add_order_operation_logs(
        db,
        rows,
        operation_type="to_picking",
        operation_attribute="修改订单基础信息",
        description=lambda order: (
            f"{description}，订单 {_order_display_number(order)} 状态：{ORDER_STATUS_WAITING_PURCHASE} -> {ORDER_STATUS_PICKING}"
        ),
        operator=operator,
        source=source,
        operated_at=now,
        extra={"purchase_order_id": purchase.id, "purchase_no": purchase.purchase_no},
    )


def _to_picking_existing_response(db: Session, order_ids: list[int]) -> OrderBatchResponse | None:
    rows = _load_orders_by_ids(db, order_ids)
    if any(_derive_order_status(row) != ORDER_STATUS_PICKING for row in rows):
        return None
    order_item_ids = _order_item_ids_for_orders(db, rows)
    purchase, _has_existing_sources = _purchase_order_from_existing_sources(db, order_item_ids)
    if not purchase:
        return None
    return OrderBatchResponse(
        updated=len(rows),
        message=f"订单已在采购单 {purchase.purchase_no} 中，并已转入配货中 {len(rows)} 条",
        purchase_order_id=purchase.id,
        purchase_no=purchase.purchase_no,
    )


def _rollback_orders_to_waiting_purchase_after_purchase_delete(
    db: Session,
    candidate_order_ids: list[int],
    purchase: PurchaseOrder,
    operator: str,
    *,
    operated_at: datetime | None = None,
) -> list[Order]:
    unique_order_ids = list(dict.fromkeys(candidate_order_ids or []))
    if not unique_order_ids:
        return []
    payment_start_at = _local_date_start_utc(PURCHASE_DELETE_ROLLBACK_PAYMENT_START_DATE)
    remaining_purchase_source_exists = exists().where(PurchaseOrderSource.order_id == Order.id)
    rows = db.scalars(
        select(Order).where(
            Order.id.in_(unique_order_ids),
            Order.biz_status == ORDER_STATUS_PICKING,
            Order.payment_at.is_not(None),
            Order.payment_at >= payment_start_at,
            ~remaining_purchase_source_exists,
        )
    ).all()
    if not rows:
        return []
    now = operated_at or datetime.utcnow()
    for row in rows:
        row.biz_status = ORDER_STATUS_WAITING_PURCHASE
        if row.local_status == "picking":
            row.local_status = "label_saved" if row.label_printed_at else "new"
        row.picking_at = None
        row.updated_at = now
    add_order_operation_logs(
        db,
        rows,
        operation_type="purchase_order_deleted_rollback",
        operation_attribute="删除采购单",
        description=lambda _order: f"采购单 {purchase.purchase_no} 已删除，订单已回到待采购",
        operator=operator,
        source=ORDER_LOG_MANUAL_SOURCE,
        operated_at=now,
        extra={
            "purchase_order_id": purchase.id,
            "purchase_no": purchase.purchase_no,
            "status_before": ORDER_STATUS_PICKING,
            "status_after": ORDER_STATUS_WAITING_PURCHASE,
            "payment_start_date": PURCHASE_DELETE_ROLLBACK_PAYMENT_START_DATE.isoformat(),
        },
    )
    return rows


def _role_menu_codes(role: Role | None, db: Session) -> list[str]:
    if not role:
        return []
    if role.code == ROLE_ADMIN:
        return ADMIN_MENU_CODES.copy()
    rows = db.scalars(
        select(RoleMenuPermission.menu_code)
        .where(RoleMenuPermission.role_id == role.id)
        .order_by(RoleMenuPermission.menu_code)
    ).all()
    allowed = {_normalize_menu_code(row) for row in rows}
    return [item["code"] for item in MENU_DEFINITIONS if item["code"] in allowed]


def _menu_codes_for_user(user: LocalUser, db: Session) -> list[str]:
    roles = _roles_for_user(user, db, enabled_only=True)
    if any(role.code == ROLE_ADMIN for role in roles):
        return ADMIN_MENU_CODES.copy()
    allowed: set[str] = set()
    for role in roles:
        allowed.update(_role_menu_codes(role, db))
    return [item["code"] for item in MENU_DEFINITIONS if item["code"] in allowed]


def _roles_for_user(user: LocalUser, db: Session, *, enabled_only: bool = False) -> list[Role]:
    rows = db.scalars(
        select(Role)
        .join(UserRole, UserRole.role_id == Role.id)
        .where(UserRole.user_id == user.id)
        .order_by(asc(Role.id))
    ).all()
    if not rows and user.role_id:
        legacy_role = db.get(Role, user.role_id)
        if legacy_role:
            rows = [legacy_role]
    if enabled_only:
        rows = [role for role in rows if role.enabled]
    return rows


def _is_admin_user(user: LocalUser, db: Session) -> bool:
    return any(role.code == ROLE_ADMIN for role in _roles_for_user(user, db))


def _role_ids_from_payload(role_ids: list[int] | None, role_id: int | None = None) -> list[int]:
    values = list(role_ids or [])
    if not values and role_id:
        values = [role_id]
    return [value for value in dict.fromkeys(values) if value]


def _roles_by_payload(role_ids: list[int] | None, role_id: int | None, db: Session) -> list[Role]:
    ids = _role_ids_from_payload(role_ids, role_id)
    if not ids:
        raise HTTPException(status_code=400, detail="请选择角色")
    roles = db.scalars(select(Role).where(Role.id.in_(ids))).all()
    role_map = {role.id: role for role in roles}
    missing_ids = [item for item in ids if item not in role_map]
    if missing_ids:
        raise HTTPException(status_code=400, detail="角色不存在")
    if any(not role.enabled for role in roles):
        raise HTTPException(status_code=400, detail="角色已停用")
    return [role_map[item] for item in ids]


def _normalize_wecom_mobile(value: str | None) -> str:
    mobile = (value or "").strip()
    if mobile and not re.fullmatch(r"1[3-9]\d{9}", mobile):
        raise HTTPException(status_code=400, detail="企微手机号格式不正确")
    return mobile


def _set_user_roles(user: LocalUser, roles: list[Role], db: Session) -> None:
    db.query(UserRole).filter(UserRole.user_id == user.id).delete(synchronize_session=False)
    for role in roles:
        db.add(UserRole(user_id=user.id, role_id=role.id))
    user.role_id = roles[0].id if roles else None
    user.updated_at = datetime.utcnow()


def _normalize_menu_code(code: str) -> str:
    return MENU_CODE_ALIASES.get(code, code)


def _set_role_menus(role: Role, menu_codes: list[str], db: Session) -> None:
    db.query(RoleMenuPermission).filter(RoleMenuPermission.role_id == role.id).delete(synchronize_session=False)
    db.flush()
    if role.code == ROLE_ADMIN:
        for code in ADMIN_MENU_CODES:
            db.add(RoleMenuPermission(role_id=role.id, menu_code=code))
        return
    normalized_codes = (_normalize_menu_code(code) for code in menu_codes)
    for normalized_code in dict.fromkeys(normalized_codes):
        if normalized_code in MENU_CODES:
            db.add(RoleMenuPermission(role_id=role.id, menu_code=normalized_code))


def _default_menus_for_role(role_code: str) -> list[str]:
    if role_code == ROLE_ADMIN:
        return ADMIN_MENU_CODES.copy()
    return DEFAULT_ROLE_MENUS.get(role_code, []).copy()


def _user_dto(user: LocalUser, db: Session) -> UserDto:
    roles = _roles_for_user(user, db)
    primary_role = next((role for role in roles if role.code == ROLE_ADMIN), roles[0] if roles else None)
    return UserDto(
        id=user.id,
        username=user.username,
        display_name=user.display_name or "",
        wecom_mobile=user.wecom_mobile or "",
        role_id=primary_role.id if primary_role else None,
        role_code=primary_role.code if primary_role else "",
        role_name=primary_role.name if primary_role else "",
        role_ids=[role.id for role in roles],
        role_codes=[role.code for role in roles],
        role_names=[role.name for role in roles],
        enabled=bool(user.enabled),
        created_at=_iso(user.created_at),
        updated_at=_iso(user.updated_at),
    )


def _user_option_dto(user: LocalUser) -> UserOptionDto:
    return UserOptionDto(id=user.id, username=user.username, display_name=user.display_name or user.username)


def _wecom_mention_user_option_dto(user: LocalUser) -> WeComMentionUserOptionDto:
    return WeComMentionUserOptionDto(
        id=user.id,
        username=user.username,
        display_name=user.display_name or user.username,
        wecom_mobile=(user.wecom_mobile or "").strip(),
    )


def _user_display_name(user: LocalUser | None) -> str:
    if not user:
        return ""
    return user.display_name or user.username


def _role_dto(role: Role, db: Session) -> RoleDto:
    return RoleDto(
        id=role.id,
        code=role.code,
        name=role.name,
        description=role.description or "",
        is_system=bool(role.is_system),
        enabled=bool(role.enabled),
        menus=_role_menu_codes(role, db),
        created_at=_iso(role.created_at),
        updated_at=_iso(role.updated_at),
    )


def _platform_setting_dto(row: PlatformSetting) -> PlatformSettingDto:
    return PlatformSettingDto(
        id=row.id,
        platform=row.platform,
        platform_name=row.platform_name or _platform_display_name(row.platform, row.platform),
        enabled=bool(row.enabled),
        updated_at=_iso(row.updated_at),
    )


def _platform_print_setting_dto(row: PlatformPrintSetting) -> PlatformPrintSettingDto:
    document_type = row.document_type or PRINT_DOCUMENT_TYPE_LABEL
    page_orientation = normalize_print_orientation(row.page_orientation or PRINT_ORIENTATION_AUTO)
    return PlatformPrintSettingDto(
        id=row.id,
        platform=row.platform,
        document_type=document_type,
        document_type_name=PRINT_DOCUMENT_TYPES.get(document_type, document_type),
        printer_name=row.printer_name or "",
        printer_system=row.printer_system or "",
        printer_device_uri=row.printer_device_uri or "",
        printer_driver_name=row.printer_driver_name or "",
        printer_port_name=row.printer_port_name or "",
        printer_fingerprint=row.printer_fingerprint or "",
        page_orientation=page_orientation,
        page_orientation_name=PRINT_ORIENTATION_TYPES.get(page_orientation, page_orientation),
        enabled=bool(row.enabled),
        remark=row.remark or "",
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _printer_name_key(value: str) -> str:
    return "".join(ch.lower() for ch in (value or "") if ch.isalnum())


def _printer_identity_from_dto(printer: PrinterDto) -> PrinterIdentity:
    return PrinterIdentity(
        name=printer.name,
        system=printer.system or "",
        device_uri=printer.device_uri or "",
        driver_name=printer.driver_name or "",
        port_name=printer.port_name or "",
        status=printer.status or "",
        online=printer.online,
    )


def _printer_fingerprint_from_dto(printer: PrinterDto) -> str:
    return printer_fingerprint(_printer_identity_from_dto(printer))


def _apply_printer_identity(row: PlatformPrintSetting, printer: PrinterDto) -> None:
    row.printer_name = printer.name
    row.printer_system = printer.system or ""
    row.printer_device_uri = printer.device_uri or ""
    row.printer_driver_name = printer.driver_name or ""
    row.printer_port_name = printer.port_name or ""
    row.printer_fingerprint = _printer_fingerprint_from_dto(printer)


def _run_printer_command(args: list[str], timeout: int = 20) -> subprocess.CompletedProcess | None:
    try:
        return subprocess.run(args, capture_output=True, text=True, timeout=timeout)
    except Exception:
        logger.exception("Printer command failed: %s", args)
        return None


def _powershell_executable() -> str | None:
    if platform.system().lower() != "windows":
        return None
    return (
        shutil.which("powershell.exe")
        or shutil.which("powershell")
        or shutil.which("pwsh.exe")
        or shutil.which("pwsh")
    )


def _parse_cups_status_printer_name(text: str) -> str:
    if text.startswith("printer "):
        parts = text.split(maxsplit=2)
        return parts[1] if len(parts) >= 2 else ""
    if not text.startswith("打印机"):
        return ""
    value = text.removeprefix("打印机").strip()
    if not value or value.startswith(("已", "正在")):
        return ""
    markers = ["现在", " 正在", " 闲置", "闲置", "已", "禁用", "启用", "，", ",", "。"]
    end_positions = [value.find(marker) for marker in markers if value.find(marker) > 0]
    return value[:min(end_positions)].strip() if end_positions else value


def _cups_default_printer(lpstat: str) -> str:
    result = _run_printer_command([lpstat, "-d"])
    output = (result.stdout or result.stderr or "").strip() if result else ""
    if not output:
        return ""
    separator = "：" if "：" in output else ":"
    if separator in output:
        return output.rsplit(separator, 1)[-1].strip()
    return ""


def _cups_printer_devices(lpstat: str) -> dict[str, str]:
    result = _run_printer_command([lpstat, "-v"])
    if not result or result.returncode != 0:
        return {}
    devices: dict[str, str] = {}
    for line in (result.stdout or "").splitlines():
        text = line.strip()
        if not text:
            continue
        if text.lower().startswith("device for ") and ":" in text:
            name, uri = text[len("device for "):].split(":", 1)
            devices[name.strip()] = uri.strip()
            continue
        match = re.match(r"^用于(.+?)的设备[:：]\s*(.+)$", text)
        if match:
            devices[match.group(1).strip()] = match.group(2).strip()
    return devices


def _cups_printer_statuses(lpstat: str) -> dict[str, dict]:
    result = _run_printer_command([lpstat, "-p"])
    if not result or result.returncode != 0:
        return {}
    statuses: dict[str, dict] = {}
    current_name = ""
    for raw_line in (result.stdout or "").splitlines():
        raw = raw_line.rstrip()
        text = raw.strip()
        if not text:
            continue
        if raw[:1].isspace() and current_name:
            current = statuses.setdefault(current_name, {"status": ""})
            current["status"] = "\n".join(part for part in [current.get("status", ""), text] if part)
            continue
        name = _parse_cups_status_printer_name(text)
        if not name:
            continue
        current_name = name
        statuses[name] = {"status": text}
    for item in statuses.values():
        status_text = str(item.get("status") or "")
        lower_status = status_text.lower()
        offline = any(token in lower_status for token in ("disabled", "offline", "not connected", "unplugged"))
        offline = offline or any(token in status_text for token in ("离线", "禁用", "未连接", "未接入"))
        item["online"] = not offline
    return statuses


def _list_cups_printers() -> list[PrinterDto]:
    lpstat = shutil.which("lpstat")
    if not lpstat:
        return []
    devices = _cups_printer_devices(lpstat)
    statuses = _cups_printer_statuses(lpstat)
    default_name = _cups_default_printer(lpstat)
    names = sorted(set(devices) | set(statuses), key=lambda value: value.lower())
    return [
        PrinterDto(
            name=name,
            display_name=name,
            system="cups",
            device_uri=devices.get(name, ""),
            status=str(statuses.get(name, {}).get("status") or ""),
            is_default=name == default_name,
            online=statuses.get(name, {}).get("online"),
            fingerprint=printer_fingerprint(
                PrinterIdentity(
                    name=name,
                    system="cups",
                    device_uri=devices.get(name, ""),
                    online=statuses.get(name, {}).get("online"),
                )
            ),
        )
        for name in names
    ]


def _list_windows_printers() -> list[PrinterDto]:
    powershell = _powershell_executable()
    if not powershell:
        return []
    script = r"""
$default = $null
try {
  $default = (Get-CimInstance Win32_Printer | Where-Object { $_.Default -eq $true } | Select-Object -First 1).Name
} catch {}
try {
  $items = @(Get-Printer -ErrorAction Stop | Sort-Object Name | ForEach-Object {
    [pscustomobject]@{
      Name=$_.Name
      DriverName=$_.DriverName
      PortName=$_.PortName
      PrinterStatus=($_.PrinterStatus -as [string])
      WorkOffline=([bool]$_.WorkOffline)
      IsDefault=($_.Name -eq $default)
    }
  })
} catch {
  $items = @(Get-CimInstance Win32_Printer | Sort-Object Name | ForEach-Object {
    [pscustomobject]@{
      Name=$_.Name
      DriverName=$_.DriverName
      PortName=$_.PortName
      PrinterStatus=($_.PrinterStatus -as [string])
      WorkOffline=([bool]$_.WorkOffline)
      IsDefault=([bool]$_.Default)
    }
  })
}
$items | ConvertTo-Json -Compress
"""
    result = _run_printer_command([powershell, "-NoProfile", "-Command", script], timeout=30)
    if not result or result.returncode != 0:
        return []
    try:
        data = json.loads((result.stdout or "[]").strip() or "[]")
    except json.JSONDecodeError:
        logger.exception("Failed to parse printer list from PowerShell")
        return []
    if isinstance(data, dict):
        data = [data]
    printers: list[PrinterDto] = []
    for item in data if isinstance(data, list) else []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("Name") or "").strip()
        if not name:
            continue
        status = str(item.get("PrinterStatus") or "")
        work_offline = bool(item.get("WorkOffline"))
        online = not (work_offline or "offline" in status.lower())
        printers.append(
            PrinterDto(
                name=name,
                display_name=name,
                system="windows",
                driver_name=str(item.get("DriverName") or ""),
                port_name=str(item.get("PortName") or ""),
                status=status,
                is_default=bool(item.get("IsDefault")),
                online=online,
                fingerprint=printer_fingerprint(
                    PrinterIdentity(
                        name=name,
                        system="windows",
                        driver_name=str(item.get("DriverName") or ""),
                        port_name=str(item.get("PortName") or ""),
                        status=status,
                        online=online,
                    )
                ),
            )
        )
    return printers


def _list_server_printers() -> list[PrinterDto]:
    if platform.system().lower() == "windows":
        return _list_windows_printers()
    return _list_cups_printers()


def _validated_server_printer(printer_name: str) -> PrinterDto:
    raw_name = (printer_name or "").strip()
    if not raw_name:
        raise HTTPException(status_code=400, detail="请选择打印机")
    printers = _list_server_printers()
    if not printers:
        raise HTTPException(status_code=400, detail="未检测到后台服务所在电脑的打印机")
    by_name = {printer.name: printer for printer in printers}
    if raw_name in by_name:
        return by_name[raw_name]
    normalized = _printer_name_key(raw_name)
    matches = [printer for printer in printers if _printer_name_key(printer.name) == normalized]
    if len(matches) == 1:
        return matches[0]
    raise HTTPException(status_code=400, detail="请选择后台服务所在电脑当前已有的打印机")


def _validated_server_printer_name(printer_name: str) -> str:
    return _validated_server_printer(printer_name).name


def _backfill_print_setting_printer_identity(db: Session) -> int:
    printers = _list_server_printers()
    if not printers:
        return 0
    by_name = {printer.name: printer for printer in printers}
    changed = 0
    rows = db.scalars(select(PlatformPrintSetting).where(PlatformPrintSetting.printer_name != "")).all()
    for row in rows:
        printer = by_name.get(row.printer_name or "")
        if not printer:
            normalized = _printer_name_key(row.printer_name or "")
            matches = [item for item in printers if _printer_name_key(item.name) == normalized]
            printer = matches[0] if len(matches) == 1 else None
        if not printer:
            continue
        fingerprint = _printer_fingerprint_from_dto(printer)
        if (
            (row.printer_system or "") == (printer.system or "")
            and (row.printer_device_uri or "") == (printer.device_uri or "")
            and (row.printer_driver_name or "") == (printer.driver_name or "")
            and (row.printer_port_name or "") == (printer.port_name or "")
            and (row.printer_fingerprint or "") == fingerprint
        ):
            continue
        _apply_printer_identity(row, printer)
        row.updated_at = datetime.utcnow()
        changed += 1
    return changed


def _platform_print_option_map(db: Session) -> dict[str, dict]:
    rows = db.scalars(
        select(PlatformPrintSetting).where(
            PlatformPrintSetting.enabled == True,
            PlatformPrintSetting.document_type == PRINT_DOCUMENT_TYPE_LABEL,
        )
    ).all()
    return {
        row.platform: {
            "page_orientation": label_orientation_for_platform(row.platform, row.page_orientation),
            "target_size_mm": label_size_mm_for_platform(row.platform),
        }
        for row in rows
    }


def _shipping_deadline_setting_dto(row: ShippingDeadlineSetting) -> ShippingDeadlineSettingDto:
    base_date_field = row.base_date_field or BASE_DATE_PLATFORM_CREATED
    return ShippingDeadlineSettingDto(
        id=row.id,
        platform=row.platform,
        platform_name=_platform_display_name(row.platform, "其他" if row.platform == OTHER_PLATFORM else row.platform),
        base_date_field=base_date_field,
        base_date_field_name=SHIPPING_DEADLINE_BASE_DATE_LABELS.get(base_date_field, base_date_field),
        offset_days=int(row.offset_days or 0),
        sort_order=int(row.sort_order or 0),
        enabled=bool(row.enabled),
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _dashboard_platform_setting_name(platform: str, fallback: str = "") -> str:
    if platform == "mercadolibre":
        return "MercadoLibre (MDK)"
    if platform == "dmsmatrix":
        return "Fruugo (DMSMatrix)"
    if platform == OTHER_DASHBOARD_PLATFORM:
        return "其他平台"
    return _platform_display_name(platform, fallback or platform)


def _dashboard_platform_setting_items(db: Session) -> list[DashboardPlatformSettingDto]:
    seed_default_platform_settings(db)
    receipt_rates = load_dashboard_receipt_rates(db)
    deadline_settings = load_shipping_deadline_settings(db)
    platform_rows = db.scalars(
        select(PlatformSetting)
        .where(PlatformSetting.enabled == True)
        .order_by(asc(PlatformSetting.sort_order), asc(PlatformSetting.platform))
    ).all()
    platforms = [(row.platform, row.platform_name or row.platform) for row in platform_rows]
    platforms.append((OTHER_DASHBOARD_PLATFORM, "其他平台"))
    return [
        DashboardPlatformSettingDto(
            platform=platform,
            platform_name=_dashboard_platform_setting_name(platform, platform_name),
            receipt_rate_pct=_dashboard_number(dashboard_receipt_rate_for(receipt_rates, platform) * 100, 2),
            fulfillment_days=int(
                (deadline_settings.get(canonical_deadline_platform(platform)) or deadline_settings.get(OTHER_PLATFORM)).offset_days
                if (deadline_settings.get(canonical_deadline_platform(platform)) or deadline_settings.get(OTHER_PLATFORM))
                else 0
            ),
        )
        for platform, platform_name in platforms
    ]


def _scheduled_task_dto(row: ScheduledTask) -> ScheduledTaskDto:
    settings = dict(row.settings or {}) if isinstance(row.settings, dict) else {}
    settings.pop("test_mode", None)
    settings.pop("max_orders", None)
    return ScheduledTaskDto(
        id=row.id,
        name=row.name,
        task_type=row.task_type or "auto_order_pipeline",
        cron_expr=row.cron_expr,
        enabled=bool(row.enabled),
        settings=settings,
        remark=row.remark or "",
        last_run_at=_iso(row.last_run_at),
        last_status=row.last_status or "",
        last_message=row.last_message or "",
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _scheduled_task_run_dto(row: ScheduledTaskRun, pdf_export_platforms: list[str] | None = None) -> ScheduledTaskRunDto:
    return ScheduledTaskRunDto(
        id=row.id,
        scheduled_task_id=row.scheduled_task_id,
        task_type=row.task_type or "",
        trigger_mode=row.trigger_mode or "",
        status=row.status or "",
        summary=row.summary or "",
        stats_json=row.stats_json if isinstance(row.stats_json, dict) else {"items": row.stats_json or []},
        pdf_export_platforms=pdf_export_platforms or [],
        attempt_no=int(row.attempt_no or 0),
        max_retry_count=int(row.max_retry_count or 0),
        parent_run_id=row.parent_run_id,
        original_run_id=row.original_run_id,
        next_retry_at=_iso(row.next_retry_at),
        retry_reason=row.retry_reason or "",
        email_sent=bool(row.email_sent),
        email_error=row.email_error or "",
        started_at=_iso(row.started_at),
        ended_at=_iso(row.ended_at),
        created_at=_iso(row.created_at),
    )


def _email_smtp_setting_dto(row: EmailSmtpSetting) -> EmailSmtpSettingDto:
    return EmailSmtpSettingDto(
        provider=row.provider or "qq",
        enabled=bool(row.enabled),
        smtp_host=row.smtp_host or "smtp.qq.com",
        smtp_port=int(row.smtp_port or 465),
        use_ssl=bool(row.use_ssl),
        sender_email=row.sender_email or "",
        sender_name=row.sender_name or "",
        notification_recipients=notification_recipient_values(row),
        has_auth_code=bool(row.encrypted_auth_code),
        last_test_at=_iso(row.last_test_at),
        last_test_status=row.last_test_status or "",
        last_test_message=row.last_test_message or "",
        updated_at=_iso(row.updated_at),
    )


def _wecom_robot_setting_dto(row: WeComRobotSetting) -> WeComRobotSettingDto:
    webhook_url_masked = ""
    if row.encrypted_webhook_url:
        try:
            webhook_url_masked = mask_wecom_webhook_url(decrypt_wecom_webhook_url(row))
        except Exception:
            webhook_url_masked = "已保存"
    return WeComRobotSettingDto(
        has_webhook_url=bool(row.encrypted_webhook_url),
        webhook_url_masked=webhook_url_masked,
        timeout_seconds=max(1, int(row.timeout_seconds or 30)),
        max_retries=max(0, int(row.max_retries or 0)),
        rate_limit_per_minute=max(1, int(row.rate_limit_per_minute or 20)),
        default_mentioned_user_ids=loads_int_list(row.default_mentioned_user_ids),
        default_mentioned_list=loads_string_list(row.default_mentioned_list),
        default_mentioned_mobile_list=loads_string_list(row.default_mentioned_mobile_list),
        default_prompt=(row.default_prompt or "").strip(),
        purchase_order_notify_enabled=bool(getattr(row, "purchase_order_notify_enabled", False)),
        updated_at=_iso(row.updated_at),
    )


def _validate_wecom_robot_payload(
    payload: WeComRobotSettingUpdateRequest,
    row: WeComRobotSetting,
) -> dict:
    current_has_webhook = bool(row.encrypted_webhook_url)
    webhook_url = (payload.webhook_url or "").strip() if payload.webhook_url is not None else ""
    if webhook_url:
        webhook_url = validate_wecom_webhook_url(webhook_url)
    elif not current_has_webhook:
        raise HTTPException(status_code=422, detail="webhook_url不能为空")

    timeout_seconds = int(payload.timeout_seconds)
    max_retries = int(payload.max_retries)
    rate_limit_per_minute = int(payload.rate_limit_per_minute)
    if timeout_seconds < 1 or timeout_seconds > 300:
        raise HTTPException(status_code=422, detail="timeout_seconds 必须在 1 到 300 之间")
    if max_retries < 0 or max_retries > 10:
        raise HTTPException(status_code=422, detail="max_retries 必须在 0 到 10 之间")
    if rate_limit_per_minute < 1 or rate_limit_per_minute > 20:
        raise HTTPException(status_code=422, detail="rate_limit_per_minute 必须在 1 到 20 之间")

    return {
        "webhook_url": webhook_url,
        "timeout_seconds": timeout_seconds,
        "max_retries": max_retries,
        "rate_limit_per_minute": rate_limit_per_minute,
        "default_mentioned_user_ids": normalize_int_list(payload.default_mentioned_user_ids),
        "default_mentioned_list": normalize_string_list(payload.default_mentioned_list),
        "default_mentioned_mobile_list": normalize_string_list(payload.default_mentioned_mobile_list),
        "default_prompt": str(payload.default_prompt or "").strip(),
        "purchase_order_notify_enabled": bool(payload.purchase_order_notify_enabled),
    }


def _translation_provider_setting_dto(row: TranslationProviderSetting) -> TranslationProviderSettingDto:
    secret_key_masked = ""
    if row.encrypted_secret_key:
        try:
            secret_key_masked = mask_translation_secret_key(decrypt_translation_secret_key(row))
        except Exception:
            secret_key_masked = "已保存"
    provider = normalize_translation_provider(row.provider)
    return TranslationProviderSettingDto(
        provider=provider,
        provider_name=row.provider_name or translation_provider_name(provider),
        enabled=bool(row.enabled),
        app_id=(row.app_id or "").strip(),
        has_secret_key=bool(row.encrypted_secret_key),
        secret_key_masked=secret_key_masked,
        endpoint=(row.endpoint or translation_provider_endpoint(provider)).strip(),
        source_language=(row.source_language or "auto").strip() or "auto",
        timeout_seconds=max(1, int(row.timeout_seconds or 30)),
        max_retries=max(0, int(row.max_retries or 0)),
        batch_size=max(1, int(row.batch_size or 80)),
        batch_chars=max(100, int(row.batch_chars or 5000)),
        provider_options=translation_provider_options_dict(row),
        last_test_at=_iso(row.last_test_at),
        last_test_status=row.last_test_status or "",
        last_test_message=row.last_test_message or "",
        updated_at=_iso(row.updated_at),
    )


def _validate_translation_provider_payload(
    payload: TranslationProviderSettingUpdateRequest,
    row: TranslationProviderSetting,
) -> dict:
    try:
        provider = normalize_translation_provider(payload.provider)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    app_id = str(payload.app_id or "").strip()
    secret_key = str(payload.secret_key or "").strip() if payload.secret_key is not None else ""
    endpoint = str(payload.endpoint or "").strip() or translation_provider_endpoint(provider)
    parsed_endpoint = urlparse(endpoint)
    if parsed_endpoint.scheme not in {"http", "https"} or not parsed_endpoint.netloc:
        raise HTTPException(status_code=422, detail="endpoint 必须是有效的 http(s) 地址")
    source_language = str(payload.source_language or "auto").strip() or "auto"
    if len(source_language) > 20:
        raise HTTPException(status_code=422, detail="source_language 不能超过 20 个字符")
    timeout_seconds = int(payload.timeout_seconds)
    max_retries = int(payload.max_retries)
    batch_size = int(payload.batch_size)
    batch_chars = int(payload.batch_chars)
    if timeout_seconds < 1 or timeout_seconds > 300:
        raise HTTPException(status_code=422, detail="timeout_seconds 必须在 1 到 300 之间")
    if max_retries < 0 or max_retries > 10:
        raise HTTPException(status_code=422, detail="max_retries 必须在 0 到 10 之间")
    if batch_size < 1 or batch_size > 200:
        raise HTTPException(status_code=422, detail="batch_size 必须在 1 到 200 之间")
    if batch_chars < 100 or batch_chars > 20000:
        raise HTTPException(status_code=422, detail="batch_chars 必须在 100 到 20000 之间")
    if payload.enabled:
        if not app_id:
            raise HTTPException(status_code=422, detail="app_id不能为空")
        if not secret_key and not row.encrypted_secret_key:
            raise HTTPException(status_code=422, detail="secret_key不能为空")
    return {
        "provider": provider,
        "provider_name": translation_provider_name(provider),
        "enabled": bool(payload.enabled),
        "app_id": app_id,
        "secret_key": secret_key,
        "endpoint": endpoint,
        "source_language": source_language,
        "timeout_seconds": timeout_seconds,
        "max_retries": max_retries,
        "batch_size": batch_size,
        "batch_chars": batch_chars,
        "provider_options_json": dumps_translation_provider_options(payload.provider_options),
    }


def _translation_client_for_db(db: Session):
    row = get_translation_provider_setting(db, DEFAULT_TRANSLATION_PROVIDER)
    return build_translation_client_from_setting(row)


TEXT_TRANSLATION_MAX_CHARS = 5000
TEXT_TRANSLATION_LOG_PLATFORM = "baidu_translate"
TEXT_TRANSLATION_LOG_OPERATION = "text_translation"


def _translation_language_codes(*, include_auto: bool = False) -> set[str]:
    codes = {str(item["code"]).strip() for item in list_translation_language_presets() if str(item.get("code") or "").strip()}
    if include_auto:
        codes.add("auto")
    return codes


def _translation_language_label(code: str) -> str:
    normalized = str(code or "").strip()
    if normalized == "auto":
        return "自动检测（auto）"
    for item in list_translation_language_presets():
        if item.get("code") == normalized:
            return str(item.get("label") or normalized)
    return normalized


def _log_text_translation_call(
    *,
    provider_row: TranslationProviderSetting | None,
    request_id: str,
    user: LocalUser,
    source_language: str,
    target_language: str,
    source_text: str,
    translated_text: str = "",
    status_value: str,
    error_message: str | None = None,
    duration_ms: int | None = None,
) -> None:
    provider = str((provider_row.provider if provider_row is not None else DEFAULT_TRANSLATION_PROVIDER) or DEFAULT_TRANSLATION_PROVIDER)
    endpoint = str((provider_row.endpoint if provider_row is not None else "") or translation_provider_endpoint(provider))
    log_api_call(
        platform=TEXT_TRANSLATION_LOG_PLATFORM,
        account_id=str(provider_row.app_id or provider) if provider_row is not None else provider,
        method="POST",
        url=endpoint,
        operation=TEXT_TRANSLATION_LOG_OPERATION,
        status=status_value,
        request_id=request_id,
        request_body={
            "provider": provider,
            "from": source_language,
            "from_label": _translation_language_label(source_language),
            "to": target_language,
            "to_label": _translation_language_label(target_language),
            "q": source_text,
        },
        response_status=200 if status_value == "success" else 400,
        response_body={"translated_text": translated_text} if translated_text else None,
        error_message=error_message,
        duration_ms=duration_ms,
        extra={
            "user_id": user.id,
            "username": user.username,
            "display_name": user.display_name or "",
            "source_char_count": len(source_text),
            "translated_char_count": len(translated_text),
        },
    )


def _translate_text_once(
    payload: TextTranslationRequest,
    *,
    user: LocalUser,
    db: Session,
) -> TextTranslationResponse:
    text_value = str(payload.text or "").strip()
    if not text_value:
        raise HTTPException(status_code=422, detail="text不能为空")
    if len(text_value) > TEXT_TRANSLATION_MAX_CHARS:
        raise HTTPException(status_code=422, detail=f"text不能超过 {TEXT_TRANSLATION_MAX_CHARS} 个字符")

    source_language = str(payload.source_language or "auto").strip() or "auto"
    target_language = str(payload.target_language or "").strip()
    if not target_language:
        raise HTTPException(status_code=422, detail="target_language不能为空")
    if target_language == "auto":
        raise HTTPException(status_code=422, detail="target_language不能为auto")

    if source_language not in _translation_language_codes(include_auto=True):
        raise HTTPException(status_code=422, detail="source_language不在百度翻译语种列表中")
    if target_language not in _translation_language_codes(include_auto=False):
        raise HTTPException(status_code=422, detail="target_language不在百度翻译语种列表中")

    request_id = uuid.uuid4().hex
    started = perf_counter()
    row: TranslationProviderSetting | None = None
    translated_text = ""
    try:
        row = get_translation_provider_setting(db, DEFAULT_TRANSLATION_PROVIDER)
        client = build_translation_client_from_setting(row)
        translated = client.translate_texts(
            [text_value],
            from_lang=source_language,
            to_lang=target_language,
        )
        translated_text = str(translated.get(text_value) or "").strip()
        if not translated_text:
            raise RuntimeError("翻译服务未返回结果")
    except Exception as exc:
        _log_text_translation_call(
            provider_row=row,
            request_id=request_id,
            user=user,
            source_language=source_language,
            target_language=target_language,
            source_text=text_value,
            status_value="failed",
            error_message=str(exc) or "文字翻译失败",
            duration_ms=int((perf_counter() - started) * 1000),
        )
        raise HTTPException(status_code=400, detail=str(exc) or "文字翻译失败") from exc

    _log_text_translation_call(
        provider_row=row,
        request_id=request_id,
        user=user,
        source_language=source_language,
        target_language=target_language,
        source_text=text_value,
        translated_text=translated_text,
        status_value="success",
        duration_ms=int((perf_counter() - started) * 1000),
    )
    return TextTranslationResponse(
        status="success",
        message="翻译完成",
        request_id=request_id,
        provider=row.provider if row is not None else DEFAULT_TRANSLATION_PROVIDER,
        source_language=source_language,
        target_language=target_language,
        translated_text=translated_text,
        source_char_count=len(text_value),
        translated_char_count=len(translated_text),
    )


def _encrypt_model_api_key(api_key: str | None) -> bytes | None:
    value = (api_key or "").strip()
    if not value:
        return None
    return get_credential_manager().encrypt_credentials({"api_key": value})


def _decrypt_model_api_key(row: ModelEndpoint) -> str:
    data = get_credential_manager().decrypt_credentials(row.encrypted_api_key)
    return str(data.get("api_key") or "")


def _mask_secret(value: str) -> str:
    if not value:
        return "-"
    if len(value) <= 8:
        return "********"
    return f"{value[:4]}****{value[-4:]}"


def _model_endpoint_dto(row: ModelEndpoint) -> ModelEndpointDto:
    api_key = ""
    if row.encrypted_api_key:
        try:
            api_key = _decrypt_model_api_key(row)
        except Exception:
            api_key = ""
    return ModelEndpointDto(
        id=row.id,
        name=row.name or "",
        base_url=row.base_url or "",
        api_key_masked=_mask_secret(api_key),
        enabled=bool(row.enabled),
        remark=row.remark or "",
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _model_setting_dto(row: ModelSetting) -> ModelSettingDto:
    endpoint = row.endpoint
    return ModelSettingDto(
        id=row.id,
        name=row.name or "",
        model=row.model or "",
        endpoint_id=row.endpoint_id,
        endpoint_name=(endpoint.name if endpoint else "") or "",
        endpoint_enabled=bool(endpoint.enabled) if endpoint else False,
        url=(endpoint.base_url if endpoint else "") or "",
        is_default=bool(row.is_default),
        supports_vision=bool(getattr(row, "supports_vision", False)),
        enabled=bool(row.enabled),
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _normalize_base_url(value: str) -> str:
    return (value or "").strip().rstrip("/")


def _model_chat_completions_url(base_url: str) -> str:
    normalized = _normalize_base_url(base_url)
    if not normalized:
        return ""
    if normalized.endswith("/chat/completions"):
        return normalized
    if normalized.endswith("/v1"):
        return f"{normalized}/chat/completions"
    return f"{normalized}/v1/chat/completions"


def _validate_model_endpoint_payload(
    payload: ModelEndpointUpsertRequest,
    *,
    current: ModelEndpoint | None = None,
) -> dict:
    name = (payload.name or "").strip()
    base_url = _normalize_base_url(payload.base_url)
    api_key = (payload.api_key or "").strip() if payload.api_key is not None else ""
    if not name:
        raise HTTPException(status_code=422, detail="接口配置名称不能为空")
    if not base_url:
        raise HTTPException(status_code=422, detail="base url不能为空")
    parsed = urlparse(base_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(status_code=422, detail="base url必须是有效的 http 或 https 地址")
    if current is None and not api_key:
        raise HTTPException(status_code=422, detail="api key不能为空")
    return {
        "name": name,
        "base_url": base_url,
        "api_key": api_key,
        "enabled": bool(payload.enabled),
        "remark": (payload.remark or "").strip(),
    }


def _ensure_unique_model_endpoint_name(db: Session, name: str, current_id: int | None = None) -> None:
    query = select(ModelEndpoint).where(ModelEndpoint.name == name)
    if current_id is not None:
        query = query.where(ModelEndpoint.id != current_id)
    if db.scalar(query.limit(1)):
        raise HTTPException(status_code=409, detail="接口配置名称已存在")


def _validate_model_setting_payload(payload: ModelSettingUpsertRequest, db: Session) -> tuple[dict, ModelEndpoint]:
    name = (payload.name or "").strip()
    model = (payload.model or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="模型名称不能为空")
    if not model:
        raise HTTPException(status_code=422, detail="模型标识不能为空")
    if payload.endpoint_id is None:
        raise HTTPException(status_code=422, detail="请选择接口配置")
    endpoint = db.get(ModelEndpoint, payload.endpoint_id)
    if not endpoint:
        raise HTTPException(status_code=422, detail="接口配置不存在")
    enabled = bool(payload.enabled)
    return {
        "name": name,
        "model": model,
        "endpoint_id": endpoint.id,
        "enabled": enabled,
        "is_default": bool(payload.is_default) if enabled else False,
        "supports_vision": bool(payload.supports_vision),
    }, endpoint


def _ensure_unique_model_setting_name(db: Session, name: str, current_id: int | None = None) -> None:
    query = select(ModelSetting).where(ModelSetting.name == name)
    if current_id is not None:
        query = query.where(ModelSetting.id != current_id)
    if db.scalar(query.limit(1)):
        raise HTTPException(status_code=409, detail="模型名称已存在")


def _clear_default_model_settings(db: Session, current_id: int | None = None) -> None:
    query = select(ModelSetting).where(ModelSetting.is_default == True)
    if current_id is not None:
        query = query.where(ModelSetting.id != current_id)
    for row in db.scalars(query).all():
        row.is_default = False
        row.updated_at = datetime.utcnow()


def _get_default_model_setting(db: Session) -> ModelSetting:
    setting = db.scalar(
        select(ModelSetting)
        .options(selectinload(ModelSetting.endpoint))
        .where(ModelSetting.enabled == True, ModelSetting.is_default == True)
        .order_by(desc(ModelSetting.updated_at), desc(ModelSetting.id))
        .limit(1)
    )
    if not setting:
        raise HTTPException(status_code=422, detail="未配置启用的默认模型")
    endpoint = setting.endpoint
    if not endpoint:
        raise HTTPException(status_code=422, detail="默认模型接口配置不存在")
    if not endpoint.enabled:
        raise HTTPException(status_code=422, detail="默认模型接口配置已禁用")
    return setting


def _post_model_chat_completion(
    setting: ModelSetting,
    messages: list[dict],
    *,
    max_tokens: int = 800,
    temperature: float = 0,
    timeout: float = 60,
) -> str:
    endpoint = setting.endpoint
    if not endpoint:
        raise HTTPException(status_code=422, detail="模型接口配置不存在")
    api_key = _decrypt_model_api_key(endpoint)
    if not api_key:
        raise HTTPException(status_code=422, detail="默认模型 api key 为空")
    upstream_url = _model_chat_completions_url(endpoint.base_url)
    if not upstream_url:
        raise HTTPException(status_code=422, detail="默认模型 base url 为空")
    try:
        with httpx.Client(timeout=timeout) as client:
            response = client.post(
                upstream_url,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": setting.model,
                    "messages": messages,
                    "max_tokens": max_tokens,
                    "temperature": temperature,
                },
            )
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"默认模型连接失败：{exc}") from exc
    if response.status_code >= 400:
        detail = response.text[:500] if response.text else response.reason_phrase
        raise HTTPException(status_code=502, detail=f"默认模型接口返回错误：{detail}")
    try:
        payload = response.json()
    except ValueError as exc:
        raise HTTPException(status_code=502, detail="默认模型返回内容不是 JSON") from exc
    choices = payload.get("choices") if isinstance(payload, dict) else None
    if not isinstance(choices, list) or not choices:
        raise HTTPException(status_code=502, detail="默认模型返回内容缺少 choices")
    first = choices[0] if isinstance(choices[0], dict) else {}
    message = first.get("message") if isinstance(first.get("message"), dict) else {}
    content = message.get("content")
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        parts = [
            str(item.get("text") or "")
            for item in content
            if isinstance(item, dict) and item.get("type") in {"text", "output_text"}
        ]
        return "".join(parts).strip()
    raise HTTPException(status_code=502, detail="默认模型返回内容缺少文本")


def _get_ai_image_model_setting(
    db: Session,
    model_setting_id: int | None,
    *,
    require_vision: bool = False,
) -> ModelSetting:
    if model_setting_id is None:
        setting = _get_default_model_setting(db)
    else:
        setting = db.scalar(
            select(ModelSetting)
            .options(selectinload(ModelSetting.endpoint))
            .where(ModelSetting.id == model_setting_id)
            .limit(1)
        )
        if not setting:
            raise HTTPException(status_code=422, detail="图片模型不存在")
        if not setting.enabled:
            raise HTTPException(status_code=422, detail="图片模型已禁用")
        if not setting.endpoint:
            raise HTTPException(status_code=422, detail="图片模型接口配置不存在")
        if not setting.endpoint.enabled:
            raise HTTPException(status_code=422, detail="图片模型接口已禁用")
    if require_vision and not bool(getattr(setting, "supports_vision", False)):
        raise HTTPException(status_code=422, detail="所选模型未启用图片理解")
    return setting


async def _save_ai_image_upload(file: UploadFile, output_dir: Path, index: int) -> Path:
    original_name = Path(file.filename or f"image-{index}").name
    target = output_dir / f"input-{index}.upload"
    total = 0
    max_bytes = 25 * 1024 * 1024
    with target.open("wb") as stream:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                raise HTTPException(status_code=400, detail="单张图片不能超过 25MB")
            stream.write(chunk)
    if not total:
        raise HTTPException(status_code=400, detail=f"图片文件为空: {original_name}")
    try:
        metadata = validate_image_file(target)
    except AiImageError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    suffix = ".jpg" if metadata["format"] == "jpeg" else f".{metadata['format']}"
    normalized = target.with_suffix(suffix)
    target.replace(normalized)
    return normalized


def _ai_image_object_key(job_id: str, group: str, index: int, path: Path) -> str:
    day = datetime.utcnow().strftime("%Y/%m/%d")
    stem = re.sub(r"[^A-Za-z0-9._-]+", "-", path.stem).strip("-._")[:80] or "image"
    suffix = path.suffix.lower() or ".png"
    return f"ai_toolbox/{day}/{job_id}/{group}/{index:02d}-{stem}{suffix}"


def _validated_ai_image_object_key(value: str) -> str:
    object_key = str(value or "").strip()
    parts = object_key.split("/")
    if (
        not object_key.startswith("ai_toolbox/")
        or "\\" in object_key
        or any(not part or part in {".", ".."} for part in parts)
        or any(not re.fullmatch(r"[A-Za-z0-9._-]+", part) for part in parts)
    ):
        raise HTTPException(status_code=400, detail="图片文件标识无效")
    return object_key


def _ai_image_download_filename(value: str, object_key: str) -> str:
    fallback = object_key.rsplit("/", 1)[-1] or "image"
    basename = str(value or "").replace("\\", "/").rsplit("/", 1)[-1]
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f\x7f]+', "_", basename).strip(" .")
    return (cleaned or fallback)[:180]


def _open_ai_image_oss_object(object_key: str):
    try:
        import oss2
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="图片存储服务依赖不可用") from exc

    try:
        config = load_oss_config()
        auth = oss2.Auth(config.access_key_id, config.access_key_secret)
        bucket = oss2.Bucket(auth, config.endpoint, config.bucket_name)
        return bucket.get_object(object_key)
    except ImageStorageError as exc:
        raise HTTPException(status_code=503, detail="图片存储服务未配置") from exc
    except Exception as exc:
        if getattr(exc, "status", None) == 404 or exc.__class__.__name__ == "NoSuchKey":
            raise HTTPException(status_code=404, detail="图片文件不存在") from exc
        logger.exception("Failed to open AI image from object storage: %s", object_key)
        raise HTTPException(status_code=502, detail="图片文件读取失败") from exc


def _iter_ai_image_oss_object(result):
    try:
        while True:
            chunk = result.read(1024 * 1024)
            if not chunk:
                break
            yield chunk
    finally:
        close = getattr(result, "close", None)
        if callable(close):
            close()


def _iter_ai_image_download_archive(output):
    try:
        while True:
            chunk = output.read(1024 * 1024)
            if not chunk:
                break
            yield chunk
    finally:
        output.close()


async def _upload_ai_image_asset(path: Path, *, job_id: str, group: str, index: int) -> AiImageAssetDto:
    metadata = validate_image_file(path)
    object_key, url = await asyncio.to_thread(
        upload_file_to_oss,
        path,
        object_key=_ai_image_object_key(job_id, group, index, path),
    )
    return AiImageAssetDto(
        name=path.name,
        url=url,
        oss_object_key=object_key,
        width=int(metadata["width"]),
        height=int(metadata["height"]),
        format=str(metadata["format"]),
        size_bytes=path.stat().st_size,
    )


def _scheduled_task_run_step_dto(row: ScheduledTaskRunStep) -> ScheduledTaskRunStepDto:
    return ScheduledTaskRunStepDto(
        id=row.id,
        run_id=row.run_id,
        step_code=row.step_code,
        step_name=row.step_name,
        status=row.status or "",
        message=row.message or "",
        stats_json=row.stats_json if isinstance(row.stats_json, dict) else {"items": row.stats_json or []},
        payload_json=row.payload_json if isinstance(row.payload_json, dict) else {"items": row.payload_json or []},
        started_at=_iso(row.started_at),
        ended_at=_iso(row.ended_at),
    )


def _scheduled_task_run_order_dto(
    row: ScheduledTaskRunOrder,
    platform_order_no: str | None = None,
    label_file_path: str | None = None,
) -> ScheduledTaskRunOrderDto:
    document_type_name = PRINT_PLATFORM_CHINESE_LABEL_NAME if row.platform == PRINT_PLATFORM_CHINESE_LABEL and int(row.order_id or 0) == 0 else "面单"
    return ScheduledTaskRunOrderDto(
        id=row.id,
        run_id=row.run_id,
        order_id=row.order_id,
        platform_order_no=platform_order_no or (PRINT_PLATFORM_CHINESE_LABEL_NAME if int(row.order_id or 0) == 0 else ""),
        platform=row.platform or "",
        document_type_name=document_type_name,
        purchase_order_id=row.purchase_order_id,
        pdf_generated=bool(row.pdf_generated),
        pdf_file_path=row.pdf_file_path or "",
        has_label_file=bool(label_file_path),
        label_file_path=label_file_path or "",
        printer_name=row.printer_name or "",
        print_job_name=row.print_job_name or "",
        print_submitted=bool(row.print_submitted),
        print_message=row.print_message or "",
        status_before=row.status_before or "",
        status_after=row.status_after or "",
        needs_reprint=bool(row.needs_reprint),
        error_message=row.error_message or "",
        created_at=_iso(row.created_at),
    )


def _scheduled_task_run_platform_document_type(platform: str) -> str:
    return PRINT_PLATFORM_CHINESE_LABEL_NAME if platform == PRINT_PLATFORM_CHINESE_LABEL else "面单"


def _scheduled_run_is_successful(row: ScheduledTaskRun | None) -> bool:
    if not row:
        return False
    summary = row.summary or ""
    return (row.status or "") in {"success", "partial_success"} or "任务完成" in summary or "成功" in summary


def _append_unique_text(values: list[str], value: str | None, *, limit: int | None = None) -> None:
    text = str(value or "").strip()
    if not text or text in values:
        return
    if limit is not None and len(values) >= limit:
        return
    values.append(text)


def _scheduled_task_run_platform_rows(
    run_id: int,
    db: Session,
    needs_reprint: bool | None = None,
) -> list[ScheduledTaskRunPlatformDto]:
    if needs_reprint is True:
        refresh_reprint_candidates(db, run_id)
    latest_labels = (
        select(
            Shipment.order_id.label("order_id"),
            LabelFile.file_path.label("file_path"),
            func.row_number()
            .over(partition_by=Shipment.order_id, order_by=LabelFile.id.desc())
            .label("rn"),
        )
        .join(LabelFile, LabelFile.shipment_id == Shipment.id)
        .subquery()
    )
    stmt = (
        select(ScheduledTaskRunOrder, Order.platform_order_no, Order.platform_order_id, latest_labels.c.file_path)
        .outerjoin(Order, Order.id == ScheduledTaskRunOrder.order_id)
        .outerjoin(latest_labels, (latest_labels.c.order_id == ScheduledTaskRunOrder.order_id) & (latest_labels.c.rn == 1))
        .where(ScheduledTaskRunOrder.run_id == run_id)
        .order_by(asc(ScheduledTaskRunOrder.id))
    )
    if needs_reprint is not None:
        stmt = stmt.where(ScheduledTaskRunOrder.needs_reprint == needs_reprint)
    rows = db.execute(stmt).all()

    grouped: dict[str, dict] = {}
    for row, platform_order_no, platform_order_id, label_file_path in rows:
        platform = (row.platform or "unknown").strip() or "unknown"
        group = grouped.setdefault(
            platform,
            {
                "run_id": run_id,
                "platform": platform,
                "document_type_name": _scheduled_task_run_platform_document_type(platform),
                "total_count": 0,
                "pdf_count": 0,
                "print_submitted_count": 0,
                "failed_count": 0,
                "reprintable_count": 0,
                "order_nos": [],
                "printer_names": [],
                "print_job_names": [],
                "messages": [],
                "pdf_file_paths": [],
                "needs_reprint": False,
                "print_submitted": False,
            },
        )
        group["total_count"] += 1
        if row.pdf_generated:
            group["pdf_count"] += 1
        if row.print_submitted:
            group["print_submitted_count"] += 1
            group["print_submitted"] = True
        if row.needs_reprint:
            group["failed_count"] += 1
            group["reprintable_count"] += 1 if row.pdf_file_path else 0
            group["needs_reprint"] = True
        order_no = platform_order_no or platform_order_id
        if not order_no and int(row.order_id or 0) == 0:
            order_no = _scheduled_task_run_platform_document_type(platform)
        _append_unique_text(group["order_nos"], order_no, limit=20)
        _append_unique_text(group["printer_names"], row.printer_name, limit=10)
        _append_unique_text(group["print_job_names"], row.print_job_name, limit=10)
        _append_unique_text(group["messages"], row.error_message or row.print_message, limit=10)
        _append_unique_text(group["pdf_file_paths"], row.pdf_file_path or label_file_path, limit=20)

    return [ScheduledTaskRunPlatformDto(**group) for group in grouped.values()]


def _exchange_rate_dto(row: ExchangeRate) -> ExchangeRateDto:
    return ExchangeRateDto(
        id=row.id,
        rate_date=row.rate_date.isoformat(),
        currency_code=row.currency_code,
        currency_name=row.currency_name or "",
        rate=str(row.rate),
        source_updated_at=_iso(row.source_updated_at),
        synced_at=_iso(row.synced_at),
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _exchange_rate_currency_setting_dto(row: ExchangeRateCurrencySetting) -> ExchangeRateCurrencySettingDto:
    return ExchangeRateCurrencySettingDto(
        id=row.id,
        currency_code=row.currency_code,
        currency_name=row.currency_name or "",
        enabled=bool(row.enabled),
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _catalog_decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value.normalize(), "f")


def _platform_product_catalog_item_dto(row: PlatformProductCatalogItem) -> PlatformProductCatalogItemDto:
    product = row.product
    shop = row.shop
    return PlatformProductCatalogItemDto(
        id=row.id,
        shop_id=row.shop_id,
        shop_name=(shop.display_name or shop.account_id) if shop else "",
        platform=row.platform or "",
        product_id=row.product_id,
        product_code=product.product_code if product else "",
        internal_product_name=product.internal_name if product else "",
        platform_product_id=row.platform_product_id or "",
        platform_sku=row.platform_sku or "",
        product_name=row.product_name or "",
        main_image_url=catalog_main_image_display_url(row.raw_payload),
        listing_status=row.listing_status or "",
        warehouse_code=row.warehouse_code or "",
        warehouse_name=row.warehouse_name or "",
        fulfillment_type=row.fulfillment_type or "",
        logistics_type=row.logistics_type or "",
        available_stock=int(row.available_stock or 0),
        reserved_stock=row.reserved_stock,
        price_amount=_catalog_decimal_text(row.price_amount),
        price_currency=row.price_currency or "CNY",
        exchange_rate=_catalog_decimal_text(row.exchange_rate),
        exchange_rate_date=row.exchange_rate_date.isoformat() if row.exchange_rate_date else None,
        current_price_cny=_catalog_decimal_text(row.current_price_cny),
        cost_cny=_catalog_decimal_text(row.cost_cny),
        commission_rate=_catalog_decimal_text(row.commission_rate),
        shipping_fee_cny=_catalog_decimal_text(row.shipping_fee_cny),
        target_margin_rate=_catalog_decimal_text(row.target_margin_rate),
        current_profit_cny=_catalog_decimal_text(row.current_profit_cny),
        current_margin_rate=_catalog_decimal_text(row.current_margin_rate),
        suggested_price_cny=_catalog_decimal_text(row.suggested_price_cny),
        calculation_status=row.calculation_status or "",
        calculation_message=row.calculation_message or "",
        last_synced_at=_iso(row.last_synced_at),
        calculated_at=_iso(row.calculated_at),
        is_active=bool(row.is_active),
    )


def _platform_product_pricing_rule_dto(row: PlatformProductPricingRule) -> PlatformProductPricingRuleDto:
    return PlatformProductPricingRuleDto(
        id=row.id,
        name=row.name or "",
        platform=row.platform or "",
        shop_id=row.shop_id,
        shop_name=(row.shop.display_name or row.shop.account_id) if row.shop else "",
        product_id=row.product_id,
        product_name=row.product.internal_name if row.product else "",
        warehouse_code=row.warehouse_code or "",
        logistics_type=row.logistics_type or "",
        commission_rate=_catalog_decimal_text(row.commission_rate) or "0",
        base_shipping_fee_cny=_catalog_decimal_text(row.base_shipping_fee_cny) or "0",
        shipping_fee_per_kg_cny=_catalog_decimal_text(row.shipping_fee_per_kg_cny) or "0",
        target_margin_rate=_catalog_decimal_text(row.target_margin_rate) or "0",
        price_increment_cny=_catalog_decimal_text(row.price_increment_cny) or "0.01",
        priority=row.priority,
        enabled=bool(row.enabled),
        remark=row.remark or "",
        updated_at=_iso(row.updated_at),
    )


def _catalog_rule_decimal(value: str | float, field_name: str, *, positive: bool = False) -> Decimal:
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise HTTPException(status_code=400, detail=f"{field_name}必须是数字") from exc
    if result < 0 or (positive and result <= 0):
        raise HTTPException(status_code=400, detail=f"{field_name}不能为负数")
    return result


def _apply_platform_product_pricing_rule(
    row: PlatformProductPricingRule,
    payload: PlatformProductPricingRuleInput,
    db: Session,
    *,
    username: str,
) -> None:
    platform_value = _normalize_catalog_platform(payload.platform or "")
    if not platform_value:
        raise HTTPException(status_code=400, detail="请选择平台")
    shop = db.get(PlatformAccount, payload.shop_id) if payload.shop_id else None
    if payload.shop_id and not shop:
        raise HTTPException(status_code=400, detail="店铺不存在")
    if shop and shop.platform != platform_value:
        raise HTTPException(status_code=400, detail="店铺与规则平台不一致")
    if payload.product_id and not db.get(Product, payload.product_id):
        raise HTTPException(status_code=400, detail="内部产品不存在")
    commission = _catalog_rule_decimal(payload.commission_rate, "佣金率")
    target_margin = _catalog_rule_decimal(payload.target_margin_rate, "目标利润率")
    if commission + target_margin >= Decimal("1"):
        raise HTTPException(status_code=400, detail="佣金率与目标利润率之和必须小于100%")
    row.name = (payload.name or "").strip()
    if not row.name:
        raise HTTPException(status_code=400, detail="规则名称不能为空")
    row.platform = platform_value
    row.shop_id = shop.id if shop else None
    row.product_id = payload.product_id
    row.warehouse_code = (payload.warehouse_code or "").strip()
    row.logistics_type = (payload.logistics_type or "").strip()
    row.commission_rate = commission
    row.base_shipping_fee_cny = _catalog_rule_decimal(payload.base_shipping_fee_cny, "起始运费")
    row.shipping_fee_per_kg_cny = _catalog_rule_decimal(payload.shipping_fee_per_kg_cny, "每公斤运费")
    row.target_margin_rate = target_margin
    row.price_increment_cny = _catalog_rule_decimal(payload.price_increment_cny, "价格取整步长", positive=True)
    row.priority = int(payload.priority or 100)
    row.enabled = bool(payload.enabled)
    row.remark = (payload.remark or "").strip()
    row.created_by = row.created_by or username
    row.updated_at = datetime.utcnow()


def _validate_cron_expr(cron_expr: str) -> str:
    expr = (cron_expr or "").strip()
    if not expr:
        raise HTTPException(status_code=400, detail="Cron 表达式不能为空")
    try:
        CronTrigger.from_crontab(expr, timezone="Asia/Shanghai")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Cron 表达式无效，请使用标准 5 段格式") from exc
    return expr


def _ensure_user_role_link(user: LocalUser, role: Role, db: Session) -> None:
    if db.scalar(select(UserRole.id).where(UserRole.user_id == user.id, UserRole.role_id == role.id).limit(1)):
        return
    db.add(UserRole(user_id=user.id, role_id=role.id))


def _role_has_menu(role: Role, menu_code: str, db: Session) -> bool:
    return bool(
        db.scalar(
            select(RoleMenuPermission.id)
            .where(RoleMenuPermission.role_id == role.id, RoleMenuPermission.menu_code == menu_code)
            .limit(1)
        )
    )


def _ensure_role_menu(role: Role, menu_code: str, db: Session) -> None:
    normalized_code = _normalize_menu_code(menu_code)
    if normalized_code not in MENU_CODES or _role_has_menu(role, normalized_code, db):
        return
    db.add(RoleMenuPermission(role_id=role.id, menu_code=normalized_code))


def _sync_role_menu_permissions(role: Role, db: Session) -> None:
    if role.code == ROLE_ADMIN:
        role.enabled = True
        _set_role_menus(role, ADMIN_MENU_CODES, db)
        return

    if not db.scalar(select(RoleMenuPermission.id).where(RoleMenuPermission.role_id == role.id).limit(1)):
        _set_role_menus(role, _default_menus_for_role(role.code), db)
        return

    if _role_has_menu(role, "order-outbound", db):
        _ensure_role_menu(role, "outbound-scans", db)
        db.query(RoleMenuPermission).filter(
            RoleMenuPermission.role_id == role.id,
            RoleMenuPermission.menu_code == "order-outbound",
        ).delete(synchronize_session=False)
    if _role_has_menu(role, "system-settings", db):
        _ensure_role_menu(role, "scheduled-task-logs", db)
        _ensure_role_menu(role, "exchange-rates", db)
    if _role_has_menu(role, "shops", db):
        _ensure_role_menu(role, "logistics-authorizations", db)
    if _role_has_menu(role, "orders", db):
        _ensure_role_menu(role, "logistics-rules", db)
    if _role_has_menu(role, "users", db):
        _ensure_role_menu(role, "permissions", db)
    if role.code == ROLE_SALES:
        _ensure_role_menu(role, "operations-daily-report", db)
        _ensure_role_menu(role, "traffic-analytics", db)
        _ensure_role_menu(role, "platform-product-catalog", db)
    if _role_has_menu(role, "traffic-analytics", db):
        _ensure_role_menu(role, "traffic-sync-status", db)


def _ensure_roles_and_user_roles(db: Session, admin_username: str) -> Role:
    role_map: dict[str, Role] = {}
    legacy_user_role = db.scalar(select(Role).where(Role.code == ROLE_USER))
    if legacy_user_role and (legacy_user_role.is_system or legacy_user_role.name == "普通用户"):
        db.query(LocalUser).filter(LocalUser.role_id == legacy_user_role.id).update(
            {LocalUser.role_id: None, LocalUser.updated_at: datetime.utcnow()},
            synchronize_session=False,
        )
        db.delete(legacy_user_role)
        db.flush()

    for item in ROLE_DEFINITIONS:
        role = db.scalar(select(Role).where(Role.code == item["code"]))
        if not role:
            role = Role(
                code=item["code"],
                name=item["name"],
                description="",
                is_system=True,
                enabled=True,
            )
            db.add(role)
            db.flush()
            _set_role_menus(role, _default_menus_for_role(role.code), db)
        else:
            if role.code == ROLE_ADMIN and role.name == "管理员":
                role.name = item["name"]
            else:
                role.name = role.name or item["name"]
            role.is_system = True
            _sync_role_menu_permissions(role, db)
        role_map[role.code] = role

    admin_role = role_map[ROLE_ADMIN]
    for user in db.scalars(select(LocalUser)).all():
        if user.username == admin_username:
            user.role_id = admin_role.id
            _ensure_user_role_link(user, admin_role, db)
        elif user.role_id:
            legacy_role = db.get(Role, user.role_id)
            if legacy_role:
                _ensure_user_role_link(user, legacy_role, db)
        user.display_name = user.display_name or user.username
        user.enabled = True if user.enabled is None else user.enabled
        user.updated_at = user.updated_at or datetime.utcnow()
    db.commit()
    return admin_role


def _required_menus_for_request(request: Request) -> set[str]:
    path = request.url.path
    if path.startswith("/api/v1/outbound-scans"):
        if request.method == "POST":
            return {"scan-outbound"}
        if path.startswith("/api/v1/outbound-scans/export"):
            return {"outbound-scans"}
        order_outbound = str(request.query_params.get("order_outbound", "")).lower() in {"1", "true", "yes"}
        return {"outbound-scans"} if order_outbound else {"scan-outbound", "outbound-scans"}
    if request.method == "GET" and path == "/api/v1/system-settings/platform-settings":
        return {"system-settings"}
    if request.method == "GET" and path in {"/api/v1/shops", "/api/shops"}:
        return {"shops"}
    if path in {"/api/v1/traffic-analytics/accounts", "/api/v1/traffic-analytics/sync"}:
        return {"traffic-analytics", "traffic-sync-status"}
    for prefix, menu_codes in API_MENU_RULES:
        if path.startswith(prefix):
            return menu_codes
    return set()


def _ensure_request_menu_access(request: Request, user: LocalUser, db: Session) -> None:
    roles = _roles_for_user(user, db, enabled_only=True)
    if not roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Role disabled or missing")
    if any(role.code == ROLE_ADMIN for role in roles):
        return
    path = request.url.path
    if path in {"/api/v1/auth/me", "/api/auth/me"}:
        return
    required = _required_menus_for_request(request)
    if not required:
        return
    allowed = set(_menu_codes_for_user(user, db))
    if allowed.isdisjoint(required):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No permission for this menu")


_access_dependencies = create_access_dependencies(
    ensure_request_menu_access=_ensure_request_menu_access,
    is_admin_user=_is_admin_user,
)
current_user = _access_dependencies.current_user
current_user_from_scheduled_task_run_pdf_download_token = (
    _access_dependencies.current_user_from_scheduled_task_run_pdf_download_token
)
require_admin = _access_dependencies.require_admin
require_internal_service_token = create_internal_service_dependency()

app.include_router(
    create_api_router(
        current_user_dependency=current_user,
        roles_for_user=_roles_for_user,
        menu_codes_for_user=_menu_codes_for_user,
        admin_role_code=ROLE_ADMIN,
        require_admin_dependency=require_admin,
        menu_definitions=MENU_DEFINITIONS,
        enabled_buyer_users=_enabled_buyer_users,
        user_option_dto=_user_option_dto,
        role_dto=_role_dto,
        roles_by_payload=_roles_by_payload,
        normalize_wecom_mobile=_normalize_wecom_mobile,
        set_user_roles=_set_user_roles,
        user_dto=_user_dto,
        set_role_menus=_set_role_menus,
        hidden_role_code=ROLE_USER,
        reserved_role_codes={ROLE_ADMIN, ROLE_USER, ROLE_PURCHASE, ROLE_SALES, ROLE_CUSTOMER_SERVICE},
        iso_formatter=_iso,
        dashboard_services=DashboardRouteServices(
            platform_setting_items=_dashboard_platform_setting_items,
            is_admin_user=_is_admin_user,
            shop_scope=_dashboard_shop_scope,
            context=_dashboard_context,
            text_datetime=_dashboard_text_datetime,
            int_value=_dashboard_int,
            text_date=_dashboard_text_date,
            mtd_comparison=_dashboard_mtd_comparison,
            last_order_date=_dashboard_last_order_date,
            period=_dashboard_period,
            comparison_period=_dashboard_comparison_period,
            monthly_sales=_dashboard_monthly_sales,
            daily_sales=_dashboard_daily_sales,
            shop_sales=_dashboard_shop_sales,
            risk_buckets=_dashboard_risk_buckets,
            risk_shops=_dashboard_risk_shops,
            operations_daily_report=_operations_daily_report,
            risk_skus=_dashboard_risk_skus,
            hot_skus=_dashboard_hot_skus,
            local_now=_local_now,
        ),
        sync_settings_services=SyncSettingsRouteServices(
            canonical_platform=_canonical_platform,
            platform_lookup_codes=_platform_lookup_codes,
            reload_jobs=reload_jobs,
        ),
    )
)


@app.on_event("startup")
async def startup() -> None:
    Base.metadata.create_all(bind=engine)
    _ensure_legacy_columns()
    _normalize_legacy_fbj_export_statuses()
    _backfill_order_internal_order_no()
    _ensure_order_number_indexes()
    _ensure_dashboard_indexes()
    _ensure_traffic_analytics_indexes()
    _ensure_order_workflow_indexes()
    _backfill_order_operation_logs()
    _repair_joom_orders()
    _repair_wildberries_order_numbers()
    db = next(get_db())
    try:
        interrupted_traffic_runs = mark_interrupted_traffic_runs(db)
        settings = get_settings()
        validate_security_settings(settings)
        admin_role = _ensure_roles_and_user_roles(db, settings.admin_username)
        admin = db.scalar(select(LocalUser).where(LocalUser.username == settings.admin_username))
        if not admin:
            admin = LocalUser(
                username=settings.admin_username,
                password_hash=hash_password(settings.admin_password),
                display_name=settings.admin_username,
                role_id=admin_role.id,
                enabled=True,
            )
            db.add(admin)
        else:
            admin.role_id = admin_role.id
            admin.enabled = True
            admin.display_name = admin.display_name or admin.username
            admin.updated_at = datetime.utcnow()
        db.flush()
        _ensure_user_role_link(admin, admin_role, db)
        seed_default_platform_settings(db)
        seed_default_dashboard_platform_settings(db)
        seed_default_shipping_deadline_settings(db)
        _seed_default_logistics_authorizations(db)
        backfilled_logistics_rule_carrier_codes = _backfill_logistics_rule_carrier_codes(db)
        backfilled_order_logistics_carrier_codes = _backfill_order_logistics_carrier_codes(db)
        db.commit()
        if interrupted_traffic_runs:
            logger.warning("Marked interrupted traffic sync runs as failed: %s", interrupted_traffic_runs)
        if backfilled_logistics_rule_carrier_codes:
            logger.info("Backfilled logistics rule carrier codes: %s", backfilled_logistics_rule_carrier_codes)
        if backfilled_order_logistics_carrier_codes:
            logger.info("Backfilled order logistics carrier codes: %s", backfilled_order_logistics_carrier_codes)
        merged_purchase_items = _merge_duplicate_purchase_order_items(db)
        if merged_purchase_items:
            db.commit()
            logger.info("Merged duplicate purchase order items: %s", merged_purchase_items)
        _backfill_order_items(db)
        changed = backfill_order_dispatch_deadlines(db)
        if changed:
            db.commit()
        print_identity_changed = _backfill_print_setting_printer_identity(db)
        if print_identity_changed:
            db.commit()
            logger.info("Backfilled print setting printer identity: %s", print_identity_changed)
    finally:
        db.close()

    db = next(get_db())
    try:
        await refresh_configs(db)
    except Exception:
        logger.exception("Startup refresh configs failed")
    finally:
        db.close()

    try:
        start_scheduler()
    except Exception:
        logger.exception("Startup scheduler failed")
    try:
        start_order_follow_up_export_worker()
    except Exception:
        logger.exception("Startup order follow up export worker failed")

    try:
        await asyncio.to_thread(process_due_task_retries)
    except Exception:
        logger.exception("Startup scheduled task retry scan failed")


@app.on_event("shutdown")
async def shutdown() -> None:
    try:
        await stop_order_follow_up_export_worker()
    except Exception:
        logger.exception("Shutdown order follow up export worker failed")
    try:
        stop_scheduler()
    except Exception:
        logger.exception("Shutdown scheduler failed")


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "service": "business"}


# Keep the historical import path for internal callers while the HTTP route lives
# in app.api.routes.dashboard.
def dashboard_analytics(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    compare_from: date | None = Query(None),
    compare_to: date | None = Query(None),
    shop_ids: list[int] | None = Query(None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> DashboardAnalyticsResponse:
    scope = _dashboard_shop_scope(db, shop_ids)
    summary, start_day, end_day, previous_start, previous_end = _dashboard_context(
        db, date_from, date_to, compare_from, compare_to, scope
    )
    return DashboardAnalyticsResponse(
        generated_at=_dashboard_text_datetime(_local_now()) or "",
        total_orders=_dashboard_int(summary.total_orders),
        first_order_date=_dashboard_text_date(summary.first_order_date),
        last_order_date=_dashboard_text_date(summary.last_order_date),
        blank_currency_orders=_dashboard_int(summary.blank_currency_orders),
        monthly_sales=_dashboard_monthly_sales(db, start_day, end_day, scope),
        daily_sales=_dashboard_daily_sales(db, start_day, end_day, scope),
        comparison_daily_sales=_dashboard_daily_sales(db, previous_start, previous_end, scope),
        shop_sales=_dashboard_shop_sales(db, start_day, end_day, scope),
        current_label=f"{start_day.isoformat()}~{end_day.isoformat()}",
        comparison_label=f"{previous_start.isoformat()}~{previous_end.isoformat()}",
        mtd_comparison=_dashboard_mtd_comparison(db, start_day, end_day, previous_start, previous_end, scope),
        risk_buckets=_dashboard_risk_buckets(db, scope),
        risk_shops=_dashboard_risk_shops(db, scope),
        risk_skus=_dashboard_risk_skus(db, scope),
        hot_skus=_dashboard_hot_skus(db, start_day, end_day, previous_start, previous_end, scope),
    )


@app.get("/api/v1/system-settings/platform-settings", response_model=list[PlatformSettingDto])
def list_platform_settings(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> list[PlatformSettingDto]:
    seed_default_platform_settings(db)
    db.commit()
    rows = db.scalars(
        select(PlatformSetting).order_by(asc(PlatformSetting.sort_order), asc(PlatformSetting.platform))
    ).all()
    return [_platform_setting_dto(row) for row in rows]


@app.put("/api/v1/system-settings/platform-settings/{platform}", response_model=PlatformSettingDto)
def update_platform_setting_enabled(
    platform: str,
    payload: PlatformSettingToggleRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformSettingDto:
    platform_code = _canonical_platform(platform)
    if not platform_code:
        raise HTTPException(status_code=400, detail="平台不能为空")
    seed_default_platform_settings(db)
    row = db.scalar(select(PlatformSetting).where(PlatformSetting.platform == platform_code))
    if not row:
        raise HTTPException(status_code=404, detail="平台不存在")
    row.enabled = bool(payload.enabled)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _platform_setting_dto(row)


@app.get("/api/v1/system-settings/print-settings", response_model=list[PlatformPrintSettingDto])
def list_platform_print_settings(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> list[PlatformPrintSettingDto]:
    rows = db.scalars(select(PlatformPrintSetting).order_by(asc(PlatformPrintSetting.platform), asc(PlatformPrintSetting.document_type))).all()
    return [_platform_print_setting_dto(row) for row in rows]


@app.get("/api/v1/system-settings/printers", response_model=list[PrinterDto])
def list_system_printers(_: LocalUser = Depends(current_user)) -> list[PrinterDto]:
    """Return printers visible to the backend service host."""
    return _list_server_printers()


@app.post("/api/v1/system-settings/printer-monitor/check", response_model=PrinterMonitorResultDto)
def check_printer_monitor_status(
    payload: PrinterMonitorRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PrinterMonitorResultDto:
    printer_name = (payload.printer_name or "").strip()
    if not printer_name:
        raise HTTPException(status_code=400, detail="打印机名称不能为空")
    task = db.get(ScheduledTask, payload.scheduled_task_id) if payload.scheduled_task_id else None
    result = monitor_printer_status(
        db,
        printer_name,
        task=task,
        recipients=payload.recipients,
        auto_recover=payload.auto_recover,
        max_retries=payload.max_retries,
    )
    return PrinterMonitorResultDto(**result)


@app.post("/api/v1/system-settings/print-settings", response_model=PlatformPrintSettingDto)
def create_platform_print_setting(
    payload: PlatformPrintSettingUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformPrintSettingDto:
    platform = (payload.platform or "").strip()
    if not platform:
        raise HTTPException(status_code=400, detail="平台不能为空")
    document_type = (payload.document_type or PRINT_DOCUMENT_TYPE_LABEL).strip() or PRINT_DOCUMENT_TYPE_LABEL
    if document_type not in PRINT_DOCUMENT_TYPES:
        raise HTTPException(status_code=400, detail="单据类型无效")
    if not is_valid_print_orientation(payload.page_orientation):
        raise HTTPException(status_code=400, detail="打印方向无效")
    exists_row = db.scalar(select(PlatformPrintSetting).where(PlatformPrintSetting.platform == platform, PlatformPrintSetting.document_type == document_type))
    if exists_row:
        raise HTTPException(status_code=400, detail="该平台单据类型打印设置已存在")
    printer = _validated_server_printer(payload.printer_name)
    row = PlatformPrintSetting(
        platform=platform,
        document_type=document_type,
        page_orientation=normalize_print_orientation(payload.page_orientation),
        enabled=payload.enabled,
        remark=(payload.remark or "").strip(),
    )
    _apply_printer_identity(row, printer)
    db.add(row)
    db.commit()
    db.refresh(row)
    return _platform_print_setting_dto(row)


@app.put("/api/v1/system-settings/print-settings/{setting_id:int}", response_model=PlatformPrintSettingDto)
def update_platform_print_setting(
    setting_id: int,
    payload: PlatformPrintSettingUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformPrintSettingDto:
    row = db.get(PlatformPrintSetting, setting_id)
    if not row:
        raise HTTPException(status_code=404, detail="打印设置不存在")
    platform = (payload.platform or "").strip()
    if not platform:
        raise HTTPException(status_code=400, detail="平台不能为空")
    document_type = (payload.document_type or PRINT_DOCUMENT_TYPE_LABEL).strip() or PRINT_DOCUMENT_TYPE_LABEL
    if document_type not in PRINT_DOCUMENT_TYPES:
        raise HTTPException(status_code=400, detail="单据类型无效")
    if not is_valid_print_orientation(payload.page_orientation):
        raise HTTPException(status_code=400, detail="打印方向无效")
    duplicate = db.scalar(
        select(PlatformPrintSetting).where(
            PlatformPrintSetting.platform == platform,
            PlatformPrintSetting.document_type == document_type,
            PlatformPrintSetting.id != setting_id,
        )
    )
    if duplicate:
        raise HTTPException(status_code=400, detail="该平台单据类型打印设置已存在")
    printer = _validated_server_printer(payload.printer_name)
    row.platform = platform
    row.document_type = document_type
    _apply_printer_identity(row, printer)
    row.page_orientation = normalize_print_orientation(payload.page_orientation)
    row.enabled = payload.enabled
    row.remark = (payload.remark or "").strip()
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _platform_print_setting_dto(row)


@app.get("/api/v1/system-settings/shipping-deadline-settings", response_model=list[ShippingDeadlineSettingDto])
def list_shipping_deadline_settings(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ShippingDeadlineSettingDto]:
    seed_default_shipping_deadline_settings(db)
    db.commit()
    rows = db.scalars(
        select(ShippingDeadlineSetting).where(ShippingDeadlineSetting.enabled == True).order_by(
            asc(ShippingDeadlineSetting.sort_order), asc(ShippingDeadlineSetting.platform)
        )
    ).all()
    return [_shipping_deadline_setting_dto(row) for row in rows]


@app.put("/api/v1/system-settings/shipping-deadline-settings", response_model=ShippingDeadlineSettingsUpdateResponse)
def update_shipping_deadline_settings(
    payload: ShippingDeadlineSettingsUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ShippingDeadlineSettingsUpdateResponse:
    if not payload.items:
        raise HTTPException(status_code=400, detail="请至少保留一条发货截止时间规则")

    normalized_items: dict[str, tuple[str, int, int]] = {}
    for index, item in enumerate(payload.items):
        platform = canonical_deadline_platform(item.platform)
        if not platform:
            raise HTTPException(status_code=400, detail="平台不能为空")
        if platform in normalized_items:
            raise HTTPException(status_code=400, detail=f"平台 {platform} 的规则重复")
        try:
            base_date_field = normalize_base_date_field(item.base_date_field)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        normalized_items[platform] = (base_date_field, int(item.offset_days or 0), index)

    existing = {
        row.platform: row
        for row in db.scalars(select(ShippingDeadlineSetting)).all()
    }
    now = datetime.utcnow()
    for platform, (base_date_field, offset_days, sort_order) in normalized_items.items():
        row = existing.get(platform)
        if not row:
            row = ShippingDeadlineSetting(platform=platform, created_at=now)
            db.add(row)
        row.base_date_field = base_date_field
        row.offset_days = offset_days
        row.sort_order = sort_order
        row.enabled = True
        row.updated_at = now

    for platform, row in existing.items():
        if platform not in normalized_items:
            row.enabled = False
            row.updated_at = now

    backfilled = backfill_order_dispatch_deadlines(db)
    db.commit()
    rows = db.scalars(
        select(ShippingDeadlineSetting).where(ShippingDeadlineSetting.enabled == True).order_by(
            asc(ShippingDeadlineSetting.sort_order), asc(ShippingDeadlineSetting.platform)
        )
    ).all()
    return ShippingDeadlineSettingsUpdateResponse(
        items=[_shipping_deadline_setting_dto(row) for row in rows],
        backfilled=backfilled,
    )


@app.get("/api/v1/system-settings/email-smtp", response_model=EmailSmtpSettingDto)
def get_email_smtp_setting(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> EmailSmtpSettingDto:
    return _email_smtp_setting_dto(get_email_setting(db))


@app.get("/api/v1/system-settings/email-providers", response_model=list[EmailProviderDto])
def list_email_providers(_: LocalUser = Depends(current_user)) -> list[EmailProviderDto]:
    return [EmailProviderDto(**item) for item in list_email_provider_presets()]


@app.get("/api/v1/system-settings/exchange-rates", response_model=ExchangeRateListResponse)
def list_exchange_rates(
    rate_date: str | None = Query(None),
    currency_code: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ExchangeRateListResponse:
    stmt = select(ExchangeRate)
    count_stmt = select(func.count(ExchangeRate.id))
    if rate_date:
        try:
            parsed_date = date.fromisoformat(rate_date)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="rate_date 必须是 YYYY-MM-DD") from exc
        stmt = stmt.where(ExchangeRate.rate_date == parsed_date)
        count_stmt = count_stmt.where(ExchangeRate.rate_date == parsed_date)
    if currency_code:
        normalized = currency_code.strip().upper()
        if normalized:
            stmt = stmt.where(ExchangeRate.currency_code == normalized)
            count_stmt = count_stmt.where(ExchangeRate.currency_code == normalized)

    total = db.scalar(count_stmt) or 0
    rows = db.scalars(
        stmt.order_by(ExchangeRate.rate_date.desc(), ExchangeRate.currency_code.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    currencies = db.scalars(
        select(ExchangeRate.currency_code).distinct().order_by(ExchangeRate.currency_code.asc())
    ).all()
    return ExchangeRateListResponse(
        items=[_exchange_rate_dto(row) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
        currencies=currencies,
    )


@app.get("/api/v1/system-settings/exchange-rate-currency-settings", response_model=list[ExchangeRateCurrencySettingDto])
def list_exchange_rate_currency_settings(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ExchangeRateCurrencySettingDto]:
    rows = db.scalars(
        select(ExchangeRateCurrencySetting)
        .where(ExchangeRateCurrencySetting.enabled == True)
        .order_by(ExchangeRateCurrencySetting.currency_code.asc())
    ).all()
    return [_exchange_rate_currency_setting_dto(row) for row in rows]


@app.put("/api/v1/system-settings/exchange-rate-currency-settings", response_model=list[ExchangeRateCurrencySettingDto])
def update_exchange_rate_currency_settings(
    payload: ExchangeRateCurrencySettingUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ExchangeRateCurrencySettingDto]:
    currencies: list[tuple[str, str]] = []
    seen: set[str] = set()
    for item in payload.currencies or []:
        code = str(item.currency_code or "").strip().upper()
        if not code or code == "CNY" or code in seen:
            continue
        seen.add(code)
        currencies.append((code, str(item.currency_name or "").strip()))

    existing = {
        row.currency_code: row
        for row in db.scalars(select(ExchangeRateCurrencySetting)).all()
    }
    for row in existing.values():
        row.enabled = False
        row.updated_at = datetime.utcnow()

    now = datetime.utcnow()
    for code, currency_name in currencies:
        row = existing.get(code)
        latest_rate = None
        if not currency_name or not row:
            latest_rate = db.scalar(
                select(ExchangeRate)
                .where(ExchangeRate.currency_code == code)
                .order_by(ExchangeRate.rate_date.desc(), ExchangeRate.updated_at.desc())
                .limit(1)
            )
        if not row:
            row = ExchangeRateCurrencySetting(currency_code=code)
        row.currency_name = currency_name or (latest_rate.currency_name if latest_rate else row.currency_name or "")
        row.enabled = True
        row.updated_at = now
        db.add(row)

    db.commit()
    rows = db.scalars(
        select(ExchangeRateCurrencySetting)
        .where(ExchangeRateCurrencySetting.enabled == True)
        .order_by(ExchangeRateCurrencySetting.currency_code.asc())
    ).all()
    return [_exchange_rate_currency_setting_dto(row) for row in rows]


@app.post("/api/v1/system-settings/exchange-rates/sync", response_model=ExchangeRateSyncResult)
async def sync_exchange_rates_now(_: LocalUser = Depends(current_user)) -> ExchangeRateSyncResult:
    result = await sync_exchange_rates_from_provider(replace_existing=True)
    return ExchangeRateSyncResult(**result)


@app.get("/api/v1/platform-product-catalog/images/{filename}")
def get_platform_product_catalog_image_file(filename: str):
    try:
        path = catalog_main_image_file_path(filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="图片文件名无效") from exc
    if not path.is_file():
        raise HTTPException(status_code=404, detail="图片不存在")
    return FileResponse(
        path,
        media_type=CATALOG_MAIN_IMAGE_CONTENT_TYPES.get(path.suffix.lower(), "application/octet-stream"),
        filename=path.name,
    )


@app.get("/api/v1/platform-product-catalog", response_model=PlatformProductCatalogListResponse)
def list_platform_product_catalog(
    platform: str | None = Query(None),
    shop_id: int | None = Query(None),
    keyword: str | None = Query(None),
    calculation_status: str | None = Query(None),
    mapped: bool | None = Query(None),
    include_inactive: bool = Query(False),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductCatalogListResponse:
    filters = []
    if platform and platform.strip():
        filters.append(PlatformProductCatalogItem.platform == platform.strip().lower())
    if shop_id:
        filters.append(PlatformProductCatalogItem.shop_id == shop_id)
    if not include_inactive:
        filters.append(PlatformProductCatalogItem.is_active == True)
    if calculation_status and calculation_status.strip():
        filters.append(PlatformProductCatalogItem.calculation_status == calculation_status.strip())
    if mapped is True:
        filters.append(PlatformProductCatalogItem.product_id.is_not(None))
    elif mapped is False:
        filters.append(PlatformProductCatalogItem.product_id.is_(None))
    if keyword and keyword.strip():
        text_value = f"%{keyword.strip()}%"
        filters.append(
            or_(
                PlatformProductCatalogItem.product_name.ilike(text_value),
                PlatformProductCatalogItem.platform_sku.ilike(text_value),
                PlatformProductCatalogItem.platform_product_id.ilike(text_value),
                Product.internal_name.ilike(text_value),
                Product.product_code.ilike(text_value),
            )
        )

    stmt = select(PlatformProductCatalogItem).outerjoin(Product, Product.id == PlatformProductCatalogItem.product_id)
    count_stmt = select(func.count(PlatformProductCatalogItem.id)).outerjoin(Product, Product.id == PlatformProductCatalogItem.product_id)
    if filters:
        stmt = stmt.where(*filters)
        count_stmt = count_stmt.where(*filters)
    total = int(db.scalar(count_stmt) or 0)
    rows = db.scalars(
        stmt.options(
            joinedload(PlatformProductCatalogItem.shop),
            joinedload(PlatformProductCatalogItem.product),
        )
        .order_by(desc(PlatformProductCatalogItem.last_synced_at), asc(PlatformProductCatalogItem.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    summary_rows = db.execute(
        select(PlatformProductCatalogItem.calculation_status, func.count(PlatformProductCatalogItem.id))
        .outerjoin(Product, Product.id == PlatformProductCatalogItem.product_id)
        .where(*filters)
        .group_by(PlatformProductCatalogItem.calculation_status)
    ).all()
    summary = {str(status or "unknown"): int(count or 0) for status, count in summary_rows}
    summary["total"] = total
    return PlatformProductCatalogListResponse(
        items=[_platform_product_catalog_item_dto(row) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
        summary=summary,
    )


@app.get("/api/v1/platform-product-catalog/options", response_model=PlatformProductCatalogOptionsResponse)
def platform_product_catalog_options(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductCatalogOptionsResponse:
    shops = db.scalars(
        select(PlatformAccount)
        .where(PlatformAccount.enabled == True, PlatformAccount.platform.in_(CATALOG_SUPPORTED_PLATFORMS))
        .order_by(asc(PlatformAccount.display_name), asc(PlatformAccount.id))
    ).all()
    products = db.scalars(select(Product).where(Product.enabled == True).order_by(asc(Product.product_code), asc(Product.id))).all()
    return PlatformProductCatalogOptionsResponse(
        shops=[
            {
                "id": row.id,
                "platform": row.platform,
                "label": row.display_name or row.account_id,
                "account_id": row.account_id,
            }
            for row in shops
        ],
        products=[{"id": row.id, "product_code": row.product_code, "internal_name": row.internal_name} for row in products],
    )


@app.post("/api/v1/platform-product-catalog/sync", response_model=PlatformProductCatalogSyncResult)
async def sync_platform_product_catalog_now(
    payload: PlatformProductCatalogSyncRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductCatalogSyncResult:
    result = await synchronize_platform_catalog(db, shop_ids=payload.shop_ids or None, mode=payload.mode)
    return PlatformProductCatalogSyncResult(**result)


@app.post("/api/v1/platform-product-catalog/recalculate", response_model=PlatformProductCatalogRecalculateResult)
def recalculate_platform_product_catalog(
    payload: PlatformProductCatalogRecalculateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductCatalogRecalculateResult:
    return PlatformProductCatalogRecalculateResult(recalculated=recalculate_catalog(db, item_ids=payload.item_ids or None))


@app.put("/api/v1/platform-product-catalog/items/{item_id:int}/mapping", response_model=PlatformProductCatalogItemDto)
def update_platform_product_catalog_mapping(
    item_id: int,
    payload: PlatformProductCatalogMappingRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductCatalogItemDto:
    item = db.get(PlatformProductCatalogItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="平台商品不存在")
    try:
        row = map_catalog_item(db, item, payload.product_id, username=user.username)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    db.refresh(row)
    return _platform_product_catalog_item_dto(row)


@app.get("/api/v1/platform-product-catalog/rules", response_model=list[PlatformProductPricingRuleDto])
def list_platform_product_pricing_rules(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[PlatformProductPricingRuleDto]:
    rows = db.scalars(
        select(PlatformProductPricingRule)
        .options(
            joinedload(PlatformProductPricingRule.shop),
            joinedload(PlatformProductPricingRule.product),
        )
        .order_by(asc(PlatformProductPricingRule.platform), asc(PlatformProductPricingRule.priority), asc(PlatformProductPricingRule.id))
    ).all()
    return [_platform_product_pricing_rule_dto(row) for row in rows]


@app.post("/api/v1/platform-product-catalog/rules", response_model=PlatformProductPricingRuleDto)
def create_platform_product_pricing_rule(
    payload: PlatformProductPricingRuleInput,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductPricingRuleDto:
    row = PlatformProductPricingRule(created_by=user.username)
    _apply_platform_product_pricing_rule(row, payload, db, username=user.username)
    db.add(row)
    db.commit()
    db.refresh(row)
    return _platform_product_pricing_rule_dto(row)


@app.put("/api/v1/platform-product-catalog/rules/{rule_id:int}", response_model=PlatformProductPricingRuleDto)
def update_platform_product_pricing_rule(
    rule_id: int,
    payload: PlatformProductPricingRuleInput,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> PlatformProductPricingRuleDto:
    row = db.get(PlatformProductPricingRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="费用规则不存在")
    _apply_platform_product_pricing_rule(row, payload, db, username=user.username)
    db.commit()
    db.refresh(row)
    return _platform_product_pricing_rule_dto(row)


@app.delete("/api/v1/platform-product-catalog/rules/{rule_id:int}", status_code=status.HTTP_204_NO_CONTENT)
def delete_platform_product_pricing_rule(
    rule_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> Response:
    row = db.get(PlatformProductPricingRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="费用规则不存在")
    db.delete(row)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.put("/api/v1/system-settings/email-smtp", response_model=EmailSmtpSettingDto)
def update_email_smtp_setting(
    payload: EmailSmtpSettingUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> EmailSmtpSettingDto:
    row = get_email_setting(db)
    provider_codes = {item["code"] for item in list_email_provider_presets()}
    provider = (payload.provider or "qq").strip().lower()
    if provider not in provider_codes:
        provider = "qq"
    smtp_host, smtp_port, use_ssl = apply_provider_preset(provider, payload.smtp_host, payload.smtp_port, payload.use_ssl)
    if provider == "custom" and not smtp_host:
        raise HTTPException(status_code=400, detail="自定义 SMTP 需要填写 SMTP 主机")
    row.provider = provider
    row.enabled = bool(payload.enabled)
    row.smtp_host = smtp_host
    row.smtp_port = smtp_port
    row.use_ssl = use_ssl
    row.sender_email = (payload.sender_email or "").strip()
    row.sender_name = (payload.sender_name or "").strip()
    row.notification_recipients = {
        "wanbang_tracking_failure": (payload.notification_recipients.wanbang_tracking_failure or "").strip(),
        "bsi_address_anomaly": (payload.notification_recipients.bsi_address_anomaly or "").strip(),
    }
    if payload.auth_code is not None and payload.auth_code.strip():
        row.encrypted_auth_code = encrypt_auth_code(payload.auth_code)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _email_smtp_setting_dto(row)


@app.post("/api/v1/system-settings/email-smtp/test", response_model=EmailSmtpSettingDto)
def test_email_smtp_setting(
    payload: EmailTestRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> EmailSmtpSettingDto:
    row = get_email_setting(db)
    recipients = parse_recipients(payload.recipient or "")
    try:
        send_email(
            row,
            recipients,
            "CaifuClaw AI 邮件通知测试",
            "这是一封 CaifuClaw AI 邮件通知测试邮件。收到此邮件说明 SMTP 发件配置可用。",
        )
        row.last_test_status = "success"
        row.last_test_message = "测试邮件发送成功"
    except Exception as exc:
        row.last_test_status = "failed"
        row.last_test_message = str(exc) or "测试邮件发送失败"
    row.last_test_at = datetime.utcnow()
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _email_smtp_setting_dto(row)


@app.get("/api/v1/system-settings/wecom-robot", response_model=WeComRobotSettingDto)
def get_wecom_robot_setting_api(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> WeComRobotSettingDto:
    return _wecom_robot_setting_dto(get_wecom_robot_setting(db))


@app.get("/api/v1/system-settings/wecom-robot/mention-users", response_model=list[WeComMentionUserOptionDto])
def list_wecom_mention_user_options(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[WeComMentionUserOptionDto]:
    rows = db.scalars(
        select(LocalUser)
        .where(LocalUser.enabled == True)
        .order_by(asc(LocalUser.id))
    ).all()
    return [_wecom_mention_user_option_dto(row) for row in rows]


@app.put("/api/v1/system-settings/wecom-robot", response_model=WeComRobotSettingDto)
def update_wecom_robot_setting_api(
    payload: WeComRobotSettingUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> WeComRobotSettingDto:
    row = get_wecom_robot_setting(db)
    data = _validate_wecom_robot_payload(payload, row)
    row.timeout_seconds = data["timeout_seconds"]
    row.max_retries = data["max_retries"]
    row.rate_limit_per_minute = data["rate_limit_per_minute"]
    row.default_mentioned_user_ids = dumps_int_list(data["default_mentioned_user_ids"])
    row.default_mentioned_list = dumps_string_list(data["default_mentioned_list"])
    row.default_mentioned_mobile_list = dumps_string_list(
        mentioned_mobile_list_from_user_ids(db, data["default_mentioned_user_ids"])
        or data["default_mentioned_mobile_list"]
    )
    row.default_prompt = data["default_prompt"]
    row.purchase_order_notify_enabled = data["purchase_order_notify_enabled"]
    if data["webhook_url"]:
        row.encrypted_webhook_url = encrypt_wecom_webhook_url(data["webhook_url"])
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _wecom_robot_setting_dto(row)


@app.post("/api/v1/system-settings/wecom-robot/test", response_model=WeComRobotTestResponse)
def test_wecom_robot_setting_api(
    payload: WeComRobotTestRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> WeComRobotTestResponse:
    try:
        result = send_wecom_robot_test_message(db, payload.content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc) or "企业微信机器人测试失败") from exc
    if result.get("status") in {"mentioned", "skipped"}:
        message = str(result.get("message") or "默认提示语为空，未发送企业微信测试消息")
        return WeComRobotTestResponse(status=str(result.get("status")), message=message)
    return WeComRobotTestResponse(status="success", message="企业微信测试消息已发送")


@app.get("/api/v1/system-settings/translation-provider-options", response_model=list[TranslationProviderOptionDto])
def list_translation_provider_options_api(_: LocalUser = Depends(current_user)) -> list[TranslationProviderOptionDto]:
    return [
        TranslationProviderOptionDto(code=str(item["code"]), name=str(item["name"]))
        for item in list_translation_provider_presets()
    ]


@app.get("/api/v1/system-settings/translation-language-options", response_model=list[TranslationLanguageOptionDto])
def list_translation_language_options_api(_: LocalUser = Depends(current_user)) -> list[TranslationLanguageOptionDto]:
    return [
        TranslationLanguageOptionDto(code=str(item["code"]), label=str(item["label"]))
        for item in list_translation_language_presets()
    ]


@app.get("/api/v1/system-settings/translation-provider", response_model=TranslationProviderSettingDto)
def get_translation_provider_setting_api(
    provider: str = Query(DEFAULT_TRANSLATION_PROVIDER),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> TranslationProviderSettingDto:
    try:
        row = get_translation_provider_setting(db, provider)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return _translation_provider_setting_dto(row)


@app.put("/api/v1/system-settings/translation-provider", response_model=TranslationProviderSettingDto)
def update_translation_provider_setting_api(
    payload: TranslationProviderSettingUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> TranslationProviderSettingDto:
    try:
        provider = normalize_translation_provider(payload.provider)
        row = get_translation_provider_setting(db, provider)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    data = _validate_translation_provider_payload(payload, row)
    row.provider = data["provider"]
    row.provider_name = data["provider_name"]
    row.enabled = data["enabled"]
    row.app_id = data["app_id"]
    row.endpoint = data["endpoint"]
    row.source_language = data["source_language"]
    row.timeout_seconds = data["timeout_seconds"]
    row.max_retries = data["max_retries"]
    row.batch_size = data["batch_size"]
    row.batch_chars = data["batch_chars"]
    row.provider_options_json = data["provider_options_json"]
    if data["secret_key"]:
        row.encrypted_secret_key = encrypt_translation_secret_key(data["secret_key"])
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _translation_provider_setting_dto(row)


@app.post("/api/v1/system-settings/translation-provider/test", response_model=TranslationProviderTestResponse)
def test_translation_provider_setting_api(
    payload: TranslationProviderTestRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> TranslationProviderTestResponse:
    text_value = str(payload.text or "").strip()
    if not text_value:
        raise HTTPException(status_code=422, detail="text不能为空")
    target_language = str(payload.target_language or "").strip()
    if not target_language:
        raise HTTPException(status_code=422, detail="target_language不能为空")
    row: TranslationProviderSetting | None = None
    try:
        row = get_translation_provider_setting(db, payload.provider)
        client = build_translation_client_from_setting(row)
        translated = client.translate_texts(
            [text_value],
            from_lang=(row.source_language or "auto").strip() or "auto",
            to_lang=target_language,
        )
        translated_text = str(translated.get(text_value) or "").strip()
        if not translated_text:
            raise RuntimeError("翻译服务未返回结果")
    except Exception as exc:
        if row is not None:
            try:
                row.last_test_at = datetime.utcnow()
                row.last_test_status = "error"
                row.last_test_message = str(exc) or "翻译测试失败"
                row.updated_at = datetime.utcnow()
                db.commit()
            except Exception:
                db.rollback()
        raise HTTPException(status_code=400, detail=str(exc) or "翻译测试失败") from exc
    row.last_test_at = datetime.utcnow()
    row.last_test_status = "success"
    row.last_test_message = translated_text
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return TranslationProviderTestResponse(status="success", message="翻译测试成功", translated_text=translated_text)


@app.post("/api/v1/ai-translation/translate", response_model=TextTranslationResponse)
def translate_text_api(
    payload: TextTranslationRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> TextTranslationResponse:
    return _translate_text_once(payload, user=user, db=db)


@app.get("/api/v1/system-settings/model-endpoints", response_model=list[ModelEndpointDto])
def list_model_endpoints(
    enabled_only: bool = Query(False),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ModelEndpointDto]:
    query = select(ModelEndpoint)
    if enabled_only:
        query = query.where(ModelEndpoint.enabled == True)
    rows = db.scalars(query.order_by(desc(ModelEndpoint.updated_at), desc(ModelEndpoint.id))).all()
    return [_model_endpoint_dto(row) for row in rows]


@app.post("/api/v1/system-settings/model-endpoints", response_model=ModelEndpointDto)
def create_model_endpoint(
    payload: ModelEndpointUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ModelEndpointDto:
    data = _validate_model_endpoint_payload(payload)
    _ensure_unique_model_endpoint_name(db, data["name"])
    row = ModelEndpoint(
        name=data["name"],
        base_url=data["base_url"],
        encrypted_api_key=_encrypt_model_api_key(data["api_key"]),
        enabled=data["enabled"],
        remark=data["remark"],
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _model_endpoint_dto(row)


@app.put("/api/v1/system-settings/model-endpoints/{endpoint_id:int}", response_model=ModelEndpointDto)
def update_model_endpoint(
    endpoint_id: int,
    payload: ModelEndpointUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ModelEndpointDto:
    row = db.get(ModelEndpoint, endpoint_id)
    if not row:
        raise HTTPException(status_code=404, detail="接口配置不存在")
    data = _validate_model_endpoint_payload(payload, current=row)
    _ensure_unique_model_endpoint_name(db, data["name"], current_id=endpoint_id)
    row.name = data["name"]
    row.base_url = data["base_url"]
    row.enabled = data["enabled"]
    row.remark = data["remark"]
    if data["api_key"]:
        row.encrypted_api_key = _encrypt_model_api_key(data["api_key"])
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _model_endpoint_dto(row)


@app.delete("/api/v1/system-settings/model-endpoints/{endpoint_id:int}")
def delete_model_endpoint(
    endpoint_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> dict:
    row = db.get(ModelEndpoint, endpoint_id)
    if not row:
        raise HTTPException(status_code=404, detail="接口配置不存在")
    in_use = db.scalar(select(ModelSetting.id).where(ModelSetting.endpoint_id == endpoint_id).limit(1))
    if in_use:
        raise HTTPException(status_code=409, detail="接口配置已被模型使用，不能删除")
    db.delete(row)
    db.commit()
    return {"status": "ok"}


@app.get("/api/v1/system-settings/model-settings", response_model=list[ModelSettingDto])
def list_model_settings(
    enabled_only: bool = Query(False),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ModelSettingDto]:
    query = select(ModelSetting).options(selectinload(ModelSetting.endpoint))
    if enabled_only:
        query = query.where(ModelSetting.enabled == True)
    rows = db.scalars(
        query.order_by(desc(ModelSetting.is_default), desc(ModelSetting.updated_at), desc(ModelSetting.id))
    ).all()
    return [_model_setting_dto(row) for row in rows]


@app.post("/api/v1/system-settings/model-settings", response_model=ModelSettingDto)
def create_model_setting(
    payload: ModelSettingUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ModelSettingDto:
    data, _endpoint = _validate_model_setting_payload(payload, db)
    _ensure_unique_model_setting_name(db, data["name"])
    if data["is_default"]:
        _clear_default_model_settings(db)
    row = ModelSetting(**data)
    db.add(row)
    db.commit()
    db.refresh(row)
    db.refresh(row, attribute_names=["endpoint"])
    return _model_setting_dto(row)


@app.put("/api/v1/system-settings/model-settings/{setting_id:int}", response_model=ModelSettingDto)
def update_model_setting(
    setting_id: int,
    payload: ModelSettingUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ModelSettingDto:
    row = db.get(ModelSetting, setting_id)
    if not row:
        raise HTTPException(status_code=404, detail="模型设置不存在")
    data, _endpoint = _validate_model_setting_payload(payload, db)
    _ensure_unique_model_setting_name(db, data["name"], current_id=setting_id)
    row.name = data["name"]
    row.model = data["model"]
    row.endpoint_id = data["endpoint_id"]
    row.enabled = data["enabled"]
    row.is_default = data["is_default"]
    row.supports_vision = data["supports_vision"]
    if row.is_default:
        _clear_default_model_settings(db, current_id=setting_id)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    db.refresh(row, attribute_names=["endpoint"])
    return _model_setting_dto(row)


@app.post("/api/v1/system-settings/model-settings/{setting_id:int}/test", response_model=ModelConnectionTestResponse)
def test_model_setting_connection(
    setting_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ModelConnectionTestResponse:
    setting = db.scalar(
        select(ModelSetting)
        .options(selectinload(ModelSetting.endpoint))
        .where(ModelSetting.id == setting_id)
    )
    if not setting:
        raise HTTPException(status_code=404, detail="模型设置不存在")
    endpoint = setting.endpoint
    if not endpoint:
        raise HTTPException(status_code=422, detail="接口配置不存在")
    if not setting.enabled:
        raise HTTPException(status_code=422, detail="模型已禁用")
    if not endpoint.enabled:
        raise HTTPException(status_code=422, detail="接口配置已禁用")
    upstream_url = _model_chat_completions_url(endpoint.base_url)
    if not upstream_url:
        raise HTTPException(status_code=422, detail="base url不能为空")
    started_at = datetime.utcnow()
    _post_model_chat_completion(
        setting,
        [{"role": "user", "content": "Please reply with OK."}],
        max_tokens=16,
        temperature=0,
        timeout=30,
    )
    duration_ms = int((datetime.utcnow() - started_at).total_seconds() * 1000)
    return ModelConnectionTestResponse(
        model_setting_id=setting.id,
        model_setting_name=setting.name,
        model=setting.model,
        endpoint_name=endpoint.name,
        upstream_url=upstream_url,
        duration_ms=duration_ms,
        message="连接正常",
    )


@app.delete("/api/v1/system-settings/model-settings/{setting_id:int}")
def delete_model_setting(
    setting_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> dict:
    row = db.get(ModelSetting, setting_id)
    if not row:
        raise HTTPException(status_code=404, detail="模型设置不存在")
    db.delete(row)
    db.commit()
    return {"status": "ok"}


@app.get("/api/v1/ai-image/download")
def download_ai_image(
    object_key: str = Query(...),
    filename: str = Query(""),
    _: LocalUser = Depends(current_user),
) -> StreamingResponse:
    safe_object_key = _validated_ai_image_object_key(object_key)
    result = _open_ai_image_oss_object(safe_object_key)
    object_headers = getattr(result, "headers", {}) or {}
    content_type = (
        object_headers.get("Content-Type")
        or object_headers.get("content-type")
        or "application/octet-stream"
    )
    content_length = object_headers.get("Content-Length") or object_headers.get("content-length")
    headers = {
        "Content-Disposition": _content_disposition_attachment(
            _ai_image_download_filename(filename, safe_object_key)
        ),
        "Cache-Control": "private, no-store",
    }
    if content_length:
        headers["Content-Length"] = str(content_length)
    return StreamingResponse(
        _iter_ai_image_oss_object(result),
        media_type=str(content_type),
        headers=headers,
    )


@app.post("/api/v1/ai-image/download-batch")
def download_ai_images_batch(
    payload: AiImageBatchDownloadRequest,
    _: LocalUser = Depends(current_user),
) -> StreamingResponse:
    if not payload.items:
        raise HTTPException(status_code=400, detail="请选择要下载的图片")
    if len(payload.items) > 50:
        raise HTTPException(status_code=400, detail="一次最多下载 50 张图片")

    output = tempfile.SpooledTemporaryFile(max_size=64 * 1024 * 1024, mode="w+b")
    used_names: set[str] = set()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for item in payload.items:
            safe_object_key = _validated_ai_image_object_key(item.object_key)
            result = _open_ai_image_oss_object(safe_object_key)
            filename = _ai_image_download_filename(item.filename, safe_object_key)
            stem = Path(filename).stem
            suffix = Path(filename).suffix
            candidate = filename
            index = 2
            while candidate.lower() in used_names:
                candidate = f"{stem} ({index}){suffix}"
                index += 1
            used_names.add(candidate.lower())
            with archive.open(candidate, mode="w") as target:
                for chunk in _iter_ai_image_oss_object(result):
                    target.write(chunk)

    output.seek(0)
    filename = f"ai-images-{datetime.now():%Y%m%d_%H%M%S}.zip"
    return StreamingResponse(
        _iter_ai_image_download_archive(output),
        media_type="application/zip",
        headers={
            "Content-Disposition": _content_disposition_attachment(filename),
            "Cache-Control": "private, no-store",
        },
    )


@app.post("/api/v1/ai-image/process", response_model=AiImageProcessResponse)
async def process_ai_image(
    operation: str = Form(...),
    prompt: str = Form(""),
    model_setting_id: int | None = Form(None),
    size: str = Form("1024x1024"),
    quality: str = Form("medium"),
    count: int = Form(1),
    output_format: str = Form("png"),
    output_compression: int | None = Form(None),
    split_mode: str = Form("long"),
    split_instruction: str = Form(""),
    split_max_height: int = Form(2048),
    split_rows: int = Form(2),
    split_columns: int = Form(2),
    merge_layout: str = Form("grid"),
    merge_columns: int = Form(2),
    merge_cell_width: int | None = Form(None),
    merge_cell_height: int | None = Form(None),
    merge_gap: int = Form(16),
    merge_background: str = Form("#ffffff"),
    merge_fit_mode: str = Form("contain"),
    image_urls: str = Form(""),
    mask_image_url: str = Form(""),
    images: list[UploadFile] = File(default=[]),
    mask_image: UploadFile | None = File(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> AiImageProcessResponse:
    normalized_operation = (operation or "").strip().lower()
    network_image_urls = [line.strip() for line in str(image_urls or "").splitlines() if line.strip()]
    network_mask_url = str(mask_image_url or "").strip()
    input_image_count = len(images) + len(network_image_urls)
    if normalized_operation not in {"generate", "edit", "split", "merge"}:
        raise HTTPException(status_code=422, detail="图片操作类型无效")
    if input_image_count > 8:
        raise HTTPException(status_code=400, detail="一次最多添加 8 张图片")
    if mask_image is not None and normalized_operation != "edit":
        raise HTTPException(status_code=400, detail="蒙版仅支持图片修改")
    if network_mask_url and normalized_operation != "edit":
        raise HTTPException(status_code=400, detail="蒙版仅支持图片修改")
    if mask_image is not None and network_mask_url:
        raise HTTPException(status_code=400, detail="蒙版只能选择本地文件或网络地址之一")
    if normalized_operation == "generate" and input_image_count:
        raise HTTPException(status_code=400, detail="文生图不需要添加输入图片")
    if normalized_operation == "edit" and not input_image_count:
        raise HTTPException(status_code=400, detail="图片修改至少需要添加一张图片")
    if normalized_operation == "split" and input_image_count != 1:
        raise HTTPException(status_code=400, detail="图片拆分需要且只能添加一张图片")
    if normalized_operation == "merge" and input_image_count < 2:
        raise HTTPException(status_code=400, detail="图片合并至少需要添加两张图片")

    job_id = uuid.uuid4().hex
    with tempfile.TemporaryDirectory(prefix="ai-image-") as temp_dir_text:
        temp_dir = Path(temp_dir_text)
        source_paths: list[Path] = []
        for index, file in enumerate(images, start=1):
            source_paths.append(await _save_ai_image_upload(file, temp_dir, index))
        for index, source_url in enumerate(network_image_urls, start=len(source_paths)):
            downloaded_path, _final_url = await asyncio.to_thread(
                download_network_image_to_directory,
                source_url,
                temp_dir,
                index=index,
            )
            source_paths.append(downloaded_path)
        if mask_image is not None:
            mask_path = await _save_ai_image_upload(mask_image, temp_dir, 99)
        elif network_mask_url:
            mask_path, _final_mask_url = await asyncio.to_thread(
                download_network_image_to_directory,
                network_mask_url,
                temp_dir,
                index=99,
            )
        else:
            mask_path = None

        try:
            setting: ModelSetting | None = None
            model_setting_name = ""
            model = ""
            if normalized_operation in {"generate", "edit"}:
                setting = _get_ai_image_model_setting(db, model_setting_id)
                endpoint = setting.endpoint
                api_key = _decrypt_model_api_key(endpoint) if endpoint is not None else ""
                if not api_key:
                    raise HTTPException(status_code=422, detail="图片模型 api key 为空")
                effective_prompt = str(prompt or "").strip()
                raw_images = await asyncio.to_thread(
                    call_image_api,
                    ImageApiConfig(base_url=endpoint.base_url, api_key=api_key, model=setting.model),
                    operation=normalized_operation,
                    prompt=effective_prompt,
                    source_paths=source_paths,
                    mask_path=mask_path,
                    size=size,
                    quality=quality,
                    count=count,
                    output_format=output_format,
                    output_compression=output_compression,
                )
                result_paths = await asyncio.to_thread(
                    write_api_images,
                    raw_images,
                    temp_dir / "outputs",
                    output_format=output_format,
                )
                model_setting_name = setting.name or ""
                model = setting.model or ""
            elif normalized_operation == "split":
                split_regions = None
                normalized_split_mode = (split_mode or "long").strip().lower()
                if normalized_split_mode == "ai":
                    setting = _get_ai_image_model_setting(db, model_setting_id, require_vision=True)
                    messages = await asyncio.to_thread(
                        build_ai_split_messages,
                        source_paths[0],
                        split_instruction,
                    )
                    model_text = await asyncio.to_thread(
                        _post_model_chat_completion,
                        setting,
                        messages,
                        max_tokens=1800,
                        temperature=0,
                        timeout=180,
                    )
                    metadata = validate_image_file(source_paths[0])
                    model_split_regions = await asyncio.to_thread(
                        parse_ai_split_regions,
                        model_text,
                        image_width=int(metadata["width"]),
                        image_height=int(metadata["height"]),
                    )
                    split_regions = await asyncio.to_thread(
                        refine_ai_split_regions,
                        source_paths[0],
                        model_split_regions,
                    )
                    logger.info(
                        "AI image split detected job_id=%s source=%sx%s proposed=%s refined=%s",
                        job_id,
                        metadata["width"],
                        metadata["height"],
                        [
                            [region.left, region.top, region.right, region.bottom]
                            for region in model_split_regions
                        ],
                        [
                            {
                                "box": [region.left, region.top, region.right, region.bottom],
                                "size": [region.right - region.left, region.bottom - region.top],
                                "confidence": region.confidence,
                            }
                            for region in split_regions
                        ],
                    )
                    model_setting_name = setting.name or ""
                    model = setting.model or ""
                result_paths = await asyncio.to_thread(
                    split_image,
                    source_paths[0],
                    temp_dir / "outputs",
                    split_mode=normalized_split_mode,
                    max_height=split_max_height,
                    rows=split_rows,
                    columns=split_columns,
                    output_format=output_format,
                    output_compression=output_compression,
                    regions=split_regions,
                )
            else:
                output_path = temp_dir / "outputs" / f"merged{output_suffix(output_format)}"
                result = await asyncio.to_thread(
                    merge_images,
                    source_paths,
                    output_path,
                    layout=merge_layout,
                    columns=merge_columns,
                    cell_width=merge_cell_width,
                    cell_height=merge_cell_height,
                    gap=merge_gap,
                    background=merge_background,
                    fit_mode=merge_fit_mode,
                    output_format=output_format,
                    output_compression=output_compression,
                )
                result_paths = [result]

            source_assets = [
                await _upload_ai_image_asset(path, job_id=job_id, group="inputs", index=index)
                for index, path in enumerate(source_paths, start=1)
            ]
            if mask_path is not None:
                source_assets.append(await _upload_ai_image_asset(mask_path, job_id=job_id, group="mask", index=1))
            assets = [
                await _upload_ai_image_asset(path, job_id=job_id, group="outputs", index=index)
                for index, path in enumerate(result_paths, start=1)
            ]
            return AiImageProcessResponse(
                operation=normalized_operation,
                model_setting_id=setting.id if setting else None,
                model_setting_name=model_setting_name,
                model=model,
                source_assets=source_assets,
                assets=assets,
            )
        except HTTPException:
            raise
        except AiImageUpstreamError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc
        except (AiImageError, ImageStorageError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            logger.exception("AI image processing failed")
            raise HTTPException(status_code=500, detail=f"图片处理失败: {exc}") from exc


@app.get("/api/v1/system-settings/scheduled-tasks", response_model=list[ScheduledTaskDto])
def list_scheduled_tasks(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> list[ScheduledTaskDto]:
    rows = db.scalars(select(ScheduledTask).order_by(desc(ScheduledTask.id))).all()
    return [_scheduled_task_dto(row) for row in rows]


@app.post("/api/v1/system-settings/scheduled-tasks", response_model=ScheduledTaskDto)
def create_scheduled_task(
    payload: ScheduledTaskUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskDto:
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="任务名称不能为空")
    task_type = (payload.task_type or "auto_order_pipeline").strip() or "auto_order_pipeline"
    if task_type != "auto_order_pipeline":
        raise HTTPException(status_code=400, detail="仅支持自动订单流水线任务")
    cron_expr = _validate_cron_expr(payload.cron_expr)
    task_settings = dict(payload.settings or {})
    task_settings.pop("test_mode", None)
    task_settings.pop("max_orders", None)
    row = ScheduledTask(
        name=name,
        task_type=task_type,
        cron_expr=cron_expr,
        enabled=payload.enabled,
        settings=task_settings,
        remark=(payload.remark or "").strip(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    reload_jobs()
    return _scheduled_task_dto(row)


@app.put("/api/v1/system-settings/scheduled-tasks/{task_id:int}", response_model=ScheduledTaskDto)
def update_scheduled_task(
    task_id: int,
    payload: ScheduledTaskUpsertRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskDto:
    row = db.get(ScheduledTask, task_id)
    if not row:
        raise HTTPException(status_code=404, detail="定时任务不存在")
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="任务名称不能为空")
    task_type = (payload.task_type or "auto_order_pipeline").strip() or "auto_order_pipeline"
    if task_type != "auto_order_pipeline":
        raise HTTPException(status_code=400, detail="仅支持自动订单流水线任务")
    row.name = name
    row.task_type = task_type
    row.cron_expr = _validate_cron_expr(payload.cron_expr)
    row.enabled = payload.enabled
    task_settings = dict(payload.settings or {})
    task_settings.pop("test_mode", None)
    task_settings.pop("max_orders", None)
    row.settings = task_settings
    row.remark = (payload.remark or "").strip()
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    reload_jobs()
    return _scheduled_task_dto(row)


@app.post("/api/v1/system-settings/scheduled-tasks/{task_id:int}/toggle", response_model=ScheduledTaskDto)
def toggle_scheduled_task(task_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> ScheduledTaskDto:
    row = db.get(ScheduledTask, task_id)
    if not row:
        raise HTTPException(status_code=404, detail="定时任务不存在")
    row.enabled = not bool(row.enabled)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    reload_jobs()
    return _scheduled_task_dto(row)


@app.post("/api/v1/system-settings/scheduled-tasks/{task_id:int}/run", response_model=ScheduledTaskDto)
def execute_scheduled_task_now(
    task_id: int,
    background_tasks: BackgroundTasks,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskDto:
    row = db.get(ScheduledTask, task_id)
    if not row:
        raise HTTPException(status_code=404, detail="定时任务不存在")
    row.last_run_at = datetime.utcnow()
    row.last_status = "running"
    row.last_message = "手动触发已提交，后台执行中"
    db.commit()
    background_tasks.add_task(run_scheduled_task_now, task_id)
    db.refresh(row)
    return _scheduled_task_dto(row)


@app.delete("/api/v1/system-settings/scheduled-tasks/{task_id:int}")
def delete_scheduled_task(task_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> dict:
    row = db.get(ScheduledTask, task_id)
    if not row:
        raise HTTPException(status_code=404, detail="定时任务不存在")
    db.delete(row)
    db.commit()
    reload_jobs()
    return {"ok": True}


@app.get("/api/v1/system-settings/scheduled-task-runs", response_model=ScheduledTaskRunListResponse)
def get_scheduled_task_runs(
    task_id: int | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    limit: int | None = Query(default=None, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskRunListResponse:
    if limit is not None:
        page = 1
        page_size = limit
    where_clause = ScheduledTaskRun.scheduled_task_id == task_id if task_id else True
    total = db.scalar(select(func.count()).select_from(ScheduledTaskRun).where(where_clause)) or 0
    stmt = select(ScheduledTaskRun).where(where_clause).order_by(desc(ScheduledTaskRun.id)).offset((page - 1) * page_size).limit(page_size)
    rows = db.scalars(stmt).all()
    run_ids = [row.id for row in rows]
    pdf_platform_map: dict[int, list[str]] = {}
    if run_ids:
        pdf_rows = db.execute(
            select(ScheduledTaskRunOrder.run_id, ScheduledTaskRunOrder.platform)
            .where(
                ScheduledTaskRunOrder.run_id.in_(run_ids),
                ScheduledTaskRunOrder.pdf_generated == True,
                ScheduledTaskRunOrder.pdf_file_path != "",
            )
            .group_by(ScheduledTaskRunOrder.run_id, ScheduledTaskRunOrder.platform)
            .order_by(asc(ScheduledTaskRunOrder.run_id), asc(func.min(ScheduledTaskRunOrder.id)))
        ).all()
        for run_id, platform in pdf_rows:
            pdf_platform_map.setdefault(run_id, []).append((platform or "unknown").strip() or "unknown")
    return ScheduledTaskRunListResponse(
        items=[_scheduled_task_run_dto(row, pdf_platform_map.get(row.id, [])) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.get("/api/v1/system-settings/scheduled-task-runs/{run_id:int}/steps", response_model=list[ScheduledTaskRunStepDto])
def get_scheduled_task_run_steps(
    run_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ScheduledTaskRunStepDto]:
    rows = db.scalars(
        select(ScheduledTaskRunStep).where(ScheduledTaskRunStep.run_id == run_id).order_by(asc(ScheduledTaskRunStep.id))
    ).all()
    return [_scheduled_task_run_step_dto(row) for row in rows]


@app.get("/api/v1/system-settings/scheduled-task-runs/{run_id:int}/orders", response_model=list[ScheduledTaskRunOrderDto])
def get_scheduled_task_run_orders(
    run_id: int,
    needs_reprint: bool | None = Query(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ScheduledTaskRunOrderDto]:
    if needs_reprint is True:
        refresh_reprint_candidates(db, run_id)
    latest_labels = (
        select(
            Shipment.order_id.label("order_id"),
            LabelFile.file_path.label("file_path"),
            func.row_number()
            .over(partition_by=Shipment.order_id, order_by=LabelFile.id.desc())
            .label("rn"),
        )
        .join(LabelFile, LabelFile.shipment_id == Shipment.id)
        .subquery()
    )
    stmt = (
        select(ScheduledTaskRunOrder, Order.platform_order_no, Order.platform_order_id, latest_labels.c.file_path)
        .outerjoin(Order, Order.id == ScheduledTaskRunOrder.order_id)
        .outerjoin(latest_labels, (latest_labels.c.order_id == ScheduledTaskRunOrder.order_id) & (latest_labels.c.rn == 1))
        .where(ScheduledTaskRunOrder.run_id == run_id)
        .order_by(asc(ScheduledTaskRunOrder.id))
    )
    if needs_reprint is not None:
        stmt = stmt.where(ScheduledTaskRunOrder.needs_reprint == needs_reprint)
    rows = db.execute(stmt).all()
    return [
        _scheduled_task_run_order_dto(row, platform_order_no or platform_order_id, label_file_path)
        for row, platform_order_no, platform_order_id, label_file_path in rows
    ]


@app.get("/api/v1/system-settings/scheduled-task-runs/{run_id:int}/platforms", response_model=list[ScheduledTaskRunPlatformDto])
def get_scheduled_task_run_platforms(
    run_id: int,
    needs_reprint: bool | None = Query(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> list[ScheduledTaskRunPlatformDto]:
    run = db.get(ScheduledTaskRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="任务运行记录不存在")
    return _scheduled_task_run_platform_rows(run_id, db, needs_reprint)


@app.get("/api/v1/system-settings/scheduled-task-runs/{run_id:int}/pdf")
def download_scheduled_task_run_pdfs(
    run_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    return _scheduled_task_run_pdf_response(run_id, db)


@app.post(
    "/api/v1/system-settings/scheduled-task-runs/{run_id:int}/pdf-download-link",
    response_model=ScheduledTaskRunPdfDownloadLinkDto,
)
def create_scheduled_task_run_pdf_download_link(
    run_id: int,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskRunPdfDownloadLinkDto:
    pdf_entries = _scheduled_task_run_pdf_entries_or_raise(run_id, db)
    expires_in_seconds = 300
    token = create_scheduled_task_run_pdf_download_token(user.username, run_id, expires_in_seconds)
    filename = _scheduled_task_run_pdf_filename(_local_now().strftime("%Y%m%d_%H%M%S"))
    return ScheduledTaskRunPdfDownloadLinkDto(
        url=f"/api/v1/system-settings/scheduled-task-runs/{run_id}/pdf-download?token={token}",
        filename=filename,
        expires_in_seconds=expires_in_seconds,
    )


@app.get("/api/v1/system-settings/scheduled-task-runs/{run_id:int}/pdf-download")
def download_scheduled_task_run_pdfs_with_token(
    run_id: int,
    _: LocalUser = Depends(current_user_from_scheduled_task_run_pdf_download_token),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    return _scheduled_task_run_pdf_response(run_id, db)


@app.post("/api/v1/system-settings/scheduled-task-run-orders/{run_order_id:int}/reprint", response_model=ScheduledTaskRunOrderDto)
def reprint_scheduled_task_run_order(
    run_order_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskRunOrderDto:
    row = db.get(ScheduledTaskRunOrder, run_order_id)
    if not row:
        raise HTTPException(status_code=404, detail="运行订单记录不存在")
    retry_run_order_print(run_order_id)
    db.expire_all()
    row = db.get(ScheduledTaskRunOrder, run_order_id)
    order = db.get(Order, row.order_id) if row and int(row.order_id or 0) > 0 else None
    return _scheduled_task_run_order_dto(row, (order.platform_order_no or order.platform_order_id) if order else None)


@app.post("/api/v1/system-settings/scheduled-task-runs/{run_id:int}/platforms/{platform}/reprint", response_model=ScheduledTaskRunPlatformDto)
def reprint_scheduled_task_run_platform(
    run_id: int,
    platform: str,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
) -> ScheduledTaskRunPlatformDto:
    run = db.get(ScheduledTaskRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="任务运行记录不存在")
    try:
        retry_run_platform_print(run_id, platform, failed_only=not _scheduled_run_is_successful(run))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    db.expire_all()
    rows = _scheduled_task_run_platform_rows(run_id, db)
    platform_key = (platform or "").strip()
    for row in rows:
        if row.platform == platform_key:
            return row
    raise HTTPException(status_code=404, detail="运行平台记录不存在")


@app.get("/api/v1/logistics-rules", response_model=LogisticsMatchRuleListResponse)
def list_logistics_match_rules(
    name: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    enabled: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    stmt = select(LogisticsMatchRule)
    count_stmt = select(func.count(LogisticsMatchRule.id))
    conditions = []
    if name and name.strip():
        like = f"%{name.strip()}%"
        conditions.append(
            or_(
                LogisticsMatchRule.name.ilike(like),
                LogisticsMatchRule.logistics_channel.ilike(like),
            )
        )
    if platform and platform.strip():
        conditions.append(LogisticsMatchRule.platform == normalize_platform_code(platform))
    if enabled is not None:
        conditions.append(LogisticsMatchRule.enabled == enabled)
    if conditions:
        stmt = stmt.where(*conditions)
        count_stmt = count_stmt.where(*conditions)
    rows = db.scalars(
        stmt.order_by(asc(LogisticsMatchRule.priority), asc(LogisticsMatchRule.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return LogisticsMatchRuleListResponse(
        items=[_logistics_match_rule_dto(row) for row in rows],
        total=int(db.scalar(count_stmt) or 0),
        page=page,
        page_size=page_size,
    )


@app.get("/api/v1/logistics-rules/channel-options", response_model=list[LogisticsChannelOptionDto])
def list_logistics_channel_options(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    rows = db.scalars(
        select(LogisticsAuthorization)
        .where(LogisticsAuthorization.enabled == True)
        .order_by(asc(LogisticsAuthorization.carrier_name), asc(LogisticsAuthorization.account_name), asc(LogisticsAuthorization.id))
    ).all()
    return _enabled_logistics_channel_options(rows)


@app.get("/api/v1/logistics-rules/shop-options", response_model=list[dict[str, str]])
async def list_logistics_rule_shop_options(
    platform: str = Query(...),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    return _logistics_shop_options_for_platform(db, platform)


@app.get("/api/v1/logistics-rules/{rule_id:int}", response_model=LogisticsMatchRuleDto)
def get_logistics_match_rule(rule_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(LogisticsMatchRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="物流规则不存在")
    return _logistics_match_rule_dto(row)


@app.post("/api/v1/logistics-rules", response_model=LogisticsMatchRuleDto, status_code=201)
def create_logistics_match_rule(
    payload: LogisticsMatchRulePayload,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = LogisticsMatchRule(created_by=user.username, created_at=datetime.utcnow(), updated_at=datetime.utcnow())
    _apply_logistics_match_rule_payload(db, row, payload)
    db.add(row)
    db.commit()
    db.refresh(row)
    return _logistics_match_rule_dto(row)


@app.put("/api/v1/logistics-rules/{rule_id:int}", response_model=LogisticsMatchRuleDto)
def update_logistics_match_rule(
    rule_id: int,
    payload: LogisticsMatchRulePayload,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = db.get(LogisticsMatchRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="物流规则不存在")
    _apply_logistics_match_rule_payload(db, row, payload)
    db.commit()
    db.refresh(row)
    return _logistics_match_rule_dto(row)


@app.post("/api/v1/logistics-rules/{rule_id:int}/toggle-enabled", response_model=LogisticsMatchRuleDto)
def toggle_logistics_match_rule(rule_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(LogisticsMatchRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="物流规则不存在")
    row.enabled = not row.enabled
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _logistics_match_rule_dto(row)


@app.delete("/api/v1/logistics-rules/{rule_id:int}")
def delete_logistics_match_rule(rule_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> dict:
    row = db.get(LogisticsMatchRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="物流规则不存在")
    db.delete(row)
    db.commit()
    return {"status": "ok"}


@app.post("/api/v1/logistics-rules/rematch", response_model=LogisticsRematchResponse)
def rematch_logistics_rules(
    payload: LogisticsRematchRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    if payload.order_ids:
        rows = _load_orders_by_ids(db, payload.order_ids)
    else:
        stmt = select(Order)
        if not payload.include_shipped:
            stmt = stmt.where(
                or_(
                    Order.biz_status.is_(None),
                    ~Order.biz_status.in_(
                        [
                            ORDER_STATUS_SHIPPED,
                            ORDER_STATUS_AWAITING_PICKUP,
                            ORDER_STATUS_DELIVERED,
                            ORDER_STATUS_VOIDED,
                            ORDER_STATUS_COMPLETED,
                        ]
                    ),
                )
            )
        rows = db.scalars(stmt.order_by(desc(Order.created_at), desc(Order.id)).limit(5000)).all()
    result = _rematch_orders(db, rows, include_manual=payload.include_manual)
    db.commit()
    return result


@app.get("/api/v1/logistics-authorizations", response_model=list[LogisticsAuthorizationDto])
def list_logistics_authorizations(
    carrier_name: str | None = Query(default=None),
    carrier_code: str | None = Query(default=None),
    enabled: bool | None = Query(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    stmt = select(LogisticsAuthorization)
    if carrier_name:
        like = f"%{carrier_name.strip()}%"
        stmt = stmt.where(
            or_(
                LogisticsAuthorization.carrier_name.ilike(like),
                LogisticsAuthorization.account_name.ilike(like),
            )
        )
    if carrier_code:
        stmt = stmt.where(LogisticsAuthorization.carrier_code == _normalize_logistics_carrier_code(carrier_code))
    if enabled is not None:
        stmt = stmt.where(LogisticsAuthorization.enabled == enabled)
    rows = db.scalars(stmt.order_by(asc(LogisticsAuthorization.id))).all()
    return [_logistics_authorization_dto(row) for row in rows]


@app.get("/api/v1/logistics-authorizations/{auth_id}", response_model=LogisticsAuthorizationDto)
def get_logistics_authorization(auth_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(LogisticsAuthorization, auth_id)
    if not row:
        raise HTTPException(status_code=404, detail="Logistics authorization not found")
    return _logistics_authorization_dto(row)


@app.get("/api/v1/logistics-authorizations/{auth_id}/credentials")
def get_logistics_authorization_credentials(auth_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(LogisticsAuthorization, auth_id)
    if not row:
        raise HTTPException(status_code=404, detail="Logistics authorization not found")
    return _logistics_credentials(row)


@app.post("/api/v1/logistics-authorizations", response_model=LogisticsAuthorizationDto, status_code=201)
def create_logistics_authorization(
    _: LocalUser = Depends(current_user),
):
    raise HTTPException(status_code=405, detail="物流授权必须使用系统预设值")


@app.put("/api/v1/logistics-authorizations/{auth_id}", response_model=LogisticsAuthorizationDto)
def update_logistics_authorization(
    auth_id: int,
    payload: LogisticsAuthorizationUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = db.get(LogisticsAuthorization, auth_id)
    if not row:
        raise HTTPException(status_code=404, detail="Logistics authorization not found")
    current_credentials = _logistics_credentials(row)
    if payload.carrier_name is not None:
        row.carrier_name = payload.carrier_name.strip()
    if payload.account_name is not None:
        account_name = payload.account_name.strip()
        if not account_name:
            raise HTTPException(status_code=400, detail="授权账号名称不能为空")
        duplicate = db.scalar(
            select(LogisticsAuthorization).where(
                LogisticsAuthorization.id != row.id,
                LogisticsAuthorization.carrier_code == row.carrier_code,
                LogisticsAuthorization.account_name == account_name,
            )
        )
        if duplicate:
            raise HTTPException(status_code=400, detail="同一物流公司下授权账号名称已存在")
        row.account_name = account_name
    if payload.enabled is not None:
        row.enabled = payload.enabled
    if payload.credential_type is not None:
        row.credential_type = payload.credential_type or "api_key"
    if payload.credentials is not None:
        current_credentials = dict(payload.credentials or {})
        row.encrypted_credentials = get_credential_manager().encrypt_credentials(current_credentials)
        row.credentials_version = datetime.utcnow().isoformat()
    if payload.config_json is not None:
        row.config_json = dict(payload.config_json or {})
    if payload.settings_json is not None:
        row.settings_json = dict(payload.settings_json or {})
    _apply_logistics_authorization_result(row, current_credentials, payload.authorization_expires_at)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _logistics_authorization_dto(row)


@app.post("/api/v1/logistics-authorizations/{auth_id}/verify", response_model=LogisticsAuthorizationVerifyResponse)
async def verify_logistics_authorization(auth_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(LogisticsAuthorization, auth_id)
    if not row:
        raise HTTPException(status_code=404, detail="Logistics authorization not found")
    credentials = _logistics_credentials(row)
    valid, missing, message = _verify_logistics_credentials(row.carrier_code, credentials)
    if valid and row.carrier_code == BSI_CARRIER_CODE:
        valid, message, refreshed_config = await verify_bsi_authorization(credentials, dict(row.config_json or {}))
        row.config_json = refreshed_config
    row.authorization_status = LOGISTICS_AUTH_SUCCESS if valid else LOGISTICS_AUTH_FAILED
    row.token_valid = valid
    row.token_message = message
    if valid:
        row.last_authorized_at = datetime.utcnow()
    row.updated_at = datetime.utcnow()
    db.commit()
    return LogisticsAuthorizationVerifyResponse(
        authorization_status=row.authorization_status or LOGISTICS_AUTH_UNAUTHORIZED,
        token_valid=valid,
        token_message=message,
        missing_fields=missing,
    )


@app.post("/api/v1/logistics-authorizations/{auth_id}/toggle-enabled", response_model=LogisticsAuthorizationDto)
def toggle_logistics_authorization(auth_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(LogisticsAuthorization, auth_id)
    if not row:
        raise HTTPException(status_code=404, detail="Logistics authorization not found")
    row.enabled = not row.enabled
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _logistics_authorization_dto(row)


@app.get("/internal/logistics-authorizations/{carrier_code}")
def internal_get_logistics_authorization(
    carrier_code: str,
    account_name: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _internal_service: bool = Depends(require_internal_service_token),
):
    stmt = select(LogisticsAuthorization).where(
        LogisticsAuthorization.carrier_code == _normalize_logistics_carrier_code(carrier_code),
        LogisticsAuthorization.enabled == True,
    )
    if account_name:
        stmt = stmt.where(LogisticsAuthorization.account_name == account_name)
    row = db.scalar(stmt.order_by(asc(LogisticsAuthorization.id)).limit(1))
    if not row:
        raise HTTPException(status_code=404, detail="Logistics authorization not found")
    return {
        "id": row.id,
        "carrier_code": row.carrier_code,
        "carrier_name": row.carrier_name,
        "account_name": row.account_name,
        "credentials": _logistics_credentials(row),
        "config_json": row.config_json or {},
        "settings_json": row.settings_json or {},
    }


def _run_manual_sync_background(platform: str | None, account_id: str | None, full_refresh: bool) -> None:
    db = SessionLocal()
    try:
        result = asyncio.run(
            sync_enabled_accounts(
                db,
                platform,
                account_id,
                full_refresh=full_refresh,
            )
        )
        logger.info(
            "Manual sync completed: platform=%s account_id=%s full_refresh=%s result=%s",
            platform or "all",
            account_id or "all",
            full_refresh,
            result,
        )
        reload_jobs()
    except httpx.HTTPStatusError as exc:
        detail = exc.response.text if exc.response is not None else str(exc)
        try:
            parsed = json.loads(detail)
            detail = parsed.get("message") or parsed.get("detail") or detail
        except (TypeError, ValueError):
            pass
        logger.exception("Manual sync platform request failed: %s", detail)
    except Exception:
        logger.exception(
            "Manual sync failed: platform=%s account_id=%s full_refresh=%s",
            platform or "all",
            account_id or "all",
            full_refresh,
        )
    finally:
        db.close()


@app.post("/api/sync/manual")
@app.post("/api/v1/sync/manual")
async def manual_sync(
    payload: ManualSyncRequest,
    background_tasks: BackgroundTasks,
    _: LocalUser = Depends(current_user),
):
    platform = payload.platform.strip() if payload.platform else None
    account_id = payload.account_id.strip() if payload.account_id else None
    background_tasks.add_task(_run_manual_sync_background, platform, account_id, payload.full_refresh)
    return {
        "status": "accepted",
        "message": "手动同步已提交，后台执行中",
        "platform": platform,
        "account_id": account_id,
        "full_refresh": payload.full_refresh,
    }


@app.post("/api/v1/sync/status-refresh")
async def manual_status_refresh(
    payload: ManualSyncRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Logistics/status refresh is limited to the auto order pipeline."""
    raise HTTPException(status_code=409, detail="物流同步仅允许由 auto_order_pipeline 定时任务执行")


@app.get("/api/v1/platforms")
async def v1_platforms(
    enabled_only: bool = Query(default=False),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    seed_default_platform_settings(db)
    db.commit()
    enabled_by_platform = {
        row.platform: bool(row.enabled)
        for row in db.scalars(select(PlatformSetting)).all()
    }
    seen: set[str] = set()
    result: list[dict] = []
    for item in PLATFORM_CATALOG:
        platform = _canonical_platform(item.get("platform") or "")
        if not platform or platform in seen:
            continue
        seen.add(platform)
        enabled = enabled_by_platform.get(platform, bool(item.get("enabled", True)))
        if enabled_only and not enabled:
            continue
        display_name = _platform_display_name(platform, item.get("display_name") or platform)
        base_url = item.get("base_url") or (item.get("settings") or {}).get("base_url") or ""
        result.append(
            {
                "platform": platform,
                "account_id": "",
                "display_name": display_name,
                "enabled": enabled,
                "auth_type": item.get("auth_type") or "api_key",
                "settings": {"base_url": base_url} if base_url else {},
            }
        )
    return result


@app.post("/api/v1/platforms/refresh")
async def v1_refresh_platforms(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    configs = await refresh_configs(db)
    reload_jobs()
    return {"status": "refreshed", "count": len(configs)}


@app.get("/api/shops", response_model=list[ShopDto])
@app.get("/api/v1/shops", response_model=list[ShopDto])
def list_shops(
    display_name: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    enabled: bool | None = Query(default=None),
    sort_by: str | None = Query(default=None),
    sort_order: str = Query(default="desc"),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    rows = db.scalars(_build_shop_query(display_name, platform, enabled, sort_by, sort_order)).all()
    return [_shop_dto(row) for row in rows]


@app.get("/api/shops/{platform}/{account_id}", response_model=ShopDto)
@app.get("/api/v1/shops/{platform}/{account_id}", response_model=ShopDto)
def get_shop(platform: str, account_id: str, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    return _shop_dto(row)


@app.get("/api/v1/shops/{platform}/{account_id}/credentials")
def get_shop_credentials(platform: str, account_id: str, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    creds = get_credential_manager().decrypt_credentials(row.encrypted_credentials)
    return creds or {}


@app.post("/api/shops", response_model=ShopDto, status_code=201)
@app.post("/api/v1/shops", response_model=ShopDto, status_code=201)
async def create_shop(payload: ShopCreateRequest, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    platform = _canonical_platform(payload.platform)
    account_id = _shop_id(payload) or _generate_shop_account_id(db, platform)
    existing = _find_shop(db, platform, account_id)
    if existing:
        raise HTTPException(status_code=400, detail="Shop already exists")
    credentials = _normalize_shop_credentials(platform, payload.credentials)

    account = PlatformAccount(
        platform=platform,
        account_id=account_id,
        display_name=payload.display_name or account_id,
        enabled=payload.enabled,
        credential_type=payload.credential_type,
        auth_type=payload.credential_type,
        encrypted_credentials=get_credential_manager().encrypt_credentials(credentials),
        settings=payload.settings,
        created_by=user.username,
        created_at=datetime.utcnow(),
        credentials_version=datetime.utcnow().isoformat(),
    )
    _apply_shop_authorization_result(account, credentials, payload.authorization_expires_at)
    db.add(account)
    setting = db.scalar(select(SyncSetting).where(SyncSetting.platform == platform, SyncSetting.account_id == account_id))
    if not setting:
        setting = SyncSetting(platform=platform, account_id=account_id)
        db.add(setting)
    setting.enabled = payload.enabled
    setting.interval_seconds = int((payload.settings or {}).get("sync_interval_seconds", 1200))
    setting.dry_run_fulfillment = bool((payload.settings or {}).get("dry_run_fulfillment", False))
    db.flush()
    db.commit()
    db.refresh(account)
    reload_jobs()
    return _shop_dto(account)


@app.put("/api/shops/{platform}/{account_id}", response_model=ShopDto)
@app.put("/api/v1/shops/{platform}/{account_id}", response_model=ShopDto)
async def update_shop(
    platform: str,
    account_id: str,
    payload: ShopUpdateRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    platform = row.platform
    previous_enabled = row.enabled
    if payload.display_name is not None:
        row.display_name = payload.display_name
    if payload.enabled is not None:
        row.enabled = payload.enabled
        setting = db.scalar(select(SyncSetting).where(SyncSetting.platform.in_(_platform_lookup_codes(platform)), SyncSetting.account_id == account_id))
        if setting:
            setting.enabled = payload.enabled
    if payload.settings is not None:
        row.settings = dict(payload.settings or {})
        setting = db.scalar(select(SyncSetting).where(SyncSetting.platform.in_(_platform_lookup_codes(platform)), SyncSetting.account_id == account_id))
        if setting:
            setting.interval_seconds = int((payload.settings or {}).get("sync_interval_seconds", setting.interval_seconds or 1200))
            setting.dry_run_fulfillment = bool((payload.settings or {}).get("dry_run_fulfillment", False))
    if payload.enabled is not None and previous_enabled != row.enabled:
        audit_sync_event(
            db,
            "shop_enabled_changed",
            platform=platform,
            account_id=account_id,
            job_type="sync_orders",
            status="enabled" if row.enabled else "disabled",
            message=f"shop {'enabled' if row.enabled else 'disabled'}",
            extra={"before": previous_enabled, "after": row.enabled},
        )
    db.commit()
    db.refresh(row)
    reload_jobs()
    return _shop_dto(row)


@app.post("/api/v1/shops/{platform}/{account_id}/reauthorize", response_model=ShopDto)
async def reauthorize_shop(
    platform: str,
    account_id: str,
    payload: ShopAuthorizationRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    platform = row.platform
    credentials = payload.credentials
    if not credentials:
        if not row.encrypted_credentials:
            raise HTTPException(status_code=400, detail="缺少授权凭据")
        credentials = get_credential_manager().decrypt_credentials(row.encrypted_credentials)
        credentials = _normalize_shop_credentials(platform, credentials)
    else:
        credentials = _normalize_shop_credentials(platform, credentials)
        row.encrypted_credentials = get_credential_manager().encrypt_credentials(credentials)
    row.credentials_version = datetime.utcnow().isoformat()
    _apply_shop_authorization_result(row, credentials, payload.authorization_expires_at)
    db.commit()
    db.refresh(row)
    return _shop_dto(row)


def _oauth_callback_response(platform: str, request: Request) -> HTMLResponse:
    params = request.query_params
    code = next(
        (
            str(params.get(key) or "").strip()
            for key in ("code", "authorization_code", "auth_code", "oauth_code")
            if str(params.get(key) or "").strip()
        ),
        "",
    )
    state = str(params.get("state") or "").strip()
    error = str(params.get("error_description") or params.get("error") or "").strip()
    return HTMLResponse(callback_html(platform, code=code, state=state, error=error))


@app.get("/api/oauth/{platform}/callback", response_class=HTMLResponse)
def oauth_callback(platform: str, request: Request) -> HTMLResponse:
    return _oauth_callback_response(platform, request)


@app.get("/api/joom/callback", response_class=HTMLResponse)
def joom_oauth_callback(request: Request) -> HTMLResponse:
    return _oauth_callback_response("joom_logistics", request)


@app.get("/api/allegro/callback", response_class=HTMLResponse)
def allegro_oauth_callback(request: Request) -> HTMLResponse:
    return _oauth_callback_response("allegro", request)


@app.get("/api/mercadolibre/callback", response_class=HTMLResponse)
def mercadolibre_oauth_callback(request: Request) -> HTMLResponse:
    return _oauth_callback_response("mercadolibre", request)


@app.post("/api/logistics/bsi/callback/{callback_token}")
async def bsi_logistics_callback(callback_token: str, request: Request) -> dict:
    if not hmac.compare_digest(callback_token, get_settings().bsi_callback_token):
        raise HTTPException(status_code=404, detail="Not found")
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    raw_order_numbers = str(payload.get("orderNumbers") or "") if isinstance(payload, dict) else ""
    order_numbers = list(dict.fromkeys(item.strip() for item in raw_order_numbers.split(",") if item.strip()))
    return {
        "status": 1,
        "msg": "",
        "notFound": [],
        "success": order_numbers,
        "fail": [],
    }


@app.post("/api/v1/shops/{platform}/{account_id}/oauth/start")
async def start_shop_oauth(
    platform: str,
    account_id: str,
    payload: ShopOAuthStartRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    oauth_platform = _oauth_platform(platform)
    if oauth_platform not in {"joom_logistics", "mercadolibre", "allegro"}:
        raise HTTPException(status_code=400, detail=f"{platform} OAuth is not supported")
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    platform = row.platform
    current_credentials = get_credential_manager().decrypt_credentials(row.encrypted_credentials) if row.encrypted_credentials else {}
    credentials = _normalize_shop_credentials(
        platform,
        _merge_non_empty_credentials(current_credentials, payload.credentials),
    )
    client_secret = _oauth_client_secret(credentials)
    if client_secret:
        credentials["client_secret"] = client_secret
    row.credential_type = "oauth2"
    row.auth_type = "oauth2"
    row.encrypted_credentials = get_credential_manager().encrypt_credentials(credentials)
    row.credentials_version = datetime.utcnow().isoformat()
    try:
        session = create_authorization_session(db, row, credentials)
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    audit_sync_event(
        db,
        "oauth_authorization_started",
        platform=oauth_platform,
        account_id=row.account_id,
        status="pending",
        extra={"state": session.state},
    )
    db.commit()
    return {
        "status": session.status,
        "state": session.state,
        "authorize_url": build_authorize_url(session),
        "expires_at": _iso(session.expires_at),
    }


@app.post("/api/v1/shops/{platform}/{account_id}/oauth/complete")
async def complete_shop_oauth(
    platform: str,
    account_id: str,
    payload: ShopOAuthCompleteRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    oauth_platform = _oauth_platform(platform)
    if oauth_platform not in {"joom_logistics", "mercadolibre", "allegro"}:
        raise HTTPException(status_code=400, detail=f"{platform} OAuth is not supported")
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    platform = row.platform
    session = get_authorization_session(
        db,
        state=payload.state,
        platform=oauth_platform,
        account_id=row.account_id,
    )
    if not session or session.platform_account_id != row.id:
        raise HTTPException(status_code=404, detail="OAuth authorization session not found")
    if session.status == "success":
        return {"status": "success", "state": session.state, "shop": _shop_dto(row)}
    if session.status == "pending" and not (payload.code or "").strip():
        row.authorization_status = SHOP_AUTH_UNAUTHORIZED
        row.token_valid = None
        row.token_message = f"{oauth_platform} 授权尚未完成，请完成授权后再同步结果"
        db.commit()
        return {"status": "pending", "state": session.state, "message": row.token_message}
    if session.status != "pending":
        message = session.error_message or f"OAuth 授权状态异常：{session.status}"
        row.authorization_status = SHOP_AUTH_FAILED
        row.token_valid = False
        row.token_message = message
        db.commit()
        raise HTTPException(status_code=400, detail=message)
    current_credentials = get_credential_manager().decrypt_credentials(row.encrypted_credentials) if row.encrypted_credentials else {}
    try:
        token_payload = await exchange_authorization_code(session, (payload.code or "").strip(), current_credentials)
    except Exception as exc:
        mark_authorization_failed(session, exc)
        row.authorization_status = SHOP_AUTH_FAILED
        row.token_valid = False
        row.token_message = str(exc)
        audit_sync_event(
            db,
            "oauth_authorization_failed",
            platform=oauth_platform,
            account_id=row.account_id,
            status="failed",
            message=str(exc),
            extra={"state": session.state},
        )
        db.commit()
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    credentials = _normalize_shop_credentials(
        platform,
        _merge_non_empty_credentials(current_credentials, token_payload),
    )
    row.credential_type = "oauth2"
    row.auth_type = "oauth2"
    row.encrypted_credentials = get_credential_manager().encrypt_credentials(credentials)
    row.credentials_version = datetime.utcnow().isoformat()
    _apply_shop_authorization_result(row, credentials, token_payload.get("expires_at") or None)
    mark_authorization_success(session)
    audit_sync_event(
        db,
        "oauth_authorization_completed",
        platform=oauth_platform,
        account_id=row.account_id,
        status="success",
        extra={"state": session.state},
    )
    db.commit()
    db.refresh(row)
    return {"status": "success", "state": session.state, "shop": _shop_dto(row)}


@app.post("/api/v1/shops/{platform}/{account_id}/toggle-enabled", response_model=ShopDto)
def toggle_shop_enabled(platform: str, account_id: str, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    previous_enabled = row.enabled
    row.enabled = not row.enabled
    setting = db.scalar(select(SyncSetting).where(SyncSetting.platform.in_(_platform_lookup_codes(row.platform)), SyncSetting.account_id == account_id))
    if setting:
        setting.enabled = row.enabled
    audit_sync_event(
        db,
        "shop_enabled_changed",
        platform=row.platform,
        account_id=account_id,
        job_type="sync_orders",
        status="enabled" if row.enabled else "disabled",
        message=f"shop {'enabled' if row.enabled else 'disabled'}",
        extra={"before": previous_enabled, "after": row.enabled},
    )
    db.commit()
    db.refresh(row)
    reload_jobs()
    return _shop_dto(row)


@app.delete("/api/shops/{platform}/{account_id}")
@app.delete("/api/v1/shops/{platform}/{account_id}")
def delete_shop(platform: str, account_id: str, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    db.delete(row)
    db.commit()
    reload_jobs()
    return {"message": "Shop deleted"}


@app.post("/api/shops/{platform}/{account_id}/credentials")
@app.post("/api/v1/shops/{platform}/{account_id}/credentials")
async def update_credentials(
    platform: str,
    account_id: str,
    payload: CredentialsRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = _find_shop(db, platform, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Shop not found")
    platform = row.platform
    credentials = _normalize_shop_credentials(platform, payload.credentials)
    row.encrypted_credentials = get_credential_manager().encrypt_credentials(credentials)
    row.credentials_version = datetime.utcnow().isoformat()
    _apply_shop_authorization_result(row, credentials, payload.authorization_expires_at)
    db.commit()
    return {"message": "Credentials updated"}


@app.post("/api/v1/shops/{platform}/{account_id}/verify-credentials")
def verify_credentials(platform: str, account_id: str, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = _find_shop(db, platform, account_id)
    if not row or not row.encrypted_credentials:
        raise HTTPException(status_code=404, detail="Credentials not found")
    credentials = get_credential_manager().decrypt_credentials(row.encrypted_credentials)
    valid, missing = _apply_shop_authorization_result(row, credentials)
    db.commit()
    return {
        "valid": valid,
        "missing": missing,
        "status": row.status,
        "authorization_status": row.authorization_status,
        "token_message": row.token_message,
    }


@app.get("/api/v1/products", response_model=ProductListResponse)
def list_products(
    keyword: str | None = None,
    product_code: str | None = None,
    internal_name: str | None = None,
    english_name: str | None = None,
    ean: str | None = None,
    shop_sku: str | None = None,
    enabled: bool | None = None,
    is_slow_moving_material: bool | None = None,
    include_options: bool = Query(default=True),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    shops = _enabled_product_shops(db) if include_options else []
    users = _enabled_buyer_users(db) if include_options else []
    stmt = _product_list_base_query(
        product_code,
        internal_name,
        english_name,
        ean,
        shop_sku,
        enabled,
        is_slow_moving_material,
        keyword,
    )
    conditions = _product_filter_conditions(
        product_code,
        internal_name,
        english_name,
        ean,
        shop_sku,
        enabled,
        is_slow_moving_material,
        keyword,
    )
    count_stmt = select(func.count(Product.id))
    if conditions:
        count_stmt = count_stmt.where(*conditions)
    total = db.scalar(count_stmt) or 0
    rows = db.scalars(
        stmt.options(selectinload(Product.mappings), joinedload(Product.buyer_user))
        .order_by(desc(Product.updated_at), desc(Product.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return ProductListResponse(
        shops=[_product_shop_dto(shop) for shop in shops],
        users=[_user_option_dto(user) for user in users],
        items=[_product_dto(row, shops) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.get("/api/v1/products/shops", response_model=list[ProductShopDto])
def list_product_shops(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    return [_product_shop_dto(shop) for shop in _enabled_product_shops(db)]


@app.get("/api/v1/products/options")
def list_product_options(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    return {
        "shops": [_product_shop_dto(shop) for shop in _enabled_product_shops(db)],
        "users": [_user_option_dto(user) for user in _enabled_buyer_users(db)],
    }


@app.get("/api/v1/products/import-template")
def download_product_import_template(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    shops = _enabled_product_shops(db)
    workbook = _build_product_import_template_workbook(shops, db)
    filename = f"product_import_template_{_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _product_workbook_response(workbook, filename)


@app.get("/api/v1/products/export-disabled", include_in_schema=False)
def export_products_disabled(
    keyword: str | None = None,
    product_code: str | None = None,
    internal_name: str | None = None,
    english_name: str | None = None,
    ean: str | None = None,
    shop_sku: str | None = None,
    enabled: bool | None = None,
    is_slow_moving_material: bool | None = None,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    shops = _enabled_product_shops(db)
    stmt = _product_list_base_query(
        product_code,
        internal_name,
        english_name,
        ean,
        shop_sku,
        enabled,
        is_slow_moving_material,
        keyword,
    )
    rows = db.scalars(
        stmt.options(selectinload(Product.mappings), joinedload(Product.buyer_user)).order_by(asc(Product.product_code))
    ).all()
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "产品目录"
    _append_product_export_header(worksheet, shops)
    for product in rows:
        for export_row in _product_export_rows(product, shops):
            worksheet.append(export_row)
    _style_product_workbook(worksheet, shops)
    filename = f"products_{_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _product_workbook_response(workbook, filename)


@app.get("/api/v1/products/{product_id:int}", response_model=ProductDto)
def get_product(product_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.scalar(select(Product).options(selectinload(Product.mappings), joinedload(Product.buyer_user)).where(Product.id == product_id))
    if not row:
        raise HTTPException(status_code=404, detail="产品不存在")
    return _product_dto(row, _enabled_product_shops(db))


@app.post("/api/v1/products", response_model=ProductDto, status_code=201)
def create_product(payload: ProductUpsertRequest, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    name = _normalize_product_name(payload.internal_name)
    if db.scalar(select(Product).where(Product.internal_name == name)):
        raise HTTPException(status_code=400, detail="中文名称已存在")
    payload.internal_name = name
    product = _create_product_from_payload(db, payload)
    db.commit()
    product = db.scalar(
        select(Product)
        .options(selectinload(Product.mappings), joinedload(Product.buyer_user))
        .where(Product.id == product.id)
    )
    return _product_dto(product, _enabled_product_shops(db))


@app.put("/api/v1/products/{product_id:int}", response_model=ProductDto)
def update_product(product_id: int, payload: ProductUpsertRequest, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    product = db.scalar(select(Product).options(selectinload(Product.mappings), joinedload(Product.buyer_user)).where(Product.id == product_id))
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")
    name = _normalize_product_name(payload.internal_name)
    conflict = db.scalar(select(Product).where(Product.internal_name == name, Product.id != product_id))
    if conflict:
        raise HTTPException(status_code=400, detail="中文名称已存在")
    product.internal_name = name
    product.english_name = _normalize_optional_text(payload.english_name, 255)
    product.cost = _decimal_or_none(payload.cost)
    product.weight = _decimal_or_none(payload.weight)
    product.gross_weight = _decimal_or_none(payload.gross_weight)
    product.package_length = _decimal_or_none(payload.package_length)
    product.package_width = _decimal_or_none(payload.package_width)
    product.package_height = _decimal_or_none(payload.package_height)
    product.ean = _normalize_optional_text(payload.ean, 64)
    product.description = _normalize_optional_text(payload.description)
    product.main_image_url = _normalize_optional_text(payload.main_image_url)
    product.is_slow_moving_material = bool(payload.is_slow_moving_material)
    product.safety_stock = _int_or_none(payload.safety_stock)
    buyer_user = _buyer_user_by_payload(db, payload.buyer_user_id)
    product.buyer_user_id = buyer_user.id if buyer_user else None
    product.enabled = payload.enabled
    product.updated_at = datetime.utcnow()
    _upsert_product_mappings(db, product, payload.mappings)
    db.commit()
    product = db.scalar(
        select(Product)
        .options(selectinload(Product.mappings), joinedload(Product.buyer_user))
        .where(Product.id == product_id)
    )
    return _product_dto(product, _enabled_product_shops(db))


@app.delete("/api/v1/products/{product_id:int}")
def delete_product(product_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")
    db.delete(product)
    db.commit()
    return {"message": "已删除"}


@app.post("/api/v1/products/{product_id:int}/toggle-enabled", response_model=ProductDto)
def toggle_product_enabled(product_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    product = db.scalar(select(Product).options(selectinload(Product.mappings), joinedload(Product.buyer_user)).where(Product.id == product_id))
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")
    product.enabled = not product.enabled
    product.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(product)
    return _product_dto(product, _enabled_product_shops(db))


@app.post("/api/v1/products/batch/enabled")
def batch_set_products_enabled(
    payload: ProductBatchRequest,
    enabled: bool = True,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    if not payload.product_ids:
        raise HTTPException(status_code=400, detail="请先选择产品")
    rows = db.scalars(select(Product).where(Product.id.in_(payload.product_ids))).all()
    for row in rows:
        row.enabled = enabled
        row.updated_at = datetime.utcnow()
    db.commit()
    return {"updated": len(rows), "message": f"已{'启用' if enabled else '禁用'} {len(rows)} 个产品"}


def _shop_header_map(shops: list[PlatformAccount]) -> dict[str, PlatformAccount]:
    mapping: dict[str, PlatformAccount] = {}
    for shop in shops:
        for name in {shop.display_name, shop.account_id}:
            if name:
                mapping[name.strip().lower()] = shop
    return mapping


def _clean_import_value(value) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    return "" if text_value.upper() == "#N/A" else text_value


def _read_product_import_rows(file_name: str, content: bytes) -> tuple[list[str], list[list[object]], int]:
    if not file_name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 xlsx 格式的 Excel 文件")
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导入 xlsx") from exc
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook["产品目录"] if "产品目录" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], [], 1
    first_header = [_clean_import_value(value) for value in rows[0]]
    if len(rows) > 1:
        second_header = [_clean_import_value(value) for value in rows[1]]
        group_markers = {"SKU CODE", "标准参数"}
        group_marker_keys = {value.lower() for value in group_markers}
        sub_header_markers = {"成本", "净重", "毛重", "包装长", "包装宽", "包装高", "EAN"}
        sub_header_marker_keys = {value.lower() for value in sub_header_markers}
        has_group_header = any(value.strip() in group_markers or value.strip().lower() in group_marker_keys for value in first_header)
        has_second_header = any(value.strip() in sub_header_markers or value.strip().lower() in sub_header_marker_keys for value in second_header)
        if has_group_header and has_second_header:
            width = max(len(first_header), len(second_header))
            headers = []
            for index in range(width):
                top_value = first_header[index] if index < len(first_header) else ""
                bottom_value = second_header[index] if index < len(second_header) else ""
                headers.append(bottom_value or top_value)
            return headers, [list(row) for row in rows[2:]], 3
    return first_header, [list(row) for row in rows[1:]], 2


def _find_header(headers: list[str], names: set[str]) -> int | None:
    normalized = {name.strip().lower() for name in names}
    return next((idx for idx, title in enumerate(headers) if title.strip().lower() in normalized), None)


def _normalize_inventory_name(value: str | None) -> str:
    name = (value or "").replace("\u3000", " ").strip()
    name = " ".join(name.split())
    if not name:
        raise HTTPException(status_code=400, detail="产品名称不能为空")
    return name


def _inventory_stock_status(stock_qty: int, safety_stock: int | None) -> str:
    if safety_stock is None:
        return ""
    return "低库存" if stock_qty <= safety_stock else "正常"


def _inventory_dto(row: ProductInventory, product: Product) -> InventoryDto:
    stock_qty = row.stock_qty or 0
    return InventoryDto(
        id=row.id,
        product_id=product.id,
        product_code=product.product_code,
        product_name=product.internal_name,
        stock_qty=stock_qty,
        last_count_qty=row.last_count_qty or 0,
        safety_stock=product.safety_stock,
        stock_status=_inventory_stock_status(stock_qty, product.safety_stock),
        remark=row.remark or "",
        updated_by=row.updated_by,
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _inventory_query(
    product_code: str | None = None,
    product_name: str | None = None,
    stock_status: str | None = None,
    hide_zero_safety_stock: bool = False,
):
    stmt = select(ProductInventory, Product).join(Product, Product.id == ProductInventory.product_id)
    if product_code:
        stmt = stmt.where(Product.product_code.ilike(f"%{product_code.strip()}%"))
    if product_name:
        stmt = stmt.where(Product.internal_name.ilike(f"%{product_name.strip()}%"))
    if hide_zero_safety_stock:
        stmt = stmt.where(Product.safety_stock.is_not(None), Product.safety_stock != 0)
    if stock_status == "low":
        stmt = stmt.where(Product.safety_stock.is_not(None), ProductInventory.stock_qty <= Product.safety_stock)
    elif stock_status == "normal":
        stmt = stmt.where(Product.safety_stock.is_not(None), ProductInventory.stock_qty > Product.safety_stock)
    return stmt


def _inventory_count_query(
    product_code: str | None = None,
    product_name: str | None = None,
    stock_status: str | None = None,
    hide_zero_safety_stock: bool = False,
):
    stmt = select(func.count(ProductInventory.id)).select_from(ProductInventory).join(Product, Product.id == ProductInventory.product_id)
    if product_code:
        stmt = stmt.where(Product.product_code.ilike(f"%{product_code.strip()}%"))
    if product_name:
        stmt = stmt.where(Product.internal_name.ilike(f"%{product_name.strip()}%"))
    if hide_zero_safety_stock:
        stmt = stmt.where(Product.safety_stock.is_not(None), Product.safety_stock != 0)
    if stock_status == "low":
        stmt = stmt.where(Product.safety_stock.is_not(None), ProductInventory.stock_qty <= Product.safety_stock)
    elif stock_status == "normal":
        stmt = stmt.where(Product.safety_stock.is_not(None), ProductInventory.stock_qty > Product.safety_stock)
    return stmt


def _apply_inventory_payload(row: ProductInventory, product: Product, payload: InventoryUpsertRequest, username: str) -> None:
    row.product_id = product.id
    row.product_name = product.internal_name
    row.stock_qty = max(0, int(payload.stock_qty or 0))
    row.last_count_qty = max(0, int(payload.last_count_qty or 0))
    if payload.safety_stock is not None:
        product.safety_stock = payload.safety_stock
    row.remark = (payload.remark or "").strip()
    row.updated_by = username
    row.updated_at = datetime.utcnow()


AUTO_PURCHASE_ORDER_REMARKS = {
    "订单列表转入配货中自动生成",
    "订单列表确认已打印自动生成",
    "定时任务自动生成",
}


def _purchase_order_remark_for_display(row: PurchaseOrder) -> str:
    remark = (row.remark or "").strip()
    return "" if remark in AUTO_PURCHASE_ORDER_REMARKS else remark


def _purchase_order_dto(row: PurchaseOrder) -> PurchaseOrderDto:
    return PurchaseOrderDto(
        id=row.id,
        purchase_no=row.purchase_no,
        purchase_date=_date_iso(row.purchase_date or row.created_at),
        source_count=row.source_count or 0,
        item_count=row.item_count or 0,
        total_required_qty=row.total_required_qty or 0,
        created_by=row.created_by,
        remark=_purchase_order_remark_for_display(row),
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _purchase_item_dto(row: PurchaseOrderItem) -> PurchaseOrderItemDto:
    return PurchaseOrderItemDto(
        id=row.id,
        product_id=row.product_id,
        product_name=row.product_name,
        required_qty=row.required_qty or 0,
        buyer_user_id=row.buyer_user_id,
        buyer=row.buyer or "",
        total_cost_record=float(row.total_cost_record) if row.total_cost_record is not None else None,
        purchase_cost=float(row.purchase_cost) if row.purchase_cost is not None else None,
        purchase_channel=row.purchase_channel or "",
        purchase_qty=row.purchase_qty or 0,
        remark=row.remark or "",
        created_at=_iso(row.created_at),
        updated_at=_iso(row.updated_at),
    )


def _purchase_source_dto(row: PurchaseOrderSource) -> PurchaseOrderSourceDto:
    return PurchaseOrderSourceDto(
        id=row.id,
        order_id=row.order_id,
        order_item_id=row.order_item_id,
        product_id=row.product_id,
        product_name=row.product_name,
        quantity=row.quantity or 0,
        created_at=_iso(row.created_at),
    )


def _purchase_order_detail_dto(row: PurchaseOrder) -> PurchaseOrderDetailDto:
    base = _purchase_order_dto(row).model_dump()
    lock = row_lock = None
    if row:
        row_lock = None
    return PurchaseOrderDetailDto(
        **base,
        lock_acquired=False,
        lock_owner="",
        lock_expires_at=None,
        items=[_purchase_item_dto(item) for item in sorted(row.items, key=lambda item: item.id)],
        sources=[_purchase_source_dto(source) for source in sorted(row.sources, key=lambda source: source.id)],
    )


PURCHASE_EDIT_LOCK_MINUTES = 30


def _cleanup_expired_purchase_order_locks(db: Session) -> None:
    now = datetime.utcnow()
    db.query(PurchaseOrderEditLock).filter(PurchaseOrderEditLock.expires_at <= now).delete(synchronize_session=False)
    db.flush()


def _current_purchase_order_lock(db: Session, purchase_order_id: int) -> PurchaseOrderEditLock | None:
    _cleanup_expired_purchase_order_locks(db)
    return db.scalar(select(PurchaseOrderEditLock).where(PurchaseOrderEditLock.purchase_order_id == purchase_order_id))


def _purchase_lock_dto(purchase_order_id: int, lock: PurchaseOrderEditLock | None, acquired: bool, message: str) -> PurchaseOrderEditLockDto:
    return PurchaseOrderEditLockDto(
        purchase_order_id=purchase_order_id,
        lock_acquired=acquired,
        lock_owner=lock.locked_by if lock else "",
        lock_expires_at=_iso(lock.expires_at) if lock else None,
        message=message,
    )


def _acquire_purchase_order_lock(db: Session, purchase_order_id: int, username: str) -> PurchaseOrderEditLockDto:
    lock = _current_purchase_order_lock(db, purchase_order_id)
    now = datetime.utcnow()
    expires_at = now + timedelta(minutes=PURCHASE_EDIT_LOCK_MINUTES)
    if lock and lock.locked_by != username:
        return _purchase_lock_dto(purchase_order_id, lock, False, f"采购单当前正在由 {lock.locked_by} 编辑")
    if not lock:
        lock = PurchaseOrderEditLock(
            purchase_order_id=purchase_order_id,
            locked_by=username,
            locked_at=now,
            expires_at=expires_at,
        )
        db.add(lock)
    else:
        lock.locked_at = now
        lock.expires_at = expires_at
        lock.updated_at = now
    db.commit()
    db.refresh(lock)
    return _purchase_lock_dto(purchase_order_id, lock, True, "已进入编辑状态")


def _release_purchase_order_lock(db: Session, purchase_order_id: int, username: str, *, allow_missing: bool = True) -> None:
    lock = _current_purchase_order_lock(db, purchase_order_id)
    if not lock:
        if allow_missing:
            return
        raise HTTPException(status_code=409, detail="采购单编辑锁不存在")
    if lock.locked_by != username:
        raise HTTPException(status_code=409, detail=f"采购单当前由 {lock.locked_by} 编辑")
    db.delete(lock)
    db.commit()


def _ensure_purchase_order_lock_owner(db: Session, purchase_order_id: int, username: str) -> None:
    lock = _current_purchase_order_lock(db, purchase_order_id)
    if not lock:
        raise HTTPException(status_code=409, detail="采购单未进入编辑状态，请重新打开")
    if lock.locked_by != username:
        raise HTTPException(status_code=409, detail=f"采购单当前由 {lock.locked_by} 编辑")
    lock.expires_at = datetime.utcnow() + timedelta(minutes=PURCHASE_EDIT_LOCK_MINUTES)
    lock.updated_at = datetime.utcnow()
    db.flush()


def _next_purchase_no(db: Session) -> str:
    prefix = f"PO{_local_now().strftime('%Y%m%d')}"
    latest_order_no = db.scalar(
        select(PurchaseOrder.purchase_no)
        .where(PurchaseOrder.purchase_no.like(f"{prefix}-%"))
        .order_by(desc(PurchaseOrder.purchase_no))
        .limit(1)
    )
    latest_log_no = db.scalar(
        select(PurchaseOrderLog.purchase_no)
        .where(PurchaseOrderLog.purchase_no.like(f"{prefix}-%"))
        .order_by(desc(PurchaseOrderLog.purchase_no))
        .limit(1)
    )
    latest = max([value for value in [latest_order_no, latest_log_no] if value], default=None)
    next_number = 1
    if latest and "-" in latest:
        suffix = latest.rsplit("-", 1)[-1]
        if suffix.isdigit():
            next_number = int(suffix) + 1
    return f"{prefix}-{next_number:03d}"


def _order_item_product_rows(db: Session, order_item_ids: list[int]):
    mapping_choice = mapping_choice_for_order_item()
    product_id = func.coalesce(mapping_choice["exact_product"].id, mapping_choice["insensitive_product"].id)
    product_name = func.coalesce(mapping_choice["exact_product"].internal_name, mapping_choice["insensitive_product"].internal_name)
    product_cost = func.coalesce(mapping_choice["exact_product"].cost, mapping_choice["insensitive_product"].cost)
    buyer_user_id = func.coalesce(mapping_choice["exact_product"].buyer_user_id, mapping_choice["insensitive_product"].buyer_user_id)
    exact_buyer = aliased(LocalUser)
    insensitive_buyer = aliased(LocalUser)
    buyer_display_name = func.coalesce(exact_buyer.display_name, insensitive_buyer.display_name)
    buyer_username = func.coalesce(exact_buyer.username, insensitive_buyer.username)
    return db.execute(
        select(
            Order,
            OrderItem,
            product_id.label("product_id"),
            product_name.label("product_name"),
            product_cost.label("product_cost"),
            buyer_user_id.label("buyer_user_id"),
            buyer_display_name.label("buyer_display_name"),
            buyer_username.label("buyer_username"),
        )
        .join(OrderItem, OrderItem.order_id == Order.id)
        .outerjoin(
            PlatformAccount,
            (PlatformAccount.platform == Order.platform) & (PlatformAccount.account_id == Order.shop_id),
        )
        .outerjoin(mapping_choice["exact_mapping"], mapping_choice["exact_condition"])
        .outerjoin(mapping_choice["exact_product"], mapping_choice["exact_product"].id == mapping_choice["exact_mapping"].product_id)
        .outerjoin(mapping_choice["insensitive_mapping"], mapping_choice["insensitive_condition"])
        .outerjoin(
            mapping_choice["insensitive_product"],
            mapping_choice["insensitive_product"].id == mapping_choice["insensitive_mapping"].product_id,
        )
        .outerjoin(exact_buyer, exact_buyer.id == mapping_choice["exact_product"].buyer_user_id)
        .outerjoin(insensitive_buyer, insensitive_buyer.id == mapping_choice["insensitive_product"].buyer_user_id)
        .where(OrderItem.id.in_(order_item_ids))
    ).all()


def _create_purchase_order(db: Session, purchase_date: date, created_by: str | None = None, remark: str = "") -> PurchaseOrder:
    purchase = PurchaseOrder(
        purchase_no=_next_purchase_no(db),
        purchase_date=purchase_date,
        source_count=0,
        item_count=0,
        total_required_qty=0,
        created_by=created_by,
        remark=(remark or "").strip(),
    )
    db.add(purchase)
    db.flush()
    return purchase


def _purchase_item_group_key(product_name: str | None) -> str:
    return " ".join((product_name or "").replace("\u3000", " ").split()).strip()


def _join_distinct_purchase_text(values) -> str | None:
    parts: list[str] = []
    for value in values:
        text_value = str(value or "").strip()
        if text_value and text_value not in parts:
            parts.append(text_value)
    return "；".join(parts) if parts else None


def _summarize_purchase_order(purchase: PurchaseOrder, db: Session, now: datetime | None = None) -> None:
    now = now or datetime.utcnow()
    purchase.source_count = db.scalar(
        select(func.count(PurchaseOrderSource.id)).where(PurchaseOrderSource.purchase_order_id == purchase.id)
    ) or 0
    purchase.item_count = db.scalar(
        select(func.count(PurchaseOrderItem.id)).where(PurchaseOrderItem.purchase_order_id == purchase.id)
    ) or 0
    purchase.total_required_qty = db.scalar(
        select(func.coalesce(func.sum(PurchaseOrderItem.required_qty), 0)).where(PurchaseOrderItem.purchase_order_id == purchase.id)
    ) or 0
    purchase.updated_at = now


def _append_purchase_order_items(
    db: Session,
    purchase: PurchaseOrder,
    rows,
    operator: str | None,
) -> int:
    grouped_items: dict[str, dict] = {}
    for order, item, product_id, product_name, product_cost, buyer_user_id, buyer_display_name, buyer_username in sorted(
        rows,
        key=lambda value: (value[3] or "", value[1].id),
    ):
        key = _purchase_item_group_key(product_name)
        if not key:
            raise ValueError(f"存在产品中文名称为空的明细: {item.id}")
        grouped = grouped_items.setdefault(
            key,
            {
                "product_id": product_id,
                "product_name": key,
                "total_cost_record": product_cost,
                "buyer_user_id": buyer_user_id,
                "buyer": buyer_display_name or buyer_username or "",
                "required_qty": 0,
                "sources": [],
            },
        )
        quantity = max(0, int(item.quantity or 1))
        grouped["required_qty"] += quantity
        grouped["sources"].append((order, item, product_id, key, quantity))

    for grouped in grouped_items.values():
        item_row = PurchaseOrderItem(
            purchase_order_id=purchase.id,
            product_id=grouped["product_id"],
            product_name=grouped["product_name"],
            required_qty=grouped["required_qty"],
            total_cost_record=grouped["total_cost_record"],
            buyer_user_id=grouped["buyer_user_id"],
            buyer=grouped["buyer"],
            purchase_qty=grouped["required_qty"],
        )
        db.add(item_row)
        db.flush()
        for order, item, source_product_id, source_product_name, quantity in grouped["sources"]:
            db.add(
                PurchaseOrderSource(
                    purchase_order_id=purchase.id,
                    purchase_order_item_id=item_row.id,
                    order_id=order.id,
                    order_item_id=item.id,
                    product_id=source_product_id,
                    product_name=source_product_name,
                    quantity=quantity,
                )
            )
    db.flush()
    _summarize_purchase_order(purchase, db)
    db.flush()
    return len(grouped_items)


def _merge_duplicate_purchase_order_items(db: Session) -> int:
    rows = db.scalars(
        select(PurchaseOrderItem).order_by(
            asc(PurchaseOrderItem.purchase_order_id),
            asc(PurchaseOrderItem.product_name),
            asc(PurchaseOrderItem.id),
        )
    ).all()
    grouped_items: dict[tuple[int, str], list[PurchaseOrderItem]] = {}
    for row in rows:
        key = _purchase_item_group_key(row.product_name)
        if not key:
            continue
        grouped_items.setdefault((row.purchase_order_id, key), []).append(row)

    merged_count = 0
    changed_purchase_ids: set[int] = set()
    now = datetime.utcnow()
    for (purchase_order_id, product_name), items in grouped_items.items():
        if len(items) <= 1:
            continue
        primary = items[0]
        primary.product_name = product_name
        primary.required_qty = sum(max(0, int(item.required_qty or 0)) for item in items)
        primary.purchase_qty = sum(max(0, int(item.purchase_qty or 0)) for item in items)
        primary.purchase_channel = _join_distinct_purchase_text(item.purchase_channel for item in items)
        primary.remark = _join_distinct_purchase_text(item.remark for item in items)
        primary.updated_at = now
        db.query(PurchaseOrderSource).filter(
            PurchaseOrderSource.purchase_order_item_id.in_([item.id for item in items])
        ).update({PurchaseOrderSource.product_name: product_name}, synchronize_session=False)
        for duplicate in items[1:]:
            if primary.product_id is None and duplicate.product_id is not None:
                primary.product_id = duplicate.product_id
            if primary.buyer_user_id is None and duplicate.buyer_user_id is not None:
                primary.buyer_user_id = duplicate.buyer_user_id
            if not primary.buyer and duplicate.buyer:
                primary.buyer = duplicate.buyer
            if primary.total_cost_record is None and duplicate.total_cost_record is not None:
                primary.total_cost_record = duplicate.total_cost_record
            if primary.purchase_cost is None and duplicate.purchase_cost is not None:
                primary.purchase_cost = duplicate.purchase_cost
            db.query(PurchaseOrderSource).filter(
                PurchaseOrderSource.purchase_order_item_id == duplicate.id
            ).update(
                {
                    PurchaseOrderSource.purchase_order_item_id: primary.id,
                    PurchaseOrderSource.product_name: product_name,
                },
                synchronize_session=False,
            )
            db.delete(duplicate)
            merged_count += 1
        changed_purchase_ids.add(purchase_order_id)

    db.flush()
    for purchase_order_id in changed_purchase_ids:
        purchase = db.get(PurchaseOrder, purchase_order_id)
        if purchase:
            _summarize_purchase_order(purchase, db, now)
    if merged_count:
        db.flush()
    return merged_count


def _generate_or_append_purchase_order_for_item_ids(
    db: Session,
    order_item_ids: list[int],
    operator: str | None,
    remark: str = "",
) -> tuple[PurchaseOrder, bool]:
    if not order_item_ids:
        raise ValueError("请先选择订单明细")
    existing_sources = db.scalars(
        select(PurchaseOrderSource.order_item_id).where(PurchaseOrderSource.order_item_id.in_(order_item_ids))
    ).all()
    if existing_sources:
        raise ValueError(f"存在已生成采购单的明细: {', '.join(map(str, existing_sources[:20]))}")
    rows = _order_item_product_rows(db, order_item_ids)
    if len(rows) != len(order_item_ids):
        found = {item.id for _, item, _, _, _, _, _, _ in rows}
        missing = [item_id for item_id in order_item_ids if item_id not in found]
        raise ValueError(f"订单明细不存在: {', '.join(map(str, missing[:20]))}")
    pending_items = [item.id for order, item, _, _, _, _, _, _ in rows if _derive_order_status(order) in ORDER_SUMMARY_EXCLUDED_STATUSES]
    if pending_items:
        raise ValueError(f"待处理订单不能生成采购单: {', '.join(map(str, pending_items[:20]))}")
    empty_product_items = [item.id for _, item, _, product_name, _, _, _, _ in rows if not product_name]
    if empty_product_items:
        raise ValueError(f"存在产品中文名称为空的明细: {', '.join(map(str, empty_product_items[:20]))}")

    purchase = _create_purchase_order(db, _local_today(), operator, remark)
    _append_purchase_order_items(db, purchase, rows, operator)
    _log_purchase_order(db, purchase, "create", operator)
    return purchase, True


def _purchase_snapshot(row: PurchaseOrder) -> dict:
    return {
        "purchase_order": _purchase_order_dto(row).model_dump(),
        "items": [_purchase_item_dto(item).model_dump() for item in row.items],
        "sources": [_purchase_source_dto(source).model_dump() for source in row.sources],
    }


def _log_purchase_order(db: Session, row: PurchaseOrder, action: str, operator: str | None, snapshot: dict | None = None) -> None:
    db.add(
        PurchaseOrderLog(
            purchase_order_id=row.id,
            purchase_no=row.purchase_no,
            action=action,
            operator=operator,
            snapshot=snapshot or _purchase_snapshot(row),
        )
    )


@app.post("/api/v1/purchase-orders/generate", response_model=PurchaseOrderDetailDto, status_code=201)
def generate_purchase_order(
    payload: PurchaseOrderGenerateRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    order_item_ids = list(dict.fromkeys(payload.order_item_ids or []))
    try:
        purchase, _ = _generate_or_append_purchase_order_for_item_ids(db, order_item_ids, user.username, payload.remark)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    db.commit()
    enqueue_purchase_order_wecom_notification(purchase.id, source="purchase_order_generate")
    row = db.scalar(
        select(PurchaseOrder)
        .options(joinedload(PurchaseOrder.items), joinedload(PurchaseOrder.sources))
        .where(PurchaseOrder.id == purchase.id)
    )
    return _purchase_order_detail_dto(row)


@app.get("/api/v1/purchase-orders", response_model=PurchaseOrderListResponse)
def list_purchase_orders(
    purchase_no: str | None = None,
    purchase_start: str | None = Query(default=None),
    purchase_end: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    stmt = select(PurchaseOrder)
    count_stmt = select(func.count(PurchaseOrder.id))
    if purchase_no:
        condition = PurchaseOrder.purchase_no.ilike(f"%{purchase_no.strip()}%")
        stmt = stmt.where(condition)
        count_stmt = count_stmt.where(condition)
    if purchase_start:
        try:
            start_date = date.fromisoformat(purchase_start)
        except ValueError:
            raise HTTPException(status_code=400, detail="采购开始日期格式无效")
        condition = PurchaseOrder.purchase_date >= start_date
        stmt = stmt.where(condition)
        count_stmt = count_stmt.where(condition)
    if purchase_end:
        try:
            end_date = date.fromisoformat(purchase_end)
        except ValueError:
            raise HTTPException(status_code=400, detail="采购结束日期格式无效")
        condition = PurchaseOrder.purchase_date <= end_date
        stmt = stmt.where(condition)
        count_stmt = count_stmt.where(condition)
    total = db.scalar(count_stmt) or 0
    rows = db.scalars(
        stmt.order_by(desc(PurchaseOrder.created_at), desc(PurchaseOrder.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return PurchaseOrderListResponse(
        items=[_purchase_order_dto(row) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.get("/api/v1/purchase-orders/export")
def export_purchase_orders(
    purchase_no: str | None = None,
    purchase_start: str | None = Query(default=None),
    purchase_end: str | None = Query(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    stmt = select(PurchaseOrder)
    if purchase_no:
        stmt = stmt.where(PurchaseOrder.purchase_no.ilike(f"%{purchase_no.strip()}%"))
    if purchase_start:
        try:
            start_date = date.fromisoformat(purchase_start)
        except ValueError:
            raise HTTPException(status_code=400, detail="采购开始日期格式无效")
        stmt = stmt.where(PurchaseOrder.purchase_date >= start_date)
    if purchase_end:
        try:
            end_date = date.fromisoformat(purchase_end)
        except ValueError:
            raise HTTPException(status_code=400, detail="采购结束日期格式无效")
        stmt = stmt.where(PurchaseOrder.purchase_date <= end_date)
    rows = db.scalars(stmt.order_by(desc(PurchaseOrder.created_at), desc(PurchaseOrder.id))).all()
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "采购单列表"
    worksheet.append(["采购单号", "采购日期", "来源明细数", "采购明细数", "总需求数量", "创建人", "创建时间", "备注"])
    for row in rows:
        dto = _purchase_order_dto(row)
        worksheet.append([
            dto.purchase_no,
            dto.purchase_date or "",
            dto.source_count,
            dto.item_count,
            dto.total_required_qty,
            dto.created_by or "",
            _excel_datetime(dto.created_at),
            dto.remark or "",
        ])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"purchase-orders-{_local_now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v1/purchase-orders/{purchase_order_id:int}", response_model=PurchaseOrderDetailDto)
def get_purchase_order(purchase_order_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.scalar(
        select(PurchaseOrder)
        .options(joinedload(PurchaseOrder.items), joinedload(PurchaseOrder.sources))
        .where(PurchaseOrder.id == purchase_order_id)
    )
    if not row:
        raise HTTPException(status_code=404, detail="采购单不存在")
    return _purchase_order_detail_dto(row)


@app.post("/api/v1/purchase-orders/{purchase_order_id:int}/lock", response_model=PurchaseOrderEditLockDto)
def acquire_purchase_order_lock(
    purchase_order_id: int,
    _: PurchaseOrderEditLockRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = db.get(PurchaseOrder, purchase_order_id)
    if not row:
        raise HTTPException(status_code=404, detail="采购单不存在")
    return _acquire_purchase_order_lock(db, purchase_order_id, user.username)


@app.delete("/api/v1/purchase-orders/{purchase_order_id:int}/lock")
def release_purchase_order_lock(purchase_order_id: int, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)) -> dict:
    row = db.get(PurchaseOrder, purchase_order_id)
    if not row:
        raise HTTPException(status_code=404, detail="采购单不存在")
    _release_purchase_order_lock(db, purchase_order_id, user.username)
    return {"ok": True}


@app.put("/api/v1/purchase-orders/{purchase_order_id:int}", response_model=PurchaseOrderDto)
def update_purchase_order(
    purchase_order_id: int,
    payload: PurchaseOrderUpdateRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = db.get(PurchaseOrder, purchase_order_id)
    if not row:
        raise HTTPException(status_code=404, detail="采购单不存在")
    _ensure_purchase_order_lock_owner(db, purchase_order_id, user.username)
    if payload.purchase_date is not None:
        if payload.purchase_date:
            try:
                row.purchase_date = date.fromisoformat(payload.purchase_date)
            except ValueError:
                raise HTTPException(status_code=400, detail="采购日期格式无效")
        else:
            row.purchase_date = None
    row.remark = (payload.remark or "").strip()
    row.updated_at = datetime.utcnow()
    _log_purchase_order(db, row, "update", user.username)
    db.commit()
    db.refresh(row)
    return _purchase_order_dto(row)


@app.put("/api/v1/purchase-orders/{purchase_order_id:int}/items/{item_id:int}", response_model=PurchaseOrderItemDto)
def update_purchase_order_item(
    purchase_order_id: int,
    item_id: int,
    payload: PurchaseOrderItemUpdateRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    item = db.scalar(
        select(PurchaseOrderItem).where(
            PurchaseOrderItem.id == item_id,
            PurchaseOrderItem.purchase_order_id == purchase_order_id,
        )
    )
    if not item:
        raise HTTPException(status_code=404, detail="采购单明细不存在")
    _ensure_purchase_order_lock_owner(db, purchase_order_id, user.username)
    purchase = db.get(PurchaseOrder, purchase_order_id)
    buyer_user = _buyer_user_by_payload(db, payload.buyer_user_id)
    item.buyer_user_id = buyer_user.id if buyer_user else None
    item.buyer = _user_display_name(buyer_user)
    item.total_cost_record = _decimal_or_none(payload.total_cost_record)
    item.purchase_cost = _decimal_or_none(payload.purchase_cost)
    item.purchase_channel = (payload.purchase_channel or "").strip()
    item.purchase_qty = max(0, int(payload.purchase_qty or 0))
    item.remark = (payload.remark or "").strip()
    item.updated_at = datetime.utcnow()
    if buyer_user and item.product_id:
        product = db.get(Product, item.product_id)
        if product:
            product.buyer_user_id = buyer_user.id
            product.updated_at = datetime.utcnow()
    if purchase:
        purchase.updated_at = datetime.utcnow()
        _log_purchase_order(db, purchase, "update_item", user.username)
    db.commit()
    db.refresh(item)
    return _purchase_item_dto(item)


@app.delete("/api/v1/purchase-orders/{purchase_order_id:int}")
def delete_purchase_order(purchase_order_id: int, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.scalar(
        select(PurchaseOrder)
        .options(joinedload(PurchaseOrder.items), joinedload(PurchaseOrder.sources))
        .where(PurchaseOrder.id == purchase_order_id)
    )
    if not row:
        raise HTTPException(status_code=404, detail="采购单不存在")
    _ensure_purchase_order_lock_owner(db, purchase_order_id, user.username)
    snapshot = _purchase_snapshot(row)
    source_order_ids = [source.order_id for source in row.sources]
    operator = operator_name(user)
    db.add(
        PurchaseOrderLog(
            purchase_order_id=row.id,
            purchase_no=row.purchase_no,
            action="delete",
            operator=user.username,
            snapshot=snapshot,
        )
    )
    db.delete(row)
    db.flush()
    rollback_rows = _rollback_orders_to_waiting_purchase_after_purchase_delete(
        db,
        source_order_ids,
        row,
        operator,
    )
    db.commit()
    message = "已删除"
    if rollback_rows:
        message += f"，已回到待采购 {len(rollback_rows)} 条"
    return {"message": message}


@app.get("/api/v1/purchase-orders/{purchase_order_id:int}/export")
def export_purchase_order(purchase_order_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.scalar(
        select(PurchaseOrder)
        .options(joinedload(PurchaseOrder.items), joinedload(PurchaseOrder.sources))
        .where(PurchaseOrder.id == purchase_order_id)
    )
    if not row:
        raise HTTPException(status_code=404, detail="采购单不存在")
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = row.purchase_no
    worksheet.append(["配货日", "产品名称", "需求数量", "采购人", "总表成本记录", "采购成本", "采购渠道", "实际采购数量", "备注"])
    for item in sorted(row.items, key=lambda value: value.product_name):
        worksheet.append([
            "",
            item.product_name,
            item.required_qty,
            item.buyer or "",
            item.total_cost_record if item.total_cost_record is not None else "",
            item.purchase_cost if item.purchase_cost is not None else "",
            item.purchase_channel or "",
            item.purchase_qty,
            item.remark or "",
        ])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{row.purchase_no}.xlsx"'},
    )


def _date_iso(value) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return (value + LOCAL_TIME_OFFSET).date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _excel_datetime(value) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        text_value = value.strip()
        parse_value = text_value[:-1] if text_value.endswith("Z") else text_value
        try:
            value = datetime.fromisoformat(parse_value)
        except ValueError:
            return text_value
    if isinstance(value, datetime):
        return (value.replace(tzinfo=None) + LOCAL_TIME_OFFSET).replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _excel_platform_datetime(value) -> str:
    text_value = _platform_time_text(value)
    if not text_value:
        return ""
    return text_value.replace("T", " ")


def _date_filter_bounds(start_value: str | None, end_value: str | None) -> tuple[datetime | None, datetime | None]:
    start_dt = None
    end_dt = None
    try:
        if start_value:
            start_dt = datetime.combine(date.fromisoformat(start_value), time.min)
        if end_value:
            end_dt = datetime.combine(date.fromisoformat(end_value) + timedelta(days=1), time.min)
    except (TypeError, ValueError):
        return None, None
    return start_dt, end_dt


def _build_purchase_details_query(
    purchase_no: str | None,
    product_name: str | None,
    buyer: str | None,
    purchase_start: str | None,
    purchase_end: str | None,
    picking_start: str | None,
    picking_end: str | None,
):
    purchase_date_expr = func.coalesce(PurchaseOrder.purchase_date, func.date(PurchaseOrder.created_at))
    picking_date_expr = func.date(Order.picking_at + text("INTERVAL '8 hours'"))
    daily_qty_expr = func.coalesce(func.sum(PurchaseOrderSource.quantity), 0)
    stock_qty_expr = func.coalesce(ProductInventory.stock_qty, 0)
    inventory_join = or_(
        and_(PurchaseOrderItem.product_id.isnot(None), ProductInventory.product_id == PurchaseOrderItem.product_id),
        and_(PurchaseOrderItem.product_id.is_(None), ProductInventory.product_name == PurchaseOrderItem.product_name),
    )
    stmt = (
        select(
            PurchaseOrder.id.label("purchase_order_id"),
            PurchaseOrderItem.id.label("item_id"),
            PurchaseOrder.purchase_no.label("purchase_no"),
            purchase_date_expr.label("purchase_date"),
            picking_date_expr.label("picking_date"),
            PurchaseOrderItem.product_name.label("product_name"),
            daily_qty_expr.label("daily_order_qty"),
            stock_qty_expr.label("stock_qty"),
            (daily_qty_expr - stock_qty_expr).label("pending_purchase_qty"),
            PurchaseOrderItem.buyer_user_id.label("buyer_user_id"),
            PurchaseOrderItem.buyer.label("buyer"),
            PurchaseOrderItem.total_cost_record.label("total_cost_record"),
            PurchaseOrderItem.purchase_cost.label("purchase_cost"),
            PurchaseOrderItem.purchase_channel.label("purchase_channel"),
            PurchaseOrderItem.purchase_qty.label("purchase_qty"),
            PurchaseOrderItem.remark.label("remark"),
        )
        .select_from(PurchaseOrder)
        .join(PurchaseOrderItem, PurchaseOrderItem.purchase_order_id == PurchaseOrder.id)
        .join(
            PurchaseOrderSource,
            and_(
                PurchaseOrderSource.purchase_order_id == PurchaseOrder.id,
                PurchaseOrderSource.purchase_order_item_id == PurchaseOrderItem.id,
            ),
        )
        .join(Order, Order.id == PurchaseOrderSource.order_id)
        .outerjoin(ProductInventory, inventory_join)
    )
    if purchase_no:
        stmt = stmt.where(PurchaseOrder.purchase_no.ilike(f"%{purchase_no.strip()}%"))
    if product_name:
        stmt = stmt.where(PurchaseOrderItem.product_name.ilike(f"%{product_name.strip()}%"))
    if buyer:
        stmt = stmt.where(PurchaseOrderItem.buyer.ilike(f"%{buyer.strip()}%"))
    purchase_start_dt, purchase_end_dt = _date_filter_bounds(purchase_start, purchase_end)
    if purchase_start_dt:
        stmt = stmt.where(purchase_date_expr >= purchase_start_dt.date())
    if purchase_end_dt:
        stmt = stmt.where(purchase_date_expr < purchase_end_dt.date())
    picking_start_dt = None
    picking_end_dt = None
    if picking_start:
        picking_start_dt, _ = _local_day_utc_bounds(picking_start)
    if picking_end:
        _, picking_end_dt = _local_day_utc_bounds(picking_end)
    if picking_start_dt:
        stmt = stmt.where(Order.picking_at >= picking_start_dt)
    if picking_end_dt:
        stmt = stmt.where(Order.picking_at < picking_end_dt)
    return stmt.group_by(
        PurchaseOrder.id,
        PurchaseOrderItem.id,
        PurchaseOrder.purchase_no,
        purchase_date_expr,
        picking_date_expr,
        PurchaseOrderItem.product_id,
        ProductInventory.stock_qty,
        PurchaseOrderItem.product_name,
        PurchaseOrderItem.buyer_user_id,
        PurchaseOrderItem.buyer,
        PurchaseOrderItem.total_cost_record,
        PurchaseOrderItem.purchase_cost,
        PurchaseOrderItem.purchase_channel,
        PurchaseOrderItem.purchase_qty,
        PurchaseOrderItem.remark,
    ).order_by(desc(purchase_date_expr), desc(PurchaseOrder.purchase_no), asc(picking_date_expr), asc(PurchaseOrderItem.product_name))


def _purchase_detail_dto(row) -> PurchaseDetailDto:
    daily_order_qty = int(row.daily_order_qty or 0)
    stock_qty = int(row.stock_qty or 0)
    pending_purchase_qty = int(row.pending_purchase_qty if row.pending_purchase_qty is not None else daily_order_qty - stock_qty)
    return PurchaseDetailDto(
        purchase_order_id=row.purchase_order_id,
        item_id=row.item_id,
        purchase_no=row.purchase_no or "",
        purchase_date=_date_iso(row.purchase_date),
        picking_date=_date_iso(row.picking_date),
        product_name=row.product_name or "",
        daily_order_qty=daily_order_qty,
        stock_qty=stock_qty,
        pending_purchase_qty=pending_purchase_qty,
        buyer_user_id=row.buyer_user_id,
        buyer=row.buyer or "",
        total_cost_record=float(row.total_cost_record) if row.total_cost_record is not None else None,
        purchase_cost=float(row.purchase_cost) if row.purchase_cost is not None else None,
        purchase_channel=row.purchase_channel or "",
        purchase_qty=int(row.purchase_qty or 0),
        remark=row.remark or "",
    )


@app.get("/api/v1/purchase-details", response_model=PurchaseDetailListResponse)
def list_purchase_details(
    purchase_no: str | None = None,
    product_name: str | None = None,
    buyer: str | None = None,
    purchase_start: str | None = Query(default=None),
    purchase_end: str | None = Query(default=None),
    picking_start: str | None = Query(default=None),
    picking_end: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    stmt = _build_purchase_details_query(purchase_no, product_name, buyer, purchase_start, purchase_end, picking_start, picking_end)
    total = db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery())) or 0
    rows = db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return PurchaseDetailListResponse(
        items=[_purchase_detail_dto(row) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.get("/api/v1/purchase-details/export")
def export_purchase_details(
    purchase_no: str | None = None,
    product_name: str | None = None,
    buyer: str | None = None,
    purchase_start: str | None = Query(default=None),
    purchase_end: str | None = Query(default=None),
    picking_start: str | None = Query(default=None),
    picking_end: str | None = Query(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    stmt = _build_purchase_details_query(purchase_no, product_name, buyer, purchase_start, purchase_end, picking_start, picking_end)
    rows = [_purchase_detail_dto(row) for row in db.execute(stmt).all()]
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "采购明细表"
    worksheet.append([
        "采购单号",
        "采购日期",
        "配货日",
        "产品名称",
        "采购数量(当日来单)",
        "库存数",
        "待采购数量",
        "采购人",
        "总表成本记录",
        "采购成本",
        "采购渠道",
        "采购数量",
        "备注",
    ])
    for row in rows:
        worksheet.append([
            row.purchase_no,
            row.purchase_date or "",
            row.picking_date or "",
            row.product_name,
            row.daily_order_qty,
            row.stock_qty,
            row.pending_purchase_qty,
            row.buyer,
            row.total_cost_record if row.total_cost_record is not None else "",
            row.purchase_cost if row.purchase_cost is not None else "",
            row.purchase_channel,
            row.purchase_qty,
            row.remark,
        ])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"purchase-details-{_local_now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v1/inventory", response_model=InventoryListResponse)
def list_inventory(
    product_code: str | None = None,
    product_name: str | None = None,
    stock_status: str | None = None,
    hide_zero_safety_stock: bool = False,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    stmt = _inventory_query(product_code, product_name, stock_status, hide_zero_safety_stock)
    total = db.scalar(_inventory_count_query(product_code, product_name, stock_status, hide_zero_safety_stock)) or 0
    rows = db.execute(
        stmt.order_by(asc(Product.internal_name), asc(ProductInventory.id)).offset((page - 1) * page_size).limit(page_size)
    ).all()
    return InventoryListResponse(
        items=[_inventory_dto(row, product) for row, product in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.post("/api/v1/inventory", response_model=InventoryDto, status_code=201)
def create_inventory(
    payload: InventoryUpsertRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    product = db.get(Product, payload.product_id)
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")
    if db.scalar(select(ProductInventory).where(ProductInventory.product_id == product.id)):
        raise HTTPException(status_code=400, detail="该产品已有库存记录")
    row = ProductInventory(product_id=product.id, product_name=product.internal_name)
    _apply_inventory_payload(row, product, payload, user.username)
    db.add(row)
    db.commit()
    db.refresh(row)
    return _inventory_dto(row, product)


@app.put("/api/v1/inventory/{inventory_id:int}", response_model=InventoryDto)
def update_inventory(
    inventory_id: int,
    payload: InventoryUpsertRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = db.get(ProductInventory, inventory_id)
    if not row:
        raise HTTPException(status_code=404, detail="库存记录不存在")
    product = db.get(Product, payload.product_id)
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")
    conflict = db.scalar(
        select(ProductInventory).where(ProductInventory.product_id == product.id, ProductInventory.id != inventory_id)
    )
    if conflict:
        raise HTTPException(status_code=400, detail="该产品已有库存记录")
    _apply_inventory_payload(row, product, payload, user.username)
    db.commit()
    db.refresh(row)
    return _inventory_dto(row, product)


@app.delete("/api/v1/inventory/{inventory_id:int}")
def delete_inventory(inventory_id: int, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    row = db.get(ProductInventory, inventory_id)
    if not row:
        raise HTTPException(status_code=404, detail="库存记录不存在")
    db.delete(row)
    db.commit()
    return {"message": "已删除"}


@app.post("/api/v1/inventory/import")
async def import_inventory(file: UploadFile = File(...), user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    content = await file.read()
    headers, rows, first_data_row = _read_product_import_rows(file.filename or "", content)
    if not headers:
        raise HTTPException(status_code=400, detail="导入文件为空")
    name_col = _find_header(headers, {"产品名称", "中文名称", "product_name", "internal_name", "name"})
    stock_col = _find_header(headers, {"库存数量", "库存", "stock_qty", "stock"})
    last_count_col = _find_header(headers, {"上次盘点", "last_count_qty", "last_count"})
    safety_stock_col = _find_header(headers, {"安全库存", "安全库存数", "safety_stock", "safe_stock"})
    remark_col = _find_header(headers, {"备注", "remark"})
    if name_col is None:
        raise HTTPException(status_code=400, detail="缺少产品名称列")
    if safety_stock_col is None:
        raise HTTPException(status_code=400, detail="缺少安全库存列")

    products = db.scalars(select(Product)).all()
    product_map = {_normalize_inventory_name(product.internal_name): product for product in products}
    created = 0
    updated = 0
    errors: list[dict] = []
    for row_index, row in enumerate(rows, start=first_data_row):
        try:
            get_value = lambda col: row[col] if col is not None and col < len(row) else None
            name = _normalize_inventory_name(_clean_import_value(get_value(name_col)))
            product = product_map.get(name)
            if not product:
                raise HTTPException(status_code=400, detail="产品目录不存在")
            stock_qty_value = _clean_import_value(get_value(stock_col)) if stock_col is not None else ""
            stock_qty = _int_or_none(stock_qty_value) if stock_qty_value else None
            if stock_qty is not None and stock_qty < 0:
                raise HTTPException(status_code=400, detail="库存数量不能小于 0")
            last_count_qty = _int_or_none(_clean_import_value(get_value(last_count_col))) or 0
            safety_stock_value = _clean_import_value(get_value(safety_stock_col))
            if not safety_stock_value:
                raise HTTPException(status_code=400, detail="安全库存不能为空")
            safety_stock = _int_or_none(safety_stock_value)
            if safety_stock < 0:
                raise HTTPException(status_code=400, detail="安全库存不能小于 0")
            remark = _clean_import_value(get_value(remark_col)) if remark_col is not None else ""
            inventory = db.scalar(select(ProductInventory).where(ProductInventory.product_id == product.id))
            if stock_qty is not None:
                effective_stock_qty = stock_qty
            elif inventory:
                effective_stock_qty = inventory.stock_qty or 0
            else:
                effective_stock_qty = 0
            if inventory:
                updated += 1
            else:
                inventory = ProductInventory(product_id=product.id, product_name=product.internal_name)
                db.add(inventory)
                created += 1
            _apply_inventory_payload(
                inventory,
                product,
                InventoryUpsertRequest(
                    product_id=product.id,
                    stock_qty=effective_stock_qty,
                    last_count_qty=last_count_qty,
                    remark=remark,
                ),
                user.username,
            )
            product.safety_stock = safety_stock
        except Exception as exc:
            errors.append({"row": row_index, "message": getattr(exc, "detail", str(exc))})
            continue
    db.commit()
    return {"created": created, "updated": updated, "failed": len(errors), "errors": errors[:100]}


@app.get("/api/v1/inventory/import-template")
def download_inventory_import_template(_: LocalUser = Depends(current_user)):
    workbook = _build_inventory_import_template_workbook()
    filename = f"inventory_import_template_{_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _product_workbook_response(workbook, filename)


@app.get("/api/v1/inventory/export")
def export_inventory(
    product_code: str | None = None,
    product_name: str | None = None,
    stock_status: str | None = None,
    hide_zero_safety_stock: bool = False,
    columns: str | None = Query(default=None),
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    rows = db.execute(
        _inventory_query(product_code, product_name, stock_status, hide_zero_safety_stock).order_by(asc(Product.internal_name))
    ).all()
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "库存"
    export_columns = _table_export_columns(INVENTORY_TABLE_KEY, INVENTORY_PRIMARY_COLUMN_KEY, INVENTORY_EXPORT_COLUMNS, user, db, columns)
    worksheet.append([column["title"] for column in export_columns])
    for row, product in rows:
        dto = _inventory_dto(row, product)
        worksheet.append([_inventory_export_value(column["key"], dto) for column in export_columns])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"inventory_{_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/v1/products/import")
async def import_products(file: UploadFile = File(...), _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    content = await file.read()
    headers, rows, first_data_row = _read_product_import_rows(file.filename or "", content)
    if not headers:
        raise HTTPException(status_code=400, detail="导入文件为空")
    name_col = _find_header(headers, {"产品中文名", "中文名称", "internal_name", "product_name", "name"})
    english_name_col = _find_header(headers, {"产品英文名", "英文名称", "english_name", "english name"})
    cost_col = _find_header(headers, {"成本", "cost"})
    weight_col = _find_header(headers, {"净重", "重量", "weight", "net_weight"})
    gross_weight_col = _find_header(headers, {"毛重", "gross_weight"})
    package_length_col = _find_header(headers, {"包装长", "package_length", "length"})
    package_width_col = _find_header(headers, {"包装宽", "package_width", "width"})
    package_height_col = _find_header(headers, {"包装高", "package_height", "height"})
    ean_col = _find_header(headers, {"EAN", "ean"})
    description_col = _find_header(headers, {"描述", "description"})
    main_image_url_col = _find_header(headers, {"图片链接", "main_image_url", "image_url"})
    slow_moving_col = _find_header(headers, {"是否呆滞料", "is_slow_moving_material", "slow_moving"})
    stock_col = _find_header(headers, {"安全库存数", "库存数", "safety_stock", "safe_stock"})
    if name_col is None:
        raise HTTPException(status_code=400, detail="缺少产品中文名列")
    shops = _enabled_product_shops(db)
    shop_map = _shop_header_map(shops)
    buyer_col = _find_header(headers, {"采购人", "buyer", "purchaser"})
    ignored_names = {
        "编码",
        "产品编号",
        "产品中文名",
        "中文名称",
        "产品英文名",
        "英文名称",
        "采购人",
        "buyer",
        "purchaser",
        "成本",
        "净重",
        "重量",
        "毛重",
        "包装长",
        "包装宽",
        "包装高",
        "ean",
        "EAN",
        "描述",
        "图片链接",
        "是否呆滞料",
        "安全库存数",
        "库存数",
        "状态",
    }
    ignored_header_keys = {name.strip().lower() for name in ignored_names}
    shop_columns = {
        idx: shop_map[title.strip().lower()]
        for idx, title in enumerate(headers)
        if title and title.strip().lower() not in ignored_header_keys and title.strip().lower() in shop_map
    }
    grouped_rows: dict[str, dict] = {}
    errors: list[dict] = []
    user_lookup = _user_lookup_for_import(db)

    def bool_or_none(value) -> bool | None:
        text_value = _clean_import_value(value)
        if not text_value:
            return None
        normalized = text_value.strip().lower()
        if normalized in {"是", "true", "yes", "y", "1"}:
            return True
        if normalized in {"否", "false", "no", "n", "0"}:
            return False
        raise HTTPException(status_code=400, detail=f"是否呆滞料格式错误: {text_value}")

    for row_index, row in enumerate(rows, start=first_data_row):
        try:
            if not any(_clean_import_value(value) for value in row):
                continue
            get_value = lambda col: row[col] if col is not None and col < len(row) else None
            name = _normalize_product_name(_clean_import_value(get_value(name_col)))
            name_key = _product_name_match_key(name)
            grouped = grouped_rows.setdefault(
                name_key,
                {
                    "first_row": row_index,
                    "name": name,
                    "english_name": None,
                    "cost": None,
                    "weight": None,
                    "gross_weight": None,
                    "package_length": None,
                    "package_width": None,
                    "package_height": None,
                    "ean": None,
                    "description": None,
                    "main_image_url": None,
                    "is_slow_moving_material": None,
                    "stock": None,
                    "buyer_user_id": None,
                    "mappings": {},
                },
            )

            def merge_field(field: str, value) -> None:
                if value is None:
                    return
                if grouped[field] is None:
                    grouped[field] = value
                elif grouped[field] != value:
                    raise HTTPException(status_code=400, detail=f"{name} 的基础字段存在冲突，请检查导入行")

            if english_name_col is not None:
                english_name_value = _clean_import_value(get_value(english_name_col))
                if english_name_value:
                    merge_field("english_name", _normalize_optional_text(english_name_value, 255))
            if cost_col is not None:
                cost_value = _clean_import_value(get_value(cost_col))
                if cost_value:
                    merge_field("cost", _decimal_or_none(cost_value))
            if weight_col is not None:
                weight_value = _clean_import_value(get_value(weight_col))
                if weight_value:
                    merge_field("weight", _decimal_or_none(weight_value))
            if gross_weight_col is not None:
                gross_weight_value = _clean_import_value(get_value(gross_weight_col))
                if gross_weight_value:
                    merge_field("gross_weight", _decimal_or_none(gross_weight_value))
            if package_length_col is not None:
                package_length_value = _clean_import_value(get_value(package_length_col))
                if package_length_value:
                    merge_field("package_length", _decimal_or_none(package_length_value))
            if package_width_col is not None:
                package_width_value = _clean_import_value(get_value(package_width_col))
                if package_width_value:
                    merge_field("package_width", _decimal_or_none(package_width_value))
            if package_height_col is not None:
                package_height_value = _clean_import_value(get_value(package_height_col))
                if package_height_value:
                    merge_field("package_height", _decimal_or_none(package_height_value))
            if ean_col is not None:
                ean_value = _clean_import_value(get_value(ean_col))
                if ean_value:
                    merge_field("ean", _normalize_optional_text(ean_value, 64))
            if description_col is not None:
                description_value = _clean_import_value(get_value(description_col))
                if description_value:
                    merge_field("description", description_value)
            if main_image_url_col is not None:
                main_image_url_value = _clean_import_value(get_value(main_image_url_col))
                if main_image_url_value:
                    merge_field("main_image_url", main_image_url_value)
            if slow_moving_col is not None:
                slow_moving_value = bool_or_none(get_value(slow_moving_col))
                if slow_moving_value is not None:
                    merge_field("is_slow_moving_material", slow_moving_value)
            if stock_col is not None:
                stock_value = _clean_import_value(get_value(stock_col))
                if stock_value:
                    merge_field("stock", _int_or_none(stock_value))
            if buyer_col is not None:
                buyer_value = _clean_import_value(get_value(buyer_col))
                if buyer_value and buyer_value not in user_lookup:
                    raise HTTPException(status_code=400, detail=f"采购人不存在或已停用: {buyer_value}")
                if buyer_value:
                    merge_field("buyer_user_id", user_lookup[buyer_value].id)
            for col_idx, shop in shop_columns.items():
                value = _clean_import_value(get_value(col_idx))
                if value:
                    mapping_values = grouped["mappings"].setdefault(str(shop.id), [])
                    if value not in mapping_values:
                        mapping_values.append(value)
        except Exception as exc:
            errors.append({"row": row_index, "message": getattr(exc, "detail", str(exc))})
            continue

    if errors:
        return {"created": 0, "updated": 0, "mapping_updated": 0, "failed": len(errors), "errors": errors[:100]}

    created = 0
    updated = 0
    mapping_updated = 0
    existing_products = db.scalars(select(Product).options(selectinload(Product.mappings))).all()
    product_lookup: dict[str, Product] = {}
    duplicate_names: dict[str, list[str]] = {}
    for product in existing_products:
        key = _product_name_match_key(product.internal_name)
        if key in product_lookup:
            duplicate_names.setdefault(key, [product_lookup[key].internal_name]).append(product.internal_name)
        else:
            product_lookup[key] = product

    for name_key, grouped in grouped_rows.items():
        name = grouped["name"]
        product_created = False
        try:
            with db.begin_nested():
                if name_key in duplicate_names:
                    raise HTTPException(status_code=400, detail=f"存在多个中文名称仅大小写不同的产品，无法判断更新对象: {', '.join(duplicate_names[name_key])}")
                product = product_lookup.get(name_key)
                if not product:
                    product = Product(product_code=_generate_product_code(db), internal_name=name, enabled=True)
                    db.add(product)
                    db.flush()
                    product_created = True
                if grouped["english_name"] is not None:
                    product.english_name = grouped["english_name"]
                if grouped["cost"] is not None:
                    product.cost = grouped["cost"]
                if grouped["weight"] is not None:
                    product.weight = grouped["weight"]
                if grouped["gross_weight"] is not None:
                    product.gross_weight = grouped["gross_weight"]
                if grouped["package_length"] is not None:
                    product.package_length = grouped["package_length"]
                if grouped["package_width"] is not None:
                    product.package_width = grouped["package_width"]
                if grouped["package_height"] is not None:
                    product.package_height = grouped["package_height"]
                if grouped["ean"] is not None:
                    product.ean = grouped["ean"]
                if grouped["description"] is not None:
                    product.description = grouped["description"]
                if grouped["main_image_url"] is not None:
                    product.main_image_url = grouped["main_image_url"]
                if grouped["is_slow_moving_material"] is not None:
                    product.is_slow_moving_material = grouped["is_slow_moving_material"]
                if grouped["stock"] is not None:
                    product.safety_stock = grouped["stock"]
                if grouped["buyer_user_id"] is not None:
                    product.buyer_user_id = grouped["buyer_user_id"]
                _upsert_product_mappings(db, product, grouped["mappings"])
                db.flush()
            if product_created:
                created += 1
                product_lookup[name_key] = product
            else:
                updated += 1
            mapping_updated += sum(len(values) for values in grouped["mappings"].values())
        except Exception as exc:
            errors.append({"row": grouped["first_row"], "message": getattr(exc, "detail", str(exc))})
            continue
    db.commit()
    return {"created": created, "updated": updated, "mapping_updated": mapping_updated, "failed": len(errors), "errors": errors[:100]}


@app.get("/api/v1/products/export")
def export_products(
    keyword: str | None = None,
    product_code: str | None = None,
    internal_name: str | None = None,
    english_name: str | None = None,
    ean: str | None = None,
    shop_sku: str | None = None,
    enabled: bool | None = None,
    is_slow_moving_material: bool | None = None,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    shops = _enabled_product_shops(db)
    stmt = _product_list_base_query(
        product_code,
        internal_name,
        english_name,
        ean,
        shop_sku,
        enabled,
        is_slow_moving_material,
        keyword,
    )
    rows = db.scalars(
        stmt.options(selectinload(Product.mappings), joinedload(Product.buyer_user)).order_by(asc(Product.product_code))
    ).all()
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "产品目录"
    _append_product_export_header(worksheet, shops)
    for product in rows:
        for export_row in _product_export_rows(product, shops):
            worksheet.append(export_row)
    _style_product_workbook(worksheet, shops)
    filename = f"products_{_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _product_workbook_response(workbook, filename)


@app.get("/api/orders", response_model=list[OrderDto])
def list_orders_legacy(_: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    response = _query_orders(db, None, None, None, None, None, None, None, None, 1, 200)
    return response.items


@app.get("/api/v1/orders", response_model=OrderListResponse)
def list_orders_v1(
    status: str | None = Query(default=None),
    risk: str | None = Query(default=None),
    shop: str | None = Query(default=None),
    shop_ids: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    number: str | None = Query(default=None),
    transaction_id: str | None = Query(default=None),
    order_no: str | None = Query(default=None),
    product_keyword: str | None = Query(default=None),
    payment_time_range: str | None = Query(default=None),
    payment_start: str | None = Query(default=None),
    payment_end: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    if page_size not in {50, 100, 500}:
        page_size = 50
    return _query_orders(
        db,
        status,
        platform,
        transaction_id,
        order_no,
        number,
        payment_time_range,
        payment_start,
        payment_end,
        page,
        page_size,
        product_keyword,
        risk_filter=risk,
        shop=shop,
        shop_keys=_shop_keys_from_ids(db, shop_ids),
    )


@app.post("/api/v1/orders/search", response_model=OrderListResponse)
def search_orders_v1(
    payload: OrderSearchRequest,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    submitted_count, numbers = _normalize_batch_order_numbers(payload.numbers)
    page_size = payload.page_size if payload.page_size in {50, 100, 500} else 50
    return _query_orders(
        db=db,
        status_filter=payload.status,
        platform=payload.platform,
        transaction_id=None,
        order_no=None,
        number=None,
        payment_time_range=None,
        payment_start=payload.payment_start,
        payment_end=payload.payment_end,
        page=payload.page,
        page_size=page_size,
        product_keyword=payload.product_keyword,
        numbers=numbers,
        submitted_number_count=submitted_count,
        risk_filter=payload.risk,
        shop=payload.shop,
        shop_keys=_shop_keys_from_ids(db, payload.shop_ids),
    )


@app.get("/api/v1/orders/status-counts")
def order_status_counts(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    counts = {
        ORDER_STATUS_PENDING_KEY: 0,
        ORDER_STATUS_WAITING_PRINT_KEY: 0,
        ORDER_STATUS_WAITING_PURCHASE_KEY: 0,
        ORDER_STATUS_PICKING_KEY: 0,
        ORDER_STATUS_SHIPPED_KEY: 0,
        ORDER_STATUS_AWAITING_PICKUP_KEY: 0,
        ORDER_STATUS_DELIVERED_KEY: 0,
        ORDER_STATUS_VOIDED_KEY: 0,
    }
    local_mapping = {
        "shipped": ORDER_STATUS_SHIPPED,
        "awaiting_pickup": ORDER_STATUS_AWAITING_PICKUP,
        "delivered": ORDER_STATUS_DELIVERED,
        "voided": ORDER_STATUS_VOIDED,
        "cancelled": ORDER_STATUS_VOIDED,
        "shipment_creating": ORDER_STATUS_PICKING,
        "label_downloading": ORDER_STATUS_PICKING,
        "label_saved": ORDER_STATUS_PICKING,
        "shipment_created": ORDER_STATUS_PICKING,
        "picking": ORDER_STATUS_PICKING,
    }
    stmt = select(Order.biz_status, Order.local_status, Order.platform_status, func.count(Order.id)).group_by(
        Order.biz_status,
        Order.local_status,
        Order.platform_status,
    )
    waiting_purchase_count = 0
    for status_row in db.execute(stmt).all():
        values = tuple(status_row)
        label_printed = False
        has_purchase_order = False
        platform_status = None
        if len(values) == 6:
            biz, local, platform_status, label_printed, has_purchase_order, cnt = values
        elif len(values) == 5:
            biz, local, label_printed, has_purchase_order, cnt = values
        elif len(values) == 4:
            biz, local, platform_status, cnt = values
        else:
            biz, local, cnt = values

        is_voided_platform_status = _is_voided_platform_status(platform_status)
        if biz == ORDER_STATUS_WAITING_PURCHASE and not is_voided_platform_status:
            if len(values) in {5, 6} and label_printed and not has_purchase_order:
                waiting_purchase_count += int(cnt or 0)
            continue
        status_label = ORDER_STATUS_VOIDED if is_voided_platform_status else biz or local_mapping.get(local or "", ORDER_STATUS_PENDING)
        status = _status_key_for_label(status_label)
        if status in counts:
            counts[status] += int(cnt or 0)
    if hasattr(db, "scalar"):
        waiting_purchase_count = int(
            db.scalar(select(func.count(Order.id)).where(_waiting_purchase_condition(), ~_voided_platform_status_condition())) or 0
        )
    counts[ORDER_STATUS_WAITING_PURCHASE_KEY] = waiting_purchase_count
    counts["platform_status_counts"] = _platform_status_counts(db)
    return counts


def _platform_status_count_label(value: str | None) -> str:
    status_value = str(value or "").strip()
    if not status_value or status_value.lower() == "none":
        return "未记录"
    return status_value


def _platform_status_counts(db: Session) -> dict[str, int]:
    counts: dict[str, int] = {}
    stmt = select(Order.platform_status, func.count(Order.id)).group_by(Order.platform_status)
    for platform_status, cnt in db.execute(stmt).all():
        label = _platform_status_count_label(platform_status)
        counts[label] = counts.get(label, 0) + int(cnt or 0)
    return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))


@app.post("/api/v1/orders/batch/risk-handling", response_model=OrderBatchResponse)
def batch_update_order_risk_handling(
    payload: OrderRiskHandlingRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    rows = _load_orders_by_ids(db, payload.order_ids)
    now = datetime.utcnow()
    deadline_24h = now + timedelta(hours=24)
    invalid_rows = []
    for row in rows:
        deadline = row.dispatch_deadline_at or _effective_shipping_deadline(row, _extract_order_fields(row.raw_payload or {}))
        if deadline is not None and deadline.tzinfo is not None:
            deadline = deadline.astimezone(timezone.utc).replace(tzinfo=None)
        if row.biz_status not in ORDER_RISK_BIZ_STATUSES or deadline is None or deadline >= deadline_24h:
            invalid_rows.append(_order_display_number(row))
    if invalid_rows:
        preview = "、".join(invalid_rows[:5])
        suffix = f" 等 {len(invalid_rows)} 条" if len(invalid_rows) > 5 else ""
        raise HTTPException(status_code=400, detail=f"风险订单状态已变化，请刷新后重试：{preview}{suffix}")

    order_ids = [row.id for row in rows]
    existing_rows = db.scalars(
        select(OrderRiskHandling).where(OrderRiskHandling.order_id.in_(order_ids))
    ).all()
    handling_map = {row.order_id: row for row in existing_rows}
    note = (payload.note or "").strip()
    operator = operator_name(user)
    for order in rows:
        handling = handling_map.get(order.id)
        if payload.handled:
            if handling is None:
                handling = OrderRiskHandling(order_id=order.id)
                db.add(handling)
                handling_map[order.id] = handling
            handling.handled_at = now
            handling.handled_by = operator
            handling.note = note
            add_order_operation_log(
                db,
                order_id=order.id,
                operation_type="risk_handled",
                operation_attribute="发货风险跟进",
                description=f"订单 {_order_display_number(order)} 已标记为风险已处理",
                operator=operator,
                source=ORDER_LOG_MANUAL_SOURCE,
                operated_at=now,
                extra={"risk_note": note},
            )
        else:
            if handling is not None:
                db.delete(handling)
            add_order_operation_log(
                db,
                order_id=order.id,
                operation_type="risk_reopened",
                operation_attribute="发货风险跟进",
                description=f"订单 {_order_display_number(order)} 已取消风险处理标记",
                operator=operator,
                source=ORDER_LOG_MANUAL_SOURCE,
                operated_at=now,
            )
    db.commit()
    return OrderBatchResponse(
        updated=len(rows),
        message=f"已{'标记' if payload.handled else '取消'}风险处理 {len(rows)} 条",
    )


@app.get("/api/v1/orders/{order_id:int}", response_model=OrderDetailDto)
def get_order_detail(
    order_id: int,
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    row = db.get(Order, order_id)
    if not row:
        raise HTTPException(status_code=404, detail="订单不存在")
    return _order_detail_dto(row, db)


@app.get("/api/v1/orders/{order_id:int}/operation-logs", response_model=OrderOperationLogListResponse)
def list_order_operation_logs(
    order_id: int,
    before_id: int | None = Query(default=None, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    operation_type: str | None = Query(default=None),
    source: str | None = Query(default=None),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    stmt = select(OrderOperationLog).where(OrderOperationLog.order_id == order_id)
    operation_type = str(operation_type or "").strip()
    source = str(source or "").strip()
    if operation_type:
        stmt = stmt.where(OrderOperationLog.operation_type == operation_type)
    if source:
        if source not in {ORDER_LOG_SYSTEM_SOURCE, ORDER_LOG_MANUAL_SOURCE, ORDER_LOG_HISTORY_SOURCE}:
            raise HTTPException(status_code=400, detail="日志来源无效")
        stmt = stmt.where(OrderOperationLog.source == source)

    if before_id:
        cursor = db.get(OrderOperationLog, before_id)
        if not cursor or cursor.order_id != order_id:
            raise HTTPException(status_code=400, detail="日志游标无效")
        stmt = stmt.where(
            or_(
                OrderOperationLog.operated_at < cursor.operated_at,
                and_(
                    OrderOperationLog.operated_at == cursor.operated_at,
                    OrderOperationLog.id < cursor.id,
                ),
            )
        )

    rows = db.scalars(
        stmt.order_by(desc(OrderOperationLog.operated_at), desc(OrderOperationLog.id)).limit(page_size + 1)
    ).all()
    has_more = len(rows) > page_size
    rows = rows[:page_size]
    items = []
    for row in rows:
        dto = _order_operation_log_dto(row)
        dto.description = _legacy_order_log_description(db, row, order)
        items.append(dto)
    return OrderOperationLogListResponse(
        items=items,
        has_more=has_more,
        next_before_id=rows[-1].id if has_more and rows else None,
    )


@app.get("/api/v1/order-summary", response_model=OrderSummaryResponse)
def list_order_summary(
    status: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    shop_ids: str | None = Query(default=None),
    number: str | None = Query(default=None),
    transaction_id: str | None = Query(default=None),
    tracking_number: str | None = Query(default=None),
    product_keyword: str | None = Query(default=None),
    warning: str | None = Query(default=None),
    payment_time_range: str | None = Query(default=None),
    payment_start: str | None = Query(default=None),
    payment_end: str | None = Query(default=None),
    picking_start: str | None = Query(default=None),
    picking_end: str | None = Query(default=None),
    old_customer_only: bool = Query(default=False),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    lazy: bool = Query(default=False),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    if page_size not in {50, 100, 500}:
        page_size = 50
    return _query_order_summary(
        db,
        status,
        platform,
        transaction_id,
        tracking_number,
        number,
        payment_time_range,
        payment_start,
        payment_end,
        picking_start,
        picking_end,
        old_customer_only,
        page,
        page_size,
        warning,
        product_keyword,
        lazy=lazy,
        shop_keys=_shop_keys_from_ids(db, shop_ids),
    )


@app.get("/api/v1/order-summary/export")
def export_order_summary(
    item_ids: str | None = Query(default=None),
    columns: str | None = Query(default=None),
    status: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    shop_ids: str | None = Query(default=None),
    number: str | None = Query(default=None),
    transaction_id: str | None = Query(default=None),
    tracking_number: str | None = Query(default=None),
    product_keyword: str | None = Query(default=None),
    warning: str | None = Query(default=None),
    payment_time_range: str | None = Query(default=None),
    payment_start: str | None = Query(default=None),
    payment_end: str | None = Query(default=None),
    picking_start: str | None = Query(default=None),
    picking_end: str | None = Query(default=None),
    old_customer_only: bool = Query(default=False),
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc

    response = _query_order_summary(
        db,
        status,
        platform,
        transaction_id,
        tracking_number,
        number,
        payment_time_range,
        payment_start,
        payment_end,
        picking_start,
        picking_end,
        old_customer_only,
        1,
        500,
        warning,
        product_keyword,
        item_ids=_to_int_list(item_ids),
        paginate=False,
        shop_keys=_shop_keys_from_ids(db, shop_ids),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "订单明细表"
    export_columns = _table_export_columns(
        ORDER_SUMMARY_TABLE_KEY,
        ORDER_SUMMARY_PRIMARY_COLUMN_KEY,
        ORDER_SUMMARY_EXPORT_COLUMNS,
        user,
        db,
        columns,
    )
    worksheet.append([column["title"] for column in export_columns])
    for row in response.items:
        worksheet.append([_order_summary_export_value(column["key"], row) for column in export_columns])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"order-summary-{_local_now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/v1/outbound-scans", response_model=OutboundScanResponse)
def create_outbound_scan(
    payload: OutboundScanRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    raw_input = payload.raw_input if payload.raw_input is not None else payload.tracking_number
    tracking_number = _clean_tracking_number(payload.tracking_number)
    if not tracking_number:
        record = _create_outbound_scan_record(
            db,
            tracking_number="",
            raw_input=raw_input or "",
            result="invalid",
            message="扫码内容为空",
            scanned_by=user.username,
        )
        return OutboundScanResponse(result=record.result, message=record.message, record=_outbound_scan_dto(record))

    order = _find_order_by_tracking_number(db, tracking_number)
    if not order:
        record = _create_outbound_scan_record(
            db,
            tracking_number=tracking_number,
            raw_input=raw_input or "",
            result="not_found",
            message="未找到对应订单",
            scanned_by=user.username,
        )
        return OutboundScanResponse(result=record.result, message=record.message, record=_outbound_scan_dto(record))

    if _has_successful_outbound_scan(db, tracking_number):
        record = _create_outbound_scan_record(
            db,
            tracking_number=tracking_number,
            raw_input=raw_input or "",
            result="duplicate",
            message="该货运单号已扫码出库",
            scanned_by=user.username,
            order=order,
        )
        return OutboundScanResponse(result=record.result, message=record.message, record=_outbound_scan_dto(record))

    if order.biz_status != ORDER_STATUS_PICKING:
        record = _create_outbound_scan_record(
            db,
            tracking_number=tracking_number,
            raw_input=raw_input or "",
            result="invalid",
            message=f"订单状态为{order.biz_status or '-'}，只有配货中订单允许扫码出库",
            scanned_by=user.username,
            order=order,
        )
        return OutboundScanResponse(result=record.result, message=record.message, record=_outbound_scan_dto(record))

    record = _create_outbound_scan_record(
        db,
        tracking_number=tracking_number,
        raw_input=raw_input or "",
        result="success",
        message="扫码记录成功",
        scanned_by=user.username,
        order=order,
    )
    now = record.scanned_at or datetime.utcnow()
    order.biz_status = ORDER_STATUS_SHIPPED
    order.local_status = "shipped"
    order.shipped_at = now
    order.marked_shipped_at = now
    if not order.shipment_tracking_number:
        shipment = _latest_shipment(db, order.id)
        if shipment and shipment.tracking_number:
            order.shipment_tracking_number = shipment.tracking_number
    order.updated_at = now
    record.order_status = _derive_order_status(order)
    add_order_operation_log(
        db,
        order_id=order.id,
        operation_type="outbound_scan",
        operation_attribute="扫码出库",
        description=f"货运单号 {tracking_number} 扫码出库成功，订单已转为已发货",
        operator=operator_name(user),
        source=ORDER_LOG_MANUAL_SOURCE,
        operated_at=now,
        event_key=f"outbound_scan:{record.id}",
        extra={"tracking_number": tracking_number, "scan_record_id": record.id, "status_after": ORDER_STATUS_SHIPPED},
    )
    db.commit()
    return OutboundScanResponse(result=record.result, message=record.message, record=_outbound_scan_dto(record))


@app.get("/api/v1/outbound-scans", response_model=OutboundScanListResponse)
def list_outbound_scans(
    number: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    shop_name: str | None = Query(default=None),
    result: str | None = Query(default=None),
    scanned_by: str | None = Query(default=None),
    scanned_start: str | None = Query(default=None),
    scanned_end: str | None = Query(default=None),
    today_only: bool = Query(default=False),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    if today_only and not scanned_start and not scanned_end:
        local_today = _local_today().isoformat()
        scanned_start = local_today
        scanned_end = local_today
    stmt = _build_outbound_scans_query(
        number,
        platform,
        shop_name,
        result,
        scanned_by,
        scanned_start,
        scanned_end,
    )
    total = db.scalar(
        _build_outbound_scans_count_query(
            number,
            platform,
            shop_name,
            result,
            scanned_by,
            scanned_start,
            scanned_end,
        )
    ) or 0
    rows = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return OutboundScanListResponse(
        items=[_outbound_scan_dto(row) for row in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.get("/api/v1/outbound-scans/stats", response_model=OutboundScanStatsResponse)
def outbound_scan_stats(
    _: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    start, end = _local_day_utc_bounds()
    stmt = (
        select(OutboundScanRecord.result, func.count(OutboundScanRecord.id))
        .where(OutboundScanRecord.scanned_at >= start, OutboundScanRecord.scanned_at < end)
        .group_by(OutboundScanRecord.result)
    )
    counts = {"success": 0, "duplicate": 0, "not_found": 0, "invalid": 0, "error": 0}
    for result, count in db.execute(stmt).all():
        if result in counts:
            counts[result] = int(count or 0)
        else:
            counts["error"] += int(count or 0)
    total = sum(counts.values())
    last_scanned_at = db.scalar(
        select(OutboundScanRecord.scanned_at)
        .where(OutboundScanRecord.scanned_at >= start, OutboundScanRecord.scanned_at < end)
        .order_by(desc(OutboundScanRecord.scanned_at), desc(OutboundScanRecord.id))
        .limit(1)
    )
    return OutboundScanStatsResponse(total=total, last_scanned_at=_iso(last_scanned_at), **counts)


@app.get("/api/v1/outbound-scans/export")
def export_outbound_scans(
    number: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    shop_name: str | None = Query(default=None),
    result: str | None = Query(default=None),
    scanned_by: str | None = Query(default=None),
    scanned_start: str | None = Query(default=None),
    scanned_end: str | None = Query(default=None),
    order_outbound: bool = Query(default=False),
    columns: str | None = Query(default=None),
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    export_result = "success" if order_outbound else result
    rows = db.scalars(
        _build_outbound_scans_query(
            number,
            platform,
            shop_name,
            export_result,
            scanned_by,
            scanned_start,
            scanned_end,
        )
    ).all()
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "出库记录" if order_outbound else "扫码记录"
    export_columns = _table_export_columns(
        OUTBOUND_SCANS_TABLE_KEY,
        OUTBOUND_SCANS_PRIMARY_COLUMN_KEY,
        OUTBOUND_SCANS_EXPORT_COLUMNS,
        user,
        db,
        columns,
    )
    if order_outbound:
        worksheet.append(["扫码时间", "货运单号", "平台", "店铺", "订单编号", "订单状态", "平台状态", "操作员"])
    else:
        worksheet.append([column["title"] for column in export_columns])
    for row in rows:
        dto = _outbound_scan_dto(row)
        if order_outbound:
            worksheet.append([
                _excel_datetime(row.scanned_at),
                dto.tracking_number,
                _platform_display_name(dto.platform) or "-",
                dto.shop_name,
                dto.platform_order_no,
                dto.order_status,
                dto.platform_status,
                dto.scanned_by,
            ])
        else:
            worksheet.append([_outbound_scan_export_value(column["key"], dto) for column in export_columns])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename_prefix = "order-outbound" if order_outbound else "outbound-scans"
    filename = f"{filename_prefix}-{_local_now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _ensure_joom_registration_only_orders(rows: list[Order]) -> None:
    registration_only_rows = [
        row
        for row in rows
        if order_is_joom_overseas_warehouse(row) or order_is_joom_fbj_warehouse(row)
    ]
    if registration_only_rows:
        raise HTTPException(status_code=409, detail="Joom FBJ/海外仓订单仅登记，不进入面单、打印和采购流程")


@app.post("/api/v1/orders/batch/to-picking", response_model=OrderBatchResponse)
def batch_to_picking(payload: OrderBatchRequest, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    rows = _load_orders_by_ids(db, payload.order_ids)
    existing_response = _to_picking_existing_response(db, payload.order_ids)
    if existing_response:
        return existing_response
    _ensure_order_statuses(rows, {ORDER_STATUS_WAITING_PURCHASE})
    _ensure_joom_registration_only_orders(rows)
    operator = operator_name(user)
    rules = load_enabled_logistics_rules(db)
    now = datetime.utcnow()
    purchasable_rows, unmatched_rule_rows = split_logistics_rule_eligible_orders(rows, rules, matched_at=now)
    if unmatched_rule_rows:
        _mark_logistics_rule_unmatched_as_shipped(
            db,
            unmatched_rule_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
    purchasable_rows, label_exempt_rows = _split_platform_label_exempt_rows(purchasable_rows)
    if label_exempt_rows:
        _mark_platform_label_exempt_rows_as_shipped(
            db,
            label_exempt_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
    if not purchasable_rows:
        db.commit()
        message_parts = []
        if label_exempt_rows:
            message_parts.append(f"无需平台物流/面单订单已跳过采购并转为已发货 {len(label_exempt_rows)} 条")
        if unmatched_rule_rows:
            message_parts.append(f"物流规则未匹配订单已跳过采购并转为已发货 {len(unmatched_rule_rows)} 条")
        return OrderBatchResponse(
            updated=len(rows),
            message="，".join(message_parts) or "没有可生成采购单的订单",
        )
    try:
        purchase = _generate_purchase_order_for_orders(db, purchasable_rows, operator, "", allow_existing=True)
        _move_orders_to_picking_after_purchase(
            db,
            purchasable_rows,
            purchase,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            description=f"已生成采购单 {purchase.purchase_no}，订单已转入配货中",
        )
        db.commit()
        enqueue_purchase_order_wecom_notification(purchase.id, source="batch_to_picking")
    except IntegrityError as exc:
        db.rollback()
        existing_response = _to_picking_existing_response(db, payload.order_ids)
        if existing_response:
            return existing_response
        raise HTTPException(status_code=409, detail="转入配货中时采购单明细已变化，请刷新后重试") from exc
    message = f"已生成采购单 {purchase.purchase_no}，并转入配货中 {len(purchasable_rows)} 条"
    if label_exempt_rows:
        message += f"，无需平台物流/面单订单已跳过采购并转为已发货 {len(label_exempt_rows)} 条"
    if unmatched_rule_rows:
        message += f"，物流规则未匹配订单已转为已发货 {len(unmatched_rule_rows)} 条"
    return OrderBatchResponse(
        updated=len(rows),
        message=message,
        purchase_order_id=purchase.id,
        purchase_no=purchase.purchase_no,
    )


@app.post("/api/v1/orders/batch/to-printing", response_model=OrderBatchResponse)
def batch_to_printing(payload: OrderBatchRequest, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    rows = _load_orders_by_ids(db, payload.order_ids)
    _ensure_order_statuses(rows, {ORDER_STATUS_PENDING})
    _ensure_joom_registration_only_orders(rows)
    joom_offline_rows = [row for row in rows if order_is_joom_offline_shipping(row)]
    if joom_offline_rows:
        preview = "、".join(_order_display_number(row) for row in joom_offline_rows[:10])
        suffix = f" 等 {len(joom_offline_rows)} 条" if len(joom_offline_rows) > 10 else ""
        raise HTTPException(
            status_code=409,
            detail=f"Joom 线下物流订单不使用平台在线面单，请等待平台返回已发货状态和货运单号：{preview}{suffix}",
        )
    operator = operator_name(user)
    now = datetime.utcnow()
    rules = load_enabled_logistics_rules(db)
    logistics_rows, unmatched_rule_rows = split_logistics_rule_eligible_orders(rows, rules, matched_at=now)
    regular_rows, label_exempt_rows = _split_platform_label_exempt_rows(logistics_rows)
    if not payload.allow_missing_tracking:
        missing_tracking = [row for row in regular_rows if not _order_tracking_number_value(db, row)]
        if missing_tracking:
            preview = "、".join(_order_display_number(row) for row in missing_tracking[:10])
            suffix = f" 等 {len(missing_tracking)} 条" if len(missing_tracking) > 10 else ""
            raise HTTPException(status_code=400, detail=f"以下订单没有货运单号，请确认后再转入待打印: {preview}{suffix}")
    for row in regular_rows:
        row.biz_status = ORDER_STATUS_WAITING_PRINT
        row.updated_at = now
    if label_exempt_rows:
        _mark_platform_label_exempt_rows_as_shipped(
            db,
            label_exempt_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
            extra={"action": "to_printing"},
        )
    if regular_rows:
        add_order_operation_logs(
            db,
            regular_rows,
            operation_type="to_printing",
            operation_attribute="修改订单基础信息",
            description=lambda order: (
                f"订单 {_order_display_number(order)} 已转入待打印，状态：{ORDER_STATUS_PENDING} -> {ORDER_STATUS_WAITING_PRINT}"
            ),
            operator=operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
            extra={
                "order_ids": [row.id for row in regular_rows],
            },
        )
    if unmatched_rule_rows:
        _mark_logistics_rule_unmatched_as_shipped(
            db,
            unmatched_rule_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
    db.commit()
    message = f"已转入待打印 {len(regular_rows)} 条"
    if label_exempt_rows:
        message += f"，无需平台物流/面单订单已跳过采购并转为已发货 {len(label_exempt_rows)} 条"
    if unmatched_rule_rows:
        message += f"，物流规则未匹配订单已转为已发货 {len(unmatched_rule_rows)} 条"
    return OrderBatchResponse(updated=len(rows), message=message)


@app.post("/api/v1/orders/batch/confirm-printed", response_model=OrderBatchResponse)
def batch_confirm_printed(payload: OrderBatchRequest, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    rows = _load_orders_by_ids(db, payload.order_ids)
    _ensure_order_statuses(rows, {ORDER_STATUS_WAITING_PRINT})
    _ensure_joom_registration_only_orders(rows)
    operator = operator_name(user)
    now = datetime.utcnow()
    rules = load_enabled_logistics_rules(db)
    printable_rows, unmatched_rule_rows = split_logistics_rule_eligible_orders(rows, rules, matched_at=now)
    for row in printable_rows:
        row.label_printed_at = row.label_printed_at or now
        row.biz_status = ORDER_STATUS_WAITING_PURCHASE
        row.updated_at = now
    if printable_rows:
        add_order_operation_logs(
            db,
            printable_rows,
            operation_type="confirm_printed",
            operation_attribute="确认已打印",
            description=lambda order: (
                f"订单 {_order_display_number(order)} 已确认面单打印完成，状态：{ORDER_STATUS_WAITING_PRINT} -> {ORDER_STATUS_WAITING_PURCHASE}"
            ),
            operator=operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
    if unmatched_rule_rows:
        _mark_logistics_rule_unmatched_as_shipped(
            db,
            unmatched_rule_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
    db.commit()
    message = f"已确认打印，并转入待采购 {len(printable_rows)} 条"
    if unmatched_rule_rows:
        message += f"，物流规则未匹配订单已转为已发货 {len(unmatched_rule_rows)} 条"
    return OrderBatchResponse(
        updated=len(rows),
        message=message,
    )


@app.post("/api/v1/orders/batch/sync-logistics", response_model=OrderBatchResponse)
async def batch_sync_logistics(payload: OrderBatchRequest, _: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    raise HTTPException(status_code=409, detail="物流同步仅允许由 auto_order_pipeline 定时任务执行")


@app.post("/api/v1/orders/batch/logistics-channel", response_model=OrderBatchResponse)
def batch_set_logistics_channel(
    payload: OrderLogisticsChannelBatchRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    rows = _load_orders_by_ids(db, payload.order_ids)
    channel = (payload.logistics_channel or "").strip()
    if not channel:
        raise HTTPException(status_code=400, detail="物流渠道不能为空")
    authorization = _logistics_authorization_for_rule_channel(db, channel)
    carrier_code = _normalize_logistics_carrier_code(authorization.carrier_code) if authorization else ""
    now = datetime.utcnow()
    for row in rows:
        apply_manual_logistics_channel(row, channel, carrier_code=carrier_code, matched_at=now)
    add_order_operation_logs(
        db,
        rows,
        operation_type="set_logistics_channel",
        operation_attribute="指定物流渠道",
        description=lambda order: f"订单 {_order_display_number(order)} 已人工指定物流渠道：{channel}",
        operator=operator_name(user),
        source=ORDER_LOG_MANUAL_SOURCE,
        operated_at=now,
        extra={"order_ids": [row.id for row in rows], "logistics_channel": channel},
    )
    db.commit()
    return OrderBatchResponse(updated=len(rows), message=f"已指定物流渠道 {len(rows)} 条")


def _order_wanbang_test_error_item(row: Order, error: str) -> OrderWanbangTestItemDto:
    return OrderWanbangTestItemDto(
        order_id=row.id,
        order_no=_order_display_number(row),
        success=False,
        error=error[:500],
    )


def _save_wanbang_test_result(
    db: Session,
    row: Order,
    label_result,
    shipment_result: ShipmentResult,
) -> str:
    content = label_result.content
    posting_number = shipment_result.platform_shipment_id or row.posting_number or row.platform_order_id
    file_path, sha256 = save_label_pdf(row.tenant_id, row.platform, row.account_id, posting_number, content)

    shipment = Shipment(
        order_id=row.id,
        platform_shipment_id=shipment_result.platform_shipment_id or posting_number,
        tracking_number=shipment_result.tracking_number or row.shipment_tracking_number or "",
        carrier=shipment_result.carrier or WANBANG_CARRIER_NAME,
        status=shipment_result.status or "label_ready",
    )
    db.add(shipment)
    db.flush()

    apply_label_result_tracking(row, shipment, label_result)
    if shipment_result.tracking_number:
        row.shipment_tracking_number = shipment_result.tracking_number
        shipment.tracking_number = shipment_result.tracking_number
    row.error_message = ""
    row.updated_at = datetime.utcnow()
    db.add(
        LabelFile(
            shipment_id=shipment.id,
            file_path=file_path,
            content_type=label_result.content_type or "application/pdf",
            sha256=sha256,
        )
    )
    return file_path


@app.post("/api/v1/orders/batch/wanbang-test", response_model=OrderWanbangTestResponse)
async def batch_wanbang_test(
    payload: OrderBatchRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    rows = _load_orders_by_ids(db, payload.order_ids)
    if not rows:
        raise HTTPException(status_code=400, detail="未选中订单")

    operator = operator_name(user)
    now = datetime.utcnow()
    items: list[OrderWanbangTestItemDto] = []
    for row in rows:
        try:
            test_result, label_result, shipment_result = await run_wanbang_test_flow_for_order(db, row)
            if not _is_real_label_pdf(label_result.content):
                raise RuntimeError("万邦返回非有效真实面单 PDF")
            label_path = _save_wanbang_test_result(db, row, label_result, shipment_result)
            items.append(
                OrderWanbangTestItemDto(
                    order_id=row.id,
                    order_no=_order_display_number(row),
                    success=True,
                    account_name=test_result.account_name,
                    process_code=test_result.process_code,
                    tracking_number=test_result.tracking_number,
                    parcel_status=test_result.parcel_status,
                    reference_id=test_result.reference_id,
                    label_ready=test_result.label_ready,
                    label_attempts=test_result.label_attempts,
                    label_bytes=test_result.label_bytes,
                    label_sha256=test_result.label_sha256,
                    label_path=label_path,
                )
            )
            add_order_operation_log(
                db,
                order_id=row.id,
                operation_type="wanbang_test",
                operation_attribute="万邦测试",
                description=(
                    f"订单 {_order_display_number(row)} 已完成万邦测试：创建包裹 {test_result.process_code}"
                    f"，状态 {test_result.parcel_status or '-'}，面单获取成功"
                ),
                operator=operator,
                source=ORDER_LOG_MANUAL_SOURCE,
                operated_at=now,
                extra={
                    "process_code": test_result.process_code,
                    "reference_id": test_result.reference_id,
                    "tracking_number": test_result.tracking_number,
                    "parcel_status": test_result.parcel_status,
                    "label_attempts": test_result.label_attempts,
                    "label_sha256": test_result.label_sha256,
                    "label_path": label_path,
                },
            )
        except Exception as exc:
            error = safe_exception_message(exc)
            row.error_message = f"万邦测试失败：{error[:500]}"
            row.updated_at = datetime.utcnow()
            items.append(_order_wanbang_test_error_item(row, error))
            add_order_operation_log(
                db,
                order_id=row.id,
                operation_type="wanbang_test",
                operation_attribute="万邦测试",
                description=f"订单 {_order_display_number(row)} 万邦测试失败：{error[:300]}",
                operator=operator,
                source=ORDER_LOG_MANUAL_SOURCE,
                operated_at=now,
                extra={"error": error[:500]},
            )

    db.commit()
    succeeded = sum(1 for item in items if item.success)
    failed = len(items) - succeeded
    message = f"万邦测试完成：成功 {succeeded} 条，失败 {failed} 条"
    return OrderWanbangTestResponse(total=len(items), succeeded=succeeded, failed=failed, message=message, items=items)


@app.post("/api/v1/orders/batch/print-label")
async def batch_print_label(
    payload: OrderBatchRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    """下载面单 PDF。

    优先读取本地缓存；本地缺失时从平台拉取并保存，再合并为单个文件返回。
    """
    rows = _load_orders_by_ids(db, payload.order_ids)
    _ensure_order_statuses(
        rows,
        {
            ORDER_STATUS_WAITING_PRINT,
            ORDER_STATUS_WAITING_PURCHASE,
            ORDER_STATUS_PICKING,
            ORDER_STATUS_SHIPPED,
            ORDER_STATUS_DELIVERED,
        },
    )
    if not rows:
        raise HTTPException(status_code=400, detail="未选中订单")
    _ensure_joom_registration_only_orders(rows)

    operator = operator_name(user)
    now = datetime.utcnow()
    rules = load_enabled_logistics_rules(db)
    printable_candidate_rows, unmatched_rule_rows = split_logistics_rule_eligible_orders(rows, rules, matched_at=now)
    joom_offline_rows = [row for row in printable_candidate_rows if order_is_joom_offline_shipping(row)]
    joom_offline_ids = {row.id for row in joom_offline_rows}
    printable_candidate_rows = [row for row in printable_candidate_rows if row.id not in joom_offline_ids]
    regular_rows, label_exempt_rows = _split_platform_label_exempt_rows(printable_candidate_rows)
    if regular_rows:
        pdf_map, cached, fetched, failed = await _ensure_labels_cached(db, regular_rows, load_bytes=True)
        print_option_map = _platform_print_option_map(db)
    else:
        pdf_map, cached, fetched, failed = {}, 0, 0, 0
        print_option_map = {}

    # 按传入顺序合并，保证预览页签里的 PDF 页面顺序与勾选顺序一致。
    ordered_parts: list[bytes] = []
    missing_labels: list[str] = []
    missing_details: list[str] = []
    allegro_unavailable_rows: list[Order] = []
    for r in regular_rows:
        pdf = pdf_map.get(r.id)
        if pdf:
            print_options = print_option_map.get(r.platform or "") or {}
            ordered_parts.append(
                orient_pdf_bytes(
                    pdf,
                    print_options.get("page_orientation"),
                    target_size_mm=print_options.get("target_size_mm") or label_size_mm_for_platform(r.platform),
                )
            )
        else:
            label = r.platform_order_no or r.posting_number or r.platform_order_id or str(r.id)
            if str(r.platform or "").strip().lower() == "allegro" and _allegro_label_fetch_unavailable_message(r.error_message):
                allegro_unavailable_rows.append(r)
                continue
            missing_labels.append(label)
            if r.error_message:
                missing_details.append(f"{label}: {r.error_message}")

    skipped_rows = label_exempt_rows + allegro_unavailable_rows
    if not ordered_parts and not skipped_rows and not joom_offline_rows:
        if unmatched_rule_rows and not regular_rows:
            _mark_logistics_rule_unmatched_as_shipped(
                db,
                unmatched_rule_rows,
                operator,
                source=ORDER_LOG_MANUAL_SOURCE,
                operated_at=now,
            )
            db.commit()
            return {
                "filename": f"labels_{_local_now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "content_type": "application/pdf",
                "pdf_base64": "",
                "cached": cached,
                "fetched": fetched,
                "failed": failed,
                "skipped": len(unmatched_rule_rows),
                "printed": 0,
                "total": len(rows),
            }
        if missing_details:
            preview = "；".join(missing_details[:3])
            suffix = f" 等 {len(missing_details)} 条" if len(missing_details) > 3 else ""
            raise HTTPException(status_code=502, detail=f"所选订单面单拉取失败：{preview}{suffix}")
        preview = "、".join(missing_labels[:5])
        suffix = f" 等 {len(missing_labels)} 条" if len(missing_labels) > 5 else ""
        raise HTTPException(status_code=502, detail=f"所选订单均没有真实面单数据：{preview}{suffix}")
    printed_ids = [r.id for r in regular_rows if pdf_map.get(r.id)]
    printed_id_set = set(printed_ids)
    previously_printed_ids = {row.id for row in rows if row.label_printed_at is not None}
    status_before_print = {row.id: row.biz_status for row in rows}
    if printed_ids:
        db.execute(
            text(
                """
                UPDATE orders
                SET label_printed_at = :now, updated_at = :now
                WHERE id = ANY(:ids) AND label_printed_at IS NULL
                """
            ),
            {"now": now, "ids": printed_ids},
        )
        for row in rows:
            if row.id in printed_id_set and row.biz_status == ORDER_STATUS_WAITING_PRINT:
                row.biz_status = ORDER_STATUS_WAITING_PURCHASE
                row.updated_at = now
    if skipped_rows:
        _mark_platform_label_exempt_rows_as_shipped(
            db,
            skipped_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
            extra={
                "action": "print_label",
                "allegro_label_unavailable_order_ids": [row.id for row in allegro_unavailable_rows],
            },
        )
        for row in allegro_unavailable_rows:
            row.error_message = ""
    printed_rows = [r for r in rows if r.id in printed_id_set]
    if printed_rows:
        add_order_operation_logs(
            db,
            printed_rows,
            operation_type="print_label",
            operation_attribute="打印面单",
            description=lambda order: (
                f"订单 {_order_display_number(order)} 面单已{'重新打印' if order.id in previously_printed_ids else '打印'}"
                + _order_tracking_log_suffix(db, order)
                + (
                    f"，状态：{status_before_print.get(order.id)} -> {order.biz_status}"
                    if status_before_print.get(order.id) and status_before_print.get(order.id) != order.biz_status
                    else ""
                )
            ),
            operator=operator_name(user),
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
            extra={
                "cached": cached,
                "fetched": fetched,
                "failed": failed,
                "skipped": len(missing_labels),
            },
        )
    if unmatched_rule_rows:
        _mark_logistics_rule_unmatched_as_shipped(
            db,
            unmatched_rule_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
    db.commit()

    merged = merge_pdf_parts(ordered_parts) if ordered_parts else b""

    filename = f"labels_{_local_now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return {
        "filename": filename,
        "content_type": "application/pdf",
        "pdf_base64": base64.b64encode(merged).decode("ascii"),
        "cached": cached,
        "fetched": fetched,
        "failed": failed,
        "skipped": len(missing_labels) + len(skipped_rows) + len(joom_offline_rows) + len(unmatched_rule_rows),
        "printed": len(ordered_parts),
        "total": len(rows),
    }


@app.post("/api/v1/orders/batch/print-chinese-label")
def batch_print_chinese_label(
    payload: OrderBatchRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    """生成本地中文拣货标签 PDF。"""
    rows = _load_orders_by_ids(db, payload.order_ids)
    _ensure_order_statuses(
        rows,
        {
            ORDER_STATUS_WAITING_PRINT,
            ORDER_STATUS_WAITING_PURCHASE,
            ORDER_STATUS_PICKING,
            ORDER_STATUS_SHIPPED,
            ORDER_STATUS_DELIVERED,
        },
    )
    if not rows:
        raise HTTPException(status_code=400, detail="未选中订单")
    _ensure_joom_registration_only_orders(rows)

    now = datetime.utcnow()
    operator = operator_name(user)
    rules = load_enabled_logistics_rules(db)
    rows, unmatched_rule_rows = split_logistics_rule_eligible_orders(rows, rules, matched_at=now)
    if unmatched_rule_rows:
        _mark_logistics_rule_unmatched_as_shipped(
            db,
            unmatched_rule_rows,
            operator,
            source=ORDER_LOG_MANUAL_SOURCE,
            operated_at=now,
        )
        db.commit()
    if not rows:
        return {
            "filename": f"chinese_labels_{_local_now().strftime('%Y%m%d_%H%M%S')}.pdf",
            "content_type": "application/pdf",
            "pdf_base64": "",
            "cached": 0,
            "fetched": 0,
            "failed": 0,
            "skipped": len(unmatched_rule_rows),
            "printed": 0,
            "total": len(unmatched_rule_rows),
        }

    product_name_map = _order_chinese_product_name_map(db, rows)
    missing_tracking: list[str] = []
    label_rows: list[ChineseLabelRow] = []
    for row in rows:
        display_number = _order_display_number(row)
        tracking_number = _order_tracking_number_value(db, row)
        product_name = product_name_map.get(row.id, "")
        if not tracking_number:
            missing_tracking.append(display_number)
        if tracking_number:
            extracted = _extract_order_fields(row.raw_payload or {})
            label_rows.append(
                ChineseLabelRow(
                    tracking_number=tracking_number,
                    deadline=resolve_chinese_label_deadline(
                        platform=row.platform,
                        payment_at=row.payment_at,
                        platform_created_at=row.platform_created_at,
                        imported_at=row.created_at,
                        fallback=_effective_shipping_deadline(row, extracted),
                    ),
                    product_name=product_name,
                )
            )

    if missing_tracking:
        details: list[str] = []
        preview = "、".join(missing_tracking[:5])
        suffix = f" 等 {len(missing_tracking)} 条" if len(missing_tracking) > 5 else ""
        details.append(f"缺少货运单号：{preview}{suffix}")
        raise HTTPException(status_code=400, detail="；".join(details))

    pdf_bytes = generate_chinese_label_pdf(label_rows)
    filename = f"chinese_labels_{_local_now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return {
        "filename": filename,
        "content_type": "application/pdf",
        "pdf_base64": base64.b64encode(pdf_bytes).decode("ascii"),
        "cached": 0,
        "fetched": len(label_rows),
        "failed": 0,
        "skipped": 0,
        "printed": len(label_rows),
        "total": len(rows),
    }


@app.post("/api/v1/orders/batch/mark-shipped", response_model=OrderBatchResponse)
def batch_mark_shipped(payload: OrderBatchRequest, user: LocalUser = Depends(current_user), db: Session = Depends(get_db)):
    rows = _load_orders_by_ids(db, payload.order_ids)
    _ensure_order_statuses(rows, {ORDER_STATUS_PICKING})
    now = datetime.utcnow()
    for row in rows:
        row.biz_status = ORDER_STATUS_SHIPPED
        row.local_status = "shipped"
        row.shipped_at = now
        shipment = _latest_shipment(db, row.id)
        if shipment and not row.shipment_tracking_number:
            row.shipment_tracking_number = shipment.tracking_number
        if shipment and not row.handover_at:
            row.handover_at = shipment.created_at
    add_order_operation_logs(
        db,
        rows,
        operation_type="mark_shipped",
        operation_attribute="修改订单基础信息",
        description=lambda order: (
            f"订单 {_order_display_number(order)} 已标记发货，状态：{ORDER_STATUS_PICKING} -> {ORDER_STATUS_SHIPPED}"
            + _order_tracking_log_suffix(db, order)
        ),
        operator=operator_name(user),
        source=ORDER_LOG_MANUAL_SOURCE,
        operated_at=now,
        extra={"order_ids": [row.id for row in rows]},
    )
    db.execute(text("UPDATE orders SET marked_shipped_at = :now, updated_at = :now WHERE id = ANY(:ids)"), {"now": now, "ids": [row.id for row in rows]})
    db.commit()
    return OrderBatchResponse(updated=len(rows), message=f"已标记发货 {len(rows)} 条")


def _order_export_response(
    rows: list[Order],
    user: LocalUser,
    db: Session,
    columns: str | None,
) -> StreamingResponse:
    export_columns = _order_list_export_columns(user, db, columns)
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 xlsx") from exc
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "订单列表"
    worksheet.append([column["title"] for column in export_columns])
    for row in rows:
        dto = _order_dto(row, db=db)
        worksheet.append([_order_list_export_value(column["key"], dto) for column in export_columns])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"orders-export-{_local_now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v1/orders/export")
def export_orders(
    order_ids: str | None = Query(default=None),
    columns: str | None = Query(default=None),
    status: str | None = Query(default=None),
    risk: str | None = Query(default=None),
    shop: str | None = Query(default=None),
    shop_ids: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    number: str | None = Query(default=None),
    transaction_id: str | None = Query(default=None),
    order_no: str | None = Query(default=None),
    product_keyword: str | None = Query(default=None),
    payment_time_range: str | None = Query(default=None),
    payment_start: str | None = Query(default=None),
    payment_end: str | None = Query(default=None),
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    selected_ids = _to_int_list(order_ids)
    if selected_ids:
        rows = _load_orders_by_ids(db, selected_ids)
    else:
        rows = db.scalars(
            _build_orders_query(
                status,
                platform,
                transaction_id,
                order_no,
                payment_time_range,
                payment_start,
                payment_end,
                number=number,
                product_keyword=product_keyword,
                risk_filter=risk,
                shop=shop,
                shop_keys=_shop_keys_from_ids(db, shop_ids),
            )
        ).all()
    return _order_export_response(rows, user, db, columns)


@app.post("/api/v1/orders/export")
def export_orders_batch(
    payload: OrderExportRequest,
    user: LocalUser = Depends(current_user),
    db: Session = Depends(get_db),
):
    if payload.order_ids:
        rows = _load_orders_by_ids(db, payload.order_ids)
    else:
        _, numbers = _normalize_batch_order_numbers(payload.numbers)
        rows = db.scalars(
            _build_orders_query(
                status_filter=payload.status,
                platform=payload.platform,
                transaction_id=None,
                order_no=None,
                payment_time_range=None,
                payment_start=payload.payment_start,
                payment_end=payload.payment_end,
                product_keyword=payload.product_keyword,
                numbers=numbers,
                risk_filter=payload.risk,
                shop=payload.shop,
                shop_keys=_shop_keys_from_ids(db, payload.shop_ids),
            )
        ).all()
    columns = ",".join(payload.columns) if payload.columns else None
    return _order_export_response(rows, user, db, columns)


@app.post("/internal/credentials/decrypt")
def internal_decrypt_credentials(
    platform: str,
    account_id: str,
    db: Session = Depends(get_db),
    _internal_service: bool = Depends(require_internal_service_token),
):
    row = _find_shop(db, platform, account_id)
    if not row or not row.encrypted_credentials:
        raise HTTPException(status_code=404, detail="Shop credentials not found")
    return {
        "platform": _canonical_platform(row.platform),
        "account_id": account_id,
        "credentials": get_credential_manager().decrypt_credentials(row.encrypted_credentials),
        "settings": row.settings or {},
    }


FILE_BROWSER_COOKIE_NAME = "caifuclaw_filebrowser_session"
FILE_BROWSER_SESSION_SECONDS = 2 * 60 * 60
FILE_BROWSER_PROXY_BASE = "http://127.0.0.1:8088/filebrowser/"
FILE_BROWSER_UPLOAD_EXCEL_ROOT = Path.home() / "outdata" / "upload_excel"
FILE_BROWSER_UPLOAD_EXCEL_URL = "/files/upload_excel/"
FILE_BROWSER_AUTH_REFRESH_MARGIN_SECONDS = 60
FILE_BROWSER_PUBLIC_API_PREFIXES = {
    "login",
    "public",
    "raw",
    "renew",
    "resources",
    "search",
    "settings",
    "share",
    "shares",
    "tus",
    "usage",
    "users",
}
FILE_BROWSER_DOWNLOAD_ONLY_EXTENSIONS = {
    ".csv",
    ".doc",
    ".docx",
    ".ods",
    ".odt",
    ".ppt",
    ".pptx",
    ".rar",
    ".tar",
    ".tgz",
    ".xls",
    ".xlsm",
    ".xlsx",
    ".zip",
}
FILE_BROWSER_FORWARD_BLOCKED_HEADERS = {
    "connection",
    "keep-alive",
    "proxy-authenticate",
    "proxy-authorization",
    "te",
    "trailer",
    "transfer-encoding",
    "upgrade",
}
FILE_BROWSER_RESPONSE_BLOCKED_HEADERS = FILE_BROWSER_FORWARD_BLOCKED_HEADERS | {"content-encoding", "content-length"}
FILE_BROWSER_CLEAN_UI_STYLE = """
<style id="caifuclaw-filebrowser-clean-ui">
  :root,
  :root.dark {
    --blue: #1677ff !important;
    --dark-blue: #0958d9 !important;
    --red: #ef4444 !important;
    --dark-red: #dc2626 !important;
    --moon-grey: #f3f6fb !important;
    --icon-red: #ef4444 !important;
    --icon-orange: #f59e0b !important;
    --icon-yellow: #f59e0b !important;
    --icon-green: #10b981 !important;
    --icon-blue: #1677ff !important;
    --icon-violet: #8b5cf6 !important;
    --input-red: #fff1f0 !important;
    --input-green: #f0fdf4 !important;
    --item-selected: #eef5ff !important;
    --action: #374151 !important;
    --background: #f3f6fb !important;
    --surfacePrimary: #fff !important;
    --surfaceSecondary: #f8fafc !important;
    --divider: #edf0f5 !important;
    --iconPrimary: #1677ff !important;
    --iconSecondary: #fff !important;
    --iconTertiary: #d1d5db !important;
    --textPrimary: #6b7280 !important;
    --textSecondary: #111827 !important;
    --hover: #eef5ff !important;
    --borderPrimary: #edf0f5 !important;
    --borderSecondary: #d9e1ea !important;
    --dividerPrimary: rgba(15, 23, 42, 0.18) !important;
    --dividerSecondary: #fff !important;
    color-scheme: light;
  }

  html,
  body,
  #app {
    min-width: 0 !important;
    min-height: 100%;
    color: #111827;
    background: #f3f6fb !important;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Hiragino Sans GB",
      "Microsoft YaHei", "Helvetica Neue", Arial, sans-serif !important;
    letter-spacing: 0 !important;
  }

  body {
    margin: 0 !important;
    padding-bottom: 0 !important;
    font-size: 14px !important;
    line-height: 1.5715 !important;
  }

  *,
  *::before,
  *::after {
    box-sizing: border-box;
  }

  body > nav,
  #app > nav,
  nav {
    display: none !important;
  }

  body > .overlay,
  #app > div > .overlay {
    background: rgba(15, 23, 42, 0.28) !important;
  }

  main {
    width: 100% !important;
    max-width: none !important;
    min-height: 100vh;
    margin: 0 !important;
    padding: 0 10px 0 !important;
    background: #f3f6fb !important;
  }

  #download {
    pointer-events: none !important;
    visibility: hidden !important;
  }

  #loading {
    background: #f3f6fb !important;
  }

  #loading .spinner > div,
  main .spinner > div {
    background: #1677ff !important;
  }

  header {
    position: fixed !important;
    top: 0 !important;
    left: 0 !important;
    z-index: 1000 !important;
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    width: 100% !important;
    height: 56px !important;
    padding: 0 16px !important;
    color: #374151 !important;
    background: #fff !important;
    border: 0 !important;
    border-bottom: 1px solid #edf0f5 !important;
    box-shadow: none !important;
  }

  header > img,
  header title,
  header #search,
  header .search-button {
    display: none !important;
  }

  header .caifuclaw-filebrowser-brand {
    display: inline-flex !important;
    align-items: center !important;
    flex: 0 0 auto !important;
    gap: 8px !important;
    height: 56px !important;
    min-width: 0 !important;
    margin: 0 10px 0 0 !important;
    padding: 0 !important;
    color: #111827 !important;
    font-size: 18px !important;
    font-weight: 600 !important;
    line-height: 24px !important;
    white-space: nowrap !important;
  }

  header .caifuclaw-filebrowser-brand img {
    display: block !important;
    flex: 0 0 auto !important;
    width: 26px !important;
    height: 26px !important;
    margin: 0 !important;
    border-radius: 0 !important;
  }

  header .caifuclaw-filebrowser-brand span {
    display: inline-block !important;
    min-width: 0 !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
  }

  .breadcrumbs,
  header .breadcrumbs {
    position: static !important;
    top: auto !important;
    display: flex !important;
    flex: 1 1 auto !important;
    align-items: center !important;
    gap: 4px !important;
    min-width: 0 !important;
    min-height: 32px !important;
    height: 32px !important;
    margin: 0 !important;
    color: #374151 !important;
    background: transparent !important;
    border: 0 !important;
    opacity: 1 !important;
    overflow-x: auto !important;
    scrollbar-width: none;
  }

  .breadcrumbs::-webkit-scrollbar,
  header .breadcrumbs::-webkit-scrollbar {
    display: none;
  }

  .breadcrumbs > a,
  .breadcrumbs > span,
  .breadcrumbs span a,
  header .breadcrumbs > a,
  header .breadcrumbs > span,
  header .breadcrumbs span a {
    display: inline-flex !important;
    align-items: center !important;
    flex: 0 0 auto !important;
    min-width: 0 !important;
    height: 32px !important;
    padding: 0 6px !important;
    color: #374151 !important;
    font-size: 14px !important;
    line-height: 20px !important;
    border-radius: 6px !important;
    text-decoration: none !important;
    transition: color 0.2s ease, background 0.2s ease;
  }

  .breadcrumbs > a:hover,
  .breadcrumbs span a:hover,
  header .breadcrumbs > a:hover,
  header .breadcrumbs span a:hover {
    color: #1677ff !important;
    background: #eef5ff !important;
  }

  .breadcrumbs .chevron,
  header .breadcrumbs .chevron {
    display: inline-flex !important;
    align-items: center !important;
    color: #9ca3af !important;
  }

  .breadcrumbs i,
  header .breadcrumbs i {
    color: inherit !important;
    font-size: 20px !important;
  }

  .action,
  header .action,
  #dropdown .action,
  .context-menu .action,
  #file-selection .action,
  #multiple-selection .action {
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    flex: 0 0 auto !important;
    width: 34px !important;
    height: 34px !important;
    margin: 0 !important;
    padding: 0 !important;
    color: #374151 !important;
    background: transparent !important;
    border: 0 !important;
    border-radius: 6px !important;
    box-shadow: none !important;
    cursor: pointer !important;
    transition: color 0.2s ease, background 0.2s ease, border-color 0.2s ease;
  }

  .action:hover,
  .action:focus-visible {
    color: #1677ff !important;
    background: #eef5ff !important;
  }

  .action:focus-visible,
  .button:focus-visible,
  .input:focus-visible,
  input:focus-visible,
  select:focus-visible,
  textarea:focus-visible,
  [role="button"]:focus-visible {
    outline: 2px solid #1677ff !important;
    outline-offset: 2px !important;
  }

  .action i,
  header .action i,
  #dropdown .action i,
  .context-menu .action i,
  #file-selection .action i,
  #multiple-selection .action i {
    margin: 0 !important;
    padding: 0 !important;
    color: inherit !important;
    font-size: 20px !important;
    line-height: 1 !important;
    border-radius: 6px !important;
    text-shadow: none !important;
  }

  header .action span:not(.counter),
  #file-selection .action span:not(.counter),
  #multiple-selection .action span:not(.counter) {
    display: none !important;
  }

  .action.disabled,
  .action[disabled],
  button[disabled],
  .button[disabled] {
    color: #9ca3af !important;
    background: transparent !important;
    opacity: 0.55 !important;
    cursor: not-allowed !important;
  }

  .action .counter {
    right: -4px !important;
    bottom: -3px !important;
    width: 18px !important;
    height: 18px !important;
    color: #fff !important;
    font-size: 11px !important;
    font-weight: 600 !important;
    line-height: 16px !important;
    background: #1677ff !important;
    border: 1px solid #fff !important;
    border-radius: 999px !important;
  }

  #dropdown {
    position: static !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-end !important;
    flex: 0 0 auto !important;
    gap: 6px !important;
    margin-left: auto !important;
    min-width: 0 !important;
    background: transparent !important;
    border: 0 !important;
    box-shadow: none !important;
    transform: none !important;
  }

  #dropdown > div {
    display: inline-flex !important;
    align-items: center !important;
    gap: 6px !important;
  }

  #dropdown .action.caifuclaw-filebrowser-batch-ready {
    color: #1677ff !important;
    background: #eef5ff !important;
  }

  header .action,
  #dropdown .action {
    display: none !important;
  }

  header .action.caifuclaw-filebrowser-toolbar-download,
  #dropdown .action.caifuclaw-filebrowser-toolbar-download {
    display: inline-flex !important;
    gap: 6px !important;
    width: auto !important;
    min-width: 78px !important;
    padding: 0 12px !important;
    color: #1677ff !important;
    font-weight: 500 !important;
    background: #eef5ff !important;
  }

  header > .action.caifuclaw-filebrowser-toolbar-download {
    margin-left: auto !important;
  }

  header .action.caifuclaw-filebrowser-toolbar-download span.caifuclaw-filebrowser-download-label,
  #dropdown .action.caifuclaw-filebrowser-toolbar-download span.caifuclaw-filebrowser-download-label {
    display: inline-block !important;
    min-width: 0 !important;
    color: inherit !important;
    font-size: 14px !important;
    line-height: 20px !important;
    white-space: nowrap !important;
  }

  header .action.caifuclaw-filebrowser-toolbar-download .counter,
  #dropdown .action.caifuclaw-filebrowser-toolbar-download .counter {
    display: none !important;
  }

  #more {
    display: none !important;
  }

  #listing {
    width: 100% !important;
    max-width: 100% !important;
    min-height: calc(100vh - 88px) !important;
    margin: 0 !important;
    color: #111827 !important;
  }

  #listing.list {
    display: flex !important;
    flex-direction: column !important;
    overflow: hidden !important;
    background: #fff !important;
    border: 1px solid #edf0f5 !important;
    border-radius: 6px 6px 0 0 !important;
  }

  #listing h2 {
    display: none !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row,
  #listing.list .item.header.caifuclaw-filebrowser-header {
    align-items: center !important;
    display: grid !important;
    grid-template-columns: 34px 52px minmax(220px, clamp(220px, 24vw, 460px)) 112px 180px 96px !important;
    column-gap: 0 !important;
  }

  #listing.list .item,
  #listing.list .item.caifuclaw-filebrowser-row {
    width: 100% !important;
    min-height: 52px !important;
    margin: 0 !important;
    padding: 0 16px !important;
    color: #374151 !important;
    background: #fff !important;
    border: 0 !important;
    border-bottom: 1px solid #e5e7eb !important;
    box-shadow: none !important;
    transition: background 0.16s ease, color 0.16s ease;
  }

  #listing.list .item:hover {
    background: #f8fafc !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row {
    cursor: pointer !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row:focus-visible {
    background: #f8fbff !important;
    outline: 2px solid rgba(22, 119, 255, 0.36) !important;
    outline-offset: -2px !important;
  }

  #listing.list .item.header,
  #listing.list .item.header.caifuclaw-filebrowser-header {
    position: sticky !important;
    top: 56px !important;
    z-index: 5 !important;
    min-height: 40px !important;
    padding: 0 16px !important;
    color: #4b5563 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    background: #fafafa !important;
    border: 0 !important;
    border-bottom: 1px solid #e5e7eb !important;
  }

  #listing.list .item.header p {
    color: #4b5563 !important;
    font-weight: 600 !important;
  }

  #listing.list .item.header i {
    color: #6b7280 !important;
    opacity: 1 !important;
  }

  #listing.list .item.header.caifuclaw-filebrowser-header .name,
  #listing.list .item.header.caifuclaw-filebrowser-header .size,
  #listing.list .item.header.caifuclaw-filebrowser-header .modified {
    cursor: pointer !important;
    user-select: none !important;
  }

  #listing.list .item.header.caifuclaw-filebrowser-header .caifuclaw-filebrowser-sort-active,
  #listing.list .item.header.caifuclaw-filebrowser-header .caifuclaw-filebrowser-sort-active i {
    color: #1677ff !important;
  }

  #listing.list .item.header.caifuclaw-filebrowser-header .caifuclaw-filebrowser-header-label {
    display: inline-block !important;
    min-width: 0 !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row.caifuclaw-filebrowser-selected {
    background: #eaf3ff !important;
    box-shadow: inset 0 0 0 1px rgba(22, 119, 255, 0.16) !important;
    color: #111827 !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row.caifuclaw-filebrowser-selected:hover {
    background: #e4efff !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row.caifuclaw-filebrowser-selected .name,
  #listing.list .item.caifuclaw-filebrowser-row.caifuclaw-filebrowser-selected .size,
  #listing.list .item.caifuclaw-filebrowser-row.caifuclaw-filebrowser-selected .modified {
    color: #111827 !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row > div:first-of-type {
    grid-column: 2 !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    justify-self: start !important;
    width: 52px !important;
  }

  #listing.list .item.caifuclaw-filebrowser-row > div:nth-of-type(2) {
    display: contents !important;
  }

  #listing.list .item.header.caifuclaw-filebrowser-header > div {
    display: contents !important;
  }

  #listing.list .item div:first-of-type i {
    margin: 0 !important;
    color: #1677ff;
    font-size: 24px !important;
  }

  #listing.list .item div:first-of-type img {
    width: 36px !important;
    height: 36px !important;
    margin: 0 !important;
    border-radius: 6px !important;
    object-fit: cover !important;
  }

  #listing.list .item .caifuclaw-filebrowser-select-cell,
  #listing.list .item .caifuclaw-filebrowser-select-all-cell {
    align-items: center;
    cursor: pointer;
    display: flex;
    grid-column: 1 !important;
    justify-content: center;
    justify-self: start;
    margin: 0;
    min-height: 34px;
    width: 34px;
    border-radius: 6px;
    transition: background 0.2s ease;
  }

  #listing.list .item .caifuclaw-filebrowser-icon-spacer {
    display: block;
    grid-column: 2 !important;
    width: 52px;
    height: 1px;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-cell {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    grid-column: 2 !important;
    width: 52px !important;
    height: 52px !important;
    color: #64748b !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-button {
    display: inline-grid !important;
    place-items: center !important;
    width: 40px !important;
    height: 40px !important;
    margin: 0 !important;
    padding: 0 !important;
    color: inherit !important;
    background: transparent !important;
    border: 0 !important;
    border-radius: 6px !important;
    cursor: zoom-in !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-button:hover .caifuclaw-filebrowser-preview-frame,
  #listing.list .item .caifuclaw-filebrowser-preview-button:focus-visible .caifuclaw-filebrowser-preview-frame {
    border-color: #1677ff !important;
    box-shadow: 0 0 0 2px rgba(22, 119, 255, 0.12) !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-button:focus-visible {
    outline: 2px solid #1677ff !important;
    outline-offset: 2px !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-frame {
    position: relative !important;
    display: inline-grid !important;
    place-items: center !important;
    width: 36px !important;
    height: 36px !important;
    overflow: hidden !important;
    color: #64748b !important;
    background: #f8fafc !important;
    border: 1px solid #d9e2ec !important;
    border-radius: 6px !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-frame img {
    position: relative !important;
    z-index: 1 !important;
    display: block !important;
    width: 100% !important;
    height: 100% !important;
    object-fit: cover !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-frame i {
    position: absolute !important;
    z-index: 0 !important;
    margin: 0 !important;
    color: inherit !important;
    font-size: 22px !important;
    line-height: 1 !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-frame i::before {
    content: none !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-cell--folder .caifuclaw-filebrowser-preview-frame {
    color: #1677ff !important;
    background: #eef5ff !important;
    border-color: #dbeafe !important;
  }

  #listing.list .item .caifuclaw-filebrowser-preview-cell--file .caifuclaw-filebrowser-preview-frame {
    color: #6b7280 !important;
    background: #f8fafc !important;
  }

  #listing.list .item .caifuclaw-filebrowser-select-cell:hover,
  #listing.list .item .caifuclaw-filebrowser-select-all-cell:hover {
    background: #eef5ff;
  }

  #listing.list .item.caifuclaw-filebrowser-row.caifuclaw-filebrowser-selected .caifuclaw-filebrowser-select-cell {
    background: rgba(22, 119, 255, 0.10);
  }

  #listing.list .item .caifuclaw-filebrowser-select-cell input,
  #listing.list .item .caifuclaw-filebrowser-select-all-cell input {
    accent-color: #1677ff;
    cursor: pointer;
    height: 16px;
    margin: 0;
    width: 16px;
  }

  #listing.list .item .caifuclaw-filebrowser-sr-only {
    clip: rect(0 0 0 0);
    clip-path: inset(50%);
    height: 1px;
    overflow: hidden;
    position: absolute;
    white-space: nowrap;
    width: 1px;
  }

  #listing.list .item .name {
    grid-column: 3 !important;
    min-width: 0 !important;
    width: auto !important;
    color: #111827 !important;
    font-weight: 500 !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    white-space: nowrap !important;
    word-break: normal !important;
  }

  #listing.list .item .caifuclaw-filebrowser-native-icon-hidden {
    display: none !important;
  }

  #listing.list .item .size {
    grid-column: 4 !important;
    justify-self: end !important;
    width: auto !important;
    color: #6b7280 !important;
    font-variant-numeric: tabular-nums;
    font-weight: 400 !important;
    text-align: right !important;
  }

  #listing.list .item .modified {
    grid-column: 5 !important;
    justify-self: center !important;
    color: #6b7280 !important;
    font-variant-numeric: tabular-nums;
    font-weight: 400 !important;
    text-align: center !important;
    white-space: nowrap !important;
    width: auto !important;
  }

  #listing.list .item .caifuclaw-filebrowser-download-title,
  #listing.list .item .caifuclaw-filebrowser-download-cell {
    grid-column: 6 !important;
    justify-self: center !important;
    width: 96px !important;
    text-align: center !important;
  }

  #listing.list .item .caifuclaw-filebrowser-download-title {
    color: #4b5563;
    font-weight: 600;
  }

  #listing.list .item .caifuclaw-filebrowser-download-cell {
    align-items: center;
    border-radius: 6px;
    color: #374151;
    display: flex;
    height: 34px;
    justify-content: center;
    margin: 0;
    text-decoration: none;
  }

  #listing.list .item .caifuclaw-filebrowser-download-cell:hover {
    color: #1677ff;
    background-color: #eef5ff;
  }

  #listing.list .item .caifuclaw-filebrowser-download-cell i {
    color: inherit;
    font-size: 20px;
    padding: 0;
  }

  #listing.list .item .caifuclaw-filebrowser-download-cell i::before {
    content: none !important;
  }

  #file-selection,
  #multiple-selection {
    position: fixed !important;
    left: 50% !important;
    z-index: 1001 !important;
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    width: min(520px, calc(100vw - 24px)) !important;
    min-height: 48px !important;
    padding: 8px 10px 8px 14px !important;
    color: #111827 !important;
    background: #fff !important;
    border: 1px solid #edf0f5 !important;
    border-radius: 6px !important;
    box-shadow: 0 6px 16px rgba(15, 23, 42, 0.10) !important;
    transform: translateX(-50%) !important;
  }

  #file-selection {
    bottom: 12px !important;
  }

  #multiple-selection {
    bottom: -72px !important;
    justify-content: space-between !important;
    transition: bottom 0.2s ease !important;
  }

  #multiple-selection.active {
    bottom: 12px !important;
  }

  #file-selection > span,
  #multiple-selection p {
    flex: 1 1 auto !important;
    min-width: 0 !important;
    margin: 0 !important;
    color: #374151 !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    white-space: nowrap !important;
  }

  #file-selection .action,
  #multiple-selection .action {
    width: 32px !important;
    height: 32px !important;
  }

  .context-menu {
    min-width: 188px !important;
    max-width: 240px !important;
    padding: 4px !important;
    color: #111827 !important;
    background: #fff !important;
    border: 1px solid #edf0f5 !important;
    border-radius: 6px !important;
    box-shadow: 0 6px 16px rgba(15, 23, 42, 0.12) !important;
    overflow: hidden !important;
  }

  .context-menu .action {
    justify-content: flex-start !important;
    width: 100% !important;
    height: 36px !important;
    padding: 0 10px !important;
    gap: 8px !important;
    color: #374151 !important;
    border-radius: 4px !important;
  }

  .context-menu .action span:not(.counter) {
    display: inline !important;
    min-width: 0 !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    white-space: nowrap !important;
  }

  .context-menu .action .counter {
    position: static !important;
    margin-left: auto !important;
  }

  #modal-background {
    background: rgba(15, 23, 42, 0.28) !important;
  }

  .card {
    color: #111827 !important;
    background: #fff !important;
    border: 1px solid #edf0f5 !important;
    border-radius: 6px !important;
    box-shadow: none !important;
  }

  .card.floating {
    width: min(420px, calc(100vw - 32px)) !important;
    max-width: min(420px, calc(100vw - 32px)) !important;
    max-height: calc(100vh - 48px) !important;
    border-radius: 6px !important;
    overflow: auto !important;
  }

  .card .card-title {
    align-items: center !important;
    min-height: 48px !important;
    padding: 14px 18px !important;
    border-bottom: 1px solid #edf0f5 !important;
  }

  .card .card-title h2,
  .card h2 {
    margin: 0 !important;
    color: #111827 !important;
    font-size: 16px !important;
    font-weight: 600 !important;
    line-height: 24px !important;
  }

  .card h3 {
    color: #111827 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
  }

  .card > div,
  .card .card-content {
    padding: 16px 18px !important;
  }

  .card .card-action {
    display: flex !important;
    justify-content: flex-end !important;
    gap: 8px !important;
    padding: 12px 18px 16px !important;
    border-top: 1px solid #edf0f5 !important;
  }

  .button,
  input.button,
  button.button {
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    min-height: 32px !important;
    padding: 0 12px !important;
    color: #fff !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    line-height: 30px !important;
    text-transform: none !important;
    background: #1677ff !important;
    border: 1px solid #1677ff !important;
    border-radius: 6px !important;
    box-shadow: none !important;
  }

  .button:hover,
  input.button:hover,
  button.button:hover {
    background: #0958d9 !important;
    border-color: #0958d9 !important;
  }

  .button--flat,
  input.button--flat,
  button.button--flat {
    color: #1677ff !important;
    background: transparent !important;
    border-color: transparent !important;
  }

  .button--flat:hover,
  input.button--flat:hover,
  button.button--flat:hover {
    color: #0958d9 !important;
    background: #eef5ff !important;
    border-color: transparent !important;
  }

  .button--red,
  .button--flat.button--red {
    color: #dc2626 !important;
    background: #fff1f0 !important;
    border-color: #fecaca !important;
  }

  .input,
  input[type="text"],
  input[type="password"],
  input[type="number"],
  input[type="search"],
  input[type="email"],
  select,
  textarea {
    min-height: 32px !important;
    color: #111827 !important;
    background: #fff !important;
    border: 1px solid #d9e1ea !important;
    border-radius: 6px !important;
    box-shadow: none !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease;
  }

  .input:hover,
  input[type="text"]:hover,
  input[type="password"]:hover,
  input[type="number"]:hover,
  input[type="search"]:hover,
  input[type="email"]:hover,
  select:hover,
  textarea:hover,
  .input:focus,
  input[type="text"]:focus,
  input[type="password"]:focus,
  input[type="number"]:focus,
  input[type="search"]:focus,
  input[type="email"]:focus,
  select:focus,
  textarea:focus {
    border-color: #1677ff !important;
    box-shadow: 0 0 0 2px rgba(22, 119, 255, 0.12) !important;
    outline: 0 !important;
  }

  table tr {
    border-bottom: 1px solid #edf0f5 !important;
  }

  table th {
    color: #4b5563 !important;
    font-weight: 600 !important;
    background: #fafafa !important;
  }

  table th,
  table td {
    padding: 8px 12px !important;
  }

  h2.message {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 8px !important;
    min-height: 220px !important;
    margin: 0 !important;
    color: #6b7280 !important;
    font-size: 15px !important;
    font-weight: 500 !important;
  }

  h2.message i {
    color: #9ca3af !important;
    font-size: 24px !important;
  }

  .upload-files .card.floating {
    right: 16px !important;
    bottom: 16px !important;
    left: auto !important;
    width: min(360px, calc(100vw - 32px)) !important;
    max-width: min(360px, calc(100vw - 32px)) !important;
    transform: none !important;
  }

  .upload-files .file .file-name {
    color: #111827 !important;
    font-size: 14px !important;
    font-weight: 500 !important;
  }

  .upload-files .file .file-progress {
    height: 6px !important;
    background: #edf0f5 !important;
    border-radius: 999px !important;
  }

  .upload-files .file .file-progress div {
    background: #1677ff !important;
  }

  .Vue-Toastification__toast {
    min-height: 48px !important;
    padding: 12px 16px !important;
    font-family: inherit !important;
    border-radius: 6px !important;
    box-shadow: 0 6px 16px rgba(15, 23, 42, 0.12) !important;
  }

  .Vue-Toastification__toast-body {
    font-size: 14px !important;
    line-height: 22px !important;
  }

  .shell,
  .shell__content {
    background: #fff !important;
    border-top: 1px solid #edf0f5 !important;
    box-shadow: none !important;
  }

  #editor-container {
    background: #f3f6fb !important;
  }

  #editor-container .bar {
    background: #fff !important;
    border-bottom: 1px solid #edf0f5 !important;
  }

  #previewer {
    background: #0f172a !important;
  }

  #previewer header {
    color: #fff !important;
    background: transparent !important;
    border: 0 !important;
    box-shadow: none !important;
  }

  #previewer header .action {
    color: #fff !important;
  }

  #previewer header .action:hover {
    background: rgba(255, 255, 255, 0.18) !important;
  }

  #previewer header title {
    display: block !important;
    min-width: 0 !important;
    color: #fff !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    white-space: nowrap !important;
  }

  .caifuclaw-filebrowser-image-preview[hidden] {
    display: none !important;
  }

  .caifuclaw-filebrowser-image-preview {
    position: fixed !important;
    inset: 0 !important;
    z-index: 1005 !important;
    display: grid !important;
    place-items: center !important;
    padding: 24px !important;
    background: rgba(15, 23, 42, 0.72) !important;
  }

  .caifuclaw-filebrowser-image-preview__dialog {
    display: grid !important;
    grid-template-rows: 52px minmax(0, 1fr) !important;
    width: min(960px, calc(100vw - 48px)) !important;
    height: min(760px, calc(100vh - 48px)) !important;
    overflow: hidden !important;
    color: #111827 !important;
    background: #fff !important;
    border: 1px solid rgba(255, 255, 255, 0.28) !important;
    border-radius: 6px !important;
    box-shadow: 0 18px 42px rgba(15, 23, 42, 0.28) !important;
  }

  .caifuclaw-filebrowser-image-preview__header {
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    min-width: 0 !important;
    padding: 0 12px 0 16px !important;
    background: #fff !important;
    border-bottom: 1px solid #edf0f5 !important;
  }

  .caifuclaw-filebrowser-image-preview__title {
    flex: 1 1 auto !important;
    min-width: 0 !important;
    margin: 0 !important;
    color: #111827 !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    line-height: 22px !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    white-space: nowrap !important;
  }

  .caifuclaw-filebrowser-image-preview__action {
    display: inline-grid !important;
    place-items: center !important;
    flex: 0 0 auto !important;
    width: 34px !important;
    height: 34px !important;
    margin: 0 !important;
    padding: 0 !important;
    color: #374151 !important;
    background: transparent !important;
    border: 0 !important;
    border-radius: 6px !important;
    text-decoration: none !important;
    cursor: pointer !important;
  }

  .caifuclaw-filebrowser-image-preview__action:hover,
  .caifuclaw-filebrowser-image-preview__action:focus-visible {
    color: #1677ff !important;
    background: #eef5ff !important;
    outline: 0 !important;
  }

  .caifuclaw-filebrowser-image-preview__action i {
    color: inherit !important;
    font-size: 20px !important;
    line-height: 1 !important;
  }

  .caifuclaw-filebrowser-image-preview__body {
    display: grid !important;
    place-items: center !important;
    min-width: 0 !important;
    min-height: 0 !important;
    padding: 16px !important;
    background: #0f172a !important;
  }

  .caifuclaw-filebrowser-image-preview__body img {
    display: block !important;
    max-width: 100% !important;
    max-height: 100% !important;
    object-fit: contain !important;
    border-radius: 4px !important;
  }

  .caifuclaw-filebrowser-image-preview__error {
    color: #fff !important;
    font-size: 14px !important;
  }

  body.caifuclaw-filebrowser-image-preview-open {
    overflow: hidden !important;
  }

  @media (max-width: 736px) {
    main {
      padding: 0 8px 0 !important;
    }

    header {
      height: 56px !important;
      padding: 0 10px !important;
    }

    header .breadcrumbs {
      height: 56px !important;
      min-width: 0 !important;
    }

    #dropdown {
      position: static !important;
      display: flex !important;
      min-width: 0 !important;
      padding: 0 !important;
      background: transparent !important;
      border: 0 !important;
      box-shadow: none !important;
      transform: none !important;
    }

    #dropdown > div {
      display: inline-flex !important;
      gap: 6px !important;
    }

    #dropdown .action {
      justify-content: center !important;
      width: auto !important;
      min-width: 78px !important;
      height: 34px !important;
      padding: 0 12px !important;
      gap: 6px !important;
      border-radius: 6px !important;
    }

    #dropdown .action.caifuclaw-filebrowser-toolbar-download {
      display: inline-flex !important;
    }

    #listing.list .item,
    #listing.list .item.caifuclaw-filebrowser-row,
    #listing.list .item.header,
    #listing.list .item.header.caifuclaw-filebrowser-header {
      padding-right: 10px !important;
      padding-left: 10px !important;
    }

    #listing.list .item .caifuclaw-filebrowser-download-title,
    #listing.list .item .caifuclaw-filebrowser-download-cell {
      width: 68px !important;
    }

    #listing.list .item.caifuclaw-filebrowser-row,
    #listing.list .item.header.caifuclaw-filebrowser-header {
      grid-template-columns: 30px 44px minmax(160px, 1fr) 0 10.75rem 68px !important;
    }

    #listing.list .item .caifuclaw-filebrowser-select-cell,
    #listing.list .item .caifuclaw-filebrowser-select-all-cell {
      width: 30px;
    }

    #listing.list .item.caifuclaw-filebrowser-row > div:first-of-type,
    #listing.list .item .caifuclaw-filebrowser-icon-spacer {
      width: 44px !important;
    }

    #listing.list .item .caifuclaw-filebrowser-preview-cell {
      width: 44px !important;
    }

    #listing.list .item .size {
      display: none !important;
    }

    #listing.list .item .modified {
      justify-self: center !important;
    }

    #file-selection,
    #multiple-selection {
      width: calc(100vw - 16px) !important;
      bottom: 8px !important;
    }

    #multiple-selection {
      bottom: -72px !important;
    }

    #multiple-selection.active {
      bottom: 8px !important;
    }

    .caifuclaw-filebrowser-image-preview {
      padding: 12px !important;
    }

    .caifuclaw-filebrowser-image-preview__dialog {
      width: calc(100vw - 24px) !important;
      height: calc(100vh - 24px) !important;
    }
  }

  @media (max-width: 450px) {
    #listing.list .item .modified {
      display: none !important;
    }

    #listing.list .item .name {
      min-width: 0 !important;
    }

    #listing.list .item.caifuclaw-filebrowser-row,
    #listing.list .item.header.caifuclaw-filebrowser-header {
      grid-template-columns: 30px 40px minmax(0, 1fr) 0 0 56px !important;
    }

    #listing.list .item.caifuclaw-filebrowser-row > div:first-of-type {
      width: 40px !important;
    }

    #listing.list .item .caifuclaw-filebrowser-icon-spacer {
      width: 40px !important;
    }

    #listing.list .item .caifuclaw-filebrowser-preview-cell {
      width: 40px !important;
    }
  }
</style>
""".strip()
FILE_BROWSER_CLEAN_UI_SCRIPT = """
<script id="caifuclaw-filebrowser-behavior">
(() => {
  const TITLE_MAP = [
    [".name", "名称", "按名称排序", "name"],
    [".size", "大小", "按大小排序", "size"],
    [".modified", "修改时间", "按修改时间排序", "modified"],
  ];
  const DEFAULT_SORT_DIRECTIONS = {
    name: "asc",
    size: "desc",
    modified: "desc",
  };
  const IMAGE_FILE_EXTENSIONS = new Set([
    ".apng",
    ".avif",
    ".bmp",
    ".gif",
    ".jpg",
    ".jpeg",
    ".png",
    ".svg",
    ".webp",
  ]);

  let currentMetadataDirectory = "";
  let directoryMetadata = new Map();
  let metadataRequestId = 0;
  let caifuclawSortState = { field: "modified", direction: "desc" };
  const selectedPaths = new Set();
  const originalPreviewCells = new WeakMap();
  const exactDateTimeFormatter = new Intl.DateTimeFormat("zh-CN", {
    timeZone: "Asia/Shanghai",
    year: "numeric",
    month: "2-digit",
    day: "2-digit",
    hour: "2-digit",
    minute: "2-digit",
    second: "2-digit",
    hour12: false,
    hourCycle: "h23",
  });

  const encodePath = (path) => path.split("/").map((part) => encodeURIComponent(part)).join("/");

  const baseUrl = () => {
    const configuredBase = window.FileBrowser?.BaseURL;
    return (typeof configuredBase === "string" ? configuredBase : "/filebrowser").replace(/\\/+$/, "");
  };

  const localBearerToken = () => {
    try {
      return window.localStorage?.getItem("token") || "";
    } catch {
      return "";
    }
  };

  const fileBrowserSessionHeaders = () => {
    const token = localBearerToken();
    return token ? { Authorization: `Bearer ${token}` } : {};
  };

  const originalFetch = window.fetch?.bind(window);
  let fileBrowserSessionRefreshPromise = null;

  const renewFileBrowserSession = () => {
    if (!originalFetch) {
      return Promise.reject(new Error("Fetch is unavailable"));
    }
    if (!fileBrowserSessionRefreshPromise) {
      fileBrowserSessionRefreshPromise = originalFetch("/api/v1/filebrowser/session", {
        method: "POST",
        credentials: "same-origin",
        headers: fileBrowserSessionHeaders(),
      })
        .then((response) => {
          if (!response.ok) {
            throw new Error(`File Browser session refresh failed: ${response.status}`);
          }
          return response;
        })
        .finally(() => {
          fileBrowserSessionRefreshPromise = null;
        });
    }
    return fileBrowserSessionRefreshPromise;
  };

  const urlForFetchInput = (input) => {
    try {
      if (typeof input === "string") {
        return new URL(input, window.location.origin);
      }
      if (input instanceof URL) {
        return input;
      }
      if (input?.url) {
        return new URL(input.url, window.location.origin);
      }
    } catch {
      return null;
    }
    return null;
  };

  const isFileBrowserApiRequest = (input) => {
    const url = urlForFetchInput(input);
    if (!url || url.origin !== window.location.origin) {
      return false;
    }
    const base = baseUrl();
    return ["resources", "usage", "raw"].some((name) => {
      const prefix = `${base}/api/${name}`;
      return url.pathname === prefix || url.pathname.startsWith(`${prefix}/`);
    });
  };

  const installFileBrowserFetchRecovery = () => {
    if (!originalFetch || window.__caifuclawFileBrowserFetchRecovery) {
      return;
    }
    window.__caifuclawFileBrowserFetchRecovery = true;
    window.fetch = (input, init) => {
      const shouldRecover = isFileBrowserApiRequest(input);
      return originalFetch(input, init)
        .then((response) => {
          if (!shouldRecover || response.status !== 401) {
            return response;
          }
          return renewFileBrowserSession()
            .then(() => originalFetch(input, init))
            .catch(() => response);
        })
        .catch((error) => {
          if (!shouldRecover) {
            throw error;
          }
          return renewFileBrowserSession().then(() => originalFetch(input, init));
        });
    };
  };

  installFileBrowserFetchRecovery();

  const currentDirectoryPath = () => {
    const base = baseUrl();
    let path = decodeURIComponent(window.location.pathname);
    if (base && path.startsWith(base)) {
      path = path.slice(base.length);
    }
    if (path.startsWith("/files")) {
      path = path.slice("/files".length);
    }
    if (!path.startsWith("/")) {
      path = `/${path}`;
    }
    return path.replace(/\\/+$/, "") || "/";
  };

  const resourcesUrlFor = (directory) => {
    const normalized = directory === "/" ? "/" : `${directory.replace(/\\/+$/, "")}/`;
    return `${baseUrl()}/api/resources${encodePath(normalized)}`;
  };

  const resourcePathFor = (name) => {
    const directory = currentDirectoryPath();
    return `${directory === "/" ? "" : directory}/${name}`;
  };

  const rowName = (row) => row.querySelector(".name")?.textContent?.trim() || row.getAttribute("aria-label") || "";

  const rowResourcePath = (row) => resourcePathFor(rowName(row));

  const listedRows = () => Array.from(document.querySelectorAll("#listing.list .item.caifuclaw-filebrowser-row:not(.header)"));

  const selectedRows = () => listedRows().filter((row) => selectedPaths.has(rowResourcePath(row)));

  const downloadUrlFor = (row) => {
    const name = rowName(row);
    const url = new URL(`${baseUrl()}/api/raw${encodePath(resourcePathFor(name))}`, window.location.origin);
    if (row.dataset.dir === "true") {
      url.searchParams.set("algo", "zip");
    }
    return url.toString();
  };

  const directoryUrlFor = (row) => {
    const path = resourcePathFor(rowName(row)).replace(/\\/+$/, "");
    return `${baseUrl()}/files${encodePath(path)}/`;
  };

  const fileExtension = (name) => {
    const index = name.lastIndexOf(".");
    return index >= 0 ? name.slice(index).toLowerCase() : "";
  };

  const isImageRow = (row) => row.dataset.dir !== "true" && IMAGE_FILE_EXTENSIONS.has(fileExtension(rowName(row)));

  const rowPreviewIcon = (row) => {
    if (row.dataset.dir === "true") {
      return "folder";
    }
    return "description";
  };

  const shouldUseImagePreviewCells = () => {
    const parts = currentDirectoryPath().split("/").filter(Boolean);
    return parts[0] === "down_image" && parts.length >= 3;
  };

  const serverReachabilityRecoveryKey = () => `caifuclaw:filebrowser-server-recovery:${window.location.pathname}`;

  const hasServerReachabilityError = () => (document.body?.textContent || "").includes("The server can't be reached");

  let serverReachabilityRecoveryStarted = false;

  const recoverServerReachabilityError = () => {
    if (serverReachabilityRecoveryStarted || !hasServerReachabilityError()) {
      return false;
    }
    try {
      if (window.sessionStorage?.getItem(serverReachabilityRecoveryKey()) === "true") {
        return false;
      }
    } catch {}
    serverReachabilityRecoveryStarted = true;
    renewFileBrowserSession()
      .then(() => {
        try {
          window.sessionStorage?.setItem(serverReachabilityRecoveryKey(), "true");
        } catch {}
        window.location.reload();
      })
      .catch(() => {
        serverReachabilityRecoveryStarted = false;
      });
    return true;
  };

  const recoverFromResourceError = (directory) => {
    if (serverReachabilityRecoveryStarted) {
      return Promise.resolve(null);
    }
    serverReachabilityRecoveryStarted = true;
    return renewFileBrowserSession()
      .then(() => fetch(resourcesUrlFor(directory), { credentials: "same-origin" }))
      .then((response) => (response.ok ? response.json() : null))
      .catch(() => null)
      .finally(() => {
        serverReachabilityRecoveryStarted = false;
      });
  };

  const clearServerReachabilityRecovery = () => {
    if (hasServerReachabilityError()) {
      return;
    }
    try {
      window.sessionStorage?.removeItem(serverReachabilityRecoveryKey());
    } catch {}
  };

  const rememberOriginalPreviewCell = (cell) => {
    if (originalPreviewCells.has(cell)) {
      return;
    }
    originalPreviewCells.set(cell, {
      ariaHidden: cell.getAttribute("aria-hidden"),
      className: cell.className,
      childNodes: Array.from(cell.childNodes).map((node) => node.cloneNode(true)),
      title: cell.getAttribute("title"),
    });
  };

  const restoreOriginalPreviewCell = (cell) => {
    const original = originalPreviewCells.get(cell);
    if (!original) {
      return;
    }
    cell.className = original.className;
    if (original.ariaHidden === null) {
      cell.removeAttribute("aria-hidden");
    } else {
      cell.setAttribute("aria-hidden", original.ariaHidden);
    }
    if (original.title === null) {
      cell.removeAttribute("title");
    } else {
      cell.setAttribute("title", original.title);
    }
    cell.replaceChildren(...original.childNodes.map((node) => node.cloneNode(true)));
    delete cell.dataset.caifuclawPreviewName;
    delete cell.dataset.caifuclawPreviewSrc;
    delete cell.dataset.caifuclawPreviewType;
  };

  const hideNativeNameIcons = (row) => {
    const nameCell = row.querySelector(".name");
    if (!nameCell) {
      return;
    }
    nameCell.querySelectorAll("i.material-icons").forEach((icon) => {
      const text = icon.textContent?.trim();
      if (["folder", "description", "insert_drive_file", "image"].includes(text)) {
        icon.classList.add("caifuclaw-filebrowser-native-icon-hidden");
        icon.setAttribute("aria-hidden", "true");
      }
    });
  };

  const formatExactDateTime = (value) => {
    if (!value) {
      return "";
    }
    const date = new Date(value);
    if (Number.isNaN(date.getTime())) {
      return "";
    }
    const parts = exactDateTimeFormatter.formatToParts(date);
    const part = (type) => parts.find((item) => item.type === type)?.value || "00";
    return `${part("year")}-${part("month")}-${part("day")} ${part("hour")}:${part("minute")}:${part("second")}`;
  };

  const setText = (element, value) => {
    if (element && element.textContent !== value) {
      element.textContent = value;
    }
  };

  const ensureHeaderLabel = (cell, label) => {
    let labelElement = cell.querySelector(".caifuclaw-filebrowser-header-label") || cell.querySelector("span");
    if (!labelElement) {
      const icon = cell.querySelector("i");
      const textNodes = Array.from(cell.childNodes).filter((node) => (
        node.nodeType === Node.TEXT_NODE && node.textContent?.trim()
      ));
      labelElement = document.createElement("span");
      labelElement.className = "caifuclaw-filebrowser-header-label";
      if (icon) {
        cell.insertBefore(labelElement, icon);
      } else {
        cell.prepend(labelElement);
      }
      textNodes.forEach((node) => node.remove());
    }
    labelElement.classList.add("caifuclaw-filebrowser-header-label");
    setText(labelElement, label);
    return labelElement;
  };

  const compareRowNames = (left, right) => rowName(left).localeCompare(rowName(right), "zh-Hans-CN", {
    numeric: true,
    sensitivity: "base",
  });

  const rowSizeValue = (row) => {
    const metadataSize = directoryMetadata.get(rowResourcePath(row))?.size;
    if (Number.isFinite(metadataSize)) {
      return Number(metadataSize);
    }
    const text = row.querySelector(".size")?.textContent?.trim() || "";
    if (!text || text === "-") {
      return 0;
    }
    const match = text.replace(/,/g, "").match(/^([0-9.]+)\\s*(B|KB|MB|GB|TB)?$/i);
    if (!match) {
      return 0;
    }
    const multiplier = {
      B: 1,
      KB: 1024,
      MB: 1024 ** 2,
      GB: 1024 ** 3,
      TB: 1024 ** 4,
    }[(match[2] || "B").toUpperCase()] || 1;
    return Number(match[1]) * multiplier;
  };

  const filenameFromContentDisposition = (value) => {
    if (!value) {
      return "";
    }
    const encodedMatch = value.match(/filename\\*=([^;]+)/i);
    if (encodedMatch) {
      const raw = encodedMatch[1].trim().replace(/^UTF-8''/i, "").replace(/^"(.*)"$/, "$1");
      try {
        return decodeURIComponent(raw);
      } catch {
        return raw;
      }
    }
    const quotedMatch = value.match(/filename="([^"]+)"/i);
    if (quotedMatch) {
      return quotedMatch[1];
    }
    const plainMatch = value.match(/filename=([^;]+)/i);
    return plainMatch ? plainMatch[1].trim() : "";
  };

  const fallbackDownloadFilename = (row) => {
    const name = rowName(row) || "download";
    if (row.dataset.dir === "true" && !name.toLowerCase().endsWith(".zip")) {
      return `${name}.zip`;
    }
    return name;
  };

  const saveBlobDownload = (blob, filename) => {
    const objectUrl = URL.createObjectURL(blob);
    const link = document.createElement("a");
    link.href = objectUrl;
    link.download = filename || "download";
    link.rel = "noopener";
    document.body.appendChild(link);
    link.click();
    link.remove();
    window.setTimeout(() => URL.revokeObjectURL(objectUrl), 60000);
  };

  const triggerIframeDownloadFallback = (url) => {
    const frame = document.createElement("iframe");
    frame.hidden = true;
    frame.src = url;
    document.body.appendChild(frame);
    window.setTimeout(() => frame.remove(), 60000);
  };

  const triggerDownload = (url, filename = "") => {
    fetch(url, { credentials: "same-origin" })
      .then((response) => {
        if (!response.ok) {
          throw new Error(`Download failed: ${response.status}`);
        }
        const headerFilename = filenameFromContentDisposition(response.headers.get("content-disposition"));
        return response.blob().then((blob) => ({ blob, filename: headerFilename || filename }));
      })
      .then(({ blob, filename: resolvedFilename }) => saveBlobDownload(blob, resolvedFilename))
      .catch(() => triggerIframeDownloadFallback(url));
  };

  const triggerRowDownload = (row) => triggerDownload(downloadUrlFor(row), fallbackDownloadFilename(row));

  const downloadSelectedRows = () => {
    selectedRows().forEach((row, index) => {
      window.setTimeout(() => triggerRowDownload(row), index * 250);
    });
  };

  const ensureImagePreviewOverlay = () => {
    let overlay = document.querySelector(".caifuclaw-filebrowser-image-preview");
    if (overlay) {
      return overlay;
    }

    overlay = document.createElement("div");
    overlay.className = "caifuclaw-filebrowser-image-preview";
    overlay.hidden = true;
    overlay.setAttribute("role", "dialog");
    overlay.setAttribute("aria-modal", "true");
    overlay.setAttribute("aria-labelledby", "caifuclaw-filebrowser-image-preview-title");

    const dialog = document.createElement("div");
    dialog.className = "caifuclaw-filebrowser-image-preview__dialog";

    const header = document.createElement("div");
    header.className = "caifuclaw-filebrowser-image-preview__header";

    const title = document.createElement("p");
    title.id = "caifuclaw-filebrowser-image-preview-title";
    title.className = "caifuclaw-filebrowser-image-preview__title";

    const downloadLink = document.createElement("a");
    downloadLink.className = "caifuclaw-filebrowser-image-preview__action";
    downloadLink.href = "#";
    downloadLink.title = "下载图片";
    downloadLink.setAttribute("aria-label", "下载图片");
    downloadLink.addEventListener("click", (event) => {
      event.preventDefault();
      const url = downloadLink.href;
      const filename = overlay.dataset.filename || "image";
      if (url && url !== "#") {
        triggerDownload(url, filename);
      }
    });
    const downloadIcon = document.createElement("i");
    downloadIcon.className = "material-icons";
    downloadIcon.textContent = "file_download";
    downloadLink.appendChild(downloadIcon);

    const closeButton = document.createElement("button");
    closeButton.type = "button";
    closeButton.className = "caifuclaw-filebrowser-image-preview__action";
    closeButton.title = "关闭预览";
    closeButton.setAttribute("aria-label", "关闭预览");
    closeButton.addEventListener("click", () => closeImagePreviewOverlay());
    const closeIcon = document.createElement("i");
    closeIcon.className = "material-icons";
    closeIcon.textContent = "close";
    closeButton.appendChild(closeIcon);

    const body = document.createElement("div");
    body.className = "caifuclaw-filebrowser-image-preview__body";

    const image = document.createElement("img");
    image.alt = "";
    image.addEventListener("error", () => {
      const error = document.createElement("span");
      error.className = "caifuclaw-filebrowser-image-preview__error";
      error.textContent = "图片加载失败";
      body.replaceChildren(error);
    });
    body.appendChild(image);

    overlay.addEventListener("click", (event) => {
      if (event.target === overlay) {
        closeImagePreviewOverlay();
      }
    });
    overlay.addEventListener("keydown", (event) => {
      if (event.key === "Escape") {
        event.preventDefault();
        closeImagePreviewOverlay();
      }
    });

    header.appendChild(title);
    header.appendChild(downloadLink);
    header.appendChild(closeButton);
    dialog.appendChild(header);
    dialog.appendChild(body);
    overlay.appendChild(dialog);
    document.body.appendChild(overlay);
    return overlay;
  };

  const closeImagePreviewOverlay = () => {
    const overlay = document.querySelector(".caifuclaw-filebrowser-image-preview");
    if (!overlay) {
      return;
    }
    overlay.hidden = true;
    document.body.classList.remove("caifuclaw-filebrowser-image-preview-open");
    const image = overlay.querySelector("img");
    if (image) {
      image.removeAttribute("src");
    }
  };

  const openImagePreviewOverlay = (src, name) => {
    const overlay = ensureImagePreviewOverlay();
    const title = overlay.querySelector(".caifuclaw-filebrowser-image-preview__title");
    const body = overlay.querySelector(".caifuclaw-filebrowser-image-preview__body");
    let image = overlay.querySelector("img");
    if (!image) {
      image = document.createElement("img");
      image.alt = "";
      image.addEventListener("error", () => {
        const error = document.createElement("span");
        error.className = "caifuclaw-filebrowser-image-preview__error";
        error.textContent = "图片加载失败";
        body?.replaceChildren(error);
      });
    }
    if (body) {
      body.replaceChildren(image);
    }
    const downloadLink = overlay.querySelector(".caifuclaw-filebrowser-image-preview__action[href]");
    setText(title, name || "图片预览");
    image.alt = name || "图片预览";
    image.src = src;
    overlay.dataset.filename = name || "image";
    if (downloadLink) {
      downloadLink.href = src;
      downloadLink.setAttribute("aria-label", `下载图片 ${name || ""}`.trim());
    }
    overlay.hidden = false;
    document.body.classList.add("caifuclaw-filebrowser-image-preview-open");
    overlay.querySelector("button")?.focus({ preventScroll: true });
  };

  const ensureHeaderBrand = () => {
    const header = document.querySelector("header");
    if (!header || header.querySelector(".caifuclaw-filebrowser-brand")) {
      return;
    }
    const brand = document.createElement("div");
    brand.className = "caifuclaw-filebrowser-brand";
    brand.setAttribute("aria-label", "CaifuClaw");

    const logo = document.createElement("img");
    logo.src = "/caifuclaw-logo.png";
    logo.alt = "";

    const name = document.createElement("span");
    name.textContent = "CaifuClaw";

    brand.appendChild(logo);
    brand.appendChild(name);
    header.insertBefore(brand, header.firstChild);
  };

  const markToolbarDownloadActions = () => {
    document.querySelectorAll("header .action, #dropdown .action").forEach((button) => {
      const iconText = button.querySelector("i")?.textContent?.trim();
      const isDownload = iconText === "file_download";
      button.classList.toggle("caifuclaw-filebrowser-toolbar-download", isDownload);
      if (!isDownload) {
        return;
      }
      let label = Array.from(button.querySelectorAll("span")).find((item) => !item.classList.contains("counter"));
      if (!label) {
        label = document.createElement("span");
        button.appendChild(label);
      }
      label.classList.add("caifuclaw-filebrowser-download-label");
      setText(label, "下载");
    });
  };

  const forceListView = () => {
    const listing = document.querySelector("#listing");
    if (!listing) {
      return;
    }
    listing.classList.add("list");
    listing.classList.remove("mosaic", "gallery");
  };

  const loadDirectoryMetadata = () => {
    const directory = currentDirectoryPath();
    if (directory === currentMetadataDirectory) {
      return;
    }
    currentMetadataDirectory = directory;
    directoryMetadata = new Map();
    selectedPaths.clear();
    const requestId = ++metadataRequestId;
    fetch(resourcesUrlFor(directory), { credentials: "same-origin" })
      .then((response) => (response.ok ? response.json() : null))
      .then((payload) => (payload ? payload : recoverFromResourceError(directory)))
      .then((payload) => {
        if (requestId !== metadataRequestId || !payload) {
          return;
        }
        directoryMetadata = new Map((payload.items || []).map((item) => [item.path, item]));
        scheduleApply();
      })
      .catch(() => {
        recoverFromResourceError(directory).then((payload) => {
          if (requestId !== metadataRequestId || !payload) {
            return;
          }
          directoryMetadata = new Map((payload.items || []).map((item) => [item.path, item]));
          scheduleApply();
        });
      });
  };

  const updateExactModifiedTime = (row) => {
    const modifiedCell = row.querySelector(".modified");
    if (!modifiedCell) {
      return;
    }
    const time = modifiedCell.querySelector("time");
    const metadata = directoryMetadata.get(rowResourcePath(row));
    const isoTime = metadata?.modified || time?.getAttribute("datetime") || "";
    const exactTime = formatExactDateTime(isoTime);
    if (!exactTime) {
      return;
    }
    if (time) {
      setText(time, exactTime);
      if (time.getAttribute("title") !== exactTime) {
        time.setAttribute("title", exactTime);
      }
      if (isoTime && time.getAttribute("datetime") !== isoTime) {
        time.setAttribute("datetime", isoTime);
      }
    } else {
      setText(modifiedCell, exactTime);
    }
    if (modifiedCell.getAttribute("title") !== exactTime) {
      modifiedCell.setAttribute("title", exactTime);
    }
  };

  const rowModifiedTimestamp = (row) => {
    const metadataTime = directoryMetadata.get(rowResourcePath(row))?.modified;
    const timeElement = row.querySelector(".modified time");
    const rawTime = metadataTime || timeElement?.getAttribute("datetime") || row.querySelector(".modified")?.textContent || "";
    const timestamp = new Date(rawTime).getTime();
    return Number.isNaN(timestamp) ? 0 : timestamp;
  };

  const compareRowsByActiveSort = (left, right) => {
    const directionFactor = caifuclawSortState.direction === "asc" ? 1 : -1;
    if (caifuclawSortState.field === "name") {
      return compareRowNames(left, right) * directionFactor;
    }
    const valueDiff = caifuclawSortState.field === "size"
      ? rowSizeValue(left) - rowSizeValue(right)
      : rowModifiedTimestamp(left) - rowModifiedTimestamp(right);
    if (valueDiff !== 0) {
      return valueDiff * directionFactor;
    }
    return compareRowNames(left, right);
  };

  const sortRowsByActiveSort = () => {
    const rows = listedRows();
    if (rows.length < 2) {
      return;
    }
    const sortedRows = [...rows].sort(compareRowsByActiveSort);
    const parent = rows[0].parentElement;
    if (!parent) {
      return;
    }
    if (sortedRows.every((row, index) => row === rows[index])) {
      return;
    }
    sortedRows.forEach((row) => parent.appendChild(row));
  };

  const updateHeaderSortCell = (cell, field, label, sortLabel) => {
    const isActive = caifuclawSortState.field === field;
    const direction = caifuclawSortState.direction;
    const nextDirection = isActive
      ? (direction === "asc" ? "desc" : "asc")
      : DEFAULT_SORT_DIRECTIONS[field];
    const activeLabel = direction === "asc" ? "升序" : "降序";
    cell.classList.toggle("caifuclaw-filebrowser-sort-active", isActive);
    cell.setAttribute("aria-sort", isActive ? (direction === "asc" ? "ascending" : "descending") : "none");
    cell.setAttribute("title", isActive ? `${label}，当前${activeLabel}，点击切换排序` : sortLabel);
    cell.setAttribute("aria-label", isActive ? `${label}，当前${activeLabel}，点击切换排序` : sortLabel);
    cell.dataset.caifuclawSortField = field;
    cell.dataset.caifuclawSortDirection = isActive ? direction : nextDirection;

    let icon = cell.querySelector("i");
    if (!icon) {
      icon = document.createElement("i");
      icon.className = "material-icons";
      cell.appendChild(icon);
    }
    icon.setAttribute("aria-hidden", "true");
    setText(icon, isActive ? (direction === "asc" ? "arrow_upward" : "arrow_downward") : "unfold_more");
  };

  const updateAllHeaderSortCells = () => {
    document.querySelectorAll("#listing.list .item.header.caifuclaw-filebrowser-header").forEach((header) => {
      for (const [selector, label, sortLabel, field] of TITLE_MAP) {
        const cell = header.querySelector(selector);
        if (cell) {
          updateHeaderSortCell(cell, field, label, sortLabel);
        }
      }
    });
  };

  const setActiveSort = (field) => {
    caifuclawSortState = {
      field,
      direction: caifuclawSortState.field === field
        ? (caifuclawSortState.direction === "asc" ? "desc" : "asc")
        : DEFAULT_SORT_DIRECTIONS[field],
    };
    sortRowsByActiveSort();
    updateAllHeaderSortCells();
    updateSelectionUi();
  };

  const bindHeaderSortCell = (cell, field) => {
    if (cell.dataset.caifuclawSortBound === "true") {
      return;
    }
    cell.dataset.caifuclawSortBound = "true";
    cell.setAttribute("role", "button");
    cell.setAttribute("tabindex", "0");
    cell.addEventListener("click", (event) => {
      event.preventDefault();
      event.stopPropagation();
      event.stopImmediatePropagation();
      setActiveSort(field);
    }, true);
    cell.addEventListener("keydown", (event) => {
      if (![" ", "Enter"].includes(event.key)) {
        return;
      }
      event.preventDefault();
      event.stopPropagation();
      setActiveSort(field);
    }, true);
  };

  const updateSelectionUi = () => {
    const rows = listedRows();
    const checkedRows = rows.filter((row) => selectedPaths.has(rowResourcePath(row)));
    rows.forEach((row) => {
      const checked = selectedPaths.has(rowResourcePath(row));
      row.classList.toggle("caifuclaw-filebrowser-selected", checked);
      row.dataset.caifuclawSelected = checked ? "true" : "false";
      row.setAttribute("aria-selected", checked ? "true" : "false");
      const input = row.querySelector(".caifuclaw-filebrowser-select");
      if (input) {
        input.checked = checked;
      }
    });

    document.querySelectorAll(".caifuclaw-filebrowser-select-all").forEach((input) => {
      input.checked = rows.length > 0 && checkedRows.length === rows.length;
      input.indeterminate = checkedRows.length > 0 && checkedRows.length < rows.length;
      input.disabled = rows.length === 0;
      input.closest("label")?.setAttribute(
        "title",
        rows.length ? `已选择 ${checkedRows.length} / ${rows.length} 项` : "当前目录没有可选择项目",
      );
    });

    document.querySelectorAll("#dropdown button.action").forEach((button) => {
      const iconText = button.querySelector("i")?.textContent?.trim();
      if (iconText !== "file_download") {
        return;
      }
      if (button.dataset.caifuclawBatchBound !== "true") {
        button.dataset.caifuclawBatchBound = "true";
        button.addEventListener("click", (event) => {
          if (!selectedRows().length) {
            return;
          }
          event.preventDefault();
          event.stopPropagation();
          event.stopImmediatePropagation();
          downloadSelectedRows();
        }, true);
      }
      button.classList.toggle("caifuclaw-filebrowser-batch-ready", checkedRows.length > 0);
      button.setAttribute("title", checkedRows.length ? `下载已选择 ${checkedRows.length} 项` : "下载");
      button.setAttribute("aria-label", checkedRows.length ? `下载已选择 ${checkedRows.length} 项` : "下载");
      const label = Array.from(button.querySelectorAll("span")).find((item) => !item.classList.contains("counter"));
      if (label) {
        setText(label, "下载");
      }
      let counter = button.querySelector(".caifuclaw-filebrowser-selected-count");
      if (checkedRows.length) {
        if (!counter) {
          counter = document.createElement("span");
          counter.className = "counter caifuclaw-filebrowser-selected-count";
          button.appendChild(counter);
        }
        setText(counter, String(checkedRows.length));
      } else if (counter) {
        counter.remove();
      }
    });
  };

  const setRowChecked = (row, checked, update = true) => {
    const path = rowResourcePath(row);
    if (checked) {
      selectedPaths.add(path);
    } else {
      selectedPaths.delete(path);
    }
    if (update) {
      updateSelectionUi();
    }
  };

  const setAllRowsChecked = (checked) => {
    listedRows().forEach((row) => setRowChecked(row, checked, false));
    updateSelectionUi();
  };

  const createCheckboxLabel = (className, inputClassName, labelText) => {
    const label = document.createElement("label");
    label.className = className;
    label.addEventListener("click", (event) => event.stopPropagation());
    label.addEventListener("keydown", (event) => event.stopPropagation());

    const input = document.createElement("input");
    input.type = "checkbox";
    input.className = inputClassName;
    input.addEventListener("click", (event) => event.stopPropagation());

    const text = document.createElement("span");
    text.className = "caifuclaw-filebrowser-sr-only";
    text.textContent = labelText;

    label.appendChild(input);
    label.appendChild(text);
    return { label, input, text };
  };

  const ensureHeaderSelectionCell = (inner) => {
    if (!inner || inner.querySelector(".caifuclaw-filebrowser-select-all-cell")) {
      return;
    }
    const { label, input } = createCheckboxLabel(
      "caifuclaw-filebrowser-select-all-cell",
      "caifuclaw-filebrowser-select-all",
      "全选当前目录",
    );
    input.addEventListener("change", (event) => {
      event.stopPropagation();
      setAllRowsChecked(input.checked);
    });
    inner.insertBefore(label, inner.firstChild);
  };

  const ensureHeaderIconSpacer = (inner) => {
    if (!inner || inner.querySelector(".caifuclaw-filebrowser-icon-spacer")) {
      return;
    }
    const spacer = document.createElement("span");
    spacer.className = "caifuclaw-filebrowser-icon-spacer";
    spacer.setAttribute("aria-hidden", "true");
    const anchor = inner.querySelector(".caifuclaw-filebrowser-select-all-cell")?.nextSibling || inner.firstChild;
    inner.insertBefore(spacer, anchor);
  };

  const ensureRowPreviewCell = (row) => {
    const name = rowName(row);
    const cell = row.querySelector(":scope > div:first-of-type");
    if (!cell) {
      return;
    }
    rememberOriginalPreviewCell(cell);
    if (!shouldUseImagePreviewCells()) {
      if (cell.dataset.caifuclawPreviewType) {
        restoreOriginalPreviewCell(cell);
      }
      return;
    }
    const previewType = isImageRow(row) ? "image" : (row.dataset.dir === "true" ? "folder" : "file");
    const previewSrc = previewType === "image" ? downloadUrlFor(row) : "";
    if (
      cell.dataset.caifuclawPreviewName === name
      && cell.dataset.caifuclawPreviewType === previewType
      && cell.dataset.caifuclawPreviewSrc === previewSrc
    ) {
      return;
    }

    cell.dataset.caifuclawPreviewName = name;
    cell.dataset.caifuclawPreviewType = previewType;
    cell.dataset.caifuclawPreviewSrc = previewSrc;

    if (previewType !== "image") {
      cell.className = "caifuclaw-filebrowser-preview-cell";
      cell.classList.add(previewType === "folder" ? "caifuclaw-filebrowser-preview-cell--folder" : "caifuclaw-filebrowser-preview-cell--file");
      cell.setAttribute("aria-hidden", "true");
      cell.setAttribute("title", row.dataset.dir === "true" ? `文件夹 ${name}` : `文件 ${name}`);
      const frame = document.createElement("span");
      frame.className = "caifuclaw-filebrowser-preview-frame";
      const icon = document.createElement("i");
      icon.className = "material-icons";
      icon.textContent = rowPreviewIcon(row);
      frame.appendChild(icon);
      cell.replaceChildren(frame);
      return;
    }

    cell.className = "caifuclaw-filebrowser-preview-cell";
    cell.classList.add("caifuclaw-filebrowser-preview-cell--image");
    cell.removeAttribute("aria-hidden");
    cell.setAttribute("title", `预览 ${name}`);

    const button = document.createElement("button");
    button.type = "button";
    button.className = "caifuclaw-filebrowser-preview-button";
    button.title = `预览 ${name}`;
    button.setAttribute("aria-label", `预览图片 ${name}`.trim());
    button.addEventListener("click", (event) => {
      event.preventDefault();
      event.stopPropagation();
      openImagePreviewOverlay(previewSrc, name);
    });
    button.addEventListener("keydown", (event) => event.stopPropagation());

    const frame = document.createElement("span");
    frame.className = "caifuclaw-filebrowser-preview-frame";

    const icon = document.createElement("i");
    icon.className = "material-icons";
    icon.textContent = "image";
    frame.appendChild(icon);

    const image = document.createElement("img");
    image.src = previewSrc;
    image.alt = "";
    image.loading = "lazy";
    image.decoding = "async";
    image.addEventListener("error", () => {
      image.hidden = true;
    });
    frame.appendChild(image);

    button.appendChild(frame);
    cell.replaceChildren(button);
  };

  const ensureRowSelectionCell = (row) => {
    const name = rowName(row);
    const existing = row.querySelector(".caifuclaw-filebrowser-select-cell");
    if (existing) {
      const text = existing.querySelector(".caifuclaw-filebrowser-sr-only");
      const nextLabel = `选择 ${name}`.trim();
      setText(text, nextLabel);
      existing.setAttribute("title", nextLabel);
      return;
    }
    const { label, input } = createCheckboxLabel(
      "caifuclaw-filebrowser-select-cell",
      "caifuclaw-filebrowser-select",
      `选择 ${name}`.trim(),
    );
    label.setAttribute("title", `选择 ${name}`.trim());
    input.addEventListener("change", (event) => {
      event.stopPropagation();
      setRowChecked(row, input.checked);
    });
    row.insertBefore(label, row.firstChild);
  };

  const isRowActionExcludedTarget = (target) => (
    target instanceof Element
    && Boolean(target.closest(
      ".caifuclaw-filebrowser-download-cell, .caifuclaw-filebrowser-select-cell, .action, input, button, select, textarea",
    ))
  );

  const runRowPrimaryAction = (row) => {
    if (row.dataset.dir === "true") {
      window.location.href = directoryUrlFor(row);
      return;
    }
    triggerRowDownload(row);
  };

  const translateHeader = (header) => {
    if (!header.classList.contains("caifuclaw-filebrowser-header")) {
      header.classList.add("caifuclaw-filebrowser-header");
    }
    for (const [selector, label, sortLabel, field] of TITLE_MAP) {
      const cell = header.querySelector(selector);
      if (!cell) {
        continue;
      }
      ensureHeaderLabel(cell, label);
      bindHeaderSortCell(cell, field);
      updateHeaderSortCell(cell, field, label, sortLabel);
    }
    const inner = header.querySelector(":scope > div");
    ensureHeaderSelectionCell(inner);
    ensureHeaderIconSpacer(inner);
    if (inner && !inner.querySelector(".caifuclaw-filebrowser-download-title")) {
      const title = document.createElement("p");
      title.className = "caifuclaw-filebrowser-download-title";
      title.textContent = "下载";
      inner.appendChild(title);
    }
  };

  const translateSectionTitles = () => {
    document.querySelectorAll("#listing h2").forEach((heading) => {
      const text = heading.textContent?.trim();
      if (text === "Folders") {
        heading.textContent = "文件夹";
      } else if (text === "Files") {
        heading.textContent = "文件";
      }
    });
  };

  const enhanceRows = () => {
    document.querySelectorAll("#listing.list .item:not(.header)").forEach((row) => {
      if (!row.classList.contains("caifuclaw-filebrowser-row")) {
        row.classList.add("caifuclaw-filebrowser-row");
      }
      row.setAttribute("tabindex", "0");
      hideNativeNameIcons(row);
      ensureRowSelectionCell(row);
      ensureRowPreviewCell(row);
      updateExactModifiedTime(row);
      const existing = row.querySelector(".caifuclaw-filebrowser-download-cell");
      const url = downloadUrlFor(row);
      const name = rowName(row);
      const title = row.dataset.dir === "true" ? "下载压缩包" : "下载文件";
      const label = `${title} ${name}`.trim();
      if (existing) {
        if (existing.href !== url) {
          existing.href = url;
        }
        if (existing.getAttribute("title") !== title) {
          existing.setAttribute("title", title);
        }
        if (existing.getAttribute("aria-label") !== label) {
          existing.setAttribute("aria-label", label);
        }
        return;
      }
      const link = document.createElement("a");
      link.className = "caifuclaw-filebrowser-download-cell";
      link.href = url;
      link.title = title;
      link.setAttribute("aria-label", label);
      link.addEventListener("click", (event) => {
        event.preventDefault();
        event.stopPropagation();
        triggerDownload(link.href, fallbackDownloadFilename(row));
      });
      link.addEventListener("keydown", (event) => event.stopPropagation());

      const icon = document.createElement("i");
      icon.className = "material-icons";
      icon.textContent = "file_download";
      link.appendChild(icon);
      row.appendChild(link);
    });
  };

  const bindRowActionInteractions = () => {
    listedRows().forEach((row) => {
      if (row.dataset.caifuclawRowActionBound === "true") {
        return;
      }
      row.dataset.caifuclawRowActionBound = "true";
      row.addEventListener("click", (event) => {
        if (isRowActionExcludedTarget(event.target)) {
          return;
        }
        event.preventDefault();
        event.stopPropagation();
        event.stopImmediatePropagation();
        runRowPrimaryAction(row);
      }, true);
      row.addEventListener("keydown", (event) => {
        if (![" ", "Enter"].includes(event.key) || isRowActionExcludedTarget(event.target)) {
          return;
        }
        event.preventDefault();
        event.stopPropagation();
        if (event.key === " ") {
          setRowChecked(row, !selectedPaths.has(rowResourcePath(row)));
          return;
        }
        runRowPrimaryAction(row);
      }, true);
    });
  };

  const applyCleanUi = () => {
    if (document.documentElement.lang !== "zh-CN") {
      document.documentElement.lang = "zh-CN";
    }
    if (recoverServerReachabilityError()) {
      return;
    }
    clearServerReachabilityRecovery();
    ensureHeaderBrand();
    markToolbarDownloadActions();
    forceListView();
    loadDirectoryMetadata();
    document.querySelectorAll("#listing.list .item.header").forEach(translateHeader);
    translateSectionTitles();
    enhanceRows();
    bindRowActionInteractions();
    sortRowsByActiveSort();
    updateSelectionUi();
  };

  let scheduled = false;
  const scheduleApply = () => {
    if (scheduled) {
      return;
    }
    scheduled = true;
    window.requestAnimationFrame(() => {
      scheduled = false;
      applyCleanUi();
    });
  };

  document.addEventListener("DOMContentLoaded", scheduleApply);
  window.addEventListener("popstate", scheduleApply);
  new MutationObserver(scheduleApply).observe(document.documentElement, {
    childList: true,
    subtree: true,
  });
  scheduleApply();
})();
</script>
""".strip()
FILE_BROWSER_DOWNLOAD_AUTO_ZIP_PATCHES = (
    (
        "setup(e){let t=X(),{t:n}=B(),r={zip:`zip`,tar:`tar`,targz:`tar.gz`,tarbz2:`tar.bz2`,tarxz:`tar.xz`,tarlz4:`tar.lz4`,tarsz:`tar.sz`,tarbr:`tar.br`,tarzst:`tar.zst`};return",
        "setup(e){let t=X(),{t:n}=B(),r={zip:`zip`,tar:`tar`,targz:`tar.gz`,tarbz2:`tar.bz2`,tarxz:`tar.xz`,tarlz4:`tar.lz4`,tarsz:`tar.sz`,tarbr:`tar.br`,tarzst:`tar.zst`};y(()=>setTimeout(()=>t.currentPrompt?.confirm(`zip`),0));return",
    ),
    (
        'setup(e){const t=Ws(),{t:n}=re(),i={zip:"zip",tar:"tar",targz:"tar.gz",tarbz2:"tar.bz2",tarxz:"tar.xz",tarlz4:"tar.lz4",tarsz:"tar.sz",tarbr:"tar.br",tarzst:"tar.zst"};return',
        'setup(e){const t=Ws(),{t:n}=re(),i={zip:"zip",tar:"tar",targz:"tar.gz",tarbz2:"tar.bz2",tarxz:"tar.xz",tarlz4:"tar.lz4",tarsz:"tar.sz",tarbr:"tar.br",tarzst:"tar.zst"};w(()=>setTimeout(()=>t.currentPrompt?.confirm("zip"),0));return',
    ),
)
FILE_BROWSER_LIST_VIEW_PATCHES = (
    (
        "class:H([`file-icons`,k(m).user?.viewMode??``])",
        "class:H([`file-icons`,`list`])",
    ),
    (
        'class:de(["file-icons",M(f).user?.viewMode??""])',
        'class:de(["file-icons","list"])',
    ),
    (
        'viewMode:{list:`mosaic`,mosaic:`mosaic gallery`,"mosaic gallery":`list`}[m.user?.viewMode??`list`]||`list`',
        "viewMode:`list`",
    ),
    (
        'viewMode:{list:"mosaic",mosaic:"mosaic gallery","mosaic gallery":"list"}[f.user?.viewMode??"list"]||"list"',
        'viewMode:"list"',
    ),
)
_filebrowser_cached_auth_token = ""
_filebrowser_cached_auth_token_expires_at = datetime.min.replace(tzinfo=timezone.utc)


def _filebrowser_default_files_url() -> str:
    today_directory = _local_now().strftime("%y%m%d")
    if (FILE_BROWSER_UPLOAD_EXCEL_ROOT / today_directory).is_dir():
        return f"{FILE_BROWSER_UPLOAD_EXCEL_URL}{today_directory}/"
    return FILE_BROWSER_UPLOAD_EXCEL_URL


@app.post("/api/v1/filebrowser/session")
def create_filebrowser_session(response: Response, user: LocalUser = Depends(current_user)) -> dict[str, str]:
    session_token = create_filebrowser_session_token(user.username, FILE_BROWSER_SESSION_SECONDS)
    response.set_cookie(
        FILE_BROWSER_COOKIE_NAME,
        session_token,
        max_age=FILE_BROWSER_SESSION_SECONDS,
        httponly=True,
        samesite="lax",
        path="/",
    )
    return {"url": _filebrowser_default_files_url()}


def _filebrowser_user_from_cookie(request: Request, db: Session) -> LocalUser:
    session_token = request.cookies.get(FILE_BROWSER_COOKIE_NAME, "").strip()
    if not session_token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing File Browser session")
    try:
        username = decode_filebrowser_session_token(session_token)
    except JWTError as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid File Browser session") from exc
    return _current_user_by_username(username, request, db)


def _filebrowser_proxy_target(path: str, query_string: bytes) -> str:
    normalized_path = path.lstrip("/")
    target = urljoin(FILE_BROWSER_PROXY_BASE, normalized_path)
    if query_string:
        target = f"{target}?{query_string.decode('latin-1')}"
    return target


def _filebrowser_public_location(path: str) -> str:
    normalized_path = path.lstrip("/")
    if not normalized_path:
        return "/"
    return f"/{normalized_path}"


def _filebrowser_public_api_proxy_path(path: str) -> str:
    normalized_path = path.lstrip("/")
    prefix = normalized_path.split("/", 1)[0]
    if prefix not in FILE_BROWSER_PUBLIC_API_PREFIXES:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not Found")
    return f"api/{normalized_path}"


def _filebrowser_legacy_redirect_url(path: str, query_string: bytes) -> str | None:
    normalized_path = path.lstrip("/")
    if not normalized_path:
        target = FILE_BROWSER_UPLOAD_EXCEL_URL
    elif normalized_path.startswith(("files/", "static/", "api/")):
        target = _filebrowser_public_location(normalized_path)
    else:
        return None
    if query_string:
        target = f"{target}?{query_string.decode('latin-1')}"
    return target


def _filebrowser_download_only_redirect_url(path: str, method: str) -> str | None:
    if method.upper() not in {"GET", "HEAD"}:
        return None
    normalized_path = path.lstrip("/")
    if not normalized_path.startswith("files/"):
        return None
    file_path = normalized_path.removeprefix("files/").rstrip("/")
    if not file_path or Path(file_path).suffix.lower() not in FILE_BROWSER_DOWNLOAD_ONLY_EXTENSIONS:
        return None
    return f"/api/raw/{quote(file_path, safe='/')}"


def _filebrowser_token_expires_at(token: str) -> datetime:
    try:
        payload_part = token.split(".", 2)[1]
        payload_part += "=" * (-len(payload_part) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_part))
        exp = int(payload.get("exp", 0))
    except (IndexError, TypeError, ValueError, json.JSONDecodeError):
        return datetime.now(timezone.utc) + timedelta(minutes=30)
    return datetime.fromtimestamp(exp - FILE_BROWSER_AUTH_REFRESH_MARGIN_SECONDS, tz=timezone.utc)


async def _get_filebrowser_auth_token(client: httpx.AsyncClient, force_refresh: bool = False) -> str:
    global _filebrowser_cached_auth_token, _filebrowser_cached_auth_token_expires_at

    now = datetime.now(timezone.utc)
    if not force_refresh and _filebrowser_cached_auth_token and _filebrowser_cached_auth_token_expires_at > now:
        return _filebrowser_cached_auth_token

    try:
        response = await client.get(
            urljoin(FILE_BROWSER_PROXY_BASE, "api/login"),
            headers={"Host": "127.0.0.1:8088"},
        )
        response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail="File Browser authorization failed") from exc

    token = response.text.strip()
    if not token:
        raise HTTPException(status_code=502, detail="File Browser authorization token is empty")
    _filebrowser_cached_auth_token = token
    _filebrowser_cached_auth_token_expires_at = _filebrowser_token_expires_at(token)
    return token


def _filebrowser_request_headers(request: Request, auth_token: str) -> dict[str, str]:
    headers: dict[str, str] = {}
    for name, value in request.headers.items():
        lowered = name.lower()
        if lowered in FILE_BROWSER_FORWARD_BLOCKED_HEADERS or lowered in {"host", "cookie"}:
            continue
        if lowered == "x-auth":
            continue
        headers[name] = value
    headers["X-Auth"] = auth_token
    headers["Host"] = "127.0.0.1:8088"
    return headers


def _patch_filebrowser_public_html(html: str) -> str:
    patched = re.sub(r'("BaseURL"\s*:\s*)"/filebrowser"', r'\1""', html)
    patched = re.sub(r'("StaticURL"\s*:\s*)"/filebrowser/static"', r'\1"/static"', patched)
    patched = patched.replace('"/filebrowser/static/', '"/static/')
    patched = patched.replace("'/filebrowser/static/", "'/static/")
    patched = patched.replace('href="/filebrowser/static/', 'href="/static/')
    patched = patched.replace('src="/filebrowser/static/', 'src="/static/')
    patched = patched.replace('data-src="/filebrowser/static/', 'data-src="/static/')
    return patched


def _inject_filebrowser_clean_ui(content: bytes, content_type: str) -> bytes:
    if "text/html" not in content_type.lower():
        return content
    html = content.decode("utf-8", errors="replace")
    if "</head>" not in html:
        return content
    original_html = html
    html = _patch_filebrowser_public_html(html)
    custom_head = []
    if "caifuclaw-filebrowser-clean-ui" not in html:
        custom_head.append(FILE_BROWSER_CLEAN_UI_STYLE)
    if "caifuclaw-filebrowser-behavior" not in html:
        custom_head.append(FILE_BROWSER_CLEAN_UI_SCRIPT)
    if not custom_head:
        return html.encode("utf-8") if html != original_html else content
    return html.replace("</head>", f"{''.join(custom_head)}</head>", 1).encode("utf-8")


def _patch_filebrowser_frontend_script(content: bytes, path: str, content_type: str) -> bytes:
    if "javascript" not in content_type.lower() or not path.startswith("static/assets/index"):
        return content
    script = content.decode("utf-8", errors="replace")
    patched = script
    for original, replacement in FILE_BROWSER_DOWNLOAD_AUTO_ZIP_PATCHES + FILE_BROWSER_LIST_VIEW_PATCHES:
        patched = patched.replace(original, replacement, 1)
    if patched == script:
        return content
    return patched.encode("utf-8")


def _is_filebrowser_frontend_script(path: str, content_type: str) -> bool:
    return "javascript" in content_type.lower() and path.lstrip("/").startswith("static/assets/index")


async def _proxy_filebrowser_request(path: str, request: Request, db: Session) -> Response:
    _filebrowser_user_from_cookie(request, db)
    if request.scope.get("path", "").startswith("/filebrowser"):
        legacy_redirect_url = _filebrowser_legacy_redirect_url(path, request.scope.get("query_string", b""))
        if legacy_redirect_url:
            return Response(status_code=status.HTTP_302_FOUND, headers={"location": legacy_redirect_url})
    download_redirect_url = _filebrowser_download_only_redirect_url(path, request.method)
    if download_redirect_url:
        return Response(status_code=status.HTTP_302_FOUND, headers={"location": download_redirect_url})
    target = _filebrowser_proxy_target(path, request.scope.get("query_string", b""))
    body = await request.body()
    async with httpx.AsyncClient(follow_redirects=False, timeout=60.0) as client:
        request_auth_token = request.headers.get("x-auth")
        auth_token = request_auth_token or await _get_filebrowser_auth_token(client)
        proxied = await client.request(
            request.method,
            target,
            content=body,
            headers=_filebrowser_request_headers(request, auth_token),
        )
        if proxied.status_code == status.HTTP_401_UNAUTHORIZED:
            auth_token = await _get_filebrowser_auth_token(client, force_refresh=True)
            proxied = await client.request(
                request.method,
                target,
                content=body,
                headers=_filebrowser_request_headers(request, auth_token),
            )

    headers = {
        name: value
        for name, value in proxied.headers.items()
        if name.lower() not in FILE_BROWSER_RESPONSE_BLOCKED_HEADERS and name.lower() != "set-cookie"
    }
    location = headers.get("location")
    if location and location.startswith(FILE_BROWSER_PROXY_BASE):
        headers["location"] = _filebrowser_public_location(location.removeprefix(FILE_BROWSER_PROXY_BASE))
    content_type = proxied.headers.get("content-type", "")
    content = _inject_filebrowser_clean_ui(proxied.content, content_type)
    content = _patch_filebrowser_frontend_script(content, path.lstrip("/"), content_type)
    if _is_filebrowser_frontend_script(path, content_type):
        headers["cache-control"] = "no-store"
    return Response(content=content, status_code=proxied.status_code, headers=headers)


@app.api_route(
    "/filebrowser/",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    include_in_schema=False,
)
async def proxy_filebrowser_root(request: Request, db: Session = Depends(get_db)):
    return await _proxy_filebrowser_request("", request, db)


@app.api_route(
    "/filebrowser/{path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    include_in_schema=False,
)
async def proxy_filebrowser(path: str, request: Request, db: Session = Depends(get_db)):
    return await _proxy_filebrowser_request(path, request, db)


@app.api_route(
    "/files/",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    include_in_schema=False,
)
async def proxy_filebrowser_public_files_root(request: Request, db: Session = Depends(get_db)):
    return await _proxy_filebrowser_request("files/", request, db)


@app.api_route(
    "/files/{path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    include_in_schema=False,
)
async def proxy_filebrowser_public_files(path: str, request: Request, db: Session = Depends(get_db)):
    return await _proxy_filebrowser_request(f"files/{path}", request, db)


@app.api_route(
    "/static/{path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    include_in_schema=False,
)
async def proxy_filebrowser_public_static(path: str, request: Request, db: Session = Depends(get_db)):
    return await _proxy_filebrowser_request(f"static/{path}", request, db)


@app.api_route(
    "/api/{path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    include_in_schema=False,
)
async def proxy_filebrowser_public_api(path: str, request: Request, db: Session = Depends(get_db)):
    return await _proxy_filebrowser_request(_filebrowser_public_api_proxy_path(path), request, db)


class SPAStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope):
        try:
            return await super().get_response(path, scope)
        except StarletteHTTPException as exc:
            if exc.status_code == 404:
                return await super().get_response("index.html", scope)
            raise


frontend_dist = Path(__file__).resolve().parents[1] / "frontend" / "dist"
if frontend_dist.exists():
    from starlette.responses import FileResponse

    FRONTEND_NO_CACHE_HEADERS = {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }
    FRONTEND_IMMUTABLE_CACHE_HEADERS = {
        "Cache-Control": "public, max-age=31536000, immutable",
    }

    def _frontend_asset_fallback(path: str) -> Path | None:
        requested = Path(path).name
        match = re.match(r"^(?P<prefix>.+)-(?P<hash>[A-Za-z0-9_-]{6,})\.(?P<ext>js|css|mjs)$", requested)
        if not match:
            return None

        prefix = match.group("prefix")
        ext = match.group("ext")
        candidates = sorted((frontend_dist / "assets").glob(f"{prefix}-*.{ext}"), key=lambda item: item.stat().st_mtime, reverse=True)
        for candidate in candidates:
            if candidate.name != requested and candidate.is_file():
                return candidate
        return None

    class FrontendAssetFiles(StaticFiles):
        async def get_response(self, path: str, scope):
            try:
                response = await super().get_response(path, scope)
            except StarletteHTTPException as exc:
                if exc.status_code != 404:
                    raise
                fallback = _frontend_asset_fallback(path)
                if fallback is None:
                    raise
                response = FileResponse(fallback)
            response.headers.update(FRONTEND_IMMUTABLE_CACHE_HEADERS)
            return response

    def frontend_index_response():
        return FileResponse(frontend_dist / "index.html", media_type="text/html", headers=FRONTEND_NO_CACHE_HEADERS)

    # Vite assets are content-hashed, so they can be cached aggressively.
    app.mount("/assets", FrontendAssetFiles(directory=frontend_dist / "assets"), name="frontend-assets")

    # Serve index.html for SPA routes
    @app.get("/", include_in_schema=False)
    async def serve_index():
        return frontend_index_response()

    @app.get("/version.json", include_in_schema=False)
    async def serve_frontend_version():
        version_file = frontend_dist / "version.json"
        if version_file.is_file():
            return FileResponse(version_file, media_type="application/json", headers=FRONTEND_NO_CACHE_HEADERS)
        return {"version": "dev", "builtAt": None}

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        # Only serve index.html for non-api paths
        file_path = frontend_dist / full_path
        if file_path.is_file():
            return FileResponse(file_path, headers=FRONTEND_NO_CACHE_HEADERS)
        return frontend_index_response()
