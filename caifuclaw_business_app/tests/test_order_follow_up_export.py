from datetime import datetime
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

import app.task_runner as task_runner
import app.order_follow_up_export as order_follow_up_export
from app.models import OrderFollowUpExportJob
from app.order_follow_up_export import (
    DIRECT_PLAN_SHEET_NAME,
    DIRECT_EXPORT_LOG_SHEET_NAME,
    EXPORT_LOG_SHEET_NAME,
    ExistingWorkbookOrder,
    ORDER_HEADERS,
    ORDER_SHEET_NAME,
    _existing_workbook_orders,
    _export_buyer_name,
    _export_logistics_channel,
    _export_order_number,
    _matching_workbook_order,
    _order_item_sku_aliases,
    _workbook_match_applies,
    _write_main_workbook,
    _write_purchase_plan,
)
from app.settings import get_settings


def _template(path):
    workbook = Workbook()
    order_sheet = workbook.active
    order_sheet.title = ORDER_SHEET_NAME
    order_sheet.append(ORDER_HEADERS + ["客户确认", "客户审核", "是否打单", "订单类型", "预警", "Shipping time"])
    order_sheet.append(
        [
            datetime(2026, 7, 18),
            "OZON",
            "Template Shop",
            datetime(2026, 7, 18, 8, 0),
            "TEMPLATE-1",
            "RU",
            "Buyer",
            "SKU-TEMPLATE",
            1,
            10,
            "CNY",
            "",
            "",
            datetime(2026, 7, 20),
            None,
            "TRACK-TEMPLATE",
            "=N2-1",
            "Template Product",
            '=IFERROR(LEFT(E2,FIND("-",E2)-1),"")',
            None,
            None,
            0,
            '=IF(V2>0,"",Q2-NOW())',
            '=IFERROR(VLOOKUP(P2,订单出库!A:B,2,FALSE),0)',
        ]
    )
    workbook.create_sheet("产品目录")
    workbook.create_sheet("订单出库")
    direct_sheet = workbook.create_sheet(DIRECT_PLAN_SHEET_NAME)
    direct_sheet["E3"] = "配货日"
    direct_sheet["F3"] = "产品名称"
    direct_sheet["G3"] = "采购数量（当日来单）"
    direct_sheet["H3"] = "库存数"
    direct_sheet["I3"] = "导出时间"
    direct_sheet["J3"] = "待采购数量"
    direct_sheet["K3"] = "总表成本记录"
    direct_sheet["R3"] = "备注"
    direct_sheet["E4"] = "2026-07-18"
    direct_sheet["F4"] = "Template Product"
    direct_sheet["G4"] = 1
    direct_sheet["H4"] = 0
    direct_sheet["I4"] = "2026-07-18 08:00:00"
    direct_sheet["J4"] = 1
    direct_sheet["E4"].number_format = r"yyyy\-m\-d"
    direct_sheet["I4"].number_format = r"yyyy\-m\-d\ h:mm"
    workbook.save(path)
    workbook.close()


def _snapshot(order_item_id: int, *, mapped: bool) -> dict:
    sku = f"SKU-{order_item_id}"
    return {
        "order_id": order_item_id + 100,
        "order_item_id": order_item_id,
        "allocation_date": "2026-07-19",
        "platform": "ozon",
        "shop_name": "CaifuClaw Shop",
        "created_time": "2026-07-19T08:10:00",
        "order_number": f"ORDER-{order_item_id}",
        "country_code": "RU",
        "buyer_name": "Buyer",
        "sku": sku,
        "quantity": 2,
        "unit_price": 25.5,
        "currency": "CNY",
        "buyer_selected_logistics": "online",
        "logistics_channel": "channel",
        "shipping_deadline": "2026-07-21T17:00:00",
        "tracking_number": f"TRACK-{order_item_id}",
        "dispatch_deadline": "2026-07-20T17:00:00",
        "product_id": order_item_id if mapped else None,
        "product_name": f"产品-{order_item_id}" if mapped else "",
        "display_product_name": f"产品-{order_item_id}" if mapped else sku,
        "mapping_status": "mapped" if mapped else "missing",
        "stock_qty": 1 if mapped else 0,
        "cost": 12.5 if mapped else None,
        "buyer": "Tony" if mapped else "",
        "order_type": 0,
        "direct_plan_eligible": True,
        "source_status": "配货中",
        "source_run_id": 160,
    }


def _settings(tmp_path):
    return get_settings().model_copy(
        update={
            "order_follow_up_export_data_root": str(tmp_path),
            "order_follow_up_export_workbook_name": "Order follow up 2026_caifuclaw.xlsx",
            "order_follow_up_export_template_workbook_name": "Order follow up 2026.xlsx",
            "order_follow_up_export_sync_dir": "result_data_sync",
            "order_follow_up_export_backup_dir": "result_data_backup",
            "order_follow_up_export_purchase_plan_dir": "pur_plan",
            "order_follow_up_export_recalculate_engine": "none",
        }
    )


def test_main_workbook_exports_missing_mapping_and_is_idempotent(tmp_path):
    sync_dir = tmp_path / "result_data_sync"
    sync_dir.mkdir()
    _template(sync_dir / "Order follow up 2026.xlsx")
    settings = _settings(tmp_path)
    items = [
        {"order_item_id": 1, "action": "append", "worksheet_row": None, "snapshot": _snapshot(1, mapped=True)},
        {"order_item_id": 2, "action": "append", "worksheet_row": None, "snapshot": _snapshot(2, mapped=False)},
    ]

    first = _write_main_workbook(101, items, settings)
    workbook = load_workbook(first.file_path, data_only=False)
    try:
        order_sheet = workbook[ORDER_SHEET_NAME]
        assert order_sheet["R3"].value == "产品-1"
        assert order_sheet["R4"].value == "未登记目录"
        assert order_sheet["S3"].value == '=IFERROR(LEFT(E3,FIND("-",E3)-1),"")'
        direct_sheet = workbook[DIRECT_PLAN_SHEET_NAME]
        direct_rows = {
            direct_sheet.cell(row_number, 6).value: direct_sheet.cell(row_number, 18).value
            for row_number in range(5, 7)
        }
        assert direct_rows["产品-1"] in {None, ""}
        assert direct_rows["SKU-2"] == "未登记目录，仅提示，不可自动采购"
        for row_number in range(5, 7):
            allocation_date = direct_sheet.cell(row_number, 5)
            exported_at = direct_sheet.cell(row_number, 9)
            assert allocation_date.value == "2026-07-19"
            assert allocation_date.data_type == "s"
            assert isinstance(exported_at.value, str)
            assert exported_at.data_type == "s"
            datetime.strptime(exported_at.value, "%Y-%m-%d %H:%M:%S")
        assert workbook[EXPORT_LOG_SHEET_NAME].sheet_state == "veryHidden"
    finally:
        workbook.close()

    second = _write_main_workbook(101, items, settings)
    assert second.changed_item_count == 0
    workbook = load_workbook(second.file_path, read_only=True, data_only=False)
    try:
        assert workbook[ORDER_SHEET_NAME].max_row == 4
        assert workbook[DIRECT_PLAN_SHEET_NAME].max_row == 6
    finally:
        workbook.close()


def test_purchase_plan_includes_missing_mapping(tmp_path):
    settings = _settings(tmp_path)
    job = OrderFollowUpExportJob(
        id=102,
        scheduled_task_run_id=160,
        created_at=datetime(2026, 7, 19, 4, 30),
    )
    items = [
        {"order_item_id": 2, "action": "append", "worksheet_row": None, "snapshot": _snapshot(2, mapped=False)}
    ]

    result = _write_purchase_plan(job, items, settings)
    assert result.file_path is not None
    workbook = load_workbook(result.file_path, read_only=True, data_only=False)
    try:
        sheet = workbook[DIRECT_PLAN_SHEET_NAME]
        assert sheet["B2"].value == "SKU-2"
        assert sheet["E2"].value == 2
        assert sheet["G2"].value == "未登记目录"
    finally:
        workbook.close()


def test_update_refreshes_existing_direct_row_without_adding_quantity(tmp_path):
    sync_dir = tmp_path / "result_data_sync"
    sync_dir.mkdir()
    _template(sync_dir / "Order follow up 2026.xlsx")
    settings = _settings(tmp_path)
    initial = _snapshot(3, mapped=False)
    _write_main_workbook(
        103,
        [{"order_item_id": 3, "action": "append", "worksheet_row": None, "snapshot": initial}],
        settings,
    )

    updated = _snapshot(3, mapped=True)
    result = _write_main_workbook(
        104,
        [{"order_item_id": 3, "action": "update", "worksheet_row": 3, "snapshot": updated}],
        settings,
    )

    assert result.direct_row_count == 0
    workbook = load_workbook(result.file_path, data_only=False)
    try:
        direct_sheet = workbook[DIRECT_PLAN_SHEET_NAME]
        assert direct_sheet.max_row == 5
        assert direct_sheet["F5"].value == "产品-3"
        assert direct_sheet["G5"].value == 2
        assert direct_sheet["H5"].value == 1
        assert direct_sheet["J5"].value == 1
        assert direct_sheet["R5"].value in {None, ""}
        direct_log = workbook[DIRECT_EXPORT_LOG_SHEET_NAME]
        assert direct_log.sheet_state == "veryHidden"
        assert [direct_log.cell(3, column).value for column in range(2, 8)] == [
            104,
            3,
            "update",
            5,
            "2026-07-19|product:3",
            0,
        ]
    finally:
        workbook.close()


def test_purchase_plan_excludes_update_items(tmp_path):
    settings = _settings(tmp_path)
    job = OrderFollowUpExportJob(
        id=105,
        scheduled_task_run_id=160,
        created_at=datetime(2026, 7, 19, 4, 30),
    )
    items = [
        {"order_item_id": 4, "action": "update", "worksheet_row": 3, "snapshot": _snapshot(4, mapped=True)},
        {"order_item_id": 5, "action": "append", "worksheet_row": None, "snapshot": _snapshot(5, mapped=True)},
        {
            "order_item_id": 6,
            "action": "append",
            "worksheet_row": None,
            "snapshot": {**_snapshot(6, mapped=True), "direct_plan_eligible": False},
        },
    ]

    result = _write_purchase_plan(job, items, settings)
    assert result.row_count == 1
    workbook = load_workbook(result.file_path, read_only=True, data_only=False)
    try:
        sheet = workbook[DIRECT_PLAN_SHEET_NAME]
        assert sheet.max_row == 2
        assert sheet["B2"].value == "产品-5"
        assert workbook["_caifuclaw_export_meta"]["B3"].value == "5"
    finally:
        workbook.close()


def test_reliable_platform_fields_use_current_raw_order_data(tmp_path):
    settings = _settings(tmp_path)
    allegro = SimpleNamespace(
        platform="allegro",
        platform_order_no="DEMO-ORDER-0039",
        posting_number="DEMO-ORDER-0039",
        platform_order_id="DEMO-ORDER-0039",
        shipment_tracking_number="DEMO-TRACKING-0008",
        buyer_name="aneta_p78",
        logistics_channel="",
        raw_payload={
            "delivery": {
                "address": {"firstName": "Aneta", "lastName": "Przybyl"},
            }
        },
    )
    item = SimpleNamespace(
        sku="DEMO-SKU-0008",
        raw_payload={
            "raw_payload": {
                "offer": {"external": {"id": "POP_Straykids_KRAMA_Compact"}},
            }
        },
    )

    assert _export_order_number(allegro) == "DEMOORDER0039"
    assert _export_buyer_name(allegro) == "Aneta Przybyl"
    assert _order_item_sku_aliases(item, allegro.platform) == [
        "POP_Straykids_KRAMA_Compact",
            "DEMO-SKU-0008",
    ]

    ozon = SimpleNamespace(
        platform="ozon",
        platform_order_no="DEMO-ORDER-0040",
        posting_number="DEMO-ORDER-0041",
        platform_order_id="DEMO-ORDER-0021",
        shipment_tracking_number="",
        logistics_channel="",
    )
    assert _export_order_number(ozon) == "DEMO-ORDER-0041"
    assert _export_logistics_channel(ozon, settings) == "ozon线上发货"


def test_existing_workbook_baseline_matches_order_alias_and_sku_alias(tmp_path):
    path = tmp_path / "baseline.xlsx"
    _template(path)
    existing_orders = _existing_workbook_orders(path)

    match = _matching_workbook_order(
        {
            "platform": "ozon",
            "order_number": "TEMPLATE-1-1",
            "order_number_aliases": ["TEMPLATE-1"],
            "tracking_number": "",
            "sku": "DEMO-SKU-0009",
            "source_sku": "DEMO-SKU-0010",
            "sku_aliases": ["SKU-NEW", "SKU-TEMPLATE"],
        },
        existing_orders,
    )

    assert match is not None
    assert match.row_number == 2
    assert match.mapping_missing is False


def test_overseas_orders_only_dedupe_against_caifuclaw_managed_rows():
    snapshot = {"is_overseas_warehouse": True}

    assert not _workbook_match_applies(
        snapshot,
        ExistingWorkbookOrder(row_number=10, mapping_missing=False, managed_by_caifuclaw=False),
    )
    assert _workbook_match_applies(
        snapshot,
        ExistingWorkbookOrder(row_number=20, mapping_missing=False, managed_by_caifuclaw=True),
    )


def test_confirmed_allegro_bsi_draft_is_shipped_after_follow_up_export(monkeypatch):
    allegro_bsi = SimpleNamespace(
        id=1,
        platform="allegro",
        platform_order_no="DEMO-ORDER-0042",
        posting_number="DEMO-ORDER-0042",
        platform_order_id="DEMO-ORDER-0042",
        bsi_order_no="DEMO-ORDER-0043",
        biz_status="待处理",
        local_status="new",
        shipped_at=None,
        marked_shipped_at=None,
        error_message="",
        updated_at=None,
    )
    regular_order = SimpleNamespace(
        id=2,
        platform="allegro",
        platform_order_no="DEMO-ORDER-0044",
        posting_number="DEMO-ORDER-0044",
        platform_order_id="DEMO-ORDER-0044",
        bsi_order_no="",
        biz_status="待处理",
        local_status="new",
        shipped_at=None,
        marked_shipped_at=None,
        error_message="",
        updated_at=None,
    )

    class FakeDb:
        def scalars(self, _statement):
            return SimpleNamespace(all=lambda: [allegro_bsi, regular_order])

    logs = []
    monkeypatch.setattr(
        order_follow_up_export,
        "confirmed_order_ids_for_follow_up_export",
        lambda _db, _order_ids, **_kwargs: {1, 2},
    )
    monkeypatch.setattr(order_follow_up_export, "add_order_operation_logs", lambda *_args, **kwargs: logs.append(kwargs))

    rows = order_follow_up_export.mark_bsi_orders_shipped_after_follow_up_export(
        FakeDb(),
        [1, 2],
        export_job_id=21,
        operated_at=datetime(2026, 7, 30, 0, 23),
    )

    assert rows == [allegro_bsi]
    assert allegro_bsi.biz_status == "已发货"
    assert allegro_bsi.local_status == "shipped"
    assert allegro_bsi.shipped_at == datetime(2026, 7, 30, 0, 23)
    assert allegro_bsi.marked_shipped_at == datetime(2026, 7, 30, 0, 23)
    assert regular_order.biz_status == "待处理"
    assert regular_order.local_status == "new"
    assert logs[0]["operation_type"] == "bsi_follow_up_export_shipped"
    assert "不获取平台面单、不打印、不生成采购单" in logs[0]["description"](allegro_bsi)


def test_export_job_confirmation_uses_exported_order_items_after_item_ids_change():
    class FakeDb:
        def execute(self, _statement):
            return SimpleNamespace(
                all=lambda: [
                    (1, "success", 100),
                    (1, "skipped", 101),
                    (2, "success", None),
                ]
            )

    confirmed = order_follow_up_export.confirmed_order_ids_for_follow_up_export(
        FakeDb(),
        [1, 2, 3],
        export_job_id=21,
    )

    assert confirmed == {1}


def test_enqueue_failure_does_not_escape_pipeline_hook(monkeypatch):
    def raise_error(_run_id):
        raise RuntimeError("database unavailable")

    monkeypatch.setattr(task_runner, "enqueue_order_follow_up_export", raise_error)
    task_runner._enqueue_order_follow_up_export_after_success(SimpleNamespace(id=160), "success")


def test_pipeline_hook_only_enqueues_success(monkeypatch):
    calls = []
    monkeypatch.setattr(task_runner, "enqueue_order_follow_up_export", calls.append)

    task_runner._enqueue_order_follow_up_export_after_success(SimpleNamespace(id=160), "failed")

    assert calls == []
