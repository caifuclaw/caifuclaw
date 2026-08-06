# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

import asyncio
import io
from types import SimpleNamespace

import openpyxl
from fastapi import UploadFile

from app.main import (
    _apply_inventory_payload,
    _build_inventory_import_template_workbook,
    _inventory_count_query,
    _inventory_query,
    _read_product_import_rows,
    import_inventory,
)
from app.schemas import InventoryUpsertRequest


def test_inventory_payload_updates_safety_stock_to_zero():
    product = SimpleNamespace(id=1, internal_name="测试产品", safety_stock=8)
    inventory = SimpleNamespace()
    payload = InventoryUpsertRequest(product_id=1, stock_qty=12, last_count_qty=10, safety_stock=0, remark="复盘")

    _apply_inventory_payload(inventory, product, payload, "tester")

    assert product.safety_stock == 0
    assert inventory.stock_qty == 12


def test_inventory_hide_zero_filter_excludes_empty_safety_stock():
    list_sql = str(_inventory_query(hide_zero_safety_stock=True).compile(compile_kwargs={"literal_binds": True}))
    count_sql = str(_inventory_count_query(hide_zero_safety_stock=True).compile(compile_kwargs={"literal_binds": True}))

    for sql in (list_sql, count_sql):
        where_clause = sql.split("WHERE", maxsplit=1)[1]
        assert "products.safety_stock IS NOT NULL AND products.safety_stock != 0" in where_clause
        assert "product_inventory.stock_qty" not in where_clause


def test_inventory_import_template_contains_supported_fields_and_guide():
    workbook = _build_inventory_import_template_workbook()
    worksheet = workbook["库存"]

    assert worksheet.max_row == 1
    assert worksheet.freeze_panes == "A2"
    assert [cell.value for cell in worksheet[1]] == ["产品名称", "库存数量", "上次盘点", "安全库存", "备注"]
    assert worksheet["A1"].comment is not None
    assert worksheet["B1"].comment is not None
    assert worksheet["D1"].comment is not None
    assert worksheet["A1"].fill.fgColor.rgb == worksheet["D1"].fill.fgColor.rgb
    assert worksheet["A1"].fill.fgColor.rgb != worksheet["B1"].fill.fgColor.rgb
    assert "填写说明" in workbook.sheetnames
    validation_ranges = {str(validation.sqref) for validation in worksheet.data_validations.dataValidation}
    assert validation_ranges == {"B2:C1048576", "D2:D1048576"}
    assert workbook["填写说明"][3][1].value == "否"
    assert workbook["填写说明"][5][1].value == "是"


def test_inventory_import_template_can_be_read_by_inventory_import_parser():
    workbook = _build_inventory_import_template_workbook()
    worksheet = workbook["库存"]
    worksheet.append(["测试产品", 12, 10, 5, "首次盘点"])
    output = io.BytesIO()
    workbook.save(output)

    headers, rows, first_data_row = _read_product_import_rows("inventory.xlsx", output.getvalue())

    assert first_data_row == 2
    assert headers == ["产品名称", "库存数量", "上次盘点", "安全库存", "备注"]
    assert rows == [["测试产品", 12, 10, 5, "首次盘点"]]


def test_inventory_import_updates_product_safety_stock():
    workbook = _build_inventory_import_template_workbook()
    workbook["库存"].append(["测试产品", 12, 10, 8, "首次盘点"])
    workbook["库存"].append(["测试产品", None, 12, 9, "复盘"])
    workbook["库存"].append(["测试产品", 20, 18, None, "缺少安全库存"])
    workbook["库存"].append(["测试产品", 20, 18, -1, "无效安全库存"])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    product = SimpleNamespace(id=1, internal_name="测试产品", safety_stock=3)
    inventory = SimpleNamespace(product_id=1, product_name="测试产品", stock_qty=0, last_count_qty=0, remark="")

    class ScalarRows:
        def all(self):
            return [product]

    class FakeDb:
        def scalars(self, _statement):
            return ScalarRows()

        def scalar(self, _statement):
            return inventory

        def commit(self):
            return None

    result = asyncio.run(
        import_inventory(
            UploadFile(output, filename="inventory.xlsx"),
            SimpleNamespace(username="tester"),
            FakeDb(),
        )
    )

    assert result["created"] == 0
    assert result["updated"] == 2
    assert result["failed"] == 2
    assert result["errors"] == [
        {"row": 4, "message": "安全库存不能为空"},
        {"row": 5, "message": "安全库存不能小于 0"},
    ]
    assert product.safety_stock == 9
    assert inventory.stock_qty == 12
    assert inventory.last_count_qty == 12

    safety_only_workbook = openpyxl.Workbook()
    safety_only_workbook.active.append(["产品名称", "安全库存"])
    safety_only_workbook.active.append(["测试产品", 11])
    safety_only_output = io.BytesIO()
    safety_only_workbook.save(safety_only_output)
    safety_only_output.seek(0)

    safety_only_result = asyncio.run(
        import_inventory(
            UploadFile(safety_only_output, filename="inventory.xlsx"),
            SimpleNamespace(username="tester"),
            FakeDb(),
        )
    )

    assert safety_only_result == {"created": 0, "updated": 1, "failed": 0, "errors": []}
    assert product.safety_stock == 11
    assert inventory.stock_qty == 12
