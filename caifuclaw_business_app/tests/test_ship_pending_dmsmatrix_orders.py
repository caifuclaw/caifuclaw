# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

from datetime import datetime
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from scripts.ship_pending_dmsmatrix_orders import (
    ORDER_SHEET,
    OUTBOUND_SHEET,
    Candidate,
    _match_excel_rows,
    _normalise_platform,
    update_workbook,
)


def _candidate(*, order_id: int = 10, item_id: int = 20) -> Candidate:
    order = SimpleNamespace(
        id=order_id,
        platform_order_no="DEMO-ORDER-0126",
        posting_number="DEMO-ORDER-0127",
        platform_order_id="DEMO-ORDER-0127",
        shipment_tracking_number="",
    )
    item = SimpleNamespace(
        id=item_id,
        sku="DEMO-SKU-0024",
        raw_payload={"offer_id": "DEMO-OFFER-0002"},
    )
    return Candidate(order=order, items=(item,))


def _workbook(path, *, shipping_time=None, duplicate=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = ORDER_SHEET
    sheet.append(["平台", "订单编号", "SKU", "Shipping time"])
    sheet.append(["dmsmatrix", "DEMO-ORDER-0126", "DEMO-SKU-0024", shipping_time])
    if duplicate:
        sheet.append(["Fruugo", "DEMO-ORDER-0127", "DEMO-OFFER-0002", None])
    outbound = workbook.create_sheet(OUTBOUND_SHEET)
    outbound.append(["物流单号", "发货日期"])
    workbook.save(path)
    workbook.close()


def test_platform_aliases_are_limited_to_dmsmatrix():
    assert _normalise_platform("DMSMatrix") == "dmsmatrix"
    assert _normalise_platform("Fruugo-DMS") == "dmsmatrix"
    assert _normalise_platform("ozon") == "ozon"


def test_excel_matching_requires_exactly_one_row(tmp_path):
    path = tmp_path / "orders.xlsx"
    _workbook(path, duplicate=True)
    workbook = load_workbook(path, data_only=False)
    try:
        try:
            _match_excel_rows(workbook[ORDER_SHEET], [_candidate()])
        except RuntimeError as exc:
            assert "matched 2 Excel rows" in str(exc)
        else:
            raise AssertionError("ambiguous Excel rows should fail")
    finally:
        workbook.close()


def test_dry_run_reports_missing_outbound_registration_without_writing(tmp_path):
    path = tmp_path / "orders.xlsx"
    backup_dir = tmp_path / "backups"
    _workbook(path)
    shipping_time = datetime(2026, 7, 29, 13, 30)

    result = update_workbook(
        path,
        [_candidate()],
        shipping_time=shipping_time,
        apply=False,
        backup_dir=backup_dir,
    )

    assert result.changed_count == 1
    assert result.matches[0].order_sheet_row == 2
    assert result.matches[0].outbound_row == 0
    assert result.backup_path is None
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        assert workbook[ORDER_SHEET]["D2"].value is None
    finally:
        workbook.close()


def test_apply_registers_outbound_shipping_time_without_overwriting(tmp_path):
    path = tmp_path / "orders.xlsx"
    backup_dir = tmp_path / "backups"
    _workbook(path)
    shipping_time = datetime(2026, 7, 29, 13, 30)

    result = update_workbook(
        path,
        [_candidate()],
        shipping_time=shipping_time,
        apply=True,
        backup_dir=backup_dir,
    )

    assert result.changed_count == 1
    assert result.backup_path is not None and result.backup_path.is_file()
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        assert workbook[ORDER_SHEET]["D2"].value is None
        assert workbook[OUTBOUND_SHEET]["A2"].value == "DEMO-ORDER-0127"
        assert workbook[OUTBOUND_SHEET]["B2"].value == shipping_time
    finally:
        workbook.close()

    second = update_workbook(
        path,
        [_candidate()],
        shipping_time=datetime(2026, 7, 30, 8, 0),
        apply=True,
        backup_dir=backup_dir,
    )
    assert second.changed_count == 0
    assert second.backup_path is None
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        assert workbook[ORDER_SHEET]["D2"].value is None
        assert workbook[OUTBOUND_SHEET]["B2"].value == shipping_time
    finally:
        workbook.close()
