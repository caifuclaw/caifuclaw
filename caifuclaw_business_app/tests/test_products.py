import io
from types import SimpleNamespace
from unittest.mock import Mock

import openpyxl

from app.main import (
    _append_product_export_header,
    _build_product_import_template_workbook,
    _product_dto,
    _product_filter_conditions,
    _product_name_match_key,
    _read_product_import_rows,
    delete_product,
)


def _product_with_mapping():
    return SimpleNamespace(
        id=1,
        product_code="DEMO-PRODUCT-0014",
        internal_name="测试产品",
        english_name="Test Product",
        cost=None,
        weight=None,
        gross_weight=None,
        package_length=None,
        package_width=None,
        package_height=None,
        ean="DEMO-PRODUCT-0015",
        description="描述",
        main_image_url="https://example.test/image.jpg",
        is_slow_moving_material=True,
        safety_stock=None,
        buyer_user_id=None,
        buyer_user=None,
        enabled=True,
        mappings=[
            SimpleNamespace(id=2, shop_id=1, shop_sku="SKU-1"),
            SimpleNamespace(id=3, shop_id=1, shop_sku=""),
        ],
        created_at=None,
        updated_at=None,
    )


def test_product_dto_keeps_mappings_when_options_are_omitted():
    dto = _product_dto(_product_with_mapping(), [])

    assert dto.mappings == {"1": ["SKU-1"]}
    assert dto.english_name == "Test Product"
    assert dto.ean == "DEMO-PRODUCT-0015"
    assert dto.description == "描述"
    assert dto.main_image_url == "https://example.test/image.jpg"
    assert dto.is_slow_moving_material is True


def test_product_dto_limits_mappings_to_visible_shops_when_provided():
    shops = [
        SimpleNamespace(id=1),
        SimpleNamespace(id=2),
    ]

    dto = _product_dto(_product_with_mapping(), shops)

    assert dto.mappings == {"1": ["SKU-1"], "2": []}


def test_product_keyword_filter_searches_five_product_fields_with_or_matching():
    conditions = _product_filter_conditions(keyword="  Test-123  ")

    assert len(conditions) == 1
    compiled = str(conditions[0].compile(compile_kwargs={"literal_binds": True}))
    assert "products.product_code" in compiled
    assert "products.internal_name" in compiled
    assert "products.english_name" in compiled
    assert "products.ean" in compiled
    assert "product_shop_mappings.shop_sku" in compiled
    assert compiled.count(" OR ") == 4
    assert "%Test-123%" in compiled


def test_product_keyword_filter_combines_with_independent_filters():
    conditions = _product_filter_conditions(keyword="product", shop_sku="DEMO-PRODUCT-0016", enabled=True)

    assert len(conditions) == 3


def test_product_import_reads_grouped_excel_headers():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "产品目录"
    worksheet.append(["编码", "产品中文名", "产品英文名", "SKU CODE", "", "标准参数", "", "描述"])
    worksheet.append(["", "", "", "Ozon店铺", "WB店铺", "成本", "净重", ""])
    worksheet.append(["P00000001", "测试产品", "Test Product", "SKU-OZON", "SKU-WB", 12.5, 0.25, "描述"])
    output = io.BytesIO()
    workbook.save(output)

    headers, rows, first_data_row = _read_product_import_rows("products.xlsx", output.getvalue())

    assert first_data_row == 3
    assert headers == ["编码", "产品中文名", "产品英文名", "Ozon店铺", "WB店铺", "成本", "净重", "描述"]
    assert rows[0][1] == "测试产品"


def test_product_import_name_match_key_ignores_case_and_outer_spaces():
    assert _product_name_match_key("  Test 产品  ") == _product_name_match_key("test 产品")


def test_delete_product_allows_regular_product():
    product = SimpleNamespace()
    db = Mock()
    db.get.return_value = product

    response = delete_product(1, _=None, db=db)

    assert response == {"message": "已删除"}
    db.delete.assert_called_once_with(product)
    db.commit.assert_called_once_with()


def test_product_import_template_is_formatted_without_sample_data(monkeypatch):
    shops = [
        SimpleNamespace(id=1, display_name="Ozon店铺", account_id="ozon"),
        SimpleNamespace(id=2, display_name="WB店铺", account_id="wb"),
    ]
    buyers = [SimpleNamespace(username="buyer01", display_name="采购员A")]
    monkeypatch.setattr("app.main._enabled_buyer_users", lambda db: buyers)

    workbook = _build_product_import_template_workbook(shops, None)
    worksheet = workbook["产品目录"]

    assert worksheet.max_row == 2
    assert worksheet.freeze_panes == "A3"
    assert worksheet["A1"].comment is not None
    assert "填写说明" in workbook.sheetnames
    assert "可选品类" not in workbook.sheetnames
    assert "可选采购人" in workbook.sheetnames
    header_values = [cell.value for row in worksheet.iter_rows(min_row=1, max_row=2) for cell in row]
    assert "编码" not in header_values
    validation_ranges = {str(validation.sqref) for validation in worksheet.data_validations.dataValidation}
    assert "O3:O1048576" in validation_ranges

    output = io.BytesIO()
    workbook.save(output)
    headers, rows, first_data_row = _read_product_import_rows("template.xlsx", output.getvalue())

    assert first_data_row == 3
    assert rows == []
    assert headers[:5] == ["产品中文名", "产品英文名", "Ozon店铺", "WB店铺", "成本"]


def test_product_export_writes_grouped_headers_with_shop_columns_before_standard_fields():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    shops = [
        SimpleNamespace(display_name="Ozon店铺", account_id="ozon"),
        SimpleNamespace(display_name="WB店铺", account_id="wb"),
    ]

    _append_product_export_header(worksheet, shops)

    assert [cell.value for cell in worksheet[1][:12]] == [
        "编码",
        "产品中文名",
        "产品英文名",
        "SKU CODE",
        None,
        "标准参数",
        None,
        None,
        None,
        None,
        None,
        None,
    ]
    assert [cell.value for cell in worksheet[2][:12]] == [
        None,
        None,
        None,
        "Ozon店铺",
        "WB店铺",
        "成本",
        "净重",
        "毛重",
        "包装长",
        "包装宽",
        "包装高",
        "EAN",
    ]
