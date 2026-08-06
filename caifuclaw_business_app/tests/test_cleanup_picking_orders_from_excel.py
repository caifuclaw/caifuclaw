# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, select
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import Order, OrderOperationLog
from scripts.cleanup_picking_orders_from_excel import (
    GREEN_FILL,
    RED_FILL,
    process_excel_cleanup,
    read_excel_targets,
)


@compiles(JSONB, "sqlite")
def _compile_jsonb_for_sqlite(_type, _compiler, **_kwargs):
    return "JSON"


def _session_factory():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine, tables=[Order.__table__, OrderOperationLog.__table__])
    return sessionmaker(bind=engine, autoflush=False, autocommit=False)


def _write_workbook(path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "订单列表"
    worksheet.append(["订单编号", "平台", "店铺", "交易号", "交运单号", "货运单号"])
    worksheet.append(["GREEN-1", "Ozon", "SHOP", "T-1", "POST-1", "TRACK-1"])
    worksheet.append(["RED-1", "Ozon", "SHOP", "T-2", "POST-2", "TRACK-2"])
    worksheet.append(["KEEP-1", "Ozon", "SHOP", "T-3", "POST-3", "TRACK-3"])
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor=GREEN_FILL)
    worksheet["A3"].fill = PatternFill(fill_type="solid", fgColor=RED_FILL)
    workbook.save(path)


def _order(order_id, order_no, posting):
    return Order(
        id=order_id,
        tenant_id="default",
        platform="ozon",
        account_id="A",
        shop_id="SHOP-ID",
        shop_name="SHOP",
        platform_order_id=order_no,
        platform_order_no=order_no,
        posting_number=posting,
        biz_status="配货中",
        local_status="picking",
        raw_payload={},
        last_api_payload={},
    )


def test_read_excel_targets_detects_colors_and_normalizes_platform(tmp_path):
    path = tmp_path / "orders.xlsx"
    _write_workbook(path)

    targets, counts, data_rows = read_excel_targets(path)

    assert data_rows == 3
    assert counts[GREEN_FILL] == 1
    assert counts[RED_FILL] == 1
    assert counts["NONE"] == 1
    assert [target.status_after for target in targets] == ["已发货", "已作废"]
    assert targets[0].platform == "ozon"
    assert targets[0].posting_number == "POST-1"


def test_cleanup_dry_run_does_not_change_orders_or_logs(tmp_path):
    path = tmp_path / "orders.xlsx"
    _write_workbook(path)
    session_factory = _session_factory()
    with session_factory() as db:
        db.add_all([_order(1, "GREEN-1", "POST-1"), _order(2, "RED-1", "POST-2")])
        db.commit()

        result = process_excel_cleanup(db, path, apply=False)

        assert result.target_count == 2
        assert result.pending_count == 2
        assert db.get(Order, 1).biz_status == "配货中"
        assert db.scalar(select(OrderOperationLog.id)) is None


def test_cleanup_apply_updates_statuses_and_records_source_audit(tmp_path):
    path = tmp_path / "orders.xlsx"
    _write_workbook(path)
    session_factory = _session_factory()
    with session_factory() as db:
        db.add_all([_order(1, "GREEN-1", "POST-1"), _order(2, "RED-1", "POST-2")])
        db.commit()

        result = process_excel_cleanup(db, path, apply=True)
        db.commit()

        green = db.get(Order, 1)
        red = db.get(Order, 2)
        assert result.updated_order_ids == [1, 2]
        assert green.biz_status == "已发货"
        assert green.local_status == "shipped"
        assert isinstance(green.shipped_at, datetime)
        assert isinstance(green.marked_shipped_at, datetime)
        assert red.biz_status == "已作废"
        assert red.local_status == "cancelled"

        logs = db.scalars(select(OrderOperationLog).order_by(OrderOperationLog.order_id)).all()
        assert len(logs) == 2
        assert logs[0].operation_type == "excel_picking_status_cleanup"
        assert logs[0].source == "system"
        assert logs[0].extra["source_sheet"] == "订单列表"
        assert logs[0].extra["source_row"] == 2
        assert logs[0].extra["fill_color"] == GREEN_FILL
        assert logs[0].extra["changes"][0]["before"] == "配货中"
        assert logs[1].extra["source_row"] == 3


def test_cleanup_is_idempotent_for_already_correct_orders(tmp_path):
    path = tmp_path / "orders.xlsx"
    _write_workbook(path)
    session_factory = _session_factory()
    with session_factory() as db:
        shipped = _order(1, "GREEN-1", "POST-1")
        shipped.biz_status = "已发货"
        shipped.local_status = "shipped"
        voided = _order(2, "RED-1", "POST-2")
        voided.biz_status = "已作废"
        voided.local_status = "cancelled"
        db.add_all([shipped, voided])
        db.commit()

        result = process_excel_cleanup(db, path, apply=True)

        assert result.pending_count == 0
        assert result.already_target_count == 2
        assert result.updated_order_ids == []
        assert db.scalar(select(OrderOperationLog.id)) is None
