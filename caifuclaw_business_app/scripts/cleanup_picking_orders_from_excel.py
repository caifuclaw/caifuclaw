"""Correct picking-order business statuses from a color-coded Excel export.

The script is dry-run by default. Use ``--apply`` only after reviewing the
matching summary. Every changed order receives an idempotent operation log
containing the source workbook hash, sheet/row, color, match key, and status
changes.
"""
from __future__ import annotations

import argparse
import hashlib
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from sqlalchemy import and_, select
from sqlalchemy.orm import Session

REPO_ROOT = Path(__file__).resolve().parents[2]
BUSINESS_ROOT = REPO_ROOT / "caifuclaw_business_app"
for _path in (REPO_ROOT, BUSINESS_ROOT):
    if str(_path) not in sys.path:
        sys.path.insert(0, str(_path))

from app.database import SessionLocal  # noqa: E402
from app.models import Order, OrderOperationLog  # noqa: E402
from app.order_operation_logs import (  # noqa: E402
    ORDER_LOG_SYSTEM_SOURCE,
    SYSTEM_OPERATOR,
    add_order_operation_log,
)


SOURCE_SHEET = "订单列表"
GREEN_FILL = "92D050"
RED_FILL = "FF0000"
TARGET_FILLS = frozenset({GREEN_FILL, RED_FILL})
ORDER_STATUS_PICKING = "配货中"
ORDER_STATUS_SHIPPED = "已发货"
ORDER_STATUS_VOIDED = "已作废"
LOCAL_STATUS_SHIPPED = "shipped"
LOCAL_STATUS_CANCELLED = "cancelled"
OPERATION_TYPE = "excel_picking_status_cleanup"
OPERATION_ATTRIBUTE = "Excel配货状态清理"

PLATFORM_MAP = {
    "joom": "joom_logistics",
    "joom logistics": "joom_logistics",
    "ozon": "ozon",
    "mercadolibre": "mercadolibre",
    "mercado libre": "mercadolibre",
    "wildberries": "wildberries",
    "allegro": "allegro",
}

REQUIRED_HEADERS = (
    "订单编号",
    "平台",
    "店铺",
    "交运单号",
)


@dataclass(frozen=True)
class ExcelStatusTarget:
    source_row: int
    fill_color: str
    platform: str
    shop_name: str
    order_no: str
    posting_number: str
    tracking_number: str
    status_after: str
    local_status_after: str

    @property
    def match_key(self) -> tuple[str, str, str, str]:
        return (self.platform, self.shop_name, self.order_no, self.posting_number)


@dataclass
class CleanupResult:
    source_file: str
    source_sha256: str
    total_data_rows: int
    color_counts: dict[str, int]
    targets: list[ExcelStatusTarget]
    matched_count: int = 0
    pending_count: int = 0
    already_target_count: int = 0
    updated_order_ids: list[int] | None = None
    skipped_order_ids: list[int] | None = None

    def __post_init__(self) -> None:
        if self.updated_order_ids is None:
            self.updated_order_ids = []
        if self.skipped_order_ids is None:
            self.skipped_order_ids = []

    @property
    def target_count(self) -> int:
        return len(self.targets)


def normalize_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_posting_number(value: object) -> str:
    value = normalize_text(value)
    return "" if value in {"", "-", "—", "–"} else value


def normalize_platform(value: object) -> str:
    value = normalize_text(value).lower()
    return PLATFORM_MAP.get(value, value)


def cell_fill_color(cell) -> str:
    """Return the supported solid fill color, or ``NONE`` for other fills."""
    fill = cell.fill
    if fill.fill_type != "solid":
        return "NONE"
    rgb = normalize_text(getattr(fill.fgColor, "rgb", ""))
    if rgb:
        rgb = rgb[-6:].upper()
        if rgb in TARGET_FILLS:
            return rgb
    return "NONE"


def source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_excel_targets(path: Path, *, sheet_name: str = SOURCE_SHEET) -> tuple[list[ExcelStatusTarget], Counter[str], int]:
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            raise ValueError(f"Worksheet is empty: {sheet_name}")
        headers = [normalize_text(cell.value) for cell in header_row]
        indexes = {header: index for index, header in enumerate(headers) if header}
        missing = [header for header in REQUIRED_HEADERS if header not in indexes]
        if missing:
            raise ValueError(f"Missing required Excel headers: {', '.join(missing)}")

        targets: list[ExcelStatusTarget] = []
        color_counts: Counter[str] = Counter()
        data_rows = 0
        for source_row, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            data_rows += 1
            color = cell_fill_color(row[0]) if row else "NONE"
            color_counts[color] += 1
            if color not in TARGET_FILLS:
                continue

            def value(header: str) -> object:
                index = indexes[header]
                return row[index].value if index < len(row) else None

            platform = normalize_platform(value("平台"))
            shop_name = normalize_text(value("店铺"))
            order_no = normalize_text(value("订单编号"))
            posting_number = normalize_posting_number(value("交运单号"))
            if not platform or not shop_name or not order_no:
                raise ValueError(
                    f"Invalid target row {source_row}: platform/shop/order number must not be empty"
                )

            is_shipped = color == GREEN_FILL
            targets.append(
                ExcelStatusTarget(
                    source_row=source_row,
                    fill_color=color,
                    platform=platform,
                    shop_name=shop_name,
                    order_no=order_no,
                    posting_number=posting_number,
                    tracking_number=normalize_text(value("货运单号")) if "货运单号" in indexes else "",
                    status_after=ORDER_STATUS_SHIPPED if is_shipped else ORDER_STATUS_VOIDED,
                    local_status_after=LOCAL_STATUS_SHIPPED if is_shipped else LOCAL_STATUS_CANCELLED,
                )
            )
        return targets, color_counts, data_rows
    finally:
        workbook.close()


def _find_order(db: Session, target: ExcelStatusTarget) -> Order:
    statement = select(Order).where(
        and_(
            Order.platform == target.platform,
            Order.shop_name == target.shop_name,
            Order.platform_order_no == target.order_no,
            Order.posting_number == target.posting_number,
        )
    )
    rows = db.scalars(statement).all()
    if len(rows) != 1:
        key = " / ".join(target.match_key)
        if not rows:
            raise ValueError(f"Excel row {target.source_row} did not match an order: {key}")
        raise ValueError(
            f"Excel row {target.source_row} matched {len(rows)} orders; expected one: {key}"
        )
    return rows[0]


def _order_display_number(order: Order) -> str:
    return order.platform_order_no or order.posting_number or order.platform_order_id or str(order.id)


def _log_event_key(source_hash: str, target: ExcelStatusTarget, order: Order) -> str:
    return f"excel:{source_hash[:24]}:{target.source_row}:{order.id}"


def process_excel_cleanup(
    db: Session,
    source_path: str | Path,
    *,
    apply: bool = False,
    sheet_name: str = SOURCE_SHEET,
) -> CleanupResult:
    path = Path(source_path).expanduser().resolve()
    targets, color_counts, data_rows = read_excel_targets(path, sheet_name=sheet_name)
    result = CleanupResult(
        source_file=str(path),
        source_sha256=source_sha256(path),
        total_data_rows=data_rows,
        color_counts=dict(color_counts),
        targets=targets,
    )

    matched_ids: set[int] = set()
    resolved: list[tuple[ExcelStatusTarget, Order]] = []
    for target in targets:
        order = _find_order(db, target)
        if order.id in matched_ids:
            raise ValueError(f"Multiple Excel target rows point to order id {order.id}")
        matched_ids.add(order.id)
        if order.biz_status not in {ORDER_STATUS_PICKING, target.status_after}:
            raise ValueError(
                f"Order {_order_display_number(order)} has status {order.biz_status!r}; "
                f"expected {ORDER_STATUS_PICKING!r} or {target.status_after!r}"
            )
        resolved.append((target, order))

    result.matched_count = len(resolved)
    pending = [(target, order) for target, order in resolved if order.biz_status == ORDER_STATUS_PICKING]
    result.pending_count = len(pending)
    result.already_target_count = len(resolved) - len(pending)
    result.skipped_order_ids.extend(order.id for _target, order in resolved if order.biz_status != ORDER_STATUS_PICKING)
    if not apply:
        return result

    now = datetime.utcnow()
    for target, order in pending:
        before_biz_status = order.biz_status or ""
        before_local_status = order.local_status or ""
        before_shipped_at = order.shipped_at
        before_marked_shipped_at = order.marked_shipped_at
        order.biz_status = target.status_after
        order.local_status = target.local_status_after
        if target.status_after == ORDER_STATUS_SHIPPED:
            order.shipped_at = order.shipped_at or now
            order.marked_shipped_at = order.marked_shipped_at or now
        order.updated_at = now

        changes = [
            {
                "field": "biz_status",
                "label": "业务状态",
                "before": before_biz_status or "-",
                "after": target.status_after,
            },
            {
                "field": "local_status",
                "label": "本地状态",
                "before": before_local_status or "-",
                "after": target.local_status_after,
            },
        ]
        if target.status_after == ORDER_STATUS_SHIPPED:
            changes.extend(
                [
                    {
                        "field": "shipped_at",
                        "label": "实际发货时间",
                        "before": before_shipped_at.isoformat() if before_shipped_at else "-",
                        "after": order.shipped_at.isoformat() if order.shipped_at else now.isoformat(),
                    },
                    {
                        "field": "marked_shipped_at",
                        "label": "标记发货时间",
                        "before": before_marked_shipped_at.isoformat() if before_marked_shipped_at else "-",
                        "after": order.marked_shipped_at.isoformat()
                        if order.marked_shipped_at
                        else now.isoformat(),
                    },
                ]
            )

        action_text = "绿色标记为已发货" if target.fill_color == GREEN_FILL else "红色标记为已取消"
        add_order_operation_log(
            db,
            order_id=order.id,
            operation_type=OPERATION_TYPE,
            operation_attribute=OPERATION_ATTRIBUTE,
            description=(
                f"Excel清理：订单 {_order_display_number(order)} {action_text}，"
                f"状态：{before_biz_status or '-'} -> {target.status_after}"
            ),
            operator=SYSTEM_OPERATOR,
            source=ORDER_LOG_SYSTEM_SOURCE,
            operated_at=now,
            event_key=_log_event_key(result.source_sha256, target, order),
            extra={
                "result": "success",
                "source_file": result.source_file,
                "source_sha256": result.source_sha256,
                "source_sheet": sheet_name,
                "source_row": target.source_row,
                "fill_color": target.fill_color,
                "match_key": {
                    "platform": target.platform,
                    "shop_name": target.shop_name,
                    "order_no": target.order_no,
                    "posting_number": target.posting_number,
                },
                "tracking_number_from_excel": target.tracking_number,
                "status_before": before_biz_status,
                "status_after": target.status_after,
                "changes": changes,
            },
        )
        result.updated_order_ids.append(order.id)
    return result


def print_result(result: CleanupResult, *, apply: bool, max_print: int) -> None:
    mode = "apply" if apply else "dry-run"
    print(
        f"mode={mode} source={result.source_file} rows={result.total_data_rows} "
        f"targets={result.target_count} matched={result.matched_count} "
        f"pending={result.pending_count} already_target={result.already_target_count}"
    )
    print(
        f"green={result.color_counts.get(GREEN_FILL, 0)} "
        f"red={result.color_counts.get(RED_FILL, 0)} "
        f"no_fill={result.color_counts.get('NONE', 0)}"
    )
    if apply:
        print(f"updated_order_ids={len(result.updated_order_ids or [])}")
    if result.skipped_order_ids:
        print(f"skipped_already_target_ids={len(result.skipped_order_ids)}")
    for order_id in (result.updated_order_ids or [])[:max(0, max_print)]:
        print(f"updated_order_id={order_id}")
    if len(result.updated_order_ids or []) > max_print:
        print(f"... {len(result.updated_order_ids or []) - max_print} more updated orders")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Clean picking-order statuses from a color-coded Excel export.")
    parser.add_argument("source", type=Path, help="Color-coded order export workbook")
    parser.add_argument("--apply", action="store_true", help="Commit status and log changes; defaults to dry-run")
    parser.add_argument("--sheet", default=SOURCE_SHEET, help=f"Worksheet name (default: {SOURCE_SHEET})")
    parser.add_argument("--max-print", type=int, default=20, help="Maximum updated order ids to print")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    with SessionLocal() as db:
        try:
            result = process_excel_cleanup(db, args.source, apply=args.apply, sheet_name=args.sheet)
            print_result(result, apply=args.apply, max_print=args.max_print)
            if args.apply:
                db.commit()
            else:
                db.rollback()
            return 0
        except Exception:
            db.rollback()
            raise


if __name__ == "__main__":
    raise SystemExit(main())
