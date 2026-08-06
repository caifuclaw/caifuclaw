"""Import 2025 H1 historical orders from the per-shop workbook tabs.

This script is deliberately independent from the regular synchronizer and the
other historical import scripts. It is insert-only: existing orders are
reported and skipped. A dry-run is the default; ``--apply`` inserts data and
then performs a best-effort platform status/tracking reconciliation for only
the rows created by that batch.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse
from uuid import uuid4

import httpx
import openpyxl
from openpyxl.utils.datetime import time_to_days, to_excel
from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError


REPO_ROOT = Path(__file__).resolve().parents[2]
BUSINESS_ROOT = REPO_ROOT / "caifuclaw_business_app"
for _path in (REPO_ROOT, BUSINESS_ROOT):
    if str(_path) not in sys.path:
        sys.path.insert(0, str(_path))

from app.country_mapping import country_name_cn, country_name_to_code  # noqa: E402
from app.credential_manager import get_credential_manager  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.models import Order, OrderItem, OrderOperationLog, PlatformAccount, Shipment  # noqa: E402
from app.order_operation_logs import ORDER_LOG_HISTORY_SOURCE, SYSTEM_OPERATOR  # noqa: E402
from app.order_types import infer_is_overseas_warehouse, normalize_fulfillment_type  # noqa: E402
from app.settings import get_settings  # noqa: E402
from app.sync_engine import _ensure_base_url  # noqa: E402
from connector_runtime.app.factory import connector_for  # noqa: E402


DEFAULT_SOURCE = Path("./demo_data/Order follow up 2025.xlsx")
DEFAULT_REPORT_DIR = BUSINESS_ROOT / "outputs" / "historical_store_tab_import"
START_AT = datetime(2025, 1, 1)
END_AT = datetime(2025, 7, 1)
SOURCE_TAG = "historical_store_tab_import_2025_h1"
DEFAULT_BIZ_STATUS = "已妥投"
DEFAULT_PLATFORM_STATUS = "delivered"
DEFAULT_LOCAL_STATUS = "shipped"

INVALID_VALUES = {"", "0", "#n/a", "#na", "n/a", "na", "none", "null", "-", "--"}
VOIDED_MARKERS = {
    "cancel",
    "canceled",
    "cancelled",
    "cancelled_by_seller",
    "cancelled_by_customer",
    "cancelled_by_merchant",
    "refund",
    "refunded",
    "paidbyjoomrefund",
    "voided",
}
VOIDED_TEXT_RE = re.compile(r"取消|作废|退款|\bcancel(?:ed|led)?\b|\brefund(?:ed)?\b|\bvoided\b", re.I)


@dataclass(frozen=True)
class SheetConfig:
    sheet: str
    platform: str
    account_id: str
    shop_name: str
    date_header: str
    order_header: str
    sku_header: str
    amount_header: str | None = None
    amount_mode: str = "first_nonzero"
    quantity_header: str | None = None
    tracking_header: str | None = None
    product_headers: tuple[str, ...] = ()
    logistics_headers: tuple[str, ...] = ()
    country_header: str | None = None
    buyer_header: str | None = None
    deadline_header: str | None = None
    shipped_header: str | None = None
    parent_header: str | None = None
    status_headers: tuple[str, ...] = ()
    fulfillment_type: str = "FBS"
    currency: str = ""
    default_country_code: str = ""
    attach_blank_order_by_tracking: bool = False


SHEET_CONFIGS: tuple[SheetConfig, ...] = (
    SheetConfig(
        sheet="FBP订单",
        platform="ozon",
        account_id="100001",
        shop_name="OZON DEMO SHOP A",
        date_header="Processing",
        order_header="Shipment number",
        sku_header="Shipment details",
        amount_header="Order total ¥",
        tracking_header=None,
        product_headers=("中文名称",),
        logistics_headers=("Delivery agent and method",),
        parent_header="客户确认",
        status_headers=("备注",),
        fulfillment_type="FBP",
        currency="CNY",
        default_country_code="RU",
    ),
    SheetConfig(
        sheet="OZON-ECOMANGO",
        platform="ozon",
        account_id="100001",
        shop_name="OZON DEMO SHOP A",
        date_header="Created Date",
        order_header="Order number",
        sku_header="Shipment details",
        amount_header="Order total",
        amount_mode="sum_lines",
        quantity_header="Product quantity",
        tracking_header="Tracking number",
        product_headers=("中文名称",),
        logistics_headers=("Delivery agent and method",),
        deadline_header="Deadline",
        shipped_header="Shipping time",
        parent_header="客户确认",
        status_headers=("预警", "status", "备注"),
        currency="CNY",
        default_country_code="RU",
    ),
    SheetConfig(
        sheet="OZON-SUPREME",
        platform="ozon",
        account_id="100002",
        shop_name="Ozon Demo Shop B",
        date_header="Created Date",
        order_header="Order number",
        sku_header="Shipment details",
        amount_header="Order total",
        amount_mode="sum_lines",
        quantity_header="数量",
        tracking_header="Tracking number",
        product_headers=("Product name",),
        logistics_headers=("Delivery agent and method",),
        deadline_header="Deadline",
        shipped_header="Shipping time",
        parent_header="客户确认",
        status_headers=("预警", "status", "备注"),
        currency="USD",
        default_country_code="RU",
    ),
    SheetConfig(
        sheet="JOOM-Joom Demo Shop",
        platform="joom_logistics",
        account_id="JOOM-DEMO-001",
        shop_name="Joom Demo Shop",
        date_header="Created Date",
        order_header="Order Number",
        sku_header="Shipment Details",
        amount_header="Order Total(USD)",
        quantity_header="Quantity",
        tracking_header="Tracking Number",
        product_headers=("Product Name",),
        logistics_headers=("Shipping Provider",),
        country_header="Country",
        shipped_header="Shipping Time",
        status_headers=("预警", "status"),
        currency="USD",
    ),
    SheetConfig(
        sheet="JOOM-DEMO-SHOP",
        platform="joom_logistics",
        account_id="J001",
        shop_name="Joom Demo Shop",
        date_header="Created Date",
        order_header="Order Number",
        sku_header="Shipment Details",
        amount_header="Order Total(USD)",
        quantity_header="Quantity",
        tracking_header="Tracking Number",
        product_headers=("Product Name",),
        logistics_headers=("Shipping Provider",),
        country_header="Country",
        shipped_header="Shipping Time",
        status_headers=("预警", "status"),
        currency="USD",
    ),
    SheetConfig(
        sheet="JOOM-FBJ",
        platform="joom_logistics",
        account_id="JOOM-DEMO-001",
        shop_name="Joom Demo Shop",
        date_header="Created Date",
        order_header="Order Number",
        sku_header="Shipment Details",
        amount_header="Order Total(USD)",
        quantity_header="Quantity",
        tracking_header="Tracking Number",
        product_headers=("Product Name",),
        logistics_headers=("Shipping Provider",),
        country_header="Country",
        currency="USD",
        fulfillment_type="FBJ",
    ),
    SheetConfig(
        sheet="MERCADO",
        platform="mercadolibre",
        account_id="mercado-demo",
        shop_name="Mercado Demo Shop",
        date_header="Order date",
        order_header="Order number",
        sku_header="SKU",
        amount_header="应收款金额",
        quantity_header="Units",
        tracking_header="Tracking number",
        product_headers=("产品名称", "产品标题"),
        country_header="站点",
        buyer_header="Buyer",
        shipped_header="发货时间",
        status_headers=("order status", "预警", "备注"),
        currency="USD",
    ),
    SheetConfig(
        sheet="Wildberries",
        platform="wildberries",
        account_id="wildberries-demo",
        shop_name="WB DEMO SHOP CN",
        date_header="付款时间",
        order_header="订单号",
        sku_header="SKU",
        tracking_header="货运单号",
        product_headers=("Product Name", "平台商品标题"),
        logistics_headers=("物流渠道",),
        shipped_header="Shipping Time",
        status_headers=("预警", "备注"),
        currency="CNY",
        default_country_code="RU",
    ),
    SheetConfig(
        sheet="WB海外仓",
        platform="wildberries",
        account_id="wildberries-demo",
        shop_name="WB DEMO SHOP CN",
        date_header="付款时间",
        order_header="订单号",
        sku_header="SKU",
        tracking_header="货运单号",
        product_headers=("Product Name", "平台商品标题"),
        logistics_headers=("物流渠道",),
        status_headers=("备注",),
        currency="CNY",
        default_country_code="RU",
        fulfillment_type="OVERSEAS_WAREHOUSE",
    ),
    SheetConfig(
        sheet="Allegro",
        platform="allegro",
        account_id="allegro0002",
        shop_name="Demo Shop",
        date_header="Order date",
        order_header="Order number",
        sku_header="SKU",
        amount_header="应收款金额",
        quantity_header="Units",
        tracking_header="Tracking number",
        product_headers=("产品名称", "产品标题"),
        logistics_headers=("物流商",),
        country_header="Country",
        buyer_header="Buyer",
        shipped_header="发货时间",
        status_headers=("预警", "备注"),
        currency="PLN",
        attach_blank_order_by_tracking=True,
    ),
    SheetConfig(
        sheet="Fruugo",
        platform="dmsmatrix",
        account_id="dms0001",
        shop_name="Fruugo-DMS",
        date_header="下单时间",
        order_header="订单号",
        sku_header="SKU",
        amount_header="收款USD",
        tracking_header="WB 追踪码",
        product_headers=("产品名称", "Product Name"),
        country_header="国家",
        buyer_header="Customer Name",
        shipped_header="发货时间",
        status_headers=("Status", "预警", "备注"),
        currency="USD",
    ),
)


@dataclass
class SourceRow:
    sheet: str
    row_no: int
    platform: str
    account_id: str
    shop_name: str
    order_no: str
    parent_order_no: str
    sku: str
    quantity: int
    amount: Decimal | None
    product_name: str
    tracking_number: str
    logistics_channel: str
    country_code: str
    buyer_name: str
    platform_created_at: datetime
    deadline_at: datetime | None
    shipped_at: datetime | None
    status_text: str
    raw_values: dict[str, Any]


@dataclass
class OrderGroup:
    config: SheetConfig
    order_no: str
    rows: list[SourceRow] = field(default_factory=list)

    @property
    def key(self) -> tuple[str, str, str]:
        return self.config.platform, self.config.account_id, normalize_identifier(self.order_no, self.config.platform)


@dataclass
class LiveUpdate:
    platform_status: str = ""
    platform_order_id: str = ""
    platform_order_no: str = ""
    posting_number: str = ""
    tracking_number: str = ""
    handover_at: datetime | None = None
    raw_payload: dict[str, Any] = field(default_factory=dict)


@dataclass
class ReconcileTarget:
    id: int
    platform: str
    account_id: str
    source_order_no: str
    platform_order_id: str
    platform_order_no: str
    posting_number: str
    platform_created_at: datetime | None

    @property
    def lookup(self) -> str:
        return self.source_order_no or self.posting_number or self.platform_order_id or self.platform_order_no


@dataclass
class ExistingMatch:
    order_id: int | None = None
    account_id: str = ""
    reason: str = ""


def utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None, microsecond=0)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_valid_value(value: Any) -> bool:
    return clean(value).lower() not in INVALID_VALUES


def normalize_identifier(value: Any, platform: str = "") -> str:
    text = clean(value).lower()
    if text in INVALID_VALUES:
        return ""
    if platform == "allegro":
        return re.sub(r"[^a-z0-9]", "", text)
    return text


def parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text_value = clean(value).replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text_value.replace(" ", "T", 1))
        except ValueError:
            for format_value in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    parsed = datetime.strptime(text_value, format_value)
                    break
                except ValueError:
                    continue
            else:
                return None
    if parsed.tzinfo:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed.replace(microsecond=0) if parsed.year >= 2000 else None


def parse_amount(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        if value.year >= 2000:
            return None
        return Decimal(str(to_excel(value)))
    if isinstance(value, date):
        if value.year >= 2000:
            return None
        return Decimal(str(to_excel(value)))
    if isinstance(value, time):
        return Decimal(str(time_to_days(value)))
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text_value = clean(value).replace("\u00a0", "").replace(" ", "")
    text_value = re.sub(r"[^0-9,\.\-+()]", "", text_value)
    if not text_value:
        return None
    negative = text_value.startswith("(") and text_value.endswith(")")
    text_value = text_value.strip("()")
    if "," in text_value and "." not in text_value:
        tail = text_value.rsplit(",", 1)[-1]
        text_value = text_value.replace(",", "." if len(tail) in {1, 2} else "")
    elif "," in text_value:
        text_value = text_value.replace(",", "")
    try:
        amount = Decimal(text_value)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def parse_quantity(value: Any) -> int:
    try:
        quantity = int(Decimal(clean(value)))
    except (InvalidOperation, ValueError):
        return 1
    return quantity if quantity > 0 else 1


def decimal_text(value: Decimal | None) -> str:
    if value is None:
        return ""
    if value == 0:
        return "0"
    return format(value.normalize(), "f")


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return decimal_text(value)
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_safe(item) for item in value]
    return value


def first(values: Iterable[Any]) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def unique_valid(values: Iterable[Any]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text_value = clean(value)
        normalized = text_value.lower()
        if normalized in INVALID_VALUES or normalized in seen:
            continue
        seen.add(normalized)
        result.append(text_value)
    return result


def header_map(values: Iterable[Any]) -> tuple[list[str], dict[str, int]]:
    headers = [clean(value) for value in values]
    positions: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header:
            positions.setdefault(header, index)
    return headers, positions


def raw_row_values(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    counts: Counter[str] = Counter()
    for index, value in enumerate(row):
        header = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
        counts[header] += 1
        key = header if counts[header] == 1 else f"{header}__{counts[header]}"
        if value not in (None, ""):
            result[key] = json_safe(value)
    return result


def resolve_country_code(value: Any, default_code: str = "") -> str:
    text_value = clean(value)
    if len(text_value) == 2 and text_value.isalpha():
        return text_value.upper()
    return country_name_to_code(text_value) or default_code


def parse_sheet_rows(
    worksheet,
    config: SheetConfig,
    *,
    start_at: datetime = START_AT,
    end_at: datetime = END_AT,
) -> tuple[list[SourceRow], list[dict[str, Any]]]:
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers, positions = header_map(next(iterator))
    except StopIteration as exc:
        raise RuntimeError(f"Excel sheet is empty: {config.sheet}") from exc

    required = [config.date_header, config.order_header, config.sku_header]
    missing_headers = [header for header in required if header not in positions]
    if missing_headers:
        raise RuntimeError(f"{config.sheet} is missing required columns: {', '.join(missing_headers)}")

    def value(row: tuple[Any, ...], header: str | None) -> Any:
        if not header or header not in positions:
            return None
        index = positions[header]
        return row[index] if index < len(row) else None

    parsed_rows: list[SourceRow] = []
    issues: list[dict[str, Any]] = []
    prior_order_by_tracking: dict[str, str] = {}
    for row_no, row in enumerate(iterator, start=2):
        if not any(item not in (None, "") for item in row):
            continue
        created_at = parse_datetime(value(row, config.date_header))
        if not created_at or not (start_at <= created_at < end_at):
            continue
        order_no = clean(value(row, config.order_header))
        tracking = clean(value(row, config.tracking_header))
        tracking_key = normalize_identifier(tracking)
        attached = False
        if not is_valid_value(order_no) and config.attach_blank_order_by_tracking and tracking_key:
            order_no = prior_order_by_tracking.get(tracking_key, "")
            attached = bool(order_no)
        if not is_valid_value(order_no):
            issues.append(
                {
                    "sheet": config.sheet,
                    "row_no": row_no,
                    "reason": "missing_order_number",
                    "tracking_number": tracking,
                }
            )
            continue
        if tracking_key and not attached:
            prior_order_by_tracking[tracking_key] = order_no

        sku = clean(value(row, config.sku_header))
        if not is_valid_value(sku):
            issues.append(
                {
                    "sheet": config.sheet,
                    "row_no": row_no,
                    "reason": "missing_sku",
                    "order_no": order_no,
                }
            )
            continue
        product_name = clean(first(value(row, header) for header in config.product_headers))
        logistics = clean(first(value(row, header) for header in config.logistics_headers))
        status_values = [clean(value(row, header)) for header in config.status_headers]
        all_row_text = " | ".join(clean(item) for item in row if item not in (None, ""))
        country_value = value(row, config.country_header)
        raw_values = raw_row_values(headers, row)
        if attached:
            raw_values["_attached_order_no"] = order_no
        parsed_rows.append(
            SourceRow(
                sheet=config.sheet,
                row_no=row_no,
                platform=config.platform,
                account_id=config.account_id,
                shop_name=config.shop_name,
                order_no=order_no,
                parent_order_no=clean(value(row, config.parent_header)),
                sku=sku,
                quantity=parse_quantity(value(row, config.quantity_header)),
                amount=parse_amount(value(row, config.amount_header)),
                product_name=product_name,
                tracking_number=tracking if is_valid_value(tracking) else "",
                logistics_channel=logistics,
                country_code=resolve_country_code(country_value, config.default_country_code),
                buyer_name=clean(value(row, config.buyer_header)),
                platform_created_at=created_at,
                deadline_at=parse_datetime(value(row, config.deadline_header)),
                shipped_at=parse_datetime(value(row, config.shipped_header)),
                status_text=" | ".join([*(item for item in status_values if item), all_row_text]),
                raw_values=raw_values,
            )
        )
    return parsed_rows, issues


def load_source_rows(
    path: Path,
    *,
    start_at: datetime = START_AT,
    end_at: datetime = END_AT,
) -> tuple[list[SourceRow], list[dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        missing_sheets = [config.sheet for config in SHEET_CONFIGS if config.sheet not in workbook.sheetnames]
        if missing_sheets:
            raise RuntimeError(f"Excel workbook is missing sheets: {', '.join(missing_sheets)}")
        rows: list[SourceRow] = []
        issues: list[dict[str, Any]] = []
        for config in SHEET_CONFIGS:
            sheet_rows, sheet_issues = parse_sheet_rows(
                workbook[config.sheet],
                config,
                start_at=start_at,
                end_at=end_at,
            )
            rows.extend(sheet_rows)
            issues.extend(sheet_issues)
        return rows, issues
    finally:
        workbook.close()


def build_groups(rows: Iterable[SourceRow]) -> list[OrderGroup]:
    configs = {config.sheet: config for config in SHEET_CONFIGS}
    groups: dict[tuple[str, str, str], OrderGroup] = {}
    for row in rows:
        config = configs[row.sheet]
        key = (config.platform, config.account_id, normalize_identifier(row.order_no, config.platform))
        group = groups.setdefault(key, OrderGroup(config=config, order_no=row.order_no))
        group.rows.append(row)
    return sorted(groups.values(), key=lambda item: (item.config.sheet, min(row.row_no for row in item.rows)))


def deduplicated_rows(group: OrderGroup) -> tuple[list[SourceRow], list[int]]:
    selected: dict[tuple[str, int, str], SourceRow] = {}
    order: list[tuple[str, int, str]] = []
    duplicate_rows: list[int] = []
    for row in group.rows:
        fingerprint = (
            normalize_identifier(row.sku),
            row.quantity,
            decimal_text(row.amount),
        )
        if fingerprint not in selected:
            order.append(fingerprint)
        else:
            duplicate_rows.append(selected[fingerprint].row_no)
        selected[fingerprint] = row
    return [selected[key] for key in order], duplicate_rows


def tracking_values(group: OrderGroup) -> list[str]:
    return unique_valid(row.tracking_number for row in group.rows)


def group_amount(group: OrderGroup) -> Decimal | None:
    rows, _ = deduplicated_rows(group)
    amounts = [row.amount for row in rows if row.amount is not None]
    if group.config.amount_mode == "sum_lines":
        return sum(amounts, Decimal("0")) if amounts else None
    return first(amount for amount in amounts if amount != 0)


def group_created_at(group: OrderGroup) -> datetime:
    return min(row.platform_created_at for row in group.rows)


def group_deadline_at(group: OrderGroup) -> datetime | None:
    return min((row.deadline_at for row in group.rows if row.deadline_at), default=None)


def group_shipped_at(group: OrderGroup) -> datetime | None:
    return min((row.shipped_at for row in group.rows if row.shipped_at), default=None)


def group_cancelled(group: OrderGroup) -> bool:
    return any(VOIDED_TEXT_RE.search(row.status_text or "") for row in group.rows)


def group_items(group: OrderGroup) -> list[dict[str, Any]]:
    rows, duplicate_rows = deduplicated_rows(group)
    items: list[dict[str, Any]] = []
    for row in rows:
        unit_price: Decimal | None = None
        if group.config.amount_mode == "sum_lines" and row.amount is not None and row.quantity:
            unit_price = row.amount / Decimal(row.quantity)
        items.append(
            {
                "sku": row.sku,
                "product_name": row.product_name,
                "quantity": row.quantity,
                "unit_price": decimal_text(unit_price) or None,
                "currency": group.config.currency,
                "row_no": row.row_no,
                "duplicate_rows_removed": duplicate_rows,
            }
        )
    return items


def initial_identity(group: OrderGroup) -> tuple[str, str, str]:
    order_no = group.order_no
    if group.config.platform == "ozon":
        parts = order_no.rsplit("-", 1)
        platform_order_no = parts[0] if len(parts) == 2 and parts[1].isdigit() else order_no
        return order_no, platform_order_no, order_no
    posting_number = order_no if group.config.platform in {"wildberries", "allegro"} else ""
    return order_no, order_no, posting_number


def group_country(group: OrderGroup) -> str:
    return clean(first(row.country_code for row in group.rows if row.country_code))


def group_buyer(group: OrderGroup) -> str:
    return clean(first(row.buyer_name for row in group.rows if row.buyer_name))


def group_logistics(group: OrderGroup) -> str:
    return clean(first(row.logistics_channel for row in group.rows if row.logistics_channel))


def raw_payload_for_group(group: OrderGroup, batch_id: str, source_file: Path, imported_at: datetime) -> dict[str, Any]:
    rows, duplicate_rows = deduplicated_rows(group)
    tracks = tracking_values(group)
    return {
        "source": SOURCE_TAG,
        "excel_import_batch_id": batch_id,
        "source_file": str(source_file),
        "source_sheet": group.config.sheet,
        "source_order_identity": group.order_no,
        "source_parent_order_numbers": unique_valid(row.parent_order_no for row in group.rows),
        "source_row_numbers": [row.row_no for row in group.rows],
        "source_rows": [row.raw_values for row in group.rows],
        "deduplicated_row_numbers": [row.row_no for row in rows],
        "duplicate_row_numbers_removed": duplicate_rows,
        "excel_tracking_values": tracks,
        "explicit_excel_cancel": group_cancelled(group),
        "status_assumed": not group_cancelled(group),
        "status_source": "excel_cancel" if group_cancelled(group) else "historical_default",
        "imported_at": imported_at.isoformat(sep=" "),
    }


def event_key(batch_id: str, group: OrderGroup) -> str:
    value = f"{SOURCE_TAG}:{batch_id}:{group.config.platform}:{group.config.account_id}:{group.order_no}"
    if len(value) <= 180:
        return value
    return f"{SOURCE_TAG}:{hashlib.sha256(value.encode()).hexdigest()}"


def load_accounts(db) -> dict[tuple[str, str], PlatformAccount]:
    return {
        (account.platform, account.account_id): account
        for account in db.scalars(select(PlatformAccount).order_by(PlatformAccount.id)).all()
    }


def build_existing_index(db) -> dict[tuple[str, str, str, str], list[Order]]:
    platforms = {config.platform for config in SHEET_CONFIGS}
    rows = db.scalars(select(Order).where(Order.platform.in_(platforms))).all()
    index: dict[tuple[str, str, str, str], list[Order]] = defaultdict(list)
    for row in rows:
        for field_name in ("platform_order_id", "platform_order_no", "posting_number"):
            value = normalize_identifier(getattr(row, field_name, ""), row.platform)
            if value:
                index[(row.platform, row.account_id, field_name, value)].append(row)
                index[(row.platform, "*", field_name, value)].append(row)
    return index


def add_order_to_index(index, order: Order) -> None:
    for field_name in ("platform_order_id", "platform_order_no", "posting_number"):
        value = normalize_identifier(getattr(order, field_name, ""), order.platform)
        if value:
            index.setdefault((order.platform, order.account_id, field_name, value), []).append(order)
            index.setdefault((order.platform, "*", field_name, value), []).append(order)


def existing_match(group: OrderGroup, index) -> ExistingMatch:
    platform = group.config.platform
    account_id = group.config.account_id
    target = normalize_identifier(group.order_no, platform)
    fields = ("posting_number", "platform_order_no", "platform_order_id") if platform == "ozon" else (
        "platform_order_id",
        "platform_order_no",
        "posting_number",
    )
    for field_name in fields:
        rows = index.get((platform, account_id, field_name, target), [])
        if rows:
            return ExistingMatch(rows[0].id, rows[0].account_id, f"{field_name}_same_account")
    for field_name in fields:
        rows = [
            row
            for row in index.get((platform, "*", field_name, target), [])
            if row.account_id != account_id
        ]
        if rows:
            return ExistingMatch(rows[0].id, rows[0].account_id, f"{field_name}_other_account")
    return ExistingMatch()


def insert_group(db, group: OrderGroup, batch_id: str, source_file: Path, imported_at: datetime) -> tuple[Order, int, int]:
    platform_order_id, platform_order_no, posting_number = initial_identity(group)
    cancelled = group_cancelled(group)
    tracks = tracking_values(group)
    primary_tracking = tracks[-1] if tracks else ""
    logistics = group_logistics(group)
    country_code = group_country(group)
    fulfillment_type = normalize_fulfillment_type(group.config.fulfillment_type)
    raw_payload = raw_payload_for_group(group, batch_id, source_file, imported_at)
    order = Order(
        tenant_id=get_settings().default_tenant_id,
        platform=group.config.platform,
        account_id=group.config.account_id,
        shop_id=group.config.account_id,
        shop_name=group.config.shop_name,
        platform_order_id=platform_order_id,
        platform_order_no=platform_order_no or None,
        posting_number=posting_number,
        buyer_name=group_buyer(group) or None,
        platform_status="cancelled" if cancelled else DEFAULT_PLATFORM_STATUS,
        biz_status="已作废" if cancelled else DEFAULT_BIZ_STATUS,
        local_status="cancelled" if cancelled else DEFAULT_LOCAL_STATUS,
        fulfillment_type=fulfillment_type,
        is_overseas_warehouse=infer_is_overseas_warehouse(group.config.platform, fulfillment_type, raw_payload),
        platform_handover_deadline=group_deadline_at(group),
        platform_created_at=group_created_at(group),
        country_code=country_code or None,
        country_name_cn=country_name_cn(country_code) if country_code else None,
        buyer_selected_logistics=logistics or None,
        logistics_channel=logistics,
        order_amount=decimal_text(group_amount(group)) or None,
        currency=group.config.currency or None,
        payment_at=group_created_at(group),
        shipping_deadline_at=group_deadline_at(group),
        shipment_tracking_number=primary_tracking or None,
        marked_shipped_at=group_shipped_at(group),
        handover_at=group_shipped_at(group),
        shipped_at=group_shipped_at(group),
        last_api_payload={},
        raw_payload=raw_payload,
        created_at=imported_at,
        updated_at=imported_at,
    )
    db.add(order)
    db.flush()

    items = group_items(group)
    for item in items:
        db.add(
            OrderItem(
                order_id=order.id,
                sku=item["sku"],
                platform_product_name=item["product_name"],
                quantity=item["quantity"],
                unit_price=item["unit_price"],
                currency=item["currency"],
                raw_payload={
                    "source": SOURCE_TAG,
                    "excel_import_batch_id": batch_id,
                    "excel_row_number": item["row_no"],
                    "duplicate_row_numbers_removed": item["duplicate_rows_removed"],
                },
                created_at=imported_at,
                updated_at=imported_at,
            )
        )

    for index, tracking_number in enumerate(tracks, start=1):
        platform_shipment_id = posting_number or platform_order_id
        if len(tracks) > 1:
            platform_shipment_id = f"{platform_shipment_id}:{index}"[:160]
        db.add(
            Shipment(
                order_id=order.id,
                platform_shipment_id=platform_shipment_id,
                tracking_number=tracking_number,
                carrier=logistics or group.config.platform,
                status="shipped",
                created_at=imported_at,
            )
        )

    db.add(
        OrderOperationLog(
            order_id=order.id,
            operation_type="historical_store_tab_import",
            operation_attribute="历史订单导入",
            description=f"从 {source_file.name}/{group.config.sheet} 导入历史订单；既有订单只跳过不更新",
            operator=SYSTEM_OPERATOR,
            source=ORDER_LOG_HISTORY_SOURCE,
            event_key=event_key(batch_id, group),
            extra={
                "source": SOURCE_TAG,
                "batch_id": batch_id,
                "sheet": group.config.sheet,
                "row_numbers": [row.row_no for row in group.rows],
                "source_order_identity": group.order_no,
                "explicit_excel_cancel": cancelled,
                "tracking_numbers": tracks,
            },
            operated_at=imported_at,
            created_at=imported_at,
        )
    )
    db.flush()
    return order, len(items), len(tracks)


def report_row(group: OrderGroup, result: str, *, match: ExistingMatch | None = None, error: str = "") -> dict[str, Any]:
    rows, duplicate_rows = deduplicated_rows(group)
    return {
        "sheet": group.config.sheet,
        "platform": group.config.platform,
        "account_id": group.config.account_id,
        "shop_name": group.config.shop_name,
        "source_order_identity": group.order_no,
        "source_rows": len(group.rows),
        "deduplicated_rows": len(rows),
        "duplicate_rows_removed": ";".join(map(str, duplicate_rows)),
        "row_numbers": ";".join(str(row.row_no) for row in group.rows),
        "amount": decimal_text(group_amount(group)),
        "currency": group.config.currency,
        "tracking_numbers": ";".join(tracking_values(group)),
        "explicit_excel_cancel": group_cancelled(group),
        "result": result,
        "existing_order_id": match.order_id if match else "",
        "existing_account_id": match.account_id if match else "",
        "match_reason": match.reason if match else "",
        "error": error,
    }


def analyze_groups(groups: list[OrderGroup], index) -> tuple[list[dict[str, Any]], Counter]:
    rows: list[dict[str, Any]] = []
    stats: Counter = Counter()
    for group in groups:
        stats["source_groups"] += 1
        stats[f"sheet:{group.config.sheet}:source"] += 1
        match = existing_match(group, index)
        if match.order_id:
            result = "existing_other_account" if match.account_id != group.config.account_id else "existing_skip"
            stats[result] += 1
        else:
            result = "would_insert"
            stats["would_insert"] += 1
            stats[f"sheet:{group.config.sheet}:would_insert"] += 1
        if group_cancelled(group):
            stats["explicit_cancel"] += 1
        if group_amount(group) is None:
            stats["missing_amount"] += 1
        tracks = tracking_values(group)
        if not tracks:
            stats["missing_tracking"] += 1
        elif len(tracks) > 1:
            stats["multiple_tracking"] += 1
        rows.append(report_row(group, result, match=match))
    return rows, stats


def apply_groups(
    groups: list[OrderGroup],
    *,
    batch_id: str,
    source_file: Path,
    chunk_size: int,
) -> tuple[list[dict[str, Any]], Counter]:
    rows: list[dict[str, Any]] = []
    stats: Counter = Counter()
    imported_at = utc_now()
    with SessionLocal() as db:
        index = build_existing_index(db)
        accounts = load_accounts(db)
        missing_accounts = sorted(
            {
                (group.config.platform, group.config.account_id)
                for group in groups
                if (group.config.platform, group.config.account_id) not in accounts
            }
        )
        if missing_accounts:
            raise RuntimeError(f"System shops not found: {missing_accounts}")

        pending_since_commit = 0
        for group in groups:
            stats["source_groups"] += 1
            match = existing_match(group, index)
            if match.order_id:
                result = "existing_other_account" if match.account_id != group.config.account_id else "existing_skip"
                stats[result] += 1
                rows.append(report_row(group, result, match=match))
                continue
            try:
                with db.begin_nested():
                    order, item_count, shipment_count = insert_group(
                        db,
                        group,
                        batch_id,
                        source_file,
                        imported_at,
                    )
            except IntegrityError as exc:
                stats["unique_conflict_skip"] += 1
                rows.append(report_row(group, "unique_conflict_skip", error=str(exc.orig)[:500]))
                continue
            except Exception as exc:
                stats["failed"] += 1
                rows.append(report_row(group, "failed", error=f"{type(exc).__name__}: {str(exc)[:500]}"))
                continue
            add_order_to_index(index, order)
            stats["inserted"] += 1
            stats["items_inserted"] += item_count
            stats["shipments_inserted"] += shipment_count
            stats[f"sheet:{group.config.sheet}:inserted"] += 1
            rows.append(report_row(group, "inserted"))
            pending_since_commit += 1
            if pending_since_commit >= chunk_size:
                db.commit()
                pending_since_commit = 0
        db.commit()
    return rows, stats


def live_update_from_connector(update: Any) -> LiveUpdate:
    return LiveUpdate(
        platform_status=clean(getattr(update, "platform_status", "")),
        platform_order_id=clean(getattr(update, "platform_order_id", "")),
        platform_order_no=clean(getattr(update, "platform_order_no", "")),
        posting_number=clean(getattr(update, "posting_number", "")),
        tracking_number=clean(getattr(update, "shipment_tracking_number", "")),
        handover_at=parse_datetime(getattr(update, "handover_at", None)),
        raw_payload=json_safe(getattr(update, "raw_payload", {}) or {}),
    )


def is_voided_status(status: str) -> bool:
    normalized = clean(status).lower().replace(" ", "_")
    return normalized in VOIDED_MARKERS or "cancel" in normalized or "refund" in normalized


def account_connector(account: PlatformAccount):
    credentials = get_credential_manager().decrypt_credentials(account.encrypted_credentials)
    settings = dict(account.settings or {})
    _ensure_base_url(settings, account.platform)
    settings["account_id"] = account.account_id
    settings["display_name"] = account.display_name or account.account_id
    if account.platform == "ozon":
        settings["status_sync_lookback_days"] = max(int(settings.get("status_sync_lookback_days", 0) or 0), 900)
    return connector_for(account.platform, credentials, settings)


def api_timestamp(value: datetime) -> str:
    aware = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
    return aware.replace(microsecond=0).isoformat().replace("+00:00", "Z")


async def connector_post_with_retries(
    connector,
    path: str,
    payload: dict[str, Any],
    *,
    attempts: int = 5,
) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(attempts):
        try:
            return await connector._post(path, payload)
        except Exception as exc:
            last_error = exc
            if attempt + 1 >= attempts:
                break
            await asyncio.sleep(min(2**attempt, 30))
    assert last_error is not None
    raise last_error


async def fetch_ozon_updates(targets: list[ReconcileTarget], connector) -> dict[str, LiveUpdate]:
    pending = {target.lookup: target for target in targets if target.lookup}
    if not pending:
        return {}
    created = [target.platform_created_at for target in targets if target.platform_created_at]
    start = min(created) - timedelta(days=2) if created else START_AT - timedelta(days=2)
    end = min(max(created) + timedelta(days=2) if created else END_AT + timedelta(days=2), utc_now())
    resolved: dict[str, LiveUpdate] = {}
    cursor_start = start
    while cursor_start < end and pending:
        cursor_end = min(cursor_start + timedelta(days=30), end)
        cursor = ""
        while True:
            payload: dict[str, Any] = {
                "dir": "ASC",
                "filter": {"since": api_timestamp(cursor_start), "to": api_timestamp(cursor_end)},
                "limit": 100,
                "with": {"analytics_data": True, "barcodes": False, "financial_data": True},
            }
            if cursor:
                payload["cursor"] = cursor
            data = await connector_post_with_retries(connector, "/v4/posting/fbs/list", payload)
            result = data.get("result") if isinstance(data, dict) and isinstance(data.get("result"), dict) else {}
            postings = data.get("postings") if isinstance(data, dict) else []
            postings = postings or result.get("postings") or []
            for posting in postings if isinstance(postings, list) else []:
                posting_number = clean(posting.get("posting_number"))
                if posting_number not in pending:
                    continue
                shipping = posting.get("shipping") if isinstance(posting.get("shipping"), dict) else {}
                shipment = posting.get("shipment") if isinstance(posting.get("shipment"), dict) else {}
                resolved[posting_number] = LiveUpdate(
                    platform_status=clean(posting.get("status")),
                    platform_order_id=clean(posting.get("order_id")),
                    platform_order_no=clean(posting.get("order_number")),
                    posting_number=posting_number,
                    tracking_number=clean(
                        posting.get("tracking_number")
                        or shipping.get("tracking_number")
                        or shipment.get("tracking_number")
                    ),
                    raw_payload=json_safe(posting),
                )
                pending.pop(posting_number, None)
            cursor = clean(data.get("cursor") if isinstance(data, dict) else "") or clean(result.get("cursor"))
            has_next = bool((data.get("has_next") if isinstance(data, dict) else False) or result.get("has_next"))
            if not cursor or not has_next:
                break
        cursor_start = cursor_end

    for posting_number in list(pending):
        try:
            data = await connector_post_with_retries(
                connector,
                "/v3/posting/fbs/get",
                {
                    "posting_number": posting_number,
                    "with": {"analytics_data": True, "barcodes": False, "financial_data": True},
                },
                attempts=3,
            )
        except Exception:
            continue
        raw = data.get("result") if isinstance(data, dict) and isinstance(data.get("result"), dict) else data
        if not isinstance(raw, dict):
            continue
        resolved[posting_number] = LiveUpdate(
            platform_status=clean(raw.get("status")),
            platform_order_id=clean(raw.get("order_id")),
            platform_order_no=clean(raw.get("order_number")),
            posting_number=clean(raw.get("posting_number")) or posting_number,
            tracking_number=clean(raw.get("tracking_number")),
            raw_payload=json_safe(raw),
        )
    return resolved


async def fetch_joom_updates(targets: list[ReconcileTarget], connector) -> dict[str, LiveUpdate]:
    pending = {target.lookup for target in targets if target.lookup}
    if not pending:
        return {}
    created = [target.platform_created_at for target in targets if target.platform_created_at]
    start = min(created) - timedelta(days=2) if created else START_AT - timedelta(days=2)
    base_url = connector.base_url.rstrip("/")
    base_host = urlparse(base_url)
    next_url = ""
    params: dict[str, Any] | None = {"updatedFrom": api_timestamp(start), "limit": 500}
    updates: dict[str, LiveUpdate] = {}
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        for _page in range(500):
            url = next_url or f"{base_url}/orders/multi"
            response = None
            for attempt in range(5):
                response = await client.get(url, headers=connector.headers, params=params if not next_url else None)
                if response.status_code not in {429, 500, 502, 503, 504}:
                    break
                retry_after = response.headers.get("Retry-After")
                try:
                    delay = float(retry_after) if retry_after else 2**attempt
                except ValueError:
                    delay = 2**attempt
                await asyncio.sleep(min(delay, 30))
            assert response is not None
            response.raise_for_status()
            payload = response.json()
            data = payload.get("data") if isinstance(payload, dict) else None
            if isinstance(data, list):
                items = data
                container = payload
            else:
                container = data if isinstance(data, dict) else payload
                items = (container.get("items") or container.get("orders") or []) if isinstance(container, dict) else []
            for item in items if isinstance(items, list) else []:
                if not isinstance(item, dict):
                    continue
                order_id = clean(item.get("id") or item.get("orderId"))
                if order_id not in pending:
                    continue
                shipment = item.get("shipment") if isinstance(item.get("shipment"), dict) else {}
                updates[order_id] = LiveUpdate(
                    platform_status=clean(item.get("status") or shipment.get("status")),
                    platform_order_id=order_id,
                    platform_order_no=order_id,
                    posting_number=order_id,
                    tracking_number=clean(connector._tracking_number(item)),
                    handover_at=parse_datetime(
                        shipment.get("fulfilledTimestamp")
                        or shipment.get("shippedTimestamp")
                        or shipment.get("timestamp")
                    ),
                    raw_payload=json_safe(connector._normalize_order_payload(item)),
                )
                pending.discard(order_id)
            paging = payload.get("paging") if isinstance(payload, dict) else None
            if not isinstance(paging, dict) and isinstance(container, dict):
                paging = container.get("paging")
            next_url = clean(paging.get("next")) if isinstance(paging, dict) else ""
            if next_url:
                parsed = urlparse(next_url)
                if parsed.scheme != base_host.scheme or parsed.netloc != base_host.netloc:
                    raise RuntimeError("Joom paging.next points to an unexpected host")
            if not next_url or not pending:
                break
            params = None
    return updates


async def fetch_allegro_updates(
    targets: list[ReconcileTarget], connector, *, concurrency: int = 8
) -> dict[str, LiveUpdate]:
    semaphore = asyncio.Semaphore(concurrency)

    async def one(target: ReconcileTarget, client: httpx.AsyncClient) -> tuple[str, LiveUpdate] | None:
        lookup = target.lookup
        lookup_id = connector._checkout_form_lookup_id(lookup)
        async with semaphore:
            response = None
            for attempt in range(5):
                response = await client.get(f"{connector.base_url}/order/checkout-forms/{lookup_id}", headers=connector.headers)
                if response.status_code not in {429, 500, 502, 503, 504}:
                    break
                retry_after = response.headers.get("Retry-After")
                try:
                    delay = float(retry_after) if retry_after else 2**attempt
                except ValueError:
                    delay = 2**attempt
                await asyncio.sleep(min(delay, 30))
            assert response is not None
            if response.status_code == 404:
                return None
            response.raise_for_status()
            form = response.json()
        if not isinstance(form, dict):
            return None
        delivery = form.get("delivery") if isinstance(form.get("delivery"), dict) else {}
        fulfillment = form.get("fulfillment") if isinstance(form.get("fulfillment"), dict) else {}
        return lookup, LiveUpdate(
            platform_status=clean(fulfillment.get("status") or form.get("status")),
            platform_order_id=clean(form.get("id")) or lookup_id,
            platform_order_no=clean(form.get("id")) or lookup_id,
            posting_number=lookup,
            tracking_number=clean(delivery.get("trackingNumber") or delivery.get("tracking_number")),
            raw_payload=json_safe(form),
        )

    async with httpx.AsyncClient(timeout=60) as client:
        results = await asyncio.gather(*(one(target, client) for target in targets), return_exceptions=True)
    return dict(item for item in results if isinstance(item, tuple))


async def fetch_mercado_updates(
    targets: list[ReconcileTarget], connector, *, concurrency: int = 8
) -> dict[str, LiveUpdate]:
    semaphore = asyncio.Semaphore(concurrency)

    async def one(target: ReconcileTarget, client: httpx.AsyncClient) -> tuple[str, LiveUpdate] | None:
        lookup = target.lookup
        async with semaphore:
            detail = await connector._fetch_order_detail(client, lookup)
            if detail:
                detail = await connector._hydrate_search_item(client, detail, fetch_details=False, fetch_shipments=True)
        if not detail:
            return None
        shipping = connector._package_shipment(detail)
        inner_orders = detail.get("orders") if isinstance(detail.get("orders"), list) else []
        order_status = clean(first(item.get("status") for item in inner_orders if isinstance(item, dict)))
        return lookup, LiveUpdate(
            platform_status=clean(shipping.get("status") or shipping.get("substatus") or detail.get("status") or order_status),
            platform_order_id=clean(detail.get("id") or detail.get("order_id") or lookup),
            platform_order_no=clean(detail.get("id") or detail.get("order_id") or lookup),
            posting_number=clean(shipping.get("id") or shipping.get("shipment_id")),
            tracking_number=clean(shipping.get("tracking_number") or shipping.get("trackingNumber")),
            raw_payload=json_safe(detail),
        )

    async with httpx.AsyncClient(timeout=connector._request_timeout()) as client:
        results = await asyncio.gather(*(one(target, client) for target in targets), return_exceptions=True)
    return dict(item for item in results if isinstance(item, tuple))


async def fetch_generic_updates(targets: list[ReconcileTarget], connector) -> dict[str, LiveUpdate]:
    lookups = [target.lookup for target in targets if target.lookup]
    updates: dict[str, LiveUpdate] = {}
    for offset in range(0, len(lookups), 100):
        result = await connector.fetch_order_status_updates(lookups[offset : offset + 100])
        for item in result:
            update = live_update_from_connector(item)
            identifiers = unique_valid(
                (update.posting_number, update.platform_order_id, update.platform_order_no)
            )
            for identifier in identifiers:
                updates.setdefault(identifier, update)
    return updates


async def fetch_live_updates(
    targets: list[ReconcileTarget],
    accounts: dict[tuple[str, str], PlatformAccount],
) -> tuple[dict[tuple[str, str, str], LiveUpdate], list[dict[str, str]]]:
    grouped: dict[tuple[str, str], list[ReconcileTarget]] = defaultdict(list)
    for target in targets:
        grouped[(target.platform, target.account_id)].append(target)
    updates: dict[tuple[str, str, str], LiveUpdate] = {}
    errors: list[dict[str, str]] = []
    for account_key, account_targets in sorted(grouped.items()):
        account = accounts.get(account_key)
        if not account or not account.encrypted_credentials:
            errors.append({"platform": account_key[0], "account_id": account_key[1], "error": "credentials unavailable"})
            continue
        try:
            connector = account_connector(account)
            if account.platform == "ozon":
                result = await fetch_ozon_updates(account_targets, connector)
            elif account.platform == "joom_logistics":
                result = await fetch_joom_updates(account_targets, connector)
            elif account.platform == "allegro":
                result = await fetch_allegro_updates(account_targets, connector)
            elif account.platform == "mercadolibre":
                result = await fetch_mercado_updates(account_targets, connector)
            else:
                result = await fetch_generic_updates(account_targets, connector)
        except Exception as exc:
            errors.append(
                {
                    "platform": account_key[0],
                    "account_id": account_key[1],
                    "error": f"{type(exc).__name__}: {str(exc)[:500]}",
                }
            )
            continue
        normalized_results = {
            normalize_identifier(key, account.platform): value for key, value in result.items() if key
        }
        for target in account_targets:
            update = normalized_results.get(normalize_identifier(target.lookup, target.platform))
            if update:
                updates[(target.platform, target.account_id, target.source_order_no)] = update
    return updates, errors


def load_batch_orders(db, batch_id: str) -> list[Order]:
    return db.scalars(
        select(Order)
        .where(Order.raw_payload["excel_import_batch_id"].as_string() == batch_id)
        .order_by(Order.id)
    ).all()


def target_from_order(order: Order) -> ReconcileTarget:
    payload = order.raw_payload if isinstance(order.raw_payload, dict) else {}
    return ReconcileTarget(
        id=order.id,
        platform=order.platform,
        account_id=order.account_id,
        source_order_no=clean(payload.get("source_order_identity")),
        platform_order_id=order.platform_order_id,
        platform_order_no=order.platform_order_no or "",
        posting_number=order.posting_number or "",
        platform_created_at=order.platform_created_at,
    )


async def reconcile_batch(
    batch_id: str,
    *,
    chunk_size: int,
    platforms: set[str] | None = None,
    account_ids: set[str] | None = None,
) -> tuple[list[dict[str, Any]], Counter, list[dict[str, str]]]:
    with SessionLocal() as db:
        orders = load_batch_orders(db, batch_id)
        if platforms:
            orders = [order for order in orders if order.platform in platforms]
        if account_ids:
            orders = [order for order in orders if order.account_id in account_ids]
        accounts = load_accounts(db)
        targets = [target_from_order(order) for order in orders]
    live_updates, errors = await fetch_live_updates(targets, accounts)
    report: list[dict[str, Any]] = []
    stats: Counter = Counter()
    reconciled_at = utc_now()
    with SessionLocal() as db:
        orders = load_batch_orders(db, batch_id)
        if platforms:
            orders = [order for order in orders if order.platform in platforms]
        if account_ids:
            orders = [order for order in orders if order.account_id in account_ids]
        pending_since_commit = 0
        for order in orders:
            payload = order.raw_payload if isinstance(order.raw_payload, dict) else {}
            source_order_no = clean(payload.get("source_order_identity"))
            update = live_updates.get((order.platform, order.account_id, source_order_no))
            before_status = order.platform_status or ""
            before_tracking = order.shipment_tracking_number or ""
            if not update:
                next_payload = dict(payload)
                next_payload["status_reconciled_at"] = reconciled_at.isoformat(sep=" ")
                next_payload["status_reconciled_found"] = False
                order.raw_payload = next_payload
                stats["not_found"] += 1
                result = "not_found_kept_existing_values"
            else:
                next_payload = dict(payload)
                next_payload.update(
                    {
                        "status_reconciled_at": reconciled_at.isoformat(sep=" "),
                        "status_reconciled_found": True,
                        "status_assumed": False,
                        "status_source": "platform_live",
                        "platform_snapshot": json_safe(update.raw_payload),
                    }
                )
                order.raw_payload = next_payload
                if update.platform_status:
                    order.platform_status = update.platform_status
                if order.platform == "ozon":
                    order.platform_order_id = update.platform_order_id or order.platform_order_id
                    order.platform_order_no = update.platform_order_no or order.platform_order_no
                    order.posting_number = update.posting_number or order.posting_number
                elif order.platform == "mercadolibre" and update.posting_number:
                    order.posting_number = update.posting_number
                if update.tracking_number and is_valid_value(update.tracking_number):
                    order.shipment_tracking_number = update.tracking_number
                    shipment_exists = db.scalar(
                        select(Shipment.id).where(
                            Shipment.order_id == order.id,
                            Shipment.tracking_number == update.tracking_number,
                        )
                    )
                    if not shipment_exists:
                        db.add(
                            Shipment(
                                order_id=order.id,
                                platform_shipment_id=(update.posting_number or order.posting_number or order.platform_order_id)[:160],
                                tracking_number=update.tracking_number,
                                carrier=order.logistics_channel or order.platform,
                                status="shipped",
                                created_at=reconciled_at,
                            )
                        )
                        stats["shipments_added"] += 1
                if update.handover_at:
                    order.handover_at = update.handover_at
                    order.shipped_at = update.handover_at
                order.last_api_payload = json_safe(update.raw_payload)
                order.logistics_last_synced_at = reconciled_at
                if is_voided_status(update.platform_status):
                    order.biz_status = "已作废"
                    order.local_status = "cancelled"
                    stats["cancelled"] += 1
                    result = "platform_cancelled"
                else:
                    result = "platform_found_updated"
                stats["found"] += 1
            order.updated_at = reconciled_at
            try:
                with db.begin_nested():
                    db.flush()
            except IntegrityError as exc:
                db.refresh(order)
                stats["identity_conflict_kept"] += 1
                result = "identity_conflict_kept_existing_values"
                error = str(exc.orig)[:500]
            else:
                error = ""
            report.append(
                {
                    "order_id": order.id,
                    "platform": order.platform,
                    "account_id": order.account_id,
                    "source_order_identity": source_order_no,
                    "before_platform_status": before_status,
                    "after_platform_status": order.platform_status or "",
                    "before_tracking_number": before_tracking,
                    "after_tracking_number": order.shipment_tracking_number or "",
                    "result": result,
                    "error": error,
                }
            )
            pending_since_commit += 1
            if pending_since_commit >= chunk_size:
                db.commit()
                pending_since_commit = 0
        db.commit()
    stats["errors"] = len(errors)
    return report, stats, errors


def current_batch_state(batch_id: str) -> tuple[list[dict[str, Any]], Counter]:
    with SessionLocal() as db:
        orders = load_batch_orders(db, batch_id)
    rows: list[dict[str, Any]] = []
    stats: Counter = Counter()
    for order in orders:
        payload = order.raw_payload if isinstance(order.raw_payload, dict) else {}
        found = payload.get("status_reconciled_found")
        if found is True:
            reconciliation_result = "found"
        elif found is False:
            reconciliation_result = "not_found_kept_existing_values"
        else:
            reconciliation_result = "not_attempted"
        stats["orders"] += 1
        stats[f"reconciliation:{reconciliation_result}"] += 1
        stats[f"biz_status:{order.biz_status or '<blank>'}"] += 1
        stats[f"platform:{order.platform}"] += 1
        if not order.payment_at:
            stats["missing_payment_at"] += 1
        if not order.order_amount:
            stats["missing_amount"] += 1
        if not order.shipment_tracking_number:
            stats["missing_tracking"] += 1
        rows.append(
            {
                "order_id": order.id,
                "source_sheet": clean(payload.get("source_sheet")),
                "platform": order.platform,
                "account_id": order.account_id,
                "source_order_identity": clean(payload.get("source_order_identity")),
                "platform_order_id": order.platform_order_id,
                "platform_order_no": order.platform_order_no or "",
                "posting_number": order.posting_number or "",
                "platform_status": order.platform_status or "",
                "biz_status": order.biz_status or "",
                "tracking_number": order.shipment_tracking_number or "",
                "payment_at": order.payment_at,
                "order_amount": order.order_amount or "",
                "currency": order.currency or "",
                "reconciliation_result": reconciliation_result,
                "reconciled_at": clean(payload.get("status_reconciled_at")),
            }
        )
    return rows, stats


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0]) if rows else ["result"]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(json_safe(value), ensure_ascii=False, indent=2), encoding="utf-8")


def parse_cli_datetime(value: str) -> datetime:
    parsed = parse_datetime(value)
    if not parsed:
        raise argparse.ArgumentTypeError(f"Invalid date: {value}")
    return parsed


def make_batch_id() -> str:
    return f"h1-2025-{utc_now().strftime('%Y%m%dT%H%M%S')}-{uuid4().hex[:8]}"


def summary_from_stats(stats: Counter) -> dict[str, Any]:
    sheets: dict[str, dict[str, int]] = {}
    categories: dict[str, dict[str, int]] = {}
    summary: dict[str, Any] = {}
    for key, value in sorted(stats.items()):
        if key.startswith("sheet:"):
            _, sheet, metric = key.split(":", 2)
            sheets.setdefault(sheet, {})[metric] = value
        elif ":" in key:
            category, name = key.split(":", 1)
            categories.setdefault(category, {})[name] = value
        else:
            summary[key] = value
    summary.update(categories)
    if sheets:
        summary["sheets"] = sheets
    return summary


def reconciliation_report_stem(platforms: list[str], account_ids: list[str]) -> str:
    scope = [*sorted(set(platforms)), *sorted(set(account_ids))]
    if not scope:
        return "status_reconciliation"
    suffix = "_".join(re.sub(r"[^a-zA-Z0-9_-]+", "-", value).strip("-") for value in scope)
    return f"status_reconciliation_{suffix}"[:180]


def write_current_batch_state(report_dir: Path, batch_id: str) -> dict[str, Any]:
    state_rows, state_stats = current_batch_state(batch_id)
    state_summary = summary_from_stats(state_stats)
    write_csv(report_dir / "batch_current_state.csv", state_rows)
    write_json(
        report_dir / "batch_current_state.json",
        {"batch_id": batch_id, "stats": state_summary},
    )
    return state_summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--start", type=parse_cli_datetime, default=START_AT)
    parser.add_argument("--end", type=parse_cli_datetime, default=END_AT, help="Exclusive end date")
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--batch-id", default="")
    parser.add_argument("--chunk-size", type=int, default=200)
    parser.add_argument("--apply", action="store_true", help="Insert missing orders; default is dry-run")
    parser.add_argument("--no-reconcile", action="store_true", help="Skip live status reconciliation after --apply")
    parser.add_argument("--reconcile-batch", default="", help="Only reconcile a previously imported batch")
    parser.add_argument("--platform", action="append", default=[], help="Limit reconciliation to a platform; repeatable")
    parser.add_argument("--account-id", action="append", default=[], help="Limit reconciliation to an account; repeatable")
    return parser


async def async_main(args: argparse.Namespace) -> int:
    batch_id = args.reconcile_batch or args.batch_id or make_batch_id()
    report_dir = args.report_dir.expanduser().resolve() / batch_id
    if args.reconcile_batch:
        status_rows, status_stats, errors = await reconcile_batch(
            batch_id,
            chunk_size=max(args.chunk_size, 1),
            platforms=set(args.platform) or None,
            account_ids=set(args.account_id) or None,
        )
        report_stem = reconciliation_report_stem(args.platform, args.account_id)
        write_csv(report_dir / f"{report_stem}.csv", status_rows)
        write_json(
            report_dir / f"{report_stem}.json",
            {"batch_id": batch_id, "stats": summary_from_stats(status_stats), "account_errors": errors},
        )
        current_state = write_current_batch_state(report_dir, batch_id)
        print(
            json.dumps(
                {
                    "batch_id": batch_id,
                    "status": summary_from_stats(status_stats),
                    "current_state": current_state,
                    "report_dir": str(report_dir),
                },
                ensure_ascii=False,
            )
        )
        return 0

    rows, issues = load_source_rows(args.source, start_at=args.start, end_at=args.end)
    groups = build_groups(rows)
    with SessionLocal() as db:
        index = build_existing_index(db)
        accounts = load_accounts(db)
    required_accounts = {(group.config.platform, group.config.account_id) for group in groups}
    missing_accounts = sorted(required_accounts - set(accounts))
    if missing_accounts:
        raise RuntimeError(f"System shops not found: {missing_accounts}")
    if issues:
        write_json(report_dir / "source_issues.json", issues)
        raise RuntimeError(f"Source contains {len(issues)} unresolved H1 rows; see {report_dir / 'source_issues.json'}")

    if args.apply:
        result_rows, stats = apply_groups(
            groups,
            batch_id=batch_id,
            source_file=args.source,
            chunk_size=max(args.chunk_size, 1),
        )
        mode = "apply"
    else:
        result_rows, stats = analyze_groups(groups, index)
        mode = "dry_run"
    write_csv(report_dir / "import_results.csv", result_rows)
    summary = {
        "batch_id": batch_id,
        "mode": mode,
        "source_file": str(args.source),
        "start": args.start.isoformat(sep=" "),
        "end_exclusive": args.end.isoformat(sep=" "),
        "source_rows": len(rows),
        "source_groups": len(groups),
        "stats": summary_from_stats(stats),
        "report_dir": str(report_dir),
    }
    write_json(report_dir / "summary.json", summary)

    if args.apply and not args.no_reconcile and stats.get("inserted", 0):
        status_rows, status_stats, errors = await reconcile_batch(
            batch_id,
            chunk_size=max(args.chunk_size, 1),
            platforms=set(args.platform) or None,
            account_ids=set(args.account_id) or None,
        )
        write_csv(report_dir / "status_reconciliation.csv", status_rows)
        write_json(
            report_dir / "status_reconciliation.json",
            {"batch_id": batch_id, "stats": summary_from_stats(status_stats), "account_errors": errors},
        )
        summary["status_reconciliation"] = summary_from_stats(status_stats)
        summary["status_account_errors"] = errors
        summary["current_state"] = write_current_batch_state(report_dir, batch_id)
        write_json(report_dir / "summary.json", summary)

    print(json.dumps(summary, ensure_ascii=False))
    return 1 if stats.get("failed", 0) else 0


def main() -> int:
    return asyncio.run(async_main(build_parser().parse_args()))


if __name__ == "__main__":
    raise SystemExit(main())
