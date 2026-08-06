"""Insert historical 2025 orders and the verified 2026 Joom gap order.

This is intentionally independent from the regular order importer and sync
jobs.  It is insert-only: existing rows are reported and skipped, never
updated.  A dry-run is the default.  ``--apply`` inserts the rows, and status
reconciliation (unless disabled) only touches rows created by this batch.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import json
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse
from uuid import uuid4

import httpx
import openpyxl
from sqlalchemy import select


REPO_ROOT = Path(__file__).resolve().parents[2]
BUSINESS_ROOT = REPO_ROOT / "caifuclaw_business_app"
for _path in (REPO_ROOT, BUSINESS_ROOT):
    if str(_path) not in sys.path:
        sys.path.insert(0, str(_path))

from app.country_mapping import country_name_cn  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.credential_manager import get_credential_manager  # noqa: E402
from app.models import Order, OrderItem, OrderOperationLog, PlatformAccount, Shipment  # noqa: E402
from app.order_operation_logs import ORDER_LOG_HISTORY_SOURCE, SYSTEM_OPERATOR  # noqa: E402
from app.order_types import infer_is_overseas_warehouse, normalize_fulfillment_type  # noqa: E402
from app.settings import get_settings  # noqa: E402
from app.sync_engine import _ensure_base_url  # noqa: E402
from connector_runtime.app.factory import connector_for  # noqa: E402


DEFAULT_2025_FILE = Path("./demo_data/Order follow up 2025.xlsx")
DEFAULT_2026_FILE = Path("./demo_data/result_data_sync/Order follow up 2026.xlsx")
DEFAULT_REPORT_DIR = REPO_ROOT / "caifuclaw_business_app" / "outputs" / "historical_order_import"
SHEET_2025 = "平台订单总表"
SHEET_2026 = "订单总表"
TARGET_2026_ORDER = "DEMO-ORDER-001"
TARGET_2026_ACCOUNT = ("joom_logistics", "JOOM-DEMO-001")
SOURCE_TAG = "historical_excel_import"
SOURCE_2025 = "order_follow_up_2025"
SOURCE_2026 = "order_follow_up_2026_gap"
DEFAULT_2025_BIZ_STATUS = "已妥投"
DEFAULT_2025_PLATFORM_STATUS = "delivered"
DEFAULT_2025_LOCAL_STATUS = "shipped"
DEFAULT_2026_BIZ_STATUS = "已发货"
DEFAULT_2026_LOCAL_STATUS = "shipped"

PLATFORM_MAP = {
    "OZON": "ozon",
    "Joom": "joom_logistics",
    "mercadolibre": "mercadolibre",
    "Wildberries": "wildberries",
    "allegro": "allegro",
}
SHOP_ALIASES = {
    ("joom_logistics", "joom demo shop"): "JOOM-DEMO-001",
    ("joom_logistics", "joom demo legacy"): "J001",
    ("allegro", "demo_shop"): "allegro0002",
}
VOIDED_STATUS_MARKERS = {
    "cancel",
    "canceled",
    "cancelled",
    "cancelled_by_seller",
    "cancelled_by_customer",
    "cancelled_by_merchant",
    "refunded",
    "refund",
    "voided",
    "paidbyjoomrefund",
}


def utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None, microsecond=0)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat(sep=" ")
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "null"} else text


def parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        text = clean(value).replace("Z", "+00:00").replace(" ", "T", 1)
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            return None
    if parsed.tzinfo:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed.replace(microsecond=0) if parsed.year >= 2000 else None


def parse_excel_local_datetime(value: Any) -> datetime | None:
    parsed = parse_datetime(value)
    if not parsed:
        return None
    if isinstance(value, datetime) and value.tzinfo is not None:
        return parsed
    text = clean(value)
    if isinstance(value, str) and (text.endswith("Z") or "+" in text[10:] or "-" in text[10:]):
        return parsed
    return parsed - timedelta(hours=8)


def iso_or_none(value: datetime | None) -> str | None:
    return value.isoformat(sep=" ") if value else None


def parse_int(value: Any, default: int = 1) -> int:
    try:
        result = int(float(value))
    except (TypeError, ValueError):
        return default
    return result if result > 0 else default


def parse_money(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    try:
        return format(Decimal(text).normalize(), "f")
    except (InvalidOperation, ValueError):
        return text


def json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_safe(item) for item in value]
    return value


@dataclass
class SourceRow:
    row_no: int
    platform_label: str
    platform: str
    shop_name: str
    account_id: str | None
    order_no: str
    country_code: str
    buyer_name: str
    sku: str
    quantity: int
    unit_price: str
    currency: str
    buyer_selected_logistics: str
    logistics_channel: str
    platform_deadline_at: datetime | None
    tracking_number: str
    dispatch_deadline_at: datetime | None
    product_name: str
    order_type: str
    excel_status: str
    picking_at: datetime | None
    platform_created_at: datetime | None
    shipped_at: datetime | None


@dataclass
class OrderGroup:
    platform: str
    account_id: str
    shop_name: str
    order_no: str
    rows: list[SourceRow] = field(default_factory=list)
    source: str = SOURCE_2025
    created_at_hint: datetime | None = None

    @property
    def key(self) -> tuple[str, str, str]:
        return self.platform, self.account_id, self.order_no


@dataclass
class LiveUpdate:
    platform_status: str
    platform_order_id: str = ""
    platform_order_no: str = ""
    posting_number: str = ""
    tracking_number: str = ""
    handover_at: datetime | None = None
    raw_payload: dict[str, Any] = field(default_factory=dict)


@dataclass
class ExistingMatch:
    row: Order | None = None
    reason: str = ""


@dataclass
class ImportStats:
    source_rows: int = 0
    parsed_rows: int = 0
    unresolved_rows: int = 0
    filtered_rows: int = 0
    groups: int = 0
    existing: int = 0
    inserted: int = 0
    skipped: int = 0
    items_inserted: int = 0
    shipments_inserted: int = 0
    status_reconciled: int = 0
    status_cancelled: int = 0
    status_not_found: int = 0
    status_errors: int = 0


def load_accounts(db) -> tuple[dict[tuple[str, str], PlatformAccount], dict[tuple[str, str], PlatformAccount]]:
    by_id: dict[tuple[str, str], PlatformAccount] = {}
    by_name: dict[tuple[str, str], PlatformAccount] = {}
    for account in db.scalars(select(PlatformAccount).order_by(PlatformAccount.id)).all():
        by_id[(account.platform, account.account_id)] = account
        for value in (account.account_id, account.display_name):
            key = clean(value).lower()
            if key:
                by_name[(account.platform, key)] = account
    return by_id, by_name


def resolve_account(platform: str, shop_name: str, by_id, by_name) -> PlatformAccount | None:
    normalized = clean(shop_name).lower()
    exact = by_name.get((platform, normalized))
    if exact:
        return exact
    alias_id = SHOP_ALIASES.get((platform, normalized))
    return by_id.get((platform, alias_id)) if alias_id else None


def _column(headers: list[str], *names: str) -> int:
    for name in names:
        if name in headers:
            return headers.index(name)
    raise RuntimeError(f"Excel sheet is missing required column: {'/'.join(names)}")


def read_source(path: Path, sheet_name: str, year: int, by_id, by_name, stats: ImportStats) -> tuple[list[SourceRow], list[SourceRow]]:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Excel sheet not found: {sheet_name}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        columns = {
            "platform": _column(headers, "平台"),
            "shop": _column(headers, "店铺名"),
            "created": _column(headers, "创建时间"),
            "order": _column(headers, "订单编号"),
            "country": _column(headers, "国家二字码"),
            "buyer": _column(headers, "客户姓名"),
            "sku": _column(headers, "SKU"),
            "quantity": _column(headers, "商品数量"),
            "unit_price": _column(headers, "商品销售单价"),
            "currency": _column(headers, "币种"),
            "selected_logistics": _column(headers, "自选物流"),
            "logistics": _column(headers, "物流渠道"),
            "platform_deadline": _column(headers, "最后发货期限"),
            "tracking": _column(headers, "货运单号"),
            "dispatch_deadline": _column(headers, "发出截止时间"),
            "product": _column(headers, "产品中文名称"),
            "order_type": _column(headers, "订单类型"),
            "warning": _column(headers, "预警"),
            "picking": _column(headers, "配货日", "配货日期"),
            "shipping": _column(headers, "Shipping time"),
        }
        parsed: list[SourceRow] = []
        unresolved: list[SourceRow] = []
        for row_no, row in enumerate(rows, start=2):
            stats.source_rows += 1

            def value(name: str) -> Any:
                index = columns[name]
                return row[index] if index < len(row) else None

            platform_label = clean(value("platform"))
            platform = PLATFORM_MAP.get(platform_label, platform_label.lower())
            shop_name = clean(value("shop"))
            order_no = clean(value("order"))
            sku = clean(value("sku"))
            if not platform or not shop_name or not order_no or not sku:
                stats.filtered_rows += 1
                continue
            created_at = parse_datetime(value("created"))
            if not created_at or created_at.year != year:
                stats.filtered_rows += 1
                continue
            account = resolve_account(platform, shop_name, by_id, by_name)
            item = SourceRow(
                row_no=row_no,
                platform_label=platform_label,
                platform=platform,
                shop_name=shop_name,
                account_id=account.account_id if account else None,
                order_no=order_no,
                country_code=clean(value("country")).upper(),
                buyer_name=clean(value("buyer")),
                sku=sku,
                quantity=parse_int(value("quantity")),
                unit_price=parse_money(value("unit_price")),
                currency=clean(value("currency")).upper(),
                buyer_selected_logistics=clean(value("selected_logistics")),
                logistics_channel=clean(value("logistics")),
                platform_deadline_at=parse_datetime(value("platform_deadline")),
                tracking_number=clean(value("tracking")),
                dispatch_deadline_at=parse_datetime(value("dispatch_deadline")),
                product_name=clean(value("product")),
                order_type=clean(value("order_type")),
                excel_status=clean(value("warning")),
                picking_at=parse_excel_local_datetime(value("picking")),
                platform_created_at=created_at,
                shipped_at=parse_datetime(value("shipping")),
            )
            stats.parsed_rows += 1
            if account:
                parsed.append(item)
            else:
                stats.unresolved_rows += 1
                unresolved.append(item)
        return parsed, unresolved
    finally:
        workbook.close()


def read_target_2026(path: Path, by_id, by_name, stats: ImportStats) -> tuple[list[SourceRow], list[SourceRow]]:
    rows, unresolved = read_source(path, SHEET_2026, 2026, by_id, by_name, stats)
    target = [row for row in rows if row.order_no == TARGET_2026_ORDER]
    if not target:
        raise RuntimeError(f"2026 workbook does not contain target order {TARGET_2026_ORDER}")
    if any((row.platform, row.account_id) != TARGET_2026_ACCOUNT for row in target):
        raise RuntimeError(f"{TARGET_2026_ORDER} is not mapped to Joom JOOM-DEMO-001 in the 2026 workbook")
    return target, unresolved


def group_rows(rows: Iterable[SourceRow], source: str) -> dict[tuple[str, str, str], OrderGroup]:
    groups: dict[tuple[str, str, str], OrderGroup] = {}
    for row in rows:
        if not row.account_id:
            continue
        key = (row.platform, row.account_id, row.order_no)
        group = groups.setdefault(
            key,
            OrderGroup(
                platform=row.platform,
                account_id=row.account_id,
                shop_name=row.shop_name,
                order_no=row.order_no,
                source=source,
            ),
        )
        group.rows.append(row)
    return groups


def first(values: Iterable[str]) -> str:
    for value in values:
        if value:
            return value
    return ""


def unique(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def tracking_values(group: OrderGroup) -> list[str]:
    return unique(row.tracking_number for row in group.rows)


def group_created(group: OrderGroup) -> datetime | None:
    return min((row.platform_created_at for row in group.rows if row.platform_created_at), default=group.created_at_hint)


def group_deadline(group: OrderGroup) -> datetime | None:
    return min((row.platform_deadline_at for row in group.rows if row.platform_deadline_at), default=None)


def group_dispatch_deadline(group: OrderGroup) -> datetime | None:
    return min((row.dispatch_deadline_at for row in group.rows if row.dispatch_deadline_at), default=None)


def group_shipped(group: OrderGroup) -> datetime | None:
    return min((row.shipped_at for row in group.rows if row.shipped_at), default=None)


def group_picking(group: OrderGroup) -> datetime | None:
    return min((row.picking_at for row in group.rows if row.picking_at), default=None)


def group_logistics(group: OrderGroup) -> tuple[str, str]:
    return first(row.buyer_selected_logistics for row in group.rows), first(row.logistics_channel for row in group.rows)


def group_country(group: OrderGroup) -> str:
    return first(row.country_code for row in group.rows).upper()


def group_buyer(group: OrderGroup) -> str:
    return first(row.buyer_name for row in group.rows)


def group_currency(group: OrderGroup) -> str:
    return first(row.currency for row in group.rows)


def group_amount(group: OrderGroup) -> str:
    total = Decimal("0")
    found = False
    for row in group.rows:
        if not row.unit_price:
            continue
        try:
            total += Decimal(row.unit_price) * row.quantity
            found = True
        except InvalidOperation:
            continue
    return format(total.normalize(), "f") if found else ""


def group_items(group: OrderGroup) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in group.rows:
        key = (row.sku, row.product_name, row.unit_price, row.currency)
        item = grouped.setdefault(
            key,
            {
                "sku": row.sku,
                "product_name": row.product_name,
                "quantity": 0,
                "unit_price": row.unit_price or None,
                "currency": row.currency,
                "row_numbers": [],
            },
        )
        item["quantity"] += row.quantity
        item["row_numbers"].append(row.row_no)
    return list(grouped.values())


def fulfillment_type(group: OrderGroup) -> str:
    text = " ".join(row.order_type.upper() for row in group.rows)
    if "FBJ" in text:
        return "FBJ"
    if "海外仓" in text or "OVERSEAS" in text:
        return "OVERSEAS_WAREHOUSE"
    if group.platform == "wildberries" and "WB" in text:
        return "OVERSEAS_WAREHOUSE"
    if group.platform == "allegro" and "PL" in text:
        return "OVERSEAS_WAREHOUSE"
    return "FBS"


def normalize_allegro(value: Any) -> str:
    return clean(value).lower().replace("-", "")


def build_indexes(db) -> dict[str, dict[tuple[str, str, str], list[Order]]]:
    indexes: dict[str, dict[tuple[str, str, str], list[Order]]] = {
        field: defaultdict(list) for field in ("platform_order_id", "platform_order_no", "posting_number", "tracking_number")
    }
    for row in db.scalars(select(Order)).all():
        for field in indexes:
            value = clean(row.shipment_tracking_number if field == "tracking_number" else getattr(row, field, ""))
            if value:
                indexes[field][(row.platform, row.account_id, value)].append(row)
    return indexes


def indexed_candidates(index, platform: str, value: str, account_id: str | None = None) -> list[Order]:
    if not value:
        return []
    result: list[Order] = []
    for (indexed_platform, indexed_account, indexed_value), rows in index.items():
        if indexed_platform != platform or (account_id is not None and indexed_account != account_id):
            continue
        if indexed_value == value:
            result.extend(rows)
    return result


def existing_match(group: OrderGroup, indexes) -> ExistingMatch:
    platform = group.platform
    account = group.account_id
    identity_fields = ["posting_number"] if platform == "ozon" else ["platform_order_id", "platform_order_no", "posting_number"]
    for field in identity_fields:
        rows = indexes[field].get((platform, account, group.order_no), [])
        if rows:
            return ExistingMatch(rows[0], f"{field}_same_account")

    if platform == "allegro":
        target = normalize_allegro(group.order_no)
        for field in ("platform_order_id", "platform_order_no", "posting_number"):
            for (indexed_platform, _indexed_account, value), rows in indexes[field].items():
                if indexed_platform == platform and normalize_allegro(value) == target and rows:
                    return ExistingMatch(rows[0], "allegro_uuid_normalized")

    for field in identity_fields:
        rows = indexed_candidates(indexes[field], platform, group.order_no)
        if rows:
            return ExistingMatch(rows[0], f"{field}_other_account")

    tracks = tracking_values(group)
    if len(tracks) == 1:
        same_account = indexes["tracking_number"].get((platform, account, tracks[0]), [])
        if len(same_account) == 1:
            return ExistingMatch(same_account[0], "unique_tracking_same_account")
        other_accounts = indexed_candidates(indexes["tracking_number"], platform, tracks[0])
        if len(other_accounts) == 1:
            return ExistingMatch(other_accounts[0], "unique_tracking_other_account")
    return ExistingMatch()


def live_update_from_connector(update: Any) -> LiveUpdate:
    return LiveUpdate(
        platform_status=clean(getattr(update, "platform_status", "")),
        platform_order_id=clean(getattr(update, "platform_order_id", "")),
        platform_order_no=clean(getattr(update, "platform_order_no", "")),
        posting_number=clean(getattr(update, "posting_number", "")),
        tracking_number=clean(getattr(update, "shipment_tracking_number", "")),
        handover_at=parse_datetime(getattr(update, "handover_at", "")),
        raw_payload=json_safe(getattr(update, "raw_payload", {}) or {}),
    )


def is_voided_status(status: str) -> bool:
    normalized = clean(status).lower().replace(" ", "_")
    return normalized in VOIDED_STATUS_MARKERS or "cancel" in normalized or "refund" in normalized


def account_connector(account: PlatformAccount):
    credentials = get_credential_manager().decrypt_credentials(account.encrypted_credentials)
    settings = dict(account.settings or {})
    _ensure_base_url(settings, account.platform)
    settings["account_id"] = account.account_id
    if account.platform == "ozon":
        settings["status_sync_lookback_days"] = max(int(settings.get("status_sync_lookback_days", 0) or 0), 900)
    return connector_for(account.platform, credentials, settings)


def ozon_timestamp(value: datetime) -> str:
    value = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
    return value.replace(microsecond=0).isoformat().replace("+00:00", "Z")


async def fetch_ozon_updates(groups: list[OrderGroup], connector) -> dict[tuple[str, str, str], LiveUpdate]:
    """Read Ozon history in API-safe windows; never use a single year-long filter."""
    targets = {group.order_no: group for group in groups}
    if not targets:
        return {}
    created = [group_created(group) for group in groups if group_created(group)]
    start = min(created) - timedelta(days=2) if created else utc_now() - timedelta(days=30)
    end = min(max(created) + timedelta(days=2) if created else utc_now(), utc_now())
    resolved: dict[tuple[str, str, str], LiveUpdate] = {}
    cursor_start = start
    while cursor_start < end and targets:
        cursor_end = min(cursor_start + timedelta(days=30), end)
        cursor = ""
        while True:
            payload: dict[str, Any] = {
                "dir": "ASC",
                "filter": {"since": ozon_timestamp(cursor_start), "to": ozon_timestamp(cursor_end)},
                "limit": 100,
                "with": {"analytics_data": True, "barcodes": False, "financial_data": True},
            }
            if cursor:
                payload["cursor"] = cursor
            data = await connector._post("/v4/posting/fbs/list", payload)
            result = data.get("result") if isinstance(data, dict) and isinstance(data.get("result"), dict) else {}
            postings = data.get("postings") if isinstance(data, dict) else []
            postings = postings or result.get("postings") or []
            for posting in postings if isinstance(postings, list) else []:
                posting_number = clean(posting.get("posting_number"))
                if posting_number not in targets:
                    continue
                group = targets.pop(posting_number)
                raw_status = clean(posting.get("status"))
                tracking = clean(posting.get("tracking_number"))
                shipping = posting.get("shipping") if isinstance(posting.get("shipping"), dict) else {}
                shipment = posting.get("shipment") if isinstance(posting.get("shipment"), dict) else {}
                tracking = tracking or clean(shipping.get("tracking_number")) or clean(shipment.get("tracking_number"))
                update = LiveUpdate(
                    platform_status=raw_status,
                    platform_order_id=clean(posting.get("order_id")),
                    platform_order_no=clean(posting.get("order_number")),
                    posting_number=posting_number,
                    tracking_number=tracking,
                    raw_payload=json_safe(posting),
                )
                resolved[group.key] = update
            cursor = clean(data.get("cursor") if isinstance(data, dict) else "") or clean(result.get("cursor"))
            has_next = bool((data.get("has_next") if isinstance(data, dict) else False) or result.get("has_next"))
            if not cursor or not has_next:
                break
        cursor_start = cursor_end
    for posting_number, group in list(targets.items()):
        try:
            data = await connector._post(
                "/v3/posting/fbs/get",
                {"posting_number": posting_number, "with": {"analytics_data": True, "barcodes": False, "financial_data": True}},
            )
            raw = data.get("result") if isinstance(data, dict) and isinstance(data.get("result"), dict) else data
            if not isinstance(raw, dict):
                continue
            resolved[group.key] = LiveUpdate(
                platform_status=clean(raw.get("status")),
                platform_order_id=clean(raw.get("order_id")),
                platform_order_no=clean(raw.get("order_number")),
                posting_number=clean(raw.get("posting_number")) or posting_number,
                tracking_number=clean(raw.get("tracking_number")),
                raw_payload=json_safe(raw),
            )
        except Exception:
            continue
    return resolved


async def fetch_joom_bulk_updates(orders: list[Order], connector) -> list[LiveUpdate]:
    targets = {order.platform_order_id for order in orders if order.platform_order_id}
    if not targets:
        return []
    created = [order.platform_created_at for order in orders if order.platform_created_at]
    start = min(created) - timedelta(days=2) if created else utc_now() - timedelta(days=700)
    base_url = connector.base_url.rstrip("/")
    base_host = urlparse(base_url)
    next_url = ""
    params: dict[str, Any] | None = {"updatedFrom": ozon_timestamp(start), "limit": 500}
    updates: list[LiveUpdate] = []
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        for page in range(500):
            url = next_url or f"{base_url}/orders/multi"
            response = None
            for attempt in range(5):
                response = await client.get(url, headers=connector.headers, params=params if not next_url else None)
                if response.status_code not in {429, 500, 502, 503, 504}:
                    break
                retry_after = response.headers.get("Retry-After")
                try:
                    delay = float(retry_after) if retry_after else 2**attempt
                except ValueError:
                    delay = 2**attempt
                await asyncio.sleep(min(delay, 30))
            assert response is not None
            response.raise_for_status()
            payload = response.json()
            data = payload.get("data") if isinstance(payload, dict) else None
            if isinstance(data, list):
                items = data
                container = payload
            else:
                container = data if isinstance(data, dict) else payload
                items = container.get("items") or container.get("orders") or [] if isinstance(container, dict) else []
            for item in items if isinstance(items, list) else []:
                if not isinstance(item, dict):
                    continue
                order_id = clean(item.get("id") or item.get("orderId"))
                if order_id not in targets:
                    continue
                shipment = item.get("shipment") if isinstance(item.get("shipment"), dict) else {}
                normalized = connector._normalize_order_payload(item)
                updates.append(
                    LiveUpdate(
                        platform_status=clean(item.get("status") or shipment.get("status")),
                        platform_order_id=order_id,
                        platform_order_no=order_id,
                        posting_number=order_id,
                        tracking_number=clean(connector._tracking_number(item)),
                        handover_at=parse_datetime(
                            shipment.get("fulfilledTimestamp")
                            or shipment.get("shippedTimestamp")
                            or shipment.get("timestamp")
                        ),
                        raw_payload=json_safe(normalized),
                    )
                )
                targets.discard(order_id)
            paging = payload.get("paging") if isinstance(payload, dict) else None
            if not isinstance(paging, dict) and isinstance(container, dict):
                paging = container.get("paging")
            next_url = clean(paging.get("next")) if isinstance(paging, dict) else ""
            if next_url:
                parsed = urlparse(next_url)
                if parsed.scheme != base_host.scheme or parsed.netloc != base_host.netloc:
                    raise RuntimeError("Joom paging.next points to an unexpected host")
            if not next_url or not targets:
                break
            params = None
    return updates


async def fetch_allegro_concurrent_updates(orders: list[Order], connector, concurrency: int = 8) -> list[LiveUpdate]:
    semaphore = asyncio.Semaphore(concurrency)

    async def one(order: Order, client: httpx.AsyncClient) -> LiveUpdate | None:
        lookup = order.posting_number or order.platform_order_id
        lookup_id = connector._checkout_form_lookup_id(lookup)
        async with semaphore:
            response = None
            for attempt in range(5):
                response = await client.get(f"{connector.base_url}/order/checkout-forms/{lookup_id}", headers=connector.headers)
                if response.status_code not in {429, 500, 502, 503, 504}:
                    break
                retry_after = response.headers.get("Retry-After")
                try:
                    delay = float(retry_after) if retry_after else 2**attempt
                except ValueError:
                    delay = 2**attempt
                await asyncio.sleep(min(delay, 30))
            assert response is not None
            if response.status_code == 404:
                return None
            response.raise_for_status()
            form = response.json()
            if not isinstance(form, dict):
                return None
            delivery = form.get("delivery") if isinstance(form.get("delivery"), dict) else {}
            fulfillment = form.get("fulfillment") if isinstance(form.get("fulfillment"), dict) else {}
            return LiveUpdate(
                platform_status=clean(fulfillment.get("status") or form.get("status")),
                platform_order_id=clean(form.get("id")) or lookup_id,
                platform_order_no=clean(form.get("id")) or lookup_id,
                posting_number=lookup,
                tracking_number=clean(delivery.get("trackingNumber") or delivery.get("tracking_number")),
                raw_payload=json_safe(form),
            )

    async with httpx.AsyncClient(timeout=60) as client:
        results = await asyncio.gather(*(one(order, client) for order in orders), return_exceptions=True)
    return [result for result in results if isinstance(result, LiveUpdate)]


async def fetch_mercado_concurrent_updates(orders: list[Order], connector, concurrency: int = 8) -> list[LiveUpdate]:
    semaphore = asyncio.Semaphore(concurrency)

    async def one(order: Order, client: httpx.AsyncClient) -> LiveUpdate | None:
        lookup_id = order.posting_number or order.platform_order_id
        async with semaphore:
            detail = await connector._fetch_order_detail(client, lookup_id)
            if detail:
                detail = await connector._hydrate_search_item(client, detail, fetch_details=False, fetch_shipments=True)
        if not detail:
            return None
        shipping = connector._package_shipment(detail)
        inner_orders = detail.get("orders") if isinstance(detail.get("orders"), list) else []
        order_status = first(clean(item.get("status")) for item in inner_orders if isinstance(item, dict))
        status = clean(shipping.get("status") or shipping.get("substatus") or detail.get("status") or order_status)
        tracking = clean(shipping.get("tracking_number") or shipping.get("trackingNumber"))
        order_id = clean(detail.get("id") or detail.get("order_id") or lookup_id)
        return LiveUpdate(
            platform_status=status,
            platform_order_id=order_id,
            platform_order_no=order_id,
            posting_number=lookup_id,
            tracking_number=tracking,
            raw_payload=json_safe(detail),
        )

    async with httpx.AsyncClient(timeout=connector._request_timeout()) as client:
        results = await asyncio.gather(*(one(order, client) for order in orders), return_exceptions=True)
    return [result for result in results if isinstance(result, LiveUpdate)]


async def fetch_live_updates(groups: list[OrderGroup], accounts: dict[tuple[str, str], PlatformAccount]) -> tuple[dict[tuple[str, str, str], LiveUpdate], list[dict[str, str]]]:
    updates: dict[tuple[str, str, str], LiveUpdate] = {}
    errors: list[dict[str, str]] = []
    by_account: dict[tuple[str, str], list[OrderGroup]] = defaultdict(list)
    for group in groups:
        by_account[(group.platform, group.account_id)].append(group)
    for account_key, account_groups in sorted(by_account.items()):
        account = accounts.get(account_key)
        if not account or not account.encrypted_credentials:
            errors.append({"platform": account_key[0], "account_id": account_key[1], "error": "account credentials unavailable"})
            continue
        connector = account_connector(account)
        lookup = [group.order_no for group in account_groups]
        try:
            if account.platform == "ozon":
                result_map = await fetch_ozon_updates(account_groups, connector)
                updates.update(result_map)
                continue
            result = await connector.fetch_order_status_updates(lookup)
            for item in result:
                update = live_update_from_connector(item)
                for group in account_groups:
                    if group.order_no in {update.posting_number, update.platform_order_id, update.platform_order_no}:
                        updates[group.key] = update
                        break
        except Exception as exc:
            errors.append({"platform": account_key[0], "account_id": account_key[1], "error": f"{type(exc).__name__}: {str(exc)[:300]}"})
    return updates, errors


def identity_for_group(group: OrderGroup, live: LiveUpdate | None) -> tuple[str, str, str, str]:
    if live:
        platform_order_id = live.platform_order_id or group.order_no
        platform_order_no = live.platform_order_no or group.order_no
        posting = live.posting_number if group.platform == "ozon" else (live.posting_number or group.order_no)
        return platform_order_id, platform_order_no, posting, live.platform_status or DEFAULT_2025_PLATFORM_STATUS
    if group.platform == "ozon":
        return group.order_no, group.order_no, group.order_no, DEFAULT_2025_PLATFORM_STATUS
    posting = group.order_no if group.platform in {"wildberries", "allegro"} else ""
    return group.order_no, group.order_no, posting, DEFAULT_2025_PLATFORM_STATUS


def raw_payload_for_group(group: OrderGroup, batch_id: str, source_file: Path, imported_at: datetime, live: LiveUpdate | None) -> dict[str, Any]:
    selected, channel = group_logistics(group)
    tracks = tracking_values(group)
    return {
        "source": SOURCE_TAG,
        "source_type": group.source,
        "excel_import_batch_id": batch_id,
        "file": str(source_file),
        "sheet": SHEET_2025 if group.source == SOURCE_2025 else SHEET_2026,
        "imported_at": imported_at.isoformat(sep=" "),
        "row_numbers": [row.row_no for row in group.rows],
        "order_no": group.order_no,
        "excel_rows": [json_safe(row.__dict__) for row in group.rows],
        "excel_tracking_values": tracks,
        "tracking_number": tracks[0] if len(tracks) == 1 else "",
        "buyer_selected_logistics": selected,
        "logistics_channel": channel,
        "status_assumed": group.source == SOURCE_2025,
        "platform_snapshot": json_safe(live.raw_payload) if live else {},
    }


def event_key(batch_id: str, group: OrderGroup) -> str:
    value = f"historical_excel_import:{batch_id}:{group.platform}:{group.account_id}:{group.order_no}"
    if len(value) <= 180:
        return value
    return f"historical_excel_import:{hashlib.sha256(value.encode()).hexdigest()}"


def insert_group(db, group: OrderGroup, batch_id: str, source_file: Path, imported_at: datetime, live: LiveUpdate | None) -> tuple[Order, int, int]:
    platform_order_id, platform_order_no, posting_number, live_status = identity_for_group(group, live)
    tracks = tracking_values(group)
    tracking = live.tracking_number if group.source == SOURCE_2026 and live and live.tracking_number else (tracks[0] if len(tracks) == 1 else "")
    selected, channel = group_logistics(group)
    fulfillment = normalize_fulfillment_type(fulfillment_type(group))
    is_overseas = infer_is_overseas_warehouse(group.platform, fulfillment, {})
    if group.source == SOURCE_2026:
        biz_status = DEFAULT_2026_BIZ_STATUS
        local_status = DEFAULT_2026_LOCAL_STATUS
        platform_status = live_status or "shipped"
    else:
        biz_status = DEFAULT_2025_BIZ_STATUS
        local_status = DEFAULT_2025_LOCAL_STATUS
        platform_status = DEFAULT_2025_PLATFORM_STATUS
    raw_payload = raw_payload_for_group(group, batch_id, source_file, imported_at, live)
    order = Order(
        tenant_id=get_settings().default_tenant_id,
        platform=group.platform,
        account_id=group.account_id,
        shop_id=group.account_id,
        shop_name=group.shop_name,
        platform_order_id=platform_order_id,
        platform_order_no=platform_order_no or None,
        posting_number=posting_number,
        buyer_name=group_buyer(group) or None,
        platform_status=platform_status,
        biz_status=biz_status,
        local_status=local_status,
        fulfillment_type=fulfillment,
        is_overseas_warehouse=is_overseas,
        platform_handover_deadline=group_deadline(group),
        platform_created_at=group_created(group),
        country_code=group_country(group) or None,
        country_name_cn=country_name_cn(group_country(group)) if group_country(group) else None,
        buyer_selected_logistics=selected or channel or None,
        logistics_channel=channel,
        order_amount=group_amount(group) or None,
        currency=group_currency(group) or None,
        payment_at=group_created(group),
        shipping_deadline_at=group_deadline(group),
        dispatch_deadline_at=group_dispatch_deadline(group),
        shipment_tracking_number=tracking or None,
        picking_at=group_picking(group),
        marked_shipped_at=group_shipped(group),
        handover_at=(live.handover_at if live else group_shipped(group)),
        shipped_at=(live.handover_at if live else group_shipped(group)),
        last_api_payload=json_safe(live.raw_payload) if live else {},
        raw_payload=raw_payload,
        created_at=imported_at,
        updated_at=imported_at,
    )
    db.add(order)
    db.flush()
    item_count = 0
    for item in group_items(group):
        db.add(
            OrderItem(
                order_id=order.id,
                sku=item["sku"],
                platform_product_name=item["product_name"],
                quantity=item["quantity"],
                unit_price=item["unit_price"],
                currency=item["currency"],
                raw_payload={
                    "source": SOURCE_TAG,
                    "excel_import_batch_id": batch_id,
                    "excel_row_numbers": item["row_numbers"],
                },
                created_at=imported_at,
                updated_at=imported_at,
            )
        )
        item_count += 1
    shipment_count = 0
    if tracking:
        db.add(
            Shipment(
                order_id=order.id,
                platform_shipment_id=posting_number or platform_order_id,
                tracking_number=tracking,
                carrier=channel or selected or group.platform,
                status="shipped",
                created_at=imported_at,
            )
        )
        shipment_count = 1
    db.add(
        OrderOperationLog(
            order_id=order.id,
            operation_type="historical_excel_import",
            operation_attribute="历史订单导入",
            description=f"从 {source_file.name} 导入历史订单，批次 {batch_id}；既有订单只跳过不更新",
            operator=SYSTEM_OPERATOR,
            source=ORDER_LOG_HISTORY_SOURCE,
            event_key=event_key(batch_id, group),
            extra={
                "source": SOURCE_TAG,
                "source_type": group.source,
                "batch_id": batch_id,
                "file": str(source_file),
                "sheet": raw_payload["sheet"],
                "row_numbers": [row.row_no for row in group.rows],
                "order_no": group.order_no,
                "status_assumed": group.source == SOURCE_2025,
                "live_status": live.platform_status if live else "",
                "tracking_number": tracking,
            },
            operated_at=imported_at,
            created_at=imported_at,
        )
    )
    return order, item_count, shipment_count


def source_report_row(batch_id: str, group: OrderGroup, result: str, match: ExistingMatch | None = None, live: LiveUpdate | None = None, error: str = "") -> list[Any]:
    return [
        batch_id,
        group.source,
        group.platform,
        group.account_id,
        group.shop_name,
        group.order_no,
        len(group.rows),
        sum(row.quantity for row in group.rows),
        ";".join(str(row.row_no) for row in group.rows),
        result,
        match.row.id if match and match.row else "",
        match.reason if match else "",
        live.platform_status if live else "",
        live.tracking_number if live else "",
        error,
    ]


async def fetch_required_2026_update(group: OrderGroup, account: PlatformAccount) -> LiveUpdate:
    connector = account_connector(account)
    updates = await connector.fetch_order_status_updates([group.order_no])
    if not updates:
        raise RuntimeError(f"Joom did not return live status for {group.order_no}")
    update = live_update_from_connector(updates[0])
    if update.platform_status == "":
        raise RuntimeError(f"Joom returned an empty platform status for {group.order_no}")
    if not update.tracking_number:
        raise RuntimeError(f"Joom returned no current tracking number for {group.order_no}")
    return update


async def reconcile_statuses(
    batch_id: str,
    *,
    report_file: Path,
    only_source: str | None = None,
    platforms: set[str] | None = None,
) -> dict[str, int]:
    """Reconcile only orders inserted by this batch; never touch pre-existing rows."""
    stats = Counter()
    with SessionLocal() as db:
        orders = db.scalars(select(Order).where(Order.raw_payload["excel_import_batch_id"].as_string() == batch_id)).all()
        accounts = {
            (account.platform, account.account_id): account
            for account in db.scalars(select(PlatformAccount)).all()
        }
    if only_source:
        orders = [row for row in orders if (row.raw_payload or {}).get("source_type") == only_source]
    if platforms:
        orders = [row for row in orders if row.platform in platforms]
    if not orders:
        return dict(stats)

    by_account: dict[tuple[str, str], list[Order]] = defaultdict(list)
    for order in orders:
        by_account[(order.platform, order.account_id)].append(order)

    updates: dict[tuple[str, str, str], LiveUpdate] = {}
    errors: list[dict[str, str]] = []
    for account_key, account_orders in sorted(by_account.items()):
        account = accounts.get(account_key)
        if not account or not account.encrypted_credentials:
            errors.append({"platform": account_key[0], "account_id": account_key[1], "error": "credentials unavailable"})
            continue
        connector = account_connector(account)
        lookups = [order.posting_number or order.platform_order_id or order.platform_order_no for order in account_orders]
        orders_by_lookup = {
            order.posting_number or order.platform_order_id or order.platform_order_no: order
            for order in account_orders
        }
        try:
            if account.platform == "ozon":
                status_groups = [
                    OrderGroup(
                        platform=order.platform,
                        account_id=order.account_id,
                        shop_name=order.shop_name or order.account_id,
                        order_no=order.posting_number or order.platform_order_id,
                        source=SOURCE_2025,
                        created_at_hint=order.platform_created_at,
                    )
                    for order in account_orders
                ]
                ozon_updates = await fetch_ozon_updates(status_groups, connector)
                orders_by_posting = {
                    order.posting_number or order.platform_order_id: order for order in account_orders
                }
                for group_key, update in ozon_updates.items():
                    order = orders_by_posting.get(group_key[2])
                    if order:
                        updates[(order.platform, order.account_id, order.platform_order_id)] = update
                continue
            if account.platform == "joom_logistics":
                platform_updates = await fetch_joom_bulk_updates(account_orders, connector)
            elif account.platform == "allegro":
                platform_updates = await fetch_allegro_concurrent_updates(account_orders, connector)
            elif account.platform == "mercadolibre":
                platform_updates = await fetch_mercado_concurrent_updates(account_orders, connector)
            else:
                platform_updates = await connector.fetch_order_status_updates(lookups)
            for item in platform_updates:
                update = live_update_from_connector(item)
                order = orders_by_lookup.get(update.posting_number)
                if order:
                    updates[(order.platform, order.account_id, order.platform_order_id)] = update
                    continue
                for order in account_orders:
                    if order.posting_number == update.posting_number or order.platform_order_id == update.platform_order_id or order.platform_order_no == update.platform_order_no:
                        updates[(order.platform, order.account_id, order.platform_order_id)] = update
                        break
        except Exception as exc:
            errors.append({"platform": account_key[0], "account_id": account_key[1], "error": f"{type(exc).__name__}: {str(exc)[:300]}"})

    imported_at = utc_now()
    with SessionLocal() as db, report_file.open("a", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow([])
        writer.writerow(["status_reconciliation", "batch_id", batch_id])
        writer.writerow(["order_id", "platform", "account_id", "platform_order_id", "before_biz_status", "after_biz_status", "platform_status", "tracking_number", "result", "error"])
        db_orders = db.scalars(select(Order).where(Order.raw_payload["excel_import_batch_id"].as_string() == batch_id)).all()
        if only_source:
            db_orders = [row for row in db_orders if (row.raw_payload or {}).get("source_type") == only_source]
        if platforms:
            db_orders = [row for row in db_orders if row.platform in platforms]
        for order in db_orders:
            update = updates.get((order.platform, order.account_id, order.platform_order_id))
            before = order.biz_status or ""
            if not update:
                stats["not_found"] += 1
                writer.writerow([order.id, order.platform, order.account_id, order.platform_order_id, before, before, "", order.shipment_tracking_number or "", "not_found_or_error", ""])
                continue
            order.platform_status = update.platform_status or order.platform_status
            order.last_api_payload = json_safe(update.raw_payload)
            order.logistics_last_synced_at = imported_at
            if update.tracking_number and update.tracking_number != order.shipment_tracking_number:
                order.shipment_tracking_number = update.tracking_number
                shipment = db.scalar(select(Shipment).where(Shipment.order_id == order.id).order_by(Shipment.id.desc()))
                if shipment:
                    shipment.tracking_number = update.tracking_number
            if update.handover_at:
                order.handover_at = update.handover_at
                order.shipped_at = update.handover_at
            if order.raw_payload is None:
                order.raw_payload = {}
            payload = dict(order.raw_payload)
            payload["status_reconciled_at"] = imported_at.isoformat(sep=" ")
            payload["status_reconciled_platform_status"] = update.platform_status
            payload["status_assumed"] = False
            order.raw_payload = payload
            cancelled = is_voided_status(update.platform_status)
            if cancelled and order.biz_status != "已作废":
                order.biz_status = "已作废"
                stats["cancelled"] += 1
            else:
                stats["kept_default"] += 1
            stats["reconciled"] += 1
            writer.writerow([order.id, order.platform, order.account_id, order.platform_order_id, before, order.biz_status or "", update.platform_status, update.tracking_number, "cancelled" if cancelled else "updated_snapshot_kept_default", ""])
        for error in errors:
            stats["errors"] += 1
            writer.writerow(["", error.get("platform", ""), error.get("account_id", ""), "", "", "", "", "", "account_error", error.get("error", "")])
        db.commit()
    return dict(stats)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Insert 2025 historical orders and the verified 2026 Joom gap order.")
    parser.add_argument("--file-2025", type=Path, default=DEFAULT_2025_FILE)
    parser.add_argument("--file-2026", type=Path, default=DEFAULT_2026_FILE)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--batch-id", default="")
    parser.add_argument("--apply", action="store_true", help="Insert rows; default is dry-run")
    parser.add_argument("--reconcile-status", action=argparse.BooleanOptionalAction, default=True, help="After apply, query only this batch and correct cancellation/status snapshots")
    parser.add_argument("--reconcile-only", action="store_true", help="Skip insertion and reconcile an existing batch-id")
    parser.add_argument("--status-platform", action="append", help="Limit reconciliation to a platform; repeat as needed")
    parser.add_argument("--source-2025-only", action="store_true")
    parser.add_argument("--source-2026-only", action="store_true")
    return parser.parse_args()


async def run(args: argparse.Namespace) -> int:
    args.report_dir.mkdir(parents=True, exist_ok=True)
    batch_id = args.batch_id or f"historical-2025-2026-{utc_now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    report_file = args.report_dir / f"{batch_id}.csv"
    if args.reconcile_only:
        if not args.batch_id:
            raise RuntimeError("--reconcile-only requires --batch-id")
        result = await reconcile_statuses(
            batch_id,
            report_file=report_file,
            only_source=SOURCE_2025,
            platforms=set(args.status_platform or []),
        )
        print(json.dumps({"mode": "RECONCILE_ONLY", "batch_id": batch_id, "status": result, "report": str(report_file)}, ensure_ascii=False, indent=2))
        return 0

    stats = ImportStats()
    with SessionLocal() as db:
        accounts_by_id, accounts_by_name = load_accounts(db)
    rows_2025, unresolved_2025 = read_source(args.file_2025, SHEET_2025, 2025, accounts_by_id, accounts_by_name, stats)
    rows_2026, unresolved_2026 = read_target_2026(args.file_2026, accounts_by_id, accounts_by_name, stats)
    groups_2025 = group_rows(rows_2025, SOURCE_2025)
    groups_2026 = group_rows(rows_2026, SOURCE_2026)
    selected_groups: list[OrderGroup] = []
    if not args.source_2026_only:
        selected_groups.extend(groups_2025.values())
    if not args.source_2025_only:
        selected_groups.extend(groups_2026.values())
    stats.groups = len(selected_groups)

    live_updates: dict[tuple[str, str, str], LiveUpdate] = {}
    target_2026 = groups_2026.get(("joom_logistics", "JOOM-DEMO-001", TARGET_2026_ORDER))
    if target_2026 and not args.source_2025_only:
        account = accounts_by_id[TARGET_2026_ACCOUNT]
        try:
            live_updates[target_2026.key] = await fetch_required_2026_update(target_2026, account)
        except Exception as exc:
            raise RuntimeError(f"Mandatory live query for {TARGET_2026_ORDER} failed: {exc}") from exc

    report_file.parent.mkdir(parents=True, exist_ok=True)
    with report_file.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["batch_id", "source_type", "platform", "account_id", "shop_name", "order_no", "row_count", "item_quantity", "row_numbers", "result", "existing_order_id", "reason", "live_status", "live_tracking", "error"])
        imported_at = utc_now()
        with SessionLocal() as db:
            indexes = build_indexes(db)
            plans: list[tuple[OrderGroup, ExistingMatch]] = []
            for group in selected_groups:
                match = existing_match(group, indexes)
                if match.row:
                    stats.existing += 1
                    stats.skipped += 1
                    writer.writerow(source_report_row(batch_id, group, "existing_skipped", match, live_updates.get(group.key)))
                    continue
                plans.append((group, match))
                writer.writerow(source_report_row(batch_id, group, "would_insert" if not args.apply else "inserted", None, live_updates.get(group.key)))
            if args.apply:
                ozon_groups = [group for group, _match in plans if group.platform == "ozon"]
                if ozon_groups:
                    ozon_updates, ozon_errors = await fetch_live_updates(ozon_groups, accounts_by_id)
                    live_updates.update(ozon_updates)
                    missing_ozon = [group.order_no for group in ozon_groups if group.key not in ozon_updates]
                    if ozon_errors or missing_ozon:
                        details = {
                            "account_errors": ozon_errors,
                            "missing_order_count": len(missing_ozon),
                            "missing_order_samples": missing_ozon[:20],
                        }
                        raise RuntimeError(f"Ozon live identity verification failed; no rows were committed: {json.dumps(details, ensure_ascii=False)}")
                for group, _match in plans:
                    live = live_updates.get(group.key)
                    if group.source == SOURCE_2026 and not live:
                        raise RuntimeError(f"Missing mandatory live update for {group.order_no}")
                    order, item_count, shipment_count = insert_group(db, group, batch_id, args.file_2025 if group.source == SOURCE_2025 else args.file_2026, imported_at, live)
                    stats.inserted += 1
                    stats.items_inserted += item_count
                    stats.shipments_inserted += shipment_count
                db.commit()
            else:
                db.rollback()

    if args.apply and args.reconcile_status and not args.source_2026_only:
        status = await reconcile_statuses(
            batch_id,
            report_file=report_file,
            only_source=SOURCE_2025,
            platforms=set(args.status_platform or []),
        )
        stats.status_reconciled = status.get("reconciled", 0)
        stats.status_cancelled = status.get("cancelled", 0)
        stats.status_not_found = status.get("not_found", 0)
        stats.status_errors = status.get("errors", 0)
    summary = {
        "mode": "APPLY" if args.apply else "DRY_RUN",
        "batch_id": batch_id,
        "source_2025": {"file": str(args.file_2025), "groups": len(groups_2025), "unresolved_rows": len(unresolved_2025)},
        "source_2026": {
            "file": str(args.file_2026),
            "target_groups": len(groups_2026),
            "unresolved_rows": len(unresolved_2026),
            "target_live": (
                {
                    "platform_status": live_updates[target_2026.key].platform_status,
                    "platform_order_id": live_updates[target_2026.key].platform_order_id,
                    "tracking_number": live_updates[target_2026.key].tracking_number,
                    "handover_at": live_updates[target_2026.key].handover_at,
                }
                if target_2026 and target_2026.key in live_updates
                else None
            ),
        },
        "stats": stats.__dict__,
        "unresolved_2025_samples": [row.__dict__ for row in unresolved_2025[:20]],
        "unresolved_2026_samples": [row.__dict__ for row in unresolved_2026[:20]],
        "report": str(report_file),
    }
    print(json.dumps(json_safe(summary), ensure_ascii=False, indent=2))
    return 0


def main() -> None:
    args = parse_args()
    raise SystemExit(asyncio.run(run(args)))


if __name__ == "__main__":
    main()
