from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import sys

import openpyxl
from sqlalchemy import func, select, text
from sqlalchemy.orm import joinedload

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.database import SessionLocal, engine
from app.models import PlatformAccount
from app.product_models import Product, ProductShopMapping


DEFAULT_EXCEL = Path(
    r".\demo_data\product_catalog.xlsx"
)

PRODUCT_SHEET = "\u4ea7\u54c1\u76ee\u5f55"
SHOP_B_NAME = "OZON DEMO SHOP A"
SHOP_C_NAME = "Ozon Demo Shop B"

COL_NAME = 0
COL_SHOP_B = 1
COL_SHOP_C = 2
COL_COST = 11
COL_WEIGHT = 12
COL_STOCK = 13


@dataclass
class ImportStats:
    rows_seen: int = 0
    products_created: int = 0
    products_updated: int = 0
    mappings_created: int = 0
    mappings_existing: int = 0
    mappings_conflicted: int = 0
    empty_name_rows: int = 0
    no_mapping_rows: int = 0
    cost_missing_as_zero: int = 0


def clean(value) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    return "" if text_value.upper() == "#N/A" else text_value


def decimal_value(value, *, zero_on_shortage: bool = False) -> tuple[Decimal | None, bool]:
    text_value = clean(value)
    if not text_value:
        return None, False
    if zero_on_shortage and ("\u7f3a\u8d27" in text_value or text_value in {"\u65e0", "N/A", "n/a"}):
        return Decimal("0"), True
    try:
        return Decimal(text_value), False
    except InvalidOperation:
        raise ValueError(f"数字格式错误: {text_value}")


def int_value(value) -> int | None:
    text_value = clean(value)
    if not text_value:
        return None
    try:
        return int(float(text_value))
    except ValueError as exc:
        raise ValueError(f"整数格式错误: {text_value}") from exc


def generate_product_code(db) -> str:
    max_code = db.scalar(select(func.max(Product.product_code)).where(Product.product_code.like("P________")))
    next_number = 1
    if max_code and len(max_code) == 9 and max_code[1:].isdigit():
        next_number = int(max_code[1:]) + 1
    while True:
        product_code = f"P{next_number:08d}"
        if not db.scalar(select(Product.id).where(Product.product_code == product_code)):
            return product_code
        next_number += 1


def find_enabled_shop(db, display_name: str) -> PlatformAccount:
    shop = db.scalar(
        select(PlatformAccount).where(
            PlatformAccount.enabled == True,
            func.lower(PlatformAccount.display_name) == display_name.lower(),
        )
    )
    if not shop:
        raise RuntimeError(f"未找到启用店铺: {display_name}")
    return shop


def ensure_database_constraint() -> None:
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE IF EXISTS product_shop_mappings DROP CONSTRAINT IF EXISTS uq_product_shop_mapping"))


def add_mapping(
    db,
    product: Product,
    shop: PlatformAccount,
    sku: str,
    stats: ImportStats,
    conflicts: list[dict],
    row_no: int,
    seen_shop_skus: dict[tuple[int, str], str],
) -> None:
    seen_key = (shop.id, sku)
    seen_product_name = seen_shop_skus.get(seen_key)
    if seen_product_name:
        if seen_product_name == product.internal_name:
            stats.mappings_existing += 1
            return
        stats.mappings_conflicted += 1
        conflicts.append(
            {
                "row": row_no,
                "shop": shop.display_name,
                "sku": sku,
                "excel_product": product.internal_name,
                "existing_product": seen_product_name,
            }
        )
        return
    existing = db.scalar(
        select(ProductShopMapping)
        .options(joinedload(ProductShopMapping.product))
        .where(ProductShopMapping.shop_id == shop.id, ProductShopMapping.shop_sku == sku)
    )
    if existing:
        if existing.product_id == product.id:
            stats.mappings_existing += 1
            return
        stats.mappings_conflicted += 1
        conflicts.append(
            {
                "row": row_no,
                "shop": shop.display_name,
                "sku": sku,
                "excel_product": product.internal_name,
                "existing_product": existing.product.internal_name if existing.product else existing.product_id,
            }
        )
        return
    db.add(ProductShopMapping(product_id=product.id, shop_id=shop.id, shop_sku=sku))
    seen_shop_skus[seen_key] = product.internal_name
    stats.mappings_created += 1


def run_import(path: Path, apply: bool, log_path: Path) -> ImportStats:
    if not path.exists():
        raise FileNotFoundError(path)
    ensure_database_constraint()

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[PRODUCT_SHEET] if PRODUCT_SHEET in workbook.sheetnames else workbook[workbook.sheetnames[1]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]

    stats = ImportStats()
    conflicts: list[dict] = []
    errors: list[dict] = []
    duplicate_names = Counter()
    row_skus_by_name = defaultdict(list)
    seen_shop_skus: dict[tuple[int, str], str] = {}

    with SessionLocal() as db:
        shop_b = find_enabled_shop(db, SHOP_B_NAME)
        shop_c = find_enabled_shop(db, SHOP_C_NAME)
        print(f"B列 -> 店铺 {shop_b.id} {shop_b.display_name}")
        print(f"C列 -> 店铺 {shop_c.id} {shop_c.display_name}")
        print("表头:", headers)
        for mapping in db.scalars(select(ProductShopMapping).options(joinedload(ProductShopMapping.product))).all():
            seen_shop_skus[(mapping.shop_id, mapping.shop_sku)] = mapping.product.internal_name if mapping.product else str(mapping.product_id)

        for row_no, row in enumerate(rows, start=2):
            stats.rows_seen += 1
            name = clean(row[COL_NAME] if COL_NAME < len(row) else None)
            if not name:
                stats.empty_name_rows += 1
                continue
            duplicate_names[name] += 1

            b_sku = clean(row[COL_SHOP_B] if COL_SHOP_B < len(row) else None)
            c_sku = clean(row[COL_SHOP_C] if COL_SHOP_C < len(row) else None)
            if not b_sku and not c_sku:
                stats.no_mapping_rows += 1

            try:
                cost, shortage_as_zero = decimal_value(row[COL_COST] if COL_COST < len(row) else None, zero_on_shortage=True)
                if shortage_as_zero:
                    stats.cost_missing_as_zero += 1
                weight, _ = decimal_value(row[COL_WEIGHT] if COL_WEIGHT < len(row) else None)
                safety_stock = int_value(row[COL_STOCK] if COL_STOCK < len(row) else None)
            except Exception as exc:
                errors.append({"row": row_no, "product": name, "message": str(exc)})
                continue

            product = db.scalar(select(Product).options(joinedload(Product.mappings)).where(Product.internal_name == name))
            if product:
                stats.products_updated += 1
            else:
                product = Product(product_code=generate_product_code(db), internal_name=name, enabled=True)
                db.add(product)
                db.flush()
                stats.products_created += 1

            product.cost = cost
            product.weight = weight
            product.safety_stock = safety_stock
            product.updated_at = datetime.utcnow()

            if b_sku:
                row_skus_by_name[(name, shop_b.id)].append(b_sku)
                add_mapping(db, product, shop_b, b_sku, stats, conflicts, row_no, seen_shop_skus)
            if c_sku:
                row_skus_by_name[(name, shop_c.id)].append(c_sku)
                add_mapping(db, product, shop_c, c_sku, stats, conflicts, row_no, seen_shop_skus)

        if apply:
            db.commit()
        else:
            db.rollback()

    multi_sku_same_product = [
        {"product": name, "shop_id": shop_id, "sku_count": len(set(skus)), "skus": sorted(set(skus))}
        for (name, shop_id), skus in row_skus_by_name.items()
        if len(set(skus)) > 1
    ]
    duplicated_name_count = sum(1 for count in duplicate_names.values() if count > 1)

    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["type", "row", "shop", "sku", "product", "message"])
        for item in conflicts:
            writer.writerow(["conflict", item["row"], item["shop"], item["sku"], item["excel_product"], f"已绑定: {item['existing_product']}"])
        for item in errors:
            writer.writerow(["error", item["row"], "", "", item["product"], item["message"]])
        writer.writerow([])
        writer.writerow(["summary", "rows_seen", stats.rows_seen])
        writer.writerow(["summary", "products_created", stats.products_created])
        writer.writerow(["summary", "products_updated", stats.products_updated])
        writer.writerow(["summary", "mappings_created", stats.mappings_created])
        writer.writerow(["summary", "mappings_existing", stats.mappings_existing])
        writer.writerow(["summary", "mappings_conflicted", stats.mappings_conflicted])
        writer.writerow(["summary", "duplicate_product_names", duplicated_name_count])
        writer.writerow(["summary", "multi_sku_same_product_shop", len(multi_sku_same_product)])

    print("模式:", "正式写入" if apply else "dry-run 未写入")
    print("读取行数:", stats.rows_seen)
    print("新增产品:", stats.products_created)
    print("更新产品:", stats.products_updated)
    print("新增映射:", stats.mappings_created)
    print("已存在映射:", stats.mappings_existing)
    print("冲突映射:", stats.mappings_conflicted)
    print("空中文名称行:", stats.empty_name_rows)
    print("无B/C映射行:", stats.no_mapping_rows)
    print("成本缺货按0:", stats.cost_missing_as_zero)
    print("重复中文名称数:", duplicated_name_count)
    print("同产品同店铺多SKU数:", len(multi_sku_same_product))
    print("日志:", log_path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="导入产品目录 B/C 列到产品表和产品店铺映射表")
    parser.add_argument("--file", type=Path, default=DEFAULT_EXCEL, help="产品目录 Excel 文件")
    parser.add_argument("--apply", action="store_true", help="正式写入；不加则 dry-run")
    parser.add_argument("--log", type=Path, default=ROOT / "logs" / "product_catalog_bc_import.csv", help="导入日志 CSV")
    args = parser.parse_args()
    run_import(args.file, args.apply, args.log)


if __name__ == "__main__":
    main()
