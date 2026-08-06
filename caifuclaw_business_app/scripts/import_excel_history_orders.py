from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time as datetime_time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

import httpx
import openpyxl
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.country_mapping import country_name_cn
from app.credential_manager import get_credential_manager
from app.database import SessionLocal
from app.models import Order, OrderItem, OrderOperationLog, PlatformAccount, Shipment
from app.order_operation_logs import ORDER_LOG_HISTORY_SOURCE, SYSTEM_OPERATOR
from app.order_types import infer_is_overseas_warehouse, normalize_fulfillment_type
from app.settings import get_settings


UTC = timezone.utc
ORDER_NUMBER_FIELDS = ("shipment_tracking_number", "posting_number", "platform_order_no", "platform_order_id")
DEFAULT_EXCEL = Path("./demo_data/result_data_sync/Order follow up 2026.xlsx")
ORDER_SHEET = "订单总表"
SOURCE_TAG = "excel_import"
SOURCE_TYPE = "order_follow_up_2026"
IMPORTED_BIZ_STATUS = "已发货"
IMPORTED_LOCAL_STATUS = "shipped"
LOCAL_TIME_OFFSET = timedelta(hours=8)
PICKING_DATE_COLUMNS = ("配货日", "配货日期")

PLATFORM_MAP = {
    "OZON": "ozon",
    "Joom": "joom_logistics",
    "mercadolibre": "mercadolibre",
    "Wildberries": "wildberries",
    "allegro": "allegro",
}

SHOP_ALIASES = {
    ("joom_logistics", "joom demo shop"): ("JOOM-DEMO-001", "Joom Demo Shop"),
    ("joom_logistics", "joom demo legacy"): ("J001", "Joom Demo Shop"),
    ("allegro", "demo_shop"): ("allegro0002", "Allegro Demo Shop"),
}

VALID_CURRENCIES = {"CNY", "RUB", "USD", "PLN", "EUR", "GBP", "BRL", "CLP", "MXN", "COP", "ARS"}


def utc_now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(value, date):
        return datetime.combine(value, datetime_time.min).isoformat(sep=" ")
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def clean_key(value: Any) -> str:
    return clean(value).lower()


def parse_datetime_value(value: Any) -> tuple[datetime | None, bool]:
    if value in (None, ""):
        return None, False
    if isinstance(value, datetime):
        parsed = value
        has_timezone = value.tzinfo is not None
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime_time.min)
        has_timezone = False
    elif isinstance(value, datetime_time):
        return None, False
    else:
        text = clean(value)
        if not text or text in {"00:00:00", "1900-1-0"}:
            return None, False
        normalized = text.replace("Z", "+00:00")
        if " " in normalized and "T" not in normalized:
            normalized = normalized.replace(" ", "T", 1)
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            return None, False
        has_timezone = parsed.tzinfo is not None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)
    parsed = parsed.replace(microsecond=0)
    if parsed.year < 2000:
        return None, has_timezone
    return parsed, has_timezone


def parse_datetime(value: Any) -> datetime | None:
    parsed, _has_timezone = parse_datetime_value(value)
    return parsed


def parse_excel_local_datetime(value: Any) -> datetime | None:
    parsed, has_timezone = parse_datetime_value(value)
    if not parsed:
        return None
    if has_timezone:
        return parsed
    return (parsed - LOCAL_TIME_OFFSET).replace(microsecond=0)


def iso_or_none(value: datetime | None) -> str | None:
    if not value:
        return None
    return value.replace(microsecond=0).isoformat() + "Z"


def decimal_or_none(value: Any) -> Decimal | None:
    text = clean(value)
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def money_text(value: Any) -> str:
    amount = decimal_or_none(value)
    if amount is None:
        return ""
    return format(amount.normalize(), "f")


def currency_code(value: Any) -> str:
    text = clean(value).upper()
    if len(text) > 12:
        return ""
    return text if text in VALID_CURRENCIES else ""


def int_or_default(value: Any, default: int = 1) -> int:
    text = clean(value)
    if not text:
        return default
    try:
        result = int(float(text))
    except ValueError:
        return default
    return result if result > 0 else default


def order_no_base(order_no: str) -> str:
    parts = clean(order_no).split("-")
    if len(parts) >= 3 and parts[-1].isdigit():
        return "-".join(parts[:-1])
    return clean(order_no)


def status_text(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    if text.startswith("1900-"):
        return ""
    return text


def excel_status_to_platform_status(value: str) -> str:
    normalized = value.strip().lower()
    if normalized == "delivered":
        return "delivered"
    if normalized in {"delayed", "urgent"}:
        return "shipped"
    return "shipped"


@dataclass
class SourceRow:
    row_no: int
    platform_label: str
    platform: str
    shop_name: str
    account: PlatformAccount | None
    mapping_mode: str
    order_no: str
    country_code: str
    buyer_name: str
    sku: str
    quantity: int
    unit_price: str
    currency: str
    buyer_selected_logistics: str
    logistics_channel: str
    platform_deadline_at: datetime | None
    tracking_number: str
    dispatch_deadline_at: datetime | None
    product_name: str
    order_type: str
    excel_status: str
    picking_at: datetime | None
    platform_created_at: datetime | None
    shipped_at: datetime | None


@dataclass
class OzonInfo:
    posting_number: str
    order_id: str
    order_number: str
    status: str
    raw_payload: dict[str, Any]


@dataclass
class ExistingMatch:
    exists: bool = False
    reason: str = ""
    order_ids: list[int] = field(default_factory=list)
    tracking_conflict: str = ""


@dataclass
class OrderGroup:
    platform: str
    account_id: str
    shop_name: str
    order_no: str
    rows: list[SourceRow] = field(default_factory=list)
    existing: ExistingMatch = field(default_factory=ExistingMatch)
    ozon_info: OzonInfo | None = None
    ozon_error: str = ""

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.platform, self.account_id, self.order_no)


@dataclass
class ImportStats:
    source_rows: int = 0
    parsed_rows: int = 0
    unresolved_rows: int = 0
    groups: int = 0
    skipped_existing_groups: int = 0
    skipped_existing_by_tracking: int = 0
    tracking_conflict_groups: int = 0
    ozon_target_groups: int = 0
    ozon_resolved_groups: int = 0
    ozon_failed_groups: int = 0
    imported_orders: int = 0
    imported_items: int = 0
    imported_shipments: int = 0
    imported_logs: int = 0
    picking_date_groups: int = 0
    updated_picking_dates: int = 0
    unchanged_picking_dates: int = 0
    skipped_missing_picking_dates: int = 0
    skipped_missing_existing_orders: int = 0
    skipped_ambiguous_existing_orders: int = 0
    skipped_non_excel_import_orders: int = 0
    tracking_index_values: int = 0
    tracking_index_conflicts: int = 0
    tracking_match_candidates: int = 0
    skipped_existing_picking_dates: int = 0
    skipped_missing_tracking_numbers: int = 0
    skipped_missing_excel_tracking: int = 0
    skipped_conflicting_excel_tracking: int = 0


def build_account_lookup(db) -> tuple[dict[tuple[str, str], PlatformAccount], dict[tuple[str, str], PlatformAccount]]:
    rows = db.scalars(select(PlatformAccount).order_by(PlatformAccount.id)).all()
    exact: dict[tuple[str, str], PlatformAccount] = {}
    by_account: dict[tuple[str, str], PlatformAccount] = {}
    for account in rows:
        by_account[(account.platform, account.account_id)] = account
        for value in {account.display_name, account.account_id}:
            key = clean_key(value)
            if key:
                exact[(account.platform, key)] = account
    return exact, by_account


def resolve_account(
    platform: str,
    shop_name: str,
    exact: dict[tuple[str, str], PlatformAccount],
    by_account: dict[tuple[str, str], PlatformAccount],
) -> tuple[PlatformAccount | None, str]:
    shop_key = clean_key(shop_name)
    account = exact.get((platform, shop_key))
    if account:
        return account, "exact"
    alias = SHOP_ALIASES.get((platform, shop_key))
    if alias:
        account_id, _display_name = alias
        return by_account.get((platform, account_id)), "alias"
    return None, "unresolved"


def find_column(headers: list[str], name: str) -> int:
    try:
        return headers.index(name)
    except ValueError as exc:
        raise RuntimeError(f"Excel sheet is missing required column: {name}") from exc


def find_any_column(headers: list[str], names: tuple[str, ...]) -> int:
    for name in names:
        if name in headers:
            return headers.index(name)
    raise RuntimeError(f"Excel sheet is missing required column: {'/'.join(names)}")


def read_source_rows(path: Path, stats: ImportStats) -> list[SourceRow]:
    if not path.exists():
        raise FileNotFoundError(path)

    with SessionLocal() as db:
        exact_accounts, accounts_by_id = build_account_lookup(db)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if ORDER_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"Excel sheet not found: {ORDER_SHEET}")
        worksheet = workbook[ORDER_SHEET]
        row_iter = worksheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(row_iter)]
        col = {
            "picking_date": find_any_column(headers, PICKING_DATE_COLUMNS),
            "platform_label": find_column(headers, "平台"),
            "shop_name": find_column(headers, "店铺名"),
            "created_at": find_column(headers, "创建时间"),
            "order_no": find_column(headers, "订单编号"),
            "country_code": find_column(headers, "国家二字码"),
            "buyer_name": find_column(headers, "客户姓名"),
            "sku": find_column(headers, "SKU"),
            "quantity": find_column(headers, "商品数量"),
            "unit_price": find_column(headers, "商品销售单价"),
            "currency": find_column(headers, "币种"),
            "buyer_selected_logistics": find_column(headers, "自选物流"),
            "logistics_channel": find_column(headers, "物流渠道"),
            "platform_deadline": find_column(headers, "最后发货期限"),
            "tracking_number": find_column(headers, "货运单号"),
            "dispatch_deadline": find_column(headers, "发出截止时间"),
            "product_name": find_column(headers, "产品中文名称"),
            "order_type": find_column(headers, "订单类型"),
            "excel_status": find_column(headers, "预警"),
            "shipping_time": find_column(headers, "Shipping time"),
        }
        rows: list[SourceRow] = []
        for row_no, row in enumerate(row_iter, start=2):
            stats.source_rows += 1
            platform_label = clean(row[col["platform_label"]] if col["platform_label"] < len(row) else None)
            platform = PLATFORM_MAP.get(platform_label, platform_label.lower())
            shop_name = clean(row[col["shop_name"]] if col["shop_name"] < len(row) else None)
            order_no = clean(row[col["order_no"]] if col["order_no"] < len(row) else None)
            sku = clean(row[col["sku"]] if col["sku"] < len(row) else None)
            if not platform or not shop_name or not order_no or not sku:
                continue
            account, mapping_mode = resolve_account(platform, shop_name, exact_accounts, accounts_by_id)
            if not account:
                stats.unresolved_rows += 1
            source_row = SourceRow(
                row_no=row_no,
                platform_label=platform_label,
                platform=platform,
                shop_name=shop_name,
                account=account,
                mapping_mode=mapping_mode,
                order_no=order_no,
                country_code=clean(row[col["country_code"]] if col["country_code"] < len(row) else None).upper(),
                buyer_name=clean(row[col["buyer_name"]] if col["buyer_name"] < len(row) else None),
                sku=sku,
                quantity=int_or_default(row[col["quantity"]] if col["quantity"] < len(row) else None),
                unit_price=money_text(row[col["unit_price"]] if col["unit_price"] < len(row) else None),
                currency=currency_code(row[col["currency"]] if col["currency"] < len(row) else None),
                buyer_selected_logistics=clean(row[col["buyer_selected_logistics"]] if col["buyer_selected_logistics"] < len(row) else None),
                logistics_channel=clean(row[col["logistics_channel"]] if col["logistics_channel"] < len(row) else None),
                platform_deadline_at=parse_datetime(row[col["platform_deadline"]] if col["platform_deadline"] < len(row) else None),
                tracking_number=clean(row[col["tracking_number"]] if col["tracking_number"] < len(row) else None),
                dispatch_deadline_at=parse_datetime(row[col["dispatch_deadline"]] if col["dispatch_deadline"] < len(row) else None),
                product_name=clean(row[col["product_name"]] if col["product_name"] < len(row) else None),
                order_type=clean(row[col["order_type"]] if col["order_type"] < len(row) else None),
                excel_status=status_text(row[col["excel_status"]] if col["excel_status"] < len(row) else None),
                picking_at=parse_excel_local_datetime(row[col["picking_date"]] if col["picking_date"] < len(row) else None),
                platform_created_at=parse_datetime(row[col["created_at"]] if col["created_at"] < len(row) else None),
                shipped_at=parse_datetime(row[col["shipping_time"]] if col["shipping_time"] < len(row) else None),
            )
            rows.append(source_row)
            stats.parsed_rows += 1
        return rows
    finally:
        workbook.close()


def read_tracking_picking_index(path: Path, stats: ImportStats) -> tuple[dict[str, datetime], dict[str, set[datetime]]]:
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if ORDER_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"Excel sheet not found: {ORDER_SHEET}")
        worksheet = workbook[ORDER_SHEET]
        row_iter = worksheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(row_iter)]
        picking_col = find_any_column(headers, PICKING_DATE_COLUMNS)
        tracking_col = find_column(headers, "货运单号")
        values: dict[str, set[datetime]] = defaultdict(set)
        for row in row_iter:
            tracking_number = clean(row[tracking_col] if tracking_col < len(row) else None)
            picking_at = parse_excel_local_datetime(row[picking_col] if picking_col < len(row) else None)
            if not tracking_number or not picking_at:
                continue
            values[tracking_number].add(picking_at)
        unique = {tracking: next(iter(dates)) for tracking, dates in values.items() if len(dates) == 1}
        conflicts = {tracking: dates for tracking, dates in values.items() if len(dates) > 1}
        stats.tracking_index_values = len(values)
        stats.tracking_index_conflicts = len(conflicts)
        return unique, conflicts
    finally:
        workbook.close()


def group_rows(rows: list[SourceRow]) -> dict[tuple[str, str, str], OrderGroup]:
    groups: dict[tuple[str, str, str], OrderGroup] = {}
    for row in rows:
        if not row.account:
            continue
        key = (row.platform, row.account.account_id, row.order_no)
        group = groups.setdefault(
            key,
            OrderGroup(
                platform=row.platform,
                account_id=row.account.account_id,
                shop_name=row.account.display_name or row.shop_name,
                order_no=row.order_no,
            ),
        )
        group.rows.append(row)
    return groups


def build_excel_tracking_index(groups: dict[tuple[str, str, str], OrderGroup]) -> dict[tuple[str, str, str], set[tuple[str, str, str]]]:
    index: dict[tuple[str, str, str], set[tuple[str, str, str]]] = defaultdict(set)
    for key, group in groups.items():
        for tracking in group_tracking_values(group):
            index[(group.platform, group.account_id, tracking)].add(key)
    return index


def build_existing_indexes(db):
    rows = db.scalars(select(Order)).all()
    indexes: dict[str, dict[tuple, list[Order]]] = {
        "platform_order_id": defaultdict(list),
        "platform_order_no": defaultdict(list),
        "posting_number": defaultdict(list),
        "tracking_number": defaultdict(list),
        "unique": defaultdict(list),
    }
    for row in rows:
        platform = row.platform or ""
        account_id = row.account_id or ""
        if row.platform_order_id:
            indexes["platform_order_id"][(platform, account_id, row.platform_order_id)].append(row)
        if row.platform_order_no:
            indexes["platform_order_no"][(platform, account_id, row.platform_order_no)].append(row)
        if row.posting_number:
            indexes["posting_number"][(platform, account_id, row.posting_number)].append(row)
        if row.shipment_tracking_number:
            indexes["tracking_number"][(platform, account_id, row.shipment_tracking_number)].append(row)
        indexes["unique"][(platform, account_id, row.platform_order_id or "", row.posting_number or "")].append(row)
    return indexes


def group_tracking_values(group: OrderGroup) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in group.rows:
        tracking = clean(row.tracking_number)
        if tracking and tracking not in seen:
            seen.add(tracking)
            values.append(tracking)
    return values


def order_ids(rows: list[Order]) -> list[int]:
    return [row.id for row in rows if row.id is not None]


def find_existing_group(
    group: OrderGroup,
    indexes: dict[str, dict[tuple, list[Order]]],
    excel_tracking_index: dict[tuple[str, str, str], set[tuple[str, str, str]]],
) -> ExistingMatch:
    platform = group.platform
    account_id = group.account_id
    order_no = group.order_no
    candidates: list[tuple[str, list[Order]]] = []

    if platform == "ozon":
        # Ozon orders are stored at posting granularity. A parent order_id/order_number
        # can have multiple postings, so only the posting itself or the full unique
        # key may mark the group as existing.
        candidates.append(("posting_number", indexes["posting_number"].get((platform, account_id, order_no), [])))
        if group.ozon_info:
            candidates.append(
                (
                    "unique",
                    indexes["unique"].get((platform, account_id, group.ozon_info.order_id, order_no), []),
                )
            )
    elif platform == "joom_logistics":
        candidates.append(("platform_order_id", indexes["platform_order_id"].get((platform, account_id, order_no), [])))
        candidates.append(("platform_order_no", indexes["platform_order_no"].get((platform, account_id, order_no), [])))
    elif platform == "mercadolibre":
        candidates.append(("platform_order_id", indexes["platform_order_id"].get((platform, account_id, order_no), [])))
        candidates.append(("platform_order_no", indexes["platform_order_no"].get((platform, account_id, order_no), [])))
    elif platform in {"wildberries", "allegro"}:
        candidates.append(("platform_order_id", indexes["platform_order_id"].get((platform, account_id, order_no), [])))
        candidates.append(("posting_number", indexes["posting_number"].get((platform, account_id, order_no), [])))
    else:
        candidates.append(("platform_order_id", indexes["platform_order_id"].get((platform, account_id, order_no), [])))

    for reason, rows in candidates:
        if rows:
            return ExistingMatch(True, reason, order_ids(rows))

    tracking_values = group_tracking_values(group)
    if len(tracking_values) == 1:
        tracking = tracking_values[0]
        excel_keys = excel_tracking_index.get((platform, account_id, tracking), set())
        db_rows = indexes["tracking_number"].get((platform, account_id, tracking), [])
        if db_rows:
            if len(excel_keys) == 1 and len(db_rows) == 1:
                return ExistingMatch(True, "unique_tracking_number", order_ids(db_rows))
            return ExistingMatch(
                False,
                "",
                [],
                f"tracking_number_not_unique: excel_groups={len(excel_keys)}, db_orders={len(db_rows)}",
            )
    elif len(tracking_values) > 1:
        return ExistingMatch(False, "", [], f"group_has_multiple_tracking_numbers: {len(tracking_values)}")
    return ExistingMatch()


def fulfillment_from_group(group: OrderGroup) -> str:
    text = " ".join(clean(row.order_type) for row in group.rows if clean(row.order_type))
    if "海外仓" in text or "FBJ" in text.upper():
        if "FBJ" in text.upper():
            return "FBJ"
        return "OVERSEAS_WAREHOUSE"
    if group.platform == "wildberries" and "WB" in text.upper():
        return "OVERSEAS_WAREHOUSE"
    if group.platform == "allegro" and "PL" in text.upper():
        return "OVERSEAS_WAREHOUSE"
    return "FBS"


def ozon_fulfillment_type(raw_payload: dict[str, Any]) -> str:
    delivery_schema = clean(raw_payload.get("delivery_schema")).lower()
    if delivery_schema == "fbo":
        return "FBO"
    delivery_method = raw_payload.get("delivery_method") if isinstance(raw_payload.get("delivery_method"), dict) else {}
    analytics_data = raw_payload.get("analytics_data") if isinstance(raw_payload.get("analytics_data"), dict) else {}
    tpl_provider = clean(delivery_method.get("tpl_provider") or analytics_data.get("tpl_provider")).lower()
    if "fbp" in tpl_provider:
        return "FBP"
    return "FBS"


def group_amount(group: OrderGroup) -> str:
    total = Decimal("0")
    has_value = False
    for row in group.rows:
        price = decimal_or_none(row.unit_price)
        if price is None:
            continue
        total += price * Decimal(row.quantity or 1)
        has_value = True
    return format(total.normalize(), "f") if has_value else ""


def first_non_empty(values: list[str]) -> str:
    for value in values:
        if value:
            return value
    return ""


def unique_non_empty(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def group_currency(group: OrderGroup) -> str:
    return first_non_empty(unique_non_empty([row.currency for row in group.rows]))


def group_country_code(group: OrderGroup) -> str:
    return first_non_empty(unique_non_empty([row.country_code for row in group.rows])).upper()


def group_buyer_name(group: OrderGroup) -> str:
    return first_non_empty(unique_non_empty([row.buyer_name for row in group.rows]))


def group_platform_created_at(group: OrderGroup) -> datetime | None:
    values = [row.platform_created_at for row in group.rows if row.platform_created_at]
    return min(values) if values else None


def group_platform_deadline_at(group: OrderGroup) -> datetime | None:
    values = [row.platform_deadline_at for row in group.rows if row.platform_deadline_at]
    return min(values) if values else None


def group_dispatch_deadline_at(group: OrderGroup) -> datetime | None:
    values = [row.dispatch_deadline_at for row in group.rows if row.dispatch_deadline_at]
    return min(values) if values else None


def group_picking_at(group: OrderGroup) -> datetime | None:
    values = [row.picking_at for row in group.rows if row.picking_at]
    return min(values) if values else None


def group_shipped_at(group: OrderGroup) -> datetime | None:
    values = [row.shipped_at for row in group.rows if row.shipped_at]
    return min(values) if values else None


def group_logistics(group: OrderGroup) -> tuple[str, str]:
    selected = first_non_empty(unique_non_empty([row.buyer_selected_logistics for row in group.rows]))
    channel = first_non_empty(unique_non_empty([row.logistics_channel for row in group.rows]))
    return selected, channel


def grouped_items(group: OrderGroup) -> list[dict[str, Any]]:
    by_key: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in group.rows:
        key = (row.sku, row.product_name, row.unit_price, row.currency)
        item = by_key.setdefault(
            key,
            {
                "sku": row.sku,
                "platform_product_name": row.product_name,
                "quantity": 0,
                "unit_price": row.unit_price or None,
                "currency": row.currency,
                "row_numbers": [],
                "raw_rows": [],
            },
        )
        item["quantity"] += row.quantity or 1
        item["row_numbers"].append(row.row_no)
        item["raw_rows"].append(source_row_payload(row))
    return list(by_key.values())


def source_row_payload(row: SourceRow) -> dict[str, Any]:
    return {
        "row_no": row.row_no,
        "platform": row.platform_label,
        "shop_name": row.shop_name,
        "created_at": iso_or_none(row.platform_created_at),
        "order_no": row.order_no,
        "country_code": row.country_code,
        "buyer_name": row.buyer_name,
        "sku": row.sku,
        "quantity": row.quantity,
        "unit_price": row.unit_price,
        "currency": row.currency,
        "buyer_selected_logistics": row.buyer_selected_logistics,
        "logistics_channel": row.logistics_channel,
        "platform_deadline_at": iso_or_none(row.platform_deadline_at),
        "tracking_number": row.tracking_number,
        "dispatch_deadline_at": iso_or_none(row.dispatch_deadline_at),
        "product_name": row.product_name,
        "order_type": row.order_type,
        "excel_status": row.excel_status,
        "picking_at": iso_or_none(row.picking_at),
        "shipped_at": iso_or_none(row.shipped_at),
    }


def order_identity(group: OrderGroup) -> tuple[str, str, str, str]:
    if group.platform == "ozon":
        if not group.ozon_info:
            raise RuntimeError(f"Ozon order_id is missing for {group.order_no}")
        return group.ozon_info.order_id, group.ozon_info.order_number or order_no_base(group.order_no), group.order_no, group.ozon_info.status
    if group.platform == "mercadolibre":
        return group.order_no, group.order_no, "", "shipped"
    return group.order_no, group.order_no, group.order_no if group.platform in {"wildberries", "allegro"} else "", "shipped"


def build_raw_payload(group: OrderGroup, batch_id: str, source_file: Path, imported_at: datetime) -> dict[str, Any]:
    platform_order_id, platform_order_no, posting_number, platform_status = order_identity(group)
    country_code = group_country_code(group)
    selected_logistics, logistics_channel = group_logistics(group)
    tracking_values = group_tracking_values(group)
    picking_at = group_picking_at(group)
    shipped_at = group_shipped_at(group)
    platform_created_at = group_platform_created_at(group)
    platform_deadline_at = group_platform_deadline_at(group)
    dispatch_deadline_at = group_dispatch_deadline_at(group)
    products = [
        {
            "offer_id": item["sku"],
            "sku": item["sku"],
            "name": item["platform_product_name"],
            "quantity": item["quantity"],
            "price": item["unit_price"],
            "currency_code": item["currency"],
            "excel_row_numbers": item["row_numbers"],
        }
        for item in grouped_items(group)
    ]
    payload: dict[str, Any] = {
        "source": SOURCE_TAG,
        "source_type": SOURCE_TYPE,
        "excel_import_batch_id": batch_id,
        "excel_imported_at": iso_or_none(imported_at),
        "excel_file": str(source_file),
        "excel_sheet": ORDER_SHEET,
        "platform": group.platform,
        "shop_name": group.shop_name,
        "account_id": group.account_id,
        "id": platform_order_id,
        "order_id": platform_order_id,
        "order_number": platform_order_no,
        "posting_number": posting_number,
        "status": platform_status,
        "platform_status": platform_status,
        "created_at": iso_or_none(platform_created_at),
        "payment_at": iso_or_none(platform_created_at),
        "order_date": iso_or_none(platform_created_at),
        "country_code": country_code,
        "customer": {"name": group_buyer_name(group)},
        "buyer": {"name": group_buyer_name(group)},
        "buyer_selected_logistics": selected_logistics,
        "logistics_channel": logistics_channel,
        "shipment_tracking_number": tracking_values[0] if len(tracking_values) == 1 else "",
        "tracking_number": tracking_values[0] if len(tracking_values) == 1 else "",
        "order_amount": group_amount(group),
        "currency_code": group_currency(group),
        "currency": group_currency(group),
        "platform_handover_deadline": iso_or_none(platform_deadline_at),
        "shipping_deadline_at": iso_or_none(platform_deadline_at),
        "dispatch_deadline_at": iso_or_none(dispatch_deadline_at),
        "picking_at": iso_or_none(picking_at),
        "shipped_at": iso_or_none(shipped_at),
        "products": products,
        "excel_rows": [source_row_payload(row) for row in group.rows],
        "excel_order_type_values": unique_non_empty([row.order_type for row in group.rows]),
        "excel_status_values": unique_non_empty([row.excel_status for row in group.rows]),
        "excel_tracking_values": tracking_values,
    }
    if group.ozon_info:
        payload["ozon_api_payload"] = group.ozon_info.raw_payload
    return payload


def add_imported_order(db, group: OrderGroup, batch_id: str, source_file: Path, imported_at: datetime) -> tuple[int, int, int]:
    first_row = group.rows[0]
    platform_order_id, platform_order_no, posting_number, platform_status = order_identity(group)
    country_code = group_country_code(group)
    selected_logistics, logistics_channel = group_logistics(group)
    tracking_values = group_tracking_values(group)
    tracking_number = tracking_values[0] if len(tracking_values) == 1 else ""
    picking_at = group_picking_at(group)
    platform_created_at = group_platform_created_at(group)
    platform_deadline_at = group_platform_deadline_at(group)
    dispatch_deadline_at = group_dispatch_deadline_at(group)
    shipped_at = group_shipped_at(group)
    raw_payload = build_raw_payload(group, batch_id, source_file, imported_at)
    fulfillment_type = ozon_fulfillment_type(group.ozon_info.raw_payload) if group.ozon_info else fulfillment_from_group(group)
    fulfillment_type = normalize_fulfillment_type(fulfillment_type)

    order = Order(
        tenant_id=get_settings().default_tenant_id,
        platform=group.platform,
        account_id=group.account_id,
        shop_id=group.account_id,
        shop_name=group.shop_name,
        platform_order_id=platform_order_id,
        platform_order_no=platform_order_no or None,
        posting_number=posting_number,
        buyer_name=group_buyer_name(group) or None,
        platform_status=platform_status or excel_status_to_platform_status(first_row.excel_status),
        biz_status=IMPORTED_BIZ_STATUS,
        local_status=IMPORTED_LOCAL_STATUS,
        fulfillment_type=fulfillment_type,
        is_overseas_warehouse=infer_is_overseas_warehouse(group.platform, fulfillment_type, raw_payload),
        platform_handover_deadline=platform_deadline_at,
        platform_created_at=platform_created_at,
        country_code=country_code or None,
        country_name_cn=country_name_cn(country_code) if country_code else None,
        buyer_selected_logistics=selected_logistics or logistics_channel or None,
        order_amount=group_amount(group) or None,
        currency=group_currency(group) or None,
        payment_at=platform_created_at,
        shipping_deadline_at=platform_deadline_at,
        dispatch_deadline_at=dispatch_deadline_at,
        shipment_tracking_number=tracking_number or None,
        picking_at=picking_at,
        marked_shipped_at=shipped_at,
        handover_at=shipped_at,
        shipped_at=shipped_at,
        raw_payload=raw_payload,
        last_api_payload=group.ozon_info.raw_payload if group.ozon_info else {},
        created_at=imported_at,
        updated_at=imported_at,
    )
    db.add(order)
    db.flush()

    item_count = 0
    for item in grouped_items(group):
        db.add(
            OrderItem(
                order_id=order.id,
                sku=item["sku"],
                platform_product_name=item["platform_product_name"],
                quantity=item["quantity"],
                unit_price=item["unit_price"],
                currency=item["currency"],
                raw_payload={
                    "source": SOURCE_TAG,
                    "excel_import_batch_id": batch_id,
                    "excel_row_numbers": item["row_numbers"],
                    "rows": item["raw_rows"],
                },
                created_at=imported_at,
                updated_at=imported_at,
            )
        )
        item_count += 1

    shipment_count = 0
    if tracking_number:
        db.add(
            Shipment(
                order_id=order.id,
                platform_shipment_id=posting_number or platform_order_id,
                tracking_number=tracking_number,
                carrier=logistics_channel or selected_logistics or group.platform,
                status="shipped",
                created_at=imported_at,
            )
        )
        shipment_count += 1

    db.add(
        OrderOperationLog(
            order_id=order.id,
            operation_type="excel_import",
            operation_attribute="历史订单导入",
            description=f"从 {source_file.name} 的 {ORDER_SHEET} 页签导入，批次 {batch_id}，订单状态设置为已发货",
            operator=SYSTEM_OPERATOR,
            source=ORDER_LOG_HISTORY_SOURCE,
            event_key=f"excel_import:{batch_id}:{group.platform}:{group.account_id}:{group.order_no}",
            extra={
                "source": SOURCE_TAG,
                "source_type": SOURCE_TYPE,
                "batch_id": batch_id,
                "file": str(source_file),
                "sheet": ORDER_SHEET,
                "row_numbers": [row.row_no for row in group.rows],
                "order_no": group.order_no,
                "tracking_numbers": tracking_values,
            },
            operated_at=imported_at,
            created_at=imported_at,
        )
    )
    return item_count, shipment_count, 1


def ozon_timestamp(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    else:
        value = value.astimezone(UTC)
    return value.replace(microsecond=0).isoformat().replace("+00:00", "Z")


class OzonResolver:
    def __init__(self, accounts: dict[str, PlatformAccount], *, timeout: float = 60, chunk_days: int = 30) -> None:
        self.accounts = accounts
        self.timeout = timeout
        self.chunk_days = chunk_days
        self.credentials = get_credential_manager()

    def resolve(self, groups: list[OrderGroup]) -> tuple[dict[tuple[str, str], OzonInfo], dict[tuple[str, str], str]]:
        by_account: dict[str, list[OrderGroup]] = defaultdict(list)
        for group in groups:
            by_account[group.account_id].append(group)

        resolved: dict[tuple[str, str], OzonInfo] = {}
        failures: dict[tuple[str, str], str] = {}
        for account_id, account_groups in by_account.items():
            account = self.accounts.get(account_id)
            if not account:
                for group in account_groups:
                    failures[(group.account_id, group.order_no)] = "Ozon account not found"
                continue
            account_resolved, account_failures = self._resolve_account(account, account_groups)
            resolved.update(account_resolved)
            failures.update(account_failures)
        return resolved, failures

    def _resolve_account(self, account: PlatformAccount, groups: list[OrderGroup]) -> tuple[dict[tuple[str, str], OzonInfo], dict[tuple[str, str], str]]:
        creds = self.credentials.decrypt_credentials(account.encrypted_credentials)
        settings = account.settings or {}
        base_url = clean(settings.get("base_url")) or "https://api-seller.ozon.ru"
        headers = {
            "Client-Id": clean(creds.get("client_id")),
            "Api-Key": clean(creds.get("api_key")),
            "Content-Type": "application/json",
        }
        targets = {group.order_no for group in groups}
        min_date = min((group_platform_created_at(group) for group in groups if group_platform_created_at(group)), default=utc_now())
        max_date = max((group_platform_created_at(group) for group in groups if group_platform_created_at(group)), default=utc_now())
        since = (min_date - timedelta(days=2)).replace(tzinfo=UTC)
        until = (max_date + timedelta(days=2)).replace(tzinfo=UTC)
        now = datetime.now(UTC)
        if until > now:
            until = now

        resolved: dict[tuple[str, str], OzonInfo] = {}
        failures: dict[tuple[str, str], str] = {}
        with httpx.Client(timeout=self.timeout) as client:
            chunk_start = since
            while chunk_start < until and targets:
                chunk_end = min(chunk_start + timedelta(days=self.chunk_days), until)
                self._resolve_by_list(
                    client,
                    base_url,
                    headers,
                    account.account_id,
                    targets,
                    chunk_start,
                    chunk_end,
                    resolved,
                )
                chunk_start = chunk_end
            for posting_number in sorted(targets):
                key = (account.account_id, posting_number)
                try:
                    info = self._resolve_by_get(client, base_url, headers, posting_number)
                    resolved[key] = info
                except Exception as exc:
                    failures[key] = str(exc)[:500]
        return resolved, failures

    def _post(self, client: httpx.Client, base_url: str, headers: dict[str, str], path: str, payload: dict[str, Any]) -> dict[str, Any]:
        response = client.post(f"{base_url.rstrip('/')}{path}", headers=headers, json=payload)
        response.raise_for_status()
        data = response.json()
        if not isinstance(data, dict):
            raise RuntimeError(f"Ozon returned non-object response for {path}")
        return data

    def _resolve_by_list(
        self,
        client: httpx.Client,
        base_url: str,
        headers: dict[str, str],
        account_id: str,
        targets: set[str],
        since: datetime,
        until: datetime,
        resolved: dict[tuple[str, str], OzonInfo],
    ) -> None:
        cursor = ""
        while targets:
            payload: dict[str, Any] = {
                "dir": "ASC",
                "filter": {"since": ozon_timestamp(since), "to": ozon_timestamp(until)},
                "limit": 100,
                "with": {"analytics_data": True, "barcodes": False, "financial_data": True},
            }
            if cursor:
                payload["cursor"] = cursor
            data = self._post(client, base_url, headers, "/v4/posting/fbs/list", payload)
            result = data.get("result") if isinstance(data.get("result"), dict) else {}
            postings = data.get("postings") or result.get("postings") or []
            if not isinstance(postings, list) or not postings:
                break
            for posting in postings:
                if not isinstance(posting, dict):
                    continue
                posting_number = clean(posting.get("posting_number"))
                if posting_number not in targets:
                    continue
                order_id = clean(posting.get("order_id"))
                if not order_id:
                    continue
                info = OzonInfo(
                    posting_number=posting_number,
                    order_id=order_id,
                    order_number=clean(posting.get("order_number")) or order_no_base(posting_number),
                    status=clean(posting.get("status")) or "shipped",
                    raw_payload=posting,
                )
                resolved[(account_id, posting_number)] = info
                targets.discard(posting_number)
            cursor = clean(data.get("cursor") or result.get("cursor"))
            has_next = bool(data.get("has_next", result.get("has_next", False)))
            if not cursor or not has_next:
                break

    def _resolve_by_get(self, client: httpx.Client, base_url: str, headers: dict[str, str], posting_number: str) -> OzonInfo:
        payload = {
            "posting_number": posting_number,
            "with": {"analytics_data": True, "barcodes": False, "financial_data": True},
        }
        data = self._post(client, base_url, headers, "/v3/posting/fbs/get", payload)
        result = data.get("result") if isinstance(data.get("result"), dict) else data
        order_id = clean(result.get("order_id"))
        if not order_id:
            raise RuntimeError("Ozon response did not contain order_id")
        return OzonInfo(
            posting_number=clean(result.get("posting_number")) or posting_number,
            order_id=order_id,
            order_number=clean(result.get("order_number")) or order_no_base(posting_number),
            status=clean(result.get("status")) or "shipped",
            raw_payload=result,
        )


def write_reports(
    report_dir: Path,
    batch_id: str,
    groups: list[OrderGroup],
    *,
    apply: bool,
) -> None:
    report_dir.mkdir(parents=True, exist_ok=True)
    prefix = report_dir / f"excel_order_import_{batch_id}"
    with (prefix.with_suffix(".groups.csv")).open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow([
            "batch_id",
            "apply",
            "platform",
            "account_id",
            "shop_name",
            "order_no",
            "row_count",
            "item_rows",
            "existing",
            "existing_reason",
            "existing_order_ids",
            "tracking_conflict",
            "ozon_order_id",
            "ozon_error",
            "tracking_numbers",
            "row_numbers",
        ])
        for group in groups:
            writer.writerow([
                batch_id,
                int(apply),
                group.platform,
                group.account_id,
                group.shop_name,
                group.order_no,
                len(group.rows),
                len(grouped_items(group)),
                int(group.existing.exists),
                group.existing.reason,
                " ".join(str(value) for value in group.existing.order_ids),
                group.existing.tracking_conflict,
                group.ozon_info.order_id if group.ozon_info else "",
                group.ozon_error,
                " ".join(group_tracking_values(group)),
                " ".join(str(row.row_no) for row in group.rows),
            ])


def excel_order_imported(row: Order) -> bool:
    raw_payload = row.raw_payload if isinstance(row.raw_payload, dict) else {}
    return raw_payload.get("source") == SOURCE_TAG or raw_payload.get("source_type") == SOURCE_TYPE


def existing_order_rows_for_group(
    db,
    group: OrderGroup,
    indexes: dict[str, dict[tuple, list[Order]]],
    excel_tracking_index: dict[tuple[str, str, str], set[tuple[str, str, str]]],
) -> list[Order]:
    match = find_existing_group(group, indexes, excel_tracking_index)
    if not match.exists or not match.order_ids:
        return []
    rows = db.scalars(select(Order).where(Order.id.in_(match.order_ids)).order_by(Order.id)).all()
    return list(rows)


def add_picking_date_report_row(writer: csv.writer, batch_id: str, group: OrderGroup, picking_at: datetime | None, status: str, orders: list[Order]) -> None:
    writer.writerow(
        [
            batch_id,
            group.platform,
            group.account_id,
            group.shop_name,
            group.order_no,
            iso_or_none(picking_at),
            status,
            " ".join(str(row.id) for row in orders if row.id is not None),
            " ".join(iso_or_none(row.picking_at) or "" for row in orders),
            " ".join(str(row.row_no) for row in group.rows),
        ]
    )


def update_picking_dates(args: argparse.Namespace) -> ImportStats:
    stats = ImportStats()
    source_file = Path(args.file).expanduser().resolve()
    batch_id = args.batch_id or f"excel-picking-{utc_now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    report_dir = Path(args.report_dir).expanduser().resolve()
    report_dir.mkdir(parents=True, exist_ok=True)

    source_rows = read_source_rows(source_file, stats)
    groups_by_key = group_rows(source_rows)
    groups = list(groups_by_key.values())
    excel_tracking_index = build_excel_tracking_index(groups_by_key)
    stats.groups = len(groups)
    stats.picking_date_groups = sum(1 for group in groups if group_picking_at(group))
    imported_at = utc_now()

    report_file = report_dir / f"excel_order_picking_update_{batch_id}.csv"
    with SessionLocal() as db, report_file.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "batch_id",
                "platform",
                "account_id",
                "shop_name",
                "order_no",
                "excel_picking_at",
                "status",
                "order_ids",
                "old_picking_at_values",
                "row_numbers",
            ]
        )
        existing_indexes = build_existing_indexes(db)
        for group in groups:
            picking_at = group_picking_at(group)
            if not picking_at:
                stats.skipped_missing_picking_dates += 1
                add_picking_date_report_row(writer, batch_id, group, picking_at, "missing_excel_picking_date", [])
                continue

            rows = existing_order_rows_for_group(db, group, existing_indexes, excel_tracking_index)
            if not rows:
                stats.skipped_missing_existing_orders += 1
                add_picking_date_report_row(writer, batch_id, group, picking_at, "existing_order_not_found", [])
                continue
            if len(rows) != 1:
                stats.skipped_ambiguous_existing_orders += 1
                add_picking_date_report_row(writer, batch_id, group, picking_at, "ambiguous_existing_orders", rows)
                continue

            order = rows[0]
            if not excel_order_imported(order):
                stats.skipped_non_excel_import_orders += 1
                add_picking_date_report_row(writer, batch_id, group, picking_at, "not_excel_imported_order", rows)
                continue
            picking_at_payload = iso_or_none(picking_at)
            raw_payload = dict(order.raw_payload) if isinstance(order.raw_payload, dict) else {}
            if order.picking_at == picking_at and raw_payload.get("picking_at") == picking_at_payload:
                stats.unchanged_picking_dates += 1
                add_picking_date_report_row(writer, batch_id, group, picking_at, "unchanged", rows)
                continue

            add_picking_date_report_row(writer, batch_id, group, picking_at, "updated" if args.apply else "would_update", rows)
            if args.apply:
                raw_payload["picking_at"] = picking_at_payload
                order.raw_payload = raw_payload
                order.picking_at = picking_at
                order.updated_at = imported_at
            stats.updated_picking_dates += 1

        if args.apply:
            db.commit()
        else:
            db.rollback()

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"Mode: {mode}")
    print(f"Batch: {batch_id}")
    print(f"Source rows: {stats.source_rows}")
    print(f"Parsed rows: {stats.parsed_rows}")
    print(f"Candidate groups: {stats.groups}")
    print(f"Groups with picking date: {stats.picking_date_groups}")
    print(f"Updated picking dates: {stats.updated_picking_dates}")
    print(f"Unchanged picking dates: {stats.unchanged_picking_dates}")
    print(f"Skipped missing picking dates: {stats.skipped_missing_picking_dates}")
    print(f"Skipped missing existing orders: {stats.skipped_missing_existing_orders}")
    print(f"Skipped ambiguous existing orders: {stats.skipped_ambiguous_existing_orders}")
    print(f"Skipped non-Excel-import orders: {stats.skipped_non_excel_import_orders}")
    print(f"Report: {report_file}")
    return stats


def order_tracking_candidates(order: Order) -> list[tuple[str, str]]:
    result: list[tuple[str, str]] = []
    seen: set[str] = set()
    for field in ORDER_NUMBER_FIELDS:
        value = clean(getattr(order, field, None))
        if value and value not in seen:
            seen.add(value)
            result.append((field, value))
    return result


def add_tracking_update_report_row(
    writer: csv.writer,
    batch_id: str,
    order: Order,
    status: str,
    *,
    matched_field: str = "",
    matched_tracking: str = "",
    picking_at: datetime | None = None,
    reason: str = "",
) -> None:
    writer.writerow(
        [
            batch_id,
            order.id,
            order.platform,
            order.account_id,
            order.shop_name or "",
            order.biz_status or "",
            order.shipment_tracking_number or "",
            order.posting_number or "",
            order.platform_order_no or "",
            order.platform_order_id or "",
            matched_field,
            matched_tracking,
            iso_or_none(picking_at),
            status,
            reason,
        ]
    )


def update_missing_picking_dates_by_tracking(args: argparse.Namespace) -> ImportStats:
    stats = ImportStats()
    source_file = Path(args.file).expanduser().resolve()
    batch_id = args.batch_id or f"excel-tracking-picking-{utc_now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    report_dir = Path(args.report_dir).expanduser().resolve()
    report_dir.mkdir(parents=True, exist_ok=True)
    tracking_index, tracking_conflicts = read_tracking_picking_index(source_file, stats)
    imported_at = utc_now()

    report_file = report_dir / f"excel_tracking_picking_update_{batch_id}.csv"
    with SessionLocal() as db, report_file.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "batch_id",
                "order_id",
                "platform",
                "account_id",
                "shop_name",
                "biz_status",
                "shipment_tracking_number",
                "posting_number",
                "platform_order_no",
                "platform_order_id",
                "matched_field",
                "matched_tracking",
                "excel_picking_at",
                "status",
                "reason",
            ]
        )
        orders = db.scalars(select(Order).where(Order.picking_at.is_(None)).order_by(Order.id)).all()
        stats.groups = len(orders)
        for order in orders:
            candidates = order_tracking_candidates(order)
            if not candidates:
                stats.skipped_missing_tracking_numbers += 1
                add_tracking_update_report_row(writer, batch_id, order, "missing_order_tracking")
                continue

            matches: list[tuple[str, str, datetime]] = []
            conflict_values: list[str] = []
            for field, tracking in candidates:
                if tracking in tracking_index:
                    matches.append((field, tracking, tracking_index[tracking]))
                if tracking in tracking_conflicts:
                    conflict_values.append(tracking)

            match_dates = {picking_at for _field, _tracking, picking_at in matches}
            if len(match_dates) > 1:
                stats.skipped_conflicting_excel_tracking += 1
                add_tracking_update_report_row(
                    writer,
                    batch_id,
                    order,
                    "conflicting_excel_tracking",
                    reason="multiple candidate tracking numbers map to different picking dates",
                )
                continue
            if not matches:
                if conflict_values:
                    stats.skipped_conflicting_excel_tracking += 1
                    add_tracking_update_report_row(
                        writer,
                        batch_id,
                        order,
                        "conflicting_excel_tracking",
                        reason=" ".join(conflict_values),
                    )
                else:
                    stats.skipped_missing_excel_tracking += 1
                    add_tracking_update_report_row(writer, batch_id, order, "excel_tracking_not_found")
                continue

            matched_field, matched_tracking, picking_at = matches[0]
            if order.picking_at:
                stats.skipped_existing_picking_dates += 1
                add_tracking_update_report_row(
                    writer,
                    batch_id,
                    order,
                    "already_has_picking_date",
                    matched_field=matched_field,
                    matched_tracking=matched_tracking,
                    picking_at=picking_at,
                )
                continue

            add_tracking_update_report_row(
                writer,
                batch_id,
                order,
                "updated" if args.apply else "would_update",
                matched_field=matched_field,
                matched_tracking=matched_tracking,
                picking_at=picking_at,
            )
            if args.apply:
                raw_payload = dict(order.raw_payload) if isinstance(order.raw_payload, dict) else {}
                raw_payload["picking_at"] = iso_or_none(picking_at)
                raw_payload["picking_at_source"] = "excel_tracking_match"
                raw_payload["picking_at_tracking_number"] = matched_tracking
                order.raw_payload = raw_payload
                order.picking_at = picking_at
                order.updated_at = imported_at
            stats.updated_picking_dates += 1

        if args.apply:
            db.commit()
        else:
            db.rollback()

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"Mode: {mode}")
    print(f"Batch: {batch_id}")
    print(f"Excel tracking values: {stats.tracking_index_values}")
    print(f"Excel tracking conflicts: {stats.tracking_index_conflicts}")
    print(f"Orders missing picking date: {stats.groups}")
    print(f"Updated picking dates: {stats.updated_picking_dates}")
    print(f"Skipped existing picking dates: {stats.skipped_existing_picking_dates}")
    print(f"Skipped missing tracking numbers: {stats.skipped_missing_tracking_numbers}")
    print(f"Skipped missing Excel tracking: {stats.skipped_missing_excel_tracking}")
    print(f"Skipped conflicting Excel tracking: {stats.skipped_conflicting_excel_tracking}")
    print(f"Report: {report_file}")
    return stats


def print_stats(stats: ImportStats, report_dir: Path, batch_id: str, apply: bool) -> None:
    mode = "APPLY" if apply else "DRY-RUN"
    print(f"Mode: {mode}")
    print(f"Batch: {batch_id}")
    print(f"Source rows: {stats.source_rows}")
    print(f"Parsed rows: {stats.parsed_rows}")
    print(f"Unresolved rows skipped: {stats.unresolved_rows}")
    print(f"Candidate groups: {stats.groups}")
    print(f"Skipped existing groups: {stats.skipped_existing_groups}")
    print(f"Skipped by unique tracking: {stats.skipped_existing_by_tracking}")
    print(f"Tracking conflict groups: {stats.tracking_conflict_groups}")
    print(f"Ozon groups needing API id: {stats.ozon_target_groups}")
    print(f"Ozon resolved groups: {stats.ozon_resolved_groups}")
    print(f"Ozon failed groups: {stats.ozon_failed_groups}")
    print(f"Imported orders: {stats.imported_orders}")
    print(f"Imported order items: {stats.imported_items}")
    print(f"Imported shipments: {stats.imported_shipments}")
    print(f"Imported operation logs: {stats.imported_logs}")
    print(f"Report: {report_dir / f'excel_order_import_{batch_id}.groups.csv'}")


def run_import(args: argparse.Namespace) -> ImportStats:
    stats = ImportStats()
    source_file = Path(args.file).expanduser().resolve()
    batch_id = args.batch_id or f"excel-{utc_now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    report_dir = Path(args.report_dir).expanduser().resolve()

    source_rows = read_source_rows(source_file, stats)
    groups_by_key = group_rows(source_rows)
    stats.groups = len(groups_by_key)
    excel_tracking_index = build_excel_tracking_index(groups_by_key)
    imported_at = utc_now()

    with SessionLocal() as db:
        existing_indexes = build_existing_indexes(db)
        groups = list(groups_by_key.values())
        for group in groups:
            group.existing = find_existing_group(group, existing_indexes, excel_tracking_index)
            if group.existing.exists:
                stats.skipped_existing_groups += 1
                if group.existing.reason == "unique_tracking_number":
                    stats.skipped_existing_by_tracking += 1
            if group.existing.tracking_conflict:
                stats.tracking_conflict_groups += 1

        ozon_targets = [group for group in groups if group.platform == "ozon" and not group.existing.exists]
        if args.limit and args.limit > 0:
            limited_keys = {group.key for group in groups[: args.limit]}
            ozon_targets = [group for group in ozon_targets if group.key in limited_keys]
        stats.ozon_target_groups = len(ozon_targets)
        if ozon_targets and not args.skip_ozon_api:
            ozon_accounts = {
                account.account_id: account
                for account in db.scalars(select(PlatformAccount).where(PlatformAccount.platform == "ozon")).all()
            }
            resolver = OzonResolver(ozon_accounts, timeout=args.ozon_timeout, chunk_days=args.ozon_chunk_days)
            resolved, failures = resolver.resolve(ozon_targets)
            for group in ozon_targets:
                group.ozon_info = resolved.get((group.account_id, group.order_no))
                group.ozon_error = failures.get((group.account_id, group.order_no), "")
            stats.ozon_resolved_groups = sum(1 for group in ozon_targets if group.ozon_info)
            stats.ozon_failed_groups = sum(1 for group in ozon_targets if not group.ozon_info)
        elif ozon_targets:
            for group in ozon_targets:
                group.ozon_error = "Ozon API lookup skipped"
            stats.ozon_failed_groups = len(ozon_targets)

        # After Ozon API resolution, run exact matching again for Ozon real order IDs.
        for group in ozon_targets:
            if group.existing.exists or not group.ozon_info:
                continue
            group.existing = find_existing_group(group, existing_indexes, excel_tracking_index)
            if group.existing.exists:
                stats.skipped_existing_groups += 1
                if group.existing.reason == "unique_tracking_number":
                    stats.skipped_existing_by_tracking += 1

        importable = [group for group in groups if not group.existing.exists]
        if args.limit and args.limit > 0:
            importable = importable[: args.limit]

        for group in importable:
            if group.platform == "ozon" and not group.ozon_info:
                continue
            if args.apply:
                item_count, shipment_count, log_count = add_imported_order(db, group, batch_id, source_file, imported_at)
                stats.imported_orders += 1
                stats.imported_items += item_count
                stats.imported_shipments += shipment_count
                stats.imported_logs += log_count
            else:
                stats.imported_orders += 1
                stats.imported_items += len(grouped_items(group))
                stats.imported_shipments += 1 if len(group_tracking_values(group)) == 1 else 0
                stats.imported_logs += 1

        if args.apply:
            db.commit()
        else:
            db.rollback()

    write_reports(report_dir, batch_id, list(groups_by_key.values()), apply=args.apply)
    print_stats(stats, report_dir, batch_id, args.apply)
    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import missing historical orders from Order follow up 2026.xlsx")
    parser.add_argument("--file", type=Path, default=DEFAULT_EXCEL, help="Excel file path")
    parser.add_argument("--batch-id", default="", help="Import batch id. Defaults to generated id.")
    parser.add_argument("--report-dir", type=Path, default=ROOT / "outputs" / "excel_order_import", help="Directory for import reports")
    parser.add_argument("--apply", action="store_true", help="Write new orders to the database. Default is dry-run.")
    parser.add_argument(
        "--update-picking-dates",
        action="store_true",
        help="Update existing Excel-imported orders from the Excel 配货日 column instead of importing new orders.",
    )
    parser.add_argument(
        "--update-missing-picking-dates-by-tracking",
        action="store_true",
        help="Update orders with missing picking dates by matching order tracking numbers to the Excel 货运单号 column.",
    )
    parser.add_argument("--skip-ozon-api", action="store_true", help="Skip Ozon API order_id lookup; Ozon rows will not be imported.")
    parser.add_argument("--ozon-timeout", type=float, default=60.0, help="Ozon API timeout in seconds")
    parser.add_argument("--ozon-chunk-days", type=int, default=30, help="Date range chunk size for Ozon listing")
    parser.add_argument("--limit", type=int, default=0, help="Limit importable groups for testing")
    return parser.parse_args()


if __name__ == "__main__":
    parsed_args = parse_args()
    if parsed_args.update_missing_picking_dates_by_tracking:
        update_missing_picking_dates_by_tracking(parsed_args)
    elif parsed_args.update_picking_dates:
        update_picking_dates(parsed_args)
    else:
        run_import(parsed_args)
