# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
import sys

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import joinedload

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.database import SessionLocal
from app.models import LocalUser, PlatformAccount
from app.product_models import Product, ProductShopMapping


DEFAULT_EXCEL = ROOT / "demo_data" / "product_catalog.xlsx"
PRODUCT_SHEET = "\u4ea7\u54c1\u76ee\u5f55"

COL_NAME = 0
COL_COST = 11
COL_WEIGHT = 12
COL_STOCK = 13
COL_BUYER = 14
STATIC_COLUMNS = {COL_NAME, 10, COL_COST, COL_WEIGHT, COL_STOCK, COL_BUYER}

SHOP_COLUMN_ALIASES = {
    "OZON DEMO SHOP A": ("OZON DEMO SHOP A", "100001"),
    "OZON DEMO SHOP B": ("Ozon Demo Shop B", "100002"),
    "JOOM DEMO SHOP": ("Joom Demo Shop", "JOOM-DEMO-001"),
    "JOOM DEMO SHOP SKU": ("Joom Demo Shop", "J001"),
    "MERCADO": ("Mercado Demo Shop", "mercado-demo"),
    "Allegro": ("Allegro", "Allegro Demo Shop", "allegro0002"),
    "FRUUGO": ("FRUUGO", "Fruugo", "fruugo-demo"),
    "Wildberries": ("Wildberries", "WB DEMO SHOP CN", "Wildberries FBS", "wildberries-demo"),
    "FALABELLA": ("FALABELLA", "Falabella", "falabella-demo"),
}

BUYER_ALIASES = {
    "Tony": "tony",
    "\u5e93\u5b58": "\u4ed3\u5e93",
}


def utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


@dataclass
class FieldValue:
    row_no: int
    value: object


@dataclass
class MappingValue:
    row_no: int
    sequence: int
    product_name: str
    shop_id: int
    shop_name: str
    sku: str


@dataclass
class ProductGroup:
    name: str
    first_row: int
    fields: dict[str, list[FieldValue]] = field(default_factory=lambda: defaultdict(list))
    mappings: dict[int, list[str]] = field(default_factory=lambda: defaultdict(list))


@dataclass
class ImportStats:
    rows_seen: int = 0
    empty_name_rows: int = 0
    unique_product_names: int = 0
    duplicate_product_names: int = 0
    products_created: int = 0
    products_updated: int = 0
    field_updates: int = 0
    field_conflicts: int = 0
    buyer_unresolved: int = 0
    mappings_created: int = 0
    mappings_existing: int = 0
    mappings_deleted: int = 0
    mappings_reassigned: int = 0
    mappings_conflicted: int = 0
    source_mapping_entries: int = 0
    source_mapping_ignored_by_later_row: int = 0
    unmapped_shop_values: int = 0
    no_mapping_rows: int = 0
    shortage_cost_as_zero: int = 0


def clean(value) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    return "" if text_value.upper() == "#N/A" else text_value


def decimal_value(value, *, zero_on_shortage: bool = False) -> tuple[Decimal | None, bool]:
    text_value = clean(value)
    if not text_value:
        return None, False
    if text_value in {"\u65e0", "\u6682\u65e0"}:
        return None, False
    if zero_on_shortage and ("\u7f3a\u8d27" in text_value or text_value.lower() in {"n/a", "none", "null"}):
        return Decimal("0"), True
    try:
        return Decimal(text_value), False
    except InvalidOperation as exc:
        raise ValueError(f"\u6570\u5b57\u683c\u5f0f\u9519\u8bef: {text_value}") from exc


def int_value(value) -> int | None:
    text_value = clean(value)
    if not text_value:
        return None
    if text_value in {"\u65e0", "\u6682\u65e0"}:
        return None
    try:
        return int(float(text_value))
    except ValueError as exc:
        raise ValueError(f"\u6574\u6570\u683c\u5f0f\u9519\u8bef: {text_value}") from exc


def unique_values(values: list[FieldValue]) -> list[object]:
    result: list[object] = []
    seen: set[str] = set()
    for item in values:
        key = str(item.value)
        if key in seen:
            continue
        seen.add(key)
        result.append(item.value)
    return result


def selected_field_value(values: list[FieldValue], policy: str) -> object | None:
    if not values:
        return None
    return values[0].value if policy == "first" else values[-1].value


def build_shop_lookup(db) -> dict[str, PlatformAccount]:
    rows = db.scalars(select(PlatformAccount).order_by(PlatformAccount.id)).all()
    lookup: dict[str, PlatformAccount] = {}
    for shop in rows:
        for value in {shop.display_name, shop.account_id}:
            key = clean(value).lower()
            if key:
                lookup[key] = shop
    return lookup


def build_buyer_lookup(db) -> dict[str, LocalUser]:
    rows = db.scalars(select(LocalUser).where(LocalUser.enabled == True).order_by(LocalUser.id)).all()
    lookup: dict[str, LocalUser] = {}
    for user in rows:
        if clean(user.username).lower() == "admin":
            continue
        for value in {user.username, user.display_name or ""}:
            key = clean(value)
            if key:
                lookup[key] = user
                lookup[key.lower()] = user
    for source, target in BUYER_ALIASES.items():
        user = lookup.get(target) or lookup.get(target.lower())
        if user:
            lookup[source] = user
            lookup[source.lower()] = user
    return lookup


def resolve_shop_columns(headers: list[str], shop_lookup: dict[str, PlatformAccount]) -> tuple[dict[int, PlatformAccount], list[str]]:
    shop_columns: dict[int, PlatformAccount] = {}
    unresolved: list[str] = []
    for index, header in enumerate(headers):
        title = clean(header)
        if not title or index in STATIC_COLUMNS:
            continue
        aliases = SHOP_COLUMN_ALIASES.get(title, (title,))
        shop = next((shop_lookup.get(clean(alias).lower()) for alias in aliases if clean(alias).lower() in shop_lookup), None)
        if shop:
            shop_columns[index] = shop
        else:
            unresolved.append(title)
    return shop_columns, unresolved


def generate_product_code(max_code: str | None, used_codes: set[str]) -> str:
    next_number = 1
    if max_code and len(max_code) == 9 and max_code[1:].isdigit():
        next_number = int(max_code[1:]) + 1
    while True:
        product_code = f"P{next_number:08d}"
        if product_code not in used_codes:
            used_codes.add(product_code)
            return product_code
        next_number += 1


def add_field(group: ProductGroup, field_name: str, row_no: int, value) -> None:
    if value is not None:
        group.fields[field_name].append(FieldValue(row_no=row_no, value=value))


def add_mapping(group: ProductGroup, shop_id: int, sku: str) -> bool:
    values = group.mappings[shop_id]
    if sku in values:
        return False
    values.append(sku)
    return True


def apply_latest_sku_owner(groups: dict[str, ProductGroup], raw_mappings: list[MappingValue], stats: ImportStats, events: list[dict]) -> None:
    by_shop_sku: dict[tuple[int, str], list[MappingValue]] = defaultdict(list)
    for mapping in raw_mappings:
        by_shop_sku[(mapping.shop_id, mapping.sku)].append(mapping)

    for (shop_id, sku), mappings in by_shop_sku.items():
        product_names = {mapping.product_name for mapping in mappings}
        winner = max(mappings, key=lambda item: (item.row_no, item.sequence))
        for mapping in mappings:
            if len(product_names) > 1 and mapping.product_name != winner.product_name:
                stats.source_mapping_ignored_by_later_row += 1
                events.append(
                    {
                        "type": "source_mapping_ignored_by_later_row",
                        "row": mapping.row_no,
                        "shop": mapping.shop_name,
                        "sku": mapping.sku,
                        "product": mapping.product_name,
                        "message": f"kept row {winner.row_no}: {winner.product_name}",
                    }
                )
                continue
            add_mapping(groups[mapping.product_name], shop_id, sku)


def read_excel_groups(path: Path, stats: ImportStats) -> tuple[dict[str, ProductGroup], dict[int, PlatformAccount], list[str], list[dict]]:
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[PRODUCT_SHEET] if PRODUCT_SHEET in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]

        with SessionLocal() as db:
            shop_columns, unresolved_headers = resolve_shop_columns(headers, build_shop_lookup(db))
            buyer_lookup = build_buyer_lookup(db)

        groups: dict[str, ProductGroup] = {}
        errors: list[dict] = []
        raw_mappings: list[MappingValue] = []
        row_counts: defaultdict[str, int] = defaultdict(int)
        sequence = 0
        for row_no, row in enumerate(rows, start=2):
            stats.rows_seen += 1
            name = clean(row[COL_NAME] if COL_NAME < len(row) else None)
            if not name:
                stats.empty_name_rows += 1
                continue
            row_counts[name] += 1
            group = groups.setdefault(name, ProductGroup(name=name, first_row=row_no))

            try:
                cost, shortage_as_zero = decimal_value(row[COL_COST] if COL_COST < len(row) else None, zero_on_shortage=True)
                if shortage_as_zero:
                    stats.shortage_cost_as_zero += 1
                weight, _ = decimal_value(row[COL_WEIGHT] if COL_WEIGHT < len(row) else None)
                stock = int_value(row[COL_STOCK] if COL_STOCK < len(row) else None)
                add_field(group, "cost", row_no, cost)
                add_field(group, "weight", row_no, weight)
                add_field(group, "safety_stock", row_no, stock)
                buyer_name = clean(row[COL_BUYER] if COL_BUYER < len(row) else None)
                if buyer_name:
                    buyer = buyer_lookup.get(buyer_name) or buyer_lookup.get(buyer_name.lower())
                    if buyer:
                        add_field(group, "buyer_user_id", row_no, buyer.id)
                    else:
                        stats.buyer_unresolved += 1
                        errors.append(
                            {
                                "type": "buyer_unresolved",
                                "row": row_no,
                                "product": name,
                                "value": buyer_name,
                                "message": "\u91c7\u8d2d\u4eba\u4e0d\u5b58\u5728\u6216\u5df2\u505c\u7528",
                            }
                        )
            except Exception as exc:
                errors.append({"type": "parse_error", "row": row_no, "product": name, "message": str(exc)})
                continue

            mapped_in_row = False
            for column_index, shop in shop_columns.items():
                sku = clean(row[column_index] if column_index < len(row) else None)
                if not sku:
                    continue
                mapped_in_row = True
                sequence += 1
                stats.source_mapping_entries += 1
                raw_mappings.append(
                    MappingValue(
                        row_no=row_no,
                        sequence=sequence,
                        product_name=name,
                        shop_id=shop.id,
                        shop_name=shop.display_name or shop.account_id,
                        sku=sku,
                    )
                )
            for column_index, header in enumerate(headers):
                if column_index in shop_columns or header not in unresolved_headers:
                    continue
                value = clean(row[column_index] if column_index < len(row) else None)
                if value:
                    stats.unmapped_shop_values += 1
            if not mapped_in_row:
                stats.no_mapping_rows += 1

        stats.unique_product_names = len(groups)
        stats.duplicate_product_names = sum(1 for count in row_counts.values() if count > 1)
        apply_latest_sku_owner(groups, raw_mappings, stats, errors)
        return groups, shop_columns, unresolved_headers, errors
    finally:
        workbook.close()


def write_backup(
    db,
    backup_prefix: Path,
    product_names: set[str],
) -> None:
    backup_prefix.parent.mkdir(parents=True, exist_ok=True)
    products = db.scalars(select(Product).where(Product.internal_name.in_(product_names)).order_by(Product.id)).all()
    product_ids = {product.id for product in products}

    with (backup_prefix.with_name(backup_prefix.name + "_products_before.csv")).open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["id", "product_code", "internal_name", "cost", "weight", "safety_stock", "buyer_user_id", "enabled", "updated_at"])
        for product in products:
            writer.writerow([
                product.id,
                product.product_code,
                product.internal_name,
                product.cost,
                product.weight,
                product.safety_stock,
                product.buyer_user_id,
                product.enabled,
                product.updated_at,
            ])

    mappings = db.scalars(
        select(ProductShopMapping)
        .options(joinedload(ProductShopMapping.product), joinedload(ProductShopMapping.shop))
        .where(ProductShopMapping.product_id.in_(product_ids))
        .order_by(ProductShopMapping.product_id, ProductShopMapping.shop_id, ProductShopMapping.id)
    ).all()
    with (backup_prefix.with_name(backup_prefix.name + "_mappings_before.csv")).open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["id", "product_id", "internal_name", "shop_id", "shop_name", "shop_sku", "updated_at"])
        for mapping in mappings:
            writer.writerow([
                mapping.id,
                mapping.product_id,
                mapping.product.internal_name if mapping.product else "",
                mapping.shop_id,
                mapping.shop.display_name if mapping.shop else "",
                mapping.shop_sku,
                mapping.updated_at,
            ])


def write_log(log_path: Path, stats: ImportStats, events: list[dict], shop_columns: dict[int, PlatformAccount], unresolved_headers: list[str]) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["type", "row", "shop", "sku", "product", "field", "value", "message"])
        for event in events:
            writer.writerow([
                event.get("type", ""),
                event.get("row", ""),
                event.get("shop", ""),
                event.get("sku", ""),
                event.get("product", ""),
                event.get("field", ""),
                event.get("value", ""),
                event.get("message", ""),
            ])
        writer.writerow([])
        writer.writerow(["shop_columns"])
        for column_index, shop in sorted(shop_columns.items()):
            writer.writerow(["shop_column", column_index + 1, shop.id, shop.display_name, shop.account_id, shop.platform])
        for header in unresolved_headers:
            writer.writerow(["unresolved_shop_column", "", "", header])
        writer.writerow([])
        for key, value in stats.__dict__.items():
            writer.writerow(["summary", key, value])


def run_import(
    path: Path,
    apply: bool,
    log_path: Path,
    field_conflict_policy: str,
    *,
    prune_shop_mappings: bool = True,
) -> ImportStats:
    stats = ImportStats()
    groups, shop_columns, unresolved_headers, events = read_excel_groups(path, stats)
    replace_shop_ids = {shop.id for shop in shop_columns.values()}
    all_desired_mapping_keys = {
        (shop_id, sku)
        for group in groups.values()
        for shop_id, sku_list in group.mappings.items()
        for sku in sku_list
    }

    with SessionLocal() as db:
        existing_products = {
            product.internal_name: product
            for product in db.scalars(select(Product).options(joinedload(Product.mappings))).unique().all()
        }
        used_codes = {code for (code,) in db.execute(select(Product.product_code)).all()}
        max_code = db.scalar(select(func.max(Product.product_code)).where(Product.product_code.like("P________")))
        existing_mappings: dict[tuple[int, str], ProductShopMapping] = {
            (mapping.shop_id, mapping.shop_sku): mapping
            for mapping in db.scalars(
                select(ProductShopMapping).options(joinedload(ProductShopMapping.product), joinedload(ProductShopMapping.shop))
            ).all()
        }

        if apply:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            write_backup(db, log_path.with_name(f"product_catalog_full_{timestamp}"), set(groups))

        for name, group in groups.items():
            product = existing_products.get(name)
            if product:
                stats.products_updated += 1
            else:
                product = Product(
                    product_code=generate_product_code(max_code, used_codes),
                    internal_name=name,
                    enabled=True,
                )
                db.add(product)
                db.flush()
                existing_products[name] = product
                stats.products_created += 1

            for field_name, values in group.fields.items():
                distinct_values = unique_values(values)
                if len(distinct_values) > 1:
                    stats.field_conflicts += 1
                    events.append(
                        {
                            "type": "field_conflict",
                            "row": group.first_row,
                            "product": name,
                            "field": field_name,
                            "value": " | ".join(str(value) for value in distinct_values),
                            "message": f"policy={field_conflict_policy}",
                        }
                    )
                if not values:
                    continue
                if len(distinct_values) > 1 and field_conflict_policy == "keep-existing":
                    continue
                selected = selected_field_value(values, field_conflict_policy)
                if getattr(product, field_name) != selected:
                    setattr(product, field_name, selected)
                    stats.field_updates += 1
            product.updated_at = utc_now()

            desired_by_shop = {shop_id: set(sku_list) for shop_id, sku_list in group.mappings.items()}
            if prune_shop_mappings:
                for mapping in list(product.mappings):
                    if mapping.shop_id not in replace_shop_ids:
                        continue
                    desired_skus = desired_by_shop.get(mapping.shop_id, set())
                    if mapping.shop_sku in desired_skus:
                        continue
                    if (mapping.shop_id, mapping.shop_sku) in all_desired_mapping_keys:
                        continue
                    shop = db.get(PlatformAccount, mapping.shop_id)
                    events.append(
                        {
                            "type": "mapping_deleted_not_in_source",
                            "row": group.first_row,
                            "shop": shop.display_name if shop else mapping.shop_id,
                            "sku": mapping.shop_sku,
                            "product": name,
                            "message": "shop column is controlled by this import",
                        }
                    )
                    existing_mappings.pop((mapping.shop_id, mapping.shop_sku), None)
                    db.delete(mapping)
                    stats.mappings_deleted += 1

            for shop_id, sku_list in group.mappings.items():
                shop = db.get(PlatformAccount, shop_id)
                for sku in sku_list:
                    existing = existing_mappings.get((shop_id, sku))
                    if existing:
                        if existing.product_id == product.id:
                            stats.mappings_existing += 1
                            existing.updated_at = utc_now()
                            continue
                        previous_product = existing.product.internal_name if existing.product else existing.product_id
                        existing.product_id = product.id
                        existing.updated_at = utc_now()
                        stats.mappings_reassigned += 1
                        events.append(
                            {
                                "type": "mapping_reassigned_by_later_row",
                                "row": group.first_row,
                                "shop": shop.display_name if shop else shop_id,
                                "sku": sku,
                                "product": name,
                                "message": f"from {previous_product}",
                            }
                        )
                        continue
                    mapping = ProductShopMapping(product_id=product.id, shop_id=shop_id, shop_sku=sku)
                    db.add(mapping)
                    db.flush()
                    existing_mappings[(shop_id, sku)] = mapping
                    stats.mappings_created += 1

        if apply:
            db.commit()
        else:
            db.rollback()

    write_log(log_path, stats, events, shop_columns, unresolved_headers)
    print("\u6a21\u5f0f:", "\u6b63\u5f0f\u5199\u5165" if apply else "dry-run \u672a\u5199\u5165")
    print("\u8bfb\u53d6\u884c\u6570:", stats.rows_seen)
    print("\u552f\u4e00\u4ea7\u54c1\u4e2d\u6587\u540d:", stats.unique_product_names)
    print("\u91cd\u590d\u4e2d\u6587\u540d\u6570:", stats.duplicate_product_names)
    print("\u65b0\u589e\u4ea7\u54c1:", stats.products_created)
    print("\u66f4\u65b0\u4ea7\u54c1:", stats.products_updated)
    print("\u57fa\u7840\u5b57\u6bb5\u66f4\u65b0:", stats.field_updates)
    print("\u57fa\u7840\u5b57\u6bb5\u51b2\u7a81:", stats.field_conflicts)
    print("\u65b0\u589e\u6620\u5c04:", stats.mappings_created)
    print("\u5df2\u5b58\u5728\u6620\u5c04:", stats.mappings_existing)
    print("\u5220\u9664\u6e90\u8868\u5df2\u4e0d\u5b58\u5728\u6620\u5c04:", stats.mappings_deleted)
    print("\u91cd\u65b0\u7ed1\u5b9a\u6620\u5c04:", stats.mappings_reassigned)
    print("\u51b2\u7a81\u6620\u5c04:", stats.mappings_conflicted)
    print("\u6e90\u8868\u6620\u5c04\u6761\u76ee:", stats.source_mapping_entries)
    print("\u6e90\u8868\u8de8\u4ea7\u54c1SKU\u5ffd\u7565:", stats.source_mapping_ignored_by_later_row)
    print("\u672a\u914d\u7f6e\u5e97\u94fa\u5217\u6709\u503c\u6570:", stats.unmapped_shop_values)
    print("\u65e0\u5e97\u94fa\u6620\u5c04\u884c:", stats.no_mapping_rows)
    print("\u6210\u672c\u7f3a\u8d27\u63090:", stats.shortage_cost_as_zero)
    print("\u65e5\u5fd7:", log_path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="\u5168\u91cf\u5bfc\u5165 Excel \u4ea7\u54c1\u76ee\u5f55\u5230\u4ea7\u54c1\u7ba1\u7406\u57fa\u7840\u6570\u636e")
    parser.add_argument("--file", type=Path, default=DEFAULT_EXCEL, help="Order follow up Excel \u6587\u4ef6")
    parser.add_argument("--apply", action="store_true", help="\u6b63\u5f0f\u5199\u5165\uff1b\u4e0d\u52a0\u5219 dry-run")
    parser.add_argument(
        "--field-conflict-policy",
        choices=("latest", "first", "keep-existing"),
        default="latest",
        help="\u540c\u4e00\u4e2d\u6587\u540d\u57fa\u7840\u5b57\u6bb5\u591a\u4e2a\u503c\u65f6\u7684\u7b56\u7565",
    )
    parser.add_argument(
        "--log",
        type=Path,
        default=ROOT / "logs" / "product_catalog_full_import.csv",
        help="\u5bfc\u5165\u65e5\u5fd7 CSV",
    )
    parser.add_argument(
        "--keep-existing-shop-mappings",
        action="store_true",
        help="\u4e0d\u5220\u9664\u6e90\u8868\u5e97\u94fa\u5217\u4e2d\u5df2\u4e0d\u5b58\u5728\u7684\u65e7 SKU \u6620\u5c04",
    )
    args = parser.parse_args()
    run_import(
        args.file,
        args.apply,
        args.log,
        args.field_conflict_policy,
        prune_shop_mappings=not args.keep_existing_shop_mappings,
    )


if __name__ == "__main__":
    main()
