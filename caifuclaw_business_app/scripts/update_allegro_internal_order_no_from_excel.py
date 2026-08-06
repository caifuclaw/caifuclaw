from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import or_, select, text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.database import SessionLocal, engine  # noqa: E402
from app.models import Order  # noqa: E402


DEFAULT_EXCEL = Path("./demo_data/allegro_orders.xlsx")
DEFAULT_REPORT_DIR = ROOT / "outputs" / "allegro_internal_order_no_updates"
REQUIRED_COLUMNS = {
    "订单编号": "order_no",
    "交易编号": "transaction_no",
    "内部单号": "internal_order_no",
    "货运单号": "tracking_no",
}


@dataclass(frozen=True)
class ExcelRow:
    row_number: int
    order_no: str
    transaction_no: str
    internal_order_no: str
    tracking_no: str


@dataclass(frozen=True)
class PlanItem:
    excel_row: ExcelRow
    status: str
    order_id: int | None = None
    current_internal_order_no: str = ""
    message: str = ""


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_excel_rows(path: Path) -> list[ExcelRow]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    column_indexes: dict[str, int] = {}
    for header, attr in REQUIRED_COLUMNS.items():
        try:
            column_indexes[attr] = header_row.index(header)
        except ValueError as exc:
            raise ValueError(f"Missing required column: {header}") from exc

    rows: list[ExcelRow] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [clean(cell.value) for cell in row]
        item = ExcelRow(
            row_number=row_number,
            order_no=values[column_indexes["order_no"]] if column_indexes["order_no"] < len(values) else "",
            transaction_no=values[column_indexes["transaction_no"]] if column_indexes["transaction_no"] < len(values) else "",
            internal_order_no=values[column_indexes["internal_order_no"]]
            if column_indexes["internal_order_no"] < len(values)
            else "",
            tracking_no=values[column_indexes["tracking_no"]] if column_indexes["tracking_no"] < len(values) else "",
        )
        if item.order_no or item.transaction_no or item.internal_order_no or item.tracking_no:
            rows.append(item)
    return rows


def ensure_internal_order_no_length() -> None:
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE IF EXISTS orders ADD COLUMN IF NOT EXISTS internal_order_no VARCHAR(32)"))
        conn.execute(text("ALTER TABLE IF EXISTS orders ALTER COLUMN internal_order_no TYPE VARCHAR(32)"))


def _orders_by_transaction(db, transaction_no: str) -> list[Order]:
    stmt = (
        select(Order)
        .where(
            Order.platform == "allegro",
            or_(
                Order.platform_order_id == transaction_no,
                Order.platform_order_no == transaction_no,
                Order.posting_number == transaction_no,
                Order.raw_payload["id"].astext == transaction_no,
            ),
        )
        .order_by(Order.id.asc())
    )
    return list(db.scalars(stmt).all())


def build_update_plan(db, rows: list[ExcelRow]) -> list[PlanItem]:
    plan: list[PlanItem] = []
    seen_internal_numbers: set[str] = set()
    seen_transactions: set[str] = set()

    for row in rows:
        if not row.internal_order_no:
            plan.append(PlanItem(row, "skipped_empty_internal_no", message="Excel internal order number is empty"))
            continue
        if not row.transaction_no:
            plan.append(PlanItem(row, "skipped_empty_transaction_no", message="Excel transaction number is empty"))
            continue
        if row.internal_order_no in seen_internal_numbers:
            plan.append(PlanItem(row, "skipped_duplicate_excel_internal_no", message="Internal order number repeats in Excel"))
            continue
        if row.transaction_no in seen_transactions:
            plan.append(PlanItem(row, "skipped_duplicate_excel_transaction_no", message="Transaction number repeats in Excel"))
            continue
        seen_internal_numbers.add(row.internal_order_no)
        seen_transactions.add(row.transaction_no)

        conflict = db.scalar(
            select(Order).where(
                Order.internal_order_no == row.internal_order_no,
                Order.platform != "allegro",
            )
        )
        if conflict:
            plan.append(
                PlanItem(
                    row,
                    "skipped_internal_no_used_by_other_platform",
                    order_id=int(conflict.id),
                    current_internal_order_no=conflict.internal_order_no or "",
                    message="Target internal order number already belongs to a non-Allegro order",
                )
            )
            continue

        orders_with_internal_no = list(
            db.scalars(select(Order).where(Order.internal_order_no == row.internal_order_no)).all()
        )
        matches = _orders_by_transaction(db, row.transaction_no)
        if not matches:
            plan.append(PlanItem(row, "skipped_no_match", message="No Allegro order matched the transaction number"))
            continue
        if len(matches) > 1:
            plan.append(
                PlanItem(
                    row,
                    "skipped_multiple_matches",
                    message="Multiple Allegro orders matched the transaction number",
                )
            )
            continue

        order = matches[0]
        conflicting_orders = [item for item in orders_with_internal_no if int(item.id) != int(order.id)]
        if conflicting_orders:
            plan.append(
                PlanItem(
                    row,
                    "skipped_internal_no_used_by_other_order",
                    order_id=int(conflicting_orders[0].id),
                    current_internal_order_no=conflicting_orders[0].internal_order_no or "",
                    message="Target internal order number already belongs to another order",
                )
            )
            continue

        current_internal_order_no = clean(order.internal_order_no)
        if current_internal_order_no == row.internal_order_no:
            plan.append(
                PlanItem(
                    row,
                    "unchanged",
                    order_id=int(order.id),
                    current_internal_order_no=current_internal_order_no,
                    message="Internal order number is already current",
                )
            )
            continue

        plan.append(
            PlanItem(
                row,
                "update",
                order_id=int(order.id),
                current_internal_order_no=current_internal_order_no,
                message="Ready to update",
            )
        )

    return plan


def write_report(path: Path, plan: list[PlanItem]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "excel_row",
                "status",
                "order_id",
                "transaction_no",
                "order_no",
                "old_internal_order_no",
                "new_internal_order_no",
                "tracking_no",
                "message",
            ],
        )
        writer.writeheader()
        for item in plan:
            writer.writerow(
                {
                    "excel_row": item.excel_row.row_number,
                    "status": item.status,
                    "order_id": item.order_id or "",
                    "transaction_no": item.excel_row.transaction_no,
                    "order_no": item.excel_row.order_no,
                    "old_internal_order_no": item.current_internal_order_no,
                    "new_internal_order_no": item.excel_row.internal_order_no,
                    "tracking_no": item.excel_row.tracking_no,
                    "message": item.message,
                }
            )


def default_report_path(apply: bool) -> Path:
    suffix = "applied" if apply else "dry_run"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return DEFAULT_REPORT_DIR / f"allegro_internal_order_no_{suffix}_{timestamp}.csv"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update Allegro orders.internal_order_no from an exported Excel file.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to the Allegro order export workbook.")
    parser.add_argument("--apply", action="store_true", help="Write changes to the database. Defaults to dry-run.")
    parser.add_argument("--report", type=Path, default=None, help="Optional CSV report path.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows = read_excel_rows(args.excel)
    report_path = args.report or default_report_path(args.apply)

    ensure_internal_order_no_length()
    with SessionLocal() as db:
        plan = build_update_plan(db, rows)
        updates = [item for item in plan if item.status == "update" and item.order_id]
        if args.apply:
            for item in updates:
                order = db.get(Order, item.order_id)
                if not order:
                    raise RuntimeError(f"Order disappeared before update: {item.order_id}")
                order.internal_order_no = item.excel_row.internal_order_no
                order.updated_at = datetime.utcnow()
            db.commit()
        else:
            db.rollback()

    write_report(report_path, plan)
    counts: dict[str, int] = {}
    for item in plan:
        counts[item.status] = counts.get(item.status, 0) + 1

    mode = "applied" if args.apply else "dry_run"
    print(f"mode={mode} excel_rows={len(rows)} updates={len(updates)} report={report_path}")
    for status in sorted(counts):
        print(f"{status}={counts[status]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
