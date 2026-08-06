# Company: 深圳智柠网络科技有限公司
# Author: mohsen liang

"""Move the two pending DMSMatrix orders to shipped and reconcile the follow-up workbook.

The script is intentionally dry-run by default.  It only targets ``dmsmatrix``
orders whose business status is ``待处理`` and refuses to continue unless the
expected number of rows is found.  Excel rows are matched by platform, order
number aliases, and SKU; shipping is registered through the existing
``订单出库`` lookup table so the formulas in ``订单总表`` stay intact.
"""

from __future__ import annotations

import argparse
import os
import shutil
import sys
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select, text

APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from app.database import SessionLocal, engine  # noqa: E402
from app.models import Order, OrderItem  # noqa: E402
from app.order_follow_up_export import _export_paths  # noqa: E402
from app.order_operation_logs import (  # noqa: E402
    ORDER_LOG_SYSTEM_SOURCE,
    SYSTEM_OPERATOR,
    add_order_operation_logs,
)
from app.settings import get_settings  # noqa: E402


PLATFORM = "dmsmatrix"
ORDER_STATUS_PENDING = "待处理"
ORDER_STATUS_SHIPPED = "已发货"
LOCAL_STATUS_SHIPPED = "shipped"
OPERATION_TYPE = "ship_pending_dmsmatrix_orders"
OPERATION_ATTRIBUTE = "DMSMatrix待处理订单处理"
ORDER_SHEET = "订单总表"
OUTBOUND_SHEET = "订单出库"
WORKBOOK_LOCK_KEY = 2026071901


@dataclass(frozen=True)
class Candidate:
    order: Order
    items: tuple[OrderItem, ...]


@dataclass(frozen=True)
class ExcelMatch:
    order_id: int
    order_item_id: int
    order_sheet_row: int
    outbound_row: int
    tracking_number: str
    outbound_date_before: object
    shipping_time_after: datetime


@dataclass(frozen=True)
class ExcelUpdate:
    path: Path
    backup_path: Path | None
    matches: tuple[ExcelMatch, ...]
    changed_count: int


def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _local_now(value: datetime) -> datetime:
    """Return the workbook timestamp in the app's configured local timezone."""
    from zoneinfo import ZoneInfo

    return value.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Asia/Shanghai")).replace(tzinfo=None)


def _normalise(value: object) -> str:
    return "" if value is None else str(value).strip().lower()


def _normalise_platform(value: object) -> str:
    platform = _normalise(value)
    return "dmsmatrix" if platform in {"dmsmatrix", "fruugo", "fruugo-dms"} else platform


def _aliases(*values: object) -> set[str]:
    return {_normalise(value) for value in values if _normalise(value)}


def _order_number_aliases(order: Order) -> set[str]:
    return _aliases(
        order.platform_order_no,
        order.posting_number,
        order.platform_order_id,
        order.shipment_tracking_number,
    )


def _item_sku_aliases(item: OrderItem) -> set[str]:
    payload = item.raw_payload if isinstance(item.raw_payload, dict) else {}
    return _aliases(
        item.sku,
        payload.get("sku"),
        payload.get("offer_id"),
    )


def _tracking_aliases(order: Order) -> set[str]:
    raw_payload = getattr(order, "raw_payload", {})
    payload = raw_payload if isinstance(raw_payload, dict) else {}
    return _aliases(
        order.shipment_tracking_number,
        order.posting_number,
        payload.get("tracking_number"),
        payload.get("label_tracking_number"),
        payload.get("waybill_number"),
    )


def load_candidates(db, *, order_ids: list[int] | None = None, expected_count: int = 2) -> list[Candidate]:
    """Load exactly the pending DMSMatrix rows that are safe to process."""
    stmt = (
        select(Order)
        .where(func.lower(Order.platform) == PLATFORM, Order.biz_status == ORDER_STATUS_PENDING)
        .order_by(Order.id)
    )
    unique_ids = list(dict.fromkeys(int(value) for value in (order_ids or []) if int(value) > 0))
    if unique_ids:
        stmt = stmt.where(Order.id.in_(unique_ids))

    rows = list(db.scalars(stmt).all())
    if expected_count >= 0 and len(rows) != expected_count:
        raise RuntimeError(
            f"Expected {expected_count} pending {PLATFORM} orders, found {len(rows)}: "
            f"{', '.join(str(row.id) for row in rows) or '-'}"
        )

    candidates: list[Candidate] = []
    for order in rows:
        items = tuple(db.scalars(select(OrderItem).where(OrderItem.order_id == order.id).order_by(OrderItem.id)).all())
        if not items:
            raise RuntimeError(f"Order {order.id} has no order items; refusing to update it")
        candidates.append(Candidate(order=order, items=items))
    return candidates


def _header_indexes(sheet) -> dict[str, int]:
    headers = {_normalise(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    required = {"平台", "订单编号", "sku", "shipping time"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"Worksheet {ORDER_SHEET} is missing required headers: {', '.join(missing)}")
    return headers


def _match_excel_rows(sheet, candidates: list[Candidate]) -> list[ExcelMatch]:
    indexes = _header_indexes(sheet)
    matches: list[ExcelMatch] = []
    used_rows: set[int] = set()
    for candidate in candidates:
        order_numbers = _order_number_aliases(candidate.order)
        for item in candidate.items:
            skus = _item_sku_aliases(item)
            found: list[int] = []
            for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if row_number in used_rows:
                    continue
                if _normalise_platform(row[indexes["平台"] - 1].value) != PLATFORM:
                    continue
                row_order = _normalise(row[indexes["订单编号"] - 1].value)
                row_sku = _normalise(row[indexes["sku"] - 1].value)
                if row_order in order_numbers and row_sku in skus:
                    found.append(row_number)
            if len(found) != 1:
                raise RuntimeError(
                    f"Order {candidate.order.id} item {item.id} matched {len(found)} Excel rows; "
                    "expected exactly one"
                )
            row_number = found[0]
            used_rows.add(row_number)
            matches.append(
                ExcelMatch(
                    order_id=candidate.order.id,
                    order_item_id=item.id,
                    order_sheet_row=row_number,
                    outbound_row=0,
                    tracking_number="",
                    outbound_date_before=None,
                    shipping_time_after=datetime.min,
                )
            )
    return matches


def _match_outbound_rows(sheet, candidates: list[Candidate], order_matches: list[ExcelMatch]) -> list[ExcelMatch]:
    headers = {_normalise(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    missing = sorted({"物流单号", "发货日期"} - set(headers))
    if missing:
        raise RuntimeError(f"Worksheet {OUTBOUND_SHEET} is missing required headers: {', '.join(missing)}")
    tracking_index = headers["物流单号"]
    date_index = headers["发货日期"]
    candidates_by_id = {candidate.order.id: candidate for candidate in candidates}
    used_rows: set[int] = set()
    result: list[ExcelMatch] = []
    for match in order_matches:
        order = candidates_by_id[match.order_id].order
        tracking_values = _tracking_aliases(order)
        if not tracking_values:
            raise RuntimeError(f"Order {order.id} has no tracking number/posting number for {OUTBOUND_SHEET}")
        found: list[int] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row_number in used_rows:
                continue
            if _normalise(row[tracking_index - 1].value) in tracking_values:
                found.append(row_number)
        if len(found) > 1:
            raise RuntimeError(
                f"Order {order.id} tracking number matched {len(found)} rows in {OUTBOUND_SHEET}; expected at most one"
            )
        outbound_row = found[0] if found else 0
        if outbound_row:
            used_rows.add(outbound_row)
            before = sheet.cell(outbound_row, date_index).value
        else:
            before = None
        tracking_number = order.posting_number or order.shipment_tracking_number or next(iter(tracking_values))
        result.append(
            ExcelMatch(
                order_id=match.order_id,
                order_item_id=match.order_item_id,
                order_sheet_row=match.order_sheet_row,
                outbound_row=outbound_row,
                tracking_number=tracking_number,
                outbound_date_before=before,
                shipping_time_after=match.shipping_time_after,
            )
        )
    return result


def _backup_path(workbook_path: Path, backup_dir: Path, now: datetime) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    return backup_dir / f"{workbook_path.stem}_{now.strftime('%Y%m%d_%H%M%S')}_dmsmatrix{workbook_path.suffix}"


def update_workbook(
    path: Path,
    candidates: list[Candidate],
    *,
    shipping_time: datetime,
    apply: bool,
    backup_dir: Path,
) -> ExcelUpdate:
    if not path.is_file():
        raise FileNotFoundError(f"Order follow-up workbook not found: {path}")

    # XLSX stores datetimes with millisecond precision.
    shipping_time = shipping_time.replace(microsecond=(shipping_time.microsecond // 1000) * 1000)

    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        if ORDER_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"Worksheet not found: {ORDER_SHEET}")
        if OUTBOUND_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"Worksheet not found: {OUTBOUND_SHEET}")
        order_sheet = workbook[ORDER_SHEET]
        outbound_sheet = workbook[OUTBOUND_SHEET]
        order_matches = _match_excel_rows(order_sheet, candidates)
        matches = tuple(
            ExcelMatch(
                order_id=match.order_id,
                order_item_id=match.order_item_id,
                order_sheet_row=match.order_sheet_row,
                outbound_row=match.outbound_row,
                tracking_number=match.tracking_number,
                outbound_date_before=match.outbound_date_before,
                shipping_time_after=shipping_time,
            )
            for match in _match_outbound_rows(outbound_sheet, candidates, order_matches)
        )
        outbound_headers = {_normalise(cell.value): index for index, cell in enumerate(outbound_sheet[1], start=1)}
        tracking_index = outbound_headers["物流单号"]
        date_index = outbound_headers["发货日期"]
        changed_by_tracking: dict[str, ExcelMatch] = {}
        for match in matches:
            if match.outbound_row != 0 and match.outbound_date_before not in (None, ""):
                continue
            changed_by_tracking.setdefault(_normalise(match.tracking_number), match)
        changed = list(changed_by_tracking.values())
        if not apply or not changed:
            return ExcelUpdate(path=path, backup_path=None, matches=matches, changed_count=len(changed))

        now = _utc_now()
        backup = _backup_path(path, backup_dir, now)
        shutil.copy2(path, backup)
        temp = path.with_name(f".{path.stem}.dmsmatrix.{os.getpid()}.tmp{path.suffix}")
        try:
            for match in changed:
                outbound_row = match.outbound_row or outbound_sheet.max_row + 1
                outbound_sheet.cell(outbound_row, tracking_index).value = match.tracking_number
                outbound_sheet.cell(outbound_row, date_index).value = match.shipping_time_after
            workbook.save(temp)
            workbook.close()
            workbook = None
            validation = load_workbook(temp, read_only=True, data_only=False, keep_links=True)
            try:
                validation_sheet = validation[OUTBOUND_SHEET]
                for match in changed:
                    outbound_row = match.outbound_row
                    if outbound_row == 0:
                        for row_number, row in enumerate(validation_sheet.iter_rows(min_row=2), start=2):
                            if _normalise(row[tracking_index - 1].value) == _normalise(match.tracking_number):
                                outbound_row = row_number
                                break
                    value = validation_sheet.cell(outbound_row, date_index).value
                    if value != match.shipping_time_after:
                        raise RuntimeError(
                            f"Excel validation failed at outbound row {outbound_row}: "
                            f"{value!r} != {match.shipping_time_after!r}"
                        )
            finally:
                validation.close()
            os.replace(temp, path)
            return ExcelUpdate(path=path, backup_path=backup, matches=matches, changed_count=len(changed))
        except Exception:
            if temp.exists():
                temp.unlink()
            if backup.exists():
                shutil.copy2(backup, path)
            raise
    finally:
        if workbook is not None:
            workbook.close()


@contextmanager
def workbook_lock():
    connection = engine.connect()
    acquired = False
    try:
        if connection.dialect.name == "postgresql":
            acquired = bool(
                connection.execute(text("SELECT pg_try_advisory_lock(:key)"), {"key": WORKBOOK_LOCK_KEY}).scalar()
            )
            if not acquired:
                raise RuntimeError("Order follow-up workbook is busy; retry after the export worker finishes")
        yield
    finally:
        if acquired:
            connection.execute(text("SELECT pg_advisory_unlock(:key)"), {"key": WORKBOOK_LOCK_KEY})
        connection.close()


def apply_database_changes(db, candidates: list[Candidate], *, shipped_at: datetime, excel_path: Path) -> None:
    order_ids = [candidate.order.id for candidate in candidates]
    rows = list(
        db.scalars(
            select(Order).where(
                Order.id.in_(order_ids),
                Order.platform == PLATFORM,
                Order.biz_status == ORDER_STATUS_PENDING,
            )
        ).all()
    )
    if {row.id for row in rows} != set(order_ids):
        raise RuntimeError("One or more target orders changed while the workbook was being prepared")

    for row in rows:
        row.biz_status = ORDER_STATUS_SHIPPED
        row.local_status = LOCAL_STATUS_SHIPPED
        row.label_printed_at = row.label_printed_at or shipped_at
        row.marked_shipped_at = row.marked_shipped_at or shipped_at
        row.shipped_at = row.shipped_at or shipped_at
        row.updated_at = shipped_at

    add_order_operation_logs(
        db,
        rows,
        operation_type=OPERATION_TYPE,
        operation_attribute=OPERATION_ATTRIBUTE,
        description=lambda order: (
            f"脚本处理：DMSMatrix订单 {order.platform_order_no or order.posting_number or order.id} "
            f"状态：{ORDER_STATUS_PENDING} -> {ORDER_STATUS_SHIPPED}；同步跟进表 {excel_path.name}"
        ),
        operator=SYSTEM_OPERATOR,
        source=ORDER_LOG_SYSTEM_SOURCE,
        operated_at=shipped_at,
        event_key=lambda order: f"{OPERATION_TYPE}:{order.id}",
        extra=lambda order: {
            "platform": PLATFORM,
            "status_before": ORDER_STATUS_PENDING,
            "status_after": ORDER_STATUS_SHIPPED,
            "workbook": str(excel_path),
            "script": Path(__file__).name,
        },
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ship the pending DMSMatrix orders and update their Excel rows.")
    parser.add_argument("--apply", action="store_true", help="Write database and workbook changes. Defaults to dry-run.")
    parser.add_argument("--order-id", type=int, action="append", default=[], help="Optional order id filter; repeatable.")
    parser.add_argument("--expected-count", type=int, default=2, help="Expected pending rows; defaults to 2.")
    parser.add_argument("--excel", type=Path, default=None, help="Override the canonical follow-up workbook path.")
    parser.add_argument("--backup-dir", type=Path, default=None, help="Override the workbook backup directory.")
    parser.add_argument("--max-print", type=int, default=20, help="Maximum target rows to print.")
    return parser.parse_args()


def print_result(candidates: list[Candidate], update: ExcelUpdate, *, apply: bool, max_print: int) -> None:
    mode = "apply" if apply else "dry-run"
    print(f"mode={mode} orders={len(candidates)} excel_rows={len(update.matches)} excel_changed={update.changed_count}")
    for match in update.matches[: max(0, max_print)]:
        before = "missing" if match.outbound_row == 0 else (
            "blank" if match.outbound_date_before in (None, "") else str(match.outbound_date_before)
        )
        print(
            f"order_id={match.order_id}\torder_item_id={match.order_item_id}\t"
            f"order_sheet_row={match.order_sheet_row}\toutbound_row={match.outbound_row or 'append'}\t"
            f"tracking_number={match.tracking_number}\toutbound_date_before={before}\t"
            f"shipping_time_after={match.shipping_time_after.isoformat(sep=' ')}"
        )
    if update.backup_path:
        print(f"excel_backup={update.backup_path}")


def main() -> int:
    args = parse_args()
    if args.expected_count < 0:
        raise SystemExit("--expected-count must be >= 0")

    settings = get_settings()
    paths = _export_paths(settings)
    excel_path = (args.excel or paths.workbook).expanduser().resolve()
    backup_dir = (args.backup_dir or paths.backup_dir).expanduser().resolve()
    shipped_at = _utc_now()
    shipped_at = shipped_at.replace(microsecond=(shipped_at.microsecond // 1000) * 1000)

    with SessionLocal() as db:
        try:
            candidates = load_candidates(db, order_ids=args.order_id, expected_count=args.expected_count)
            with workbook_lock():
                update = update_workbook(
                    excel_path,
                    candidates,
                    shipping_time=_local_now(shipped_at),
                    apply=args.apply,
                    backup_dir=backup_dir,
                )
                if args.apply:
                    try:
                        apply_database_changes(db, candidates, shipped_at=shipped_at, excel_path=excel_path)
                        db.commit()
                    except Exception:
                        db.rollback()
                        if update.backup_path and update.backup_path.is_file():
                            shutil.copy2(update.backup_path, excel_path)
                        raise
                else:
                    db.rollback()
            print_result(candidates, update, apply=args.apply, max_print=args.max_print)
            return 0
        except Exception:
            db.rollback()
            raise


if __name__ == "__main__":
    raise SystemExit(main())
